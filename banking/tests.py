import os
import tempfile
import zipfile
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.test.utils import override_settings
from django.urls import reverse

from config.storage import ENCRYPTION_MARKER
from portfolio.ownership import AssetOwnershipCategory

from .management.commands.import_monica_bank_positions import MONICA_BANK_INVESTMENT_POSITIONS
from .models import (
    BankBalance,
    BankInvestmentPosition,
    BankMovement,
    BankStatementImport,
)
from .services import (
    build_bank_account_overview,
    build_banking_dashboard,
    build_banking_ownership_overview,
    build_card_spending_overview,
    classify_movement,
    import_statement,
    load_rows_from_xls,
    parse_spanish_decimal,
    parse_statement_file,
)


class BankingServicesTests(TestCase):
    def _build_broken_dimension_xlsx(self) -> bytes:
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "movimientos"
        worksheet["C2"] = "Titular"
        worksheet["D2"] = "Saldo disponible"
        worksheet["E2"] = "Saldo real"
        worksheet["C3"] = "JOAQUIN PIQUER MARTI"
        worksheet["D3"] = "100.104,89 EUR"
        worksheet["E3"] = "100.104,89 EUR"
        worksheet["C5"] = "Cuenta"
        worksheet["D5"] = "Retenciones"
        worksheet["E5"] = "Saldo consolidado"
        worksheet["C6"] = "ES19 0049 4898 94 2916239927"
        worksheet["D6"] = "0,00 EUR"
        worksheet["E6"] = "100.104,89 EUR"
        worksheet["A7"] = "Movimientos Fecha desde 07/01/2026 Fecha Hasta 06/04/2026"
        worksheet["K7"] = "06/04/2026 | 21:35:02"
        headers = [
            "Fecha Operación",
            "Fecha Valor",
            "Concepto",
            "Importe",
            "Divisa",
            "Saldo",
            "Divisa",
            "Código",
            "Número de documento",
            "Referencia 1",
            "Referencia 2",
            "Información adicional",
        ]
        for column_index, value in enumerate(headers, start=1):
            worksheet.cell(row=8, column=column_index, value=value)
        worksheet.append(
            [
                "19/03/2026",
                "19/03/2026",
                "Transferencia De Neos Additives S.l., Concepto Pago De Deuda.",
                50000.0,
                "EUR",
                100104.89,
                "EUR",
                "071",
                "",
                "",
                "",
                "",
            ]
        )

        payload = BytesIO()
        workbook.save(payload)

        broken_payload = BytesIO()
        with zipfile.ZipFile(BytesIO(payload.getvalue()), "r") as source_zip:
            with zipfile.ZipFile(broken_payload, "w", zipfile.ZIP_DEFLATED) as target_zip:
                for zip_info in source_zip.infolist():
                    file_bytes = source_zip.read(zip_info.filename)
                    if zip_info.filename == "xl/worksheets/sheet1.xml":
                        file_bytes = file_bytes.replace(
                            b'<dimension ref="A1:L9"/>',
                            b'<dimension ref="A1"/>',
                        )
                    target_zip.writestr(zip_info, file_bytes)

        return broken_payload.getvalue()

    def _build_statement_html(self, holder_name: str = "", iban: str = "ES12 3456 7890 1234") -> str:
        holder_row = f"<tr><td>Titular: {holder_name}</td></tr>" if holder_name else ""
        return f"""
<html>
  <body>
    <table>
      <tr><td>Cuenta: {iban}</td></tr>
      {holder_row}
      <tr><td>Divisa: EUR</td></tr>
      <tr><td>Desde 01/04/2026 hasta 05/04/2026</td></tr>
      <tr>
        <td>F. Operativa</td>
        <td>Concepto</td>
        <td>F. Valor</td>
        <td>Importe</td>
        <td>Saldo</td>
        <td>Referencia 1</td>
        <td>Referencia 2</td>
      </tr>
      <tr>
        <td>05/04/2026</td>
        <td>NOMINA ABRIL</td>
        <td>05/04/2026</td>
        <td>1.200,50</td>
        <td>3.400,75</td>
        <td>ABC</td>
        <td>DEF</td>
      </tr>
    </table>
  </body>
</html>
"""

    def _build_card_statement_html(self) -> str:
        return """
<html>
  <body>
    <table>
      <tr><td>Saldos i moviments</td></tr>
      <tr><td>Contracte</td><td>004273514883</td><td>Compte relacionat</td><td>0081-0278-14-0006302439</td></tr>
      <tr><td>Titular</td><td>PIQUER MARTI ,JOAQUIN</td></tr>
      <tr><td>Targeta:</td><td>5402________3026</td><td>BS CARD MASTERCARD</td></tr>
      <tr><td>Titular targeta</td><td>PIQUER MARTI ,JOAQUIN</td></tr>
      <tr><td>MOVIMIENTOS DE DEBITO</td></tr>
      <tr>
        <td>DATA</td>
        <td>CONCEPTE</td>
        <td>LOCALITAT</td>
        <td>SIT. MOV.</td>
        <td>IMPORT</td>
      </tr>
      <tr>
        <td>05/04</td>
        <td>WWW.AMAZON</td>
        <td>LUXEMBOURG</td>
        <td>AUT</td>
        <td>19,99</td>
      </tr>
      <tr>
        <td>03/04</td>
        <td>APPLE.COM/BILL</td>
        <td>CORK</td>
        <td>AUT</td>
        <td>2,99</td>
      </tr>
      <tr><td></td><td></td><td>TOTAL OPERACIONS</td><td>22,98</td><td>EUR</td></tr>
      <tr>
        <td>DATA</td>
        <td>CONCEPTE</td>
        <td>LOCALITAT</td>
        <td>IMPORT</td>
      </tr>
      <tr>
        <td>31/03</td>
        <td>WWW.AMAZON* NB99W7NE4</td>
        <td>LUXEMBOURG</td>
        <td>7,53</td>
      </tr>
      <tr>
        <td>30/03</td>
        <td>MERCADONA AVDA CONSTITUCI</td>
        <td>ONDA</td>
        <td>55,07</td>
      </tr>
    </table>
  </body>
</html>
"""

    def test_parse_spanish_decimal(self):
        self.assertEqual(parse_spanish_decimal("3.126,77"), Decimal("3126.77"))
        self.assertEqual(parse_spanish_decimal("-171,05"), Decimal("-171.05"))
        self.assertEqual(parse_spanish_decimal("3126.77"), Decimal("3126.77"))

    def test_parse_statement_file_accepts_xlsx_with_broken_dimension_metadata(self):
        document = SimpleUploadedFile(
            "MovimientosCuenta.xlsx",
            self._build_broken_dimension_xlsx(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        parsed = parse_statement_file(document)

        self.assertEqual(parsed["metadata"]["holder_name"], "JOAQUIN PIQUER MARTI")
        self.assertEqual(parsed["metadata"]["iban"], "ES19 0049 4898 94 2916239927")
        self.assertEqual(parsed["metadata"]["account_label"], "Cuenta 9927")
        self.assertEqual(parsed["metadata"]["period_start"], date(2026, 1, 7))
        self.assertEqual(parsed["metadata"]["period_end"], date(2026, 4, 6))
        self.assertEqual(parsed["metadata"]["closing_balance"], Decimal("100104.89"))
        self.assertEqual(len(parsed["movements"]), 1)
        self.assertEqual(parsed["movements"][0].amount, Decimal("50000"))

    def test_classify_plan_contribution(self):
        classified = classify_movement(
            "APORTACION PERIODICA POL. 32000006 CERT. 95526 BS PLAN AHORRO SEMESTRAL",
            Decimal("-171.05"),
        )
        self.assertEqual(classified.group, BankMovement.MovementGroup.PENSION)
        self.assertEqual(classified.bucket, "Aportaciones a planes")

    def test_classify_dividend(self):
        classified = classify_movement("COBRO DIVIDENDO IBERDROLA", Decimal("82.30"))
        self.assertEqual(classified.group, BankMovement.MovementGroup.DIVIDEND)
        self.assertEqual(classified.bucket, "Dividendos de acciones")

    def test_classify_card_settlement_as_non_consumption_expense_bucket(self):
        classified = classify_movement("LIQUIDACION TARJETA VISA", Decimal("-325.40"))
        self.assertEqual(classified.group, BankMovement.MovementGroup.EXPENSE)
        self.assertEqual(classified.bucket, "Liquidacion de tarjeta")

    def test_classify_rent_income(self):
        classified = classify_movement("ALQUILER PISO ADRIAN", Decimal("650.00"))
        self.assertEqual(classified.group, BankMovement.MovementGroup.INCOME)
        self.assertEqual(classified.bucket, "Alquiler piso (Adrian)")

    def test_build_dashboard_groups_months_and_expenses(self):
        feb_statement = BankStatementImport.objects.create(
            source_filename="feb.xls",
            source_file="banking/statements/feb.xls",
            file_checksum="checksum-feb",
            account_label="Cuenta 1234",
            period_end="2026-02-28",
            total_income=Decimal("3126.77"),
            total_expenses=Decimal("93.79"),
            total_pension_contributions=Decimal("0.00"),
            total_dividends=Decimal("82.30"),
            closing_balance=Decimal("10234.55"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        mar_statement = BankStatementImport.objects.create(
            source_filename="mar.xls",
            source_file="banking/statements/mar.xls",
            file_checksum="checksum-mar",
            account_label="Cuenta 5678",
            period_end="2026-03-31",
            total_income=Decimal("650.00"),
            total_expenses=Decimal("45.00"),
            total_pension_contributions=Decimal("0.00"),
            total_dividends=Decimal("0.00"),
            closing_balance=Decimal("9000.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        BankMovement.objects.create(
            statement_import=feb_statement,
            booking_date="2026-02-26",
            concept="NOMINA",
            normalized_concept="NOMINA",
            amount=Decimal("3126.77"),
            movement_group=BankMovement.MovementGroup.INCOME,
            concept_bucket="Nomina",
        )
        BankMovement.objects.create(
            statement_import=feb_statement,
            booking_date="2026-02-27",
            concept="COBRO DIVIDENDO IBERDROLA",
            normalized_concept="COBRO DIVIDENDO IBERDROLA",
            amount=Decimal("82.30"),
            movement_group=BankMovement.MovementGroup.DIVIDEND,
            concept_bucket="Dividendos de acciones",
        )
        BankMovement.objects.create(
            statement_import=mar_statement,
            booking_date="2026-03-05",
            concept="ALQUILER PISO ADRIAN",
            normalized_concept="ALQUILER PISO ADRIAN",
            amount=Decimal("650.00"),
            movement_group=BankMovement.MovementGroup.INCOME,
            concept_bucket="Alquiler piso (Adrian)",
        )
        BankMovement.objects.create(
            statement_import=feb_statement,
            booking_date="2026-02-23",
            concept="COMPRA TARJ. MERCADONA",
            normalized_concept="COMPRA TARJ. MERCADONA",
            amount=Decimal("-73.80"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Supermercado",
        )
        BankMovement.objects.create(
            statement_import=feb_statement,
            booking_date="2026-02-20",
            concept="NETFLIX",
            normalized_concept="NETFLIX",
            amount=Decimal("-19.99"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Suscripciones",
        )
        BankMovement.objects.create(
            statement_import=mar_statement,
            booking_date="2026-03-03",
            concept="PAGO BIZUM",
            normalized_concept="PAGO BIZUM",
            amount=Decimal("-45.00"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Bizum",
        )

        dashboard = build_banking_dashboard()

        self.assertEqual(dashboard["statement_summary"]["months_count"], 2)
        self.assertEqual(dashboard["monthly_summaries"][0]["label"], "2026-03")
        self.assertEqual(dashboard["monthly_summaries"][0]["net_cash_flow"], Decimal("605.00"))
        self.assertEqual(dashboard["monthly_summaries"][1]["net_cash_flow"], Decimal("3115.28"))
        self.assertEqual(dashboard["income_months"], ["2026-02", "2026-03"])
        self.assertEqual(dashboard["income_matrix"][0]["concept"], "Nomina")
        self.assertEqual(dashboard["income_matrix"][0]["total"], Decimal("3126.77"))
        self.assertEqual(dashboard["income_matrix"][1]["total"], Decimal("650.00"))
        self.assertEqual(dashboard["income_matrix"][2]["total"], Decimal("82.30"))
        self.assertEqual(dashboard["expense_months"], ["2026-02", "2026-03"])
        self.assertEqual(dashboard["expense_matrix"][0]["concept"], "Supermercado")
        self.assertEqual(dashboard["accounts_summary"]["accounts_count"], 2)
        self.assertEqual(dashboard["accounts_summary"]["current_balance"], Decimal("19234.55"))
        self.assertEqual(dashboard["tracked_accounts"][0]["current_balance"], Decimal("10234.55"))

    def test_build_dashboard_splits_single_multi_month_import_by_movement_month(self):
        statement = BankStatementImport.objects.create(
            source_filename="q1.xls",
            source_file="banking/statements/q1.xls",
            file_checksum="checksum-q1",
            account_label="Cuenta 1234",
            period_start="2026-01-01",
            period_end="2026-03-31",
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        BankMovement.objects.create(
            statement_import=statement,
            booking_date="2026-01-15",
            concept="NOMINA",
            normalized_concept="NOMINA",
            amount=Decimal("2500.00"),
            balance=Decimal("5000.00"),
            movement_group=BankMovement.MovementGroup.INCOME,
            concept_bucket="Nomina",
        )
        BankMovement.objects.create(
            statement_import=statement,
            booking_date="2026-01-20",
            concept="MERCADONA",
            normalized_concept="MERCADONA",
            amount=Decimal("-100.00"),
            balance=Decimal("4900.00"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Supermercado",
        )
        BankMovement.objects.create(
            statement_import=statement,
            booking_date="2026-02-10",
            concept="ALQUILER PISO ADRIAN",
            normalized_concept="ALQUILER PISO ADRIAN",
            amount=Decimal("650.00"),
            balance=Decimal("5550.00"),
            movement_group=BankMovement.MovementGroup.INCOME,
            concept_bucket="Alquiler piso (Adrian)",
        )
        BankMovement.objects.create(
            statement_import=statement,
            booking_date="2026-02-25",
            concept="PAGO BIZUM",
            normalized_concept="PAGO BIZUM",
            amount=Decimal("-50.00"),
            balance=Decimal("5500.00"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Bizum",
        )
        BankMovement.objects.create(
            statement_import=statement,
            booking_date="2026-03-05",
            concept="COBRO DIVIDENDO IBERDROLA",
            normalized_concept="COBRO DIVIDENDO IBERDROLA",
            amount=Decimal("82.30"),
            balance=Decimal("5582.30"),
            movement_group=BankMovement.MovementGroup.DIVIDEND,
            concept_bucket="Dividendos de acciones",
        )
        BankMovement.objects.create(
            statement_import=statement,
            booking_date="2026-03-12",
            concept="APORTACION PERIODICA POL.",
            normalized_concept="APORTACION PERIODICA POL.",
            amount=Decimal("-171.05"),
            balance=Decimal("5411.25"),
            movement_group=BankMovement.MovementGroup.PENSION,
            concept_bucket="Aportaciones a planes",
        )

        dashboard = build_banking_dashboard()

        self.assertEqual(dashboard["statement_summary"]["months_count"], 3)
        self.assertEqual(dashboard["statement_summary"]["total_income"], Decimal("3150.00"))
        self.assertEqual(dashboard["statement_summary"]["total_expenses"], Decimal("150.00"))
        self.assertEqual(dashboard["statement_summary"]["total_dividends"], Decimal("82.30"))
        self.assertEqual(dashboard["statement_summary"]["total_pension_contributions"], Decimal("171.05"))
        self.assertEqual([row["label"] for row in dashboard["monthly_summaries"]], ["2026-03", "2026-02", "2026-01"])
        self.assertEqual(dashboard["monthly_summaries"][0]["closing_balance"], Decimal("5411.25"))
        self.assertEqual(dashboard["monthly_summaries"][1]["closing_balance"], Decimal("5500.00"))
        self.assertEqual(dashboard["monthly_summaries"][2]["closing_balance"], Decimal("4900.00"))
        self.assertEqual(dashboard["income_months"], ["2026-01", "2026-02", "2026-03"])
        self.assertEqual(dashboard["income_matrix"][0]["values"], [Decimal("2500.00"), Decimal("0.00"), Decimal("0.00")])
        self.assertEqual(dashboard["income_matrix"][1]["values"], [Decimal("0.00"), Decimal("650.00"), Decimal("0.00")])
        self.assertEqual(dashboard["income_matrix"][2]["values"], [Decimal("0.00"), Decimal("0.00"), Decimal("82.30")])

    def test_build_bank_account_overview_merges_imported_and_manual_data(self):
        BankBalance.objects.create(
            ownership_category=AssetOwnershipCategory.MONICA,
            institution="Banco Sabadell",
            account_name="Cuenta 1234",
            deposited_amount=Decimal("5000.00"),
            current_balance=Decimal("5100.00"),
            annual_interest_income=Decimal("25.00"),
            notes="Cuenta principal",
        )
        BankStatementImport.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            source_filename="mar.xls",
            source_file="banking/statements/mar.xls",
            file_checksum="overview-account-mar",
            account_label="Cuenta 1234",
            period_end="2026-03-31",
            closing_balance=Decimal("6200.00"),
            total_income=Decimal("3000.00"),
            total_expenses=Decimal("900.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        overview = build_bank_account_overview()

        self.assertEqual(overview["summary"]["accounts_count"], 1)
        self.assertEqual(overview["summary"]["current_balance"], Decimal("6200.00"))
        self.assertEqual(overview["summary"]["deposited_amount"], Decimal("5000.00"))
        self.assertEqual(overview["summary"]["annual_interest_income"], Decimal("25.00"))
        self.assertEqual(overview["accounts"][0]["source_label"], "Manual + extracto")
        self.assertEqual(overview["accounts"][0]["ownership_category"], AssetOwnershipCategory.MONICA)
        self.assertEqual(overview["accounts"][0]["statement_count"], 1)

    def test_build_banking_ownership_overview_separates_joint_ximo_and_monica(self):
        tracked_accounts = [
            {
                "account_name": "Cuenta conjunta",
                "institution": "Banco Sabadell",
                "ownership_category": AssetOwnershipCategory.JOINT,
                "current_balance": Decimal("5000.00"),
                "annual_interest_income": Decimal("20.00"),
            },
            {
                "account_name": "Cuenta herencia Ximo",
                "institution": "CaixaBank",
                "ownership_category": AssetOwnershipCategory.XIMO,
                "current_balance": Decimal("9000.00"),
                "annual_interest_income": Decimal("45.00"),
            },
        ]
        tracked_cards = [
            {
                "card_name": "Visa Monica",
                "institution": "Banco Sabadell",
                "ownership_category": AssetOwnershipCategory.MONICA,
                "latest_spent": Decimal("430.00"),
            }
        ]
        BankInvestmentPosition.objects.create(
            ownership_category=AssetOwnershipCategory.XIMO,
            institution="CaixaBank",
            product_name="Deposito herencia",
            current_value=Decimal("12000.00"),
            annual_income=Decimal("180.00"),
        )

        overview = build_banking_ownership_overview(tracked_accounts, tracked_cards)
        groups = {group["ownership_category"]: group for group in overview["groups"]}

        self.assertEqual(groups[AssetOwnershipCategory.JOINT]["total_bank_value"], Decimal("5000.00"))
        self.assertEqual(groups[AssetOwnershipCategory.XIMO]["total_bank_value"], Decimal("21000.00"))
        self.assertEqual(groups[AssetOwnershipCategory.XIMO]["annual_income"], Decimal("225.00"))
        self.assertEqual(groups[AssetOwnershipCategory.MONICA]["cards_count"], 1)

    def test_build_card_spending_overview_separates_card_expenses_from_account_flow(self):
        account_statement = BankStatementImport.objects.create(
            source_filename="cuenta-mar.xls",
            source_file="banking/statements/cuenta-mar.xls",
            file_checksum="dashboard-account-mar",
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
            account_label="Cuenta 1234",
            period_end="2026-03-31",
            total_income=Decimal("2000.00"),
            total_expenses=Decimal("500.00"),
            closing_balance=Decimal("4500.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        card_statement = BankStatementImport.objects.create(
            source_filename="visa-mar.xls",
            source_file="banking/statements/visa-mar.xls",
            file_checksum="dashboard-card-mar",
            statement_kind=BankStatementImport.StatementKind.CARD,
            account_label="Visa Monica",
            period_end="2026-03-31",
            total_income=Decimal("50.00"),
            total_expenses=Decimal("700.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        BankMovement.objects.create(
            statement_import=account_statement,
            booking_date="2026-03-10",
            concept="NOMINA",
            normalized_concept="NOMINA",
            amount=Decimal("2000.00"),
            movement_group=BankMovement.MovementGroup.INCOME,
            concept_bucket="Nomina",
        )
        BankMovement.objects.create(
            statement_import=account_statement,
            booking_date="2026-03-11",
            concept="PAGO BIZUM",
            normalized_concept="PAGO BIZUM",
            amount=Decimal("-500.00"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Bizum",
        )
        BankMovement.objects.create(
            statement_import=card_statement,
            booking_date="2026-03-12",
            concept="MERCADONA",
            normalized_concept="MERCADONA",
            amount=Decimal("-650.00"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Supermercado",
        )
        BankMovement.objects.create(
            statement_import=card_statement,
            booking_date="2026-03-13",
            concept="DEVOLUCION AMAZON",
            normalized_concept="DEVOLUCION AMAZON",
            amount=Decimal("50.00"),
            movement_group=BankMovement.MovementGroup.INCOME,
            concept_bucket="Devoluciones y abonos",
        )

        dashboard = build_banking_dashboard()
        card_overview = build_card_spending_overview()

        self.assertEqual(dashboard["statement_summary"]["total_expenses"], Decimal("500.00"))
        self.assertEqual(dashboard["card_summary"]["total_spent"], Decimal("700.00"))
        self.assertEqual(dashboard["card_summary"]["total_refunds"], Decimal("50.00"))
        self.assertEqual(dashboard["tracked_cards"][0]["card_name"], "Visa Monica")
        self.assertEqual(card_overview["monthly_summaries"][0]["net_spent"], Decimal("600.00"))
        self.assertEqual(card_overview["expense_matrix"][0]["concept"], "Supermercado")

    def test_build_dashboard_tracks_statement_continuity_gaps_and_overlaps(self):
        jan_statement = BankStatementImport.objects.create(
            source_filename="ene.xls",
            source_file="banking/statements/ene.xls",
            file_checksum="continuity-ene",
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
            account_label="Cuenta 1234",
            period_start="2026-01-01",
            period_end="2026-01-31",
            closing_balance=Decimal("1000.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        feb_overlap_statement = BankStatementImport.objects.create(
            source_filename="feb-solape.xls",
            source_file="banking/statements/feb-solape.xls",
            file_checksum="continuity-feb-overlap",
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
            account_label="Cuenta 1234",
            period_start="2026-01-25",
            period_end="2026-02-20",
            closing_balance=Decimal("800.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        apr_gap_statement = BankStatementImport.objects.create(
            source_filename="abr-hueco.xls",
            source_file="banking/statements/abr-hueco.xls",
            file_checksum="continuity-abr-gap",
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
            account_label="Cuenta 1234",
            period_start="2026-04-01",
            period_end="2026-04-30",
            closing_balance=Decimal("1300.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        for statement, booking_date, amount, balance in (
            (jan_statement, "2026-01-31", Decimal("1000.00"), Decimal("1000.00")),
            (feb_overlap_statement, "2026-02-20", Decimal("-200.00"), Decimal("800.00")),
            (apr_gap_statement, "2026-04-30", Decimal("500.00"), Decimal("1300.00")),
        ):
            BankMovement.objects.create(
                statement_import=statement,
                booking_date=booking_date,
                concept="MOVIMIENTO",
                normalized_concept="MOVIMIENTO",
                amount=amount,
                balance=balance,
                movement_group=BankMovement.MovementGroup.INCOME if amount >= 0 else BankMovement.MovementGroup.EXPENSE,
                concept_bucket="Nomina" if amount >= 0 else "Otros gastos",
            )

        dashboard = build_banking_dashboard()

        self.assertEqual(dashboard["continuity_summary"]["groups_count"], 1)
        self.assertEqual(dashboard["continuity_summary"]["groups_with_gap_count"], 1)
        self.assertEqual(dashboard["continuity_summary"]["groups_with_overlap_count"], 1)
        self.assertEqual(dashboard["continuity_groups"][0]["status_label"], "Huecos y solapes")
        self.assertTrue(dashboard["tracked_accounts"][0]["continuity_has_issues"])
        self.assertTrue(dashboard["tracked_accounts"][0]["continuity_note"])

    def test_build_dashboard_adds_annual_analysis_and_bank_grouping(self):
        jan_statement = BankStatementImport.objects.create(
            source_filename="ene.xls",
            source_file="banking/statements/ene.xls",
            file_checksum="annual-ene",
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
            institution="Banco Sabadell",
            account_label="Cuenta nomina",
            period_start="2026-01-01",
            period_end="2026-01-31",
            total_income=Decimal("3000.00"),
            total_expenses=Decimal("800.00"),
            total_pension_contributions=Decimal("100.00"),
            total_dividends=Decimal("50.00"),
            closing_balance=Decimal("4200.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        feb_statement = BankStatementImport.objects.create(
            source_filename="feb.xls",
            source_file="banking/statements/feb.xls",
            file_checksum="annual-feb",
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
            institution="Banco Sabadell",
            account_label="Cuenta nomina",
            period_start="2026-02-01",
            period_end="2026-02-28",
            total_income=Decimal("3200.00"),
            total_expenses=Decimal("650.00"),
            total_pension_contributions=Decimal("0.00"),
            total_dividends=Decimal("0.00"),
            closing_balance=Decimal("5000.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        card_statement = BankStatementImport.objects.create(
            source_filename="visa-feb.xls",
            source_file="banking/statements/visa-feb.xls",
            file_checksum="annual-card-feb",
            statement_kind=BankStatementImport.StatementKind.CARD,
            institution="Banco Sabadell",
            account_label="Visa hogar",
            period_start="2026-02-01",
            period_end="2026-02-28",
            total_income=Decimal("50.00"),
            total_expenses=Decimal("400.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        for statement_import, booking_date, concept, amount, movement_group, concept_bucket in (
            (jan_statement, "2026-01-05", "NOMINA", Decimal("3000.00"), BankMovement.MovementGroup.INCOME, "Nomina"),
            (
                jan_statement,
                "2026-01-09",
                "COBRO DIVIDENDO IBERDROLA",
                Decimal("50.00"),
                BankMovement.MovementGroup.DIVIDEND,
                "Dividendos de acciones",
            ),
            (
                jan_statement,
                "2026-01-15",
                "MERCADONA",
                Decimal("-800.00"),
                BankMovement.MovementGroup.EXPENSE,
                "Supermercado",
            ),
            (
                jan_statement,
                "2026-01-25",
                "APORTACION PERIODICA POL.",
                Decimal("-100.00"),
                BankMovement.MovementGroup.PENSION,
                "Aportaciones a planes",
            ),
            (feb_statement, "2026-02-05", "NOMINA", Decimal("3200.00"), BankMovement.MovementGroup.INCOME, "Nomina"),
            (
                feb_statement,
                "2026-02-10",
                "LIQUIDACION TARJETA VISA",
                Decimal("-450.00"),
                BankMovement.MovementGroup.EXPENSE,
                "Liquidacion de tarjeta",
            ),
            (
                feb_statement,
                "2026-02-14",
                "ALQUILER",
                Decimal("-200.00"),
                BankMovement.MovementGroup.EXPENSE,
                "Otros gastos",
            ),
            (
                card_statement,
                "2026-02-12",
                "AMAZON",
                Decimal("-400.00"),
                BankMovement.MovementGroup.EXPENSE,
                "Compras online",
            ),
            (
                card_statement,
                "2026-02-17",
                "DEVOLUCION AMAZON",
                Decimal("50.00"),
                BankMovement.MovementGroup.INCOME,
                "Devoluciones y abonos",
            ),
        ):
            BankMovement.objects.create(
                statement_import=statement_import,
                booking_date=booking_date,
                concept=concept,
                normalized_concept=concept,
                amount=amount,
                movement_group=movement_group,
                concept_bucket=concept_bucket,
            )

        dashboard = build_banking_dashboard()
        current_year = dashboard["annual_overview"]["current_year"]
        institution = dashboard["institution_overview"]["institutions"][0]

        self.assertEqual(current_year["year"], 2026)
        self.assertEqual(current_year["gross_inflows_total"], Decimal("6250.00"))
        self.assertEqual(current_year["household_expenses_total"], Decimal("1350.00"))
        self.assertEqual(current_year["average_monthly_savings"], Decimal("2450.00"))
        self.assertEqual(current_year["net_value_flow_total"], Decimal("4800.00"))
        self.assertEqual(current_year["months"][0]["label"], "2026-01")
        self.assertEqual(current_year["months"][1]["cumulative_net_value_flow"], Decimal("4800.00"))
        self.assertEqual(institution["institution"], "Banco Sabadell")
        self.assertEqual(institution["accounts_count"], 1)
        self.assertEqual(institution["cards_count"], 1)
        self.assertEqual(institution["visible_balance"], Decimal("5000.00"))
        self.assertEqual(institution["status_label"], "Cobertura amplia")

    def test_build_dashboard_reconciles_card_settlements_without_double_counting_spend(self):
        account_statement = BankStatementImport.objects.create(
            source_filename="cuenta-abr.xls",
            source_file="banking/statements/cuenta-abr.xls",
            file_checksum="reconciled-account-abr",
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
            account_label="Cuenta 1234",
            period_start="2026-04-01",
            period_end="2026-04-30",
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        card_statement = BankStatementImport.objects.create(
            source_filename="visa-abr.xls",
            source_file="banking/statements/visa-abr.xls",
            file_checksum="reconciled-card-abr",
            statement_kind=BankStatementImport.StatementKind.CARD,
            account_label="Visa Monica",
            period_start="2026-04-01",
            period_end="2026-04-30",
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        BankMovement.objects.create(
            statement_import=account_statement,
            booking_date="2026-04-05",
            concept="LIQUIDACION TARJETA VISA",
            normalized_concept="LIQUIDACION TARJETA VISA",
            amount=Decimal("-700.00"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Liquidacion de tarjeta",
        )
        BankMovement.objects.create(
            statement_import=account_statement,
            booking_date="2026-04-12",
            concept="MERCADONA",
            normalized_concept="MERCADONA",
            amount=Decimal("-120.00"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Supermercado",
        )
        BankMovement.objects.create(
            statement_import=card_statement,
            booking_date="2026-04-08",
            concept="AMAZON",
            normalized_concept="AMAZON",
            amount=Decimal("-650.00"),
            movement_group=BankMovement.MovementGroup.EXPENSE,
            concept_bucket="Compras online",
        )
        BankMovement.objects.create(
            statement_import=card_statement,
            booking_date="2026-04-10",
            concept="DEVOLUCION AMAZON",
            normalized_concept="DEVOLUCION AMAZON",
            amount=Decimal("50.00"),
            movement_group=BankMovement.MovementGroup.INCOME,
            concept_bucket="Devoluciones y abonos",
        )

        dashboard = build_banking_dashboard()

        self.assertEqual(dashboard["reconciled_summary"]["card_settlements_total"], Decimal("700.00"))
        self.assertEqual(dashboard["reconciled_summary"]["cash_account_expenses_total"], Decimal("120.00"))
        self.assertEqual(dashboard["reconciled_summary"]["card_spending_total"], Decimal("650.00"))
        self.assertEqual(dashboard["reconciled_summary"]["card_refunds_total"], Decimal("50.00"))
        self.assertEqual(dashboard["reconciled_summary"]["household_expenses_total"], Decimal("720.00"))
        self.assertEqual(dashboard["reconciled_monthly_summaries"][0]["household_expenses"], Decimal("720.00"))
        self.assertEqual(dashboard["reconciled_summary"]["top_family_label"], "Compras y hogar")
        self.assertEqual(dashboard["reconciled_expense_matrix"][0]["concept"], "Compras y hogar")
        self.assertEqual(dashboard["reconciled_expense_matrix"][0]["total"], Decimal("600.00"))
        self.assertEqual(dashboard["reconciled_expense_matrix"][1]["concept"], "Alimentacion")

    def test_build_dashboard_groups_reconciled_expenses_into_families(self):
        statement = BankStatementImport.objects.create(
            source_filename="familias-may.xls",
            source_file="banking/statements/familias-may.xls",
            file_checksum="familias-may",
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
            account_label="Cuenta hogar",
            period_start="2026-05-01",
            period_end="2026-05-31",
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        for booking_date, concept, amount, concept_bucket in (
            ("2026-05-04", "MERCADONA ONDA", Decimal("-120.00"), "Supermercado"),
            ("2026-05-08", "RESIDENCIA MEDITERRANEO MARC", Decimal("-850.00"), "Otros gastos"),
            ("2026-05-12", "NETFLIX", Decimal("-19.99"), "Suscripciones"),
        ):
            BankMovement.objects.create(
                statement_import=statement,
                booking_date=booking_date,
                concept=concept,
                normalized_concept=concept,
                amount=amount,
                movement_group=BankMovement.MovementGroup.EXPENSE,
                concept_bucket=concept_bucket,
            )

        dashboard = build_banking_dashboard()
        family_rows = {row["concept"]: row for row in dashboard["reconciled_expense_matrix"]}

        self.assertEqual(dashboard["reconciled_summary"]["household_expenses_total"], Decimal("989.99"))
        self.assertEqual(dashboard["reconciled_summary"]["top_family_label"], "Estudios y residencia")
        self.assertEqual(family_rows["Estudios y residencia"]["total"], Decimal("850.00"))
        self.assertEqual(family_rows["Alimentacion"]["total"], Decimal("120.00"))
        self.assertEqual(family_rows["Ocio y digital"]["total"], Decimal("19.99"))

    def test_bank_investment_position_uses_current_value_when_cost_basis_missing(self):
        position = BankInvestmentPosition.objects.create(
            institution="Banco Sabadell",
            product_name="Cuenta Ahorro 5, CIALP",
            product_type=BankInvestmentPosition.ProductType.SAVINGS_PLAN,
            current_value=Decimal("5045.51"),
        )
        self.assertEqual(position.invested_amount, Decimal("5045.51"))

    def test_import_monica_bank_positions_creates_investment_funds_with_cost_basis(self):
        call_command("import_monica_bank_positions")

        positions = BankInvestmentPosition.objects.filter(
            ownership_category=AssetOwnershipCategory.MONICA,
            institution="Ibercaja",
        ).order_by("product_name")

        self.assertEqual(positions.count(), len(MONICA_BANK_INVESTMENT_POSITIONS))
        first_position = positions.first()
        self.assertIsNotNone(first_position)
        self.assertEqual(first_position.product_type, BankInvestmentPosition.ProductType.INVESTMENT_FUND)
        self.assertEqual(first_position.price_date, date(2026, 4, 16))
        self.assertIn("ISIN:", first_position.notes)
        self.assertIn("Referencia cartera Ibercaja: 4414450943", first_position.notes)

        sabadell_like = positions.get(product_name="Ibercaja Renta Fija 2027, FI Clase A")
        self.assertEqual(sabadell_like.current_value, Decimal("76536.83"))
        self.assertEqual(sabadell_like.invested_amount_override, Decimal("73978.94"))
        self.assertEqual(sabadell_like.units, Decimal("11739.7653"))

    def test_parse_statement_file_from_html_xls_export(self):
        html = """
<html>
  <body>
    <table>
      <tr><td>Cuenta: ES12 3456 7890 1234</td></tr>
      <tr><td>Titular: Ada Lovelace</td></tr>
      <tr><td>Divisa: EUR</td></tr>
      <tr><td>Desde 01/04/2026 hasta 05/04/2026</td></tr>
      <tr>
        <td>F. Operativa</td>
        <td>Concepto</td>
        <td>F. Valor</td>
        <td>Importe</td>
        <td>Saldo</td>
        <td>Referencia 1</td>
        <td>Referencia 2</td>
      </tr>
      <tr>
        <td>05/04/2026</td>
        <td>NOMINA ABRIL</td>
        <td>05/04/2026</td>
        <td>1.200,50</td>
        <td>3.400,75</td>
        <td>ABC</td>
        <td>DEF</td>
      </tr>
      <tr>
        <td>04/04/2026</td>
        <td>MERCADONA</td>
        <td>04/04/2026</td>
        <td>-45,10</td>
        <td>2.200,25</td>
        <td></td>
        <td></td>
      </tr>
    </table>
  </body>
</html>
"""
        with tempfile.TemporaryDirectory() as temp_dir:
            statement_path = Path(temp_dir) / "statement.xls"
            statement_path.write_text(html, encoding="utf-8")

            parsed = parse_statement_file(str(statement_path))

        self.assertEqual(parsed["metadata"]["iban"], "ES12 3456 7890 1234")
        self.assertEqual(parsed["metadata"]["holder_name"], "Ada Lovelace")
        self.assertEqual(parsed["metadata"]["period_start"].isoformat(), "2026-04-01")
        self.assertEqual(parsed["metadata"]["period_end"].isoformat(), "2026-04-05")
        self.assertEqual(parsed["metadata"]["opening_balance"], Decimal("2245.35"))
        self.assertEqual(parsed["metadata"]["closing_balance"], Decimal("3400.75"))
        self.assertEqual(len(parsed["movements"]), 2)
        self.assertEqual(parsed["movements"][0].amount, Decimal("1200.50"))
        self.assertEqual(parsed["movements"][1].amount, Decimal("-45.10"))

    def test_parse_statement_file_accepts_flexible_header_names(self):
        html = """
<html>
  <body>
    <table>
      <tr><td>IBAN: ES98 7654 3210 9876</td></tr>
      <tr><td>Titulares: Ximo y Monica</td></tr>
      <tr><td>Moneda: EUR</td></tr>
      <tr><td>Del 01/04/2026 al 05/04/2026</td></tr>
      <tr>
        <td>Fecha operacion</td>
        <td>Descripcion</td>
        <td>Fecha valor</td>
        <td>Importe EUR</td>
        <td>Saldo contable</td>
      </tr>
      <tr>
        <td>05/04/2026</td>
        <td>NOMINA ABRIL</td>
        <td>05/04/2026</td>
        <td>1.200,50</td>
        <td>3.400,75</td>
      </tr>
      <tr>
        <td>04/04/2026</td>
        <td>MERCADONA</td>
        <td>04/04/2026</td>
        <td>-45,10</td>
        <td>2.200,25</td>
      </tr>
    </table>
  </body>
</html>
"""
        with tempfile.TemporaryDirectory() as temp_dir:
            statement_path = Path(temp_dir) / "statement-flex.xls"
            statement_path.write_text(html, encoding="utf-8")

            parsed = parse_statement_file(str(statement_path))

        self.assertEqual(parsed["metadata"]["iban"], "ES98 7654 3210 9876")
        self.assertEqual(parsed["metadata"]["holder_name"], "Ximo y Monica")
        self.assertEqual(parsed["metadata"]["period_start"].isoformat(), "2026-04-01")
        self.assertEqual(parsed["metadata"]["period_end"].isoformat(), "2026-04-05")
        self.assertEqual(parsed["movements"][0].amount, Decimal("1200.50"))
        self.assertEqual(parsed["movements"][1].amount, Decimal("-45.10"))

    def test_parse_statement_file_accepts_debit_and_credit_columns(self):
        html = """
<html>
  <body>
    <table>
      <tr><td>Cuenta: ES12 3456 7890 1234</td></tr>
      <tr><td>Titular: Monica</td></tr>
      <tr><td>Divisa: EUR</td></tr>
      <tr><td>Desde 01/04/2026 hasta 05/04/2026</td></tr>
      <tr>
        <td>Fecha</td>
        <td>Descripcion</td>
        <td>Cargo</td>
        <td>Abono</td>
        <td>Saldo</td>
      </tr>
      <tr>
        <td>05/04/2026</td>
        <td>NOMINA ABRIL</td>
        <td></td>
        <td>1.200,50</td>
        <td>3.400,75</td>
      </tr>
      <tr>
        <td>04/04/2026</td>
        <td>MERCADONA</td>
        <td>45,10</td>
        <td></td>
        <td>2.200,25</td>
      </tr>
    </table>
  </body>
</html>
"""
        with tempfile.TemporaryDirectory() as temp_dir:
            statement_path = Path(temp_dir) / "statement-debit-credit.xls"
            statement_path.write_text(html, encoding="utf-8")

            parsed = parse_statement_file(str(statement_path))

        self.assertEqual(len(parsed["movements"]), 2)
        self.assertEqual(parsed["movements"][0].amount, Decimal("1200.50"))
        self.assertEqual(parsed["movements"][1].amount, Decimal("-45.10"))
        self.assertEqual(parsed["metadata"]["holder_name"], "Monica")

    def test_parse_statement_file_accepts_card_layout_with_short_dates(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            statement_path = Path(temp_dir) / "05042026_5402________3026.xls"
            statement_path.write_text(self._build_card_statement_html(), encoding="utf-8")

            parsed = parse_statement_file(
                str(statement_path),
                statement_kind=BankStatementImport.StatementKind.CARD,
            )

        self.assertEqual(parsed["metadata"]["account_label"], "BS CARD MASTERCARD 3026")
        self.assertEqual(parsed["metadata"]["holder_name"], "PIQUER MARTI ,JOAQUIN")
        self.assertEqual(parsed["metadata"]["period_start"].isoformat(), "2026-03-30")
        self.assertEqual(parsed["metadata"]["period_end"].isoformat(), "2026-04-05")
        self.assertEqual(len(parsed["movements"]), 4)
        self.assertEqual(parsed["movements"][0].amount, Decimal("-19.99"))
        self.assertEqual(parsed["movements"][0].reference_1, "LUXEMBOURG")
        self.assertEqual(parsed["movements"][0].reference_2, "AUT")
        self.assertEqual(parsed["movements"][-1].amount, Decimal("-55.07"))

    def test_parse_statement_file_accepts_card_xlsx_with_excel_dates_and_explicit_period(self):
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Tarjeta:", "5402________3026", "BS CARD MASTERCARD"])
        worksheet.append(["Titular tarjeta", "PIQUER MARTI ,JOAQUIN"])
        worksheet.append(["Periodo analizado", date(2026, 4, 1), date(2026, 4, 30)])
        worksheet.append(["F. Operativa", "Concepto", "F. Valor", "Importe"])
        worksheet.append([date(2026, 4, 5), "AMAZON", date(2026, 4, 5), -19.99])
        worksheet.append([date(2026, 4, 2), "MERCADONA", date(2026, 4, 2), -55.07])

        payload = BytesIO()
        workbook.save(payload)
        workbook.close()

        with tempfile.TemporaryDirectory() as temp_dir:
            statement_path = Path(temp_dir) / "visa.xlsx"
            statement_path.write_bytes(payload.getvalue())

            parsed = parse_statement_file(
                str(statement_path),
                statement_kind=BankStatementImport.StatementKind.CARD,
            )

        self.assertEqual(parsed["metadata"]["account_label"], "BS CARD MASTERCARD 3026")
        self.assertEqual(parsed["metadata"]["holder_name"], "PIQUER MARTI ,JOAQUIN")
        self.assertEqual(parsed["metadata"]["period_start"].isoformat(), "2026-04-01")
        self.assertEqual(parsed["metadata"]["period_end"].isoformat(), "2026-04-30")
        self.assertEqual(len(parsed["movements"]), 2)
        self.assertEqual(parsed["movements"][0].amount, Decimal("-19.99"))
        self.assertEqual(parsed["movements"][1].amount, Decimal("-55.07"))

    def test_parse_statement_file_prefers_explicit_card_period_over_movement_range(self):
        html = """
<html>
  <body>
    <table>
      <tr><td>Targeta:</td><td>5402________3026</td><td>BS CARD MASTERCARD</td></tr>
      <tr><td>Titular targeta</td><td>PIQUER MARTI ,JOAQUIN</td></tr>
      <tr><td>Periodo analizado</td><td>01/03/2026</td><td>31/03/2026</td></tr>
      <tr><td>MOVIMIENTOS DE DEBITO</td></tr>
      <tr>
        <td>DATA</td>
        <td>CONCEPTE</td>
        <td>LOCALITAT</td>
        <td>IMPORT</td>
      </tr>
      <tr>
        <td>31/03</td>
        <td>AMAZON</td>
        <td>LUXEMBOURG</td>
        <td>7,53</td>
      </tr>
      <tr>
        <td>30/03</td>
        <td>MERCADONA</td>
        <td>ONDA</td>
        <td>55,07</td>
      </tr>
    </table>
  </body>
</html>
"""
        with tempfile.TemporaryDirectory() as temp_dir:
            statement_path = Path(temp_dir) / "periodo-tarjeta.xls"
            statement_path.write_text(html, encoding="utf-8")

            parsed = parse_statement_file(
                str(statement_path),
                statement_kind=BankStatementImport.StatementKind.CARD,
            )

        self.assertEqual(parsed["metadata"]["period_start"].isoformat(), "2026-03-01")
        self.assertEqual(parsed["metadata"]["period_end"].isoformat(), "2026-03-31")
        self.assertEqual(len(parsed["movements"]), 2)
        self.assertEqual(parsed["movements"][0].amount, Decimal("-7.53"))
        self.assertEqual(parsed["movements"][1].amount, Decimal("-55.07"))

    def test_parse_statement_file_accepts_card_layout_with_amount_shifted_right(self):
        html = """
<html>
  <body>
    <table>
      <tr><td>Saldos y movimientos</td></tr>
      <tr><td>Tarjeta:</td><td>5402________3026</td><td>BS CARD MASTERCARD</td></tr>
      <tr><td>Titular tarjeta</td><td>PIQUER MARTI ,JOAQUIN</td></tr>
      <tr><td>MOVIMIENTOS DE DEBITO</td></tr>
      <tr>
        <td>FECHA</td>
        <td>CONCEPTO</td>
        <td>LOCALIDAD</td>
        <td>IMPORTE</td>
        <td></td>
        <td></td>
      </tr>
      <tr>
        <td>28/01</td>
        <td>AMAZON* 2F5Q77O95</td>
        <td>LUXEMBOURG</td>
        <td></td>
        <td>54,44</td>
        <td>EUR</td>
      </tr>
      <tr>
        <td>12/01</td>
        <td>AMAZON* ZC8UT0VX4</td>
        <td>LUXEMBOURG</td>
        <td></td>
        <td>-22,99</td>
        <td>EUR</td>
      </tr>
    </table>
  </body>
</html>
"""
        with tempfile.TemporaryDirectory() as temp_dir:
            statement_path = Path(temp_dir) / "06042026_5402________3026.xls"
            statement_path.write_text(html, encoding="utf-8")

            parsed = parse_statement_file(
                str(statement_path),
                statement_kind=BankStatementImport.StatementKind.CARD,
            )

        self.assertEqual(parsed["metadata"]["account_label"], "BS CARD MASTERCARD 3026")
        self.assertEqual(parsed["metadata"]["period_start"].isoformat(), "2026-01-12")
        self.assertEqual(parsed["metadata"]["period_end"].isoformat(), "2026-01-28")
        self.assertEqual(len(parsed["movements"]), 2)
        self.assertEqual(parsed["movements"][0].amount, Decimal("-54.44"))
        self.assertEqual(parsed["movements"][1].amount, Decimal("-22.99"))

    def test_load_rows_from_xls_reads_legacy_workbook_with_xlrd(self):
        import xlrd
        from xlrd.xldate import xldate_from_date_tuple

        class FakeSheet:
            def __init__(self, values, cell_types):
                self._values = values
                self._cell_types = cell_types
                self.nrows = len(values)
                self.ncols = len(values[0])

            def cell_value(self, row_index, column_index):
                return self._values[row_index][column_index]

            def cell_type(self, row_index, column_index):
                return self._cell_types[row_index][column_index]

        class FakeWorkbook:
            datemode = 0

            def __init__(self, sheet):
                self._sheet = sheet
                self.released = False

            def sheet_by_index(self, index):
                return self._sheet

            def release_resources(self):
                self.released = True

        booking_date = xldate_from_date_tuple((2026, 4, 5), 0)
        value_date = xldate_from_date_tuple((2026, 4, 5), 0)
        values = [
            ["F. Operativa", "Concepto", "F. Valor", "Importe", "Saldo"],
            [booking_date, "NOMINA", value_date, 3126.77, 10000.55],
        ]
        cell_types = [
            [xlrd.XL_CELL_TEXT] * 5,
            [xlrd.XL_CELL_DATE, xlrd.XL_CELL_TEXT, xlrd.XL_CELL_DATE, xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_NUMBER],
        ]
        workbook = FakeWorkbook(FakeSheet(values, cell_types))

        with tempfile.TemporaryDirectory() as temp_dir:
            statement_path = Path(temp_dir) / "legacy.xls"
            statement_path.write_bytes(b"legacy-xls")

            with patch("xlrd.open_workbook", return_value=workbook):
                rows = load_rows_from_xls(str(statement_path))

        self.assertEqual(
            rows,
            [
                ["F. Operativa", "Concepto", "F. Valor", "Importe", "Saldo"],
                ["05/04/2026", "NOMINA", "05/04/2026", "3126.77", "10000.55"],
            ],
        )
        self.assertTrue(workbook.released)

    def test_import_statement_infers_ownership_from_holder_name(self):
        with tempfile.TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media):
                statement = BankStatementImport.objects.create(
                    source_filename="ximo.xls",
                    source_file=SimpleUploadedFile(
                        "ximo.xls",
                        self._build_statement_html(holder_name="Ximo").encode("utf-8"),
                        content_type="application/vnd.ms-excel",
                    ),
                    file_checksum="checksum-ximo-import",
                )

                import_statement(statement)
                statement.refresh_from_db()

        self.assertEqual(statement.ownership_category, AssetOwnershipCategory.XIMO)

    def test_import_statement_reuses_existing_account_ownership(self):
        with tempfile.TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media):
                BankStatementImport.objects.create(
                    source_filename="anterior.xls",
                    source_file=SimpleUploadedFile(
                        "anterior.xls",
                        b"anterior",
                        content_type="application/vnd.ms-excel",
                    ),
                    file_checksum="checksum-anterior",
                    ownership_category=AssetOwnershipCategory.MONICA,
                    iban="ES12 3456 7890 1234",
                    account_label="Cuenta 1234",
                    import_status=BankStatementImport.ImportStatus.IMPORTED,
                )
                statement = BankStatementImport.objects.create(
                    source_filename="nuevo.xls",
                    source_file=SimpleUploadedFile(
                        "nuevo.xls",
                        self._build_statement_html(holder_name="").encode("utf-8"),
                        content_type="application/vnd.ms-excel",
                    ),
                    file_checksum="checksum-nuevo",
                )

                import_statement(statement)
                statement.refresh_from_db()

        self.assertEqual(statement.ownership_category, AssetOwnershipCategory.MONICA)


class BankingImportDeletionTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="banking-owner",
            password="secret123",
        )
        self.client.force_login(self.user)

    def _create_statement_with_file(self):
        uploaded_file = SimpleUploadedFile(
            "sample.xls",
            b"fake-xls-content",
            content_type="application/vnd.ms-excel",
        )
        statement = BankStatementImport.objects.create(
            source_filename="sample.xls",
            source_file=uploaded_file,
            file_checksum="checksum-sample",
            account_label="Cuenta prueba",
            period_end="2026-03-31",
            total_income=Decimal("100.00"),
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        BankMovement.objects.create(
            statement_import=statement,
            booking_date="2026-03-20",
            concept="NOMINA",
            normalized_concept="NOMINA",
            amount=Decimal("100.00"),
            movement_group=BankMovement.MovementGroup.INCOME,
            concept_bucket="Nomina",
        )
        return statement

    def test_delete_single_statement_removes_record_movements_and_file(self):
        with tempfile.TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media):
                statement = self._create_statement_with_file()
                file_path = statement.source_file.path

                response = self.client.post(
                    "/banking/",
                    {
                        "action": "delete_statement",
                        "statement_id": statement.id,
                    },
                )

                self.assertEqual(response.status_code, 302)
                self.assertFalse(BankStatementImport.objects.filter(pk=statement.id).exists())
                self.assertEqual(BankMovement.objects.count(), 0)
                self.assertFalse(os.path.exists(file_path))

    def test_delete_all_statements_is_blocked_to_avoid_bulk_errors(self):
        with tempfile.TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media):
                first = self._create_statement_with_file()
                second = BankStatementImport.objects.create(
                    source_filename="sample-2.xls",
                    source_file=SimpleUploadedFile(
                        "sample-2.xls",
                        b"fake-xls-content-2",
                        content_type="application/vnd.ms-excel",
                    ),
                    file_checksum="checksum-sample-2",
                    account_label="Cuenta prueba 2",
                    period_end="2026-02-28",
                    total_expenses=Decimal("50.00"),
                    import_status=BankStatementImport.ImportStatus.IMPORTED,
                )
                BankMovement.objects.create(
                    statement_import=second,
                    booking_date="2026-02-20",
                    concept="MERCADONA",
                    normalized_concept="MERCADONA",
                    amount=Decimal("-50.00"),
                    movement_group=BankMovement.MovementGroup.EXPENSE,
                    concept_bucket="Supermercado",
                )
                first_path = first.source_file.path
                second_path = second.source_file.path

                response = self.client.post(
                    "/banking/",
                    {"action": "delete_all_statements"},
                )

                self.assertEqual(response.status_code, 302)
                self.assertEqual(BankStatementImport.objects.count(), 2)
                self.assertEqual(BankMovement.objects.count(), 2)
                self.assertTrue(os.path.exists(first_path))
                self.assertTrue(os.path.exists(second_path))
                self.assertTrue(
                    BankStatementImport.objects.filter(file_checksum__in=["checksum-sample", "checksum-sample-2"]).exists()
                )


class BankingViewTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="banking-editor",
            password="secret123",
        )
        self.client.force_login(self.user)

    def _card_statement_html(self, holder_name: str = "Monica") -> str:
        return f"""
<html>
  <body>
    <table>
      <tr><td>Cuenta: VISA MONICA</td></tr>
      <tr><td>Titular: {holder_name}</td></tr>
      <tr><td>Divisa: EUR</td></tr>
      <tr><td>Desde 01/04/2026 hasta 05/04/2026</td></tr>
      <tr>
        <td>F. Operativa</td>
        <td>Concepto</td>
        <td>F. Valor</td>
        <td>Importe</td>
      </tr>
      <tr>
        <td>05/04/2026</td>
        <td>MERCADONA</td>
        <td>05/04/2026</td>
        <td>-120,50</td>
      </tr>
    </table>
  </body>
</html>
"""

    def _legacy_card_statement_html(self) -> str:
        return """
<html>
  <body>
    <table>
      <tr><td>Saldos i moviments</td></tr>
      <tr><td>Contracte</td><td>004273514883</td><td>Compte relacionat</td><td>0081-0278-14-0006302439</td></tr>
      <tr><td>Titular</td><td>PIQUER MARTI ,JOAQUIN</td></tr>
      <tr><td>Targeta:</td><td>5402________3026</td><td>BS CARD MASTERCARD</td></tr>
      <tr><td>Titular targeta</td><td>PIQUER MARTI ,JOAQUIN</td></tr>
      <tr><td>MOVIMIENTOS DE DEBITO</td></tr>
      <tr>
        <td>DATA</td>
        <td>CONCEPTE</td>
        <td>LOCALITAT</td>
        <td>SIT. MOV.</td>
        <td>IMPORT</td>
      </tr>
      <tr>
        <td>05/04</td>
        <td>WWW.AMAZON</td>
        <td>LUXEMBOURG</td>
        <td>AUT</td>
        <td>19,99</td>
      </tr>
      <tr>
        <td>03/04</td>
        <td>APPLE.COM/BILL</td>
        <td>CORK</td>
        <td>AUT</td>
        <td>2,99</td>
      </tr>
      <tr>
        <td>DATA</td>
        <td>CONCEPTE</td>
        <td>LOCALITAT</td>
        <td>IMPORT</td>
      </tr>
      <tr>
        <td>31/03</td>
        <td>WWW.AMAZON* NB99W7NE4</td>
        <td>LUXEMBOURG</td>
        <td>7,53</td>
      </tr>
      <tr>
        <td>30/03</td>
        <td>MERCADONA AVDA CONSTITUCI</td>
        <td>ONDA</td>
        <td>55,07</td>
      </tr>
    </table>
  </body>
</html>
"""

    def test_bank_page_shows_robot_automation_section(self):
        response = self.client.get(reverse("banking:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Asistente local")
        self.assertContains(response, "Enlazar una cuenta o tarjeta")
        self.assertContains(response, "Guardar banco en este ordenador")
        self.assertContains(response, "Actualizar todas las conexiones")
        self.assertContains(response, "Operativa de importacion")
        self.assertContains(response, "Gasto familiar consolidado")
        self.assertNotContains(response, "Gasto mensual en tarjetas")
        self.assertNotContains(response, "Gasto de tarjetas por concepto")
        self.assertNotContains(response, "Borrar todas las importaciones")
        self.assertNotContains(response, "Open Banking")
        self.assertNotContains(response, "GoCardless")

    def test_bank_page_shows_monica_investment_funds_with_owner_and_value_date(self):
        BankInvestmentPosition.objects.create(
            ownership_category=AssetOwnershipCategory.MONICA,
            institution="Ibercaja",
            product_name="Ibercaja Renta Fija 2027, FI Clase A",
            product_type=BankInvestmentPosition.ProductType.INVESTMENT_FUND,
            invested_amount_override=Decimal("73978.94"),
            current_value=Decimal("76536.83"),
            units=Decimal("11739.7653"),
            price_date=date(2026, 4, 16),
            annual_income=Decimal("0.00"),
        )

        response = self.client.get(reverse("banking:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Productos bancarios de inversion")
        self.assertContains(response, "Monica")
        self.assertContains(response, "Fondo de inversion")
        self.assertContains(response, "2026-04-16")
        self.assertRegex(response.content.decode("utf-8"), r"76(?:[\.,]?536)[\.,]83")

    def test_local_robot_import_endpoint_imports_statement_with_login(self):
        document = SimpleUploadedFile(
            "robot-local-visa.xls",
            self._legacy_card_statement_html().encode("utf-8"),
            content_type="application/vnd.ms-excel",
        )

        response = self.client.post(
            reverse("banking:robot_local_import"),
            {
                "statement_kind": BankStatementImport.StatementKind.CARD,
                "ownership_category": AssetOwnershipCategory.XIMO,
                "institution": "Banco Sabadell",
                "account_label": "Tarjeta principal",
                "files": document,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["imported_count"], 1)
        statement = BankStatementImport.objects.get(source_filename="robot-local-visa.xls")
        self.assertEqual(statement.import_source, BankStatementImport.ImportSource.ROBOT)
        self.assertEqual(statement.statement_kind, BankStatementImport.StatementKind.CARD)
        self.assertEqual(statement.ownership_category, AssetOwnershipCategory.XIMO)

    def test_robot_installer_download_requires_login_and_returns_script(self):
        self.client.logout()

        anonymous_response = self.client.get(reverse("banking:robot_installer"))
        self.assertEqual(anonymous_response.status_code, 302)

        self.client.force_login(self.user)
        response = self.client.get(reverse("banking:robot_installer"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment;", response.headers["Content-Disposition"])
        self.assertIn("instalar_robot_bancario.ps1", response.headers["Content-Disposition"])

    @override_settings(BANK_ROBOT_IMPORT_TOKEN="robot-token")
    def test_robot_upload_endpoint_imports_statement_without_login(self):
        self.client.logout()
        document = SimpleUploadedFile(
            "robot-visa.xls",
            self._legacy_card_statement_html().encode("utf-8"),
            content_type="application/vnd.ms-excel",
        )

        response = self.client.post(
            reverse("banking:robot_upload"),
            {
                "statement_kind": BankStatementImport.StatementKind.CARD,
                "ownership_category": AssetOwnershipCategory.XIMO,
                "institution": "Banco Sabadell",
                "files": document,
            },
            HTTP_X_BANK_ROBOT_TOKEN="robot-token",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["imported_count"], 1)
        statement = BankStatementImport.objects.get(source_filename="robot-visa.xls")
        self.assertEqual(statement.import_source, BankStatementImport.ImportSource.ROBOT)
        self.assertEqual(statement.statement_kind, BankStatementImport.StatementKind.CARD)
        self.assertEqual(statement.ownership_category, AssetOwnershipCategory.XIMO)

    @override_settings(BANK_ROBOT_IMPORT_TOKEN="robot-token")
    def test_robot_upload_endpoint_rejects_invalid_token(self):
        self.client.logout()
        document = SimpleUploadedFile(
            "robot-cuenta.xls",
            self._card_statement_html(holder_name="Monica").encode("utf-8"),
            content_type="application/vnd.ms-excel",
        )

        response = self.client.post(
            reverse("banking:robot_upload"),
            {
                "statement_kind": BankStatementImport.StatementKind.CARD,
                "files": document,
            },
            HTTP_X_BANK_ROBOT_TOKEN="wrong-token",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(BankStatementImport.objects.filter(source_filename="robot-cuenta.xls").exists())

    def test_can_create_bank_account_with_selected_owner(self):
        response = self.client.post(
            reverse("banking:list"),
            {
                "action": "save_account",
                "ownership_category": AssetOwnershipCategory.MONICA,
                "institution": "Banco Sabadell",
                "account_name": "Cuenta ahorro",
                "deposited_amount": "10.000,00",
                "current_balance": "10.250,40",
                "annual_interest_income": "125,50",
                "notes": "Cuenta personal",
            },
        )

        self.assertRedirects(response, reverse("banking:list"))
        account = BankBalance.objects.get(institution="Banco Sabadell", account_name="Cuenta ahorro")
        self.assertEqual(account.ownership_category, AssetOwnershipCategory.MONICA)
        self.assertEqual(account.deposited_amount, Decimal("10000.00"))
        self.assertEqual(account.current_balance, Decimal("10250.40"))
        self.assertEqual(account.annual_interest_income, Decimal("125.50"))

    def test_can_import_card_statement_with_explicit_type(self):
        document = SimpleUploadedFile(
            "visa.xls",
            self._card_statement_html(holder_name="Monica").encode("utf-8"),
            content_type="application/vnd.ms-excel",
        )

        response = self.client.post(
            reverse("banking:list"),
            {
                "action": "import",
                "statement_kind": BankStatementImport.StatementKind.CARD,
                "files": document,
            },
        )

        self.assertRedirects(response, reverse("banking:list"))
        statement = BankStatementImport.objects.get(source_filename="visa.xls")
        self.assertEqual(statement.statement_kind, BankStatementImport.StatementKind.CARD)

    def test_can_import_legacy_card_statement_with_short_dates(self):
        document = SimpleUploadedFile(
            "05042026_5402________3026.xls",
            self._legacy_card_statement_html().encode("utf-8"),
            content_type="application/vnd.ms-excel",
        )

        response = self.client.post(
            reverse("banking:list"),
            {
                "action": "import",
                "statement_kind": BankStatementImport.StatementKind.CARD,
                "files": document,
            },
        )

        self.assertRedirects(response, reverse("banking:list"))
        statement = BankStatementImport.objects.get(source_filename="05042026_5402________3026.xls")
        self.assertEqual(statement.statement_kind, BankStatementImport.StatementKind.CARD)
        self.assertEqual(statement.account_label, "BS CARD MASTERCARD 3026")
        self.assertEqual(statement.period_end.isoformat(), "2026-04-05")
        self.assertEqual(statement.total_expenses, Decimal("85.58"))
        self.assertEqual(statement.movements.count(), 4)

    def test_can_update_bank_account_owner_from_list(self):
        account = BankBalance.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            institution="Banco Sabadell",
            account_name="Cuenta 1234",
            deposited_amount=Decimal("5000.00"),
            current_balance=Decimal("5250.00"),
            annual_interest_income=Decimal("15.00"),
        )

        response = self.client.post(
            reverse("banking:list"),
            {
                "action": "update_account_ownership",
                "account_id": account.id,
                "ownership_category": AssetOwnershipCategory.XIMO,
            },
        )

        self.assertRedirects(response, reverse("banking:list"))
        account.refresh_from_db()
        self.assertEqual(account.ownership_category, AssetOwnershipCategory.XIMO)

    def test_can_update_statement_owner_for_same_iban(self):
        first = BankStatementImport.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            source_filename="ene.xls",
            source_file=SimpleUploadedFile("ene.xls", b"ene", content_type="application/vnd.ms-excel"),
            file_checksum="statement-ene",
            iban="ES12 3456 7890 1234",
            account_label="Cuenta 1234",
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        second = BankStatementImport.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            source_filename="feb.xls",
            source_file=SimpleUploadedFile("feb.xls", b"feb", content_type="application/vnd.ms-excel"),
            file_checksum="statement-feb",
            iban="ES12 3456 7890 1234",
            account_label="Cuenta 1234",
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        response = self.client.post(
            reverse("banking:list"),
            {
                "action": "update_statement_ownership",
                "statement_id": first.id,
                "ownership_category": AssetOwnershipCategory.MONICA,
            },
        )

        self.assertRedirects(response, reverse("banking:list"))
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.ownership_category, AssetOwnershipCategory.MONICA)
        self.assertEqual(second.ownership_category, AssetOwnershipCategory.MONICA)

    def test_updating_statement_owner_syncs_matching_manual_account(self):
        account = BankBalance.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            institution="Banco Sabadell",
            account_name="Cuenta 1234",
            deposited_amount=Decimal("5000.00"),
            current_balance=Decimal("5250.00"),
            annual_interest_income=Decimal("15.00"),
        )
        statement = BankStatementImport.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            source_filename="mar.xls",
            source_file=SimpleUploadedFile("mar.xls", b"mar", content_type="application/vnd.ms-excel"),
            file_checksum="statement-sync-mar",
            iban="ES12 3456 7890 1234",
            account_label="Cuenta 1234",
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        response = self.client.post(
            reverse("banking:list"),
            {
                "action": "update_statement_ownership",
                "statement_id": statement.id,
                "ownership_category": AssetOwnershipCategory.XIMO,
            },
        )

        self.assertRedirects(response, reverse("banking:list"))
        account.refresh_from_db()
        statement.refresh_from_db()
        self.assertEqual(account.ownership_category, AssetOwnershipCategory.XIMO)
        self.assertEqual(statement.ownership_category, AssetOwnershipCategory.XIMO)

    def test_updating_card_owner_does_not_modify_manual_bank_accounts(self):
        account = BankBalance.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            institution="Banco Sabadell",
            account_name="Visa Monica",
            deposited_amount=Decimal("5000.00"),
            current_balance=Decimal("5250.00"),
            annual_interest_income=Decimal("15.00"),
        )
        statement = BankStatementImport.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            source_filename="visa.xls",
            source_file=SimpleUploadedFile("visa.xls", b"visa", content_type="application/vnd.ms-excel"),
            file_checksum="statement-card-owner",
            statement_kind=BankStatementImport.StatementKind.CARD,
            account_label="Visa Monica",
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )

        response = self.client.post(
            reverse("banking:list"),
            {
                "action": "update_statement_ownership",
                "statement_id": statement.id,
                "ownership_category": AssetOwnershipCategory.MONICA,
            },
        )

        self.assertRedirects(response, reverse("banking:list"))
        account.refresh_from_db()
        statement.refresh_from_db()
        self.assertEqual(account.ownership_category, AssetOwnershipCategory.JOINT)
        self.assertEqual(statement.ownership_category, AssetOwnershipCategory.MONICA)


class EncryptedMediaTests(TestCase):
    encryption_key = "4nTlab58x8n6qwc_cJ3mt-SN5QSDkQ6L7fL2JH57UNM="

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="secure-media-user",
            password="secret123",
        )

    def _statement_html(self):
        return """
<html>
  <body>
    <table>
      <tr><td>Cuenta: ES12 3456 7890 1234</td></tr>
      <tr><td>Titular: Monica</td></tr>
      <tr><td>Divisa: EUR</td></tr>
      <tr><td>Desde 01/04/2026 hasta 05/04/2026</td></tr>
      <tr>
        <td>F. Operativa</td>
        <td>Concepto</td>
        <td>F. Valor</td>
        <td>Importe</td>
        <td>Saldo</td>
        <td>Referencia 1</td>
        <td>Referencia 2</td>
      </tr>
      <tr>
        <td>05/04/2026</td>
        <td>NOMINA ABRIL</td>
        <td>05/04/2026</td>
        <td>1.200,50</td>
        <td>3.400,75</td>
        <td>ABC</td>
        <td>DEF</td>
      </tr>
    </table>
  </body>
</html>
"""

    def test_uploaded_statement_is_stored_encrypted_and_can_still_be_imported(self):
        with tempfile.TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media, APP_MEDIA_ENCRYPTION_KEY=self.encryption_key):
                payload = self._statement_html().encode("utf-8")
                statement = BankStatementImport.objects.create(
                    source_filename="secure.xls",
                    source_file=SimpleUploadedFile(
                        "secure.xls",
                        payload,
                        content_type="application/vnd.ms-excel",
                    ),
                    file_checksum="checksum-secure",
                )

                raw_on_disk = Path(statement.source_file.path).read_bytes()
                self.assertTrue(raw_on_disk.startswith(ENCRYPTION_MARKER))
                self.assertNotIn(b"NOMINA ABRIL", raw_on_disk)

                import_statement(statement)
                statement.refresh_from_db()

        self.assertEqual(statement.import_status, BankStatementImport.ImportStatus.IMPORTED)
        self.assertEqual(statement.ownership_category, AssetOwnershipCategory.MONICA)

    def test_secure_media_download_requires_login_and_returns_decrypted_content(self):
        with tempfile.TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media, APP_MEDIA_ENCRYPTION_KEY=self.encryption_key):
                payload = self._statement_html().encode("utf-8")
                statement = BankStatementImport.objects.create(
                    source_filename="secure.xls",
                    source_file=SimpleUploadedFile(
                        "secure.xls",
                        payload,
                        content_type="application/vnd.ms-excel",
                    ),
                    file_checksum="checksum-secure-download",
                )

                self.assertTrue(statement.source_file.url.startswith("/secure-media/"))

                anonymous_response = self.client.get(statement.source_file.url)
                self.assertRedirects(
                    anonymous_response,
                    f"{reverse('login')}?next={statement.source_file.url}",
                )

                self.client.force_login(self.user)
                response = self.client.get(statement.source_file.url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(b"".join(response.streaming_content), payload)
