from __future__ import annotations

import json
import os
import re
import subprocess
import unicodedata
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from html.parser import HTMLParser
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from portfolio.ownership import AssetOwnershipCategory

from .models import BankBalance, BankConnection, BankExternalAccount, BankMovement, BankStatementImport


ZERO = Decimal("0.00")
HEADER_ROW = ["F. Operativa", "Concepto", "F. Valor", "Importe", "Saldo"]
DEFAULT_STATEMENT_COLUMN_MAP = {
    "booking_date": 0,
    "concept": 1,
    "value_date": 2,
    "amount": 3,
    "balance": 4,
    "reference_1": 5,
    "reference_2": 6,
}
CARD_DEBIT_SECTION_LABELS = {
    "MOVIMIENTOS DE DEBITO",
    "MOVIMIENTOS DE DEBIT",
    "MOVIMENTOS DE DEBITO",
    "MOVIMENTS DE DEBIT",
}
CARD_CREDIT_SECTION_LABELS = {
    "MOVIMIENTOS DE CREDITO",
    "MOVIMIENTOS DE CREDITO",
    "MOVIMENTS DE CREDIT",
    "MOVIMENTOS DE ABONO",
    "MOVIMIENTOS DE ABONO",
}
CARD_ACCOUNT_KEYWORDS = ("CARD", "TARJETA", "TARGETA", "VISA", "MASTERCARD", "AMEX")
GOCARDLESS_API_BASE_URL = os.environ.get(
    "GOCARDLESS_BANK_DATA_BASE_URL",
    "https://bankaccountdata.gocardless.com/api/v2",
).rstrip("/")
PLAN_KEYWORDS = ("PLAN AHORRO", "PLAN DE PENSION", "PENSION", "APORTACION PERIODICA POL.")
DIVIDEND_KEYWORDS = ("DIVID", "DIVIDENDO", "COBRO DIVIDENDO", "ABONO DIVIDENDO")
RENT_INCOME_KEYWORDS = ("ALQUILER", "PISO", "ADRIAN")
TRACKED_INCOME_BUCKETS = (
    "Nomina",
    "Alquiler piso (Adrian)",
    "Dividendos de acciones",
)
INCOME_BUCKETS = (
    ("NOMINA", "Nomina"),
    ("ANUL COMPRA TARJ.", "Devoluciones y abonos"),
    ("ABONO", "Abonos"),
    ("TRANSFER", "Transferencias recibidas"),
    ("TRASPASO", "Transferencias recibidas"),
    ("BIZUM", "Ingresos Bizum"),
)
EXPENSE_BUCKETS = (
    (("TARJETA CREDITO",), "Tarjeta de credito"),
    (("PAGO BIZUM",), "Bizum"),
    (("MERCADONA", "LIDL", "ALDI", "CARREFOUR", "EROSKI", "CONSUM"), "Supermercado"),
    (("NETFLIX", "SPOTIFY", "APPLE.COM/BILL", "ELPAIS", "PRIME VIDEO", "DISNEY"), "Suscripciones"),
    (("NATURGY", "GAS COMERCIALIZADORA", "ENDESA", "IBERDROLA", "AGUA "), "Suministros"),
    (("APROOP TELECOM", "MOVISTAR", "VODAFONE", "ORANGE", "DIGI", "TELECOM"), "Telecomunicaciones"),
    (("PARKING", "REPSOL", "CEPSA", "BP ", "GASOLINERA", "RENFE", "UBER", "CABIFY"), "Transporte"),
    (("HOLIDAY INN", "BOOKING", "AIRBNB", "HOTEL"), "Viajes"),
    (("MUCCA", "RESTAURANTE", "CAFE", "BAR ", "BURGER", "MCDONALD"), "Restauracion"),
    (("AMAZON",), "Compras online"),
    (("ASOCIACION", "SIND"), "Cuotas y asociaciones"),
    (("ADEUDO RECIBO",), "Recibos domiciliados"),
)


class StatementImportError(Exception):
    pass


@dataclass
class ParsedMovement:
    booking_date: date
    value_date: date | None
    concept: str
    amount: Decimal
    balance: Decimal | None
    reference_1: str
    reference_2: str


@dataclass
class ClassifiedMovement:
    group: str
    bucket: str


def build_uploaded_file_checksum(uploaded_file) -> str:
    digest = sha256()
    for chunk in uploaded_file.chunks():
        digest.update(chunk)
    uploaded_file.seek(0)
    return digest.hexdigest()


def normalize_lookup_text(value: str) -> str:
    return re.sub(r"\s+", " ", normalize_header_text(value)).strip()


def normalize_iban(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def infer_ownership_category_from_holder_name(holder_name: str) -> str | None:
    normalized_holder = normalize_lookup_text(holder_name)
    has_ximo = "XIMO" in normalized_holder
    has_monica = "MONICA" in normalized_holder

    if has_ximo and has_monica:
        return AssetOwnershipCategory.JOINT
    if has_ximo:
        return AssetOwnershipCategory.XIMO
    if has_monica:
        return AssetOwnershipCategory.MONICA
    return None


def resolve_statement_ownership_category(statement_import: BankStatementImport, metadata: dict) -> str:
    if statement_import.ownership_category in {
        AssetOwnershipCategory.XIMO,
        AssetOwnershipCategory.MONICA,
    }:
        return statement_import.ownership_category

    holder_category = infer_ownership_category_from_holder_name(metadata.get("holder_name", ""))
    if holder_category:
        return holder_category

    imported_statements = list(
        BankStatementImport.objects.filter(
            import_status=BankStatementImport.ImportStatus.IMPORTED,
        )
        .exclude(pk=statement_import.pk)
        .order_by("-processed_at", "-imported_at", "-id")
    )

    target_iban = normalize_iban(metadata.get("iban", ""))
    if target_iban:
        for previous_statement in imported_statements:
            if normalize_iban(previous_statement.iban) == target_iban:
                return previous_statement.ownership_category

    target_account_label = normalize_lookup_text(metadata.get("account_label", ""))
    if target_account_label:
        for previous_statement in imported_statements:
            if normalize_lookup_text(previous_statement.account_label) == target_account_label:
                return previous_statement.ownership_category

        matching_account = (
            BankBalance.objects.order_by("-updated_at", "-id")
            .filter(account_name__iexact=metadata.get("account_label", "").strip())
            .first()
        )
        if matching_account:
            return matching_account.ownership_category

    return AssetOwnershipCategory.JOINT


def import_statement(statement_import: BankStatementImport) -> BankStatementImport:
    try:
        parsed = parse_statement_file(
            statement_import.source_file,
            statement_kind=statement_import.statement_kind,
        )
    except Exception as exc:
        statement_import.import_status = BankStatementImport.ImportStatus.FAILED
        statement_import.error_message = str(exc)
        statement_import.processed_at = timezone.now()
        statement_import.save(update_fields=["import_status", "error_message", "processed_at"])
        raise StatementImportError(str(exc)) from exc

    return apply_parsed_statement(statement_import, parsed)


def apply_parsed_statement(statement_import: BankStatementImport, parsed: dict) -> BankStatementImport:
    with transaction.atomic():
        statement_import.movements.all().delete()

        total_income = ZERO
        total_expenses = ZERO
        total_pension = ZERO
        total_dividends = ZERO
        movements_to_create = []

        for movement in parsed["movements"]:
            classified = classify_movement(movement.concept, movement.amount)
            amount_abs = abs(movement.amount)

            if classified.group == BankMovement.MovementGroup.DIVIDEND:
                total_dividends += amount_abs
            elif classified.group == BankMovement.MovementGroup.PENSION:
                total_pension += amount_abs
            elif classified.group == BankMovement.MovementGroup.EXPENSE:
                total_expenses += amount_abs
            else:
                total_income += amount_abs

            movements_to_create.append(
                BankMovement(
                    statement_import=statement_import,
                    booking_date=movement.booking_date,
                    value_date=movement.value_date,
                    concept=movement.concept,
                    normalized_concept=normalize_concept(movement.concept),
                    amount=movement.amount,
                    balance=movement.balance,
                    reference_1=movement.reference_1,
                    reference_2=movement.reference_2,
                    movement_group=classified.group,
                    concept_bucket=classified.bucket,
                )
            )

        BankMovement.objects.bulk_create(movements_to_create)

        statement_import.iban = parsed["metadata"].get("iban", "")
        statement_import.holder_name = parsed["metadata"].get("holder_name", "")
        statement_import.currency = parsed["metadata"].get("currency", "EUR")
        statement_import.period_start = parsed["metadata"].get("period_start")
        statement_import.period_end = parsed["metadata"].get("period_end")
        statement_import.account_label = parsed["metadata"].get("account_label", statement_import.account_label)
        if statement_import.external_account_id:
            statement_import.ownership_category = statement_import.external_account.ownership_category
        elif statement_import.connection_id:
            statement_import.ownership_category = statement_import.connection.ownership_category
        else:
            statement_import.ownership_category = resolve_statement_ownership_category(statement_import, parsed["metadata"])
        statement_import.opening_balance = parsed["metadata"].get("opening_balance")
        statement_import.closing_balance = parsed["metadata"].get("closing_balance")
        statement_import.total_income = total_income
        statement_import.total_expenses = total_expenses
        statement_import.total_pension_contributions = total_pension
        statement_import.total_dividends = total_dividends
        statement_import.import_status = BankStatementImport.ImportStatus.IMPORTED
        statement_import.error_message = ""
        statement_import.processed_at = timezone.now()
        statement_import.save()

    return statement_import


def get_source_name(source) -> str:
    if isinstance(source, (str, Path)):
        return str(source)
    return str(getattr(source, "name", "uploaded_file"))


def read_source_bytes(source) -> bytes:
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()

    if hasattr(source, "open"):
        source.open("rb")
    try:
        if hasattr(source, "seek"):
            source.seek(0)
        payload = source.read()
        if isinstance(payload, str):
            payload = payload.encode("utf-8")
        return payload
    finally:
        if hasattr(source, "seek"):
            source.seek(0)
        if hasattr(source, "close"):
            source.close()


def parse_statement_file(source, statement_kind: str | None = None) -> dict:
    rows = load_rows_from_workbook(source)
    if not rows:
        raise ValidationError("El extracto esta vacio.")

    source_name = get_source_name(source)
    metadata = extract_statement_metadata(rows, source_name=source_name)

    try:
        return parse_standard_statement_rows(rows, metadata)
    except ValidationError as standard_error:
        if statement_kind == BankStatementImport.StatementKind.ACCOUNT:
            raise standard_error
        if statement_kind == BankStatementImport.StatementKind.CARD or looks_like_card_statement(rows):
            return parse_card_statement_rows(rows, metadata, source_name=source_name)
        raise standard_error


def parse_standard_statement_rows(rows: list[list[str]], metadata: dict) -> dict:
    data_start_index, column_map = find_statement_layout(rows)
    movements = []

    for raw_row in rows[data_start_index:]:
        row = list(raw_row)
        if not row_looks_like_statement_data(row, column_map):
            continue
        movements.append(
            ParsedMovement(
                booking_date=parse_spanish_date(get_row_cell(row, column_map.get("booking_date"))),
                value_date=parse_optional_date(get_row_cell(row, column_map.get("value_date"))),
                concept=get_row_cell(row, column_map.get("concept")),
                amount=parse_row_amount(row, column_map),
                balance=parse_optional_decimal(get_row_cell(row, column_map.get("balance"))),
                reference_1=get_row_cell(row, column_map.get("reference_1")),
                reference_2=get_row_cell(row, column_map.get("reference_2")),
            )
        )

    if not movements:
        raise ValidationError("No se han encontrado movimientos bancarios en el extracto.")

    chronological = list(reversed(movements))
    first_balance = chronological[0].balance
    if first_balance is not None:
        metadata["opening_balance"] = first_balance - chronological[0].amount
    metadata["closing_balance"] = movements[0].balance

    return {"metadata": metadata, "movements": movements}


def load_rows_from_workbook(source) -> list[list[str]]:
    source_name = get_source_name(source)
    suffix = Path(source_name).suffix.lower()
    payload = read_source_bytes(source)
    if suffix == ".xlsx":
        return load_rows_from_xlsx(payload)
    if suffix == ".xls":
        return load_rows_from_xls(payload, source_name=source_name)
    raise ValidationError("Tipo de fichero no compatible. Sube extractos XLS o XLSX.")


def load_rows_from_xlsx(source) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("Se necesita openpyxl para importar ficheros XLSX.") from exc

    workbook_source = BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
    workbook = load_workbook(workbook_source, data_only=True, read_only=True)
    worksheet = workbook.worksheets[0]
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])
    workbook.close()
    return rows


def load_rows_from_xls(source, source_name: str = "uploaded.xls") -> list[list[str]]:
    payload = read_source_bytes(source) if not isinstance(source, (bytes, bytearray)) else bytes(source)
    if payload[:4] == b"PK\x03\x04":
        return load_rows_from_xlsx(payload)

    errors = []

    try:
        return load_rows_from_xls_with_xlrd(payload)
    except ValidationError as exc:
        errors.append(str(exc))

    try:
        return load_rows_from_html_xls(payload)
    except ValidationError as exc:
        errors.append(str(exc))

    if os.name == "nt":
        try:
            return load_rows_from_excel_com(payload, source_name=source_name)
        except ValidationError as exc:
            errors.append(str(exc))

    detail = "; ".join(dict.fromkeys(error for error in errors if error))
    message = "No se ha podido leer automaticamente el fichero XLS."
    if detail:
        message = f"{message} {detail}"
    raise ValidationError(message)


def load_rows_from_xls_with_xlrd(source) -> list[list[str]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValidationError("Se necesita xlrd para importar ficheros XLS antiguos.") from exc

    try:
        if isinstance(source, (bytes, bytearray)):
            workbook = xlrd.open_workbook(file_contents=bytes(source), on_demand=True)
        else:
            workbook = xlrd.open_workbook(source, on_demand=True)
    except Exception as exc:
        raise ValidationError(f"xlrd no ha podido leer el fichero XLS: {exc}") from exc

    try:
        worksheet = workbook.sheet_by_index(0)
        rows = []
        for row_index in range(worksheet.nrows):
            row = []
            for column_index in range(worksheet.ncols):
                row.append(normalize_xls_cell(workbook, worksheet, row_index, column_index))
            rows.append(row)
        return rows
    finally:
        release = getattr(workbook, "release_resources", None)
        if callable(release):
            release()


def normalize_xls_cell(workbook, worksheet, row_index: int, column_index: int) -> str:
    import xlrd

    cell_type = worksheet.cell_type(row_index, column_index)
    value = worksheet.cell_value(row_index, column_index)

    if cell_type in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return ""
    if cell_type == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(value, workbook.datemode).strftime("%d/%m/%Y")
    if cell_type == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if value else "FALSE"
    if cell_type == xlrd.XL_CELL_NUMBER:
        return format_excel_number(value)
    return str(value).strip()


def format_excel_number(value) -> str:
    number = Decimal(str(value))
    if number == number.to_integral():
        return str(int(number))
    return format(number.normalize(), "f")


class _HTMLTableRowParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows: list[list[str]] = []
        self._table_depth = 0
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._table_depth += 1
            return
        if not self._table_depth:
            return
        if tag == "tr":
            self._current_row = []
            return
        if tag in {"td", "th"}:
            self._current_cell = []
            return
        if tag == "br" and self._current_cell is not None:
            self._current_cell.append(" ")

    def handle_endtag(self, tag):
        if tag == "table" and self._table_depth:
            self._table_depth -= 1
            return
        if not self._table_depth:
            return
        if tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            text = re.sub(r"\s+", " ", "".join(self._current_cell)).strip()
            self._current_row.append(text)
            self._current_cell = None
            return
        if tag == "tr" and self._current_row is not None:
            if any(cell.strip() for cell in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None

    def handle_data(self, data):
        if self._current_cell is not None:
            self._current_cell.append(data)


def load_rows_from_html_xls(source) -> list[list[str]]:
    raw_bytes = read_source_bytes(source) if not isinstance(source, (bytes, bytearray)) else bytes(source)
    text = ""
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    lowered = text.lower()
    if "<table" not in lowered or "</table>" not in lowered:
        raise ValidationError("El fichero XLS no es una exportacion HTML con tabla legible.")

    parser = _HTMLTableRowParser()
    parser.feed(text)
    parser.close()
    if not parser.rows:
        raise ValidationError("No se han encontrado filas en el XLS basado en HTML.")
    return parser.rows


def load_rows_from_excel_com(source, source_name: str = "uploaded.xls") -> list[list[str]]:
    if os.name != "nt":
        raise ValidationError("La importacion de XLS necesita Windows con Microsoft Excel instalado.")

    payload = read_source_bytes(source) if not isinstance(source, (bytes, bytearray)) else bytes(source)
    suffix = Path(source_name).suffix or ".xls"
    with NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
        temp_file.write(payload)
        temp_path = temp_file.name

    escaped_path = temp_path.replace("'", "''")
    script = f"""
$ErrorActionPreference = 'Stop'
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
    $workbook = $excel.Workbooks.Open('{escaped_path}')
    $worksheet = $workbook.Worksheets.Item(1)
    $used = $worksheet.UsedRange
    $rows = @()
    for ($r = 1; $r -le $used.Rows.Count; $r++) {{
        $values = @()
        for ($c = 1; $c -le $used.Columns.Count; $c++) {{
            $values += [string]$worksheet.Cells.Item($r, $c).Text
        }}
        $rows += ,$values
    }}
    $workbook.Close($false)
    $rows | ConvertTo-Json -Depth 4 -Compress
}} finally {{
    $excel.Quit()
}}
"""
    try:
        completed = subprocess.run(
            ["powershell.exe", "-NoProfile", "-Command", script],
            capture_output=True,
            text=True,
            check=False,
        )
        if completed.returncode != 0:
            stderr = completed.stderr.strip() or "No se ha podido abrir el fichero XLS con Excel."
            raise ValidationError(stderr)

        payload = completed.stdout.strip()
        if not payload:
            raise ValidationError("Excel no ha devuelto datos para el fichero XLS subido.")

        rows = json.loads(payload)
        if rows and isinstance(rows[0], str):
            rows = [rows]
        normalized_rows = []
        for row in rows:
            if isinstance(row, dict) and "value" in row:
                row = row["value"]
            elif not isinstance(row, list):
                row = [row]
            normalized_rows.append(["" if cell is None else str(cell).strip() for cell in row])
        return normalized_rows
    finally:
        Path(temp_path).unlink(missing_ok=True)


def looks_like_iban(value: str) -> bool:
    compact = normalize_iban(value)
    return bool(re.match(r"^[A-Z]{2}\d{2}[A-Z0-9]{8,30}$", compact))


def infer_reference_date_from_source_name(source_name: str) -> date | None:
    stem = Path(source_name).stem
    for match in re.finditer(r"(?<!\d)(\d{2})(\d{2})(\d{4})(?!\d)", stem):
        day, month, year = match.groups()
        try:
            return date(int(year), int(month), int(day))
        except ValueError:
            continue
    return None


def extract_statement_metadata(rows: list[list[str]], source_name: str = "") -> dict:
    metadata = {
        "currency": "EUR",
        "account_label": "",
        "iban": "",
        "holder_name": "",
        "period_start": None,
        "period_end": None,
        "opening_balance": None,
        "closing_balance": None,
        "reference_date": infer_reference_date_from_source_name(source_name),
    }
    date_pattern = re.compile(
        r"(?:Desde|Del)\s+(\d{2}/\d{2}/\d{4})\s+(?:hasta|al)\s+(\d{2}/\d{2}/\d{4})",
        re.IGNORECASE,
    )

    for row in rows:
        text = " ".join(str(cell).strip() for cell in row if str(cell).strip())
        if not text:
            continue
        row_label = normalize_header_label(get_row_cell(row, 0))
        second_cell = get_row_cell(row, 1)
        third_cell = get_row_cell(row, 2)

        match = date_pattern.search(text)
        if match:
            metadata["period_start"] = parse_spanish_date(match.group(1))
            metadata["period_end"] = parse_spanish_date(match.group(2))
            metadata["reference_date"] = metadata["period_end"]

        normalized_text = normalize_header_text(text)
        if row_label in {"TARGETA", "TARJETA"}:
            card_number = second_cell
            card_name = third_cell
            if card_name:
                last_digits = re.sub(r"\D", "", card_number)[-4:]
                metadata["account_label"] = f"{card_name} {last_digits}".strip() if last_digits else card_name
            elif card_number:
                metadata["account_label"] = f"Tarjeta {re.sub(r'\\D', '', card_number)[-4:]}"
            continue
        if row_label in {"TITULAR TARGETA", "TITULAR TARJETA"} and second_cell:
            metadata["holder_name"] = second_cell
            continue

        if normalized_text.startswith("CUENTA:"):
            account_value = text.split(":", 1)[1].strip()
            if looks_like_iban(account_value):
                metadata["iban"] = account_value
                if not metadata["account_label"]:
                    metadata["account_label"] = f"Cuenta {metadata['iban'][-4:]}"
            elif not metadata["account_label"]:
                metadata["account_label"] = account_value
        elif normalized_text.startswith("IBAN:"):
            metadata["iban"] = text.split(":", 1)[1].strip()
            if metadata["iban"] and not metadata["account_label"]:
                metadata["account_label"] = f"Cuenta {metadata['iban'][-4:]}"
        elif normalized_text.startswith("DIVISA:"):
            metadata["currency"] = text.split("Divisa:", 1)[1].strip() or "EUR"
        elif normalized_text.startswith("MONEDA:"):
            metadata["currency"] = text.split(":", 1)[1].strip() or "EUR"
        elif normalized_text.startswith("TITULAR:") or normalized_text.startswith("TITULARES:"):
            metadata["holder_name"] = text.split(":", 1)[1].strip()

    return metadata


def normalize_header_label(text: str) -> str:
    normalized = normalize_header_text(text)
    normalized = re.sub(r"[^A-Z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def get_row_cell(row: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return str(row[index]).strip()


def is_booking_date_header(label: str) -> bool:
    if not label:
        return False
    if label in {"F OPERATIVA", "F OPERACION", "FECHA OPERATIVA", "FECHA OPERACION", "FECHA", "DATA"}:
        return True
    return label.startswith("FECHA OPER")


def is_value_date_header(label: str) -> bool:
    if not label:
        return False
    if label in {"F VALOR", "FECHA VALOR", "VALOR"}:
        return True
    return "FECHA VALOR" in label


def is_concept_header(label: str) -> bool:
    if not label:
        return False
    return any(token in label for token in ("CONCEPTO", "CONCEPTE", "DESCRIPCION", "DETALLE")) or label == "MOVIMIENTO"


def is_amount_header(label: str) -> bool:
    if not label:
        return False
    return label.startswith("IMPORTE") or label.startswith("IMPORT")


def is_debit_header(label: str) -> bool:
    if not label:
        return False
    return label in {"CARGO", "CARGOS", "DEBE", "DEBITO"} or label.startswith("CARGO ")


def is_credit_header(label: str) -> bool:
    if not label:
        return False
    return label in {"ABONO", "ABONOS", "HABER", "CREDITO", "INGRESO"} or label.startswith("ABONO ")


def is_balance_header(label: str) -> bool:
    if not label:
        return False
    return label.startswith("SALDO") or label == "DISPONIBLE"


def is_reference_1_header(label: str) -> bool:
    if not label:
        return False
    return label in {"REFERENCIA 1", "REF 1", "REFERENCIA1", "LOCALITAT", "LOCALIDAD"}


def is_reference_2_header(label: str) -> bool:
    if not label:
        return False
    return label in {"REFERENCIA 2", "REF 2", "REFERENCIA2", "SIT MOV", "SIT MOV.", "ESTADO", "ESTADO MOV"}


def build_statement_column_map(row: list[str]) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for index, cell in enumerate(row):
        label = normalize_header_label(str(cell))
        if not label:
            continue
        if "reference_2" not in column_map and is_reference_2_header(label):
            column_map["reference_2"] = index
        elif "reference_1" not in column_map and is_reference_1_header(label):
            column_map["reference_1"] = index
        elif "balance" not in column_map and is_balance_header(label):
            column_map["balance"] = index
        elif "amount" not in column_map and is_amount_header(label):
            column_map["amount"] = index
        elif "debit_amount" not in column_map and is_debit_header(label):
            column_map["debit_amount"] = index
        elif "credit_amount" not in column_map and is_credit_header(label):
            column_map["credit_amount"] = index
        elif "value_date" not in column_map and is_value_date_header(label):
            column_map["value_date"] = index
        elif "concept" not in column_map and is_concept_header(label):
            column_map["concept"] = index
        elif "booking_date" not in column_map and is_booking_date_header(label):
            column_map["booking_date"] = index
    return column_map


def has_required_statement_columns(column_map: dict[str, int]) -> bool:
    has_amount = "amount" in column_map or "debit_amount" in column_map or "credit_amount" in column_map
    return "booking_date" in column_map and "concept" in column_map and has_amount


def parse_optional_date(value: str) -> date | None:
    return parse_spanish_date(value) if str(value).strip() else None


def parse_optional_decimal(value: str) -> Decimal | None:
    return parse_spanish_decimal(value) if str(value).strip() else None


def parse_row_amount(row: list[str], column_map: dict[str, int]) -> Decimal:
    if "amount" in column_map:
        return parse_spanish_decimal(get_row_cell(row, column_map["amount"]))

    debit = parse_optional_decimal(get_row_cell(row, column_map.get("debit_amount"))) or ZERO
    credit = parse_optional_decimal(get_row_cell(row, column_map.get("credit_amount"))) or ZERO
    return credit - debit


def row_looks_like_statement_data(row: list[str], column_map: dict[str, int]) -> bool:
    booking_value = get_row_cell(row, column_map.get("booking_date"))
    concept_value = get_row_cell(row, column_map.get("concept"))
    if not booking_value or not concept_value:
        return False

    try:
        parse_spanish_date(booking_value)
    except Exception:
        return False

    has_amount_value = any(
        get_row_cell(row, column_map.get(field_name))
        for field_name in ("amount", "debit_amount", "credit_amount")
    )
    if not has_amount_value:
        return False

    try:
        parse_row_amount(row, column_map)
    except Exception:
        return False

    return True


def looks_like_card_statement(rows: list[list[str]]) -> bool:
    for row in rows:
        text = " ".join(str(cell).strip() for cell in row if str(cell).strip())
        normalized_text = normalize_header_text(text)
        if not normalized_text:
            continue
        if (
            normalized_text in CARD_DEBIT_SECTION_LABELS
            or normalized_text in CARD_CREDIT_SECTION_LABELS
            or normalized_text.startswith("TARGETA:")
            or normalized_text.startswith("TARJETA:")
            or normalized_text.startswith("TITULAR TARGETA")
            or normalized_text.startswith("TITULAR TARJETA")
        ):
            return True
    return False


def parse_card_statement_date(value: str, reference_date: date | None) -> date:
    text = str(value).strip()
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", text):
        return parse_spanish_date(text)
    if re.fullmatch(r"\d{2}/\d{2}", text):
        anchor = reference_date or timezone.localdate()
        day, month = (int(part) for part in text.split("/"))
        candidate = date(anchor.year, month, day)
        if candidate > anchor:
            candidate = date(anchor.year - 1, month, day)
        return candidate
    raise ValueError(f"Fecha no compatible: {value}")


def parse_card_row_amount(row: list[str], column_map: dict[str, int], section_sign: int | None) -> Decimal:
    raw_amount = get_row_cell(row, column_map.get("amount"))
    amount = parse_spanish_decimal(raw_amount)
    if amount < ZERO or raw_amount.strip().startswith("-"):
        return amount
    if section_sign == 1:
        return abs(amount)
    return -abs(amount)


def row_looks_like_card_statement_data(
    row: list[str],
    column_map: dict[str, int],
    reference_date: date | None,
) -> bool:
    booking_value = get_row_cell(row, column_map.get("booking_date"))
    concept_value = get_row_cell(row, column_map.get("concept"))
    if not booking_value or not concept_value:
        return False

    try:
        parse_card_statement_date(booking_value, reference_date)
    except Exception:
        return False

    amount_value = get_row_cell(row, column_map.get("amount"))
    if not amount_value:
        return False

    try:
        parse_spanish_decimal(amount_value)
    except Exception:
        return False

    return True


def parse_card_statement_rows(rows: list[list[str]], metadata: dict, source_name: str = "") -> dict:
    reference_date = metadata.get("period_end") or metadata.get("reference_date") or infer_reference_date_from_source_name(source_name)
    current_column_map: dict[str, int] | None = None
    current_section_sign = -1
    movements: list[ParsedMovement] = []

    for raw_row in rows:
        row = list(raw_row)
        text = " ".join(str(cell).strip() for cell in row if str(cell).strip())
        normalized_text = normalize_header_text(text)
        if normalized_text in CARD_DEBIT_SECTION_LABELS:
            current_section_sign = -1
            continue
        if normalized_text in CARD_CREDIT_SECTION_LABELS:
            current_section_sign = 1
            continue

        candidate_column_map = build_statement_column_map(row)
        if has_required_statement_columns(candidate_column_map):
            current_column_map = candidate_column_map
            continue

        if not current_column_map:
            continue
        if not row_looks_like_card_statement_data(row, current_column_map, reference_date):
            continue

        booking_date = parse_card_statement_date(get_row_cell(row, current_column_map.get("booking_date")), reference_date)
        movements.append(
            ParsedMovement(
                booking_date=booking_date,
                value_date=parse_optional_date(get_row_cell(row, current_column_map.get("value_date"))),
                concept=get_row_cell(row, current_column_map.get("concept")),
                amount=parse_card_row_amount(row, current_column_map, current_section_sign),
                balance=None,
                reference_1=get_row_cell(row, current_column_map.get("reference_1")),
                reference_2=get_row_cell(row, current_column_map.get("reference_2")),
            )
        )

    if not movements:
        raise ValidationError("No se han encontrado movimientos de tarjeta en el documento.")

    if not metadata.get("period_start"):
        metadata["period_start"] = min((movement.booking_date for movement in movements), default=None)
    if not metadata.get("period_end"):
        metadata["period_end"] = max((movement.booking_date for movement in movements), default=None)
    if not metadata.get("reference_date") and metadata.get("period_end"):
        metadata["reference_date"] = metadata["period_end"]

    return {"metadata": metadata, "movements": movements}


def find_statement_layout(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows):
        padded = [str(cell).strip() for cell in row[:5]]
        if padded == HEADER_ROW:
            return index + 1, DEFAULT_STATEMENT_COLUMN_MAP.copy()

    for index, row in enumerate(rows):
        column_map = build_statement_column_map(row)
        if has_required_statement_columns(column_map):
            return index + 1, column_map

    for index, row in enumerate(rows):
        if row_looks_like_statement_data(row, DEFAULT_STATEMENT_COLUMN_MAP):
            return index, DEFAULT_STATEMENT_COLUMN_MAP.copy()

    raise ValidationError("No se ha encontrado la fila de cabecera del extracto.")


def parse_spanish_date(value: str):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha no compatible: {value}")


def parse_spanish_decimal(value: str) -> Decimal:
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace("EUR", "").replace("\u20ac", "")
    if not text:
        return ZERO
    if "," in text:
        normalized = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        normalized = text.replace(".", "")
    elif "." in text:
        integer_part, decimal_part = text.rsplit(".", 1)
        normalized = text if len(decimal_part) <= 2 else f"{integer_part}{decimal_part}"
    else:
        normalized = text
    return Decimal(normalized)


def normalize_concept(concept: str) -> str:
    return re.sub(r"\s+", " ", concept.strip().upper())


def normalize_header_text(text: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return ascii_text.strip().upper()


def month_label_for_date(value: date) -> str:
    return value.strftime("%Y-%m")


def classify_movement(concept: str, amount: Decimal) -> ClassifiedMovement:
    normalized = normalize_concept(concept)

    if any(keyword in normalized for keyword in PLAN_KEYWORDS):
        return ClassifiedMovement(
            group=BankMovement.MovementGroup.PENSION,
            bucket="Aportaciones a planes",
        )

    if any(keyword in normalized for keyword in DIVIDEND_KEYWORDS):
        return ClassifiedMovement(
            group=BankMovement.MovementGroup.DIVIDEND,
            bucket="Dividendos de acciones",
        )

    if amount >= ZERO:
        if any(keyword in normalized for keyword in RENT_INCOME_KEYWORDS):
            return ClassifiedMovement(
                group=BankMovement.MovementGroup.INCOME,
                bucket="Alquiler piso (Adrian)",
            )
        for keyword, bucket in INCOME_BUCKETS:
            if keyword in normalized:
                return ClassifiedMovement(
                    group=BankMovement.MovementGroup.INCOME,
                    bucket=bucket,
                )
        return ClassifiedMovement(
            group=BankMovement.MovementGroup.INCOME,
            bucket="Otros ingresos",
        )

    for keywords, bucket in EXPENSE_BUCKETS:
        if any(keyword in normalized for keyword in keywords):
            return ClassifiedMovement(
                group=BankMovement.MovementGroup.EXPENSE,
                bucket=bucket,
            )

    return ClassifiedMovement(
        group=BankMovement.MovementGroup.EXPENSE,
        bucket="Otros gastos",
    )


def build_banking_dashboard() -> dict:
    account_overview = build_bank_account_overview()
    card_overview = build_card_spending_overview()
    statements = list(
        BankStatementImport.objects.prefetch_related("movements").order_by("-period_end", "-imported_at")
    )
    imported_statements = [
        statement
        for statement in statements
        if statement.import_status == BankStatementImport.ImportStatus.IMPORTED
        and statement.statement_kind == BankStatementImport.StatementKind.ACCOUNT
    ]

    monthly_labels = set()
    accounts = {statement.iban or statement.account_name for statement in imported_statements}
    month_map = defaultdict(
        lambda: {
            "label": "",
            "accounts": set(),
            "income": ZERO,
            "expenses": ZERO,
            "pension_contributions": ZERO,
            "dividends": ZERO,
            "latest_balances": {},
        }
    )
    income_map = defaultdict(lambda: defaultdict(lambda: ZERO))
    expense_map = defaultdict(lambda: defaultdict(lambda: ZERO))
    overall_summary = {
        "statements_count": len(imported_statements),
        "months_count": 0,
        "accounts_count": len(accounts),
        "total_income": ZERO,
        "total_expenses": ZERO,
        "total_pension_contributions": ZERO,
        "total_dividends": ZERO,
    }

    for statement in imported_statements:
        account_key = statement.iban or statement.account_name
        account_name = statement.account_name
        for movement in statement.movements.all():
            movement_month = month_label_for_date(movement.booking_date)
            monthly_labels.add(movement_month)
            bucket = month_map[movement_month]
            bucket["label"] = movement_month
            bucket["accounts"].add(account_name)

            amount_abs = abs(movement.amount)
            if movement.movement_group == BankMovement.MovementGroup.DIVIDEND:
                bucket["dividends"] += amount_abs
                overall_summary["total_dividends"] += amount_abs
            elif movement.movement_group == BankMovement.MovementGroup.PENSION:
                bucket["pension_contributions"] += amount_abs
                overall_summary["total_pension_contributions"] += amount_abs
            elif movement.movement_group == BankMovement.MovementGroup.EXPENSE:
                bucket["expenses"] += amount_abs
                overall_summary["total_expenses"] += amount_abs
            else:
                bucket["income"] += amount_abs
                overall_summary["total_income"] += amount_abs

            if (
                movement.movement_group in (BankMovement.MovementGroup.INCOME, BankMovement.MovementGroup.DIVIDEND)
                and movement.concept_bucket in TRACKED_INCOME_BUCKETS
            ):
                income_map[movement.concept_bucket][movement_month] += amount_abs
            if movement.movement_group == BankMovement.MovementGroup.EXPENSE:
                expense_map[movement.concept_bucket][movement_month] += amount_abs

            if movement.balance is not None:
                latest = bucket["latest_balances"].get(account_key)
                movement_position = (movement.booking_date, movement.id)
                if latest is None or movement_position > latest["position"]:
                    bucket["latest_balances"][account_key] = {
                        "position": movement_position,
                        "balance": movement.balance,
                    }

    if not monthly_labels:
        monthly_labels = {statement.month_label for statement in imported_statements if statement.period_end}

    overall_summary["months_count"] = len(monthly_labels)

    monthly_summaries = []
    for month_label in sorted(month_map.keys(), reverse=True):
        row = month_map[month_label]
        row["accounts_count"] = len(row["accounts"])
        row["net_cash_flow"] = row["income"] + row["dividends"] - row["expenses"] - row["pension_contributions"]
        row["closing_balance"] = sum(
            (item["balance"] for item in row["latest_balances"].values()),
            ZERO,
        )
        row.pop("latest_balances", None)
        monthly_summaries.append(row)

    expense_months = sorted(monthly_labels)
    income_matrix = []
    expense_matrix = []

    for concept_bucket in TRACKED_INCOME_BUCKETS:
        values = income_map[concept_bucket]
        row_values = [values.get(month_label, ZERO) for month_label in expense_months]
        income_matrix.append(
            {
                "concept": concept_bucket,
                "values": row_values,
                "total": sum(row_values, ZERO),
            }
        )

    for concept_bucket, values in sorted(
        expense_map.items(),
        key=lambda item: sum(item[1].values()),
        reverse=True,
    ):
        row_values = [values.get(month_label, ZERO) for month_label in expense_months]
        expense_matrix.append(
            {
                "concept": concept_bucket,
                "values": row_values,
                "total": sum(row_values, ZERO),
            }
        )

    return {
        "accounts_summary": account_overview["summary"],
        "tracked_accounts": account_overview["accounts"],
        "card_summary": card_overview["summary"],
        "tracked_cards": card_overview["cards"],
        "statement_summary": overall_summary,
        "monthly_summaries": monthly_summaries,
        "income_months": expense_months,
        "income_matrix": income_matrix,
        "expense_months": expense_months,
        "expense_matrix": expense_matrix,
        "card_monthly_summaries": card_overview["monthly_summaries"],
        "card_expense_months": card_overview["expense_months"],
        "card_expense_matrix": card_overview["expense_matrix"],
        "recent_imports": statements[:12],
    }


def build_bank_account_overview() -> dict:
    manual_accounts = list(BankBalance.objects.order_by("institution", "account_name", "id"))
    imported_statements = list(
        BankStatementImport.objects.filter(
            import_status=BankStatementImport.ImportStatus.IMPORTED,
            statement_kind=BankStatementImport.StatementKind.ACCOUNT,
        )
        .exclude(closing_balance__isnull=True)
        .order_by("period_end", "imported_at", "id")
    )

    latest_statements_by_account: dict[str, BankStatementImport] = {}
    statement_counts: dict[str, int] = defaultdict(int)

    for statement in imported_statements:
        account_key = normalize_iban(statement.iban) or normalize_lookup_text(statement.account_name) or str(statement.pk)
        latest_statements_by_account[account_key] = statement
        statement_counts[account_key] += 1

    remaining_manual_accounts = manual_accounts.copy()
    tracked_accounts = []

    for account_key, statement in latest_statements_by_account.items():
        normalized_statement_name = normalize_lookup_text(statement.account_name)
        manual_match = next(
            (
                account
                for account in remaining_manual_accounts
                if normalize_lookup_text(account.account_name) == normalized_statement_name
            ),
            None,
        )
        if manual_match:
            remaining_manual_accounts.remove(manual_match)

        ownership_category = (
            manual_match.ownership_category
            if manual_match and manual_match.ownership_category
            else statement.ownership_category
        )
        latest_net_cash_flow = (
            statement.total_income
            + statement.total_dividends
            - statement.total_expenses
            - statement.total_pension_contributions
        )

        tracked_accounts.append(
            {
                "account_name": statement.account_name,
                "institution": manual_match.institution if manual_match else statement.institution,
                "ownership_category": ownership_category,
                "ownership_label": AssetOwnershipCategory(ownership_category).label,
                "source_label": "Manual + extracto" if manual_match else "Extracto importado",
                "current_balance": statement.closing_balance or ZERO,
                "deposited_amount": manual_match.deposited_amount if manual_match else None,
                "annual_interest_income": manual_match.annual_interest_income if manual_match else ZERO,
                "latest_month": statement.month_label if statement.period_end else None,
                "latest_statement_id": statement.id,
                "statement_count": statement_counts[account_key],
                "latest_net_cash_flow": latest_net_cash_flow,
                "notes": manual_match.notes if manual_match else "",
                "account_id": manual_match.id if manual_match else None,
                "has_imported_data": True,
                "has_manual_data": manual_match is not None,
                "edit_action": "update_statement_ownership",
                "edit_target_id": statement.id,
            }
        )

    for account in remaining_manual_accounts:
        tracked_accounts.append(
            {
                "account_name": account.account_name,
                "institution": account.institution,
                "ownership_category": account.ownership_category,
                "ownership_label": account.get_ownership_category_display(),
                "source_label": "Manual",
                "current_balance": account.current_balance,
                "deposited_amount": account.deposited_amount,
                "annual_interest_income": account.annual_interest_income,
                "latest_month": None,
                "latest_statement_id": None,
                "statement_count": 0,
                "latest_net_cash_flow": None,
                "notes": account.notes,
                "account_id": account.id,
                "has_imported_data": False,
                "has_manual_data": True,
                "edit_action": "update_account_ownership",
                "edit_target_id": account.id,
            }
        )

    tracked_accounts.sort(
        key=lambda item: (
            item["current_balance"],
            item["latest_month"] or "",
            item["account_name"],
        ),
        reverse=True,
    )

    summary = {
        "accounts_count": len(tracked_accounts),
        "current_balance": sum((account["current_balance"] for account in tracked_accounts), ZERO),
        "deposited_amount": sum(
            ((account["deposited_amount"] or ZERO) for account in tracked_accounts if account["has_manual_data"]),
            ZERO,
        ),
        "annual_interest_income": sum(
            ((account["annual_interest_income"] or ZERO) for account in tracked_accounts if account["has_manual_data"]),
            ZERO,
        ),
        "imported_accounts_count": sum(1 for account in tracked_accounts if account["has_imported_data"]),
        "manual_accounts_count": sum(1 for account in tracked_accounts if account["has_manual_data"]),
        "latest_month": max(
            (account["latest_month"] for account in tracked_accounts if account["latest_month"]),
            default=None,
        ),
    }

    return {
        "accounts": tracked_accounts,
        "summary": summary,
    }


def build_card_spending_overview() -> dict:
    card_statements = list(
        BankStatementImport.objects.filter(
            import_status=BankStatementImport.ImportStatus.IMPORTED,
            statement_kind=BankStatementImport.StatementKind.CARD,
        )
        .prefetch_related("movements")
        .order_by("period_end", "imported_at", "id")
    )
    if not card_statements:
        return {
            "cards": [],
            "summary": {
                "cards_count": 0,
                "statements_count": 0,
                "months_count": 0,
                "total_spent": ZERO,
                "total_refunds": ZERO,
                "latest_month": None,
            },
            "monthly_summaries": [],
            "expense_months": [],
            "expense_matrix": [],
        }

    latest_by_card: dict[str, BankStatementImport] = {}
    statement_counts: dict[str, int] = defaultdict(int)
    month_map = defaultdict(
        lambda: {
            "label": "",
            "cards": set(),
            "expenses": ZERO,
            "refunds": ZERO,
        }
    )
    expense_map = defaultdict(lambda: defaultdict(lambda: ZERO))
    monthly_labels = set()

    for statement in card_statements:
        card_key = normalize_iban(statement.iban) or normalize_lookup_text(statement.account_name) or str(statement.pk)
        latest_by_card[card_key] = statement
        statement_counts[card_key] += 1

        statement_month = statement.month_label if statement.period_end else None
        if statement_month:
            monthly_labels.add(statement_month)

        for movement in statement.movements.all():
            movement_month = month_label_for_date(movement.booking_date)
            monthly_labels.add(movement_month)
            bucket = month_map[movement_month]
            bucket["label"] = movement_month
            bucket["cards"].add(card_key)

            amount_abs = abs(movement.amount)
            if movement.movement_group == BankMovement.MovementGroup.EXPENSE:
                bucket["expenses"] += amount_abs
                expense_map[movement.concept_bucket][movement_month] += amount_abs
            elif movement.movement_group == BankMovement.MovementGroup.INCOME:
                bucket["refunds"] += amount_abs

    tracked_cards = []
    for card_key, statement in latest_by_card.items():
        tracked_cards.append(
            {
                "card_name": statement.account_name,
                "ownership_category": statement.ownership_category,
                "ownership_label": statement.get_ownership_category_display(),
                "latest_month": statement.month_label if statement.period_end else None,
                "latest_spent": statement.total_expenses,
                "latest_refunds": statement.total_income,
                "statement_count": statement_counts[card_key],
                "statement_id": statement.id,
                "source_filename": statement.source_filename,
            }
        )

    tracked_cards.sort(
        key=lambda item: (item["latest_spent"], item["latest_month"] or "", item["card_name"]),
        reverse=True,
    )

    monthly_summaries = []
    for month_label in sorted(month_map.keys(), reverse=True):
        row = month_map[month_label]
        row["cards_count"] = len(row["cards"])
        row["net_spent"] = row["expenses"] - row["refunds"]
        monthly_summaries.append(row)

    expense_months = sorted(monthly_labels)
    expense_matrix = []
    for concept_bucket, values in sorted(
        expense_map.items(),
        key=lambda item: sum(item[1].values()),
        reverse=True,
    ):
        row_values = [values.get(month_label, ZERO) for month_label in expense_months]
        expense_matrix.append(
            {
                "concept": concept_bucket,
                "values": row_values,
                "total": sum(row_values, ZERO),
            }
        )

    summary = {
        "cards_count": len(tracked_cards),
        "statements_count": len(card_statements),
        "months_count": len(monthly_labels),
        "total_spent": sum((statement.total_expenses for statement in card_statements), ZERO),
        "total_refunds": sum((statement.total_income for statement in card_statements), ZERO),
        "latest_month": max((card["latest_month"] for card in tracked_cards if card["latest_month"]), default=None),
    }

    return {
        "cards": tracked_cards,
        "summary": summary,
        "monthly_summaries": monthly_summaries,
        "expense_months": expense_months,
        "expense_matrix": expense_matrix,
    }


def get_gocardless_credentials() -> tuple[str, str]:
    secret_id = os.environ.get("GOCARDLESS_BANK_DATA_SECRET_ID", "").strip()
    secret_key = os.environ.get("GOCARDLESS_BANK_DATA_SECRET_KEY", "").strip()
    if not secret_id or not secret_key:
        raise ValidationError(
            "Faltan GOCARDLESS_BANK_DATA_SECRET_ID y GOCARDLESS_BANK_DATA_SECRET_KEY en el entorno."
        )
    return secret_id, secret_key


def gocardless_api_request(
    method: str,
    path: str,
    payload: dict | None = None,
    access_token: str | None = None,
    query: dict | None = None,
):
    path = path if path.startswith("/") else f"/{path}"
    url = f"{GOCARDLESS_API_BASE_URL}{path}"
    if query:
        encoded_query = urllib_parse.urlencode({key: value for key, value in query.items() if value not in {None, ""}})
        if encoded_query:
            url = f"{url}?{encoded_query}"

    data = None
    headers = {
        "Accept": "application/json",
    }
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    if access_token:
        headers["Authorization"] = f"Bearer {access_token}"

    request = urllib_request.Request(url, data=data, headers=headers, method=method.upper())
    try:
        with urllib_request.urlopen(request, timeout=30) as response:
            raw_body = response.read()
    except urllib_error.HTTPError as exc:
        response_body = exc.read().decode("utf-8", errors="replace")
        try:
            parsed_body = json.loads(response_body)
        except Exception:
            parsed_body = response_body
        raise ValidationError(f"Open Banking ha devuelto un error: {parsed_body}") from exc
    except urllib_error.URLError as exc:
        raise ValidationError(f"No se ha podido conectar con Open Banking: {exc.reason}") from exc

    if not raw_body:
        return {}
    return json.loads(raw_body.decode("utf-8"))


def get_gocardless_access_token() -> str:
    secret_id, secret_key = get_gocardless_credentials()
    refresh_payload = gocardless_api_request(
        "POST",
        "/token/new/",
        payload={"secret_id": secret_id, "secret_key": secret_key},
    )
    token_payload = gocardless_api_request(
        "POST",
        "/token/refresh/",
        payload={"refresh": refresh_payload["refresh"]},
    )
    return token_payload["access"]


def search_gocardless_institutions(country_code: str, query: str = "") -> list[dict]:
    access_token = get_gocardless_access_token()
    institutions = gocardless_api_request(
        "GET",
        "/institutions/",
        access_token=access_token,
        query={"country": country_code.lower()},
    )
    normalized_query = normalize_lookup_text(query)
    if not normalized_query:
        return institutions[:20]
    filtered = [
        institution
        for institution in institutions
        if normalized_query in normalize_lookup_text(institution.get("name", ""))
    ]
    return filtered[:20]


def create_open_banking_connection(connection: BankConnection, redirect_url: str) -> str:
    if connection.provider != BankConnection.Provider.GOCARDLESS:
        raise ValidationError("Proveedor de Open Banking no compatible.")

    access_token = get_gocardless_access_token()
    agreement_payload = gocardless_api_request(
        "POST",
        "/agreements/enduser/",
        access_token=access_token,
        payload={
            "institution_id": connection.institution_id,
            "max_historical_days": 90,
            "access_valid_for_days": 90,
            "access_scope": ["balances", "details", "transactions"],
        },
    )
    requisition_payload = gocardless_api_request(
        "POST",
        "/requisitions/",
        access_token=access_token,
        payload={
            "redirect": redirect_url,
            "institution_id": connection.institution_id,
            "reference": connection.reference,
            "agreement": agreement_payload["id"],
            "user_language": "ES",
        },
    )

    connection.agreement_id = agreement_payload.get("id", "")
    connection.requisition_id = requisition_payload.get("id", "")
    connection.requisition_link = requisition_payload.get("link", "")
    connection.requisition_status = requisition_payload.get("status", "")
    connection.last_error = ""
    connection.save(
        update_fields=[
            "agreement_id",
            "requisition_id",
            "requisition_link",
            "requisition_status",
            "last_error",
            "updated_at",
        ]
    )
    return connection.requisition_link


def infer_external_account_kind(details: dict, account_label: str) -> str:
    normalized_text = normalize_lookup_text(
        " ".join(
            [
                str(details.get("name", "")),
                str(details.get("displayName", "")),
                str(details.get("product", "")),
                str(details.get("details", "")),
                str(details.get("cashAccountType", "")),
                str(details.get("linkedAccounts", "")),
                account_label,
            ]
        )
    )
    if any(keyword in normalized_text for keyword in CARD_ACCOUNT_KEYWORDS):
        return BankStatementImport.StatementKind.CARD
    if details.get("linkedAccounts") and not details.get("iban"):
        return BankStatementImport.StatementKind.CARD
    return BankStatementImport.StatementKind.ACCOUNT


def choose_external_account_label(details: dict, connection: BankConnection) -> str:
    for key in ("displayName", "name", "product", "details"):
        value = str(details.get(key, "")).strip()
        if value:
            return value[:180]
    if details.get("iban"):
        return f"Cuenta {str(details['iban'])[-4:]}"
    return f"{connection.institution_name} {details.get('resourceId', '').strip()}".strip()


def parse_balance_amount(balance_value) -> Decimal | None:
    if balance_value is None or balance_value == "":
        return None
    if isinstance(balance_value, dict):
        if "amount" in balance_value:
            return Decimal(str(balance_value.get("amount")))
        nested = balance_value.get("balanceAmount")
        if isinstance(nested, dict) and "amount" in nested:
            return Decimal(str(nested.get("amount")))
        return None
    return Decimal(str(balance_value))


def pick_closing_balance(balance_payload: dict) -> Decimal | None:
    balances = balance_payload.get("balances", []) if isinstance(balance_payload, dict) else []
    if not balances:
        return None
    priority = ("closingAvailable", "expected", "closingBooked", "interimAvailable")
    best = None
    best_rank = len(priority) + 1
    for item in balances:
        balance_type = str(item.get("balanceType", ""))
        amount = parse_balance_amount(item)
        if amount is None:
            continue
        try:
            rank = priority.index(balance_type)
        except ValueError:
            rank = len(priority)
        if best is None or rank < best_rank:
            best = amount
            best_rank = rank
    return best


def parse_open_banking_transaction(raw_transaction: dict) -> ParsedMovement:
    amount = parse_balance_amount(raw_transaction.get("transactionAmount")) or ZERO
    concept = (
        raw_transaction.get("remittanceInformationUnstructured")
        or raw_transaction.get("additionalInformation")
        or raw_transaction.get("bankTransactionCode")
        or raw_transaction.get("proprietaryBankTransactionCode")
        or raw_transaction.get("creditorName")
        or raw_transaction.get("debtorName")
        or "Movimiento bancario"
    )
    reference_1 = raw_transaction.get("creditorName") or raw_transaction.get("debtorName") or ""
    reference_2 = (
        raw_transaction.get("entryReference")
        or raw_transaction.get("internalTransactionId")
        or raw_transaction.get("transactionId")
        or ""
    )
    balance = parse_balance_amount(raw_transaction.get("balanceAfterTransaction"))
    return ParsedMovement(
        booking_date=parse_spanish_date(
            raw_transaction.get("bookingDate") or raw_transaction.get("valueDate") or raw_transaction.get("bookingDateTime")
        ),
        value_date=parse_optional_date(raw_transaction.get("valueDate", "")),
        concept=str(concept).strip(),
        amount=amount,
        balance=balance,
        reference_1=str(reference_1).strip(),
        reference_2=str(reference_2).strip(),
    )


def build_open_banking_statement_checksum(external_account: BankExternalAccount, month_label: str) -> str:
    identity = (
        f"open-banking:{external_account.provider_account_id}:{external_account.statement_kind}:{month_label}"
    )
    return sha256(identity.encode("utf-8")).hexdigest()


def upsert_external_account(connection: BankConnection, account_id: str, details: dict) -> BankExternalAccount:
    account_label = choose_external_account_label(details, connection)
    external_account, created = BankExternalAccount.objects.get_or_create(
        provider_account_id=account_id,
        defaults={
            "connection": connection,
            "ownership_category": connection.ownership_category,
            "statement_kind": infer_external_account_kind(details, account_label),
            "institution": connection.institution_name,
            "account_label": account_label,
            "iban": details.get("iban", "") or details.get("bban", ""),
            "currency": details.get("currency", "EUR") or "EUR",
            "holder_name": details.get("ownerName", ""),
            "linked_account_name": details.get("linkedAccounts", ""),
            "raw_details": details,
        },
    )
    if created:
        return external_account

    updated_fields = []
    if external_account.connection_id != connection.id:
        external_account.connection = connection
        updated_fields.append("connection")
    if external_account.institution != connection.institution_name:
        external_account.institution = connection.institution_name
        updated_fields.append("institution")
    if external_account.account_label != account_label:
        external_account.account_label = account_label
        updated_fields.append("account_label")
    if details.get("iban", "") or details.get("bban", ""):
        iban = details.get("iban", "") or details.get("bban", "")
        if external_account.iban != iban:
            external_account.iban = iban
            updated_fields.append("iban")
    currency = details.get("currency", "EUR") or "EUR"
    if external_account.currency != currency:
        external_account.currency = currency
        updated_fields.append("currency")
    holder_name = details.get("ownerName", "")
    if holder_name and external_account.holder_name != holder_name:
        external_account.holder_name = holder_name
        updated_fields.append("holder_name")
    linked_account_name = details.get("linkedAccounts", "")
    if external_account.linked_account_name != linked_account_name:
        external_account.linked_account_name = linked_account_name
        updated_fields.append("linked_account_name")
    inferred_kind = infer_external_account_kind(details, account_label)
    if not external_account.statement_imports.exists() and external_account.statement_kind != inferred_kind:
        external_account.statement_kind = inferred_kind
        updated_fields.append("statement_kind")
    if external_account.raw_details != details:
        external_account.raw_details = details
        updated_fields.append("raw_details")
    if updated_fields:
        updated_fields.append("updated_at")
        external_account.save(update_fields=updated_fields)
    return external_account


def build_open_banking_parsed_payload(
    external_account: BankExternalAccount,
    movements: list[ParsedMovement],
    closing_balance: Decimal | None,
) -> dict:
    period_start = min((movement.booking_date for movement in movements), default=None)
    period_end = max((movement.booking_date for movement in movements), default=None)
    metadata = {
        "currency": external_account.currency or "EUR",
        "account_label": external_account.account_label,
        "iban": external_account.iban,
        "holder_name": external_account.holder_name,
        "period_start": period_start,
        "period_end": period_end,
        "opening_balance": None,
        "closing_balance": closing_balance,
    }
    return {
        "metadata": metadata,
        "movements": sorted(movements, key=lambda movement: (movement.booking_date, movement.concept), reverse=True),
    }


def sync_open_banking_connection(connection: BankConnection) -> dict:
    if not connection.active:
        return {"imported_statements": 0, "external_accounts": 0}
    if not connection.requisition_id:
        raise ValidationError("La conexion todavia no tiene una requisicion activa.")

    access_token = get_gocardless_access_token()
    requisition_payload = gocardless_api_request(
        "GET",
        f"/requisitions/{connection.requisition_id}/",
        access_token=access_token,
    )
    connection.requisition_status = requisition_payload.get("status", connection.requisition_status)
    accounts = requisition_payload.get("accounts", [])

    imported_statements = 0
    imported_accounts = 0
    now = timezone.now()

    with transaction.atomic():
        for account_id in accounts:
            details_response = gocardless_api_request(
                "GET",
                f"/accounts/{account_id}/details/",
                access_token=access_token,
            )
            details = details_response.get("account", details_response)
            external_account = upsert_external_account(connection, account_id, details)
            imported_accounts += 1

            if not external_account.is_active:
                continue

            balances_payload = gocardless_api_request(
                "GET",
                f"/accounts/{account_id}/balances/",
                access_token=access_token,
            )
            transactions_payload = gocardless_api_request(
                "GET",
                f"/accounts/{account_id}/transactions/",
                access_token=access_token,
            )
            booked_transactions = transactions_payload.get("transactions", {}).get("booked", [])
            parsed_movements = [parse_open_banking_transaction(item) for item in booked_transactions]
            closing_balance = pick_closing_balance(balances_payload)

            month_map: dict[str, list[ParsedMovement]] = defaultdict(list)
            for movement in parsed_movements:
                month_map[month_label_for_date(movement.booking_date)].append(movement)

            for month_label, month_movements in month_map.items():
                checksum = build_open_banking_statement_checksum(external_account, month_label)
                statement_import, _created = BankStatementImport.objects.get_or_create(
                    file_checksum=checksum,
                    defaults={
                        "connection": connection,
                        "external_account": external_account,
                        "ownership_category": external_account.ownership_category,
                        "import_source": BankStatementImport.ImportSource.OPEN_BANKING,
                        "institution": external_account.institution or connection.institution_name,
                        "account_label": external_account.account_label,
                        "iban": external_account.iban,
                        "statement_kind": external_account.statement_kind,
                        "currency": external_account.currency or "EUR",
                        "holder_name": external_account.holder_name,
                        "source_filename": f"{connection.institution_name}-{external_account.account_label}-{month_label}.openbanking",
                    },
                )
                statement_import.connection = connection
                statement_import.external_account = external_account
                statement_import.ownership_category = external_account.ownership_category
                statement_import.import_source = BankStatementImport.ImportSource.OPEN_BANKING
                statement_import.institution = external_account.institution or connection.institution_name
                statement_import.account_label = external_account.account_label
                statement_import.iban = external_account.iban
                statement_import.statement_kind = external_account.statement_kind
                statement_import.currency = external_account.currency or "EUR"
                statement_import.holder_name = external_account.holder_name
                statement_import.source_filename = (
                    f"{connection.institution_name}-{external_account.account_label}-{month_label}.openbanking"
                )[:255]
                closing_balance_for_month = closing_balance if month_label == max(month_map.keys()) else None
                parsed_payload = build_open_banking_parsed_payload(
                    external_account=external_account,
                    movements=month_movements,
                    closing_balance=closing_balance_for_month,
                )
                apply_parsed_statement(statement_import, parsed_payload)
                imported_statements += 1

            if closing_balance is not None and external_account.statement_kind == BankStatementImport.StatementKind.ACCOUNT:
                manual_account = BankBalance.objects.filter(
                    institution=external_account.institution or connection.institution_name,
                    account_name=external_account.account_label,
                ).first()
                if manual_account:
                    manual_account.ownership_category = external_account.ownership_category
                    manual_account.current_balance = closing_balance
                    if not manual_account.notes.strip():
                        manual_account.notes = "Saldo sincronizado automaticamente por Open Banking."
                    manual_account.save(update_fields=["ownership_category", "current_balance", "notes", "updated_at"])
                else:
                    BankBalance.objects.create(
                        ownership_category=external_account.ownership_category,
                        institution=external_account.institution or connection.institution_name,
                        account_name=external_account.account_label,
                        deposited_amount=ZERO,
                        current_balance=closing_balance,
                        annual_interest_income=ZERO,
                        notes="Saldo sincronizado automaticamente por Open Banking.",
                    )

            external_account.last_imported_at = now
            external_account.save(update_fields=["last_imported_at", "updated_at"])

        connection.last_synced_at = now
        connection.last_error = ""
        connection.save(update_fields=["requisition_status", "last_synced_at", "last_error", "updated_at"])

    return {
        "imported_statements": imported_statements,
        "external_accounts": imported_accounts,
    }


def sync_open_banking_connections() -> dict:
    summary = {"connections": 0, "external_accounts": 0, "imported_statements": 0, "errors": []}
    for connection in BankConnection.objects.filter(active=True):
        try:
            result = sync_open_banking_connection(connection)
        except Exception as exc:
            connection.last_error = str(exc)
            connection.save(update_fields=["last_error", "updated_at"])
            summary["errors"].append(f"{connection.institution_name}: {exc}")
            continue
        summary["connections"] += 1
        summary["external_accounts"] += result["external_accounts"]
        summary["imported_statements"] += result["imported_statements"]
    return summary


def build_open_banking_dashboard() -> dict:
    connections = list(BankConnection.objects.prefetch_related("external_accounts").order_by("institution_name", "id"))
    external_accounts = list(BankExternalAccount.objects.select_related("connection").order_by("institution", "account_label", "id"))
    summary = {
        "connections_count": len(connections),
        "active_connections_count": sum(1 for connection in connections if connection.active),
        "external_accounts_count": len(external_accounts),
        "cards_count": sum(
            1 for account in external_accounts if account.statement_kind == BankStatementImport.StatementKind.CARD
        ),
        "accounts_count": sum(
            1 for account in external_accounts if account.statement_kind == BankStatementImport.StatementKind.ACCOUNT
        ),
        "last_synced_at": max((connection.last_synced_at for connection in connections if connection.last_synced_at), default=None),
    }
    return {
        "summary": summary,
        "connections": connections,
        "external_accounts": external_accounts,
    }
