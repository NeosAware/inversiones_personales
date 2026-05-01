import io
import json
import os
import tempfile
from calendar import monthrange
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch
from urllib.error import HTTPError

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from portfolio.ownership import AssetOwnershipCategory

from .expert_consensus import attach_expert_consensus_to_dashboard, build_bridgewater_signal
from .broker_costs import estimate_broker_costs
from .llm_analysis import build_card_llm_context, enrich_dashboard_with_ai_analysis
from .management.commands.import_monica_equity_positions import MONICA_EQUITY_POSITIONS
from .models import (
    EquityClosedPosition,
    EquityExpectationReview,
    EquityNightlyAnalysisRun,
    EquityNightlyAnalysisSnapshot,
    EquityOptimizationRun,
    EquityPosition,
    EquityPriceHistory,
    EquityPurchaseForecastBaseline,
    EquityTicketSnapshot,
)
from .nightly_analysis import (
    build_positions_analysis_signature,
    build_dashboard_from_nightly_cache,
    build_current_dashboard_llm_summary,
    build_ibex_recommendation_date_map,
    build_nightly_completion_note,
    capture_purchase_forecast_baseline,
    load_cached_ibex_card,
    persist_nightly_analysis_dashboard,
    run_nightly_equity_analysis,
    serialize_cached_value,
)
from .optimization_runs import launch_equity_optimization_run, launch_equity_optimization_run_pair, process_equity_optimization_run
from .optimization_runs import (
    build_scheduled_optimization_persistence_context,
    launch_scheduled_equity_optimization_runs,
    purge_stale_scheduled_optimization_runs,
)
from .services import (
    EURIBOR_REFERENCE_NAME,
    EURIBOR_REFERENCE_SYMBOL,
    MarketSeries,
    MarketHistoryPoint,
    SPAIN_ELECTRICITY_DEMAND_NAME,
    SPAIN_ELECTRICITY_DEMAND_SYMBOL,
    SPAIN_GAS_CONSUMPTION_NAME,
    SPAIN_GAS_CONSUMPTION_SYMBOL,
    ZERO,
    apply_expert_consensus_adjustments_to_dashboard,
    apply_expectation_review_memory_to_card,
    apply_optimizer_expectation_review_adjustment,
    apply_news_context_adjustments_to_dashboard,
    build_candidate_purchase_timing_plan,
    build_equity_allocation_plan,
    build_equity_analysis_dashboard,
    build_equity_decision_rows,
    build_equity_history_cards,
    build_equity_investment_journey_context,
    build_optimizer_expectation_review_signal,
    build_reference_correlation,
    build_reference_cycle_template_from_series,
    build_candlestick_metrics,
    build_cycle_zoomed_monthly_projection_path,
    build_five_year_cycle_projection,
    build_equity_round_investment_plan,
    build_equity_sale_preview,
    build_expectation_review_dashboard,
    build_scenario_expectation_table,
    build_equity_ticket_tracking_context,
    build_equity_ticket_tracking_item,
    build_optimizer_purchase_discipline_review,
    build_ticket_expected_series,
    densify_projected_tracking_series,
    build_portfolio_correlation_context,
    build_portfolio_expectation_horizons,
    build_tracking_rebased_comparison_series,
    build_value_tracking_chart,
    build_equity_optimizer_candidate,
    build_owned_cycle_trade_timing_plan,
    add_calendar_months,
    archive_equity_position_sale,
    build_trade_alert,
    build_reference_suggestions_for_equity,
    clear_market_data_caches,
    capture_equity_ticket_snapshots,
    reconcile_trade_alert_with_expected_return,
    refresh_card_projection_visuals,
    synchronize_projection_path_with_cycle_zoom,
    fetch_market_series,
    filter_positive_optimizer_candidates,
    find_equity_company_profile,
    format_axis_value,
    format_percentage_axis_value,
    load_ibex_reference_workbook_snapshot,
    sync_equity_market_data,
)


def build_test_reference_workbook() -> str:
    from openpyxl import Workbook

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.append(["Ticker", "Empresa", "Sector", "Rent. 2025", "Ind. referencia", "Fuente", "Correl.", "PER 2025e", "Div. yield", "Capitaliz. (MrdEUR)", "Notas", "Ticker Yahoo"])
    summary.append(["SAN", "Banco Santander", "Banca", 0.923, "Euribor 12m / Tipos BCE", "BCE", 0.78, 7.2, 0.048, 92, "Banco de prueba", "SAN.MC"])

    quotes = workbook.create_sheet("Cotizaciones")
    quotes.append(["Ticker", "Empresa", "Sector", "2019", "2020", "2021", "2022", "2023", "2024", "2025"])
    quotes.append(["SAN", "Banco Santander", "Banca", 4.18, 2.31, 3.08, 2.83, 4.06, 5.42, 10.42])

    indicators = workbook.create_sheet("Indicadores")
    indicators.append(["Indicador", "Fuente", "Sector relacionado", "2019", "2020", "2021", "2022", "2023", "2024", "2025", "Var. 3a", "Var. total"])
    indicators.append(["Euribor 12m (%)", "BCE", "Banca", -0.24, -0.50, -0.50, 2.59, 4.16, 2.70, 2.10, "", ""])
    indicators.append(["IBEX 35 (puntos)", "BME", "Todos", 9549, 8073, 8713, 8229, 10102, 11595, 17315, "", ""])
    indicators.append(["Precio Brent (USD/barril prom.)", "ICE", "Consumo", 64, 42, 70, 101, 82, 79, 74, "", ""])

    correlations = workbook.create_sheet("Correlaciones")
    correlations.append(["Sector \\ Indicador", "Euribor 12m", "Brent"])
    correlations.append(["Banca", 0.78, 0.15])

    workbook.save(path)
    workbook.close()
    return path


def build_compound_market_series(
    symbol: str,
    name: str,
    growth: Decimal,
    months: int = 36,
    start_year: int = 2023,
    start_month: int = 1,
    start_price: Decimal = Decimal("10.0000"),
) -> MarketSeries:
    points = []
    price = start_price
    for index in range(months):
        month_number = (start_month - 1) + index
        year = start_year + (month_number // 12)
        month = (month_number % 12) + 1
        month_end = monthrange(year, month)[1]
        price = (price * growth).quantize(Decimal("0.0001"))
        points.append(
            {
                "date": date(year, month, month_end),
                "open": price,
                "high": (price * Decimal("1.0100")).quantize(Decimal("0.0001")),
                "low": (price * Decimal("0.9900")).quantize(Decimal("0.0001")),
                "close": price,
            }
        )

    return MarketSeries(
        symbol=symbol,
        name=name,
        latest_price=points[-1]["close"],
        latest_date=points[-1]["date"],
        points=points,
    )


def build_market_series_from_monthly_factors(
    symbol: str,
    name: str,
    monthly_factors: list[Decimal | str],
    *,
    start_year: int = 2000,
    start_month: int = 1,
    start_price: Decimal = Decimal("100.0000"),
) -> MarketSeries:
    points = []
    price = start_price.quantize(Decimal("0.0001"))
    month_offset = 0
    points.append(
        {
            "date": date(start_year, start_month, monthrange(start_year, start_month)[1]),
            "open": price,
            "high": price,
            "low": price,
            "close": price,
        }
    )
    for raw_factor in monthly_factors:
        factor = Decimal(str(raw_factor))
        price = (price * factor).quantize(Decimal("0.0001"))
        month_offset += 1
        month_number = (start_month - 1) + month_offset
        year = start_year + (month_number // 12)
        month = (month_number % 12) + 1
        month_end = monthrange(year, month)[1]
        points.append(
            {
                "date": date(year, month, month_end),
                "open": price,
                "high": price,
                "low": price,
                "close": price,
            }
        )

    return MarketSeries(
        symbol=symbol,
        name=name,
        latest_price=points[-1]["close"],
        latest_date=points[-1]["date"],
        points=points,
    )


def populate_position_history(
    position: EquityPosition,
    *,
    growth: Decimal = Decimal("1.0150"),
    benchmark_growth: Decimal = Decimal("1.0080"),
    months: int = 60,
):
    stock_series = build_compound_market_series(
        position.quote_symbol or f"{position.ticker}.MC",
        position.company_name,
        growth,
        months=months,
        start_year=2021,
        start_month=1,
        start_price=position.average_cost_per_share or Decimal("10.0000"),
    )
    benchmark_series = build_compound_market_series(
        position.benchmark_symbol or "^IBEX",
        position.benchmark_name or "IBEX 35",
        benchmark_growth,
        months=months,
        start_year=2021,
        start_month=1,
        start_price=Decimal("100.0000"),
    )
    for stock_point, benchmark_point in zip(stock_series.points, benchmark_series.points):
        EquityPriceHistory.objects.create(
            position=position,
            price_date=stock_point["date"],
            open_price=stock_point["open"],
            high_price=stock_point["high"],
            low_price=stock_point["low"],
            close_price=stock_point["close"],
            benchmark_close=benchmark_point["close"],
        )
    position.current_price_per_share = stock_series.latest_price
    position.latest_price_date = stock_series.latest_date
    position.save(update_fields=["current_price_per_share", "latest_price_date"])
    return stock_series


def populate_position_history_from_closes(
    position: EquityPosition,
    closes: list[Decimal | str],
    *,
    benchmark_closes: list[Decimal | str] | None = None,
    start_year: int = 2025,
    start_month: int = 1,
):
    normalized_closes = [Decimal(str(value)).quantize(Decimal("0.0001")) for value in closes]
    if benchmark_closes is None:
        benchmark_price = Decimal("100.0000")
        benchmark_closes = []
        for index in range(len(normalized_closes)):
            if index == 0:
                benchmark_closes.append(benchmark_price)
                continue
            benchmark_growth = Decimal("1.0060") if index % 2 else Decimal("0.9970")
            benchmark_price = (benchmark_price * benchmark_growth).quantize(Decimal("0.0001"))
            benchmark_closes.append(benchmark_price)
    normalized_benchmark_closes = [
        Decimal(str(value)).quantize(Decimal("0.0001")) for value in benchmark_closes
    ]

    for index, (close_price, benchmark_close) in enumerate(
        zip(normalized_closes, normalized_benchmark_closes)
    ):
        month_number = (start_month - 1) + index
        year = start_year + (month_number // 12)
        month = (month_number % 12) + 1
        month_end = monthrange(year, month)[1]
        EquityPriceHistory.objects.create(
            position=position,
            price_date=date(year, month, month_end),
            open_price=close_price,
            high_price=(close_price * Decimal("1.0100")).quantize(Decimal("0.0001")),
            low_price=(close_price * Decimal("0.9900")).quantize(Decimal("0.0001")),
            close_price=close_price,
            benchmark_close=benchmark_close,
        )

    position.current_price_per_share = normalized_closes[-1]
    position.latest_price_date = date(
        start_year + ((start_month - 1 + len(normalized_closes) - 1) // 12),
        ((start_month - 1 + len(normalized_closes) - 1) % 12) + 1,
        monthrange(
            start_year + ((start_month - 1 + len(normalized_closes) - 1) // 12),
            ((start_month - 1 + len(normalized_closes) - 1) % 12) + 1,
        )[1],
    )
    position.save(update_fields=["current_price_per_share", "latest_price_date"])
    return normalized_closes


class FakeHTTPResponse(io.BytesIO):
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()
        return False


@override_settings(EQUITIES_FETCH_FUNDAMENTALS=False)
class EquitiesServicesTests(TestCase):
    def tearDown(self):
        load_ibex_reference_workbook_snapshot.cache_clear()
        clear_market_data_caches()
        super().tearDown()

    def test_find_equity_company_profile_by_ibex_name_returns_indra_defaults(self):
        profile = find_equity_company_profile("Indra")

        self.assertIsNotNone(profile)
        self.assertEqual(profile["ticker"], "IDR")
        self.assertEqual(profile["quote_symbol"], "IDR.MC")
        self.assertEqual(profile["default_reference"]["benchmark_name"], "IBEX 35")

    def test_build_reference_suggestions_for_bank_prioritizes_euribor(self):
        suggestions = build_reference_suggestions_for_equity("Banco Santander", "SAN")

        self.assertGreaterEqual(len(suggestions), 2)
        self.assertEqual(suggestions[0]["benchmark_name"], EURIBOR_REFERENCE_NAME)
        self.assertEqual(suggestions[0]["benchmark_symbol"], EURIBOR_REFERENCE_SYMBOL)

    def test_build_reference_suggestions_for_enagas_prioritizes_gas_consumption(self):
        suggestions = build_reference_suggestions_for_equity("Enagas", "ENG")

        self.assertGreaterEqual(len(suggestions), 2)
        self.assertEqual(suggestions[0]["benchmark_name"], SPAIN_GAS_CONSUMPTION_NAME)
        self.assertEqual(suggestions[0]["benchmark_symbol"], SPAIN_GAS_CONSUMPTION_SYMBOL)

    def test_reference_correlation_uses_absolute_changes_for_euribor_series(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="SAN",
            quote_symbol="SAN.MC",
            reference_profile=EquityPosition.ReferenceProfile.EURIBOR_12M,
            benchmark_symbol=EURIBOR_REFERENCE_SYMBOL,
            benchmark_name=EURIBOR_REFERENCE_NAME,
            company_name="Banco Santander, S.A.",
            shares=Decimal("20"),
            average_cost_per_share=Decimal("3.0000"),
            current_price_per_share=Decimal("3.0000"),
        )
        stock_price = Decimal("3.0000")
        euribor_value = Decimal("-0.45")
        stock_return_pattern = [Decimal("2.8"), Decimal("-0.6"), Decimal("3.1"), Decimal("0.9")]
        euribor_delta_pattern = [Decimal("0.08"), Decimal("-0.02"), Decimal("0.09"), Decimal("0.01")]
        for index in range(36):
            year = 2023 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            stock_price = (stock_price * (Decimal("1.00") + (stock_return_pattern[index % 4] / Decimal("100")))).quantize(Decimal("0.0001"))
            euribor_value = (euribor_value + euribor_delta_pattern[index % 4]).quantize(Decimal("0.0001"))
            position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=stock_price,
                benchmark_close=euribor_value,
            )

        correlation = build_reference_correlation(list(position.price_history.order_by("price_date")), position)

        self.assertEqual(correlation["change_mode"], "absolute")
        self.assertIsNotNone(correlation["coefficient"])
        self.assertIsNotNone(correlation["beta"])

    def test_five_year_projection_ignores_open_month_noise(self):
        position = EquityPosition.objects.create(
            broker="Seguimiento",
            ticker="EST",
            quote_symbol="EST.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Estable SA",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )

        monthly_history = []
        price = Decimal("10.0000")
        for index in range(63):
            month_number = index
            year = 2021 + (month_number // 12)
            month = (month_number % 12) + 1
            price = (price * Decimal("1.0060")).quantize(Decimal("0.0001"))
            monthly_history.append(
                MarketHistoryPoint(
                    price_date=date(year, month, monthrange(year, month)[1]),
                    close_price=price,
                )
            )
        self.assertEqual(monthly_history[-1].price_date, date(2026, 3, 31))

        noisy_drop = monthly_history + [
            MarketHistoryPoint(price_date=date(2026, 4, 24), close_price=Decimal("6.0000")),
        ]
        noisy_rebound = monthly_history + [
            MarketHistoryPoint(price_date=date(2026, 4, 28), close_price=Decimal("14.0000")),
        ]
        correlation = {
            "coefficient": Decimal("0.40"),
            "observations_count": 60,
            "stability_gap": Decimal("0.05"),
            "stability_label": "Estable",
        }

        with (
            patch("equities.services.django_timezone.localdate", return_value=date(2026, 4, 28)),
            patch("equities.services.build_multifactor_reference_projection_bundle", return_value={"available": False}),
        ):
            drop_projection = build_five_year_cycle_projection(noisy_drop, position, correlation, include_visuals=False)
            rebound_projection = build_five_year_cycle_projection(noisy_rebound, position, correlation, include_visuals=False)

        self.assertTrue(drop_projection["uses_completed_month_anchor"])
        self.assertTrue(rebound_projection["uses_completed_month_anchor"])
        self.assertEqual(drop_projection["latest_date"], date(2026, 3, 31))
        self.assertEqual(rebound_projection["latest_date"], date(2026, 3, 31))
        self.assertEqual(drop_projection["annual_return_pct"], rebound_projection["annual_return_pct"])
        self.assertEqual(drop_projection["five_year_return_pct"], rebound_projection["five_year_return_pct"])
        self.assertIn("ultimo mes cerrado", drop_projection["explanation"])

    def test_five_year_expectation_review_waits_for_stable_sample(self):
        signal = {
            "available": True,
            "horizon_years": 5,
            "sample_count": 2,
            "latest_return_pct": Decimal("-4.00"),
            "average_return_pct": Decimal("2.00"),
            "trend_return_pct": Decimal("-12.00"),
            "recent_delta_pct": Decimal("-12.00"),
            "spread_pct": Decimal("12.00"),
            "first_review_date": date(2026, 4, 24),
            "latest_review_date": date(2026, 4, 28),
            "date_span_days": 4,
        }

        adjusted = apply_optimizer_expectation_review_adjustment(Decimal("8.00"), signal)

        self.assertTrue(adjusted["deferred_for_stability"])
        self.assertEqual(adjusted["adjusted_return_pct"], Decimal("8.00"))
        self.assertEqual(adjusted["adjustment_pct"], ZERO)

    def test_expectation_review_signal_measures_recent_real_deviation(self):
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 24),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
        )
        reviews = []
        for index, review_date in enumerate((date(2026, 4, 10), date(2026, 4, 17), date(2026, 4, 24)), start=1):
            reviews.append(
                EquityExpectationReview.objects.create(
                    run=run,
                    analysis_date=review_date,
                    review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
                    scope=EquityExpectationReview.Scope.IBEX,
                    analysis_key=f"ibex:SAN:{index}",
                    ticker="SAN",
                    quote_symbol="SAN.MC",
                    company_name="Banco Santander",
                    current_price=Decimal("100.0000"),
                    expected_return_pct_1y=Decimal("-5.00"),
                    expected_return_pct_5y=Decimal("-10.00"),
                )
            )

        signal = build_optimizer_expectation_review_signal(
            reviews,
            current_price=Decimal("98.0000"),
            current_date=date(2026, 5, 1),
        )

        feedback = signal["1y"]["reality_feedback"]
        self.assertTrue(feedback["available"])
        self.assertEqual(feedback["sample_count"], 3)
        self.assertLess(feedback["latest_gap_pct"], ZERO)
        self.assertLess(feedback["bias_adjustment_pct"], ZERO)
        self.assertGreater(feedback["mean_absolute_error_pct"], Decimal("1.00"))

    def test_expectation_review_signal_uses_full_historical_memory_not_only_recent_points(self):
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 3, 13),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
        )
        reviews = []
        for index in range(10):
            review_date = date(2026, 1, 3) + timedelta(days=7 * index)
            reviews.append(
                EquityExpectationReview.objects.create(
                    run=run,
                    analysis_date=review_date,
                    review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
                    scope=EquityExpectationReview.Scope.IBEX,
                    analysis_key=f"ibex:SAN:history:{index}",
                    ticker="SAN",
                    quote_symbol="SAN.MC",
                    company_name="Banco Santander",
                    current_price=Decimal("100.0000"),
                    expected_return_pct_1y=Decimal("10.00"),
                    expected_return_pct_5y=Decimal("25.00"),
                )
            )

        signal = build_optimizer_expectation_review_signal(
            reviews,
            current_price=Decimal("94.0000"),
            current_date=date(2026, 5, 1),
        )

        self.assertEqual(signal["1y"]["sample_count"], 10)
        feedback = signal["1y"]["reality_feedback"]
        self.assertEqual(feedback["sample_count"], 10)
        self.assertEqual(feedback["oldest_analysis_date"], date(2026, 1, 3))
        self.assertGreaterEqual(feedback["memory_span_days"], 63)
        self.assertLess(feedback["average_gap_pct"], ZERO)

    def test_expectation_review_signal_uses_matured_market_history_when_available(self):
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2025, 1, 2),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
        )
        review = EquityExpectationReview.objects.create(
            run=run,
            analysis_date=date(2024, 1, 2),
            review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
            scope=EquityExpectationReview.Scope.IBEX,
            analysis_key="ibex:SAN:matured",
            ticker="SAN",
            quote_symbol="SAN.MC",
            company_name="Banco Santander",
            current_price=Decimal("100.0000"),
            expected_return_pct_1y=Decimal("10.00"),
            expected_return_pct_5y=Decimal("25.00"),
        )

        signal = build_optimizer_expectation_review_signal(
            [review],
            current_price=Decimal("150.0000"),
            current_date=date(2026, 5, 1),
            market_points=[
                {"date": date(2024, 1, 2), "close": Decimal("100.0000")},
                {"date": date(2025, 1, 1), "close": Decimal("90.0000")},
            ],
        )

        feedback = signal["1y"]["reality_feedback"]
        self.assertTrue(feedback["available"])
        self.assertEqual(feedback["sample_count"], 1)
        self.assertEqual(feedback["latest_gap_pct"], Decimal("-20.00"))
        self.assertEqual(feedback["rows"][0]["actual_price"], Decimal("90.0000"))

    def test_expectation_memory_adjustment_cools_new_projection_with_weekly_miss(self):
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 24),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
        )
        for index, review_date in enumerate((date(2026, 4, 10), date(2026, 4, 17), date(2026, 4, 24)), start=1):
            EquityExpectationReview.objects.create(
                run=run,
                analysis_date=review_date,
                review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
                scope=EquityExpectationReview.Scope.IBEX,
                analysis_key=f"ibex:SAN:{index}",
                ticker="SAN",
                quote_symbol="SAN.MC",
                company_name="Banco Santander",
                current_price=Decimal("100.0000"),
                expected_return_pct_1y=Decimal("-5.00"),
                expected_return_pct_5y=Decimal("-10.00"),
            )
        position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ticker="SAN",
            quote_symbol="SAN.MC",
            company_name="Banco Santander",
            shares=ZERO,
            average_cost_per_share=ZERO,
            current_price_per_share=Decimal("98.0000"),
        )
        card = {
            "position": position,
            "end_date": date(2026, 5, 1),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Vigilar", "tone": "watch", "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "latest_price": Decimal("98.0000"),
                "base_return_pct": ZERO,
                "price_return_pct": ZERO,
                "low_return_pct": Decimal("-8.00"),
                "high_return_pct": Decimal("8.00"),
                "price_low_return_pct": Decimal("-8.00"),
                "price_high_return_pct": Decimal("8.00"),
                "projected_price": Decimal("98.0000"),
                "confidence_label": "Alta",
                "confidence_score_pct": Decimal("82.00"),
                "safety_score": Decimal("76.00"),
                "net_income_yield_pct": ZERO,
                "transaction_drag_pct": ZERO,
                "annualized_volatility_pct": Decimal("18.00"),
                "positive_year_ratio_pct": Decimal("52.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Transicion",
                "current_drawdown_pct": Decimal("-4.00"),
                "max_drawdown_pct": Decimal("-25.00"),
            },
            "cycle_projection_5y": {"available": False},
            "six_month_snapshot": {"available": False},
            "period_snapshots": [],
            "correlation": {},
            "relative_trend": {},
            "valuation": {},
            "technical_signal": {"available": False},
        }

        apply_expectation_review_memory_to_card(card)

        memory = card["projection"]["historical_memory_adjustment"]
        self.assertTrue(memory["applied"])
        self.assertLess(card["projection"]["base_return_pct"], ZERO)
        self.assertLess(card["projection"]["projected_price"], Decimal("98.0000"))
        self.assertTrue(memory["reality_feedback"]["available"])
        self.assertLess(card["projection_reliability"]["score"], Decimal("82.00"))

    def test_purchase_discipline_penalizes_unreliable_model_memory(self):
        position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ticker="SAN",
            quote_symbol="SAN.MC",
            company_name="Banco Santander",
            shares=ZERO,
            average_cost_per_share=ZERO,
            current_price_per_share=Decimal("10.0000"),
        )
        base_candidate = {
            "position": position,
            "sector_label": "Banca",
            "primary_signal_pct": Decimal("8.00"),
            "scenario_expected_return_pct": Decimal("12.00"),
            "annualized_target_gap_pct": Decimal("4.00"),
            "safety_score": Decimal("78.00"),
            "reliability_score": Decimal("82.00"),
            "low_return_pct": Decimal("-6.00"),
            "downside_stress_return_pct": Decimal("-5.00"),
            "annualized_volatility_pct": Decimal("14.00"),
            "uncertainty_penalty_pct": Decimal("0.80"),
            "holding_annualized_return_pct": Decimal("24.00"),
            "annualized_target_return_pct": Decimal("20.00"),
            "purchase_timing": {"available": True, "mode_label": "Comprar ya"},
        }
        reliable_candidate = {
            **base_candidate,
            "expectation_review_signal": {
                "1y": {
                    "reality_feedback": {
                        "available": True,
                        "sample_count": 12,
                        "mean_absolute_error_pct": Decimal("1.20"),
                        "average_gap_pct": Decimal("0.40"),
                        "direction_hit_rate_pct": Decimal("74.00"),
                    }
                }
            },
        }
        unreliable_candidate = {
            **base_candidate,
            "expectation_review_signal": {
                "1y": {
                    "reality_feedback": {
                        "available": True,
                        "sample_count": 12,
                        "mean_absolute_error_pct": Decimal("8.00"),
                        "average_gap_pct": Decimal("-7.00"),
                        "direction_hit_rate_pct": Decimal("30.00"),
                    }
                }
            },
        }

        reliable = build_optimizer_purchase_discipline_review(reliable_candidate)
        unreliable = build_optimizer_purchase_discipline_review(unreliable_candidate)

        self.assertGreater(reliable["score"], unreliable["score"])
        self.assertEqual(unreliable["memory_label"], "Memoria penaliza")
        self.assertLess(unreliable["memory_score"], Decimal("58.00"))

    def test_cycle_projection_5y_builds_multifactor_model_with_bce_forward_signal(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            reference_profile=EquityPosition.ReferenceProfile.SPAIN_HOUSE_PRICE,
            benchmark_symbol="EUROSTAT:prc_hpi_q:ES:TOTAL:I15_Q",
            benchmark_name="Precio vivienda Espana",
            company_name="ACS, Actividades de Construccion y Servicios, S.A.",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("20.0000"),
            current_price_per_share=Decimal("20.0000"),
        )
        stock_price = Decimal("20.0000")
        benchmark_price = Decimal("100.0000")
        stock_return_pattern = [Decimal("3.2"), Decimal("-0.7"), Decimal("2.8"), Decimal("0.9")]
        benchmark_return_pattern = [Decimal("1.1"), Decimal("0.2"), Decimal("1.0"), Decimal("0.4")]
        for index in range(120):
            year = 2016 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            stock_price = (stock_price * (Decimal("1.00") + (stock_return_pattern[index % 4] / Decimal("100")))).quantize(Decimal("0.0001"))
            benchmark_price = (benchmark_price * (Decimal("1.00") + (benchmark_return_pattern[index % 4] / Decimal("100")))).quantize(Decimal("0.0001"))
            position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=stock_price,
                benchmark_close=benchmark_price,
            )

        def build_quarterly_series(symbol: str, name: str, start_value: Decimal, growth_pattern: list[Decimal]) -> MarketSeries:
            points = []
            value = start_value
            for index in range(40):
                year = 2016 + (index // 4)
                quarter = (index % 4) + 1
                month = quarter * 3
                month_end = {3: 31, 6: 30, 9: 30, 12: 31}[month]
                value = (value * growth_pattern[index % len(growth_pattern)]).quantize(Decimal("0.0001"))
                points.append(
                    {
                        "date": date(year, month, month_end),
                        "open": value,
                        "high": value,
                        "low": value,
                        "close": value,
                    }
                )
            return MarketSeries(symbol=symbol, name=name, latest_price=points[-1]["close"], latest_date=points[-1]["date"], points=points)

        def build_monthly_absolute_series(symbol: str, name: str, start_value: Decimal, deltas: list[Decimal]) -> MarketSeries:
            points = []
            value = start_value
            for index in range(120):
                year = 2016 + (index // 12)
                month = (index % 12) + 1
                month_end = monthrange(year, month)[1]
                value = (value + deltas[index % len(deltas)]).quantize(Decimal("0.0001"))
                points.append(
                    {
                        "date": date(year, month, month_end),
                        "open": value,
                        "high": value,
                        "low": value,
                        "close": value,
                    }
                )
            return MarketSeries(symbol=symbol, name=name, latest_price=points[-1]["close"], latest_date=points[-1]["date"], points=points)

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            if reference_profile == EquityPosition.ReferenceProfile.SPAIN_HOUSE_PRICE:
                return build_quarterly_series(
                    benchmark_symbol,
                    benchmark_name,
                    Decimal("100.0000"),
                    [Decimal("1.0180"), Decimal("1.0100"), Decimal("1.0160"), Decimal("1.0080")],
                )
            if reference_profile == EquityPosition.ReferenceProfile.EURIBOR_12M:
                return build_monthly_absolute_series(
                    benchmark_symbol,
                    benchmark_name,
                    Decimal("-0.4000"),
                    [Decimal("0.08"), Decimal("-0.02"), Decimal("0.10"), Decimal("0.01")],
                )
            return build_compound_market_series(
                benchmark_symbol or "^IBEX",
                benchmark_name or "IBEX 35",
                growth=Decimal("1.0080"),
                months=120,
                start_year=2016,
                start_month=1,
                start_price=Decimal("100.0000"),
            )

        forward_signal_series = MarketSeries(
            symbol="ECB:YC.B.U2.EUR.4F.G_N_C.SV_C_YM.IF_5Y",
            name="Curva BCE forward 5A",
            latest_price=Decimal("2.4000"),
            latest_date=date(2025, 12, 31),
            points=[
                {
                    "date": date(2025, 12, 31),
                    "open": None,
                    "high": None,
                    "low": None,
                    "close": Decimal("2.4000"),
                }
            ],
        )

        with (
            patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series),
            patch("equities.services.fetch_ecb_yield_curve_series", return_value=forward_signal_series),
        ):
            card = build_equity_history_cards([position])[0]

        cycle_projection = card["cycle_projection_5y"]
        self.assertTrue(cycle_projection["available"])
        self.assertTrue(cycle_projection["factor_model_available"])
        self.assertGreaterEqual(len(cycle_projection["factors"]), 2)
        self.assertGreater(cycle_projection["factor_blend_ratio_pct"], ZERO)
        self.assertGreaterEqual(cycle_projection["forward_signal_count"], 1)
        self.assertTrue(cycle_projection["comparison_chart"]["available"])

        euribor_factor = next(
            factor for factor in cycle_projection["factors"] if factor["reference_label"] == EURIBOR_REFERENCE_NAME
        )
        self.assertTrue(euribor_factor["forward_signal"]["available"])
        self.assertEqual(euribor_factor["change_unit_label"], "pp")
        self.assertIsNotNone(euribor_factor["projected_change_5y"])
        self.assertTrue(euribor_factor["projection_chart"]["available"])
        self.assertTrue(euribor_factor["coefficient_chart"]["available"])
        self.assertEqual(len(euribor_factor["correction_rows"]), 5)
        self.assertTrue(any(row["anchor_path_value"] is not None for row in euribor_factor["correction_rows"]))
        self.assertIsNotNone(euribor_factor["weight_inputs"]["coefficient_component"])

    def test_reference_cycle_template_detects_similar_ibex_shape_in_long_history(self):
        recent_shape_factors = [
            Decimal("0.9650"),
            Decimal("0.9820"),
            Decimal("1.0180"),
            Decimal("1.0240"),
            Decimal("0.9890"),
            Decimal("1.0210"),
            Decimal("1.0170"),
            Decimal("0.9780"),
            Decimal("1.0120"),
            Decimal("1.0260"),
            Decimal("0.9920"),
            Decimal("1.0150"),
            Decimal("1.0080"),
            Decimal("0.9810"),
            Decimal("1.0200"),
            Decimal("1.0140"),
            Decimal("0.9870"),
            Decimal("1.0190"),
        ]
        expected_step_returns = [
            Decimal("-6.00"),
            Decimal("14.00"),
            Decimal("5.00"),
            Decimal("-3.00"),
            Decimal("11.00"),
            Decimal("6.00"),
            Decimal("-2.50"),
            Decimal("9.00"),
            Decimal("4.50"),
            Decimal("8.00"),
        ]

        def expand_half_year_returns(step_returns: list[Decimal]) -> list[Decimal]:
            monthly_factors = []
            for step_return in step_returns:
                monthly_factor = Decimal(str(round((1 + (float(step_return) / 100)) ** (1 / 6), 8)))
                monthly_factors.extend([monthly_factor] * 6)
            return monthly_factors

        prefix_noise = [Decimal("1.0040"), Decimal("0.9980"), Decimal("1.0060"), Decimal("0.9970")] * 6
        bridge_noise = [Decimal("1.0120"), Decimal("0.9890"), Decimal("1.0110"), Decimal("0.9920")] * 6
        monthly_factors = (
            prefix_noise
            + recent_shape_factors
            + expand_half_year_returns(expected_step_returns)
            + bridge_noise
            + recent_shape_factors
        )
        reference_series = build_market_series_from_monthly_factors(
            "^IBEX",
            "IBEX 35",
            monthly_factors,
            start_year=2010,
            start_month=1,
            start_price=Decimal("100.0000"),
        )
        candidate_anchor_date = reference_series.points[len(prefix_noise) + len(recent_shape_factors)]["date"]

        with patch("equities.services.REFERENCE_CYCLE_TEMPLATE_MAX_MATCHES", 1):
            template = build_reference_cycle_template_from_series(
                reference_series,
                latest_date=reference_series.latest_date,
                years=5,
                step_months=6,
            )

        self.assertTrue(template["available"])
        self.assertEqual(template["selected_windows_count"], 1)
        self.assertEqual(template["selected_anchor_dates"], [candidate_anchor_date])
        self.assertEqual(template["recent_months"], 18)
        self.assertEqual(template["shape_window_end_date"], reference_series.latest_date)
        self.assertGreater(template["years_covered"], Decimal("11.00"))
        for actual_return, expected_return in zip(template["step_return_pcts"], expected_step_returns):
            self.assertLess(abs(actual_return - expected_return), Decimal("0.20"))

    def test_cycle_projection_5y_uses_longest_reference_history_for_ibex_shape(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="REP",
            quote_symbol="REP.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Repsol, S.A.",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("12.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        stock_price = Decimal("12.00")
        benchmark_price = Decimal("100.00")
        pattern = [Decimal("1.025"), Decimal("0.985"), Decimal("1.030"), Decimal("0.992")]
        benchmark_pattern = [Decimal("1.010"), Decimal("1.004"), Decimal("1.012"), Decimal("0.998")]
        for index in range(120):
            year = 2016 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            stock_price = (stock_price * pattern[index % len(pattern)]).quantize(Decimal("0.0001"))
            benchmark_price = (benchmark_price * benchmark_pattern[index % len(benchmark_pattern)]).quantize(Decimal("0.0001"))
            position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=stock_price,
                benchmark_close=benchmark_price,
            )

        long_reference_series = build_market_series_from_monthly_factors(
            "^IBEX",
            "IBEX 35",
            ([Decimal("1.0180"), Decimal("0.9840"), Decimal("1.0220"), Decimal("0.9910"), Decimal("1.0160"), Decimal("0.9870")] * 44)[:264],
            start_year=2000,
            start_month=1,
            start_price=Decimal("100.0000"),
        )

        with (
            patch("equities.services.fetch_reference_series_for_choice", return_value=long_reference_series),
            patch("equities.services.fetch_ecb_yield_curve_series", return_value=None),
        ):
            card = build_equity_history_cards([position])[0]

        cycle_projection = card["cycle_projection_5y"]
        reference_template = cycle_projection["reference_cycle_template"]
        self.assertTrue(cycle_projection["available"])
        self.assertTrue(reference_template["available"])
        self.assertEqual(reference_template["reference_label"], "IBEX 35")
        self.assertGreater(reference_template["years_covered"], Decimal("20.00"))
        self.assertGreater(reference_template["years_covered"], cycle_projection["analysis_years_used"])
        self.assertEqual(reference_template["shape_window_end_date"], long_reference_series.latest_date)
        self.assertIn("ventanas historicas parecidas", cycle_projection["explanation"])
        self.assertEqual(cycle_projection["history_window_label"], "Ultimos 5 anos visibles")

    def test_equity_history_card_exposes_12m_and_5y_scenarios(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position, months=96)

        card = build_equity_history_cards([position])[0]

        self.assertTrue(card["projection"]["available"])
        self.assertEqual([row["label"] for row in card["projection"]["scenarios"]], ["Bajista", "Base", "Alcista"])
        self.assertEqual(len(card["cycle_projection_5y"]["scenarios"]), 3)
        self.assertIsNotNone(card["projection"]["band_pct"])
        self.assertIsNotNone(card["cycle_projection_5y"]["scenario_spread_annual_pct"])
        self.assertGreater(
            sum((row["probability_pct"] for row in card["projection"]["scenarios"]), ZERO),
            Decimal("99.0"),
        )
        self.assertTrue(card["scenario_tables"]["projection_12m"]["available"])
        self.assertTrue(card["scenario_tables"]["cycle_5y"]["available"])
        self.assertEqual(
            sum((row["contribution_return_pct"] for row in card["scenario_tables"]["projection_12m"]["rows"]), ZERO),
            card["scenario_tables"]["projection_12m"]["expected_return_pct"],
        )

    def test_scenario_expectation_table_calculates_weighted_hope_from_probability_times_return(self):
        table = build_scenario_expectation_table(
            [
                {"key": "bear", "label": "Bajista", "probability_pct": Decimal("33.0"), "total_return_pct": Decimal("-20.0"), "projected_price": Decimal("8.0000")},
                {"key": "base", "label": "Base", "probability_pct": Decimal("33.0"), "total_return_pct": Decimal("-5.0"), "projected_price": Decimal("9.5000")},
                {"key": "bull", "label": "Alcista", "probability_pct": Decimal("33.0"), "total_return_pct": Decimal("10.0"), "projected_price": Decimal("11.0000")},
            ],
            return_key="total_return_pct",
            fallback_value=Decimal("-5.0"),
        )

        self.assertTrue(table["available"])
        self.assertEqual(table["expected_return_pct"], Decimal("-5.00"))
        self.assertEqual(
            [row["contribution_return_pct"] for row in table["rows"]],
            [Decimal("-6.67"), Decimal("-1.67"), Decimal("3.34")],
        )
        self.assertEqual(table["expected_projected_price"], Decimal("9.5000"))

    def test_material_news_event_penalizes_projection_confidence_and_widens_ranges(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="REP",
            quote_symbol="REP.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Repsol",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("20.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position, growth=Decimal("1.0180"), benchmark_growth=Decimal("1.0070"), months=108)

        dashboard = build_equity_analysis_dashboard([position])
        card = dashboard["history_cards"][0]
        original_confidence_score = card["projection"]["confidence_score_pct"]
        original_band_pct = card["projection"]["band_pct"]
        original_reliability_score = card["projection_reliability"]["score"]
        original_cycle_spread = card["cycle_projection_5y"]["scenario_spread_annual_pct"]

        card["news_context"] = {
            "available": True,
            "label": "Contexto adverso",
            "score": Decimal("-4.20"),
            "items_count": 4,
            "top_tags": ["geopolitica", "energia"],
            "material_event": True,
            "material_note": "El shock geopolitico altera el escenario base y obliga a penalizar confianza.",
            "note": "Shock adverso de energia y geostrategia.",
            "company_signal": {"available": True, "label": "Empresa Adversa", "score": Decimal("-2.00"), "items": []},
            "sector_signal": {"available": True, "label": "Sector Adversa", "score": Decimal("-1.60"), "items": []},
            "market_signal": {"available": True, "label": "Mercado Adversa", "score": Decimal("-3.20"), "items": []},
            "top_items": [],
            "captured_at_label": "2026-04-18 01:10",
        }

        summary = apply_news_context_adjustments_to_dashboard(dashboard)
        adjusted_card = dashboard["history_cards"][0]

        self.assertEqual(summary["adjusted_cards_count"], 1)
        self.assertTrue(adjusted_card["projection"]["news_adjustment"]["applied"])
        self.assertTrue(adjusted_card["cycle_projection_5y"]["news_adjustment"]["applied"])
        self.assertLess(adjusted_card["projection"]["confidence_score_pct"], original_confidence_score)
        self.assertGreater(adjusted_card["projection"]["band_pct"], original_band_pct)
        self.assertLess(adjusted_card["projection_reliability"]["score"], original_reliability_score)
        self.assertGreater(adjusted_card["cycle_projection_5y"]["scenario_spread_annual_pct"], original_cycle_spread)
        self.assertTrue(adjusted_card["projection_12m_chart"]["available"])
        self.assertTrue(adjusted_card["presentation_projection"]["available"])
        self.assertEqual(
            adjusted_card["presentation_projection"]["visible_projected_price"],
            adjusted_card["projection"]["monthly_path"][-1]["projected_price"],
        )
        self.assertGreater(
            adjusted_card["projection"]["scenarios"][0]["probability_pct"],
            adjusted_card["projection"]["scenarios"][-1]["probability_pct"],
        )

    def test_attach_expert_consensus_weights_current_signal_by_historical_accuracy(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("34.0000"),
            current_price_per_share=Decimal("34.0000"),
            annual_dividend_income=Decimal("12.00"),
            annual_maintenance_cost=Decimal("3.00"),
        )
        populate_position_history(position, growth=Decimal("1.0140"), benchmark_growth=Decimal("1.0060"), months=96)
        dashboard = build_equity_analysis_dashboard([position])
        history_card = dashboard["history_cards"][0]

        run_a = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2025, 9, 1),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            completed_at=timezone.now(),
        )
        run_b = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2025, 9, 2),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            completed_at=timezone.now(),
        )
        EquityNightlyAnalysisSnapshot.objects.create(
            run=run_a,
            analysis_date=run_a.analysis_date,
            scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
            analysis_key="ibex:ACS:jpmorgan",
            ticker="ACS",
            quote_symbol="ACS.MC",
            company_name="ACS",
            status_key="ibex",
            sector_label="Infraestructuras",
            agent_provider="core",
            analysis_payload=serialize_cached_value(
                {
                    "expert_consensus": {
                        "company_signal": {
                            "items": [
                                {
                                    "title": "JPMorgan refuerza su recomendacion sobre ACS",
                                    "expert_source": "JPMorgan",
                                    "source_key": "jpmorgan",
                                    "score": Decimal("2.40"),
                                    "published_on": "2025-09-01",
                                    "target_symbol": "ACS.MC",
                                }
                            ]
                        }
                    }
                }
            ),
        )
        EquityNightlyAnalysisSnapshot.objects.create(
            run=run_b,
            analysis_date=run_b.analysis_date,
            scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
            analysis_key="ibex:ACS:redburn",
            ticker="ACS",
            quote_symbol="ACS.MC",
            company_name="ACS",
            status_key="ibex",
            sector_label="Infraestructuras",
            agent_provider="core",
            analysis_payload=serialize_cached_value(
                {
                    "expert_consensus": {
                        "company_signal": {
                            "items": [
                                {
                                    "title": "Redburn se pone bajista con ACS",
                                    "expert_source": "Redburn",
                                    "source_key": "redburn",
                                    "score": Decimal("-2.40"),
                                    "published_on": "2025-09-02",
                                    "target_symbol": "ACS.MC",
                                }
                            ]
                        }
                    }
                }
            ),
        )

        current_items = [
            {
                "title": "JPMorgan mantiene compra en ACS",
                "description": "El banco mantiene una lectura positiva.",
                "link": "https://example.com/jpmorgan-acs",
                "source": "Reuters",
                "expert_source": "JPMorgan",
                "source_key": "jpmorgan",
                "published_at": timezone.make_aware(datetime(2026, 4, 24, 9, 0)),
                "published_label": "2026-04-24",
                "published_on": "2026-04-24",
                "captured_on": "2026-04-25",
                "score": Decimal("2.40"),
                "tone": "positive",
                "target_symbol": "ACS.MC",
                "target_label": "ACS",
            },
            {
                "title": "Redburn insiste en vender ACS",
                "description": "La firma mantiene una lectura mas defensiva.",
                "link": "https://example.com/redburn-acs",
                "source": "Bloomberg",
                "expert_source": "Redburn",
                "source_key": "redburn",
                "published_at": timezone.make_aware(datetime(2026, 4, 24, 8, 0)),
                "published_label": "2026-04-24",
                "published_on": "2026-04-24",
                "captured_on": "2026-04-25",
                "score": Decimal("-2.40"),
                "tone": "negative",
                "target_symbol": "ACS.MC",
                "target_label": "ACS",
            },
        ]

        def fake_market_series(symbol: str, range_key: str = "max", interval: str = "1d"):
            if symbol in {"^GSPC", "^IXIC"}:
                return MarketSeries(
                    symbol=symbol,
                    name=symbol,
                    latest_price=Decimal("5500.0000"),
                    latest_date=date(2026, 4, 25),
                    points=[
                        {"date": date(2025, 4, 25), "close": Decimal("5000.0000")},
                        {"date": date(2025, 10, 25), "close": Decimal("5200.0000")},
                        {"date": date(2026, 1, 25), "close": Decimal("5300.0000")},
                        {"date": date(2026, 4, 25), "close": Decimal("5500.0000")},
                    ],
                )
            return MarketSeries(
                symbol=symbol,
                name=symbol,
                latest_price=Decimal("12.2000"),
                latest_date=date(2026, 1, 2),
                points=[
                    {"date": date(2025, 9, 1), "close": Decimal("10.0000")},
                    {"date": date(2025, 9, 2), "close": Decimal("10.0500")},
                    {"date": date(2025, 12, 30), "close": Decimal("12.1000")},
                    {"date": date(2026, 1, 2), "close": Decimal("12.2000")},
                ],
            )

        with (
            patch("equities.expert_consensus.fetch_company_expert_items", return_value=current_items),
            patch("equities.expert_consensus.fetch_market_expert_items", return_value=[]),
            patch("equities.expert_consensus.fetch_market_series", side_effect=fake_market_series),
            patch(
                "equities.expert_consensus.build_bridgewater_signal",
                return_value={
                    "available": True,
                    "label": "Bridgewater favorable",
                    "score": Decimal("2.10"),
                    "quality_score": Decimal("60.00"),
                    "quality_label": "Media",
                    "items_count": 1,
                    "positive_count": 1,
                    "negative_count": 0,
                    "neutral_count": 0,
                    "note": "Bridgewater ve un entorno de soft landing.",
                    "items": [
                        {
                            "title": "Bridgewater Daily Observations",
                            "expert_source": "Bridgewater",
                            "source_key": "bridgewater",
                            "score": Decimal("1.40"),
                            "tone": "positive",
                            "published_on": "2026-04-24",
                            "target_symbol": "^IBEX",
                        }
                    ],
                    "source_rows": [
                        {
                            "source": "Bridgewater",
                            "source_key": "bridgewater",
                            "quality_score": Decimal("60.00"),
                            "quality_label": "Media",
                            "source_weight": Decimal("0.98"),
                            "observations_count": 0,
                            "hit_rate_pct": None,
                            "current_items_count": 1,
                            "current_score": Decimal("1.40"),
                            "weighted_score": Decimal("1.37"),
                        }
                    ],
                },
            ),
        ):
            summary = attach_expert_consensus_to_dashboard(dashboard)

        self.assertEqual(summary["ranked_sources_count"], 2)
        self.assertGreater(history_card["expert_consensus"]["score"], ZERO)
        self.assertTrue(history_card["expert_consensus"]["wall_street_signal"]["available"])
        self.assertGreater(history_card["expert_consensus"]["wall_street_signal"]["score"], ZERO)
        self.assertTrue(history_card["expert_consensus"]["bridgewater_signal"]["available"])
        self.assertEqual(history_card["expert_consensus"]["bridgewater_signal"]["label"], "Bridgewater favorable")
        self.assertEqual(history_card["expert_consensus"]["source_rows"][0]["source"], "JPMorgan")
        self.assertGreater(
            history_card["expert_consensus"]["source_rows"][0]["quality_score"],
            history_card["expert_consensus"]["source_rows"][1]["quality_score"],
        )
        self.assertIn("JPMorgan", history_card["expert_consensus"]["best_sources"])

    def test_build_bridgewater_signal_reads_local_reports(self):
        fd, pdf_path = tempfile.mkstemp(suffix=".pdf")
        os.close(fd)
        report_path = Path(pdf_path)
        try:
            with (
                override_settings(
                    EQUITIES_BRIDGEWATER_REPORT_PATHS=[str(report_path)],
                    EQUITIES_BRIDGEWATER_REPORT_DIRS=[str(report_path.parent)],
                    EQUITIES_BRIDGEWATER_SCAN_LIMIT=4,
                ),
                patch(
                    "equities.expert_consensus.read_pdf_pages",
                    return_value=[
                        (
                            "Bridgewater Daily Observations. "
                            "U.S. equities remain attractive in a soft landing with disinflation. "
                            "Wall Street risk assets could outperform if productivity stays firm."
                        )
                    ],
                ),
            ):
                signal = build_bridgewater_signal()
        finally:
            report_path.unlink(missing_ok=True)

        self.assertTrue(signal["available"])
        self.assertEqual(signal["label"], "Bridgewater favorable")
        self.assertEqual(signal["source_rows"][0]["source"], "Bridgewater")
        self.assertEqual(signal["items"][0]["expert_source"], "Bridgewater")
        self.assertGreater(signal["wall_street_score"], ZERO)

    def test_expert_consensus_can_penalize_projection_when_it_conflicts_with_model(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="REP",
            quote_symbol="REP.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Repsol",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("20.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position, growth=Decimal("1.0180"), benchmark_growth=Decimal("1.0070"), months=108)

        dashboard = build_equity_analysis_dashboard([position])
        card = dashboard["history_cards"][0]
        original_price_return_pct = card["projection"]["price_return_pct"]
        original_band_pct = card["projection"]["band_pct"]
        original_reliability_score = card["projection_reliability"]["score"]
        original_cycle_spread = card["cycle_projection_5y"]["scenario_spread_annual_pct"]

        card["expert_consensus"] = {
            "available": True,
            "label": "Consenso experto adverso",
            "score": Decimal("-4.10"),
            "quality_score": Decimal("82.00"),
            "quality_label": "Alta",
            "items_count": 4,
            "note": "Las fuentes con mejor track record siguen viendo un sesgo bajista.",
            "best_sources": ["JPMorgan", "Morgan Stanley"],
            "source_rows": [],
            "company_signal": {"available": True, "label": "Empresa Consenso adverso", "score": Decimal("-4.30"), "items": []},
            "market_signal": {"available": True, "label": "Mercado Consenso mixto", "score": Decimal("-2.20"), "items": []},
            "top_items": [],
            "captured_at_label": "2026-04-25 01:20",
        }

        summary = apply_expert_consensus_adjustments_to_dashboard(dashboard)
        adjusted_card = dashboard["history_cards"][0]

        self.assertEqual(summary["adjusted_cards_count"], 1)
        self.assertTrue(adjusted_card["projection"]["expert_adjustment"]["applied"])
        self.assertTrue(adjusted_card["cycle_projection_5y"]["expert_adjustment"]["applied"])
        self.assertLess(adjusted_card["projection"]["price_return_pct"], original_price_return_pct)
        self.assertGreater(adjusted_card["projection"]["band_pct"], original_band_pct)
        self.assertLess(adjusted_card["projection_reliability"]["score"], original_reliability_score)
        self.assertGreater(adjusted_card["cycle_projection_5y"]["scenario_spread_annual_pct"], original_cycle_spread)
        self.assertTrue(adjusted_card["projection_12m_chart"]["available"])
        self.assertTrue(adjusted_card["presentation_projection"]["available"])
        self.assertEqual(
            adjusted_card["presentation_projection"]["visible_projected_price"],
            adjusted_card["projection"]["monthly_path"][-1]["projected_price"],
        )

    def test_information_basis_summary_highlights_geopolitical_and_expert_sources(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("34.0000"),
            current_price_per_share=Decimal("34.0000"),
        )
        populate_position_history(position, growth=Decimal("1.0160"), benchmark_growth=Decimal("1.0070"), months=84)
        card = build_equity_history_cards([position])[0]
        card["news_context"] = {
            "available": True,
            "label": "Contexto adverso",
            "score": Decimal("-3.20"),
            "items_count": 2,
            "top_tags": ["geopolitica", "energia"],
            "material_event": True,
            "material_note": "La tension geopolitica reciente obliga a revisar el timing.",
            "note": "Empresa adversa | mercado adversa",
            "top_items": [
                {
                    "title": "ACS cae por tension geopolitica y repunte del crudo",
                    "source": "Reuters",
                    "published_label": "2026-04-26",
                    "tone": "negative",
                    "score": Decimal("-2.40"),
                    "tags": ["geopolitica", "energia"],
                }
            ],
        }
        card["expert_consensus"] = {
            "available": True,
            "label": "Consenso experto mixto",
            "score": Decimal("1.80"),
            "quality_score": Decimal("78.00"),
            "quality_label": "Alta",
            "items_count": 2,
            "best_sources": ["JPMorgan"],
            "note": "Las fuentes mejor rankeadas siguen viendo valor aunque con mas cautela.",
            "top_items": [
                {
                    "title": "JPMorgan mantiene compra sobre ACS pero enfria el timing",
                    "source": "Reuters",
                    "expert_source": "JPMorgan",
                    "published_label": "2026-04-26",
                    "tone": "positive",
                    "score": Decimal("1.90"),
                    "tags": [],
                }
            ],
            "wall_street_signal": {
                "available": True,
                "label": "Wall Street favorable",
                "score": Decimal("2.10"),
                "note": "Wall Street sigue empujando el apetito por riesgo.",
                "top_tags": [],
                "items": [
                    {
                        "published_label": "2026-04-26",
                    }
                ],
                "source_rows": [
                    {
                        "source": "S&P 500",
                    }
                ],
            },
            "bridgewater_signal": {
                "available": True,
                "label": "Bridgewater mixta",
                "score": Decimal("-0.40"),
                "note": "Bridgewater mantiene una lectura macro menos limpia.",
                "top_tags": ["tipos"],
                "items": [
                    {
                        "published_label": "2026-04-26",
                    }
                ],
                "source_rows": [
                    {
                        "source": "Bridgewater",
                    }
                ],
            },
        }

        refresh_card_projection_visuals(card, history=[])

        self.assertTrue(card["information_basis"]["available"])
        self.assertTrue(card["information_basis"]["geopolitical_flag"])
        self.assertTrue(card["information_basis"]["macro_flag"])
        self.assertIn("Reuters", card["information_basis"]["source_labels"])
        self.assertIn("JPMorgan via Reuters", card["information_basis"]["source_labels"])
        self.assertIn("S&P 500", card["information_basis"]["source_labels"])
        self.assertIn("timing", card["information_basis"]["summary"].lower())
        self.assertTrue(card["information_basis"]["bullet_points"])
        self.assertIn("Reuters", card["information_basis"]["geopolitical_sources"])
        self.assertIn("S&P 500", card["information_basis"]["macro_sources"])

    def test_build_reference_suggestions_for_iberdrola_prioritizes_electricity_demand(self):
        suggestions = build_reference_suggestions_for_equity("Iberdrola", "IBE")

        self.assertGreaterEqual(len(suggestions), 2)
        self.assertEqual(suggestions[0]["benchmark_name"], SPAIN_ELECTRICITY_DEMAND_NAME)
        self.assertEqual(suggestions[0]["benchmark_symbol"], SPAIN_ELECTRICITY_DEMAND_SYMBOL)

    def test_sync_equity_market_data_updates_latest_price_and_history(self):
        position = EquityPosition.objects.create(
            broker="Banco Sabadell",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola S.A.",
            shares=Decimal("100"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )

        stock_series = MarketSeries(
            symbol="IBE.MC",
            name="Iberdrola S.A.",
            latest_price=Decimal("19.2500"),
            latest_date=date(2026, 3, 22),
            points=[
                {
                    "date": date(2026, 3, 20),
                    "open": Decimal("17.8000"),
                    "high": Decimal("18.4000"),
                    "low": Decimal("17.7000"),
                    "close": Decimal("18.0000"),
                },
                {
                    "date": date(2026, 3, 21),
                    "open": Decimal("18.3000"),
                    "high": Decimal("19.2000"),
                    "low": Decimal("18.1000"),
                    "close": Decimal("19.0000"),
                },
            ],
        )
        benchmark_series = MarketSeries(
            symbol="^IBEX",
            name="IBEX 35",
            latest_price=Decimal("16714.0000"),
            latest_date=date(2026, 3, 22),
            points=[
                {"date": date(2026, 3, 20), "close": Decimal("16500.0000")},
                {"date": date(2026, 3, 21), "close": Decimal("16600.0000")},
            ],
        )

        with (
            patch("equities.services.fetch_market_series", return_value=stock_series),
            patch("equities.services.fetch_reference_series", return_value=benchmark_series),
        ):
            sync_equity_market_data(position)

        position.refresh_from_db()
        self.assertEqual(position.current_price_per_share, Decimal("19.2500"))
        self.assertEqual(position.latest_price_date, date(2026, 3, 22))
        self.assertEqual(position.price_history.count(), 2)
        latest_point = position.price_history.order_by("-price_date").first()
        self.assertEqual(latest_point.open_price, Decimal("18.3000"))
        self.assertEqual(latest_point.high_price, Decimal("19.2000"))
        self.assertEqual(latest_point.low_price, Decimal("18.1000"))
        self.assertEqual(latest_point.close_price, Decimal("19.0000"))
        self.assertEqual(latest_point.benchmark_close, Decimal("16600.0000"))

    def test_build_equity_history_cards_returns_relative_performance(self):
        position = EquityPosition.objects.create(
            broker="Banco Sabadell",
            ticker="ENG",
            quote_symbol="ENG.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Enagas, S.A.",
            shares=Decimal("50"),
            average_cost_per_share=Decimal("6.8700"),
            current_price_per_share=Decimal("14.7000"),
        )
        position.price_history.create(price_date=date(2026, 3, 20), close_price=Decimal("10.0000"), benchmark_close=Decimal("100.0000"))
        position.price_history.create(price_date=date(2026, 3, 21), close_price=Decimal("12.0000"), benchmark_close=Decimal("110.0000"))

        cards = build_equity_history_cards([position])

        self.assertEqual(len(cards), 1)
        self.assertTrue(cards[0]["has_history"])
        self.assertEqual(cards[0]["stock_return_pct"], Decimal("20.00"))
        self.assertEqual(cards[0]["benchmark_return_pct"], Decimal("10.00"))
        self.assertTrue(cards[0]["stock_line"])
        self.assertEqual(cards[0]["selected_period"]["label"], "1Y")

    def test_santander_broker_costs_reduce_net_annual_income(self):
        position = EquityPosition.objects.create(
            broker="Banco Santander",
            trade_channel=EquityPosition.TradeChannel.APP,
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola S.A.",
            shares=Decimal("100"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
            annual_dividend_income=Decimal("40.00"),
            annual_maintenance_cost=Decimal("0.00"),
        )

        self.assertEqual(position.purchase_total_cost, Decimal("5.00"))
        self.assertEqual(position.estimated_broker_costs["annual_custody_cost"], Decimal("20.00"))
        self.assertEqual(position.estimated_broker_costs["annual_dividend_fee"], Decimal("2.00"))
        self.assertEqual(position.net_dividend_income, Decimal("38.00"))
        self.assertEqual(position.net_annual_income, Decimal("16.00"))

    def test_history_cards_include_suggested_reference_correlations(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="SAN",
            quote_symbol="SAN.MC",
            reference_profile=EquityPosition.ReferenceProfile.EURIBOR_12M,
            benchmark_symbol=EURIBOR_REFERENCE_SYMBOL,
            benchmark_name=EURIBOR_REFERENCE_NAME,
            company_name="Banco Santander, S.A.",
            shares=Decimal("30"),
            average_cost_per_share=Decimal("4.0000"),
            current_price_per_share=Decimal("5.0000"),
        )
        for price_date, close_price, benchmark_close in (
            (date(2026, 1, 31), Decimal("4.00"), Decimal("2.10")),
            (date(2026, 2, 28), Decimal("4.20"), Decimal("2.20")),
            (date(2026, 3, 31), Decimal("4.50"), Decimal("2.35")),
            (date(2026, 4, 30), Decimal("4.80"), Decimal("2.50")),
            (date(2026, 5, 31), Decimal("5.00"), Decimal("2.65")),
        ):
            position.price_history.create(
                price_date=price_date,
                close_price=close_price,
                benchmark_close=benchmark_close,
            )

        euribor_series = MarketSeries(
            symbol=EURIBOR_REFERENCE_SYMBOL,
            name=EURIBOR_REFERENCE_NAME,
            latest_price=Decimal("2.65"),
            latest_date=date(2026, 5, 1),
            points=[
                {"date": date(2026, 1, 1), "close": Decimal("2.10")},
                {"date": date(2026, 2, 1), "close": Decimal("2.20")},
                {"date": date(2026, 3, 1), "close": Decimal("2.35")},
                {"date": date(2026, 4, 1), "close": Decimal("2.50")},
                {"date": date(2026, 5, 1), "close": Decimal("2.65")},
            ],
        )
        generic_series = MarketSeries(
            symbol="^IBEX",
            name="IBEX 35",
            latest_price=Decimal("14000.00"),
            latest_date=date(2026, 5, 31),
            points=[
                {"date": date(2026, 1, 31), "close": Decimal("12000.00")},
                {"date": date(2026, 2, 28), "close": Decimal("12300.00")},
                {"date": date(2026, 3, 31), "close": Decimal("12700.00")},
                {"date": date(2026, 4, 30), "close": Decimal("13200.00")},
                {"date": date(2026, 5, 31), "close": Decimal("14000.00")},
            ],
        )

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            if reference_profile == EquityPosition.ReferenceProfile.EURIBOR_12M:
                return euribor_series
            return generic_series

        with patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series):
            cards = build_equity_history_cards([position])

        suggested_reference_names = [item["benchmark_name"] for item in cards[0]["suggested_references"]]
        self.assertIn(EURIBOR_REFERENCE_NAME, suggested_reference_names)
        euribor_reference = next(
            item for item in cards[0]["suggested_references"] if item["benchmark_name"] == EURIBOR_REFERENCE_NAME
        )
        self.assertIsNotNone(euribor_reference["correlation"]["coefficient"])
        self.assertTrue(cards[0]["best_correlation_chart"]["available"])
        self.assertIn(cards[0]["best_correlation_chart"]["reference_label"], suggested_reference_names)

    def test_history_cards_include_one_year_projection_from_six_months_and_reference(self):
        position = EquityPosition.objects.create(
            broker="Banco Sabadell",
            ticker="IDR",
            quote_symbol="IDR.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Indra Sistemas, S.A.",
            shares=Decimal("20"),
            average_cost_per_share=Decimal("18.0000"),
            current_price_per_share=Decimal("22.0000"),
        )
        for price_date, close_price, benchmark_close in (
            (date(2025, 12, 31), Decimal("15.00"), Decimal("100.00")),
            (date(2026, 1, 31), Decimal("16.00"), Decimal("102.00")),
            (date(2026, 2, 28), Decimal("17.20"), Decimal("104.50")),
            (date(2026, 3, 31), Decimal("18.60"), Decimal("107.00")),
            (date(2026, 4, 30), Decimal("19.80"), Decimal("109.80")),
            (date(2026, 5, 31), Decimal("21.10"), Decimal("112.40")),
            (date(2026, 6, 30), Decimal("22.00"), Decimal("115.00")),
        ):
            position.price_history.create(
                price_date=price_date,
                close_price=close_price,
                benchmark_close=benchmark_close,
            )

        cards = build_equity_history_cards([position])

        projection = cards[0]["projection"]
        self.assertTrue(projection["available"])
        self.assertGreater(projection["base_return_pct"], Decimal("0"))
        self.assertGreater(projection["projected_price"], Decimal("22.00"))
        self.assertEqual(len(projection["monthly_path"]), 12)
        self.assertEqual(len(projection["quarterly_path"]), 4)
        self.assertIsNotNone(projection["quarterly_path"][0]["projected_date"])
        monthly_deltas = [
            (current["projected_price"] - previous["projected_price"]).quantize(Decimal("0.0001"))
            for previous, current in zip(projection["monthly_path"], projection["monthly_path"][1:])
        ]
        self.assertGreater(len(set(monthly_deltas)), 1)
        self.assertIn("IBEX 35", projection["explanation"])
        self.assertTrue(cards[0]["projection_line"])
        self.assertTrue(cards[0]["historical_chart"]["available"])
        self.assertTrue(cards[0]["projection_12m_chart"]["available"])
        self.assertTrue(cards[0]["historical_chart"]["x_markers"])
        self.assertTrue(cards[0]["projection_12m_chart"]["x_markers"])

    def test_candlestick_metrics_detect_buy_bias_from_historical_velas(self):
        history = []
        start_date = date(2025, 1, 31)
        close_price = Decimal("9.80")
        for index in range(18):
            current_date = start_date + timedelta(days=index * 31)
            close_price = (close_price + Decimal("0.35")).quantize(Decimal("0.0001"))
            open_price = (close_price - Decimal("0.18")).quantize(Decimal("0.0001"))
            high_price = (close_price + Decimal("0.28")).quantize(Decimal("0.0001"))
            low_price = (open_price - Decimal("0.10")).quantize(Decimal("0.0001"))
            if index == 17:
                open_price = Decimal("15.4000")
                close_price = Decimal("16.4500")
                high_price = Decimal("16.8000")
                low_price = Decimal("15.2500")
            history.append(
                EquityPriceHistory(
                    price_date=current_date,
                    open_price=open_price,
                    high_price=high_price,
                    low_price=low_price,
                    close_price=close_price,
                )
            )

        metrics = build_candlestick_metrics(history)

        self.assertTrue(metrics["available"])
        self.assertGreater(metrics["signal_score"], ZERO)
        self.assertIn(metrics["signal_label"], {"Compra tecnica", "Sesgo comprador"})
        self.assertGreater(metrics["rsi_14"], Decimal("55.00"))
        self.assertNotEqual(metrics["breakout_label"], "Sin ruptura")

    def test_history_cards_include_technical_signal_and_projection_overlay(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="AMS",
            quote_symbol="AMS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Amadeus IT Group",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("54.0000"),
            current_price_per_share=Decimal("54.0000"),
        )
        stock_price = Decimal("42.00")
        benchmark_price = Decimal("100.00")
        for index in range(30):
            year = 2024 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            stock_price = (stock_price * Decimal("1.0240")).quantize(Decimal("0.0001"))
            benchmark_price = (benchmark_price * Decimal("1.0090")).quantize(Decimal("0.0001"))
            open_price = (stock_price * Decimal("0.9880")).quantize(Decimal("0.0001"))
            high_price = (stock_price * Decimal("1.0180")).quantize(Decimal("0.0001"))
            low_price = (open_price * Decimal("0.9920")).quantize(Decimal("0.0001"))
            if index == 29:
                open_price = (stock_price * Decimal("0.9760")).quantize(Decimal("0.0001"))
                high_price = (stock_price * Decimal("1.0320")).quantize(Decimal("0.0001"))
                low_price = (open_price * Decimal("0.9940")).quantize(Decimal("0.0001"))
            position.price_history.create(
                price_date=date(year, month, month_end),
                open_price=open_price,
                high_price=high_price,
                low_price=low_price,
                close_price=stock_price,
                benchmark_close=benchmark_price,
            )

        cards = build_equity_history_cards([position])

        technical_signal = cards[0]["technical_signal"]
        projection = cards[0]["projection"]
        self.assertTrue(technical_signal["available"])
        self.assertGreater(technical_signal["signal_score"], ZERO)
        self.assertTrue(projection["technical_adjustment"]["applied"])
        self.assertGreater(projection["technical_adjustment"]["return_adjustment_pct"], ZERO)
        self.assertEqual(cards[0]["trade_alert"]["technical_label"], technical_signal["signal_label"])

    def test_projection_chart_uses_last_year_while_history_keeps_full_cycle(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola, S.A.",
            shares=Decimal("20"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        stock_price = Decimal("10.00")
        benchmark_price = Decimal("100.00")
        for index in range(30):
            year = 2024 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            stock_price = (stock_price * Decimal("1.015")).quantize(Decimal("0.0001"))
            benchmark_price = (benchmark_price * Decimal("1.008")).quantize(Decimal("0.0001"))
            position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=stock_price,
                benchmark_close=benchmark_price,
            )

        cards = build_equity_history_cards([position])

        chart = cards[0]["projection_12m_chart"]
        self.assertTrue(chart["available"])
        self.assertEqual(chart["history_window_label"], "Ultimo ano visible")
        self.assertEqual(chart["start_label"], "2025-06-30")
        self.assertEqual(cards[0]["historical_chart"]["start_label"], "2024-01-31")
        self.assertGreater(cards[0]["historical_chart"]["points_count"], chart["points_count"])

    def test_cycle_zoomed_projection_path_preserves_first_year_shape_of_5y_cycle(self):
        path = build_cycle_zoomed_monthly_projection_path(
            Decimal("10.0000"),
            Decimal("11.5000"),
            anchor_date=date(2026, 4, 30),
            cycle_projection={
                "available": True,
                "path": [
                    {
                        "label": "6M",
                        "projected_date": date(2026, 10, 30),
                        "projected_price": Decimal("12.0000"),
                    },
                    {
                        "label": "1A",
                        "projected_date": date(2027, 4, 30),
                        "projected_price": Decimal("10.8000"),
                    },
                ],
            },
        )

        self.assertEqual(len(path), 12)
        self.assertEqual(path[-1]["label"], "1A")
        self.assertEqual(path[-1]["projected_price"], Decimal("11.5000"))
        month_6 = next(step for step in path if step["label"] == "6M")
        self.assertGreater(month_6["projected_price"], path[-1]["projected_price"])
        self.assertTrue(path[0]["projected_date"].isoformat().startswith("2026-05-"))

    def test_projection_sync_uses_first_year_of_5y_cycle_even_if_old_12m_target_was_positive(self):
        projection = {
            "available": True,
            "projected_price": Decimal("11.5000"),
            "price_return_pct": Decimal("15.00"),
            "base_return_pct": Decimal("16.20"),
            "band_pct": Decimal("10.00"),
            "confidence_label": "Media",
            "safety_score": Decimal("62.00"),
            "net_income_yield_pct": Decimal("1.50"),
            "transaction_drag_pct": Decimal("0.30"),
            "monthly_path": [],
            "quarterly_path": [],
        }
        cycle_projection = {
            "available": True,
            "model_window_label": "10.0 anos de historico",
            "path": [
                {
                    "label": "6M",
                    "projected_date": date(2026, 10, 30),
                    "projected_price": Decimal("9.1000"),
                },
                {
                    "label": "1A",
                    "projected_date": date(2027, 4, 30),
                    "projected_price": Decimal("8.8000"),
                },
                {
                    "label": "18M",
                    "projected_date": date(2027, 10, 30),
                    "projected_price": Decimal("9.4000"),
                },
            ],
        }

        synchronize_projection_path_with_cycle_zoom(
            projection,
            cycle_projection,
            current_price=Decimal("10.0000"),
            anchor_date=date(2026, 4, 30),
        )

        self.assertTrue(projection["uses_cycle_zoom_shape"])
        self.assertEqual(projection["projected_price"], Decimal("11.5000"))
        self.assertEqual(projection["monthly_path"][-1]["projected_price"], Decimal("8.8000"))
        self.assertTrue(projection["cycle_sync"]["applied"])
        self.assertEqual(projection["cycle_sync"]["projected_price"], Decimal("8.8000"))

    def test_projection_presentation_summary_uses_same_12m_path_shown_in_chart(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="SAB",
            quote_symbol="SAB.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Banco de Sabadell",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        populate_position_history(position, growth=Decimal("1.0100"), benchmark_growth=Decimal("1.0040"), months=36)
        card = build_equity_history_cards([position])[0]
        latest_date = card["end_date"]

        card["projection"].update(
            {
                "projected_price": Decimal("11.5000"),
                "price_return_pct": Decimal("15.00"),
                "base_return_pct": Decimal("16.20"),
                "net_income_yield_pct": Decimal("1.50"),
                "transaction_drag_pct": Decimal("0.30"),
                "monthly_path": [],
                "quarterly_path": [],
            }
        )
        card["cycle_projection_5y"] = {
            "available": True,
            "model_window_label": "10.0 anos de historico",
            "path": [
                {
                    "label": "6M",
                    "projected_date": add_calendar_months(latest_date, 6),
                    "projected_price": Decimal("9.1000"),
                },
                {
                    "label": "1A",
                    "projected_date": add_calendar_months(latest_date, 12),
                    "projected_price": Decimal("8.8000"),
                },
                {
                    "label": "18M",
                    "projected_date": add_calendar_months(latest_date, 18),
                    "projected_price": Decimal("9.4000"),
                },
            ],
        }

        synchronize_projection_path_with_cycle_zoom(
            card["projection"],
            card["cycle_projection_5y"],
            current_price=Decimal("10.0000"),
            anchor_date=latest_date,
        )
        refreshed_card = refresh_card_projection_visuals(card)

        self.assertTrue(refreshed_card["presentation_projection"]["available"])
        self.assertEqual(
            refreshed_card["presentation_projection"]["visible_projected_price"],
            Decimal("8.8000"),
        )
        self.assertLess(refreshed_card["presentation_projection"]["visible_total_return_pct"], ZERO)

    def test_refresh_card_projection_visuals_downgrades_buy_alert_when_visible_12m_closes_lower(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="ANA",
            quote_symbol="ANA.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Acciona",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        populate_position_history(position, growth=Decimal("1.0110"), benchmark_growth=Decimal("1.0040"), months=36)
        card = build_equity_history_cards([position])[0]
        latest_date = card["end_date"]
        card["projection"].update(
            {
                "projected_price": Decimal("12.8000"),
                "price_return_pct": Decimal("28.00"),
                "base_return_pct": Decimal("29.20"),
                "net_income_yield_pct": Decimal("1.50"),
                "transaction_drag_pct": Decimal("0.30"),
                "monthly_path": [
                    {
                        "label": "3M",
                        "projected_date": add_calendar_months(latest_date, 3),
                        "projected_price": Decimal("9.7000"),
                    },
                    {
                        "label": "6M",
                        "projected_date": add_calendar_months(latest_date, 6),
                        "projected_price": Decimal("9.3000"),
                    },
                    {
                        "label": "9M",
                        "projected_date": add_calendar_months(latest_date, 9),
                        "projected_price": Decimal("9.0000"),
                    },
                    {
                        "label": "12M",
                        "projected_date": add_calendar_months(latest_date, 12),
                        "projected_price": Decimal("8.8000"),
                    },
                ],
                "quarterly_path": [],
            }
        )
        card["trade_alert"] = {
            **card["trade_alert"],
            "label": "Comprar",
            "tone": "buy",
            "score": Decimal("4.25"),
            "trigger_label": "3 meses con alpha positiva",
            "note": "La pendiente relativa todavia apoya compras.",
        }

        refreshed_card = refresh_card_projection_visuals(card)

        self.assertEqual(refreshed_card["trade_alert"]["label"], "Vigilar")
        self.assertEqual(refreshed_card["trade_alert"]["tone"], "watch")
        self.assertEqual(refreshed_card["trade_alert"]["trigger_label"], "La senda visible 12M sigue bajista")
        self.assertTrue(refreshed_card["trade_alert"]["coherence_adjusted"])
        self.assertLess(refreshed_card["presentation_projection"]["visible_total_return_pct"], ZERO)
        self.assertLess(refreshed_card["presentation_projection"]["visible_price_return_pct"], ZERO)
        self.assertEqual(
            refreshed_card["presentation_projection"]["visible_projected_price"],
            Decimal("8.8000"),
        )
        self.assertIn(
            "misma senda naranja",
            refreshed_card["presentation_projection"]["consistency_note"].lower(),
        )
        self.assertTrue(refreshed_card["projection_12m_chart"]["available"])
        self.assertEqual(
            refreshed_card["projection_12m_chart"]["projection_end_label"],
            add_calendar_months(latest_date, 12).isoformat(),
        )

    def test_reconcile_trade_alert_with_expected_return_downgrades_buy_when_expectation_is_negative(self):
        adjusted_alert = reconcile_trade_alert_with_expected_return(
            {
                "label": "Comprar",
                "tone": "buy",
                "score": Decimal("4.25"),
                "trigger_label": "8 trimestres con alpha positiva",
                "note": "La pendiente relativa todavia apoya compras.",
            },
            expected_return_pct=Decimal("-24.00"),
            horizon_label="1A",
        )

        self.assertEqual(adjusted_alert["label"], "Vigilar")
        self.assertEqual(adjusted_alert["tone"], "watch")
        self.assertEqual(adjusted_alert["trigger_label"], "La esperanza 1A sigue negativa")
        self.assertTrue(adjusted_alert["coherence_adjusted"])
        self.assertIn("-24.0 %", adjusted_alert["note"])

    def test_cycle_projection_5y_uses_last_five_years_on_chart_and_ten_years_for_model(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="REP",
            quote_symbol="REP.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Repsol, S.A.",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("12.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        stock_price = Decimal("12.00")
        benchmark_price = Decimal("100.00")
        pattern = [Decimal("1.025"), Decimal("0.985"), Decimal("1.030"), Decimal("0.992")]
        benchmark_pattern = [Decimal("1.010"), Decimal("1.004"), Decimal("1.012"), Decimal("0.998")]
        for index in range(120):
            year = 2016 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            stock_price = (stock_price * pattern[index % len(pattern)]).quantize(Decimal("0.0001"))
            benchmark_price = (benchmark_price * benchmark_pattern[index % len(benchmark_pattern)]).quantize(Decimal("0.0001"))
            position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=stock_price,
                benchmark_close=benchmark_price,
            )

        reference_series = build_market_series_from_monthly_factors(
            "^IBEX",
            "IBEX 35",
            ([Decimal("1.0150"), Decimal("0.9860"), Decimal("1.0180"), Decimal("0.9930"), Decimal("1.0110"), Decimal("0.9890")] * 24)[:144],
            start_year=2008,
            start_month=1,
            start_price=Decimal("100.0000"),
        )

        with patch("equities.services.fetch_reference_series_for_choice", return_value=reference_series):
            cards = build_equity_history_cards([position])

        cycle_projection = cards[0]["cycle_projection_5y"]
        cycle_chart = cards[0]["cycle_projection_5y_chart"]
        self.assertTrue(cycle_projection["available"])
        self.assertTrue(cycle_chart["available"])
        self.assertEqual(cycle_chart["history_window_label"], "Ultimos 5 anos visibles")
        self.assertIn("anos de historico", cycle_chart["model_window_label"])
        self.assertGreater(cycle_projection["analysis_years_used"], Decimal("9.80"))
        self.assertEqual(cycle_chart["start_label"], "2021-01-31")
        self.assertTrue(cycle_chart["projection_end_label"].startswith("2030-"))
        self.assertEqual(cycle_projection["path"][-1]["label"], "5A")
        self.assertIn("historico disponible", cycle_projection["explanation"])
        self.assertTrue(cards[0]["projection"]["uses_cycle_zoom_shape"])
        self.assertEqual(cards[0]["projection_12m_chart"]["projection_window_label"], "Zoom 12M del patron 5A")
        self.assertEqual(cards[0]["projection_12m_chart"]["model_window_label"], cycle_chart["model_window_label"])
        first_year_step = next(step for step in cycle_projection["path"] if step["label"] == "1A")
        self.assertEqual(cards[0]["projection"]["monthly_path"][-1]["projected_price"], first_year_step["projected_price"])
        self.assertTrue(cards[0]["projection"]["cycle_sync"]["applied"])
        self.assertEqual(cards[0]["projection"]["cycle_sync"]["source_label"], "1A del patron 5A")
        path_deltas = [
            (current["projected_price"] - previous["projected_price"]).quantize(Decimal("0.01"))
            for previous, current in zip(cycle_projection["path"], cycle_projection["path"][1:])
        ]
        self.assertGreater(len(set(path_deltas)), 1)

    def test_decision_rows_include_five_year_projection_and_yearly_margins(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="REP",
            quote_symbol="REP.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Repsol, S.A.",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("100.0000"),
            current_price_per_share=Decimal("100.0000"),
        )

        row = build_equity_decision_rows(
            [
                {
                    "position": position,
                    "has_history": True,
                    "status_key": "ibex",
                    "status_label": "Solo radar",
                    "status_note": "",
                    "detail_anchor": "",
                    "reference_label": "IBEX 35",
                    "correlation": {"coefficient": Decimal("0.74")},
                    "projection": {
                        "available": True,
                        "base_return_pct": Decimal("12.50"),
                        "projected_price": Decimal("112.5000"),
                        "years_covered": Decimal("10.00"),
                        "scenarios": [
                            {"key": "bear", "label": "Bajista", "probability_pct": Decimal("50.0"), "total_return_pct": Decimal("-10.0")},
                            {"key": "base", "label": "Base", "probability_pct": Decimal("30.0"), "total_return_pct": Decimal("12.5")},
                            {"key": "bull", "label": "Alcista", "probability_pct": Decimal("20.0"), "total_return_pct": Decimal("30.0")},
                        ],
                        "safety_score": Decimal("68.00"),
                        "safety_label": "Alta",
                        "benefit_risk_ratio": Decimal("1.90"),
                        "cycle_phase": "Expansion",
                        "decision_score": Decimal("82.00"),
                    },
                    "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
                    "trade_alert": {
                        "label": "Comprar",
                        "tone": "buy",
                        "score": Decimal("75.00"),
                        "note": "Buena combinacion de retorno y seguridad.",
                        "trigger_label": "Retorno 12M y seguridad apoyan compra",
                    },
                    "coefficient_alert": {
                        "label": "Solido",
                        "tone": "good",
                        "trigger_label": "Coeficiente consistente",
                    },
                    "reference_playbook": {"best_candidate": None},
                    "cycle_projection_5y": {
                        "available": True,
                        "cycle_phase": "Expansion",
                        "five_year_return_pct": Decimal("61.0510"),
                        "scenarios": [
                            {
                                "key": "bear",
                                "label": "Bajista",
                                "probability_pct": Decimal("30.0"),
                                "annual_return_pct": Decimal("1.93"),
                                "year_2_return_pct": Decimal("4.0"),
                                "year_3_return_pct": Decimal("6.0"),
                                "year_4_return_pct": Decimal("8.0"),
                                "year_5_return_pct": Decimal("10.0"),
                                "five_year_return_pct": Decimal("10.0"),
                            },
                            {
                                "key": "base",
                                "label": "Base",
                                "probability_pct": Decimal("40.0"),
                                "annual_return_pct": Decimal("10.00"),
                                "year_2_return_pct": Decimal("21.0"),
                                "year_3_return_pct": Decimal("34.0"),
                                "year_4_return_pct": Decimal("47.0"),
                                "year_5_return_pct": Decimal("61.0510"),
                                "five_year_return_pct": Decimal("61.0510"),
                            },
                            {
                                "key": "bull",
                                "label": "Alcista",
                                "probability_pct": Decimal("30.0"),
                                "annual_return_pct": Decimal("13.68"),
                                "year_2_return_pct": Decimal("32.0"),
                                "year_3_return_pct": Decimal("50.0"),
                                "year_4_return_pct": Decimal("69.0"),
                                "year_5_return_pct": Decimal("90.0"),
                                "five_year_return_pct": Decimal("90.0"),
                            },
                        ],
                        "path": [
                            {"label": "6M", "projected_price": Decimal("104.8809")},
                            {"label": "1A", "projected_price": Decimal("110.0000")},
                            {"label": "18M", "projected_price": Decimal("115.3687")},
                            {"label": "2A", "projected_price": Decimal("121.0000")},
                            {"label": "30M", "projected_price": Decimal("126.9016")},
                            {"label": "3A", "projected_price": Decimal("133.1000")},
                            {"label": "42M", "projected_price": Decimal("139.5968")},
                            {"label": "4A", "projected_price": Decimal("146.4100")},
                            {"label": "54M", "projected_price": Decimal("153.5574")},
                            {"label": "5A", "projected_price": Decimal("161.0510")},
                        ],
                    },
                }
            ]
        )[0]

        self.assertEqual(row["cycle_return_5y_pct"], Decimal("61.05"))
        self.assertEqual(row["cycle_return_annual_pct"], Decimal("10.00"))
        self.assertEqual(row["expected_return_1y_pct"], Decimal("4.75"))
        self.assertEqual(row["expected_return_2y_pct"], Decimal("19.20"))
        self.assertEqual(row["expected_return_3y_pct"], Decimal("30.40"))
        self.assertEqual(row["expected_return_4y_pct"], Decimal("41.90"))
        self.assertEqual(row["expected_return_5y_pct"], Decimal("54.42"))
        self.assertEqual(
            [item["label"] for item in row["cycle_yearly_margins"]],
            ["AÑO 1", "AÑO 2", "AÑO 3", "AÑO 4", "AÑO 5"],
        )
        self.assertEqual(
            [item["margin_pct"] for item in row["cycle_yearly_margins"]],
            [Decimal("12.50"), Decimal("7.56"), Decimal("10.00"), Decimal("10.00"), Decimal("10.00")],
        )
        self.assertEqual(row["cycle_yearly_margins"][0]["margin_pct"], row["projected_return_pct"])

    def test_decision_rows_downgrade_buy_alert_when_expected_1y_return_is_negative(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("100.0000"),
            current_price_per_share=Decimal("100.0000"),
        )

        row = build_equity_decision_rows(
            [
                {
                    "position": position,
                    "has_history": True,
                    "status_key": "ibex",
                    "status_label": "Solo radar",
                    "status_note": "",
                    "detail_anchor": "",
                    "reference_label": "IBEX 35",
                    "correlation": {"coefficient": Decimal("0.74")},
                    "projection": {
                        "available": True,
                        "base_return_pct": Decimal("8.00"),
                        "projected_price": Decimal("108.0000"),
                        "years_covered": Decimal("10.00"),
                        "scenarios": [
                            {"key": "bear", "label": "Bajista", "probability_pct": Decimal("40.0"), "total_return_pct": Decimal("-30.0")},
                            {"key": "base", "label": "Base", "probability_pct": Decimal("40.0"), "total_return_pct": Decimal("-10.0")},
                            {"key": "bull", "label": "Alcista", "probability_pct": Decimal("20.0"), "total_return_pct": Decimal("8.0")},
                        ],
                        "safety_score": Decimal("68.00"),
                        "safety_label": "Alta",
                        "benefit_risk_ratio": Decimal("1.90"),
                        "cycle_phase": "Expansion",
                        "decision_score": Decimal("82.00"),
                    },
                    "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
                    "trade_alert": {
                        "label": "Comprar",
                        "tone": "buy",
                        "score": Decimal("75.00"),
                        "note": "La tendencia relativa sigue apoyando compra.",
                        "trigger_label": "8 trimestres con alpha positiva",
                    },
                    "coefficient_alert": {
                        "label": "Solido",
                        "tone": "good",
                        "trigger_label": "Coeficiente consistente",
                    },
                    "reference_playbook": {"best_candidate": None},
                    "cycle_projection_5y": {
                        "available": True,
                        "cycle_phase": "Expansion",
                        "five_year_return_pct": Decimal("61.0510"),
                        "path": [
                            {"label": "6M", "projected_price": Decimal("104.8809")},
                            {"label": "1A", "projected_price": Decimal("110.0000")},
                            {"label": "18M", "projected_price": Decimal("115.3687")},
                            {"label": "2A", "projected_price": Decimal("121.0000")},
                            {"label": "30M", "projected_price": Decimal("126.9016")},
                            {"label": "3A", "projected_price": Decimal("133.1000")},
                            {"label": "42M", "projected_price": Decimal("139.5968")},
                            {"label": "4A", "projected_price": Decimal("146.4100")},
                            {"label": "54M", "projected_price": Decimal("153.5574")},
                            {"label": "5A", "projected_price": Decimal("161.0510")},
                        ],
                    },
                }
            ]
        )[0]

        self.assertLess(row["expected_return_1y_pct"], ZERO)
        self.assertEqual(row["trade_alert_label"], "Vigilar")
        self.assertEqual(row["trade_alert_tone"], "watch")
        self.assertEqual(row["trade_alert_trigger"], "La esperanza 1A sigue negativa")

    def test_history_cards_and_decision_rows_include_per_valuation(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="SAN",
            quote_symbol="SAN.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Banco Santander",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("4.5000"),
            current_price_per_share=Decimal("4.5000"),
        )
        populate_position_history(position, growth=Decimal("1.0120"), benchmark_growth=Decimal("1.0060"), months=60)
        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        workbook_snapshot = {
            "available": True,
            "companies": [
                {
                    "ticker": "SAN",
                    "company_name": "Banco Santander",
                    "sector": "Banca",
                    "per_2025": Decimal("8.00"),
                }
            ],
            "companies_by_key": {
                "SAN": {
                    "ticker": "SAN",
                    "company_name": "Banco Santander",
                    "sector": "Banca",
                    "per_2025": Decimal("8.00"),
                },
                "BANCO SANTANDER": {
                    "ticker": "SAN",
                    "company_name": "Banco Santander",
                    "sector": "Banca",
                    "per_2025": Decimal("8.00"),
                },
                "SAN MC": {
                    "ticker": "SAN",
                    "company_name": "Banco Santander",
                    "sector": "Banca",
                    "per_2025": Decimal("8.00"),
                },
            },
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=workbook_snapshot),
            patch("equities.services.should_fetch_equity_fundamentals", return_value=False),
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
        ):
            card = build_equity_history_cards([position])[0]
            row = build_equity_decision_rows([card])[0]

        self.assertEqual(card["valuation"]["per_value"], Decimal("8.00"))
        self.assertEqual(card["valuation"]["label"], "Ajustada")
        self.assertEqual(card["projection"]["per_value"], Decimal("8.00"))
        self.assertEqual(row["per_value"], Decimal("8.00"))
        self.assertEqual(row["valuation_label"], "Ajustada")

    def test_capture_purchase_forecast_baseline_uses_latest_nightly_snapshot_before_purchase(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola, S.A.",
            opened_on=date(2026, 4, 17),
            shares=Decimal("20"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.4000"),
        )
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 17),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        cached_position = EquityPosition(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola, S.A.",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        card = {
            "position": cached_position,
            "reference_label": "IBEX 35",
            "projection": {
                "available": True,
                "projected_price": Decimal("11.2500"),
                "base_return_pct": Decimal("12.50"),
                "safety_score": Decimal("68.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
            "trade_alert": {"label": "Comprar"},
            "cycle_projection_5y": {
                "available": True,
                "path": [
                    {"label": "1A", "projected_price": Decimal("11.0000")},
                    {"label": "2A", "projected_price": Decimal("12.3750")},
                    {"label": "3A", "projected_price": Decimal("13.6125")},
                    {"label": "4A", "projected_price": Decimal("14.9738")},
                    {"label": "5A", "projected_price": Decimal("16.4712")},
                ],
            },
        }
        EquityNightlyAnalysisSnapshot.objects.create(
            run=run,
            analysis_date=run.analysis_date,
            scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
            analysis_key="ibex:IBE",
            ticker="IBE",
            quote_symbol="IBE.MC",
            company_name="Iberdrola, S.A.",
            status_key="ibex",
            sector_label="Utilities",
            agent_provider="core",
            analysis_payload=serialize_cached_value(card),
        )

        baseline = capture_purchase_forecast_baseline(position)

        self.assertIsNotNone(baseline)
        self.assertEqual(EquityPurchaseForecastBaseline.objects.count(), 1)
        self.assertEqual(baseline.source_analysis_date, date(2026, 4, 17))
        self.assertEqual(baseline.baseline_date, date(2026, 4, 17))
        self.assertEqual(baseline.trade_alert_label, "Comprar")
        self.assertEqual(baseline.projected_price_1y, Decimal("11.2500"))
        self.assertEqual(baseline.projected_return_pct_1y, Decimal("12.50"))
        self.assertEqual(baseline.projected_price_2y, Decimal("12.3750"))
        self.assertEqual(baseline.projected_return_pct_2y, Decimal("23.75"))
        self.assertEqual(len(baseline.projected_path_5y), 5)
        self.assertEqual(baseline.projected_path_5y[0]["label"], "1A")
        self.assertEqual(baseline.projected_path_5y[-1]["projected_price"], "16.4712")

    def test_capture_purchase_forecast_baseline_falls_back_to_first_available_run_after_purchase(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola, S.A.",
            opened_on=date(2026, 4, 15),
            shares=Decimal("20"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.4000"),
        )
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 17),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        cached_position = EquityPosition(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola, S.A.",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        card = {
            "position": cached_position,
            "reference_label": "IBEX 35",
            "projection": {
                "available": True,
                "projected_price": Decimal("11.2500"),
                "base_return_pct": Decimal("12.50"),
                "safety_score": Decimal("68.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
            "trade_alert": {"label": "Comprar"},
            "cycle_projection_5y": {
                "available": True,
                "path": [
                    {"label": "1A", "projected_price": Decimal("11.0000")},
                    {"label": "2A", "projected_price": Decimal("12.3750")},
                    {"label": "3A", "projected_price": Decimal("13.6125")},
                    {"label": "4A", "projected_price": Decimal("14.9738")},
                    {"label": "5A", "projected_price": Decimal("16.4712")},
                ],
            },
        }
        EquityNightlyAnalysisSnapshot.objects.create(
            run=run,
            analysis_date=run.analysis_date,
            scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
            analysis_key="ibex:IBE",
            ticker="IBE",
            quote_symbol="IBE.MC",
            company_name="Iberdrola, S.A.",
            status_key="ibex",
            sector_label="Utilities",
            agent_provider="core",
            analysis_payload=serialize_cached_value(card),
        )

        baseline = capture_purchase_forecast_baseline(position)

        self.assertIsNotNone(baseline)
        self.assertEqual(baseline.source_analysis_date, date(2026, 4, 17))
        self.assertEqual(baseline.baseline_date, date(2026, 4, 17))

    def test_capture_purchase_forecast_baseline_preserves_existing_snapshot_unless_overwritten(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola, S.A.",
            opened_on=date(2026, 4, 15),
            shares=Decimal("20"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.4000"),
        )
        first_run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 17),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        second_run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 18),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        for run, price_1y, price_5y in (
            (first_run, Decimal("11.2500"), Decimal("16.4712")),
            (second_run, Decimal("13.5000"), Decimal("25.9000")),
        ):
            cached_position = EquityPosition(
                broker="Interactive Brokers",
                ticker="IBE",
                quote_symbol="IBE.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Iberdrola, S.A.",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("10.0000"),
                current_price_per_share=Decimal("10.0000"),
            )
            card = {
                "position": cached_position,
                "reference_label": "IBEX 35",
                "projection": {
                    "available": True,
                    "projected_price": price_1y,
                    "base_return_pct": Decimal("12.50"),
                    "safety_score": Decimal("68.00"),
                },
                "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
                "trade_alert": {"label": "Comprar"},
                "cycle_projection_5y": {
                    "available": True,
                    "path": [
                        {"label": "1A", "projected_price": price_1y},
                        {"label": "5A", "projected_price": price_5y},
                    ],
                },
            }
            EquityNightlyAnalysisSnapshot.objects.create(
                run=run,
                analysis_date=run.analysis_date,
                scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
                analysis_key=f"ibex:IBE:{run.analysis_date.isoformat()}",
                ticker="IBE",
                quote_symbol="IBE.MC",
                company_name="Iberdrola, S.A.",
                status_key="ibex",
                sector_label="Utilities",
                agent_provider="core",
                analysis_payload=serialize_cached_value(card),
            )

        baseline = capture_purchase_forecast_baseline(position)
        self.assertEqual(baseline.source_analysis_date, date(2026, 4, 17))
        self.assertEqual(baseline.projected_price_5y, Decimal("16.4712"))

        preserved = capture_purchase_forecast_baseline(position)
        self.assertEqual(preserved.id, baseline.id)
        self.assertEqual(preserved.source_analysis_date, date(2026, 4, 17))
        self.assertEqual(preserved.projected_price_5y, Decimal("16.4712"))

        overwritten = capture_purchase_forecast_baseline(
            position,
            baseline_date=date(2026, 4, 18),
            overwrite=True,
        )
        self.assertEqual(overwritten.id, baseline.id)
        self.assertEqual(overwritten.source_analysis_date, date(2026, 4, 18))
        self.assertEqual(overwritten.projected_price_5y, Decimal("25.9000"))

    def test_build_ibex_recommendation_date_map_tracks_last_buy_and_sell_starts(self):
        for analysis_day, label in (
            (date(2026, 4, 14), "Vigilar"),
            (date(2026, 4, 15), "Comprar"),
            (date(2026, 4, 16), "Comprar"),
            (date(2026, 4, 17), "Vender"),
            (date(2026, 4, 18), "Vender"),
        ):
            run = EquityNightlyAnalysisRun.objects.create(
                analysis_date=analysis_day,
                status=EquityNightlyAnalysisRun.Status.COMPLETED,
                started_at=timezone.now(),
                completed_at=timezone.now(),
            )
            EquityNightlyAnalysisSnapshot.objects.create(
                run=run,
                analysis_date=analysis_day,
                scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
                analysis_key=f"ibex:ACS:{analysis_day.isoformat()}",
                ticker="ACS",
                quote_symbol="ACS.MC",
                company_name="ACS",
                status_key="ibex",
                sector_label="Construccion",
                agent_provider="core",
                analysis_payload=serialize_cached_value({"trade_alert": {"label": label}}),
            )

        recommendation_dates = build_ibex_recommendation_date_map(["ACS"])

        self.assertEqual(recommendation_dates["ACS"]["buy_recommended_on"], date(2026, 4, 15))
        self.assertEqual(recommendation_dates["ACS"]["sell_recommended_on"], date(2026, 4, 17))

    def test_projection_includes_dividends_and_broker_drag(self):
        position = EquityPosition.objects.create(
            broker="Banco Santander",
            trade_channel=EquityPosition.TradeChannel.APP,
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola S.A.",
            shares=Decimal("80"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("13.0000"),
            annual_dividend_income=Decimal("55.00"),
        )
        for price_date, close_price, benchmark_close in (
            (date(2024, 12, 31), Decimal("9.60"), Decimal("100.00")),
            (date(2025, 3, 31), Decimal("10.10"), Decimal("102.00")),
            (date(2025, 6, 30), Decimal("10.70"), Decimal("104.00")),
            (date(2025, 9, 30), Decimal("11.20"), Decimal("105.50")),
            (date(2025, 12, 31), Decimal("11.70"), Decimal("107.20")),
            (date(2026, 3, 31), Decimal("12.20"), Decimal("109.80")),
            (date(2026, 6, 30), Decimal("13.00"), Decimal("112.40")),
        ):
            position.price_history.create(
                price_date=price_date,
                close_price=close_price,
                benchmark_close=benchmark_close,
            )

        cards = build_equity_history_cards([position])

        projection = cards[0]["projection"]
        self.assertTrue(projection["available"])
        self.assertGreater(projection["gross_dividend_yield_pct"], Decimal("0"))
        self.assertGreater(projection["net_income_yield_pct"], Decimal("0"))
        self.assertGreater(projection["transaction_drag_pct"], Decimal("0"))
        self.assertEqual(
            projection["base_return_pct"].quantize(Decimal("0.01")),
            (
                projection["price_return_pct"]
                + projection["net_income_yield_pct"]
                - projection["transaction_drag_pct"]
            ).quantize(Decimal("0.01")),
        )

    def test_watchlist_projection_normalizes_costs_to_analysis_ticket(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Banco Sabadell",
            trade_channel=EquityPosition.TradeChannel.APP,
            ticker="SCYR",
            quote_symbol="SCYR.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Sacyr, S.A.",
            shares=Decimal("1"),
            average_cost_per_share=Decimal("4.6500"),
            current_price_per_share=Decimal("4.6500"),
            annual_dividend_income=Decimal("0"),
        )
        for price_date, close_price, benchmark_close in (
            (date(2025, 12, 31), Decimal("3.10"), Decimal("100.00")),
            (date(2026, 1, 31), Decimal("3.25"), Decimal("102.00")),
            (date(2026, 2, 28), Decimal("3.45"), Decimal("104.00")),
            (date(2026, 3, 31), Decimal("3.70"), Decimal("106.00")),
            (date(2026, 4, 30), Decimal("4.10"), Decimal("108.00")),
            (date(2026, 5, 31), Decimal("4.35"), Decimal("110.00")),
            (date(2026, 6, 30), Decimal("4.65"), Decimal("112.00")),
        ):
            position.price_history.create(
                price_date=price_date,
                close_price=close_price,
                benchmark_close=benchmark_close,
            )

        cards = build_equity_history_cards([position])

        projection = cards[0]["projection"]
        self.assertTrue(projection["available"])
        self.assertEqual(projection["analysis_value_source"], "normalized_watchlist")
        self.assertEqual(projection["analysis_value_amount"], Decimal("10000.00"))
        self.assertGreater(projection["base_return_pct"], Decimal("0"))
        self.assertEqual(cards[0]["broker_costs"]["roundtrip_total_cost"], Decimal("32.00"))
        self.assertEqual(cards[0]["broker_costs"]["annual_cost_used"], Decimal("25.00"))

    def test_history_cards_include_projection_backtest_accuracy(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola S.A.",
            shares=Decimal("15"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )

        stock_price = Decimal("10.00")
        benchmark_price = Decimal("100.00")
        for index in range(24):
            year = 2024 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            stock_price = (stock_price * Decimal("1.02")).quantize(Decimal("0.0001"))
            benchmark_price = (benchmark_price * Decimal("1.01")).quantize(Decimal("0.0001"))
            position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=stock_price,
                benchmark_close=benchmark_price,
            )

        cards = build_equity_history_cards([position])

        backtest = cards[0]["projection_backtest"]
        self.assertTrue(backtest["available"])
        self.assertGreaterEqual(backtest["comparisons_count"], 5)
        self.assertLess(backtest["mean_absolute_error_pct"], Decimal("12.00"))
        self.assertGreater(backtest["direction_hit_rate_pct"], Decimal("80.00"))
        self.assertTrue(backtest["rows"])
        self.assertTrue(backtest["monthly_chart"]["available"])
        self.assertTrue(backtest["monthly_chart"]["forecast_line"])
        self.assertTrue(backtest["monthly_chart"]["actual_line"])

    def test_history_cards_include_fundamentals_summary_when_enabled(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="SAN",
            quote_symbol="SAN.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Banco Santander, S.A.",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("4.0000"),
            current_price_per_share=Decimal("5.0000"),
        )
        for price_date, close_price, benchmark_close in (
            (date(2025, 12, 31), Decimal("4.00"), Decimal("100.00")),
            (date(2026, 1, 31), Decimal("4.20"), Decimal("101.20")),
            (date(2026, 2, 28), Decimal("4.35"), Decimal("102.40")),
            (date(2026, 3, 31), Decimal("4.55"), Decimal("104.10")),
            (date(2026, 4, 30), Decimal("4.80"), Decimal("105.30")),
            (date(2026, 5, 31), Decimal("5.00"), Decimal("107.00")),
        ):
            position.price_history.create(
                price_date=price_date,
                close_price=close_price,
                benchmark_close=benchmark_close,
            )

        fundamentals_snapshot = {
            "symbol": "SAN.MC",
            "available": True,
            "currency_code": "EUR",
            "net_income_rows": [
                {
                    "as_of_date": date(2023, 12, 31),
                    "period_type": "12M",
                    "currency_code": "EUR",
                    "value": Decimal("11076000000"),
                },
                {
                    "as_of_date": date(2024, 12, 31),
                    "period_type": "12M",
                    "currency_code": "EUR",
                    "value": Decimal("12574000000"),
                },
                {
                    "as_of_date": date(2025, 12, 31),
                    "period_type": "12M",
                    "currency_code": "EUR",
                    "value": Decimal("14101000000"),
                },
            ],
            "market_cap_rows": [
                {
                    "as_of_date": date(2026, 4, 10),
                    "period_type": "TTM",
                    "currency_code": "EUR",
                    "value": Decimal("151560161340"),
                }
            ],
            "market_cap": Decimal("151560161340"),
            "market_cap_as_of_date": date(2026, 4, 10),
        }

        with patch("equities.services.fetch_equity_fundamentals", return_value=fundamentals_snapshot):
            cards = build_equity_history_cards([position], include_fundamentals=True)

        fundamentals = cards[0]["fundamentals"]
        self.assertTrue(fundamentals["available"])
        self.assertEqual(fundamentals["trend_label"], "Mejora")
        self.assertEqual(fundamentals["market_cap_label"], "151.6 mM EUR")
        self.assertEqual(fundamentals["net_income_rows"][0]["year_label"], "2025")
        self.assertEqual(fundamentals["net_income_rows"][-1]["year_label"], "2023")

    def test_history_cards_include_trade_alert_from_prolonged_relative_trend(self):
        buy_position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IDR",
            quote_symbol="IDR.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Indra Sistemas, S.A.",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("15.00"),
        )
        sell_position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="TEF",
            quote_symbol="TEF.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Telefonica, S.A.",
            shares=Decimal("25"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("20.00"),
        )

        benchmark_price = Decimal("100.0000")
        buy_price = Decimal("10.0000")
        sell_price = Decimal("10.0000")
        for index in range(30):
            year = 2023 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            benchmark_price = (benchmark_price * Decimal("1.0080")).quantize(Decimal("0.0001"))
            buy_price = (buy_price * Decimal("1.0240")).quantize(Decimal("0.0001"))
            sell_price = (sell_price * Decimal("0.9920")).quantize(Decimal("0.0001"))
            buy_position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=buy_price,
                benchmark_close=benchmark_price,
            )
            sell_position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=sell_price,
                benchmark_close=benchmark_price,
            )

        buy_position.current_price_per_share = buy_price
        buy_position.save(update_fields=["current_price_per_share"])
        sell_position.current_price_per_share = sell_price
        sell_position.save(update_fields=["current_price_per_share"])

        cards = build_equity_history_cards([buy_position, sell_position])
        cards_by_ticker = {card["position"].ticker: card for card in cards}

        self.assertEqual(cards_by_ticker["IDR"]["trade_alert"]["label"], "Comprar")
        self.assertEqual(cards_by_ticker["TEF"]["trade_alert"]["label"], "Vender")
        self.assertGreater(cards_by_ticker["IDR"]["trade_alert"]["score"], Decimal("0"))
        self.assertLess(cards_by_ticker["TEF"]["trade_alert"]["score"], Decimal("0"))

    def test_history_cards_flag_sell_when_reference_coefficient_breaks_down_for_months(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="REP",
            quote_symbol="REP.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Repsol, S.A.",
            shares=Decimal("30"),
            average_cost_per_share=Decimal("12.0000"),
            current_price_per_share=Decimal("12.0000"),
        )

        benchmark_price = Decimal("100.0000")
        stock_price = Decimal("12.0000")
        benchmark_return_pattern = [
            Decimal("0.0060"),
            Decimal("0.0120"),
            Decimal("0.0080"),
            Decimal("0.0150"),
            Decimal("0.0070"),
            Decimal("0.0110"),
        ]
        for index in range(36):
            year = 2023 + (index // 12)
            month = (index % 12) + 1
            month_end = monthrange(year, month)[1]
            benchmark_return = benchmark_return_pattern[index % len(benchmark_return_pattern)]
            if index < 24:
                stock_return = (benchmark_return * Decimal("1.30")) + Decimal("0.0010")
            else:
                stock_return = -(benchmark_return * Decimal("1.10"))
            benchmark_price = (benchmark_price * (Decimal("1.00") + benchmark_return)).quantize(Decimal("0.0001"))
            stock_price = (stock_price * (Decimal("1.00") + stock_return)).quantize(Decimal("0.0001"))
            position.price_history.create(
                price_date=date(year, month, month_end),
                close_price=stock_price,
                benchmark_close=benchmark_price,
            )

        position.current_price_per_share = stock_price
        position.save(update_fields=["current_price_per_share"])

        cards = build_equity_history_cards([position])

        coefficient_alert = cards[0]["coefficient_alert"]
        self.assertTrue(coefficient_alert["available"])
        self.assertEqual(coefficient_alert["label"], "Vender")
        self.assertEqual(coefficient_alert["tone"], "sell")
        self.assertGreaterEqual(coefficient_alert["deterioration_streak"], 3)
        self.assertIn("Reciente", coefficient_alert["trigger_label"])

    def test_trade_alert_stays_in_watch_when_trend_and_12m_net_diverge(self):
        position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Banco Sabadell",
            ticker="SCYR",
            quote_symbol="SCYR.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Sacyr, S.A.",
            shares=Decimal("1"),
            average_cost_per_share=Decimal("4.6500"),
            current_price_per_share=Decimal("4.6500"),
        )

        alert = build_trade_alert(
            position,
            projection={
                "available": True,
                "safety_score": Decimal("60.00"),
                "base_return_pct": Decimal("-8.00"),
            },
            correlation={"coefficient": Decimal("0.76")},
            reliability={"score": Decimal("66.00")},
            relative_trend={
                "label": "Mejora moderada",
                "periods_label": "meses",
                "positive_streak": 3,
                "negative_streak": 0,
                "prolonged_positive": True,
                "prolonged_negative": False,
                "recent_gap_avg_pct": Decimal("5.20"),
                "gap_slope_pct": Decimal("0.45"),
            },
            six_month_snapshot={"available": True, "alpha_pct": Decimal("3.50")},
            one_year_snapshot={"available": True, "alpha_pct": Decimal("6.20")},
        )

        self.assertEqual(alert["label"], "Vigilar")
        self.assertEqual(alert["tone"], "watch")
        self.assertIn("neto 12M sigue en negativo", alert["note"])

    def test_trade_alert_can_use_technical_signal_as_extra_confirmation(self):
        position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Interactive Brokers",
            ticker="AMS",
            quote_symbol="AMS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Amadeus IT Group",
            shares=Decimal("1"),
            average_cost_per_share=Decimal("70.0000"),
            current_price_per_share=Decimal("70.0000"),
        )
        base_kwargs = {
            "projection": {
                "available": True,
                "safety_score": Decimal("60.00"),
                "base_return_pct": Decimal("8.00"),
            },
            "correlation": {"coefficient": Decimal("0.30")},
            "reliability": {"score": Decimal("66.00")},
            "relative_trend": {
                "label": "Mejora inicial",
                "periods_label": "meses",
                "positive_streak": 2,
                "negative_streak": 0,
                "prolonged_positive": False,
                "prolonged_negative": False,
                "recent_gap_avg_pct": Decimal("1.50"),
                "gap_slope_pct": Decimal("0.25"),
            },
            "six_month_snapshot": {"available": True, "alpha_pct": Decimal("2.00")},
            "one_year_snapshot": {"available": False},
            "valuation": {"score": Decimal("1.00")},
        }

        alert_without_technical = build_trade_alert(position, **base_kwargs)
        alert_with_technical = build_trade_alert(
            position,
            **base_kwargs,
            technical_signal={
                "available": True,
                "signal_label": "Compra tecnica",
                "signal_score": Decimal("4.50"),
                "confidence_label": "Alta",
            },
        )

        self.assertEqual(alert_without_technical["label"], "Vigilar")
        self.assertEqual(alert_with_technical["label"], "Comprar")
        self.assertGreater(alert_with_technical["score"], alert_without_technical["score"])
        self.assertEqual(alert_with_technical["technical_label"], "Compra tecnica")

    def test_watchlist_positions_do_not_count_into_portfolio_totals(self):
        EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            broker="Banco Sabadell",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola S.A.",
            shares=Decimal("10"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Banco Sabadell",
            ticker="FER",
            quote_symbol="FER.MC",
            reference_profile=EquityPosition.ReferenceProfile.SPAIN_HOUSE_PRICE,
            benchmark_symbol="EUROSTAT:prc_hpi_q:ES:TOTAL:I15_Q",
            benchmark_name="Precio vivienda Espana",
            company_name="Ferrovial",
            shares=Decimal("0.0000"),
            average_cost_per_share=Decimal("38.0000"),
            current_price_per_share=Decimal("39.5000"),
        )

        dashboard = build_equity_analysis_dashboard(list(EquityPosition.objects.prefetch_related("price_history")))

        self.assertEqual(dashboard["overview"]["owned_positions_count"], 1)
        self.assertEqual(dashboard["overview"]["watchlist_positions_count"], 1)
        self.assertEqual(dashboard["overview"]["invested_amount"], Decimal("100.0000"))
        self.assertEqual(dashboard["overview"]["current_value"], Decimal("120.0000"))

    def test_dashboard_overview_includes_first_owned_sale_recommendation(self):
        iberdrola = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        enagas = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="ENG",
            quote_symbol="ENG.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Enagas",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("14.0000"),
            current_price_per_share=Decimal("15.0000"),
        )
        populate_position_history(iberdrola, growth=Decimal("1.0150"), benchmark_growth=Decimal("1.0060"), months=36)
        populate_position_history(enagas, growth=Decimal("1.0140"), benchmark_growth=Decimal("1.0060"), months=36)
        plan_payloads = {
            "IBE": {
                "available": True,
                "mode": "sale_reentry",
                "sale_month_number": 12,
                "sale_date": date(2027, 4, 17),
                "sale_window_label": "abril 2027 (mes 12)",
                "signal_value_pct": Decimal("-0.13"),
                "summary": "Salida tactica en abril 2027.",
            },
            "ENG": {
                "available": True,
                "mode": "sale_review",
                "sale_month_number": 9,
                "sale_date": date(2027, 1, 17),
                "sale_window_label": "enero 2027 (mes 9)",
                "signal_value_pct": Decimal("-0.22"),
                "summary": "Salida tactica en enero 2027.",
            },
        }

        with patch(
            "equities.services.build_owned_cycle_trade_timing_plan",
            side_effect=lambda card: plan_payloads[card["position"].ticker],
        ):
            dashboard = build_equity_analysis_dashboard(
                list(EquityPosition.objects.prefetch_related("price_history"))
            )

        recommendation = dashboard["overview"]["next_sale_recommendation"]
        self.assertTrue(recommendation["available"])
        self.assertEqual(recommendation["ticker"], "ENG")
        self.assertEqual(recommendation["company_name"], "Enagas")
        self.assertEqual(recommendation["sale_window_label"], "enero 2027 (mes 9)")

    def test_dashboard_builds_comparable_return_summary_for_owned_positions(self):
        owned = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2025, 4, 12),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.5000"),
            annual_dividend_income=Decimal("12.00"),
            annual_maintenance_cost=Decimal("5.00"),
        )
        owned.price_history.create(price_date=date(2025, 4, 12), close_price=Decimal("10.0000"), benchmark_close=Decimal("100.0000"))
        owned.price_history.create(price_date=date(2026, 4, 12), close_price=Decimal("12.5000"), benchmark_close=Decimal("110.0000"))

        dashboard = build_equity_analysis_dashboard([owned])

        comparable_summary = dashboard["overview"]["comparable_summary"]
        self.assertTrue(comparable_summary["available"])
        self.assertEqual(comparable_summary["positions_count"], 1)
        self.assertEqual(comparable_summary["best_ticker"], "IBE")
        self.assertIsNotNone(comparable_summary["weighted_monthly_return_pct"])
        self.assertIsNotNone(comparable_summary["weighted_annual_return_pct"])

    def test_dashboard_overview_exposes_projection_horizons_for_3m_6m_9m_12m(self):
        owned = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
            annual_dividend_income=Decimal("12.00"),
            annual_maintenance_cost=Decimal("3.00"),
        )
        populate_position_history(
            owned,
            growth=Decimal("1.0140"),
            benchmark_growth=Decimal("1.0060"),
            months=36,
        )

        dashboard = build_equity_analysis_dashboard(
            list(EquityPosition.objects.prefetch_related("price_history"))
        )

        projection_horizons = dashboard["overview"]["projection_horizons"]
        self.assertEqual([item["label"] for item in projection_horizons], ["3M", "6M", "9M", "12M"])
        self.assertTrue(all(item["return_pct"] is not None for item in projection_horizons))
        self.assertEqual(
            dashboard["overview"]["weighted_projected_return_12m"],
            next(item["return_pct"] for item in projection_horizons if item["label"] == "12M"),
        )

    def test_build_current_dashboard_llm_summary_disables_llm_when_provider_is_unavailable(self):
        position = EquityPosition(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        dashboard = {
            "history_cards": [{"position": position, "ai_analysis": {}}],
            "ibex_universe_cards": [],
        }
        config = SimpleNamespace(
            provider="off",
            label="Analista cuantitativo",
            model="",
            monthly_budget_usd=ZERO,
            available=False,
            reason="Proveedor desactivado.",
        )

        summary = build_current_dashboard_llm_summary(
            dashboard,
            config=config,
            analysis_date=date(2026, 4, 26),
            estimated_cost_usd="0",
            latest_llm_run=None,
            refresh_performed=False,
            news_summary={},
            expert_summary={},
        )

        self.assertFalse(summary["enabled"])
        self.assertFalse(summary["reused"])
        self.assertEqual(summary["reason"], "Proveedor desactivado.")

    def test_build_nightly_completion_note_truncates_reused_ai_status_note(self):
        llm_summary = {
            "enabled": True,
            "reused": True,
            "refresh_performed": False,
            "label": "Claude Sonnet 4 muy largo para status note",
            "completed_count": 37,
            "total_count": 37,
            "source_analysis_date_label": "2026-04-22",
            "next_refresh_date_label": "2026-04-29",
            "news_enabled": True,
            "news_items_count": 14,
            "estimated_cost_usd": "12.3456",
        }

        note = build_nightly_completion_note(llm_summary)

        self.assertLessEqual(len(note), 255)
        self.assertTrue(note.endswith("..."))

    def test_portfolio_expectation_horizons_aggregate_company_expectations_by_current_value(self):
        owned_ibe = EquityPosition(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        owned_eng = EquityPosition(
            broker="Interactive Brokers",
            ticker="ENG",
            quote_symbol="ENG.MC",
            company_name="Enagas",
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("15.0000"),
            current_price_per_share=Decimal("15.0000"),
        )
        watchlist = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            company_name="ACS",
            shares=ZERO,
            average_cost_per_share=Decimal("45.0000"),
            current_price_per_share=Decimal("45.0000"),
        )
        history_cards = [
            {"position": owned_ibe},
            {"position": owned_eng},
            {"position": watchlist},
        ]
        decision_rows = [
            {
                "ticker": "IBE",
                "expected_return_1y_pct": Decimal("10.00"),
                "expected_return_2y_pct": Decimal("20.00"),
                "expected_return_3y_pct": Decimal("30.00"),
                "expected_return_4y_pct": Decimal("40.00"),
                "expected_return_5y_pct": Decimal("50.00"),
            },
            {
                "ticker": "ENG",
                "expected_return_1y_pct": Decimal("20.00"),
                "expected_return_2y_pct": Decimal("40.00"),
                "expected_return_3y_pct": Decimal("50.00"),
                "expected_return_4y_pct": Decimal("60.00"),
                "expected_return_5y_pct": Decimal("70.00"),
            },
            {
                "ticker": "ACS",
                "expected_return_1y_pct": Decimal("99.00"),
                "expected_return_2y_pct": Decimal("99.00"),
                "expected_return_3y_pct": Decimal("99.00"),
                "expected_return_4y_pct": Decimal("99.00"),
                "expected_return_5y_pct": Decimal("99.00"),
            },
        ]

        expectation_horizons = build_portfolio_expectation_horizons(history_cards, decision_rows)

        self.assertEqual([item["label"] for item in expectation_horizons], ["1A", "2A", "3A", "4A", "5A"])
        self.assertEqual(next(item["return_pct"] for item in expectation_horizons if item["label"] == "1A"), Decimal("17.50"))
        self.assertEqual(next(item["return_pct"] for item in expectation_horizons if item["label"] == "2A"), Decimal("35.00"))
        self.assertEqual(next(item["return_pct"] for item in expectation_horizons if item["label"] == "3A"), Decimal("45.00"))
        self.assertEqual(next(item["return_pct"] for item in expectation_horizons if item["label"] == "4A"), Decimal("55.00"))
        self.assertEqual(next(item["return_pct"] for item in expectation_horizons if item["label"] == "5A"), Decimal("65.00"))
        one_year = next(item for item in expectation_horizons if item["label"] == "1A")
        self.assertEqual(one_year["projected_total_value"], Decimal("470.00"))
        self.assertEqual(one_year["positions_count"], 2)

    @override_settings(EQUITIES_REFERENCE_WORKBOOK="")
    def test_dashboard_builds_reference_guide_from_workbook(self):
        workbook_path = build_test_reference_workbook()
        self.addCleanup(lambda: os.path.exists(workbook_path) and os.remove(workbook_path))
        load_ibex_reference_workbook_snapshot.cache_clear()

        with override_settings(EQUITIES_REFERENCE_WORKBOOK=workbook_path):
            position = EquityPosition.objects.create(
                broker="Interactive Brokers",
                ticker="SAN",
                quote_symbol="SAN.MC",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Banco Santander",
                shares=Decimal("10"),
                average_cost_per_share=Decimal("4.0000"),
                current_price_per_share=Decimal("5.0000"),
            )

            dashboard = build_equity_analysis_dashboard([position])

        self.assertTrue(dashboard["reference_guide_summary"]["workbook_loaded"])
        self.assertEqual(len(dashboard["tracked_reference_rows"]), 1)
        tracked_row = dashboard["tracked_reference_rows"][0]
        self.assertEqual(tracked_row["company_name"], "Banco Santander")
        self.assertEqual(tracked_row["best_candidate"]["name"], "Euribor 12m (%)")
        self.assertTrue(tracked_row["best_candidate"]["supports_chart"])
        self.assertEqual(dashboard["history_cards"][0]["reference_playbook"]["best_candidate"]["name"], "Euribor 12m (%)")

    def test_ticket_tracking_captures_daily_snapshots_and_builds_global_chart(self):
        positions = []
        for ticker, company_name, cost, current in (
            ("IBE", "Iberdrola", Decimal("10.0000"), Decimal("12.0000")),
            ("ENG", "Enagas", Decimal("14.0000"), Decimal("15.5000")),
        ):
            position = EquityPosition.objects.create(
                broker="Interactive Brokers",
                ticker=ticker,
                quote_symbol=f"{ticker}.MC",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name=company_name,
                shares=Decimal("20.0000"),
                average_cost_per_share=cost,
                current_price_per_share=current,
                annual_dividend_income=Decimal("40.00"),
            )
            stock_series = build_compound_market_series(
                f"{ticker}.MC",
                company_name,
                growth=Decimal("1.0180"),
                start_price=cost,
            )
            reference_series = build_compound_market_series(
                "^IBEX",
                "IBEX 35",
                growth=Decimal("1.0070"),
                start_price=Decimal("100.0000"),
            )
            for stock_point, reference_point in zip(stock_series.points, reference_series.points):
                position.price_history.create(
                    price_date=stock_point["date"],
                    open_price=stock_point["open"],
                    high_price=stock_point["high"],
                    low_price=stock_point["low"],
                    close_price=stock_point["close"],
                    benchmark_close=reference_point["close"],
                )
            positions.append(position)

        first_cards = build_equity_history_cards(positions)
        capture_equity_ticket_snapshots(first_cards, snapshot_date=date(2026, 4, 12))

        positions[0].current_price_per_share = Decimal("12.4000")
        positions[0].save(update_fields=["current_price_per_share", "updated_at"])
        positions[1].current_price_per_share = Decimal("15.9000")
        positions[1].save(update_fields=["current_price_per_share", "updated_at"])

        second_cards = build_equity_history_cards(list(EquityPosition.objects.prefetch_related("price_history")))
        capture_equity_ticket_snapshots(second_cards, snapshot_date=date(2026, 4, 13))
        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            tracking = build_equity_ticket_tracking_context(second_cards)

        self.assertEqual(EquityTicketSnapshot.objects.count(), 4)
        self.assertTrue(tracking["available"])
        self.assertEqual(tracking["tracked_ticket_count"], 2)
        self.assertTrue(tracking["global"]["available"])
        self.assertTrue(tracking["global"]["chart"]["available"])
        self.assertTrue(tracking["global"]["chart_5y"]["available"])
        self.assertTrue(tracking["global"]["net_chart_12m"]["available"])
        self.assertTrue(tracking["global"]["net_chart_5y"]["available"])
        self.assertTrue(tracking["global"]["return_chart_12m"]["available"])
        self.assertTrue(tracking["global"]["return_chart_5y"]["available"])
        self.assertTrue(tracking["global"]["cumulative_alpha_chart"]["available"])
        self.assertFalse(tracking["global"]["weekly_alpha_chart"]["available"])
        self.assertTrue(tracking["global"]["benchmark"]["available"])
        self.assertTrue(tracking["global"]["chart"]["benchmark_line"])
        self.assertEqual(tracking["global"]["chart"]["expected_line"], "")
        self.assertTrue(tracking["global"]["return_chart_12m"]["benchmark_line"])
        self.assertEqual(tracking["snapshot_days_count"], 2)
        self.assertEqual(len(tracking["tickets"]), 2)
        self.assertTrue(all(item["chart"]["available"] for item in tracking["tickets"]))
        self.assertTrue(
            all(len(item["expected_series_dense"]) > len(item["expected_series"]) for item in tracking["tickets"])
        )
        self.assertTrue(
            all(len(item["expected_series_5y_dense"]) > len(item["expected_series_5y"]) for item in tracking["tickets"])
        )
        self.assertIsNotNone(tracking["global"]["expected_today_value"])
        self.assertIsNotNone(tracking["global"]["expected_total_value_5y"])
        self.assertIsNotNone(tracking["global"]["expected_net_value_12m"])
        self.assertIsNotNone(tracking["global"]["expected_return_pct_5y"])
        self.assertTrue(tracking["global"]["chart"]["x_markers"])
        self.assertTrue(tracking["global"]["chart_5y"]["x_markers"])
        self.assertEqual(tracking["global"]["net_gain_value"], Decimal("16.00"))
        self.assertEqual(tracking["global"]["invested_return_pct"], Decimal("2.91"))
        self.assertEqual(tracking["global"]["annualized_return_pct"], Decimal("3512395.03"))
        self.assertEqual(tracking["global"]["daily_change_pct"], Decimal("2.91"))

    def test_ticket_tracking_recalibrates_expected_curve_when_reality_lags(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("100.0000"),
            current_price_per_share=Decimal("92.0000"),
            annual_dividend_income=Decimal("0.00"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 1),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("1000.00"),
            projected_market_value_12m=Decimal("1360.00"),
            projected_total_value_12m=Decimal("1360.00"),
            projected_price_12m=Decimal("136.0000"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 8),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("920.00"),
            projected_market_value_12m=Decimal("1360.00"),
            projected_total_value_12m=Decimal("1360.00"),
            projected_price_12m=Decimal("136.0000"),
        )
        snapshots = list(position.ticket_snapshots.order_by("snapshot_date"))
        raw_expected_series, raw_current_expected_value, _ = build_ticket_expected_series(
            snapshots,
            Decimal("1360.00"),
        )

        with patch("equities.services.build_owned_cycle_trade_timing_plan", return_value={"available": False}), patch(
            "equities.services.build_purchase_trade_rotation_guidance",
            return_value={"available": False},
        ):
            ticket = build_equity_ticket_tracking_item(
                {
                    "position": position,
                    "reference_label": "IBEX 35",
                    "projection": {"projected_price": Decimal("136.0000"), "base_return_pct": Decimal("36.00")},
                    "projection_reliability": {"label": "Baja", "score": Decimal("42.00")},
                    "cycle_projection_5y": {"available": False, "path": []},
                },
                snapshots,
            )

        self.assertIsNotNone(ticket)
        self.assertEqual(ticket["expected_series_calibration"]["label"], "Enfria")
        self.assertEqual(ticket["expected_series_calibration"]["tracked_days"], 7)
        self.assertLess(ticket["current_expected_value"], raw_current_expected_value)
        self.assertGreater(ticket["current_expected_value"], Decimal("920.00"))
        self.assertLess(ticket["expected_market_value_12m"], Decimal("1360.00"))

    def test_ticket_expected_series_uses_quarterly_shape_instead_of_linear_path(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("100.0000"),
            current_price_per_share=Decimal("100.0000"),
            annual_dividend_income=Decimal("0.00"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 25),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("1000.00"),
            projected_market_value_12m=Decimal("1400.00"),
            projected_total_value_12m=Decimal("1400.00"),
            projected_price_12m=Decimal("140.0000"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 5, 25),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("1010.00"),
            projected_market_value_12m=Decimal("1400.00"),
            projected_total_value_12m=Decimal("1400.00"),
            projected_price_12m=Decimal("140.0000"),
        )
        snapshots = list(position.ticket_snapshots.order_by("snapshot_date"))

        linear_series, _, _ = build_ticket_expected_series(snapshots, Decimal("1400.00"))
        shaped_series, _, _ = build_ticket_expected_series(
            snapshots,
            Decimal("1400.00"),
            card={
                "position": position,
                "projection": {
                    "latest_price": Decimal("100.0000"),
                    "projected_price": Decimal("140.0000"),
                    "quarterly_path": [
                        {"label": "3M", "projected_date": date(2026, 7, 25), "projected_price": Decimal("105.5316")},
                        {"label": "6M", "projected_date": date(2026, 10, 24), "projected_price": Decimal("113.6244")},
                        {"label": "9M", "projected_date": date(2027, 1, 23), "projected_price": Decimal("124.4352")},
                        {"label": "12M", "projected_date": date(2027, 4, 25), "projected_price": Decimal("140.0000")},
                    ],
                },
            },
        )

        linear_3m = next(point["value"] for point in linear_series if point["date"] == date(2026, 7, 25))
        shaped_3m = next(point["value"] for point in shaped_series if point["date"] == date(2026, 7, 25))
        linear_9m = next(point["value"] for point in linear_series if point["date"] == date(2027, 1, 23))
        shaped_9m = next(point["value"] for point in shaped_series if point["date"] == date(2027, 1, 23))

        self.assertLess(shaped_3m, linear_3m)
        self.assertNotEqual(shaped_9m, linear_9m)

    def test_ticket_expected_series_prefers_monthly_projection_path_for_tracking_shape(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("100.0000"),
            current_price_per_share=Decimal("100.0000"),
            annual_dividend_income=Decimal("0.00"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 25),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("1000.00"),
            projected_market_value_12m=Decimal("1400.00"),
            projected_total_value_12m=Decimal("1400.00"),
            projected_price_12m=Decimal("140.0000"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 5, 25),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("1010.00"),
            projected_market_value_12m=Decimal("1400.00"),
            projected_total_value_12m=Decimal("1400.00"),
            projected_price_12m=Decimal("140.0000"),
        )
        snapshots = list(position.ticket_snapshots.order_by("snapshot_date"))

        quarterly_series, _, _ = build_ticket_expected_series(
            snapshots,
            Decimal("1400.00"),
            card={
                "position": position,
                "projection": {
                    "latest_price": Decimal("100.0000"),
                    "projected_price": Decimal("140.0000"),
                    "quarterly_path": [
                        {"label": "3M", "projected_date": date(2026, 7, 25), "projected_price": Decimal("109.0000")},
                        {"label": "6M", "projected_date": date(2026, 10, 25), "projected_price": Decimal("118.0000")},
                        {"label": "9M", "projected_date": date(2027, 1, 25), "projected_price": Decimal("128.0000")},
                        {"label": "1A", "projected_date": date(2027, 4, 25), "projected_price": Decimal("140.0000")},
                    ],
                },
            },
        )
        monthly_series, _, _ = build_ticket_expected_series(
            snapshots,
            Decimal("1400.00"),
            card={
                "position": position,
                "projection": {
                    "latest_price": Decimal("100.0000"),
                    "projected_price": Decimal("140.0000"),
                    "monthly_path": [
                        {"label": "1M", "projected_date": date(2026, 5, 25), "projected_price": Decimal("102.0000")},
                        {"label": "2M", "projected_date": date(2026, 6, 25), "projected_price": Decimal("103.5000")},
                        {"label": "3M", "projected_date": date(2026, 7, 25), "projected_price": Decimal("105.5000")},
                        {"label": "4M", "projected_date": date(2026, 8, 25), "projected_price": Decimal("108.5000")},
                        {"label": "5M", "projected_date": date(2026, 9, 25), "projected_price": Decimal("112.0000")},
                        {"label": "6M", "projected_date": date(2026, 10, 25), "projected_price": Decimal("115.5000")},
                        {"label": "7M", "projected_date": date(2026, 11, 25), "projected_price": Decimal("118.5000")},
                        {"label": "8M", "projected_date": date(2026, 12, 25), "projected_price": Decimal("122.5000")},
                        {"label": "9M", "projected_date": date(2027, 1, 25), "projected_price": Decimal("127.0000")},
                        {"label": "10M", "projected_date": date(2027, 2, 25), "projected_price": Decimal("131.0000")},
                        {"label": "11M", "projected_date": date(2027, 3, 25), "projected_price": Decimal("135.0000")},
                        {"label": "1A", "projected_date": date(2027, 4, 25), "projected_price": Decimal("140.0000")},
                    ],
                    "quarterly_path": [
                        {"label": "3M", "projected_date": date(2026, 7, 25), "projected_price": Decimal("109.0000")},
                        {"label": "6M", "projected_date": date(2026, 10, 25), "projected_price": Decimal("118.0000")},
                        {"label": "9M", "projected_date": date(2027, 1, 25), "projected_price": Decimal("128.0000")},
                        {"label": "1A", "projected_date": date(2027, 4, 25), "projected_price": Decimal("140.0000")},
                    ],
                },
            },
        )

        monthly_first_anchor = next(point for point in monthly_series if point["date"] == date(2026, 5, 25) and point.get("is_anchor"))
        quarterly_first_future = next(point for point in quarterly_series if point["date"] == date(2026, 7, 25) and point.get("is_anchor"))

        self.assertEqual(monthly_first_anchor["label"], "1M")
        self.assertLess(monthly_first_anchor["value"], quarterly_first_future["value"])
        self.assertGreaterEqual(sum(1 for point in monthly_series if point.get("is_anchor")), 12)

    def test_ticket_tracking_recalibrates_expected_curve_when_reality_runs_ahead(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="ENG",
            quote_symbol="ENG.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Enagas",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("100.0000"),
            current_price_per_share=Decimal("111.0000"),
            annual_dividend_income=Decimal("0.00"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 1),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("1000.00"),
            projected_market_value_12m=Decimal("1120.00"),
            projected_total_value_12m=Decimal("1120.00"),
            projected_price_12m=Decimal("112.0000"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 8),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("1110.00"),
            projected_market_value_12m=Decimal("1120.00"),
            projected_total_value_12m=Decimal("1120.00"),
            projected_price_12m=Decimal("112.0000"),
        )
        snapshots = list(position.ticket_snapshots.order_by("snapshot_date"))
        raw_expected_series, raw_current_expected_value, _ = build_ticket_expected_series(
            snapshots,
            Decimal("1120.00"),
        )

        with patch("equities.services.build_owned_cycle_trade_timing_plan", return_value={"available": False}), patch(
            "equities.services.build_purchase_trade_rotation_guidance",
            return_value={"available": False},
        ):
            ticket = build_equity_ticket_tracking_item(
                {
                    "position": position,
                    "reference_label": "IBEX 35",
                    "projection": {"projected_price": Decimal("112.0000"), "base_return_pct": Decimal("12.00")},
                    "projection_reliability": {"label": "Baja", "score": Decimal("42.00")},
                    "cycle_projection_5y": {"available": False, "path": []},
                },
                snapshots,
            )

        self.assertIsNotNone(ticket)
        self.assertEqual(ticket["expected_series_calibration"]["label"], "Acelera")
        self.assertGreater(ticket["current_expected_value"], raw_current_expected_value)
        self.assertLess(ticket["current_expected_value"], Decimal("1110.00"))
        self.assertGreater(ticket["expected_market_value_12m"], Decimal("1120.00"))

    def test_ticket_tracking_moderates_overoptimistic_targets_when_backtest_is_weak(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="RED",
            quote_symbol="RED.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Redeia",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("100.0000"),
            current_price_per_share=Decimal("100.0000"),
            annual_dividend_income=Decimal("0.00"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 25),
            invested_amount=Decimal("1000.00"),
            current_value=Decimal("1000.00"),
            projected_market_value_12m=Decimal("1400.00"),
            projected_total_value_12m=Decimal("1450.00"),
            projected_price_12m=Decimal("140.0000"),
        )

        with patch("equities.services.build_owned_cycle_trade_timing_plan", return_value={"available": False}), patch(
            "equities.services.build_purchase_trade_rotation_guidance",
            return_value={"available": False},
        ):
            ticket = build_equity_ticket_tracking_item(
                {
                    "position": position,
                    "reference_label": "IBEX 35",
                    "projection": {
                        "projected_price": Decimal("140.0000"),
                        "base_return_pct": Decimal("45.00"),
                        "safety_score": Decimal("52.00"),
                        "cagr_pct": Decimal("6.50"),
                    },
                    "projection_reliability": {"label": "Baja", "score": Decimal("42.00")},
                    "projection_backtest": {
                        "available": False,
                        "precision_label": "Sin historico suficiente",
                        "monthly_chart": {"available": False},
                    },
                    "cycle_projection_5y": {
                        "available": True,
                        "path": [
                            {"label": "1A", "projected_price": Decimal("145.0000")},
                            {"label": "5A", "projected_price": Decimal("200.0000")},
                        ],
                    },
                },
                list(position.ticket_snapshots.order_by("snapshot_date")),
            )

        self.assertIsNotNone(ticket)
        self.assertTrue(ticket["expected_target_moderation_12m"]["available"])
        self.assertLess(ticket["expected_market_value_12m"], Decimal("1400.00"))
        self.assertLess(ticket["expected_total_value_12m"], Decimal("1450.00"))
        self.assertLess(ticket["expected_total_value_5y"], Decimal("2000.00"))

    def test_build_value_tracking_chart_reduces_overlapping_actual_markers(self):
        chart = build_value_tracking_chart(
            actual_series=[
                {"date": date(2026, 4, 20), "value": Decimal("-1.40")},
                {"date": date(2026, 4, 21), "value": Decimal("-0.60")},
                {"date": date(2026, 4, 22), "value": Decimal("0.20")},
                {"date": date(2026, 4, 23), "value": Decimal("0.90")},
                {"date": date(2026, 4, 24), "value": Decimal("1.40")},
            ],
            expected_series=[
                {"date": date(2026, 4, 20), "value": Decimal("-1.40")},
                {"date": date(2027, 4, 19), "value": Decimal("30.30")},
            ],
            value_suffix="%",
            axis_formatter=lambda value: f"{Decimal(str(value)).quantize(Decimal('0.1'))}",
        )

        self.assertTrue(chart["available"])
        self.assertEqual(len(chart["actual_points"]), 5)
        self.assertLess(len(chart["actual_display_points"]), len(chart["actual_points"]))
        self.assertEqual(chart["actual_display_points"][-1]["date_label"], "2026-04-24")
        self.assertTrue(chart["actual_display_points"][-1]["is_latest"])
        self.assertIn("2026-04-24", chart["actual_display_points"][-1]["tooltip"])

    def test_build_value_tracking_chart_expands_short_real_window_against_long_projection(self):
        chart = build_value_tracking_chart(
            actual_series=[
                {"date": date(2026, 4, 25), "value": Decimal("-4.43")},
                {"date": date(2026, 4, 28), "value": Decimal("-3.20")},
                {"date": date(2026, 4, 30), "value": Decimal("-2.40")},
            ],
            expected_series=densify_projected_tracking_series(
                [
                    {"date": date(2026, 4, 25), "value": Decimal("-4.43"), "label": "Hoy", "is_anchor": True},
                    {"date": date(2026, 7, 25), "value": Decimal("2.00"), "label": "3M", "is_anchor": True},
                    {"date": date(2026, 10, 25), "value": Decimal("6.10"), "label": "6M", "is_anchor": True},
                    {"date": date(2027, 1, 25), "value": Decimal("10.50"), "label": "9M", "is_anchor": True},
                    {"date": date(2027, 4, 25), "value": Decimal("15.10"), "label": "1A", "is_anchor": True},
                ]
            ),
            benchmark_series=[
                {"date": date(2026, 4, 25), "value": Decimal("-3.10")},
                {"date": date(2026, 4, 30), "value": Decimal("-3.60")},
                {"date": date(2027, 4, 25), "value": Decimal("1.50")},
            ],
            value_suffix="%",
            axis_formatter=format_percentage_axis_value,
        )

        self.assertTrue(chart["available"])
        self.assertTrue(chart["segmented_time_axis"])
        self.assertIsNotNone(chart["projection_zone_start_x"])
        self.assertTrue(chart["scale_note"])
        self.assertTrue(chart["latest_gap_line"])
        self.assertTrue(chart["expected_display_points"])
        self.assertIn("Hoy", [marker["label"] for marker in chart["x_markers"]])
        self.assertIn("1A", [marker["label"] for marker in chart["x_markers"]])
        self.assertGreater(float(chart["actual_display_points"][-1]["x"]), 180.0)

    def test_build_value_tracking_chart_supports_single_real_series(self):
        chart = build_value_tracking_chart(
            actual_series=[
                {"date": date(2026, 4, 14), "value": Decimal("-0.80")},
                {"date": date(2026, 4, 17), "value": Decimal("0.60")},
                {"date": date(2026, 4, 21), "value": Decimal("1.25")},
            ],
            expected_series=[],
            value_suffix="%",
            axis_formatter=lambda value: f"{Decimal(str(value)).quantize(Decimal('0.1'))}",
        )

        self.assertTrue(chart["available"])
        self.assertTrue(chart["actual_line"])
        self.assertEqual(chart["expected_line"], "")
        self.assertEqual(chart["projection_end_label"], "2026-04-21")
        self.assertIsNotNone(chart["zero_y"])

    def test_build_value_tracking_chart_can_render_month_markers_and_benchmark_points(self):
        chart = build_value_tracking_chart(
            actual_series=[
                {"date": date(2026, 4, 21), "value": Decimal("18300.0")},
                {"date": date(2026, 5, 19), "value": Decimal("18180.0")},
                {"date": date(2026, 6, 17), "value": Decimal("18440.0")},
                {"date": date(2026, 7, 14), "value": Decimal("18520.0")},
            ],
            expected_series=[],
            benchmark_series=[
                {"date": date(2026, 4, 21), "value": Decimal("18300.0")},
                {"date": date(2026, 5, 19), "value": Decimal("18240.0")},
                {"date": date(2026, 6, 17), "value": Decimal("18360.0")},
                {"date": date(2026, 7, 14), "value": Decimal("18410.0")},
            ],
            value_suffix="",
            axis_formatter=format_axis_value,
            time_marker_mode="month",
            grid_marker_mode="month",
        )

        self.assertTrue(chart["available"])
        self.assertTrue(chart["grid_markers"])
        self.assertTrue(chart["benchmark_display_points"])
        self.assertEqual(chart["x_markers"][0]["label"], "Abr 26")
        self.assertIn("May 26", [marker["label"] for marker in chart["x_markers"]])
        self.assertTrue(any(marker["draw_grid"] for marker in chart["grid_markers"]))

    def test_build_tracking_rebased_comparison_series_resets_scale_on_capital_change(self):
        comparison = build_tracking_rebased_comparison_series(
            actual_series=[
                {"date": date(2026, 4, 21), "value": Decimal("100.00")},
                {"date": date(2026, 4, 22), "value": Decimal("110.00")},
                {"date": date(2026, 4, 23), "value": Decimal("220.00")},
                {"date": date(2026, 4, 24), "value": Decimal("210.00")},
            ],
            invested_series=[
                {"date": date(2026, 4, 21), "value": Decimal("100.00")},
                {"date": date(2026, 4, 22), "value": Decimal("100.00")},
                {"date": date(2026, 4, 23), "value": Decimal("200.00")},
                {"date": date(2026, 4, 24), "value": Decimal("200.00")},
            ],
            benchmark_series=[
                {"date": date(2026, 4, 21), "value": Decimal("1000.00")},
                {"date": date(2026, 4, 22), "value": Decimal("1050.00")},
                {"date": date(2026, 4, 23), "value": Decimal("1100.00")},
                {"date": date(2026, 4, 24), "value": Decimal("1150.00")},
            ],
        )

        self.assertTrue(comparison["available"])
        self.assertEqual(
            [point["value"] for point in comparison["portfolio_series"]],
            [
                Decimal("1000.00"),
                Decimal("1100.00"),
                Decimal("1100.00"),
                Decimal("1050.00"),
            ],
        )
        self.assertEqual(comparison["capital_change_dates"], [date(2026, 4, 23)])
        self.assertEqual(comparison["latest_gap_value"], Decimal("-100.00"))

    def test_ticket_tracking_global_chart_keeps_first_real_portfolio_date_when_positions_enter_later(self):
        def create_position(ticker, company_name, start_price):
            position = EquityPosition.objects.create(
                broker="Interactive Brokers",
                ticker=ticker,
                quote_symbol=f"{ticker}.MC",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name=company_name,
                shares=Decimal("20.0000"),
                average_cost_per_share=start_price,
                current_price_per_share=(start_price * Decimal("1.1000")).quantize(Decimal("0.0001")),
                annual_dividend_income=Decimal("20.00"),
            )
            stock_series = build_compound_market_series(
                f"{ticker}.MC",
                company_name,
                growth=Decimal("1.0150"),
                start_price=start_price,
            )
            reference_series = build_compound_market_series(
                "^IBEX",
                "IBEX 35",
                growth=Decimal("1.0060"),
                start_price=Decimal("100.0000"),
            )
            for stock_point, reference_point in zip(stock_series.points, reference_series.points):
                position.price_history.create(
                    price_date=stock_point["date"],
                    open_price=stock_point["open"],
                    high_price=stock_point["high"],
                    low_price=stock_point["low"],
                    close_price=stock_point["close"],
                    benchmark_close=reference_point["close"],
                )
            return position

        first_positions = [
            create_position("IBE", "Iberdrola", Decimal("10.0000")),
            create_position("ENG", "Enagas", Decimal("14.0000")),
        ]
        first_cards = build_equity_history_cards(first_positions)
        capture_equity_ticket_snapshots(first_cards, snapshot_date=date(2026, 4, 12))

        create_position("SCYR", "Sacyr", Decimal("4.0000"))
        second_cards = build_equity_history_cards(list(EquityPosition.objects.prefetch_related("price_history")))
        capture_equity_ticket_snapshots(second_cards, snapshot_date=date(2026, 4, 13))

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            tracking = build_equity_ticket_tracking_context(second_cards)

        self.assertEqual(tracking["anchor_date"], date(2026, 4, 12))
        self.assertEqual(tracking["shared_anchor_date"], date(2026, 4, 13))
        self.assertEqual(tracking["snapshot_days_count"], 2)
        self.assertEqual(tracking["tracked_ticket_count"], 3)
        self.assertTrue(all(item["shared_baseline_snapshot"].snapshot_date == date(2026, 4, 13) for item in tracking["tickets"]))
        baselines_by_ticker = {item["position"].ticker: item["baseline_snapshot"].snapshot_date for item in tracking["tickets"]}
        self.assertEqual(baselines_by_ticker["IBE"], date(2026, 4, 12))
        self.assertEqual(baselines_by_ticker["ENG"], date(2026, 4, 12))
        self.assertEqual(baselines_by_ticker["SCYR"], date(2026, 4, 13))
        self.assertEqual(tracking["global"]["baseline_value"], Decimal("616.00"))
        self.assertEqual(tracking["global"]["net_gain_value"], Decimal("0.00"))
        self.assertEqual(tracking["global"]["daily_change_pct"], Decimal("0.00"))

    def test_ticket_tracking_5y_uses_purchase_baseline_path_instead_of_live_recalculation(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.3000"),
            annual_dividend_income=Decimal("20.00"),
        )
        stock_series = build_compound_market_series(
            "IBE.MC",
            "Iberdrola",
            growth=Decimal("1.0180"),
            start_price=Decimal("10.0000"),
        )
        reference_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        for stock_point, reference_point in zip(stock_series.points, reference_series.points):
            position.price_history.create(
                price_date=stock_point["date"],
                open_price=stock_point["open"],
                high_price=stock_point["high"],
                low_price=stock_point["low"],
                close_price=stock_point["close"],
                benchmark_close=reference_point["close"],
            )

        capture_equity_ticket_snapshots(
            [
                {
                    "position": position,
                    "projection": {"projected_price": Decimal("11.0000"), "base_return_pct": Decimal("10.00")},
                    "cycle_projection_5y": {"available": True, "path": [{"label": "5A", "projected_price": Decimal("60.0000")}]},
                }
            ],
            snapshot_date=date(2026, 4, 17),
        )
        EquityPurchaseForecastBaseline.objects.create(
            position=position,
            source_analysis_date=date(2026, 4, 17),
            baseline_date=date(2026, 4, 17),
            analysis_scope="ibex",
            analysis_key="ibex:IBE",
            baseline_price=Decimal("10.3000"),
            projected_price_1y=Decimal("11.5000"),
            projected_price_5y=Decimal("20.0000"),
            projected_return_pct_1y=Decimal("11.65"),
            projected_return_pct_5y=Decimal("94.17"),
            projected_path_5y=[
                {"label": "6M", "projected_price": "9.8000", "projected_date": "2026-10-17"},
                {"label": "1A", "projected_price": "11.5000", "projected_date": "2027-04-17"},
                {"label": "2A", "projected_price": "13.0000", "projected_date": "2028-04-17"},
                {"label": "3A", "projected_price": "15.5000", "projected_date": "2029-04-17"},
                {"label": "4A", "projected_price": "17.0000", "projected_date": "2030-04-17"},
                {"label": "5A", "projected_price": "20.0000", "projected_date": "2031-04-17"},
            ],
        )
        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            tracking = build_equity_ticket_tracking_context(
                [
                    {
                        "position": position,
                        "reference_label": "IBEX 35",
                        "projection": {"projected_price": Decimal("12.0000"), "base_return_pct": Decimal("16.50")},
                        "cycle_projection_5y": {
                            "available": True,
                            "path": [
                                {"label": "1A", "projected_price": Decimal("25.0000")},
                                {"label": "5A", "projected_price": Decimal("60.0000")},
                            ],
                        },
                    }
                ]
            )

        ticket = tracking["tickets"][0]
        self.assertEqual(ticket["expected_total_value_5y"], Decimal("400.00"))
        self.assertEqual(ticket["expected_series_5y"][1]["value"], Decimal("196.00"))
        self.assertEqual(ticket["expected_series_5y"][-1]["value"], Decimal("400.00"))

    def test_ticket_tracking_exposes_initial_and_current_unit_prices(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2026, 4, 16),
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.3000"),
            latest_price_date=date(2026, 4, 17),
            annual_dividend_income=Decimal("20.00"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 16),
            invested_amount=Decimal("200.00"),
            current_value=Decimal("200.00"),
            projected_market_value_12m=Decimal("224.00"),
            projected_total_value_12m=Decimal("224.00"),
            projected_price_12m=Decimal("11.2000"),
        )
        EquityTicketSnapshot.objects.create(
            position=position,
            snapshot_date=date(2026, 4, 17),
            invested_amount=Decimal("200.00"),
            current_value=Decimal("246.00"),
            projected_market_value_12m=Decimal("224.00"),
            projected_total_value_12m=Decimal("224.00"),
            projected_price_12m=Decimal("11.2000"),
        )
        EquityPurchaseForecastBaseline.objects.create(
            position=position,
            source_analysis_date=date(2026, 4, 16),
            baseline_date=date(2026, 4, 16),
            reference_label="IBEX 35",
            baseline_price=Decimal("10.0000"),
            projected_price_1y=Decimal("11.2000"),
            projected_return_pct_1y=Decimal("12.00"),
        )
        ticket = build_equity_ticket_tracking_item(
            {
                "position": position,
                "reference_label": "IBEX 35",
                "trade_alert": {"label": "Comprar"},
                "projection": {"projected_price": Decimal("11.2000"), "base_return_pct": Decimal("12.00")},
                "cycle_projection_5y": {"available": False, "path": []},
            },
            list(position.ticket_snapshots.order_by("snapshot_date")),
            purchase_baseline=position.purchase_forecast_baseline,
        )

        self.assertEqual(ticket["initial_unit_price"], Decimal("10.0000"))
        self.assertEqual(ticket["initial_unit_price_date"], date(2026, 4, 16))
        self.assertEqual(ticket["initial_unit_price_note"], "Compra o alta web")
        self.assertEqual(ticket["current_unit_price"], Decimal("12.3000"))
        self.assertEqual(ticket["current_unit_price_date"], date(2026, 4, 17))
        self.assertEqual(ticket["current_unit_price_note"], "Cotizacion mas reciente")

    def test_ticket_tracking_keeps_each_ticket_return_from_its_first_snapshot(self):
        def create_position(ticker, company_name, start_price):
            position = EquityPosition.objects.create(
                broker="Interactive Brokers",
                ticker=ticker,
                quote_symbol=f"{ticker}.MC",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name=company_name,
                shares=Decimal("20.0000"),
                average_cost_per_share=start_price,
                current_price_per_share=start_price,
                annual_dividend_income=Decimal("20.00"),
            )
            stock_series = build_compound_market_series(
                f"{ticker}.MC",
                company_name,
                growth=Decimal("1.0150"),
                start_price=start_price,
            )
            reference_series = build_compound_market_series(
                "^IBEX",
                "IBEX 35",
                growth=Decimal("1.0060"),
                start_price=Decimal("100.0000"),
            )
            for stock_point, reference_point in zip(stock_series.points, reference_series.points):
                position.price_history.create(
                    price_date=stock_point["date"],
                    open_price=stock_point["open"],
                    high_price=stock_point["high"],
                    low_price=stock_point["low"],
                    close_price=stock_point["close"],
                    benchmark_close=reference_point["close"],
                )
            return position

        iberdrola = create_position("IBE", "Iberdrola", Decimal("10.0000"))
        enagas = create_position("ENG", "Enagas", Decimal("14.0000"))
        iberdrola.current_price_per_share = Decimal("12.0000")
        iberdrola.save(update_fields=["current_price_per_share", "updated_at"])
        enagas.current_price_per_share = Decimal("15.5000")
        enagas.save(update_fields=["current_price_per_share", "updated_at"])
        first_cards = build_equity_history_cards([iberdrola, enagas])
        capture_equity_ticket_snapshots(first_cards, snapshot_date=date(2026, 4, 12))

        scyr = create_position("SCYR", "Sacyr", Decimal("4.0000"))
        iberdrola.current_price_per_share = Decimal("12.6000")
        iberdrola.save(update_fields=["current_price_per_share", "updated_at"])
        scyr.current_price_per_share = Decimal("4.3000")
        scyr.save(update_fields=["current_price_per_share", "updated_at"])
        second_cards = build_equity_history_cards(list(EquityPosition.objects.prefetch_related("price_history")))
        capture_equity_ticket_snapshots(second_cards, snapshot_date=date(2026, 4, 13))

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            tracking = build_equity_ticket_tracking_context(second_cards)

        tickets_by_ticker = {item["position"].ticker: item for item in tracking["tickets"]}
        self.assertEqual(tickets_by_ticker["IBE"]["baseline_snapshot"].snapshot_date, date(2026, 4, 12))
        self.assertEqual(tickets_by_ticker["IBE"]["shared_baseline_snapshot"].snapshot_date, date(2026, 4, 13))
        self.assertEqual(tickets_by_ticker["IBE"]["actual_change_pct"], Decimal("5.00"))
        self.assertEqual(tickets_by_ticker["SCYR"]["actual_change_pct"], Decimal("0.00"))
        self.assertEqual(tracking["global"]["baseline_value"], Decimal("636.00"))
        self.assertEqual(tracking["global"]["latest_value"], Decimal("648.00"))
        self.assertEqual(tracking["global"]["net_gain_value"], Decimal("12.00"))
        self.assertEqual(tracking["global"]["invested_return_pct"], Decimal("1.89"))
        self.assertEqual(tracking["global"]["daily_change_pct"], Decimal("1.89"))

    def test_ticket_tracking_portfolio_summary_aggregates_real_history_and_future_projection(self):
        iberdrola = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("11.0000"),
        )
        enagas = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="ENG",
            quote_symbol="ENG.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Enagas",
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("20.0000"),
            current_price_per_share=Decimal("19.0000"),
        )

        populate_position_history_from_closes(
            iberdrola,
            [Decimal("10.00"), Decimal("10.40"), Decimal("10.80"), Decimal("11.00")],
            benchmark_closes=[Decimal("100.00"), Decimal("102.00"), Decimal("104.00"), Decimal("106.00")],
            start_year=2026,
            start_month=1,
        )
        populate_position_history_from_closes(
            enagas,
            [Decimal("20.00"), Decimal("19.50"), Decimal("19.20"), Decimal("19.00")],
            benchmark_closes=[Decimal("100.00"), Decimal("102.00"), Decimal("104.00"), Decimal("106.00")],
            start_year=2026,
            start_month=1,
        )

        cards = build_equity_history_cards(
            list(EquityPosition.objects.prefetch_related("price_history"))
        )
        capture_equity_ticket_snapshots(cards, snapshot_date=date(2026, 4, 30))
        tracking = build_equity_ticket_tracking_context(cards)

        portfolio_summary = tracking["global"]["portfolio_summary"]
        expected_total_value_12m = sum(
            (
                (card["position"].shares * Decimal(str(card["projection"]["projected_price"]))).quantize(Decimal("0.01"))
            )
            for card in cards
            if card.get("projection", {}).get("projected_price") is not None
        ) if any(card.get("projection", {}).get("projected_price") is not None for card in cards) else ZERO
        self.assertTrue(portfolio_summary["available"])
        self.assertTrue(portfolio_summary["net_chart_12m"]["available"])
        self.assertTrue(portfolio_summary["return_chart_12m"]["available"])
        self.assertEqual(portfolio_summary["actual_series_12m"][0]["value"], Decimal("500.00"))
        self.assertEqual(portfolio_summary["actual_series_12m"][-1]["value"], Decimal("490.00"))
        self.assertEqual(portfolio_summary["expected_series_12m"][0]["value"], Decimal("490.00"))
        self.assertEqual(portfolio_summary["expected_series_12m"][-1]["value"], expected_total_value_12m)
        self.assertIn("Peso actual:", portfolio_summary["weight_mix_label"])

    def test_ticket_tracking_portfolio_summary_uses_common_history_start_for_all_positions(self):
        iberdrola = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        sabadell = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="SAB",
            quote_symbol="SAB.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Banco Sabadell",
            shares=Decimal("15.0000"),
            average_cost_per_share=Decimal("1.0000"),
            current_price_per_share=Decimal("1.2500"),
        )

        populate_position_history_from_closes(
            iberdrola,
            [Decimal("10.00"), Decimal("10.50"), Decimal("11.10"), Decimal("11.70"), Decimal("12.00")],
            start_year=2025,
            start_month=12,
        )
        populate_position_history_from_closes(
            sabadell,
            [Decimal("1.00"), Decimal("1.10"), Decimal("1.25")],
            start_year=2026,
            start_month=2,
        )

        cards = build_equity_history_cards(
            list(EquityPosition.objects.prefetch_related("price_history"))
        )
        capture_equity_ticket_snapshots(cards, snapshot_date=date(2026, 4, 30))
        tracking = build_equity_ticket_tracking_context(cards)

        portfolio_summary = tracking["global"]["portfolio_summary"]
        self.assertTrue(portfolio_summary["available"])
        self.assertEqual(portfolio_summary["actual_series_12m"][0]["date"], date(2026, 2, 28))
        self.assertEqual(portfolio_summary["range_label_12m"], "Historico comun desde 2026-02-28")

    def test_ticket_tracking_builds_weekly_and_cumulative_alpha_charts(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("20.00"),
        )
        stock_series = build_compound_market_series(
            "IBE.MC",
            "Iberdrola",
            growth=Decimal("1.0180"),
            start_price=Decimal("10.0000"),
        )
        for stock_point in stock_series.points:
            position.price_history.create(
                price_date=stock_point["date"],
                open_price=stock_point["open"],
                high_price=stock_point["high"],
                low_price=stock_point["low"],
                close_price=stock_point["close"],
                benchmark_close=Decimal("100.0000"),
            )

        snapshot_dates = [date(2026, 4, 13) + timedelta(days=offset) for offset in range(9)]
        current_price = Decimal("10.0000")
        for snapshot_date in snapshot_dates:
            current_price = (current_price + Decimal("0.1500")).quantize(Decimal("0.0001"))
            position.current_price_per_share = current_price
            position.save(update_fields=["current_price_per_share", "updated_at"])
            cards = build_equity_history_cards([position])
            capture_equity_ticket_snapshots(cards, snapshot_date=snapshot_date)

        benchmark_points = []
        benchmark_price = Decimal("100.0000")
        for snapshot_date in snapshot_dates:
            benchmark_price = (benchmark_price + Decimal("0.4000")).quantize(Decimal("0.0001"))
            benchmark_points.append(
                {
                    "date": snapshot_date,
                    "open": benchmark_price,
                    "high": benchmark_price,
                    "low": benchmark_price,
                    "close": benchmark_price,
                }
            )
        benchmark_series = MarketSeries(
            symbol="^IBEX",
            name="IBEX 35",
            latest_price=benchmark_points[-1]["close"],
            latest_date=benchmark_points[-1]["date"],
            points=benchmark_points,
        )

        cards = build_equity_history_cards(list(EquityPosition.objects.prefetch_related("price_history")))
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            tracking = build_equity_ticket_tracking_context(cards)

        self.assertTrue(tracking["global"]["cumulative_alpha_chart"]["available"])
        self.assertTrue(tracking["global"]["weekly_alpha_chart"]["available"])
        self.assertTrue(tracking["global"]["cumulative_alpha_chart"]["actual_line"])
        self.assertTrue(tracking["global"]["weekly_alpha_chart"]["actual_line"])
        self.assertIsNotNone(tracking["global"]["cumulative_alpha_chart"]["zero_y"])

    def test_build_portfolio_correlation_context_calculates_matrix_and_dalio_risk(self):
        close_sets = (
            ("IBE", "Iberdrola", [Decimal("20.0000"), Decimal("19.4000"), Decimal("20.1760"), Decimal("19.7725"), Decimal("20.7611"), Decimal("20.1383"), Decimal("21.3466"), Decimal("20.9197")]),
            ("ENG", "Enagas", [Decimal("18.0000"), Decimal("17.8200"), Decimal("18.3546"), Decimal("18.1709"), Decimal("18.7160"), Decimal("18.3417"), Decimal("19.0754"), Decimal("18.6940")]),
            ("BBVA", "BBVA", [Decimal("12.0000"), Decimal("13.0200"), Decimal("12.4341"), Decimal("13.2413"), Decimal("13.0427"), Decimal("14.0209"), Decimal("13.4601"), Decimal("14.2677")]),
            ("SAN", "Banco Santander", [Decimal("10.0000"), Decimal("11.0000"), Decimal("10.4500"), Decimal("11.2860"), Decimal("11.0603"), Decimal("11.9451"), Decimal("11.3478"), Decimal("12.1421")]),
        )

        for ticker, company_name, closes in close_sets:
            position = EquityPosition.objects.create(
                ownership_category=AssetOwnershipCategory.JOINT,
                broker="Interactive Brokers",
                ticker=ticker,
                quote_symbol=f"{ticker}.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name=company_name,
                shares=Decimal("10.0000"),
                average_cost_per_share=closes[0],
                current_price_per_share=closes[-1],
            )
            populate_position_history_from_closes(position, closes)

        history_cards = build_equity_history_cards(
            list(EquityPosition.objects.prefetch_related("price_history"))
        )
        correlation = build_portfolio_correlation_context(history_cards)

        self.assertTrue(correlation["available"])
        self.assertEqual(correlation["positions_count"], 4)
        self.assertEqual(correlation["pair_count"], 6)
        self.assertEqual(correlation["same_sector_pairs_count"], 1)
        self.assertEqual(correlation["related_sector_pairs_count"], 1)
        self.assertEqual(correlation["distinct_sector_pairs_count"], 4)
        self.assertEqual(correlation["average_correlation"], Decimal("-0.28"))
        self.assertEqual(correlation["average_positive_correlation_pct"], Decimal("34.55"))
        self.assertEqual(correlation["estimated_loss_probability_pct"], Decimal("33.18"))
        self.assertEqual(correlation["highest_pair"]["pair_label"], "BBVA / SAN")
        self.assertEqual(correlation["most_diversifying_pair"]["pair_label"], "IBE / SAN")
        self.assertEqual(correlation["heatmap_rows"][0]["cells"][0]["label"], "1.00")
        self.assertEqual(len(correlation["heatmap_headers"]), 4)
        self.assertEqual(len(correlation["heatmap_rows"]), 4)
        self.assertTrue(correlation["risk_curve_chart"]["available"])
        self.assertIsNotNone(correlation["risk_curve_chart"]["current_marker"])

    def test_build_owned_cycle_trade_timing_plan_detects_monthly_trend_turns(self):
        position = EquityPosition(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2023, 1, 31),
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("100.0000"),
            latest_price_date=date(2026, 4, 17),
        )
        card = {
            "position": position,
            "cycle_projection_5y": {
                "available": True,
                "path": [
                    {"label": "6M", "projected_price": Decimal("118.0000")},
                    {"label": "12M", "projected_price": Decimal("126.0000")},
                    {"label": "18M", "projected_price": Decimal("112.0000")},
                    {"label": "24M", "projected_price": Decimal("96.0000")},
                    {"label": "30M", "projected_price": Decimal("101.0000")},
                    {"label": "36M", "projected_price": Decimal("118.0000")},
                    {"label": "42M", "projected_price": Decimal("132.0000")},
                    {"label": "48M", "projected_price": Decimal("145.0000")},
                    {"label": "54M", "projected_price": Decimal("156.0000")},
                    {"label": "60M", "projected_price": Decimal("168.0000")},
                ],
            },
        }

        plan = build_owned_cycle_trade_timing_plan(card)

        self.assertTrue(plan["available"])
        self.assertEqual(plan["mode"], "sale_reentry")
        self.assertEqual(plan["sale_month_number"], 12)
        self.assertEqual(plan["sale_window_label"], "abril 2027 (mes 12)")
        self.assertEqual(plan["reentry_month_number"], 26)
        self.assertEqual(plan["reentry_window_label"], "junio 2028 (mes 26)")
        self.assertEqual(plan["signal_value_pct"], Decimal("-0.13"))
        self.assertEqual(plan["pre_sale_return_pct"], Decimal("26.00"))
        self.assertIn("pendiente desestacionalizada de 5 meses", plan["summary"].lower())

    def test_ticket_tracking_includes_sale_and_reentry_plan_from_current_cycle_trend(self):
        position = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2023, 1, 31),
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
            annual_dividend_income=Decimal("20.00"),
        )
        populate_position_history(position, growth=Decimal("1.0150"), benchmark_growth=Decimal("1.0060"), months=36)
        alternative = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Seguimiento",
            ticker="ELE",
            quote_symbol="ELE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Endesa",
            shares=Decimal("0.0000"),
            average_cost_per_share=Decimal("18.0000"),
            current_price_per_share=Decimal("18.0000"),
        )
        populate_position_history(alternative, growth=Decimal("1.0220"), benchmark_growth=Decimal("1.0060"), months=36)

        cards = build_equity_history_cards([position, alternative])
        capture_equity_ticket_snapshots(cards, snapshot_date=date(2026, 4, 16))
        position.current_price_per_share = Decimal("12.3000")
        position.save(update_fields=["current_price_per_share", "updated_at"])
        refreshed_cards = build_equity_history_cards(list(EquityPosition.objects.prefetch_related("price_history")))
        capture_equity_ticket_snapshots(refreshed_cards, snapshot_date=date(2026, 4, 17))

        EquityPurchaseForecastBaseline.objects.create(
            position=position,
            source_analysis_date=date(2026, 4, 16),
            baseline_date=date(2026, 4, 16),
            reference_label="IBEX 35",
            trade_alert_label="Comprar",
            reliability_label="Alta",
            baseline_price=Decimal("10.0000"),
            projected_price_1y=Decimal("11.2000"),
            projected_price_2y=Decimal("9.9000"),
            projected_price_3y=Decimal("11.7000"),
            projected_price_4y=Decimal("12.4000"),
            projected_price_5y=Decimal("13.2000"),
            projected_return_pct_1y=Decimal("12.00"),
            projected_return_pct_2y=Decimal("-1.00"),
            projected_return_pct_3y=Decimal("17.00"),
            projected_return_pct_4y=Decimal("24.00"),
            projected_return_pct_5y=Decimal("32.00"),
        )

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        trade_plan_payload = {
            "available": True,
            "mode": "sale_reentry",
            "analysis_basis_label": "Pendiente 5M desestacionalizada sobre la senda 5A vigente",
            "sale_month_number": 12,
            "sale_year_number": 1,
            "sale_window_label": "abril 2027 (mes 12)",
            "sale_date": date(2027, 4, 17),
            "sale_date_label": "2027-04-17",
            "reentry_month_number": 26,
            "reentry_year_number": 3,
            "reentry_window_label": "junio 2028 (mes 26)",
            "reentry_date": date(2028, 6, 17),
            "reentry_date_label": "2028-06-17",
            "summary": "La pendiente desestacionalizada de 5 meses gira a negativo y luego vuelve a positivo.",
            "signal_label": "Pendiente 5M negativa",
            "signal_value_pct": Decimal("-0.13"),
            "monthly_rows": [],
            "yearly_rows": [],
            "drawdown_month_number": 12,
            "drawdown_year_number": 1,
            "drawdown_margin_pct": Decimal("-0.13"),
            "pre_sale_return_pct": Decimal("8.00"),
        }
        with (
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
            patch("equities.services.build_owned_cycle_trade_timing_plan", return_value=trade_plan_payload),
        ):
            tracking = build_equity_ticket_tracking_context(refreshed_cards, optimizer_cards=refreshed_cards)

        trade_plan = tracking["tickets"][0]["trade_plan"]
        self.assertTrue(trade_plan["available"])
        self.assertEqual(trade_plan["mode"], "sale_reentry")
        self.assertEqual(trade_plan["sale_month_number"], 12)
        self.assertEqual(trade_plan["sale_window_label"], "abril 2027 (mes 12)")
        self.assertEqual(trade_plan["reentry_month_number"], 26)
        self.assertEqual(trade_plan["reentry_window_label"], "junio 2028 (mes 26)")
        self.assertEqual(trade_plan["drawdown_month_number"], 12)
        self.assertEqual(trade_plan["drawdown_margin_pct"], Decimal("-0.13"))
        self.assertEqual(tracking["tickets"][0]["rotation_plan"]["action"], "rotar")
        self.assertEqual(tracking["tickets"][0]["rotation_plan"]["alternative_ticker"], "ELE")
        sale_timeline = tracking["sale_timeline"]
        self.assertTrue(sale_timeline["available"])
        self.assertEqual(sale_timeline["horizon_months"], 24)
        self.assertEqual(sale_timeline["scheduled_count"], 1)
        self.assertEqual(sale_timeline["alert_count"], 0)
        self.assertEqual(sale_timeline["unscheduled_count"], 0)
        self.assertEqual(sale_timeline["next_row"]["ticker"], "IBE")
        self.assertEqual(sale_timeline["next_row"]["sale_window_label"], "abril 2027 (mes 12)")
        self.assertEqual(sale_timeline["next_row"]["projected_sale_price"], Decimal("13.2840"))
        expected_sale_preview = build_equity_sale_preview(
            position,
            sale_price_per_share=Decimal("13.2840"),
            closed_on=date(2027, 4, 17),
        )
        self.assertEqual(sale_timeline["next_row"]["estimated_net_result"], expected_sale_preview["net_result"])
        self.assertEqual(sale_timeline["projected_net_result_total"], expected_sale_preview["net_result"])

    def test_round_investment_plan_respects_existing_weights_and_review_dates(self):
        owned = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.OWNED,
                broker="Interactive Brokers",
                ticker="IBE",
                quote_symbol="IBE.MC",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Iberdrola",
                shares=Decimal("1000.0000"),
                average_cost_per_share=Decimal("20.0000"),
                current_price_per_share=Decimal("22.0000"),
            ),
            "status_key": "owned",
            "status_label": "Comprada",
            "sector_label": "Electrica",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("20.00"),
                "price_return_pct": Decimal("16.00"),
                "price_low_return_pct": Decimal("-8.00"),
                "price_high_return_pct": Decimal("28.00"),
                "projected_price": Decimal("26.4000"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.20"),
                "net_income_yield_pct": Decimal("2.80"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-18.00"),
            },
            "cycle_projection_5y": {"available": True, "annual_return_pct": Decimal("6.00"), "five_year_return_pct": Decimal("34.00")},
        }
        endesa = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Seguimiento",
                ticker="ELE",
                quote_symbol="ELE.MC",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Endesa",
                shares=Decimal("0.0000"),
                average_cost_per_share=Decimal("18.0000"),
                current_price_per_share=Decimal("18.0000"),
            ),
            "status_key": "watchlist",
            "status_label": "Seguimiento",
            "sector_label": "Electrica",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("18.00"),
                "price_return_pct": Decimal("14.50"),
                "price_low_return_pct": Decimal("-6.00"),
                "price_high_return_pct": Decimal("24.00"),
                "projected_price": Decimal("20.6100"),
                "confidence_label": "Alta",
                "safety_score": Decimal("75.00"),
                "gross_dividend_yield_pct": Decimal("3.50"),
                "net_income_yield_pct": Decimal("3.00"),
                "transaction_drag_pct": Decimal("0.10"),
                "annualized_volatility_pct": Decimal("11.50"),
                "positive_year_ratio_pct": Decimal("71.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-17.00"),
            },
            "cycle_projection_5y": {"available": True, "annual_return_pct": Decimal("5.80"), "five_year_return_pct": Decimal("32.00")},
        }
        indra = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Seguimiento",
                ticker="IDR",
                quote_symbol="IDR.MC",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Indra",
                shares=Decimal("0.0000"),
                average_cost_per_share=Decimal("18.0000"),
                current_price_per_share=Decimal("18.0000"),
            ),
            "status_key": "watchlist",
            "status_label": "Seguimiento",
            "sector_label": "Tecnologia y defensa",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("86.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("24.00"),
                "price_return_pct": Decimal("21.50"),
                "price_low_return_pct": Decimal("-7.00"),
                "price_high_return_pct": Decimal("34.00"),
                "projected_price": Decimal("22.3200"),
                "confidence_label": "Alta",
                "safety_score": Decimal("78.00"),
                "gross_dividend_yield_pct": Decimal("1.20"),
                "net_income_yield_pct": Decimal("1.00"),
                "transaction_drag_pct": Decimal("0.10"),
                "annualized_volatility_pct": Decimal("16.00"),
                "positive_year_ratio_pct": Decimal("73.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-4.00"),
                "max_drawdown_pct": Decimal("-22.00"),
            },
            "cycle_projection_5y": {"available": True, "annual_return_pct": Decimal("7.00"), "five_year_return_pct": Decimal("41.00")},
        }

        plan = build_equity_round_investment_plan(
            [owned],
            [owned, endesa, indra],
            Decimal("70000"),
            Decimal("10000"),
            Decimal("30"),
            as_of=date(2026, 4, 17),
        )

        self.assertTrue(plan["available"])
        self.assertEqual(plan["current_overweights_count"], 1)
        self.assertEqual(plan["rounds"][0]["round_date_label"], "2026-04-17")
        self.assertEqual(plan["rounds"][1]["round_date_label"], "2026-04-21")
        self.assertTrue(all(item["amount"] <= Decimal("10000") for item in plan["rounds"]))
        self.assertTrue(all(item["post_weight_pct"] <= Decimal("30.00") for item in plan["rounds"]))
        self.assertTrue(all(item["ticker"] != "IBE" for item in plan["rounds"]))

    def test_investment_journey_builds_active_and_closed_ticket_history(self):
        active = EquityPosition.objects.create(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 15),
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.5000"),
            annual_dividend_income=Decimal("30.00"),
            annual_maintenance_cost=Decimal("6.00"),
        )
        sold = EquityPosition.objects.create(
            broker="Banco Sabadell",
            ticker="ACS",
            quote_symbol="ACS.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            opened_on=date(2023, 2, 1),
            shares=Decimal("15.0000"),
            average_cost_per_share=Decimal("20.0000"),
            current_price_per_share=Decimal("24.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("10.00"),
        )
        for position, start_price, growth in (
            (active, Decimal("10.00"), Decimal("1.0100")),
            (sold, Decimal("20.00"), Decimal("1.0120")),
        ):
            stock_series = build_compound_market_series(
                position.quote_symbol,
                position.company_name,
                growth=growth,
                start_price=start_price,
            )
            reference_series = build_compound_market_series(
                "^IBEX",
                "IBEX 35",
                growth=Decimal("1.0060"),
                start_price=Decimal("100.0000"),
            )
            for stock_point, reference_point in zip(stock_series.points, reference_series.points):
                position.price_history.create(
                    price_date=stock_point["date"],
                    open_price=stock_point["open"],
                    high_price=stock_point["high"],
                    low_price=stock_point["low"],
                    close_price=stock_point["close"],
                    benchmark_close=reference_point["close"],
                )

        archive_equity_position_sale(
            sold,
            closed_on=date(2025, 9, 30),
            sale_price_per_share=Decimal("25.0000"),
            notes="Cierre completo",
        )
        context = build_equity_investment_journey_context(
            list(EquityPosition.objects.prefetch_related("price_history")),
            list(EquityClosedPosition.objects.all()),
        )

        self.assertTrue(context["available"])
        self.assertEqual(context["active_count"], 1)
        self.assertEqual(context["closed_count"], 1)
        self.assertTrue(context["value_chart"]["available"])
        self.assertTrue(context["profit_chart"]["available"])
        self.assertTrue(context["annual_result_chart"]["available"])
        self.assertTrue(context["annual_rows"])
        self.assertIsNotNone(context["current_year_row"])
        self.assertEqual(context["closed_tickets"][0]["status_label"], "Vendida")
        self.assertIsNotNone(context["cumulative_margin_pct"])
        self.assertIsNotNone(context["monthly_equivalent_return_pct"])
        self.assertIsNotNone(context["active_tickets"][0]["monthly_equivalent_return_pct"])
        self.assertGreaterEqual(context["costs_total"], Decimal("0.00"))
        self.assertEqual(
            context["purchase_cost_total"] + context["maintenance_cost_total"] + context["sale_cost_total"],
            context["costs_total"],
        )
        self.assertEqual(
            context["costs_paid_total"] + context["open_sale_cost_reserve_total"],
            context["costs_total"],
        )
        self.assertIsNotNone(context["average_annual_result"])

    def test_sale_preview_calculates_net_result_for_a_specific_sale_price(self):
        position = EquityPosition.objects.create(
            broker="Banco Santander",
            trade_channel=EquityPosition.TradeChannel.APP,
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2025, 4, 12),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.5000"),
            annual_dividend_income=Decimal("12.00"),
            annual_maintenance_cost=Decimal("5.00"),
        )

        preview = build_equity_sale_preview(
            position,
            sale_price_per_share=Decimal("12.5000"),
            closed_on=date(2026, 4, 12),
        )

        self.assertTrue(preview["available"])
        self.assertEqual(preview["purchase_cost"], Decimal("3.20"))
        self.assertEqual(preview["sale_total_cost"], Decimal("3.00"))
        self.assertEqual(preview["net_exit_value"], Decimal("122.00"))
        self.assertEqual(preview["dividend_total"], Decimal("10.00"))
        self.assertEqual(preview["maintenance_total"], Decimal("5.00"))
        self.assertEqual(preview["net_result"], Decimal("23.80"))
        self.assertGreater(preview["monthly_equivalent_return_pct"], Decimal("0.00"))
        self.assertEqual(preview["annualized_margin_pct"], preview["cumulative_margin_pct"])

    def test_sale_preview_handles_recent_positions_with_extreme_annualization(self):
        position = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.MONICA,
            broker="Cartera Monica",
            ticker="SAN",
            quote_symbol="SAN.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Banco Santander",
            opened_on=date(2026, 4, 19),
            shares=Decimal("2550.0000"),
            average_cost_per_share=Decimal("11.0420"),
            current_price_per_share=Decimal("11.0420"),
        )
        populate_position_history(position)

        preview = build_equity_sale_preview(
            position,
            closed_on=date(2026, 4, 21),
        )

        self.assertTrue(preview["available"])
        self.assertEqual(preview["holding_days"], 2)
        self.assertGreater(preview["annualized_margin_pct"], preview["cumulative_margin_pct"])

    def test_allocation_plan_respects_max_company_weight_and_sorts_by_projection(self):
        stronger = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IDR",
                company_name="Indra",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("18"),
            ),
            "reference_label": "IBEX 35",
            "projection": {
                "available": True,
                "base_return_pct": Decimal("24.0"),
                "projected_price": Decimal("22.32"),
                "confidence_label": "Alta",
                "coefficient": Decimal("0.55"),
            },
        }
        medium = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.OWNED,
                ticker="IBE",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
                benchmark_name="Demanda electrica Espana",
                benchmark_symbol="REE:demand:es:peninsular",
                shares=Decimal("10"),
                average_cost_per_share=Decimal("10"),
                current_price_per_share=Decimal("14"),
            ),
            "reference_label": "Demanda electrica Espana",
            "projection": {
                "available": True,
                "base_return_pct": Decimal("16.0"),
                "projected_price": Decimal("16.24"),
                "confidence_label": "Media",
                "coefficient": Decimal("0.31"),
            },
        }
        weak = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="MRL",
                company_name="Merlin",
                reference_profile=EquityPosition.ReferenceProfile.SPAIN_HOUSE_PRICE,
                benchmark_name="Precio vivienda Espana",
                benchmark_symbol="EUROSTAT:prc_hpi_q:ES:TOTAL:I15_Q",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("10"),
            ),
            "reference_label": "Precio vivienda Espana",
            "projection": {
                "available": True,
                "base_return_pct": Decimal("8.0"),
                "projected_price": Decimal("10.80"),
                "confidence_label": "Media",
                "safety_score": Decimal("60.00"),
                "coefficient": Decimal("0.10"),
            },
            "projection_reliability": {"label": "Media", "score": Decimal("62.00")},
        }

        plan = build_equity_allocation_plan([medium, weak, stronger], Decimal("100000"), Decimal("40"))

        self.assertTrue(plan["available"])
        self.assertEqual(len(plan["allocations"]), 3)
        self.assertEqual(plan["allocations"][0]["position"].ticker, "IDR")
        self.assertEqual(plan["allocations"][0]["allocated_amount"], Decimal("40000"))
        self.assertEqual(plan["allocations"][1]["allocated_amount"], Decimal("40000"))
        self.assertEqual(plan["allocations"][2]["allocated_amount"], Decimal("20000"))
        self.assertEqual(plan["cash_reserve_amount"], Decimal("0"))
        self.assertGreater(plan["projected_gain_total"], Decimal("0"))

    def test_allocation_plan_uses_five_year_cycle_as_secondary_signal(self):
        near_term_only = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="LOG",
                company_name="Logista",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("26"),
            ),
            "sector_label": "Distribucion",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Vigilar", "tone": "watch", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("14.40"),
                "projected_price": Decimal("29.74"),
                "confidence_label": "Alta",
                "coefficient": Decimal("0.48"),
                "safety_score": Decimal("74.00"),
                "net_income_yield_pct": Decimal("2.10"),
                "gross_dividend_yield_pct": Decimal("2.90"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("13.00"),
                "positive_year_ratio_pct": Decimal("67.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-19.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("-4.50"),
                "five_year_return_pct": Decimal("-20.00"),
            },
        }
        balanced_cycle = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IBE",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
                benchmark_name="Demanda electrica Espana",
                benchmark_symbol="REE:demand:es:peninsular",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "sector_label": "Energia",
            "reference_label": "Demanda electrica Espana",
            "trade_alert": {"label": "Vigilar", "tone": "watch", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("13.90"),
                "projected_price": Decimal("15.95"),
                "confidence_label": "Alta",
                "coefficient": Decimal("0.52"),
                "safety_score": Decimal("75.00"),
                "net_income_yield_pct": Decimal("2.20"),
                "gross_dividend_yield_pct": Decimal("3.10"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("12.50"),
                "positive_year_ratio_pct": Decimal("69.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-18.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.20"),
                "five_year_return_pct": Decimal("35.00"),
            },
        }

        plan = build_equity_allocation_plan([near_term_only, balanced_cycle], Decimal("100000"), Decimal("60"))

        self.assertTrue(plan["available"])
        self.assertEqual(plan["allocations"][0]["position"].ticker, "IBE")
        self.assertGreater(plan["allocations"][0]["cycle_support_score"], Decimal("0"))
        self.assertGreater(plan["weighted_cycle_return_annual_pct"], Decimal("0"))

    def test_allocation_plan_can_prioritize_five_year_cycle(self):
        short_term_better = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="REP",
                company_name="Repsol",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "sector_label": "Energia",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Vigilar", "tone": "watch", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("18.50"),
                "projected_price": Decimal("16.59"),
                "confidence_label": "Alta",
                "coefficient": Decimal("0.50"),
                "safety_score": Decimal("68.00"),
                "net_income_yield_pct": Decimal("1.80"),
                "gross_dividend_yield_pct": Decimal("2.80"),
                "transaction_drag_pct": Decimal("0.30"),
                "annualized_volatility_pct": Decimal("18.00"),
                "positive_year_ratio_pct": Decimal("58.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Transicion",
                "current_drawdown_pct": Decimal("-5.00"),
                "max_drawdown_pct": Decimal("-28.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("80.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("-2.20"),
                "five_year_return_pct": Decimal("-10.50"),
            },
        }
        long_term_better = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IBE",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
                benchmark_name="Demanda electrica Espana",
                benchmark_symbol="REE:demand:es:peninsular",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "sector_label": "Electrica",
            "reference_label": "Demanda electrica Espana",
            "trade_alert": {"label": "Vigilar", "tone": "watch", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("8.80"),
                "projected_price": Decimal("15.23"),
                "confidence_label": "Alta",
                "coefficient": Decimal("0.52"),
                "safety_score": Decimal("76.00"),
                "net_income_yield_pct": Decimal("2.40"),
                "gross_dividend_yield_pct": Decimal("3.20"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-18.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.60"),
                "five_year_return_pct": Decimal("37.50"),
            },
        }

        plan = build_equity_allocation_plan(
            [short_term_better, long_term_better],
            Decimal("100000"),
            Decimal("60"),
            strategy_mode="5y_primary",
        )

        self.assertTrue(plan["available"])
        self.assertEqual(plan["strategy_mode"], "5y_primary")
        self.assertEqual(plan["strategy_label"], "5A principal")
        self.assertEqual(plan["allocations"][0]["position"].ticker, "IBE")
        self.assertGreater(plan["allocations"][0]["cycle_return_annual_pct"], Decimal("0"))

    def test_allocation_plan_prefers_more_robust_scenarios_over_flashy_central_case(self):
        robust_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                trade_channel=EquityPosition.TradeChannel.APP,
                ticker="IBE",
                quote_symbol="IBE.MC",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
                annual_maintenance_cost=Decimal("0"),
            ),
            "sector_label": "Electrica",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("16.00"),
                "price_return_pct": Decimal("13.80"),
                "price_low_return_pct": Decimal("5.00"),
                "price_high_return_pct": Decimal("23.00"),
                "projected_price": Decimal("15.9320"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.20"),
                "net_income_yield_pct": Decimal("2.40"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-18.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("8.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "total_return_pct": Decimal("16.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("24.00")},
                ],
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.10"),
                "five_year_return_pct": Decimal("34.50"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("4.10")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "annual_return_pct": Decimal("6.10")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("7.70")},
                ],
            },
            "external_signal": {"label": "Neutral", "score": Decimal("0.0"), "items_count": 1, "note": ""},
        }
        flashy_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                trade_channel=EquityPosition.TradeChannel.APP,
                ticker="REP",
                quote_symbol="REP.MC",
                company_name="Repsol",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
                annual_maintenance_cost=Decimal("0"),
            ),
            "sector_label": "Energia",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Vigilar", "tone": "watch", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("18.50"),
                "price_return_pct": Decimal("15.60"),
                "price_low_return_pct": Decimal("-20.00"),
                "price_high_return_pct": Decimal("42.00"),
                "projected_price": Decimal("16.1840"),
                "confidence_label": "Alta",
                "safety_score": Decimal("69.00"),
                "gross_dividend_yield_pct": Decimal("3.60"),
                "net_income_yield_pct": Decimal("2.70"),
                "transaction_drag_pct": Decimal("0.30"),
                "annualized_volatility_pct": Decimal("18.50"),
                "positive_year_ratio_pct": Decimal("58.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Transicion",
                "current_drawdown_pct": Decimal("-5.00"),
                "max_drawdown_pct": Decimal("-28.00"),
                "news_adjustment": {
                    "applied": True,
                    "confidence_penalty_pct": Decimal("12.00"),
                    "band_multiplier": Decimal("1.35"),
                },
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("36.0"), "total_return_pct": Decimal("-22.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("34.0"), "total_return_pct": Decimal("18.50")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("30.0"), "total_return_pct": Decimal("46.00")},
                ],
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("80.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.40"),
                "five_year_return_pct": Decimal("36.50"),
                "news_adjustment": {"applied": True, "spread_multiplier": Decimal("1.30")},
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("36.0"), "annual_return_pct": Decimal("-2.50")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("34.0"), "annual_return_pct": Decimal("6.40")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("30.0"), "annual_return_pct": Decimal("12.50")},
                ],
            },
            "external_signal": {"label": "Prensa adversa", "score": Decimal("-3.40"), "items_count": 4, "note": "Contexto inestable."},
        }

        flashy_candidate = build_equity_optimizer_candidate(flashy_card)
        plan = build_equity_allocation_plan([flashy_card, robust_card], Decimal("100000"), Decimal("60"))

        self.assertIsNotNone(flashy_candidate)
        self.assertGreater(flashy_candidate["uncertainty_penalty_pct"], ZERO)
        self.assertLess(flashy_candidate["robust_return_signal_pct"], flashy_candidate["risk_adjusted_return_pct"])
        self.assertTrue(plan["available"])
        self.assertEqual(plan["allocations"][0]["position"].ticker, "IBE")
        self.assertGreater(plan["allocations"][0]["scenario_expected_return_pct"], Decimal("0"))
        self.assertGreater(plan["weighted_expected_return_pct"], Decimal("0"))
        self.assertTrue(
            all(item["position"].ticker != "REP" for item in plan["allocations"])
            or any(item["uncertainty_penalty_pct"] > ZERO for item in plan["allocations"])
        )

    def test_allocation_plan_5y_primary_penalizes_shock_adjusted_cycle_scenarios(self):
        stable_compounder = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                trade_channel=EquityPosition.TradeChannel.APP,
                ticker="IBE",
                quote_symbol="IBE.MC",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
                annual_maintenance_cost=Decimal("0"),
            ),
            "sector_label": "Electrica",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("11.50"),
                "price_return_pct": Decimal("9.30"),
                "price_low_return_pct": Decimal("3.00"),
                "price_high_return_pct": Decimal("16.00"),
                "projected_price": Decimal("15.3020"),
                "confidence_label": "Alta",
                "safety_score": Decimal("77.00"),
                "gross_dividend_yield_pct": Decimal("3.40"),
                "net_income_yield_pct": Decimal("2.50"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("11.50"),
                "positive_year_ratio_pct": Decimal("71.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-16.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("83.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.20"),
                "five_year_return_pct": Decimal("35.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("4.50")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "annual_return_pct": Decimal("6.20")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("7.60")},
                ],
            },
        }
        shocky_cycle = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                trade_channel=EquityPosition.TradeChannel.APP,
                ticker="REP",
                quote_symbol="REP.MC",
                company_name="Repsol",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
                annual_maintenance_cost=Decimal("0"),
            ),
            "sector_label": "Energia",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Vigilar", "tone": "watch", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("13.00"),
                "price_return_pct": Decimal("10.20"),
                "price_low_return_pct": Decimal("-8.00"),
                "price_high_return_pct": Decimal("22.00"),
                "projected_price": Decimal("15.4280"),
                "confidence_label": "Alta",
                "safety_score": Decimal("70.00"),
                "gross_dividend_yield_pct": Decimal("3.80"),
                "net_income_yield_pct": Decimal("2.90"),
                "transaction_drag_pct": Decimal("0.30"),
                "annualized_volatility_pct": Decimal("18.00"),
                "positive_year_ratio_pct": Decimal("59.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Transicion",
                "current_drawdown_pct": Decimal("-5.00"),
                "max_drawdown_pct": Decimal("-26.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("7.10"),
                "five_year_return_pct": Decimal("40.80"),
                "news_adjustment": {"applied": True, "spread_multiplier": Decimal("1.35")},
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("36.0"), "annual_return_pct": Decimal("-3.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("34.0"), "annual_return_pct": Decimal("7.10")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("30.0"), "annual_return_pct": Decimal("13.20")},
                ],
            },
            "external_signal": {"label": "Prensa adversa", "score": Decimal("-2.60"), "items_count": 3, "note": "Mayor incertidumbre."},
        }

        shocky_candidate = build_equity_optimizer_candidate(shocky_cycle, strategy_mode="5y_primary")
        plan = build_equity_allocation_plan(
            [shocky_cycle, stable_compounder],
            Decimal("100000"),
            Decimal("60"),
            strategy_mode="5y_primary",
        )

        self.assertIsNotNone(shocky_candidate)
        self.assertGreater(shocky_candidate["uncertainty_penalty_pct"], ZERO)
        self.assertLess(shocky_candidate["robust_cycle_support_score"], shocky_candidate["cycle_support_score"])
        self.assertTrue(plan["available"])
        self.assertEqual(plan["allocations"][0]["position"].ticker, "IBE")
        self.assertGreater(plan["allocations"][0]["robust_cycle_support_score"], ZERO)
        self.assertTrue(all(item["position"].ticker != "REP" for item in plan["allocations"]))

    def test_allocation_plan_can_limit_max_positions_per_sector(self):
        utility_best = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IBE",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
                benchmark_name="Demanda electrica Espana",
                benchmark_symbol="REE:demand:es:peninsular",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "sector_label": "Electrica",
            "reference_label": "Demanda electrica Espana",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("16.0"),
                "projected_price": Decimal("16.24"),
                "confidence_label": "Media",
                "coefficient": Decimal("0.31"),
                "safety_score": Decimal("70.00"),
                "net_income_yield_pct": Decimal("2.50"),
                "gross_dividend_yield_pct": Decimal("3.40"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("14.00"),
                "positive_year_ratio_pct": Decimal("68.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-4.00"),
                "max_drawdown_pct": Decimal("-18.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("80.00")},
        }
        utility_weaker = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="ELE",
                company_name="Endesa",
                reference_profile=EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
                benchmark_name="Demanda electrica Espana",
                benchmark_symbol="REE:demand:es:peninsular",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("18"),
            ),
            "sector_label": "Electrica",
            "reference_label": "Demanda electrica Espana",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("10.0"),
                "projected_price": Decimal("19.80"),
                "confidence_label": "Media",
                "coefficient": Decimal("0.28"),
                "safety_score": Decimal("68.00"),
                "net_income_yield_pct": Decimal("2.90"),
                "gross_dividend_yield_pct": Decimal("4.10"),
                "transaction_drag_pct": Decimal("0.15"),
                "annualized_volatility_pct": Decimal("13.00"),
                "positive_year_ratio_pct": Decimal("66.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-3.00"),
                "max_drawdown_pct": Decimal("-16.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
        }
        defense = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IDR",
                company_name="Indra",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("18"),
            ),
            "sector_label": "Tecnologia y defensa",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("24.0"),
                "projected_price": Decimal("22.32"),
                "confidence_label": "Alta",
                "coefficient": Decimal("0.55"),
                "safety_score": Decimal("74.00"),
                "net_income_yield_pct": Decimal("1.20"),
                "gross_dividend_yield_pct": Decimal("1.50"),
                "transaction_drag_pct": Decimal("0.10"),
                "annualized_volatility_pct": Decimal("17.00"),
                "positive_year_ratio_pct": Decimal("72.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-5.00"),
                "max_drawdown_pct": Decimal("-22.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
        }

        plan = build_equity_allocation_plan([utility_weaker, utility_best, defense], Decimal("100000"), Decimal("50"), 0, 1)

        self.assertTrue(plan["available"])
        self.assertEqual(len(plan["allocations"]), 2)
        self.assertEqual(plan["max_sector_positions"], 1)
        self.assertEqual(plan["sector_filtered_count"], 1)
        self.assertEqual({item["sector_label"] for item in plan["allocations"]}, {"Electrica", "Tecnologia y defensa"})
        self.assertEqual({item["position"].ticker for item in plan["allocations"]}, {"IBE", "IDR"})

    def test_allocation_plan_can_limit_total_number_of_companies(self):
        strongest = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IDR",
                company_name="Indra",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("18"),
            ),
            "sector_label": "Tecnologia y defensa",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("24.0"),
                "projected_price": Decimal("22.32"),
                "confidence_label": "Alta",
                "coefficient": Decimal("0.55"),
                "safety_score": Decimal("74.00"),
                "net_income_yield_pct": Decimal("1.20"),
                "gross_dividend_yield_pct": Decimal("1.50"),
                "transaction_drag_pct": Decimal("0.10"),
                "annualized_volatility_pct": Decimal("17.00"),
                "positive_year_ratio_pct": Decimal("72.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-5.00"),
                "max_drawdown_pct": Decimal("-22.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
        }
        second = {
            **strongest,
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IBE",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("11"),
            ),
            "sector_label": "Electrica",
            "projection": {**strongest["projection"], "base_return_pct": Decimal("17.0"), "projected_price": Decimal("12.87")},
        }
        third = {
            **strongest,
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="REP",
                company_name="Repsol",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "sector_label": "Energia",
            "projection": {**strongest["projection"], "base_return_pct": Decimal("12.0"), "projected_price": Decimal("15.68")},
        }

        plan = build_equity_allocation_plan(
            [third, second, strongest],
            Decimal("100000"),
            Decimal("50"),
            2,
            0,
        )

        self.assertTrue(plan["available"])
        self.assertEqual(plan["max_total_positions"], 2)
        self.assertEqual(plan["position_cap_filtered_count"], 1)
        self.assertEqual(len(plan["allocations"]), 2)
        self.assertEqual({item["position"].ticker for item in plan["allocations"]}, {"IDR", "IBE"})

    def test_allocation_plan_can_filter_allowed_sectors(self):
        strongest = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IDR",
                company_name="Indra",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("18"),
            ),
            "sector_label": "Tecnologia y defensa",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("24.0"),
                "projected_price": Decimal("22.32"),
                "confidence_label": "Alta",
                "coefficient": Decimal("0.55"),
                "safety_score": Decimal("74.00"),
                "net_income_yield_pct": Decimal("1.20"),
                "gross_dividend_yield_pct": Decimal("1.50"),
                "transaction_drag_pct": Decimal("0.10"),
                "annualized_volatility_pct": Decimal("17.00"),
                "positive_year_ratio_pct": Decimal("72.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-5.00"),
                "max_drawdown_pct": Decimal("-22.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
        }
        electric = {
            **strongest,
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IBE",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("11"),
            ),
            "sector_label": "Electrica",
            "projection": {**strongest["projection"], "base_return_pct": Decimal("17.0"), "projected_price": Decimal("12.87")},
        }
        energy = {
            **strongest,
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="REP",
                company_name="Repsol",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "sector_label": "Energia",
            "projection": {**strongest["projection"], "base_return_pct": Decimal("12.0"), "projected_price": Decimal("15.68")},
        }

        plan = build_equity_allocation_plan(
            [strongest, electric, energy],
            Decimal("100000"),
            Decimal("50"),
            0,
            0,
            selected_sectors=["Electrica"],
        )

        self.assertTrue(plan["available"])
        self.assertEqual(plan["selected_sectors"], ["Electrica"])
        self.assertIn("Electrica", plan["selected_sector_note"])
        self.assertEqual(len(plan["allocations"]), 1)
        self.assertEqual(plan["allocations"][0]["position"].ticker, "IBE")

    def test_allocation_plan_can_exclude_owned_positions_manually(self):
        owned_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.OWNED,
                ticker="IBE",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("10"),
                average_cost_per_share=Decimal("10"),
                current_price_per_share=Decimal("11"),
            ),
            "sector_label": "Electrica",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("15.00"),
                "price_return_pct": Decimal("11.00"),
                "price_low_return_pct": Decimal("2.00"),
                "price_high_return_pct": Decimal("22.00"),
                "projected_price": Decimal("12.6500"),
                "confidence_label": "Alta",
                "safety_score": Decimal("74.00"),
                "gross_dividend_yield_pct": Decimal("3.50"),
                "net_income_yield_pct": Decimal("3.00"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("11.50"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.50"),
                "max_drawdown_pct": Decimal("-18.00"),
            },
        }
        watchlist_card = {
            **owned_card,
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="ACS",
                company_name="ACS",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("44"),
            ),
            "sector_label": "Construccion e infraestructuras",
            "projection": {
                **owned_card["projection"],
                "base_return_pct": Decimal("12.00"),
                "projected_price": Decimal("49.2800"),
            },
        }

        plan = build_equity_allocation_plan(
            [owned_card, watchlist_card],
            Decimal("100000"),
            Decimal("60"),
            selected_owned_tickers=[],
            selected_owned_tickers_applied=True,
        )

        self.assertTrue(plan["available"])
        self.assertTrue(plan["selected_owned_tickers_applied"])
        self.assertEqual(plan["selected_owned_tickers"], [])
        self.assertIn("desactivado", plan["selected_owned_ticker_note"].lower())
        self.assertEqual([item["position"].ticker for item in plan["allocations"]], ["ACS"])

    def test_allocation_plan_can_keep_cash_when_no_positive_candidates(self):
        losing_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="TEF",
                company_name="Telefonica",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("4"),
            ),
            "reference_label": "IBEX 35",
            "projection": {
                "available": True,
                "base_return_pct": Decimal("-6.0"),
                "projected_price": Decimal("3.76"),
                "confidence_label": "Media",
                "coefficient": Decimal("0.22"),
            },
        }

        plan = build_equity_allocation_plan([losing_card], Decimal("50000"), Decimal("25"))

        self.assertFalse(plan["available"])
        self.assertIn("ninguna accion", plan["reason"].lower())

    def test_optimizer_soft_penalty_can_exclude_fragile_candidate_without_hard_gate(self):
        fragile_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="SLR",
                company_name="Solaria",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("8"),
            ),
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Vigilar", "tone": "watch", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("21.00"),
                "price_return_pct": Decimal("18.00"),
                "price_low_return_pct": Decimal("-10.00"),
                "price_high_return_pct": Decimal("34.00"),
                "projected_price": Decimal("9.6800"),
                "confidence_label": "Alta",
                "safety_score": Decimal("18.00"),
                "gross_dividend_yield_pct": Decimal("0.00"),
                "net_income_yield_pct": Decimal("0.00"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("66.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Transicion",
                "current_drawdown_pct": Decimal("-4.00"),
                "max_drawdown_pct": Decimal("-18.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("25.0"), "total_return_pct": Decimal("-10.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("45.0"), "total_return_pct": Decimal("21.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("30.0"), "total_return_pct": Decimal("34.00")},
                ],
            },
            "projection_reliability": {"label": "Baja", "score": Decimal("42.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("8.50"),
                "five_year_return_pct": Decimal("50.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("25.0"), "annual_return_pct": Decimal("1.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("45.0"), "annual_return_pct": Decimal("8.50")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("30.0"), "annual_return_pct": Decimal("13.00")},
                ],
            },
        }
        robust_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IBE",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("14.00"),
                "price_return_pct": Decimal("11.50"),
                "price_low_return_pct": Decimal("2.00"),
                "price_high_return_pct": Decimal("20.00"),
                "projected_price": Decimal("15.6100"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.40"),
                "net_income_yield_pct": Decimal("2.60"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("11.00"),
                "positive_year_ratio_pct": Decimal("72.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-16.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("2.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "total_return_pct": Decimal("14.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("20.00")},
                ],
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.20"),
                "five_year_return_pct": Decimal("35.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("4.50")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "annual_return_pct": Decimal("6.20")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("7.60")},
                ],
            },
        }

        fragile_candidate = build_equity_optimizer_candidate(fragile_card)
        robust_candidate = build_equity_optimizer_candidate(robust_card)
        filtered = filter_positive_optimizer_candidates([fragile_candidate, robust_candidate], "12m_primary")
        plan = build_equity_allocation_plan([fragile_card, robust_card], Decimal("100000"), Decimal("60"))

        self.assertIsNotNone(fragile_candidate)
        self.assertGreater(fragile_candidate["quality_floor_penalty_pct"], ZERO)
        self.assertEqual(fragile_candidate["decision_action_label"], "Vigilar")
        self.assertLess(fragile_candidate["decision_action_adjustment"], ZERO)
        self.assertEqual([item["position"].ticker for item in filtered], ["IBE"])
        self.assertTrue(plan["available"])
        self.assertEqual([item["position"].ticker for item in plan["allocations"]], ["IBE"])

    def test_optimizer_soft_penalty_still_keeps_decent_candidate_below_reference_levels(self):
        middling_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="AMS",
                company_name="Amadeus",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("60"),
            ),
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("14.00"),
                "price_return_pct": Decimal("9.50"),
                "price_low_return_pct": Decimal("-2.00"),
                "price_high_return_pct": Decimal("20.00"),
                "projected_price": Decimal("68.4000"),
                "confidence_label": "Media",
                "safety_score": Decimal("54.00"),
                "gross_dividend_yield_pct": Decimal("1.00"),
                "net_income_yield_pct": Decimal("0.80"),
                "transaction_drag_pct": Decimal("0.25"),
                "annualized_volatility_pct": Decimal("13.00"),
                "positive_year_ratio_pct": Decimal("66.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-5.00"),
                "max_drawdown_pct": Decimal("-21.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("25.0"), "total_return_pct": Decimal("-2.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("50.0"), "total_return_pct": Decimal("14.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("25.0"), "total_return_pct": Decimal("20.00")},
                ],
            },
            "projection_reliability": {"label": "Media", "score": Decimal("57.00")},
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.20"),
                "five_year_return_pct": Decimal("35.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("25.0"), "annual_return_pct": Decimal("2.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("50.0"), "annual_return_pct": Decimal("6.20")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("25.0"), "annual_return_pct": Decimal("8.20")},
                ],
            },
        }

        candidate = build_equity_optimizer_candidate(middling_card)
        filtered = filter_positive_optimizer_candidates([candidate], "12m_primary")

        self.assertIsNotNone(candidate)
        self.assertGreater(candidate["quality_floor_penalty_pct"], ZERO)
        self.assertEqual([item["position"].ticker for item in filtered], ["AMS"])

    def test_allocation_plan_recalculates_costs_and_dividends_for_assigned_capital(self):
        anchor_day = date(2026, 4, 25)
        santander_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Banco Santander",
                trade_channel=EquityPosition.TradeChannel.OFFICE,
                ticker="IBE",
                quote_symbol="IBE.MC",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
                benchmark_name="Demanda electrica Espana",
                benchmark_symbol="REE:demand:es:peninsular",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("11.00"),
                annual_maintenance_cost=Decimal("0"),
                latest_price_date=anchor_day,
            ),
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia defensiva."},
            "reference_label": "Demanda electrica Espana",
            "projection_reliability": {
                "label": "Alta",
                "score": Decimal("80.00"),
            },
            "projection": {
                "available": True,
                "base_return_pct": Decimal("8.70"),
                "price_return_pct": Decimal("6.50"),
                "price_low_return_pct": Decimal("-6.00"),
                "price_high_return_pct": Decimal("16.00"),
                "projected_price": Decimal("11.7150"),
                "latest_price": Decimal("11.0000"),
                "latest_date": anchor_day,
                "confidence_label": "Alta",
                "safety_score": Decimal("74.00"),
                "gross_dividend_yield_pct": Decimal("4.00"),
                "net_income_yield_pct": Decimal("3.70"),
                "transaction_drag_pct": Decimal("1.50"),
                "annualized_volatility_pct": Decimal("14.00"),
                "positive_year_ratio_pct": Decimal("72.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-4.00"),
                "max_drawdown_pct": Decimal("-22.00"),
                "monthly_path": [
                    {"label": "1M", "projected_date": anchor_day + timedelta(days=30), "projected_price": Decimal("10.6000")},
                    {"label": "3M", "projected_date": anchor_day + timedelta(days=90), "projected_price": Decimal("9.5000")},
                    {"label": "6M", "projected_date": anchor_day + timedelta(days=180), "projected_price": Decimal("10.7000")},
                    {"label": "9M", "projected_date": anchor_day + timedelta(days=270), "projected_price": Decimal("11.3000")},
                    {"label": "12M", "projected_date": anchor_day + timedelta(days=365), "projected_price": Decimal("11.7150")},
                ],
            },
        }

        plan = build_equity_allocation_plan([santander_card], Decimal("10000"), Decimal("100"))

        self.assertTrue(plan["available"])
        self.assertEqual(plan["roundtrip_cost_total"], Decimal("150.00"))
        self.assertEqual(plan["annual_cost_total"], Decimal("27.00"))
        self.assertEqual(plan["net_dividend_income_total"], Decimal("398.00"))
        allocation = plan["allocations"][0]
        self.assertEqual(allocation["roundtrip_total_cost"], Decimal("150.00"))
        self.assertEqual(allocation["annual_cost_used"], Decimal("27.00"))
        self.assertEqual(allocation["expected_net_dividend_income"], Decimal("398.00"))
        self.assertEqual(allocation["net_projected_return_pct"], Decimal("8.71"))
        self.assertEqual(allocation["low_return_pct"], Decimal("-3.79"))
        self.assertLess(allocation["net_projected_return_pct"], Decimal("10.50"))

    def test_optimizer_uses_visible_12m_path_when_raw_projection_and_chart_diverge(self):
        latest_date = timezone.localdate()
        position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Interactive Brokers",
            trade_channel=EquityPosition.TradeChannel.APP,
            ticker="ANA",
            quote_symbol="ANA.MC",
            company_name="Acciona",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_name="IBEX 35",
            benchmark_symbol="^IBEX",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("0"),
            current_price_per_share=Decimal("100.0000"),
            annual_maintenance_cost=Decimal("0"),
        )
        card = {
            "position": position,
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Construccion",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "score": Decimal("4.20"), "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("40.00"),
                "price_return_pct": Decimal("40.00"),
                "low_return_pct": Decimal("8.00"),
                "high_return_pct": Decimal("62.00"),
                "projected_price": Decimal("140.0000"),
                "latest_price": Decimal("100.0000"),
                "latest_date": latest_date,
                "confidence_label": "Alta",
                "safety_score": Decimal("72.00"),
                "gross_dividend_yield_pct": Decimal("0.00"),
                "net_income_yield_pct": Decimal("0.00"),
                "transaction_drag_pct": Decimal("0.00"),
                "annualized_volatility_pct": Decimal("15.00"),
                "positive_year_ratio_pct": Decimal("62.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-3.00"),
                "max_drawdown_pct": Decimal("-25.00"),
                "monthly_path": [
                    {"label": "3M", "projected_date": add_calendar_months(latest_date, 3), "projected_price": Decimal("94.0000")},
                    {"label": "6M", "projected_date": add_calendar_months(latest_date, 6), "projected_price": Decimal("90.0000")},
                    {"label": "9M", "projected_date": add_calendar_months(latest_date, 9), "projected_price": Decimal("85.0000")},
                    {"label": "12M", "projected_date": add_calendar_months(latest_date, 12), "projected_price": Decimal("80.0000")},
                ],
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("33.0"), "total_return_pct": Decimal("15.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("34.0"), "total_return_pct": Decimal("40.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("33.0"), "total_return_pct": Decimal("55.00")},
                ],
            },
            "cycle_projection_5y": {
                "available": False,
                "path": [],
                "scenarios": [],
            },
        }

        refresh_card_projection_visuals(card, history=[])
        candidate = build_equity_optimizer_candidate(card)
        plan = build_equity_allocation_plan([card], Decimal("10000"), Decimal("100"))

        self.assertIsNotNone(candidate)
        self.assertEqual(candidate["base_return_pct"], card["presentation_projection"]["visible_total_return_pct"])
        self.assertLess(candidate["base_return_pct"], ZERO)
        self.assertLess(candidate["scenario_expected_return_pct"], ZERO)
        self.assertFalse(plan["available"])
        self.assertIn("retorno robusto positivo", plan["reason"])

    def test_optimizer_candidate_cools_current_expectation_with_negative_expectation_history(self):
        review_dates = [date(2026, 3, 4), date(2026, 3, 11), date(2026, 3, 18)]
        for index, review_date in enumerate(review_dates, start=1):
            run = EquityNightlyAnalysisRun.objects.create(
                analysis_date=review_date,
                status=EquityNightlyAnalysisRun.Status.COMPLETED,
            )
            EquityExpectationReview.objects.create(
                run=run,
                analysis_date=review_date,
                review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
                scope=EquityExpectationReview.Scope.IBEX,
                analysis_key=f"ibe-{index}",
                ticker="IBE",
                quote_symbol="IBE.MC",
                company_name="Iberdrola",
                expected_return_pct_1y=Decimal(str((24, 12, 4)[index - 1])),
                expected_return_pct_5y=Decimal(str((48, 34, 20)[index - 1])),
            )

        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IBE",
                quote_symbol="IBE.MC",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("16.00"),
                "price_return_pct": Decimal("13.00"),
                "price_low_return_pct": Decimal("2.00"),
                "price_high_return_pct": Decimal("24.00"),
                "projected_price": Decimal("16.2400"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.40"),
                "net_income_yield_pct": Decimal("2.60"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("11.00"),
                "positive_year_ratio_pct": Decimal("72.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-16.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("2.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "total_return_pct": Decimal("16.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("22.00")},
                ],
            },
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.20"),
                "five_year_return_pct": Decimal("35.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("4.50")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "annual_return_pct": Decimal("6.20")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("7.60")},
                ],
            },
        }

        candidate = build_equity_optimizer_candidate(card)

        self.assertIsNotNone(candidate)
        self.assertTrue(candidate["expectation_review_signal"]["available"])
        self.assertEqual(candidate["expectation_review_signal"]["1y"]["sample_count"], 3)
        self.assertLess(candidate["scenario_expectation_review"]["adjustment_pct"], ZERO)
        self.assertLess(candidate["scenario_expected_return_pct"], candidate["raw_scenario_expected_return_pct"])

    def test_allocation_plan_uses_expectation_history_to_demote_deteriorating_candidate(self):
        review_dates = [date(2026, 3, 4), date(2026, 3, 11), date(2026, 3, 18)]
        ana_values = [(28, 62), (16, 44), (4, 22)]
        ibe_values = [(14, 38), (15, 40), (16, 42)]
        for index, review_date in enumerate(review_dates, start=1):
            run = EquityNightlyAnalysisRun.objects.create(
                analysis_date=review_date,
                status=EquityNightlyAnalysisRun.Status.COMPLETED,
            )
            EquityExpectationReview.objects.create(
                run=run,
                analysis_date=review_date,
                review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
                scope=EquityExpectationReview.Scope.IBEX,
                analysis_key=f"ana-{index}",
                ticker="ANA",
                quote_symbol="ANA.MC",
                company_name="Acciona",
                expected_return_pct_1y=Decimal(str(ana_values[index - 1][0])),
                expected_return_pct_5y=Decimal(str(ana_values[index - 1][1])),
            )
            EquityExpectationReview.objects.create(
                run=run,
                analysis_date=review_date,
                review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
                scope=EquityExpectationReview.Scope.IBEX,
                analysis_key=f"ibe-{index}",
                ticker="IBE",
                quote_symbol="IBE.MC",
                company_name="Iberdrola",
                expected_return_pct_1y=Decimal(str(ibe_values[index - 1][0])),
                expected_return_pct_5y=Decimal(str(ibe_values[index - 1][1])),
            )

        deteriorating_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="ANA",
                quote_symbol="ANA.MC",
                company_name="Acciona",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("100"),
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Construccion",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("18.00"),
                "price_return_pct": Decimal("16.50"),
                "price_low_return_pct": Decimal("1.00"),
                "price_high_return_pct": Decimal("28.00"),
                "projected_price": Decimal("118.0000"),
                "confidence_label": "Alta",
                "safety_score": Decimal("74.00"),
                "gross_dividend_yield_pct": Decimal("0.80"),
                "net_income_yield_pct": Decimal("0.50"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("14.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-4.00"),
                "max_drawdown_pct": Decimal("-20.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("1.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "total_return_pct": Decimal("18.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("24.00")},
                ],
            },
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("7.00"),
                "five_year_return_pct": Decimal("40.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("4.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "annual_return_pct": Decimal("7.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("9.00")},
                ],
            },
        }
        stable_card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                ticker="IBE",
                quote_symbol="IBE.MC",
                company_name="Iberdrola",
                reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                benchmark_name="IBEX 35",
                benchmark_symbol="^IBEX",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("14"),
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Electrica",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("16.00"),
                "price_return_pct": Decimal("13.00"),
                "price_low_return_pct": Decimal("3.00"),
                "price_high_return_pct": Decimal("21.00"),
                "projected_price": Decimal("16.2400"),
                "confidence_label": "Alta",
                "safety_score": Decimal("77.00"),
                "gross_dividend_yield_pct": Decimal("3.40"),
                "net_income_yield_pct": Decimal("2.60"),
                "transaction_drag_pct": Decimal("0.20"),
                "annualized_volatility_pct": Decimal("11.00"),
                "positive_year_ratio_pct": Decimal("72.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-16.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("3.00")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "total_return_pct": Decimal("16.00")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "total_return_pct": Decimal("21.00")},
                ],
            },
            "cycle_projection_5y": {
                "available": True,
                "annual_return_pct": Decimal("6.20"),
                "five_year_return_pct": Decimal("35.00"),
                "scenarios": [
                    {"key": "bear", "label": "Bajista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("4.50")},
                    {"key": "base", "label": "Base", "probability_pct": Decimal("52.0"), "annual_return_pct": Decimal("6.20")},
                    {"key": "bull", "label": "Alcista", "probability_pct": Decimal("24.0"), "annual_return_pct": Decimal("7.60")},
                ],
            },
        }

        plan = build_equity_allocation_plan([deteriorating_card, stable_card], Decimal("100000"), Decimal("60"))

        self.assertTrue(plan["available"])
        self.assertEqual(plan["allocations"][0]["position"].ticker, "IBE")
        self.assertLess(
            plan["allocations"][1]["scenario_expected_return_pct"],
            plan["allocations"][1]["raw_scenario_expected_return_pct"],
        )

    def test_allocation_plan_discards_small_tickets_when_fixed_costs_are_too_high(self):
        base_projection = {
            "available": True,
            "base_return_pct": Decimal("9.50"),
            "price_return_pct": Decimal("7.40"),
            "price_low_return_pct": Decimal("-4.50"),
            "price_high_return_pct": Decimal("15.20"),
            "projected_price": Decimal("11.7400"),
            "confidence_label": "Alta",
            "safety_score": Decimal("72.00"),
            "gross_dividend_yield_pct": Decimal("3.80"),
            "net_income_yield_pct": Decimal("3.20"),
            "transaction_drag_pct": Decimal("1.50"),
            "annualized_volatility_pct": Decimal("14.00"),
            "positive_year_ratio_pct": Decimal("68.00"),
            "years_covered": Decimal("10.00"),
            "cycle_phase": "Expansion",
            "current_drawdown_pct": Decimal("-3.50"),
            "max_drawdown_pct": Decimal("-18.00"),
        }
        cards = []
        for ticker in ("IBE", "ELE", "ENG"):
            cards.append(
                {
                    "position": EquityPosition(
                        position_kind=EquityPosition.PositionKind.WATCHLIST,
                        broker="Banco Santander",
                        trade_channel=EquityPosition.TradeChannel.OFFICE,
                        ticker=ticker,
                        quote_symbol=f"{ticker}.MC",
                        company_name=ticker,
                        reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
                        benchmark_name="IBEX 35",
                        benchmark_symbol="^IBEX",
                        shares=Decimal("0"),
                        average_cost_per_share=Decimal("0"),
                        current_price_per_share=Decimal("11.00"),
                        annual_maintenance_cost=Decimal("0"),
                    ),
                    "status_key": "ibex",
                    "status_label": "Radar IBEX",
                    "reference_label": "IBEX 35",
                    "trade_alert": {"label": "Comprar", "tone": "buy", "note": ""},
                    "projection_reliability": {
                        "label": "Alta",
                        "score": Decimal("80.00"),
                    },
                    "projection": dict(base_projection),
                }
            )

        plan = build_equity_allocation_plan(cards, Decimal("3000"), Decimal("50"))

        self.assertTrue(plan["available"])
        self.assertEqual(len(plan["allocations"]), 2)
        self.assertEqual(plan["ticket_filtered_count"], 1)
        self.assertIn("coste fijo+variable", plan["ticket_filter_note"].lower())
        self.assertTrue(all(item["purchase_total_cost"] == Decimal("13.00") for item in plan["allocations"]))
        self.assertTrue(all(item["allocated_amount"] == Decimal("1500.00") for item in plan["allocations"]))

    def test_unknown_broker_uses_santander_fallback_costs(self):
        costs = estimate_broker_costs(
            broker_name="Broker Desconocido",
            trade_channel="app",
            trade_amount=Decimal("10000"),
            valuation_amount=Decimal("10000"),
            annual_dividend_income=Decimal("400"),
            quote_symbol="IBE.MC",
        )

        self.assertEqual(costs["profile_key"], "santander_fallback")
        self.assertEqual(costs["purchase_total_cost"], Decimal("26.00"))
        self.assertEqual(costs["sale_total_cost"], Decimal("6.00"))
        self.assertEqual(costs["annual_custody_cost"], Decimal("25.00"))
        self.assertIn("Santander", costs["pdf_source_label"])

    @override_settings(EQUITIES_IBEX_UNIVERSE_LIMIT=1)
    def test_process_equity_optimization_run_persists_summary_and_report(self):
        analysis_day = date(2026, 4, 25)
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-TEST-001",
            label="Cartera de prueba",
            total_investment=Decimal("100000"),
            max_company_pct=Decimal("20"),
            max_total_positions=8,
            max_sector_positions=1,
            selected_sectors=["Energia"],
            selected_owned_tickers=["IBE"],
        )
        position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("11.0000"),
            current_price_per_share=Decimal("11.0000"),
        )
        card = {
            "position": position,
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Energia",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("8.50"),
                "price_return_pct": Decimal("6.20"),
                "price_low_return_pct": Decimal("-4.00"),
                "price_high_return_pct": Decimal("13.00"),
                "projected_price": Decimal("11.6800"),
                "confidence_label": "Alta",
                "safety_score": Decimal("74.00"),
                "gross_dividend_yield_pct": Decimal("4.10"),
                "net_income_yield_pct": Decimal("3.30"),
                "transaction_drag_pct": Decimal("0.90"),
                "annualized_volatility_pct": Decimal("13.00"),
                "positive_year_ratio_pct": Decimal("68.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-3.00"),
                "max_drawdown_pct": Decimal("-20.00"),
                "monthly_path": [
                    {"label": "1M", "projected_date": analysis_day + timedelta(days=30), "projected_price": Decimal("10.8200")},
                    {"label": "3M", "projected_date": analysis_day + timedelta(days=90), "projected_price": Decimal("10.5000")},
                    {"label": "6M", "projected_date": analysis_day + timedelta(days=180), "projected_price": Decimal("11.2400")},
                    {"label": "9M", "projected_date": analysis_day + timedelta(days=270), "projected_price": Decimal("11.5600")},
                    {"label": "12M", "projected_date": analysis_day + timedelta(days=365), "projected_price": Decimal("11.6800")},
                ],
            },
        }
        dashboard = {
            "optimizer_cards": [card],
            "history_cards": [],
            "ibex_universe_summary": {"analyzed_count": 35},
        }

        with (
            patch("equities.optimization_runs.sync_all_equities_market_data", return_value=[]),
            patch("equities.optimization_runs.build_equity_analysis_dashboard", return_value=dashboard) as mocked_dashboard,
            patch("equities.optimization_runs.build_news_signal_map", return_value={"IBE": {"label": "Prensa favorable", "score": Decimal("2.10"), "items_count": 2, "items": [], "note": "Buen tono", "available": True, "positive_count": 2, "negative_count": 0, "neutral_count": 0}}),
            patch("equities.optimization_runs.build_report_entries", return_value=[]),
            patch("equities.optimization_runs.build_report_html", return_value="<html>informe</html>"),
            patch("equities.optimization_runs.build_report_pdf_html", return_value="<html>pdf</html>"),
        ):
            process_equity_optimization_run(run.id)

        run.refresh_from_db()
        self.assertIsNone(mocked_dashboard.call_args.kwargs["ibex_company_limit"])
        self.assertEqual(run.status, EquityOptimizationRun.Status.COMPLETED)
        self.assertEqual(run.report_html, "<html>informe</html>")
        self.assertEqual(run.report_pdf_html, "<html>pdf</html>")
        self.assertTrue(run.summary_data["available"])
        self.assertEqual(run.summary_data["max_total_positions"], 8)
        self.assertEqual(run.summary_data["selected_sectors"], ["Energia"])
        self.assertEqual(run.summary_data["selected_owned_tickers"], ["IBE"])
        self.assertEqual(run.summary_data["top_pick_name"], "Iberdrola")
        self.assertTrue(run.summary_data["top_pick_buy_window_label"])
        self.assertIsNotNone(run.summary_data["top_pick_buy_price"])
        self.assertTrue(run.summary_data["top_pick_exit_window_label"])
        self.assertIsNotNone(run.summary_data["top_pick_exit_price"])
        self.assertIsNotNone(run.summary_data["top_pick_interval_return_pct"])
        self.assertIsNotNone(run.summary_data["top_pick_holding_annualized_return_pct"])
        self.assertIsNone(run.summary_data["weighted_cycle_return_annual_pct"])
        self.assertTrue(run.allocations_data[0]["purchase_timing"]["available"])
        self.assertTrue(run.allocations_data[0]["purchase_timing"]["buy_window_label"])
        self.assertIsNotNone(run.allocations_data[0]["purchase_timing"]["buy_price"])
        self.assertTrue(run.allocations_data[0]["purchase_timing"]["exit_window_label"])
        self.assertIsNotNone(run.allocations_data[0]["purchase_timing"]["exit_price"])
        self.assertIsNotNone(run.allocations_data[0]["purchase_timing"]["interval_return_pct"])
        self.assertIsNotNone(run.allocations_data[0]["purchase_timing"]["holding_annualized_return_pct"])
        self.assertEqual(run.progress_data["percent"], 100)
        self.assertEqual(run.progress_data["note"], "Optimizacion completada")
        self.assertEqual(run.progress_data["stage_key"], "report")
        self.assertTrue(run.progress_data["preview_candidates"])
        self.assertTrue(run.progress_data["preview_allocations"])
        self.assertTrue(all(stage["status"] == "completed" for stage in run.progress_data["stages"]))

    @override_settings(EQUITIES_OPTIMIZATION_ASYNC=False)
    def test_launch_optimization_run_still_queues_background_job_when_async_flag_is_off(self):
        with (
            patch("equities.optimization_runs.enqueue_equity_optimization_run") as mocked_enqueue,
            patch("equities.optimization_runs.process_equity_optimization_run") as mocked_process,
        ):
            run = launch_equity_optimization_run(
                total_investment=Decimal("80000"),
                max_company_pct=Decimal("20"),
                max_total_positions=6,
                max_sector_positions=1,
                selected_sectors=["Banca"],
                selected_owned_tickers=["IBE"],
            )

        self.assertEqual(run.status, EquityOptimizationRun.Status.PENDING)
        mocked_enqueue.assert_called_once_with(run.id)
        mocked_process.assert_not_called()

    @override_settings(EQUITIES_OPTIMIZATION_ASYNC=False)
    def test_launch_optimization_run_pair_creates_two_strategy_runs(self):
        with (
            patch("equities.optimization_runs.enqueue_equity_optimization_run") as mocked_enqueue,
            patch("equities.optimization_runs.process_equity_optimization_run") as mocked_process,
        ):
            runs = launch_equity_optimization_run_pair(
                total_investment=Decimal("80000"),
                max_company_pct=Decimal("20"),
                max_total_positions=6,
                max_sector_positions=1,
                selected_sectors=["Banca"],
                selected_owned_tickers=["IBE"],
                reference_label="Cartera dual",
            )

        self.assertEqual(len(runs), 2)
        self.assertEqual({run.progress_data["strategy_mode"] for run in runs}, {"12m_primary", "5y_primary"})
        self.assertEqual({run.progress_data["strategy_label"] for run in runs}, {"12M principal", "5A principal"})
        self.assertTrue(any(run.reference_code.endswith("-12M") for run in runs))
        self.assertTrue(any(run.reference_code.endswith("-5A") for run in runs))
        self.assertEqual(mocked_enqueue.call_count, 2)
        mocked_process.assert_not_called()

    @override_settings(
        EQUITIES_OPTIMIZATION_ASYNC=False,
        EQUITIES_SCHEDULED_OPTIMIZATION_ENABLED=True,
        EQUITIES_SCHEDULED_OPTIMIZATION_ISO_WEEKDAYS=(2, 4),
    )
    def test_launch_scheduled_optimization_runs_creates_pair_once_per_day(self):
        analysis_day = date(2026, 4, 21)

        with (
            patch("equities.optimization_runs.enqueue_equity_optimization_run") as mocked_enqueue,
            patch("equities.optimization_runs.process_equity_optimization_run") as mocked_process,
        ):
            first_runs = launch_scheduled_equity_optimization_runs(
                analysis_date=analysis_day,
                force=False,
            )
            second_runs = launch_scheduled_equity_optimization_runs(
                analysis_date=analysis_day,
                force=False,
            )

        self.assertEqual(len(first_runs), 2)
        self.assertEqual(len(second_runs), 2)
        self.assertEqual(EquityOptimizationRun.objects.count(), 2)
        self.assertEqual(mocked_enqueue.call_count, 2)
        mocked_process.assert_not_called()
        self.assertEqual(
            {run.progress_data["scheduled_run_key"] for run in first_runs},
            {"scheduled-optimization:2026-04-21"},
        )
        self.assertEqual(
            {run.progress_data["schedule_kind"] for run in first_runs},
            {"nightly"},
        )
        self.assertEqual(
            {run.progress_data["scheduled_weekdays_label"] for run in first_runs},
            {"martes y jueves"},
        )
        self.assertEqual({Decimal(str(run.max_company_pct)) for run in first_runs}, {Decimal("30.00")})
        self.assertEqual({run.max_total_positions for run in first_runs}, {5})
        self.assertEqual({run.max_sector_positions for run in first_runs}, {2})
        self.assertEqual({run.id for run in first_runs}, {run.id for run in second_runs})

    @override_settings(
        EQUITIES_OPTIMIZATION_ASYNC=False,
        EQUITIES_SCHEDULED_OPTIMIZATION_ENABLED=True,
        EQUITIES_SCHEDULED_OPTIMIZATION_ISO_WEEKDAYS=(2, 4),
    )
    def test_launch_scheduled_optimization_runs_replaces_failed_attempts(self):
        analysis_day = date(2026, 4, 21)
        failed_runs = [
            EquityOptimizationRun.objects.create(
                reference_code=f"OPT-FAILED-{suffix}",
                label=f"Fallida {suffix}",
                total_investment=Decimal("100000"),
                max_company_pct=Decimal("30"),
                max_total_positions=5,
                max_sector_positions=2,
                status=EquityOptimizationRun.Status.FAILED,
                progress_data={
                    "strategy_mode": strategy_mode,
                    "strategy_label": strategy_label,
                    "schedule_kind": "nightly",
                    "scheduled_run_key": "scheduled-optimization:2026-04-21",
                    "scheduled_analysis_date": "2026-04-21",
                    "scheduled_weekdays_label": "martes y jueves",
                },
            )
            for suffix, strategy_mode, strategy_label in (
                ("12M", "12m_primary", "12M principal"),
                ("5A", "5y_primary", "5A principal"),
            )
        ]

        with (
            patch("equities.optimization_runs.enqueue_equity_optimization_run") as mocked_enqueue,
            patch("equities.optimization_runs.process_equity_optimization_run") as mocked_process,
        ):
            runs = launch_scheduled_equity_optimization_runs(
                analysis_date=analysis_day,
                force=False,
            )

        self.assertEqual(len(runs), 2)
        self.assertEqual(EquityOptimizationRun.objects.count(), 2)
        self.assertFalse(EquityOptimizationRun.objects.filter(id__in=[run.id for run in failed_runs]).exists())
        self.assertEqual(mocked_enqueue.call_count, 2)
        mocked_process.assert_not_called()

    def test_purge_stale_scheduled_optimizations_removes_old_and_wrong_policy_runs(self):
        recent_valid = EquityOptimizationRun.objects.create(
            reference_code="OPT-RECENT-VALID",
            label="Reciente valida",
            total_investment=Decimal("100000"),
            max_company_pct=Decimal("30"),
            max_total_positions=5,
            max_sector_positions=2,
            status=EquityOptimizationRun.Status.COMPLETED,
            progress_data={
                "schedule_kind": "nightly",
                "scheduled_run_key": "scheduled-optimization:2026-04-16",
                "scheduled_analysis_date": "2026-04-16",
            },
        )
        old_run = EquityOptimizationRun.objects.create(
            reference_code="OPT-OLD-RUN",
            label="Antigua",
            total_investment=Decimal("100000"),
            max_company_pct=Decimal("30"),
            max_total_positions=5,
            max_sector_positions=2,
            status=EquityOptimizationRun.Status.COMPLETED,
            progress_data={
                "schedule_kind": "nightly",
                "scheduled_run_key": "scheduled-optimization:2025-12-01",
                "scheduled_analysis_date": "2025-12-01",
            },
        )
        wrong_policy_run = EquityOptimizationRun.objects.create(
            reference_code="OPT-WRONG-POLICY",
            label="Politica vieja",
            total_investment=Decimal("100000"),
            max_company_pct=Decimal("20"),
            max_total_positions=0,
            max_sector_positions=0,
            status=EquityOptimizationRun.Status.COMPLETED,
            progress_data={
                "schedule_kind": "nightly",
                "scheduled_run_key": "scheduled-optimization:2026-04-10",
                "scheduled_analysis_date": "2026-04-10",
            },
        )

        deleted_count = purge_stale_scheduled_optimization_runs(as_of=date(2026, 4, 17))

        self.assertGreaterEqual(deleted_count, 2)
        self.assertTrue(EquityOptimizationRun.objects.filter(pk=recent_valid.pk).exists())
        self.assertFalse(EquityOptimizationRun.objects.filter(pk=old_run.pk).exists())
        self.assertFalse(EquityOptimizationRun.objects.filter(pk=wrong_policy_run.pk).exists())

    @override_settings(
        EQUITIES_SCHEDULED_OPTIMIZATION_ENABLED=True,
        EQUITIES_SCHEDULED_OPTIMIZATION_ISO_WEEKDAYS=(2, 4),
    )
    def test_build_scheduled_optimization_persistence_context_aggregates_recent_repetition(self):
        today = timezone.localdate()
        scheduled_days = [today - timedelta(days=3), today - timedelta(days=10), today - timedelta(days=38), today - timedelta(days=110)]
        strategy_runs = [
            ("12M principal", "12m_primary"),
            ("5A principal", "5y_primary"),
        ]
        EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("11.0000"),
        )

        for index, analysis_day in enumerate(scheduled_days, start=1):
            for strategy_label, strategy_mode in strategy_runs:
                run = EquityOptimizationRun.objects.create(
                    reference_code=f"OPT-SCHED-{index}-{strategy_mode}",
                    label=f"Programada {analysis_day.isoformat()} - {strategy_label}",
                    total_investment=Decimal("100000"),
                    max_company_pct=Decimal("30"),
                    max_total_positions=5,
                    max_sector_positions=2,
                    status=EquityOptimizationRun.Status.COMPLETED,
                    progress_data={
                        "strategy_label": strategy_label,
                        "strategy_mode": strategy_mode,
                        "schedule_kind": "nightly",
                        "scheduled_run_key": f"scheduled-optimization:{analysis_day.isoformat()}",
                        "scheduled_analysis_date": analysis_day.isoformat(),
                        "scheduled_weekdays_label": "martes y jueves",
                    },
                    summary_data={
                        "available": True,
                        "strategy_label": strategy_label,
                        "scheduled_analysis_date": analysis_day.isoformat(),
                    },
                    allocations_data=[
                        {
                            "rank": 1 if strategy_mode == "12m_primary" else 2,
                            "company_name": "Iberdrola",
                            "ticker": "IBE",
                            "net_projected_return_pct": 12.0 if strategy_mode == "12m_primary" else 10.0,
                            "cycle_return_5y_pct": 74.0 if strategy_mode == "12m_primary" else 70.0,
                            "allocated_amount": 25000.0,
                            "allocated_weight_pct": 25.0,
                            "reliability_label": "Alta",
                            "reliability_score": 82.0 if strategy_mode == "12m_primary" else 78.0,
                            "purchase_timing": {
                                "available": True,
                                "buy_date": (analysis_day + timedelta(days=30)).isoformat(),
                                "buy_window_label": f"mayo {analysis_day.year}",
                                "buy_price": 10.45,
                                "exit_date": (analysis_day + timedelta(days=365)).isoformat(),
                                "exit_window_label": f"abril {analysis_day.year + 1}",
                                "exit_price": 11.68,
                                "interval_window_label": f"mayo {analysis_day.year} -> abril {analysis_day.year + 1}",
                                "interval_return_pct": 11.6,
                                "holding_months": 11,
                                "mode_label": "Esperar correccion",
                            },
                            "cycle_yearly_margins": [
                                {"year_number": 1, "label": "AÑO 1", "margin_pct": 12.0 if strategy_mode == "12m_primary" else 10.0},
                                {"year_number": 2, "label": "AÑO 2", "margin_pct": 7.0},
                                {"year_number": 3, "label": "AÑO 3", "margin_pct": 8.0},
                                {"year_number": 4, "label": "AÑO 4", "margin_pct": 9.0},
                                {"year_number": 5, "label": "AÑO 5", "margin_pct": 10.0},
                            ],
                        },
                        {
                            "rank": 4,
                            "company_name": "Repsol",
                            "ticker": "REP",
                            "net_projected_return_pct": 6.0,
                            "cycle_return_5y_pct": 28.0,
                            "allocated_amount": 18000.0,
                            "allocated_weight_pct": 18.0,
                            "reliability_label": "Media",
                            "reliability_score": 64.0,
                            "purchase_timing": {
                                "available": True,
                                "buy_date": (analysis_day + timedelta(days=18)).isoformat(),
                                "buy_window_label": f"mayo {analysis_day.year}",
                                "buy_price": 13.10,
                                "exit_date": (analysis_day + timedelta(days=300)).isoformat(),
                                "exit_window_label": f"febrero {analysis_day.year + 1}",
                                "exit_price": 14.25,
                                "interval_window_label": f"mayo {analysis_day.year} -> febrero {analysis_day.year + 1}",
                                "interval_return_pct": 7.4,
                                "holding_months": 9,
                                "mode_label": "Comprar ya",
                            },
                            "cycle_yearly_margins": [
                                {"year_number": 1, "label": "AÑO 1", "margin_pct": 6.0},
                                {"year_number": 2, "label": "AÑO 2", "margin_pct": 4.0},
                            ],
                        },
                        {
                            "rank": 5,
                            "company_name": "ACS",
                            "ticker": "ACS",
                            "net_projected_return_pct": 8.0,
                            "cycle_return_5y_pct": 40.0,
                            "allocated_amount": 22000.0,
                            "allocated_weight_pct": 22.0,
                            "reliability_label": "Media",
                            "reliability_score": 60.0,
                            "purchase_timing": {
                                "available": True,
                                "buy_date": (analysis_day + timedelta(days=12)).isoformat(),
                                "buy_window_label": f"mayo {analysis_day.year}",
                                "buy_price": 39.25,
                                "exit_date": (analysis_day + timedelta(days=210)).isoformat(),
                                "exit_window_label": f"noviembre {analysis_day.year}",
                                "exit_price": 43.50,
                                "interval_window_label": f"mayo {analysis_day.year} -> noviembre {analysis_day.year}",
                                "interval_return_pct": 10.8,
                                "holding_months": 6,
                                "mode_label": "Comprar ya",
                            },
                            "cycle_yearly_margins": [
                                {"year_number": 1, "label": "AÃ‘O 1", "margin_pct": 8.0},
                                {"year_number": 2, "label": "AÃ‘O 2", "margin_pct": 5.0},
                            ],
                        },
                    ],
                )
                EquityOptimizationRun.objects.filter(pk=run.pk).update(
                    created_at=timezone.make_aware(datetime.combine(analysis_day, datetime.min.time())),
                    completed_at=timezone.make_aware(datetime.combine(analysis_day, datetime.min.time())),
                )

        context = build_scheduled_optimization_persistence_context(
            as_of=today,
            live_quote_map={
                "ACS": {
                    "ticker": "ACS",
                    "company_name": "ACS, Actividades de Construccion y Servicios, S.A.",
                    "current_price": Decimal("40.1000"),
                    "current_price_date": today,
                    "current_price_date_label": today.isoformat(),
                },
                "IBE": {
                    "ticker": "IBE",
                    "company_name": "Iberdrola, S.A.",
                    "current_price": Decimal("11.0200"),
                    "current_price_date": today,
                    "current_price_date_label": today.isoformat(),
                },
                "REP": {
                    "ticker": "REP",
                    "company_name": "Repsol, S.A.",
                    "current_price": Decimal("13.6200"),
                    "current_price_date": today,
                    "current_price_date_label": today.isoformat(),
                },
            },
        )

        self.assertTrue(context["available"])
        self.assertEqual(context["runs_count_3m"], 6)
        self.assertEqual(context["distinct_days_count_3m"], 3)
        self.assertEqual(context["policy"]["max_total_positions"], 5)
        self.assertEqual(context["policy"]["max_sector_positions"], 2)
        self.assertEqual(context["policy"]["max_company_pct"], Decimal("30"))
        top_row = next(item for item in context["rows"] if item["ticker"] == "IBE")
        self.assertEqual(top_row["ticker"], "IBE")
        self.assertEqual(top_row["appearances_3m"], 6)
        self.assertEqual(top_row["presence_pct_3m"], Decimal("100.0"))
        self.assertEqual(top_row["distinct_days_3m"], 3)
        self.assertEqual(top_row["day_presence_pct_3m"], Decimal("100.0"))
        self.assertEqual(top_row["top3_3m"], 6)
        self.assertEqual(top_row["top3_pct_3m"], Decimal("100.0"))
        self.assertEqual(top_row["persistence_label"], "Media")
        self.assertEqual(top_row["strategy_labels_3m_label"], "12M principal, 5A principal")
        self.assertEqual(top_row["average_return_12m_3m"], Decimal("11.0"))
        self.assertEqual(top_row["average_return_5y_3m"], Decimal("72.0"))
        self.assertEqual(top_row["average_reliability_label_3m"], "Alta")
        self.assertEqual(top_row["average_reliability_score_3m"], Decimal("80.0"))
        self.assertTrue(top_row["buy_recommendation_available"])
        self.assertEqual(top_row["average_buy_price_3m"], Decimal("10.4500"))
        self.assertEqual(top_row["average_sell_price_3m"], Decimal("11.6800"))
        self.assertEqual(top_row["latest_sell_window_label"], f"abril {today.year + 1}")
        self.assertEqual(top_row["average_interval_return_pct_3m"], Decimal("11.6"))
        self.assertEqual(top_row["average_annualized_return_pct_3m"], Decimal("12.7"))
        self.assertEqual(top_row["average_allocated_amount_3m"], Decimal("25000.00"))
        self.assertEqual(top_row["current_price"], Decimal("11.0200"))
        self.assertEqual(top_row["current_position_label"], "Dentro del tramo esperado")
        self.assertIsNotNone(top_row["current_vs_entry_pct"])
        self.assertIsNotNone(top_row["remaining_to_exit_pct"])
        self.assertEqual([item["ticker"] for item in context["rows"][:3]], ["ACS", "IBE", "REP"])
        self.assertEqual(
            [item["margin_pct"] for item in top_row["average_year_margins"]],
            [Decimal("11.0"), Decimal("7.0"), Decimal("8.0"), Decimal("9.0"), Decimal("10.0")],
        )
        self.assertTrue(context["top_non_owned_recommendation"]["available"])
        self.assertEqual(context["top_non_owned_recommendation"]["ticker"], "ACS")
        self.assertEqual(context["top_non_owned_recommendation"]["sell_window_label"], f"noviembre {today.year}")
        self.assertEqual(context["top_non_owned_recommendation"]["interval_return_pct"], Decimal("10.8"))
        self.assertEqual(context["top_non_owned_recommendation"]["holding_annualized_return_pct"], Decimal("22.7664"))
        self.assertEqual(context["top_non_owned_recommendation"]["current_price"], Decimal("40.1000"))
        self.assertEqual(context["top_non_owned_recommendation"]["current_position_label"], "Dentro del tramo esperado")

    def test_dashboard_can_extend_optimizer_to_full_ibex_universe(self):
        acerinox = find_equity_company_profile("Acerinox")
        acs = find_equity_company_profile("ACS")
        companies = [
            {
                "ticker": acerinox["ticker"],
                "company_name": acerinox["company_name"],
                "quote_symbol": acerinox["quote_symbol"],
                "sector": acerinox["sector_label"],
                "dividend_yield": Decimal("4.20"),
                "catalog_profile": acerinox,
            },
            {
                "ticker": acs["ticker"],
                "company_name": acs["company_name"],
                "quote_symbol": acs["quote_symbol"],
                "sector": acs["sector_label"],
                "dividend_yield": Decimal("3.10"),
                "catalog_profile": acs,
            },
        ]

        def fake_market_series(symbol, range_key="10y", interval="1d"):
            growth = Decimal("1.0210") if symbol.startswith("ACS") else Decimal("1.0170")
            return build_compound_market_series(symbol, symbol, growth=growth, start_price=Decimal("12.0000"))

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            return build_compound_market_series(
                benchmark_symbol or "^IBEX",
                benchmark_name or "Referencia",
                growth=Decimal("1.0070"),
                start_price=Decimal("100.0000"),
            )

        empty_workbook = {
            "available": False,
            "path": "",
            "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=empty_workbook),
            patch("equities.services.build_ibex_universe_companies", return_value=companies),
            patch("equities.services.fetch_market_series", side_effect=fake_market_series),
            patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series),
        ):
            dashboard = build_equity_analysis_dashboard(
                [],
                include_ibex_universe=True,
                ibex_company_limit=2,
            )

        self.assertEqual(len(dashboard["ibex_universe_rows"]), 2)
        self.assertEqual(len(dashboard["optimizer_cards"]), 2)
        self.assertTrue(all(row["status_key"] == "ibex" for row in dashboard["ibex_universe_rows"]))
        self.assertGreaterEqual(dashboard["ibex_universe_summary"]["buy_alert_count"], 1)
        self.assertEqual(dashboard["ibex_universe_summary"]["broker_assumption"], "Interactive Brokers")

    def test_ibex_universe_master_list_includes_registered_positions_without_duplication(self):
        iberdrola = find_equity_company_profile("Iberdrola")
        acs = find_equity_company_profile("ACS")
        tracked = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Seguimiento",
            ticker=iberdrola["ticker"],
            quote_symbol=iberdrola["quote_symbol"],
            reference_profile=EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
            benchmark_symbol=SPAIN_ELECTRICITY_DEMAND_SYMBOL,
            benchmark_name=SPAIN_ELECTRICITY_DEMAND_NAME,
            company_name=iberdrola["company_name"],
            shares=Decimal("0"),
            average_cost_per_share=Decimal("12.0000"),
            current_price_per_share=Decimal("12.2000"),
        )
        stock_series = build_compound_market_series(
            tracked.quote_symbol,
            tracked.company_name,
            growth=Decimal("1.0100"),
            start_price=Decimal("12.0000"),
        )
        reference_series = build_compound_market_series(
            tracked.benchmark_symbol,
            tracked.benchmark_name,
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        for stock_point, reference_point in zip(stock_series.points, reference_series.points):
            tracked.price_history.create(
                price_date=stock_point["date"],
                open_price=stock_point["open"],
                high_price=stock_point["high"],
                low_price=stock_point["low"],
                close_price=stock_point["close"],
                benchmark_close=reference_point["close"],
            )

        companies = [
            {
                "ticker": iberdrola["ticker"],
                "company_name": iberdrola["company_name"],
                "quote_symbol": iberdrola["quote_symbol"],
                "sector": iberdrola["sector_label"],
                "dividend_yield": Decimal("4.20"),
                "catalog_profile": iberdrola,
            },
            {
                "ticker": acs["ticker"],
                "company_name": acs["company_name"],
                "quote_symbol": acs["quote_symbol"],
                "sector": acs["sector_label"],
                "dividend_yield": Decimal("3.10"),
                "catalog_profile": acs,
            },
        ]

        def fake_market_series(symbol, range_key="10y", interval="1d"):
            growth = Decimal("1.0210") if symbol.startswith("ACS") else Decimal("1.0170")
            return build_compound_market_series(symbol, symbol, growth=growth, start_price=Decimal("12.0000"))

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            return build_compound_market_series(
                benchmark_symbol or "^IBEX",
                benchmark_name or "Referencia",
                growth=Decimal("1.0070"),
                start_price=Decimal("100.0000"),
            )

        empty_workbook = {
            "available": False,
            "path": "",
            "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=empty_workbook),
            patch("equities.services.build_ibex_universe_companies", return_value=companies),
            patch("equities.services.fetch_market_series", side_effect=fake_market_series),
            patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series),
        ):
            dashboard = build_equity_analysis_dashboard([tracked], include_ibex_universe=True, ibex_company_limit=5)

        rows = dashboard["ibex_universe_rows"]
        self.assertEqual(len(rows), 2)
        tracked_row = next(row for row in rows if row["ticker"] == "IBE")
        self.assertEqual(tracked_row["status_label"], "En seguimiento")
        self.assertEqual(dashboard["ibex_universe_summary"]["registered_watchlist_count"], 1)
        self.assertEqual(dashboard["ibex_universe_summary"]["radar_only_count"], 1)
        self.assertEqual(len(dashboard["optimizer_cards"]), 2)

    def test_dashboard_uses_summary_mode_for_full_ibex_universe_optimizer(self):
        acs = find_equity_company_profile("ACS")
        company = {
            "ticker": acs["ticker"],
            "company_name": acs["company_name"],
            "quote_symbol": acs["quote_symbol"],
            "sector": acs["sector_label"],
            "dividend_yield": Decimal("3.10"),
            "catalog_profile": acs,
        }

        def fake_market_series(symbol, range_key="10y", interval="1d"):
            growth = Decimal("1.0210") if symbol.startswith("ACS") else Decimal("1.0070")
            start_price = Decimal("12.0000") if symbol.startswith("ACS") else Decimal("100.0000")
            return build_compound_market_series(symbol, symbol, growth=growth, start_price=start_price)

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            return build_compound_market_series(
                benchmark_symbol or "^IBEX",
                benchmark_name or "Referencia",
                growth=Decimal("1.0070"),
                start_price=Decimal("100.0000"),
            )

        empty_workbook = {
            "available": False,
            "path": "",
            "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=empty_workbook),
            patch("equities.services.build_ibex_universe_companies", return_value=[company]),
            patch("equities.services.fetch_market_series", side_effect=fake_market_series),
            patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series),
        ):
            dashboard = build_equity_analysis_dashboard(
                [],
                include_ibex_universe=True,
                ibex_company_limit=1,
            )

        self.assertEqual(len(dashboard["optimizer_cards"]), 1)
        card = dashboard["optimizer_cards"][0]
        self.assertFalse(card["historical_chart"]["available"])
        self.assertFalse(card["best_correlation_chart"]["available"])
        self.assertFalse(card["projection_12m_chart"]["available"])
        self.assertFalse(card["cycle_projection_5y_chart"]["available"])
        self.assertFalse(card["projection_backtest"]["monthly_chart"]["available"])
        self.assertEqual(card["suggested_references"], [])

    def test_fetch_market_series_reuses_cache_within_same_bucket(self):
        payload = {
            "chart": {
                "error": None,
                "result": [
                    {
                        "meta": {
                            "regularMarketPrice": 12.34,
                            "regularMarketTime": 1712707200,
                            "shortName": "Iberdrola",
                        },
                        "timestamp": [1712620800, 1712707200],
                        "indicators": {
                            "quote": [
                                {
                                    "open": [12.0, 12.2],
                                    "high": [12.3, 12.5],
                                    "low": [11.9, 12.1],
                                    "close": [12.15, 12.34],
                                }
                            ]
                        },
                    }
                ],
            }
        }

        with (
            patch("equities.services.build_market_data_cache_bucket", return_value=777),
            patch(
                "equities.services.urlopen",
                side_effect=[
                    FakeHTTPResponse(json.dumps(payload).encode("utf-8")),
                ],
            ) as mocked_urlopen,
        ):
            first_series = fetch_market_series("ibe.mc")
            second_series = fetch_market_series("IBE.MC")

        self.assertEqual(mocked_urlopen.call_count, 1)
        self.assertEqual(first_series.latest_price, Decimal("12.3400"))
        self.assertEqual(second_series.latest_price, Decimal("12.3400"))
        self.assertIsNot(first_series, second_series)
        self.assertIsNot(first_series.points, second_series.points)

    @override_settings(
        AI_LLM_PROVIDER="openai",
        OPENAI_API_KEY="test-openai-key",
        OPENAI_DEFAULT_MODEL="gpt-4o-mini",
        OPENAI_MAX_TOKENS=768,
    )
    def test_enrich_dashboard_with_ai_analysis_uses_openai_account(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        history_card = build_equity_history_cards([position])[0]
        dashboard = {
            "history_cards": [history_card],
            "ibex_universe_cards": [],
        }
        payload = {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "summary": "La lectura 12M sigue positiva y el backtest aguanta razonablemente bien.",
                                "action_label": "Mantener",
                                "action_note": "Mantener mientras no pierda fiabilidad ni empeore el alpha.",
                                "confidence_label": "Media",
                                "drivers": [
                                    "El retorno esperado 12M sigue en positivo",
                                    "La seguridad no esta deteriorada",
                                ],
                                "risks": [
                                    "La volatilidad puede ampliar el rango",
                                    "La validacion historica no es perfecta",
                                ],
                                "backtest_note": "El modelo acierta la direccion mas veces de las que falla.",
                                "cycle_note": "El ciclo 5A aun acompana, aunque con posibles fases de correccion.",
                            }
                        )
                    }
                }
            ],
            "usage": {
                "prompt_tokens": 1200,
                "completion_tokens": 300,
            },
        }

        with patch(
            "equities.llm_analysis.urlopen",
            return_value=FakeHTTPResponse(json.dumps(payload).encode("utf-8")),
        ) as mocked_urlopen:
            summary = enrich_dashboard_with_ai_analysis(
                dashboard,
                analysis_date=timezone.localdate(),
            )

        self.assertEqual(mocked_urlopen.call_count, 1)
        self.assertTrue(summary["enabled"])
        self.assertEqual(summary["provider"], "openai")
        self.assertEqual(summary["completed_count"], 1)
        self.assertEqual(summary["failed_count"], 0)
        self.assertEqual(summary["estimated_cost_usd"], "0.0004")
        self.assertTrue(history_card["ai_analysis"]["available"])
        self.assertEqual(history_card["ai_analysis"]["action_label"], "Mantener")
        self.assertEqual(history_card["ai_analysis"]["confidence_label"], "Media")
        self.assertIn("ChatGPT", history_card["ai_analysis"]["model_label"])

    @override_settings(
        AI_LLM_PROVIDER="anthropic",
        ANTHROPIC_API_KEY="test-anthropic-key",
        CLAUDE_DEFAULT_MODEL="claude-sonnet-4-20250514",
        CLAUDE_MAX_TOKENS=768,
    )
    def test_enrich_dashboard_with_ai_analysis_uses_claude_account(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("34.0000"),
            current_price_per_share=Decimal("34.0000"),
            annual_dividend_income=Decimal("0.50"),
            annual_maintenance_cost=Decimal("0.00"),
        )
        populate_position_history(position)
        history_card = build_equity_history_cards([position])[0]
        dashboard = {
            "history_cards": [],
            "ibex_universe_cards": [history_card],
        }
        payload = {
            "content": [
                {
                    "type": "text",
                    "text": json.dumps(
                        {
                            "summary": "La lectura 12M sigue constructiva y el ciclo 5A aun acompana.",
                            "action_label": "Comprar",
                            "action_note": "Comprar con disciplina porque el modelo sigue favorable.",
                            "confidence_label": "Alta",
                            "drivers": [
                                "La seguridad sigue por encima de la media",
                                "El ciclo largo no esta roto",
                            ],
                            "risks": [
                                "Una correccion de mercado puede enfriar la entrada",
                            ],
                            "backtest_note": "La comprobacion historica acompana la decision actual.",
                            "cycle_note": "El escenario 5A sigue sesgado a expansion con pausas.",
                        }
                    ),
                }
            ],
            "usage": {
                "input_tokens": 1500,
                "output_tokens": 260,
            },
        }

        with patch(
            "equities.llm_analysis.urlopen",
            return_value=FakeHTTPResponse(json.dumps(payload).encode("utf-8")),
        ) as mocked_urlopen:
            summary = enrich_dashboard_with_ai_analysis(
                dashboard,
                analysis_date=timezone.localdate(),
            )

        self.assertEqual(mocked_urlopen.call_count, 1)
        self.assertTrue(summary["enabled"])
        self.assertEqual(summary["provider"], "anthropic")
        self.assertEqual(summary["completed_count"], 1)
        self.assertEqual(summary["estimated_cost_usd"], "0.0084")
        self.assertTrue(history_card["ai_analysis"]["available"])
        self.assertEqual(history_card["ai_analysis"]["action_label"], "Comprar")
        self.assertEqual(history_card["ai_analysis"]["confidence_label"], "Alta")
        self.assertIn("Claude", history_card["ai_analysis"]["model_label"])

    def test_build_card_llm_context_includes_news_context_when_available(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("34.0000"),
            current_price_per_share=Decimal("34.0000"),
        )
        populate_position_history(position)
        history_card = build_equity_history_cards([position])[0]
        history_card["news_context"] = {
            "available": True,
            "label": "Contexto adverso",
            "score": Decimal("-3.40"),
            "items_count": 3,
            "top_tags": ["geopolitica", "energia"],
            "material_event": True,
            "material_note": "Se detecta un evento geopolitico reciente con capacidad de alterar el escenario base.",
            "captured_at_label": "2026-04-18 01:30",
            "note": "Empresa adversa | mercado adversa",
            "company_signal": {
                "available": True,
                "label": "Empresa Adversa",
                "score": Decimal("-2.10"),
                "items_count": 1,
                "positive_count": 0,
                "negative_count": 1,
                "neutral_count": 0,
                "top_tags": ["geopolitica"],
                "note": "Hay presion sobre la compania.",
                "items": [
                    {
                        "title": "ACS cae por tension geopolitica",
                        "source": "Fuente Test",
                        "published_label": "2026-04-18",
                        "tone": "negative",
                        "score": Decimal("-2.10"),
                        "tags": ["geopolitica"],
                    }
                ],
            },
            "sector_signal": {
                "available": True,
                "label": "Sector Adversa",
                "score": Decimal("-1.20"),
                "items_count": 1,
                "positive_count": 0,
                "negative_count": 1,
                "neutral_count": 0,
                "top_tags": ["energia"],
                "note": "El sector sufre por costes energeticos.",
                "items": [],
            },
            "market_signal": {
                "available": True,
                "label": "Mercado Adversa",
                "score": Decimal("-2.80"),
                "items_count": 1,
                "positive_count": 0,
                "negative_count": 1,
                "neutral_count": 0,
                "top_tags": ["geopolitica", "energia"],
                "note": "El mercado se pone defensivo.",
                "items": [],
            },
            "top_items": [
                {
                    "title": "ACS cae por tension geopolitica",
                    "source": "Fuente Test",
                    "published_label": "2026-04-18",
                    "tone": "negative",
                    "score": Decimal("-2.10"),
                    "tags": ["geopolitica"],
                }
            ],
        }
        history_card["expert_consensus"] = {
            "available": True,
            "label": "Consenso experto favorable",
            "score": Decimal("2.60"),
            "quality_score": Decimal("78.00"),
            "quality_label": "Alta",
            "items_count": 2,
            "best_sources": ["JPMorgan", "Goldman Sachs"],
            "captured_at_label": "2026-04-18 01:35",
            "note": "Las casas con mejor track record refuerzan una lectura positiva.",
            "company_signal": {
                "available": True,
                "label": "Empresa Consenso favorable",
                "score": Decimal("3.10"),
                "items_count": 1,
                "positive_count": 1,
                "negative_count": 0,
                "neutral_count": 0,
                "note": "La empresa recibe apoyo de varias firmas.",
                "items": [
                    {
                        "title": "JPMorgan mantiene compra sobre ACS",
                        "source": "Reuters",
                        "published_label": "2026-04-18",
                        "tone": "positive",
                        "score": Decimal("2.40"),
                    }
                ],
            },
            "market_signal": {
                "available": True,
                "label": "Mercado Consenso mixto",
                "score": Decimal("1.10"),
                "items_count": 1,
                "positive_count": 1,
                "negative_count": 0,
                "neutral_count": 0,
                "note": "El mercado acompana pero con menos conviccion.",
                "items": [],
            },
            "wall_street_signal": {
                "available": True,
                "label": "Wall Street favorable",
                "score": Decimal("2.40"),
                "quality_score": Decimal("80.00"),
                "quality_label": "Alta",
                "items_count": 2,
                "positive_count": 2,
                "negative_count": 0,
                "neutral_count": 0,
                "note": "S&P 500 y Nasdaq acompanan el escenario.",
                "items": [
                    {
                        "title": "S&P 500: 3M 4.20% | 12M 9.80%",
                        "source": "Mercado USA",
                        "published_label": "2026-04-18",
                        "tone": "positive",
                        "score": Decimal("1.10"),
                    }
                ],
                "source_rows": [
                    {
                        "source": "S&P 500",
                        "quality_label": "Alta",
                        "quality_score": Decimal("80.00"),
                        "source_weight": Decimal("1.10"),
                        "observations_count": 0,
                        "hit_rate_pct": None,
                        "current_items_count": 1,
                        "current_score": Decimal("1.10"),
                        "weighted_score": Decimal("1.21"),
                    }
                ],
            },
            "bridgewater_signal": {
                "available": True,
                "label": "Bridgewater favorable",
                "score": Decimal("1.90"),
                "quality_score": Decimal("60.00"),
                "quality_label": "Media",
                "items_count": 1,
                "positive_count": 1,
                "negative_count": 0,
                "neutral_count": 0,
                "note": "Bridgewater mantiene una lectura macro constructiva.",
                "items": [
                    {
                        "title": "Bridgewater Daily Observations",
                        "source": "Informe local",
                        "published_label": "2026-04-18",
                        "tone": "positive",
                        "score": Decimal("1.20"),
                    }
                ],
                "source_rows": [
                    {
                        "source": "Bridgewater",
                        "quality_label": "Media",
                        "quality_score": Decimal("60.00"),
                        "source_weight": Decimal("0.98"),
                        "observations_count": 0,
                        "hit_rate_pct": None,
                        "current_items_count": 1,
                        "current_score": Decimal("1.20"),
                        "weighted_score": Decimal("1.18"),
                    }
                ],
            },
            "source_rows": [
                {
                    "source": "JPMorgan",
                    "quality_label": "Alta",
                    "quality_score": Decimal("81.00"),
                    "source_weight": Decimal("1.11"),
                    "observations_count": 6,
                    "hit_rate_pct": Decimal("66.70"),
                    "current_items_count": 1,
                    "current_score": Decimal("2.40"),
                    "weighted_score": Decimal("2.66"),
                }
            ],
            "top_items": [
                {
                    "title": "JPMorgan mantiene compra sobre ACS",
                    "source": "Reuters",
                    "expert_source": "JPMorgan",
                    "published_label": "2026-04-18",
                    "tone": "positive",
                    "score": Decimal("2.40"),
                    "quality_label": "Alta",
                }
            ],
        }
        refresh_card_projection_visuals(history_card)

        context = build_card_llm_context(history_card, analysis_date=date(2026, 4, 18), scope="ibex")

        self.assertTrue(context["news_context"]["available"])
        self.assertTrue(context["news_context"]["material_event"])
        self.assertEqual(context["news_context"]["top_tags"][0], "geopolitica")
        self.assertEqual(context["news_context"]["company_signal"]["label"], "Empresa Adversa")
        self.assertEqual(context["news_context"]["top_headlines"][0]["title"], "ACS cae por tension geopolitica")
        self.assertTrue(context["expert_consensus"]["available"])
        self.assertEqual(context["expert_consensus"]["quality_label"], "Alta")
        self.assertEqual(context["expert_consensus"]["best_sources"][0], "JPMorgan")
        self.assertEqual(context["expert_consensus"]["source_rows"][0]["source"], "JPMorgan")
        self.assertEqual(context["expert_consensus"]["top_forecasts"][0]["title"], "JPMorgan mantiene compra sobre ACS")
        self.assertTrue(context["expert_consensus"]["wall_street_signal"]["available"])
        self.assertEqual(context["expert_consensus"]["wall_street_signal"]["label"], "Wall Street favorable")
        self.assertTrue(context["expert_consensus"]["bridgewater_signal"]["available"])
        self.assertEqual(context["expert_consensus"]["bridgewater_signal"]["label"], "Bridgewater favorable")
        self.assertTrue(context["information_basis"]["available"])
        self.assertEqual(context["information_basis"]["rows"][0]["source_label"], "Fuente Test")
        self.assertIn("geopolitica", context["information_basis"]["summary"].lower())
        self.assertTrue(context["information_basis"]["geopolitical_flag"])
        self.assertTrue(context["information_basis"]["macro_flag"])
        self.assertTrue(context["information_basis"]["highlights"])
        self.assertTrue(context["technical_view"]["available"])
        self.assertTrue(context["technical_view"]["signal_label"])
        self.assertTrue(context["technical_view"]["trend_label"])
        self.assertTrue(context["technical_view"]["note"])

    @override_settings(
        AI_LLM_PROVIDER="anthropic",
        ANTHROPIC_API_KEY="test-anthropic-key",
        CLAUDE_DEFAULT_MODEL="claude-sonnet-4-20250514",
        AI_LLM_RETRY_ATTEMPTS=3,
        AI_LLM_RATE_LIMIT_RETRY_SECONDS=15,
    )
    def test_enrich_dashboard_with_ai_analysis_retries_rate_limited_claude_calls(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="LOG",
            quote_symbol="LOG.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Logista",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("24.0000"),
            current_price_per_share=Decimal("24.0000"),
            annual_dividend_income=Decimal("0.50"),
            annual_maintenance_cost=Decimal("0.00"),
        )
        populate_position_history(position)
        history_card = build_equity_history_cards([position])[0]
        dashboard = {
            "history_cards": [],
            "ibex_universe_cards": [history_card],
        }
        success_payload = {
            "content": [
                {
                    "type": "text",
                    "text": json.dumps(
                        {
                            "summary": "La lectura cuantitativa vuelve a ser util tras el reintento.",
                            "action_label": "Mantener",
                            "action_note": "Mantener mientras no empeore la fiabilidad.",
                            "confidence_label": "Media",
                            "drivers": [
                                "La rentabilidad esperada sigue positiva",
                            ],
                            "risks": [
                                "La volatilidad sigue presente",
                            ],
                            "backtest_note": "El backtest sigue siendo razonable.",
                            "cycle_note": "El ciclo no se ha roto.",
                        }
                    ),
                }
            ],
            "usage": {
                "input_tokens": 900,
                "output_tokens": 180,
            },
        }
        rate_limit_body = json.dumps(
            {
                "type": "error",
                "error": {
                    "type": "rate_limit_error",
                    "message": "This request would exceed your organization's rate limit.",
                },
            }
        ).encode("utf-8")
        rate_limit_error = HTTPError(
            url="https://api.anthropic.com/v1/messages",
            code=429,
            msg="Too Many Requests",
            hdrs={"Retry-After": "1"},
            fp=io.BytesIO(rate_limit_body),
        )

        with (
            patch(
                "equities.llm_analysis.urlopen",
                side_effect=[rate_limit_error, FakeHTTPResponse(json.dumps(success_payload).encode("utf-8"))],
            ) as mocked_urlopen,
            patch("equities.llm_analysis.time.sleep") as mocked_sleep,
        ):
            summary = enrich_dashboard_with_ai_analysis(
                dashboard,
                analysis_date=timezone.localdate(),
            )

        self.assertEqual(mocked_urlopen.call_count, 2)
        mocked_sleep.assert_called_once_with(15)
        self.assertEqual(summary["completed_count"], 1)
        self.assertEqual(summary["failed_count"], 0)
        self.assertTrue(history_card["ai_analysis"]["available"])
        self.assertEqual(history_card["ai_analysis"]["action_label"], "Mantener")

    def test_run_nightly_equity_analysis_persists_full_ibex_cache(self):
        analysis_day = timezone.localdate()
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        tracked_card = build_equity_history_cards([position])[0]
        tracked_card["news_context"] = {
            "available": True,
            "label": "Contexto adverso",
            "score": Decimal("-2.40"),
            "items_count": 1,
            "top_tags": ["geopolitica"],
            "material_event": True,
            "material_note": "Se detecta un evento geopolitico reciente.",
            "note": "Mercado adversa",
            "top_items": [
                {
                    "title": "Iberdrola vigila la tension geopolitica",
                    "source": "Reuters",
                    "published_label": "2026-04-16",
                    "tone": "negative",
                    "score": Decimal("-1.60"),
                    "tags": ["geopolitica"],
                }
            ],
        }
        tracked_card["expert_consensus"] = {
            "available": True,
            "label": "Consenso experto favorable",
            "score": Decimal("2.10"),
            "quality_score": Decimal("76.00"),
            "quality_label": "Alta",
            "items_count": 1,
            "best_sources": ["JPMorgan"],
            "note": "JPMorgan mantiene una lectura constructiva.",
            "top_items": [
                {
                    "title": "JPMorgan mantiene compra sobre Iberdrola",
                    "source": "Reuters",
                    "expert_source": "JPMorgan",
                    "published_label": "2026-04-16",
                    "tone": "positive",
                    "score": Decimal("2.10"),
                    "tags": [],
                }
            ],
            "wall_street_signal": {"available": False},
            "bridgewater_signal": {"available": False},
        }
        ibex_position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("34.0000"),
            current_price_per_share=Decimal("34.0000"),
            annual_dividend_income=Decimal("0.50"),
        )
        ibex_card = {
            **tracked_card,
            "position": ibex_position,
            "status_key": "ibex",
            "status_label": "Solo radar",
            "detail_anchor": "",
            "sector_label": "Construccion",
        }
        tracked_card["ai_analysis"] = {
            "available": True,
            "provider": "anthropic",
            "label": "Claude claude-sonnet-4-20250514",
            "model": "claude-sonnet-4-20250514",
            "model_label": "Claude claude-sonnet-4-20250514",
            "summary": "Lectura Claude para posicion en cartera.",
            "action_label": "Mantener",
            "action_note": "Mantener mientras no cambie la tesis.",
            "confidence_label": "Media",
            "drivers": ["Retorno esperado positivo"],
            "risks": ["Backtest mejorable"],
            "backtest_note": "El backtest sigue siendo util.",
            "cycle_note": "El ciclo 5A sigue acompasado.",
            "consistency_label": "Alineado",
            "consistency_note": "La IA coincide con el motor cuantitativo.",
            "generated_at": timezone.now().isoformat(),
            "generated_at_label": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
            "usage": {"input_tokens": 1100, "output_tokens": 320, "estimated_cost_usd": "0.0081"},
        }
        ibex_card["ai_analysis"] = {
            "available": True,
            "provider": "anthropic",
            "label": "Claude claude-sonnet-4-20250514",
            "model": "claude-sonnet-4-20250514",
            "model_label": "Claude claude-sonnet-4-20250514",
            "summary": "Lectura Claude para radar IBEX.",
            "action_label": "Comprar",
            "action_note": "Comprar con disciplina si aguanta la tendencia.",
            "confidence_label": "Media",
            "drivers": ["Rentabilidad 12M positiva"],
            "risks": ["Ciclo todavia sensible"],
            "backtest_note": "La validacion historica acompana.",
            "cycle_note": "El ciclo largo sigue favorable.",
            "consistency_label": "Alineado",
            "consistency_note": "La IA coincide con el motor cuantitativo.",
            "generated_at": timezone.now().isoformat(),
            "generated_at_label": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
            "usage": {"input_tokens": 1100, "output_tokens": 320, "estimated_cost_usd": "0.0081"},
        }
        dashboard = {
            "history_cards": [tracked_card],
            "owned_history_cards": [tracked_card],
            "ibex_universe_cards": [ibex_card],
            "ibex_universe_summary": {
                "available": True,
                "analyzed_count": 1,
                "buy_alert_count": 1,
                "sell_alert_count": 0,
                "watch_alert_count": 0,
                "registered_count": 0,
                "registered_owned_count": 0,
                "registered_watchlist_count": 0,
                "radar_only_count": 1,
                "failed_count": 0,
                "failures": [],
                "broker_assumption": "Interactive Brokers",
                "trade_channel_label": "App",
                "top_pick": {"ticker": "ACS", "company_name": "ACS"},
            },
            "reference_guide_summary": {
                "available": True,
                "workbook_loaded": False,
                "source_label": "",
                "tracked_count": 1,
                "owned_count": 1,
                "watchlist_count": 0,
                "guide_only_count": 0,
            },
        }

        with (
            patch("equities.nightly_analysis.sync_all_equities_market_data", return_value=[]),
            patch("equities.nightly_analysis.build_equity_analysis_dashboard", return_value=dashboard) as mocked_dashboard,
            patch(
                "equities.nightly_analysis.attach_llm_news_context_to_dashboard",
                return_value={"enabled": True, "signals_count": 2, "items_count": 0, "material_event_count": 0},
            ),
            patch(
                "equities.nightly_analysis.resolve_ai_provider_config",
                return_value=type(
                    "ProviderConfig",
                    (),
                    {
                        "available": True,
                        "provider": "anthropic",
                        "label": "Claude claude-sonnet-4-20250514",
                        "model": "claude-sonnet-4-20250514",
                        "monthly_budget_usd": Decimal("50.00"),
                    },
                )(),
            ),
            patch(
                "equities.nightly_analysis.enrich_dashboard_with_ai_analysis",
                return_value={
                    "enabled": True,
                    "provider": "anthropic",
                    "label": "Claude claude-sonnet-4-20250514",
                    "model": "claude-sonnet-4-20250514",
                    "completed_count": 2,
                    "failed_count": 0,
                    "skipped_budget_count": 0,
                    "total_count": 2,
                    "input_tokens": 2200,
                    "output_tokens": 640,
                    "estimated_cost_usd": "0.0162",
                    "monthly_budget_usd": "50.0000",
                    "monthly_cost_before_run_usd": "0.0000",
                    "monthly_cost_after_run_usd": "0.0162",
                    "failures": [],
                },
            ) as mocked_llm,
        ):
            run = run_nightly_equity_analysis(
                analysis_date=analysis_day,
                force=True,
            )

        self.assertIsNotNone(run)
        self.assertEqual(run.status, EquityNightlyAnalysisRun.Status.COMPLETED)
        self.assertEqual(run.snapshots.count(), 2)
        self.assertEqual(run.agent_provider, "anthropic")
        self.assertIn("Claude", run.agent_label)
        self.assertEqual(
            mocked_dashboard.call_args.kwargs,
            {
                "include_ibex_universe": True,
                "ibex_company_limit": None,
                "ibex_include_visuals": True,
                "ibex_include_reference_suggestions": True,
                "ibex_include_fundamentals": True,
            },
        )
        mocked_llm.assert_called_once()
        self.assertEqual(run.summary_data["llm"]["provider"], "anthropic")
        self.assertEqual(run.summary_data["llm"]["completed_count"], 2)
        self.assertTrue(
            EquityNightlyAnalysisSnapshot.objects.filter(
                run=run,
                scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
                ticker="ACS",
            ).exists()
        )

    def test_run_nightly_equity_analysis_persists_expert_consensus_and_adjustments(self):
        analysis_day = date(2026, 4, 25)
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="REP",
            quote_symbol="REP.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Repsol",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position, growth=Decimal("1.0180"), benchmark_growth=Decimal("1.0070"), months=108)
        dashboard = build_equity_analysis_dashboard([position])

        def attach_expert_context(current_dashboard):
            summary = {
                "enabled": True,
                "signals_count": 1,
                "items_count": 3,
                "ranked_sources_count": 2,
                "strong_consensus_count": 1,
            }
            current_dashboard["history_cards"][0]["expert_consensus"] = {
                "available": True,
                "label": "Consenso experto adverso",
                "score": Decimal("-3.80"),
                "quality_score": Decimal("81.00"),
                "quality_label": "Alta",
                "items_count": 3,
                "note": "Las firmas con mejor track record mantienen una lectura mas cauta.",
                "best_sources": ["JPMorgan", "Goldman Sachs"],
                "source_rows": [
                    {
                        "source": "JPMorgan",
                        "quality_label": "Alta",
                        "quality_score": Decimal("81.00"),
                        "source_weight": Decimal("1.11"),
                        "observations_count": 6,
                        "hit_rate_pct": Decimal("66.70"),
                        "current_items_count": 1,
                        "current_score": Decimal("-2.30"),
                        "weighted_score": Decimal("-2.55"),
                    }
                ],
                "company_signal": {"available": True, "label": "Empresa Consenso adverso", "score": Decimal("-4.20"), "items": []},
                "market_signal": {"available": True, "label": "Mercado Consenso mixto", "score": Decimal("-2.10"), "items": []},
                "wall_street_signal": {"available": True, "label": "Wall Street adversa", "score": Decimal("-2.60"), "items": []},
                "bridgewater_signal": {"available": True, "label": "Bridgewater mixta", "score": Decimal("-0.90"), "items": []},
                "top_items": [],
                "captured_at_label": "2026-04-25 01:10",
            }
            current_dashboard["expert_consensus_summary"] = summary
            return summary

        with (
            patch("equities.nightly_analysis.sync_all_equities_market_data", return_value=[]),
            patch("equities.nightly_analysis.build_equity_analysis_dashboard", return_value=dashboard),
            patch(
                "equities.nightly_analysis.attach_llm_news_context_to_dashboard",
                return_value={"enabled": True, "signals_count": 1, "items_count": 0, "material_event_count": 0},
            ),
            patch("equities.nightly_analysis.attach_expert_consensus_to_dashboard", side_effect=attach_expert_context),
            patch(
                "equities.nightly_analysis.resolve_ai_provider_config",
                return_value=type(
                    "ProviderConfig",
                    (),
                    {
                        "available": False,
                        "provider": "core",
                        "label": "Analista nocturno",
                        "model": "",
                        "reason": "disabled",
                        "monthly_budget_usd": ZERO,
                    },
                )(),
            ),
        ):
            run = run_nightly_equity_analysis(
                analysis_date=analysis_day,
                force=True,
            )

        self.assertIsNotNone(run)
        self.assertEqual(run.summary_data["expert_consensus_summary"]["ranked_sources_count"], 2)
        self.assertEqual(run.summary_data["llm"]["expert_strong_consensus_count"], 1)
        snapshot = run.snapshots.get(scope=EquityNightlyAnalysisSnapshot.Scope.TRACKED, ticker="REP")
        self.assertTrue(snapshot.analysis_payload["expert_consensus"]["available"])
        self.assertTrue(snapshot.analysis_payload["projection"]["expert_adjustment"]["applied"])
        self.assertTrue(snapshot.analysis_payload["cycle_projection_5y"]["expert_adjustment"]["applied"])

    @override_settings(EQUITIES_NIGHTLY_LLM_REFRESH_ISO_WEEKDAYS=(2, 4))
    def test_run_nightly_equity_analysis_persists_expectation_reviews_for_scheduled_day(self):
        analysis_day = date(2026, 4, 16)
        self.assertEqual(analysis_day.isoweekday(), 4)
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        card = build_equity_history_cards([position])[0]
        dashboard = {
            "history_cards": [card],
            "owned_history_cards": [card],
            "ibex_universe_cards": [],
            "ibex_universe_summary": {"available": False},
            "reference_guide_summary": {"available": False},
        }

        with (
            patch("equities.nightly_analysis.sync_all_equities_market_data", return_value=[]),
            patch("equities.nightly_analysis.build_equity_analysis_dashboard", return_value=dashboard),
            patch(
                "equities.nightly_analysis.attach_llm_news_context_to_dashboard",
                return_value={"enabled": True, "signals_count": 0, "items_count": 0, "material_event_count": 0},
            ),
            patch(
                "equities.nightly_analysis.attach_expert_consensus_to_dashboard",
                return_value={"enabled": True, "signals_count": 0, "items_count": 0, "ranked_sources_count": 0, "strong_consensus_count": 0},
            ),
            patch(
                "equities.nightly_analysis.resolve_ai_provider_config",
                return_value=type(
                    "ProviderConfig",
                    (),
                    {
                        "available": False,
                        "provider": "core",
                        "label": "Analista nocturno",
                        "model": "",
                        "reason": "disabled",
                        "monthly_budget_usd": ZERO,
                    },
                )(),
            ),
        ):
            run = run_nightly_equity_analysis(
                analysis_date=analysis_day,
                force=True,
            )

        self.assertIsNotNone(run)
        self.assertEqual(run.status, EquityNightlyAnalysisRun.Status.COMPLETED)
        self.assertEqual(run.summary_data["expectation_review_kind"], "scheduled")
        self.assertEqual(run.summary_data["expectation_review_count"], 1)
        review = EquityExpectationReview.objects.get(run=run, ticker="IBE")
        self.assertEqual(review.review_kind, EquityExpectationReview.ReviewKind.SCHEDULED)
        self.assertEqual(review.scope, EquityExpectationReview.Scope.TRACKED)
        self.assertIsNotNone(review.expected_return_pct_1y)
        self.assertIsNotNone(review.expected_return_pct_5y)
        self.assertTrue(review.projection_12m_scenario_rows)
        self.assertTrue(review.cycle_5y_scenario_rows)

    def test_build_expectation_review_dashboard_builds_chart_and_corrective_equation(self):
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 24),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            completed_at=timezone.now(),
            summary_data={},
        )
        review_dates = (date(2026, 1, 15), date(2026, 2, 12), date(2026, 3, 12))
        current_prices = (Decimal("100.00"), Decimal("105.00"), Decimal("110.00"))
        expected_returns = (Decimal("24.00"), Decimal("18.00"), Decimal("12.00"))

        for review_date, current_price, expected_return in zip(review_dates, current_prices, expected_returns):
            EquityExpectationReview.objects.create(
                run=run,
                analysis_date=review_date,
                review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
                scope=EquityExpectationReview.Scope.IBEX,
                analysis_key=f"ibex:ANA:{review_date.isoformat()}",
                ticker="ANA",
                quote_symbol="ANA.MC",
                company_name="Acciona",
                current_price=current_price,
                expected_return_pct_1y=expected_return,
                projected_return_pct_1y=expected_return,
            )

        series = MarketSeries(
            symbol="ANA.MC",
            name="Acciona",
            latest_price=Decimal("118.00"),
            latest_date=date(2026, 4, 26),
            points=[{"date": date(2026, 4, 26), "close": Decimal("118.00")}],
        )

        with patch("equities.services.fetch_market_series", return_value=series):
            dashboard = build_expectation_review_dashboard(as_of=date(2026, 4, 26))

        self.assertTrue(dashboard["available"])
        self.assertEqual(dashboard["companies_count"], 1)
        company = dashboard["companies"][0]
        self.assertEqual(company["ticker"], "ANA")
        self.assertTrue(company["chart"]["available"])
        self.assertTrue(company["equation"]["available"])
        self.assertEqual(company["equation"]["sample_count"], 3)
        self.assertIn("Real observado", company["equation"]["formula_label"])
        self.assertEqual(company["reviews_count"], 3)
        self.assertEqual(len(company["rows"]), 3)

    def test_build_expectation_review_dashboard_builds_preview_chart_without_real_history(self):
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 26),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            completed_at=timezone.now(),
            summary_data={},
        )
        EquityExpectationReview.objects.create(
            run=run,
            analysis_date=date(2026, 4, 26),
            review_kind=EquityExpectationReview.ReviewKind.FORCED,
            scope=EquityExpectationReview.Scope.IBEX,
            analysis_key="ibex:CABK:2026-04-26",
            ticker="CABK",
            quote_symbol="CABK.MC",
            company_name="CaixaBank",
            current_price=Decimal("10.37"),
            expected_return_pct_1y=Decimal("-0.80"),
            expected_return_pct_2y=Decimal("18.40"),
            expected_return_pct_3y=Decimal("31.20"),
            expected_return_pct_4y=Decimal("46.50"),
            expected_return_pct_5y=Decimal("58.10"),
            projected_return_pct_1y=Decimal("-0.80"),
            projected_return_pct_5y=Decimal("58.10"),
        )

        series = MarketSeries(
            symbol="CABK.MC",
            name="CaixaBank",
            latest_price=Decimal("10.37"),
            latest_date=date(2026, 4, 26),
            points=[{"date": date(2026, 4, 26), "close": Decimal("10.37")}],
        )

        with patch("equities.services.fetch_market_series", return_value=series):
            dashboard = build_expectation_review_dashboard(as_of=date(2026, 4, 26))

        self.assertTrue(dashboard["available"])
        self.assertEqual(dashboard["companies_count"], 1)
        company = dashboard["companies"][0]
        self.assertTrue(company["preview_mode"])
        self.assertTrue(company["chart"]["available"])
        self.assertEqual(company["chart"]["actual_line"], "")
        self.assertTrue(company["chart"]["expected_line"])
        self.assertTrue(company["chart"]["has_expected_series"])
        self.assertFalse(company["chart"]["has_actual_series"])
        self.assertEqual(
            [marker["label"] for marker in company["chart"]["x_markers"]],
            ["Hoy", "1A", "2A", "3A", "4A", "5A"],
        )
        self.assertIn("ultima esperanza guardada", company["preview_note"])

    @override_settings(
        AI_LLM_PROVIDER="anthropic",
        ANTHROPIC_API_KEY="test-anthropic-key",
        CLAUDE_DEFAULT_MODEL="claude-sonnet-4-20250514",
        EQUITIES_NIGHTLY_LLM_REFRESH_ISO_WEEKDAYS=(2, 4),
    )
    def test_run_nightly_equity_analysis_reuses_last_claude_refresh_outside_scheduled_days(self):
        refresh_day = date(2026, 4, 16)
        analysis_day = date(2026, 4, 17)
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        previous_card = build_equity_history_cards([position])[0]
        previous_card["ai_analysis"] = {
            "available": True,
            "provider": "anthropic",
            "label": "Claude claude-sonnet-4-20250514",
            "model": "claude-sonnet-4-20250514",
            "model_label": "Claude claude-sonnet-4-20250514",
            "summary": "Lectura Claude guardada para reutilizar entre semana.",
            "action_label": "Mantener",
            "action_note": "Mantener con disciplina.",
            "confidence_label": "Media",
            "drivers": ["Retorno esperado positivo"],
            "risks": ["Volatilidad controlable"],
            "backtest_note": "El backtest sigue razonable.",
            "cycle_note": "El ciclo aun acompana.",
            "consistency_label": "Alineado",
            "consistency_note": "La IA coincide con el motor cuantitativo.",
            "generated_at": timezone.now().isoformat(),
            "generated_at_label": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
            "usage": {"input_tokens": 300, "output_tokens": 120, "estimated_cost_usd": "0.0030"},
        }
        persist_nightly_analysis_dashboard(
            {
                "history_cards": [previous_card],
                "ibex_universe_cards": [],
                "ibex_universe_summary": {
                    "available": False,
                    "analyzed_count": 0,
                    "buy_alert_count": 0,
                    "sell_alert_count": 0,
                    "watch_alert_count": 0,
                    "registered_count": 0,
                    "registered_owned_count": 0,
                    "registered_watchlist_count": 0,
                    "radar_only_count": 0,
                    "failed_count": 0,
                    "failures": [],
                    "broker_assumption": "",
                    "trade_channel_label": "",
                    "top_pick": None,
                },
                "reference_guide_summary": {
                    "available": True,
                    "workbook_loaded": False,
                    "source_label": "",
                    "tracked_count": 1,
                    "owned_count": 1,
                    "watchlist_count": 0,
                    "guide_only_count": 0,
                },
            },
            [position],
            analysis_date=refresh_day,
            agent_provider="anthropic",
            agent_label="Claude claude-sonnet-4-20250514",
            llm_summary={
                "enabled": True,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "completed_count": 1,
                "failed_count": 0,
                "skipped_budget_count": 0,
                "total_count": 1,
                "input_tokens": 300,
                "output_tokens": 120,
                "estimated_cost_usd": "0.0030",
                "monthly_budget_usd": "50.0000",
                "monthly_cost_before_run_usd": "0.0000",
                "monthly_cost_after_run_usd": "0.0030",
                "failures": [],
                "refresh_performed": True,
                "source_analysis_date": refresh_day.isoformat(),
                "source_analysis_date_label": refresh_day.isoformat(),
            },
        )

        fresh_card = build_equity_history_cards([position])[0]
        dashboard = {
            "history_cards": [fresh_card],
            "owned_history_cards": [fresh_card],
            "ibex_universe_cards": [],
            "ibex_universe_summary": {
                "available": False,
                "analyzed_count": 0,
                "buy_alert_count": 0,
                "sell_alert_count": 0,
                "watch_alert_count": 0,
                "registered_count": 0,
                "registered_owned_count": 0,
                "registered_watchlist_count": 0,
                "radar_only_count": 0,
                "failed_count": 0,
                "failures": [],
                "broker_assumption": "",
                "trade_channel_label": "",
                "top_pick": None,
            },
            "reference_guide_summary": {
                "available": True,
                "workbook_loaded": False,
                "source_label": "",
                "tracked_count": 1,
                "owned_count": 1,
                "watchlist_count": 0,
                "guide_only_count": 0,
            },
        }

        with (
            patch("equities.nightly_analysis.sync_all_equities_market_data", return_value=[]),
            patch("equities.nightly_analysis.build_equity_analysis_dashboard", return_value=dashboard),
            patch(
                "equities.nightly_analysis.attach_llm_news_context_to_dashboard",
                return_value={"enabled": True, "signals_count": 1, "items_count": 0, "material_event_count": 0},
            ),
            patch("equities.nightly_analysis.enrich_dashboard_with_ai_analysis") as mocked_llm,
        ):
            run = run_nightly_equity_analysis(
                analysis_date=analysis_day,
                force=False,
            )

        self.assertIsNotNone(run)
        mocked_llm.assert_not_called()
        self.assertTrue(run.summary_data["llm"]["reused"])
        self.assertEqual(run.summary_data["llm"]["source_analysis_date"], refresh_day.isoformat())
        self.assertEqual(run.summary_data["llm"]["next_refresh_date"], "2026-04-21")
        snapshot = run.snapshots.get(scope=EquityNightlyAnalysisSnapshot.Scope.TRACKED, ticker="IBE")
        self.assertTrue(snapshot.analysis_payload["ai_analysis"]["available"])
        self.assertEqual(
            snapshot.analysis_payload["ai_analysis"]["summary"],
            "Lectura Claude guardada para reutilizar entre semana.",
        )
        self.assertIn("reutilizando la ultima lectura IA", run.status_note)

    @override_settings(
        AI_LLM_PROVIDER="anthropic",
        ANTHROPIC_API_KEY="test-anthropic-key",
        CLAUDE_DEFAULT_MODEL="claude-sonnet-4-20250514",
        EQUITIES_NIGHTLY_LLM_REFRESH_ISO_WEEKDAYS=(2, 4),
    )
    def test_run_nightly_equity_analysis_refreshes_claude_on_material_news_shock_outside_schedule(self):
        analysis_day = date(2026, 4, 17)
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        fresh_card = build_equity_history_cards([position])[0]
        dashboard = {
            "history_cards": [fresh_card],
            "owned_history_cards": [fresh_card],
            "ibex_universe_cards": [],
            "ibex_universe_summary": {
                "available": False,
                "analyzed_count": 0,
                "buy_alert_count": 0,
                "sell_alert_count": 0,
                "watch_alert_count": 0,
                "registered_count": 0,
                "registered_owned_count": 0,
                "registered_watchlist_count": 0,
                "radar_only_count": 0,
                "failed_count": 0,
                "failures": [],
                "broker_assumption": "",
                "trade_channel_label": "",
                "top_pick": None,
            },
            "reference_guide_summary": {
                "available": True,
                "workbook_loaded": False,
                "source_label": "",
                "tracked_count": 1,
                "owned_count": 1,
                "watchlist_count": 0,
                "guide_only_count": 0,
            },
        }

        def attach_news_context(current_dashboard):
            current_dashboard["history_cards"][0]["news_context"] = {
                "available": True,
                "label": "Contexto adverso",
                "score": Decimal("-4.20"),
                "items_count": 3,
                "top_tags": ["geopolitica", "energia"],
                "material_event": True,
                "material_note": "Se detecta un evento geopolitico reciente con capacidad de alterar el escenario base.",
                "note": "Empresa adversa | mercado adversa",
                "company_signal": {"available": True, "label": "Empresa Adversa", "score": Decimal("-2.30"), "items_count": 1, "positive_count": 0, "negative_count": 1, "neutral_count": 0, "top_tags": ["geopolitica"], "note": "Presion sobre la compania.", "items": []},
                "sector_signal": {"available": True, "label": "Sector Adversa", "score": Decimal("-1.40"), "items_count": 1, "positive_count": 0, "negative_count": 1, "neutral_count": 0, "top_tags": ["energia"], "note": "Presion sectorial.", "items": []},
                "market_signal": {"available": True, "label": "Mercado Adversa", "score": Decimal("-3.20"), "items_count": 1, "positive_count": 0, "negative_count": 1, "neutral_count": 0, "top_tags": ["geopolitica", "energia"], "note": "Shock macro.", "items": []},
                "top_items": [],
                "captured_at_label": "2026-04-17 01:10",
            }
            return {
                "enabled": True,
                "signals_count": 1,
                "items_count": 3,
                "material_event_count": 1,
            }

        def successful_refresh(current_dashboard, analysis_date):
            current_dashboard["history_cards"][0]["ai_analysis"] = {
                "available": True,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "model_label": "Claude claude-sonnet-4-20250514",
                "summary": "Claude relee la accion por el shock informativo reciente.",
                "action_label": "Vigilar",
                "action_note": "Reducir confianza hasta que se estabilice el contexto.",
                "confidence_label": "Media",
                "drivers": ["La lectura cuantitativa base sigue viva"],
                "risks": ["El shock geopolitico puede desordenar el corto plazo"],
                "backtest_note": "El backtest no captura bien cisnes negros.",
                "cycle_note": "El ciclo 5A sigue util pero puede quedar temporalmente desfasado.",
                "news_note": "El evento geopolitico reciente obliga a vigilar mas de cerca.",
                "consistency_label": "Mixto",
                "consistency_note": "La IA introduce cautela extra por contexto web.",
                "generated_at": timezone.now().isoformat(),
                "generated_at_label": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
                "usage": {"input_tokens": 450, "output_tokens": 160, "estimated_cost_usd": "0.0042"},
            }
            return {
                "enabled": True,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "completed_count": 1,
                "failed_count": 0,
                "skipped_budget_count": 0,
                "total_count": 1,
                "input_tokens": 450,
                "output_tokens": 160,
                "estimated_cost_usd": "0.0042",
                "monthly_budget_usd": "50.0000",
                "monthly_cost_before_run_usd": "0.0000",
                "monthly_cost_after_run_usd": "0.0042",
                "failures": [],
            }

        with (
            patch("equities.nightly_analysis.sync_all_equities_market_data", return_value=[]),
            patch("equities.nightly_analysis.build_equity_analysis_dashboard", return_value=dashboard),
            patch("equities.nightly_analysis.attach_llm_news_context_to_dashboard", side_effect=attach_news_context),
            patch("equities.nightly_analysis.enrich_dashboard_with_ai_analysis", side_effect=successful_refresh) as mocked_llm,
        ):
            run = run_nightly_equity_analysis(
                analysis_date=analysis_day,
                force=False,
            )

        self.assertIsNotNone(run)
        mocked_llm.assert_called_once()
        self.assertTrue(run.summary_data["llm"]["refresh_performed"])
        self.assertEqual(run.summary_data["llm"]["refresh_reason"], "news_shock")
        self.assertEqual(run.summary_data["llm"]["material_news_event_count"], 1)
        self.assertIn("evento informativo material", run.status_note)
        snapshot = run.snapshots.get(scope=EquityNightlyAnalysisSnapshot.Scope.TRACKED, ticker="ACS")
        self.assertTrue(snapshot.analysis_payload["ai_analysis"]["available"])
        self.assertTrue(snapshot.analysis_payload["news_context"]["material_event"])
        self.assertTrue(snapshot.analysis_payload["projection"]["news_adjustment"]["applied"])
        self.assertTrue(snapshot.analysis_payload["cycle_projection_5y"]["news_adjustment"]["applied"])

    @override_settings(
        AI_LLM_PROVIDER="anthropic",
        ANTHROPIC_API_KEY="test-anthropic-key",
        CLAUDE_DEFAULT_MODEL="claude-sonnet-4-20250514",
        EQUITIES_NIGHTLY_LLM_REFRESH_ISO_WEEKDAYS=(2, 4),
    )
    def test_run_nightly_equity_analysis_keeps_previous_claude_when_refresh_fails(self):
        previous_day = date(2026, 4, 16)
        analysis_day = date(2026, 4, 21)
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        previous_card = build_equity_history_cards([position])[0]
        previous_card["ai_analysis"] = {
            "available": True,
            "provider": "anthropic",
            "label": "Claude claude-sonnet-4-20250514",
            "model": "claude-sonnet-4-20250514",
            "model_label": "Claude claude-sonnet-4-20250514",
            "summary": "Lectura Claude anterior que no debe perderse.",
            "action_label": "Mantener",
            "action_note": "Mantener mientras no cambie la tesis.",
            "confidence_label": "Alta",
            "drivers": ["Retorno esperado positivo"],
            "risks": ["Backtest mejorable"],
            "backtest_note": "El backtest sigue siendo util.",
            "cycle_note": "El ciclo sigue constructivo.",
            "consistency_label": "Alineado",
            "consistency_note": "La IA coincide con el motor cuantitativo.",
            "generated_at": timezone.now().isoformat(),
            "generated_at_label": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
            "usage": {"input_tokens": 320, "output_tokens": 110, "estimated_cost_usd": "0.0031"},
        }
        persist_nightly_analysis_dashboard(
            {
                "history_cards": [previous_card],
                "ibex_universe_cards": [],
                "ibex_universe_summary": {
                    "available": False,
                    "analyzed_count": 0,
                    "buy_alert_count": 0,
                    "sell_alert_count": 0,
                    "watch_alert_count": 0,
                    "registered_count": 0,
                    "registered_owned_count": 0,
                    "registered_watchlist_count": 0,
                    "radar_only_count": 0,
                    "failed_count": 0,
                    "failures": [],
                    "broker_assumption": "",
                    "trade_channel_label": "",
                    "top_pick": None,
                },
                "reference_guide_summary": {
                    "available": True,
                    "workbook_loaded": False,
                    "source_label": "",
                    "tracked_count": 1,
                    "owned_count": 1,
                    "watchlist_count": 0,
                    "guide_only_count": 0,
                },
            },
            [position],
            analysis_date=previous_day,
            agent_provider="anthropic",
            agent_label="Claude claude-sonnet-4-20250514",
            llm_summary={
                "enabled": True,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "completed_count": 1,
                "failed_count": 0,
                "skipped_budget_count": 0,
                "total_count": 1,
                "input_tokens": 320,
                "output_tokens": 110,
                "estimated_cost_usd": "0.0031",
                "monthly_budget_usd": "50.0000",
                "monthly_cost_before_run_usd": "0.0000",
                "monthly_cost_after_run_usd": "0.0031",
                "failures": [],
                "refresh_performed": True,
                "source_analysis_date": previous_day.isoformat(),
                "source_analysis_date_label": previous_day.isoformat(),
            },
        )

        fresh_card = build_equity_history_cards([position])[0]
        dashboard = {
            "history_cards": [fresh_card],
            "owned_history_cards": [fresh_card],
            "ibex_universe_cards": [],
            "ibex_universe_summary": {
                "available": False,
                "analyzed_count": 0,
                "buy_alert_count": 0,
                "sell_alert_count": 0,
                "watch_alert_count": 0,
                "registered_count": 0,
                "registered_owned_count": 0,
                "registered_watchlist_count": 0,
                "radar_only_count": 0,
                "failed_count": 0,
                "failures": [],
                "broker_assumption": "",
                "trade_channel_label": "",
                "top_pick": None,
            },
            "reference_guide_summary": {
                "available": True,
                "workbook_loaded": False,
                "source_label": "",
                "tracked_count": 1,
                "owned_count": 1,
                "watchlist_count": 0,
                "guide_only_count": 0,
            },
        }

        def failing_refresh(current_dashboard, analysis_date):
            current_dashboard["history_cards"][0]["ai_analysis"] = {
                "available": False,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "model_label": "Claude claude-sonnet-4-20250514",
                "note": "Analisis IA no disponible: HTTP 429",
            }
            return {
                "enabled": True,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "completed_count": 0,
                "failed_count": 1,
                "skipped_budget_count": 0,
                "total_count": 1,
                "input_tokens": 0,
                "output_tokens": 0,
                "estimated_cost_usd": "0.0000",
                "monthly_budget_usd": "50.0000",
                "monthly_cost_before_run_usd": "0.0031",
                "monthly_cost_after_run_usd": "0.0031",
                "failures": [{"ticker": "IBE", "company_name": "Iberdrola", "error": "HTTP 429"}],
            }

        with (
            patch("equities.nightly_analysis.sync_all_equities_market_data", return_value=[]),
            patch("equities.nightly_analysis.build_equity_analysis_dashboard", return_value=dashboard),
            patch(
                "equities.nightly_analysis.attach_llm_news_context_to_dashboard",
                return_value={"enabled": True, "signals_count": 1, "items_count": 0, "material_event_count": 0},
            ),
            patch("equities.nightly_analysis.enrich_dashboard_with_ai_analysis", side_effect=failing_refresh) as mocked_llm,
        ):
            run = run_nightly_equity_analysis(
                analysis_date=analysis_day,
                force=False,
            )

        self.assertIsNotNone(run)
        mocked_llm.assert_called_once()
        self.assertEqual(run.summary_data["llm"]["completed_count"], 1)
        self.assertEqual(run.summary_data["llm"]["failed_count"], 0)
        self.assertEqual(run.summary_data["llm"]["refresh_failed_count"], 1)
        self.assertEqual(run.summary_data["llm"]["retained_previous_count"], 1)
        snapshot = run.snapshots.get(scope=EquityNightlyAnalysisSnapshot.Scope.TRACKED, ticker="IBE")
        self.assertTrue(snapshot.analysis_payload["ai_analysis"]["available"])
        self.assertEqual(
            snapshot.analysis_payload["ai_analysis"]["summary"],
            "Lectura Claude anterior que no debe perderse.",
        )
        self.assertIn("respaldo previo", run.status_note)

    def test_management_command_invokes_nightly_analysis_runner(self):
        with (
            patch("equities.management.commands.run_equity_nightly_analysis.run_nightly_equity_analysis") as mocked_runner,
            patch("equities.management.commands.run_equity_nightly_analysis.launch_scheduled_equity_optimization_runs") as mocked_scheduler,
        ):
            mocked_run = type(
                "NightlyRun",
                (),
                {
                    "analysis_date": timezone.localdate(),
                    "snapshots": type("Snapshots", (), {"count": staticmethod(lambda: 3)})(),
                    "agent_label": "Analista nocturno",
                    "agent_provider": "core",
                },
            )()
            mocked_runner.return_value = mocked_run
            mocked_scheduler.return_value = []
            call_command("run_equity_nightly_analysis", "--force")

        mocked_runner.assert_called_once()
        mocked_scheduler.assert_called_once()
        self.assertTrue(mocked_scheduler.call_args.kwargs["run_inline"])

    def test_scheduled_optimization_management_command_invokes_scheduler(self):
        with patch("equities.management.commands.run_scheduled_equity_optimizations.launch_scheduled_equity_optimization_runs") as mocked_scheduler:
            mocked_scheduler.return_value = [
                EquityOptimizationRun(
                    reference_code="OPT-SCHED-001",
                    label="Programada - 12M principal",
                    total_investment=Decimal("100000"),
                    max_company_pct=Decimal("20"),
                    max_total_positions=0,
                    max_sector_positions=0,
                    progress_data={"scheduled_analysis_date": "2026-04-21"},
                ),
                EquityOptimizationRun(
                    reference_code="OPT-SCHED-002",
                    label="Programada - 5A principal",
                    total_investment=Decimal("100000"),
                    max_company_pct=Decimal("20"),
                    max_total_positions=0,
                    max_sector_positions=0,
                    progress_data={"scheduled_analysis_date": "2026-04-21"},
                ),
            ]
            call_command("run_scheduled_equity_optimizations", "--analysis-date", "2026-04-21")

        mocked_scheduler.assert_called_once()
        self.assertTrue(mocked_scheduler.call_args.kwargs["run_inline"])

    def test_scheduled_optimization_management_command_can_use_background_mode(self):
        with patch("equities.management.commands.run_scheduled_equity_optimizations.launch_scheduled_equity_optimization_runs") as mocked_scheduler:
            mocked_scheduler.return_value = [
                EquityOptimizationRun(
                    reference_code="OPT-SCHED-BG-001",
                    label="Programada - 12M principal",
                    total_investment=Decimal("100000"),
                    max_company_pct=Decimal("20"),
                    max_total_positions=0,
                    max_sector_positions=0,
                    progress_data={"scheduled_analysis_date": "2026-04-21"},
                ),
            ]
            call_command("run_scheduled_equity_optimizations", "--analysis-date", "2026-04-21", "--background")

        mocked_scheduler.assert_called_once()
        self.assertFalse(mocked_scheduler.call_args.kwargs["run_inline"])

    def test_build_dashboard_from_nightly_cache_uses_live_position_values(self):
        analysis_day = timezone.localdate()
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        tracked_card = build_equity_history_cards([position])[0]
        tracked_card["news_context"] = {
            "available": True,
            "label": "Contexto adverso",
            "score": Decimal("-2.60"),
            "items_count": 1,
            "top_tags": ["geopolitica"],
            "material_event": True,
            "material_note": "La tension geopolitica obliga a revisar el timing.",
            "note": "La compania queda expuesta a un entorno menos limpio.",
            "top_items": [
                {
                    "title": "Iberdrola cede por tension geopolitica en Europa",
                    "source": "Reuters",
                    "published_label": "2026-04-26",
                    "tone": "negative",
                    "score": Decimal("-2.10"),
                    "tags": ["geopolitica"],
                }
            ],
        }
        tracked_card["expert_consensus"] = {
            "available": True,
            "label": "Consenso mixto",
            "score": Decimal("1.20"),
            "quality_score": Decimal("72.00"),
            "quality_label": "Media",
            "items_count": 1,
            "best_sources": ["JPMorgan"],
            "note": "El consenso acompana pero con mas cautela.",
            "top_items": [
                {
                    "title": "JPMorgan mantiene sobreponderar en Iberdrola",
                    "source": "Reuters",
                    "expert_source": "JPMorgan",
                    "published_label": "2026-04-26",
                    "tone": "positive",
                    "score": Decimal("1.20"),
                    "tags": ["tipos"],
                }
            ],
            "wall_street_signal": {
                "available": True,
                "label": "Wall Street favorable",
                "score": Decimal("1.80"),
                "note": "Wall Street sigue empujando el apetito por riesgo.",
                "top_tags": [],
                "items": [{"published_label": "2026-04-26"}],
                "source_rows": [{"source": "S&P 500"}],
            },
        }
        ibex_position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("34.0000"),
            current_price_per_share=Decimal("34.0000"),
            annual_dividend_income=Decimal("0.50"),
        )
        ibex_card = {
            **tracked_card,
            "position": ibex_position,
            "status_key": "ibex",
            "status_label": "Solo radar",
            "detail_anchor": "",
            "sector_label": "Construccion",
        }
        dashboard = {
            "history_cards": [tracked_card],
            "ibex_universe_cards": [ibex_card],
            "ibex_universe_summary": {
                "available": True,
                "analyzed_count": 1,
                "buy_alert_count": 1,
                "sell_alert_count": 0,
                "watch_alert_count": 0,
                "registered_count": 0,
                "registered_owned_count": 0,
                "registered_watchlist_count": 0,
                "radar_only_count": 1,
                "failed_count": 0,
                "failures": [],
                "broker_assumption": "Interactive Brokers",
                "trade_channel_label": "App",
                "top_pick": {"ticker": "ACS", "company_name": "ACS"},
            },
            "reference_guide_summary": {
                "available": True,
                "workbook_loaded": False,
                "source_label": "",
                "tracked_count": 1,
                "owned_count": 1,
                "watchlist_count": 0,
                "guide_only_count": 0,
            },
        }
        persist_nightly_analysis_dashboard(
            dashboard,
            [position],
            analysis_date=analysis_day,
            agent_provider="core",
            agent_label="Analista nocturno",
        )

        position.current_price_per_share = Decimal("19.5000")
        position.latest_price_date = date(2026, 4, 16)
        position.save(update_fields=["current_price_per_share", "latest_price_date"])

        cached_dashboard = build_dashboard_from_nightly_cache(
            [position],
            include_ibex_universe=True,
        )

        self.assertIsNotNone(cached_dashboard)
        self.assertEqual(cached_dashboard["history_cards"][0]["position"].current_price_per_share, Decimal("19.5000"))
        self.assertTrue(cached_dashboard["history_cards"][0]["projection_12m_chart"]["available"])
        self.assertTrue(cached_dashboard["history_cards"][0]["information_basis"]["available"])
        self.assertIn("Reuters", cached_dashboard["history_cards"][0]["information_basis"]["source_labels"])
        self.assertEqual(cached_dashboard["ibex_universe_cards"][0]["position"].ticker, "ACS")
        self.assertTrue(cached_dashboard["nightly_analysis"]["available"])

    def test_process_equity_optimization_run_uses_nightly_cache_when_available(self):
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-CACHE-001",
            label="Cartera cacheada",
            total_investment=Decimal("100000"),
            max_company_pct=Decimal("20"),
            max_total_positions=8,
            max_sector_positions=1,
            selected_sectors=["Energia"],
        )
        position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("11.0000"),
            current_price_per_share=Decimal("11.0000"),
        )
        card = {
            "position": position,
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Energia",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("8.50"),
                "price_return_pct": Decimal("6.20"),
                "price_low_return_pct": Decimal("-4.00"),
                "price_high_return_pct": Decimal("13.00"),
                "projected_price": Decimal("11.6800"),
                "confidence_label": "Alta",
                "safety_score": Decimal("74.00"),
                "gross_dividend_yield_pct": Decimal("4.10"),
                "net_income_yield_pct": Decimal("3.30"),
                "transaction_drag_pct": Decimal("0.90"),
                "annualized_volatility_pct": Decimal("13.00"),
                "positive_year_ratio_pct": Decimal("68.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-3.00"),
                "max_drawdown_pct": Decimal("-20.00"),
            },
            "cycle_projection_5y": {"available": False},
        }
        dashboard = {
            "optimizer_cards": [card],
            "history_cards": [],
            "ibex_universe_summary": {"analyzed_count": 35},
        }

        with (
            patch("equities.optimization_runs.sync_all_equities_market_data", return_value=[]),
            patch("equities.optimization_runs.build_dashboard_from_nightly_cache", return_value=dashboard) as mocked_cached_dashboard,
            patch("equities.optimization_runs.build_equity_analysis_dashboard") as mocked_live_dashboard,
            patch("equities.optimization_runs.build_news_signal_map", return_value={"IBE": {"label": "Prensa favorable", "score": Decimal("2.10"), "items_count": 2, "items": [], "note": "Buen tono", "available": True, "positive_count": 2, "negative_count": 0, "neutral_count": 0}}),
            patch("equities.optimization_runs.build_report_entries", return_value=[]),
            patch("equities.optimization_runs.build_report_html", return_value="<html>informe</html>"),
            patch("equities.optimization_runs.build_report_pdf_html", return_value="<html>pdf</html>"),
        ):
            process_equity_optimization_run(run.id)

        run.refresh_from_db()
        self.assertEqual(run.status, EquityOptimizationRun.Status.COMPLETED)
        mocked_cached_dashboard.assert_called_once()
        mocked_live_dashboard.assert_not_called()

    def test_optimizer_plan_exposes_purchase_timing_for_each_allocation(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="ACS",
                quote_symbol="ACS.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="ACS",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("40.0000"),
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Infraestructuras",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("14.50"),
                "price_return_pct": Decimal("11.00"),
                "price_low_return_pct": Decimal("-3.00"),
                "price_high_return_pct": Decimal("18.00"),
                "projected_price": Decimal("46.2000"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.00"),
                "net_income_yield_pct": Decimal("2.40"),
                "transaction_drag_pct": Decimal("0.80"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.50"),
                "max_drawdown_pct": Decimal("-18.00"),
                "monthly_path": [
                    {"label": "1M", "projected_date": anchor_day + timedelta(days=30), "projected_price": Decimal("39.1000")},
                    {"label": "3M", "projected_date": anchor_day + timedelta(days=90), "projected_price": Decimal("37.8000")},
                    {"label": "6M", "projected_date": anchor_day + timedelta(days=180), "projected_price": Decimal("42.9000")},
                    {"label": "9M", "projected_date": anchor_day + timedelta(days=270), "projected_price": Decimal("44.8000")},
                    {"label": "12M", "projected_date": anchor_day + timedelta(days=365), "projected_price": Decimal("46.2000")},
                ],
            },
            "cycle_projection_5y": {"available": False},
        }

        plan = build_equity_allocation_plan([card], Decimal("80000"), Decimal("40"))

        self.assertTrue(plan["available"])
        self.assertEqual(len(plan["allocations"]), 1)
        purchase_timing = plan["allocations"][0]["purchase_timing"]
        self.assertTrue(purchase_timing["available"])
        self.assertEqual(purchase_timing["buy_month_number"], 3)
        self.assertTrue(purchase_timing["buy_window_label"])
        self.assertIsNotNone(purchase_timing["buy_price"])
        self.assertEqual(purchase_timing["exit_month_number"], 12)
        self.assertTrue(purchase_timing["exit_window_label"])
        self.assertIsNotNone(purchase_timing["exit_price"])
        self.assertEqual(purchase_timing["holding_months"], 9)
        self.assertIsNotNone(purchase_timing["interval_return_pct"])
        self.assertEqual(plan["top_pick_purchase_timing"]["buy_window_label"], purchase_timing["buy_window_label"])
        self.assertEqual(plan["top_pick_purchase_timing"]["exit_window_label"], purchase_timing["exit_window_label"])

    def test_optimizer_plan_keeps_entry_within_12m_and_allows_later_exit_when_cycle_improves(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="ACS",
                quote_symbol="ACS.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="ACS",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("40.0000"),
                latest_price_date=anchor_day,
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Infraestructuras",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("14.50"),
                "price_return_pct": Decimal("11.00"),
                "price_low_return_pct": Decimal("-3.00"),
                "price_high_return_pct": Decimal("18.00"),
                "projected_price": Decimal("46.2000"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.00"),
                "net_income_yield_pct": Decimal("2.40"),
                "transaction_drag_pct": Decimal("0.80"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.50"),
                "max_drawdown_pct": Decimal("-18.00"),
                "latest_date": anchor_day,
                "latest_price": Decimal("40.0000"),
                "monthly_path": [
                    {"label": "1M", "projected_date": anchor_day + timedelta(days=30), "projected_price": Decimal("39.1000")},
                    {"label": "3M", "projected_date": anchor_day + timedelta(days=90), "projected_price": Decimal("37.8000")},
                    {"label": "6M", "projected_date": anchor_day + timedelta(days=180), "projected_price": Decimal("42.9000")},
                    {"label": "9M", "projected_date": anchor_day + timedelta(days=270), "projected_price": Decimal("44.8000")},
                    {"label": "12M", "projected_date": anchor_day + timedelta(days=365), "projected_price": Decimal("46.2000")},
                ],
            },
            "cycle_projection_5y": {
                "available": True,
                "path": [
                    {"label": "6M", "projected_date": anchor_day + timedelta(days=180), "projected_price": Decimal("38.4000")},
                    {"label": "12M", "projected_date": anchor_day + timedelta(days=365), "projected_price": Decimal("42.5000")},
                    {"label": "24M", "projected_date": anchor_day + timedelta(days=730), "projected_price": Decimal("56.5000")},
                    {"label": "36M", "projected_date": anchor_day + timedelta(days=1095), "projected_price": Decimal("61.0000")},
                ],
            },
        }

        plan = build_equity_allocation_plan([card], Decimal("80000"), Decimal("40"))

        self.assertTrue(plan["available"])
        purchase_timing = plan["allocations"][0]["purchase_timing"]
        self.assertTrue(purchase_timing["available"])
        self.assertLessEqual(purchase_timing["entry_month_number"], 12)
        self.assertGreater(purchase_timing["exit_month_number"], 12)
        self.assertEqual(purchase_timing["analysis_basis_key"], "cycle_5y")
        self.assertEqual(purchase_timing["plan_horizon_months"], 12)
        self.assertEqual(purchase_timing["entry_horizon_months"], 12)
        self.assertGreaterEqual(purchase_timing["exit_horizon_months"], purchase_timing["exit_month_number"])
        self.assertIn("mas alla del ano de entrada", purchase_timing["summary"])

    def test_optimizer_purchase_timing_can_select_non_april_month_from_12m_path(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="ACS",
                quote_symbol="ACS.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="ACS",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("40.0000"),
                latest_price_date=anchor_day,
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Infraestructuras",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("15.00"),
                "price_return_pct": Decimal("12.00"),
                "price_low_return_pct": Decimal("-8.00"),
                "price_high_return_pct": Decimal("18.00"),
                "projected_price": Decimal("46.5000"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.00"),
                "net_income_yield_pct": Decimal("2.40"),
                "transaction_drag_pct": Decimal("0.80"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.50"),
                "max_drawdown_pct": Decimal("-18.00"),
                "latest_date": anchor_day,
                "latest_price": Decimal("40.0000"),
                "monthly_path": [
                    {"label": "1M", "projected_date": add_calendar_months(anchor_day, 1), "projected_price": Decimal("39.2000")},
                    {"label": "4M", "projected_date": add_calendar_months(anchor_day, 4), "projected_price": Decimal("36.0000")},
                    {"label": "6M", "projected_date": add_calendar_months(anchor_day, 6), "projected_price": Decimal("33.0000")},
                    {"label": "9M", "projected_date": add_calendar_months(anchor_day, 9), "projected_price": Decimal("42.5000")},
                    {"label": "12M", "projected_date": add_calendar_months(anchor_day, 12), "projected_price": Decimal("46.5000")},
                ],
            },
            "cycle_projection_5y": {"available": False},
        }

        purchase_timing = build_candidate_purchase_timing_plan(card)

        self.assertTrue(purchase_timing["available"])
        self.assertEqual(purchase_timing["analysis_basis_key"], "projection_12m")
        self.assertEqual(purchase_timing["buy_month_number"], 6)
        self.assertEqual(purchase_timing["buy_date"], add_calendar_months(anchor_day, 6))
        self.assertEqual(purchase_timing["buy_date"].month, 10)
        self.assertIn("octubre 2026", purchase_timing["buy_window_label"])

    def test_optimizer_purchase_timing_can_select_non_april_month_from_cycle_path(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="ACS",
                quote_symbol="ACS.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="ACS",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("40.0000"),
                latest_price_date=anchor_day,
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Infraestructuras",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("8.00"),
                "price_return_pct": Decimal("6.00"),
                "price_low_return_pct": Decimal("-5.00"),
                "price_high_return_pct": Decimal("12.00"),
                "projected_price": Decimal("42.4000"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.00"),
                "net_income_yield_pct": Decimal("2.40"),
                "transaction_drag_pct": Decimal("0.80"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.50"),
                "max_drawdown_pct": Decimal("-18.00"),
                "latest_date": anchor_day,
                "latest_price": Decimal("40.0000"),
                "monthly_path": [
                    {"label": "1M", "projected_date": add_calendar_months(anchor_day, 1), "projected_price": Decimal("40.4000")},
                    {"label": "6M", "projected_date": add_calendar_months(anchor_day, 6), "projected_price": Decimal("41.0000")},
                    {"label": "12M", "projected_date": add_calendar_months(anchor_day, 12), "projected_price": Decimal("42.4000")},
                ],
            },
            "cycle_projection_5y": {
                "available": True,
                "path": [
                    {"label": "4M", "projected_date": add_calendar_months(anchor_day, 4), "projected_price": Decimal("36.5000")},
                    {"label": "8M", "projected_date": add_calendar_months(anchor_day, 8), "projected_price": Decimal("31.0000")},
                    {"label": "20M", "projected_date": add_calendar_months(anchor_day, 20), "projected_price": Decimal("49.5000")},
                    {"label": "30M", "projected_date": add_calendar_months(anchor_day, 30), "projected_price": Decimal("58.0000")},
                ],
            },
        }

        purchase_timing = build_candidate_purchase_timing_plan(card, strategy_mode="5y_primary")

        self.assertTrue(purchase_timing["available"])
        self.assertEqual(purchase_timing["analysis_basis_key"], "cycle_5y")
        self.assertLessEqual(purchase_timing["buy_month_number"], 8)
        self.assertGreaterEqual(purchase_timing["buy_month_number"], 4)
        self.assertNotEqual(purchase_timing["buy_date"].month, 4)
        self.assertNotIn("abril 2027", purchase_timing["buy_window_label"])

    def test_optimizer_purchase_timing_uses_detailed_12m_entry_shape_inside_5y_cycle(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="IBE",
                quote_symbol="IBE.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Iberdrola",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("20.0000"),
                latest_price_date=anchor_day,
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Electrica",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("11.00"),
                "price_return_pct": Decimal("9.00"),
                "price_low_return_pct": Decimal("-6.00"),
                "price_high_return_pct": Decimal("15.00"),
                "projected_price": Decimal("21.8000"),
                "confidence_label": "Alta",
                "safety_score": Decimal("78.00"),
                "gross_dividend_yield_pct": Decimal("4.00"),
                "net_income_yield_pct": Decimal("3.10"),
                "transaction_drag_pct": Decimal("0.70"),
                "annualized_volatility_pct": Decimal("10.00"),
                "positive_year_ratio_pct": Decimal("74.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-1.50"),
                "max_drawdown_pct": Decimal("-15.00"),
                "latest_date": anchor_day,
                "latest_price": Decimal("20.0000"),
                "monthly_path": [
                    {"label": "1M", "projected_date": add_calendar_months(anchor_day, 1), "projected_price": Decimal("19.7000")},
                    {"label": "3M", "projected_date": add_calendar_months(anchor_day, 3), "projected_price": Decimal("18.9000")},
                    {"label": "5M", "projected_date": add_calendar_months(anchor_day, 5), "projected_price": Decimal("17.9000")},
                    {"label": "7M", "projected_date": add_calendar_months(anchor_day, 7), "projected_price": Decimal("18.6000")},
                    {"label": "9M", "projected_date": add_calendar_months(anchor_day, 9), "projected_price": Decimal("20.1000")},
                    {"label": "12M", "projected_date": add_calendar_months(anchor_day, 12), "projected_price": Decimal("21.8000")},
                ],
            },
            "cycle_projection_5y": {
                "available": True,
                "path": [
                    {"label": "6M", "projected_date": add_calendar_months(anchor_day, 6), "projected_price": Decimal("19.0000")},
                    {"label": "1A", "projected_date": add_calendar_months(anchor_day, 12), "projected_price": Decimal("18.2000")},
                    {"label": "2A", "projected_date": add_calendar_months(anchor_day, 24), "projected_price": Decimal("25.0000")},
                    {"label": "3A", "projected_date": add_calendar_months(anchor_day, 36), "projected_price": Decimal("28.5000")},
                    {"label": "5A", "projected_date": add_calendar_months(anchor_day, 60), "projected_price": Decimal("34.0000")},
                ],
            },
        }

        purchase_timing = build_candidate_purchase_timing_plan(card, strategy_mode="5y_primary")

        self.assertTrue(purchase_timing["available"])
        self.assertEqual(purchase_timing["analysis_basis_key"], "cycle_5y")
        self.assertLess(purchase_timing["buy_month_number"], 12)
        self.assertNotEqual(purchase_timing["buy_date"].month, 4)
        self.assertNotIn("abril 2027", purchase_timing["buy_window_label"])
        self.assertIn("entrada afinada", purchase_timing["analysis_basis_label"].lower())

    def test_optimizer_plan_exposes_annualized_return_for_trade_window(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="ACS",
                quote_symbol="ACS.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="ACS",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("40.0000"),
                latest_price_date=anchor_day,
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Infraestructuras",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("84.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("12.00"),
                "price_return_pct": Decimal("8.00"),
                "price_low_return_pct": Decimal("-4.00"),
                "price_high_return_pct": Decimal("18.00"),
                "projected_price": Decimal("40.5000"),
                "confidence_label": "Alta",
                "safety_score": Decimal("76.00"),
                "gross_dividend_yield_pct": Decimal("3.00"),
                "net_income_yield_pct": Decimal("2.40"),
                "transaction_drag_pct": Decimal("0.80"),
                "annualized_volatility_pct": Decimal("12.00"),
                "positive_year_ratio_pct": Decimal("70.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.50"),
                "max_drawdown_pct": Decimal("-18.00"),
                "latest_date": anchor_day,
                "latest_price": Decimal("40.0000"),
                "monthly_path": [
                    {"label": "1M", "projected_date": anchor_day + timedelta(days=30), "projected_price": Decimal("39.1000")},
                    {"label": "3M", "projected_date": anchor_day + timedelta(days=90), "projected_price": Decimal("37.8000")},
                    {"label": "6M", "projected_date": anchor_day + timedelta(days=180), "projected_price": Decimal("42.9000")},
                    {"label": "9M", "projected_date": anchor_day + timedelta(days=270), "projected_price": Decimal("44.8000")},
                    {"label": "12M", "projected_date": anchor_day + timedelta(days=365), "projected_price": Decimal("46.2000")},
                ],
            },
            "cycle_projection_5y": {"available": False},
        }

        plan = build_equity_allocation_plan([card], Decimal("80000"), Decimal("40"))

        self.assertTrue(plan["available"])
        purchase_timing = plan["allocations"][0]["purchase_timing"]
        self.assertTrue(purchase_timing["available"])
        self.assertEqual(purchase_timing["exit_month_number"], 12)
        self.assertIsNotNone(purchase_timing["holding_annualized_return_pct"])
        self.assertIn("anualizado", purchase_timing["summary"])
        self.assertIsNotNone(plan["weighted_holding_annualized_return_pct"])
        self.assertGreaterEqual(
            plan["weighted_holding_annualized_return_pct"],
            plan["target_holding_annualized_return_pct"],
        )
        self.assertEqual(plan["allocations_with_timing_count"], 1)
        self.assertEqual(plan["allocations_meeting_target_count"], 1)

    def test_optimizer_plan_keeps_cash_when_tactical_window_does_not_reach_20_percent_annualized_target(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="ENG",
                quote_symbol="ENG.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Enagas",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("20.0000"),
                latest_price_date=anchor_day,
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Energia",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("82.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("11.50"),
                "price_return_pct": Decimal("9.80"),
                "price_low_return_pct": Decimal("-2.00"),
                "price_high_return_pct": Decimal("14.00"),
                "projected_price": Decimal("21.9600"),
                "latest_price": Decimal("20.0000"),
                "latest_date": anchor_day,
                "confidence_label": "Alta",
                "safety_score": Decimal("78.00"),
                "gross_dividend_yield_pct": Decimal("3.20"),
                "net_income_yield_pct": Decimal("2.30"),
                "transaction_drag_pct": Decimal("0.60"),
                "annualized_volatility_pct": Decimal("9.00"),
                "positive_year_ratio_pct": Decimal("69.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-2.00"),
                "max_drawdown_pct": Decimal("-16.00"),
                "monthly_path": [
                    {"label": "3M", "projected_date": anchor_day + timedelta(days=90), "projected_price": Decimal("19.8000")},
                    {"label": "6M", "projected_date": anchor_day + timedelta(days=180), "projected_price": Decimal("20.4000")},
                    {"label": "9M", "projected_date": anchor_day + timedelta(days=270), "projected_price": Decimal("21.1000")},
                    {"label": "12M", "projected_date": anchor_day + timedelta(days=365), "projected_price": Decimal("21.9600")},
                ],
            },
            "cycle_projection_5y": {"available": False},
        }

        plan = build_equity_allocation_plan([card], Decimal("50000"), Decimal("100"))

        self.assertFalse(plan["available"])
        self.assertIn("20 % anualizado", plan["reason"])

    def test_optimizer_plan_keeps_cash_when_candidate_is_too_risky_for_conservative_profile(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="ANA",
                quote_symbol="ANA.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Acciona",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("40.0000"),
                latest_price_date=anchor_day,
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Infraestructuras",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("83.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("27.40"),
                "price_return_pct": Decimal("25.00"),
                "price_low_return_pct": Decimal("-18.00"),
                "price_high_return_pct": Decimal("34.00"),
                "projected_price": Decimal("50.0000"),
                "latest_price": Decimal("40.0000"),
                "latest_date": anchor_day,
                "confidence_label": "Alta",
                "safety_score": Decimal("60.00"),
                "gross_dividend_yield_pct": Decimal("2.40"),
                "net_income_yield_pct": Decimal("1.80"),
                "transaction_drag_pct": Decimal("0.60"),
                "annualized_volatility_pct": Decimal("31.00"),
                "positive_year_ratio_pct": Decimal("66.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-7.00"),
                "max_drawdown_pct": Decimal("-29.00"),
                "monthly_path": [
                    {"label": "1M", "projected_date": anchor_day + timedelta(days=30), "projected_price": Decimal("39.2000")},
                    {"label": "3M", "projected_date": anchor_day + timedelta(days=90), "projected_price": Decimal("36.0000")},
                    {"label": "6M", "projected_date": anchor_day + timedelta(days=180), "projected_price": Decimal("41.5000")},
                    {"label": "9M", "projected_date": anchor_day + timedelta(days=270), "projected_price": Decimal("45.0000")},
                    {"label": "12M", "projected_date": anchor_day + timedelta(days=365), "projected_price": Decimal("50.0000")},
                ],
            },
            "cycle_projection_5y": {"available": False},
        }

        plan = build_equity_allocation_plan([card], Decimal("50000"), Decimal("100"))

        self.assertFalse(plan["available"])
        self.assertIn("perfil adverso al riesgo", plan["reason"].lower())

    def test_optimizer_plan_exposes_conservative_profile_when_candidate_is_defensive_enough(self):
        anchor_day = date(2026, 4, 25)
        card = {
            "position": EquityPosition(
                position_kind=EquityPosition.PositionKind.WATCHLIST,
                broker="Interactive Brokers",
                ticker="IBE",
                quote_symbol="IBE.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Iberdrola",
                shares=Decimal("0"),
                average_cost_per_share=Decimal("0"),
                current_price_per_share=Decimal("15.0000"),
                latest_price_date=anchor_day,
            ),
            "status_key": "ibex",
            "status_label": "Radar IBEX",
            "sector_label": "Utilities",
            "reference_label": "IBEX 35",
            "trade_alert": {"label": "Comprar", "tone": "buy", "note": "Tendencia favorable."},
            "projection_reliability": {"label": "Alta", "score": Decimal("86.00")},
            "projection": {
                "available": True,
                "base_return_pct": Decimal("24.20"),
                "price_return_pct": Decimal("22.00"),
                "price_low_return_pct": Decimal("-6.00"),
                "price_high_return_pct": Decimal("29.00"),
                "projected_price": Decimal("18.3000"),
                "latest_price": Decimal("15.0000"),
                "latest_date": anchor_day,
                "confidence_label": "Alta",
                "safety_score": Decimal("84.00"),
                "gross_dividend_yield_pct": Decimal("2.80"),
                "net_income_yield_pct": Decimal("2.10"),
                "transaction_drag_pct": Decimal("0.50"),
                "annualized_volatility_pct": Decimal("11.50"),
                "positive_year_ratio_pct": Decimal("74.00"),
                "years_covered": Decimal("10.00"),
                "cycle_phase": "Expansion",
                "current_drawdown_pct": Decimal("-3.00"),
                "max_drawdown_pct": Decimal("-15.00"),
                "monthly_path": [
                    {"label": "1M", "projected_date": anchor_day + timedelta(days=30), "projected_price": Decimal("14.9000")},
                    {"label": "3M", "projected_date": anchor_day + timedelta(days=90), "projected_price": Decimal("14.2000")},
                    {"label": "6M", "projected_date": anchor_day + timedelta(days=180), "projected_price": Decimal("15.8000")},
                    {"label": "9M", "projected_date": anchor_day + timedelta(days=270), "projected_price": Decimal("16.9000")},
                    {"label": "12M", "projected_date": anchor_day + timedelta(days=365), "projected_price": Decimal("18.3000")},
                ],
            },
            "cycle_projection_5y": {"available": False},
        }

        plan = build_equity_allocation_plan([card], Decimal("50000"), Decimal("100"))

        self.assertTrue(plan["available"])
        self.assertEqual(plan["risk_profile_label"], "Adverso al riesgo")
        self.assertIn("alerta Comprar", plan["conservative_profile_note"])
        self.assertEqual(plan["weighted_conservative_profile_compliance_pct"], Decimal("100.00"))
        self.assertTrue(plan["allocations"][0]["passes_conservative_profile"])

@override_settings(EQUITIES_AUTO_SYNC_ON_VIEW=False, EQUITIES_IBEX_UNIVERSE_ANALYSIS=False)
@override_settings(EQUITIES_FETCH_FUNDAMENTALS=False)
class EquitiesViewTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="equity-owner",
            password="StrongPass123!",
        )
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])
        self.client.force_login(self.user)

    def tearDown(self):
        load_ibex_reference_workbook_snapshot.cache_clear()
        clear_market_data_caches()
        super().tearDown()

    def test_equities_page_shows_nightly_analysis_status_in_hero(self):
        analysis_day = timezone.localdate()
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        tracked_card = build_equity_history_cards([position])[0]
        tracked_card["ai_analysis"] = {
            "available": True,
            "provider": "anthropic",
            "label": "Claude claude-sonnet-4-20250514",
            "model": "claude-sonnet-4-20250514",
            "model_label": "Claude claude-sonnet-4-20250514",
            "summary": "La tesis nocturna sigue favoreciendo mantener por retorno esperado y ciclo aun estable.",
            "action_label": "Mantener",
            "action_note": "Mantener si la fiabilidad no cae.",
            "confidence_label": "Media",
            "drivers": ["Retorno 12M positivo"],
            "risks": ["El backtest aun no es de nivel alto"],
            "backtest_note": "La validacion historica es util pero no impecable.",
            "cycle_note": "El ciclo 5A sigue acompasado con posibles tramos de correccion.",
            "consistency_label": "Mixto",
            "consistency_note": "La IA matiza la alerta cuantitativa.",
            "generated_at": timezone.now().isoformat(),
            "generated_at_label": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
            "usage": {"input_tokens": 400, "output_tokens": 120, "estimated_cost_usd": "0.0030"},
        }
        persist_nightly_analysis_dashboard(
            {
                "history_cards": [tracked_card],
                "ibex_universe_cards": [],
                "ibex_universe_summary": {
                    "available": False,
                    "analyzed_count": 0,
                    "buy_alert_count": 0,
                    "sell_alert_count": 0,
                    "watch_alert_count": 0,
                    "registered_count": 0,
                    "registered_owned_count": 0,
                    "registered_watchlist_count": 0,
                    "radar_only_count": 0,
                    "failed_count": 0,
                    "failures": [],
                    "broker_assumption": "",
                    "trade_channel_label": "",
                    "top_pick": None,
                },
                "reference_guide_summary": {
                    "available": True,
                    "workbook_loaded": False,
                    "source_label": "",
                    "tracked_count": 1,
                    "owned_count": 1,
                    "watchlist_count": 0,
                    "guide_only_count": 0,
                },
            },
            [position],
            analysis_date=analysis_day,
            agent_provider="anthropic",
            agent_label="Claude claude-sonnet-4-20250514",
            llm_summary={
                "enabled": True,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "completed_count": 1,
                "failed_count": 0,
                "skipped_budget_count": 0,
                "total_count": 1,
                "input_tokens": 400,
                "output_tokens": 120,
                "estimated_cost_usd": "0.0030",
                "monthly_budget_usd": "50.0000",
                "monthly_cost_before_run_usd": "0.0000",
                "monthly_cost_after_run_usd": "0.0030",
                "failures": [],
            },
        )

        with patch(
            "equities.services.build_owned_cycle_trade_timing_plan",
            return_value={
                "available": True,
                "mode": "sale_reentry",
                "sale_month_number": 12,
                "sale_date": date(2027, 4, 17),
                "sale_window_label": "abril 2027 (mes 12)",
                "signal_value_pct": Decimal("-0.13"),
                "summary": "Salida tactica en abril 2027.",
            },
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertContains(response, "Analisis exhaustivo OK")
        self.assertContains(response, "Ultima ejecucion")
        self.assertContains(response, "Claude claude-sonnet-4-20250514")
        self.assertContains(response, "IA 1/1")
        self.assertContains(response, "Primera venta IBE abril 2027 (mes 12)")
        self.assertContains(response, "La primera salida tactica sugerida hoy seria Iberdrola en abril 2027 (mes 12).")

    def test_equities_page_reused_ai_status_does_not_label_pending_as_api_failures(self):
        analysis_day = timezone.localdate()
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        tracked_card = build_equity_history_cards([position])[0]
        tracked_card["ai_analysis"] = {
            "available": True,
            "provider": "anthropic",
            "label": "Claude claude-sonnet-4-20250514",
            "model": "claude-sonnet-4-20250514",
            "model_label": "Claude claude-sonnet-4-20250514",
            "summary": "Lectura Claude reutilizada desde la ultima actualizacion valida.",
            "action_label": "Mantener",
            "action_note": "Mantener si no cambia la tesis.",
            "confidence_label": "Media",
            "drivers": ["Retorno esperado positivo"],
            "risks": ["Backtest mejorable"],
            "backtest_note": "La validacion historica sigue siendo util.",
            "cycle_note": "El ciclo largo sigue acompasado.",
            "consistency_label": "Alineado",
            "consistency_note": "La IA coincide con el motor cuantitativo.",
            "generated_at": timezone.now().isoformat(),
            "generated_at_label": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
            "usage": {"input_tokens": 400, "output_tokens": 120, "estimated_cost_usd": "0.0030"},
        }
        persist_nightly_analysis_dashboard(
            {
                "history_cards": [tracked_card],
                "ibex_universe_cards": [],
                "ibex_universe_summary": {
                    "available": False,
                    "analyzed_count": 0,
                    "buy_alert_count": 0,
                    "sell_alert_count": 0,
                    "watch_alert_count": 0,
                    "registered_count": 0,
                    "registered_owned_count": 0,
                    "registered_watchlist_count": 0,
                    "radar_only_count": 0,
                    "failed_count": 0,
                    "failures": [],
                    "broker_assumption": "",
                    "trade_channel_label": "",
                    "top_pick": None,
                },
                "reference_guide_summary": {
                    "available": True,
                    "workbook_loaded": False,
                    "source_label": "",
                    "tracked_count": 1,
                    "owned_count": 1,
                    "watchlist_count": 0,
                    "guide_only_count": 0,
                },
            },
            [position],
            analysis_date=analysis_day,
            agent_provider="anthropic",
            agent_label="Claude claude-sonnet-4-20250514",
            llm_summary={
                "enabled": True,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "completed_count": 1,
                "failed_count": 11,
                "refresh_failed_count": 0,
                "retained_previous_count": 1,
                "pending_count": 11,
                "skipped_budget_count": 0,
                "total_count": 12,
                "input_tokens": 0,
                "output_tokens": 0,
                "estimated_cost_usd": "0.0000",
                "monthly_budget_usd": "50.0000",
                "monthly_cost_before_run_usd": "0.0030",
                "monthly_cost_after_run_usd": "0.0030",
                "failures": [],
                "reused": True,
                "refresh_performed": False,
                "source_analysis_date": "2026-04-16",
                "source_analysis_date_label": "2026-04-16",
                "next_refresh_date": "2026-04-21",
                "next_refresh_date_label": "2026-04-21",
            },
        )

        response = self.client.get(reverse("equities:list"))

        self.assertContains(response, "Quedan 11 valores sin lectura IA previa.")
        self.assertNotContains(response, "Hubo 11 fallo(s) de API.")

    def test_ibex_detail_view_uses_cached_nightly_snapshot(self):
        analysis_day = timezone.localdate()
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position)
        tracked_card = build_equity_history_cards([position])[0]
        ibex_position = EquityPosition(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("34.0000"),
            current_price_per_share=Decimal("34.0000"),
            annual_dividend_income=Decimal("0.50"),
        )
        ibex_card = {
            **tracked_card,
            "position": ibex_position,
            "status_key": "ibex",
            "status_label": "Solo radar",
            "detail_anchor": "",
            "sector_label": "Construccion",
            "ai_analysis": {
                "available": True,
                "provider": "anthropic",
                "label": "Claude claude-sonnet-4-20250514",
                "model": "claude-sonnet-4-20250514",
                "model_label": "Claude claude-sonnet-4-20250514",
                "summary": "La lectura IA ve una oportunidad razonable mientras la validacion historica no empeore.",
                "action_label": "Comprar",
                "action_note": "Comprar si se mantiene la pendiente relativa.",
                "confidence_label": "Media",
                "drivers": ["Retorno esperado en positivo"],
                "risks": ["La fiabilidad no es plena"],
                "backtest_note": "El modelo acierta mas de lo que falla, pero no con precision alta.",
                "cycle_note": "El ciclo 5A sigue constructivo con correcciones intermedias.",
                "consistency_label": "Alineado",
                "consistency_note": "La IA coincide con el motor cuantitativo.",
                "generated_at": timezone.now().isoformat(),
                "generated_at_label": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
                "usage": {"input_tokens": 300, "output_tokens": 110, "estimated_cost_usd": "0.0020"},
            },
        }
        persist_nightly_analysis_dashboard(
            {
                "history_cards": [tracked_card],
                "ibex_universe_cards": [ibex_card],
                "ibex_universe_summary": {
                    "available": True,
                    "analyzed_count": 1,
                    "buy_alert_count": 1,
                    "sell_alert_count": 0,
                    "watch_alert_count": 0,
                    "registered_count": 0,
                    "registered_owned_count": 0,
                    "registered_watchlist_count": 0,
                    "radar_only_count": 1,
                    "failed_count": 0,
                    "failures": [],
                    "broker_assumption": "Interactive Brokers",
                    "trade_channel_label": "App",
                    "top_pick": {"ticker": "ACS", "company_name": "ACS"},
                },
                "reference_guide_summary": {
                    "available": True,
                    "workbook_loaded": False,
                    "source_label": "",
                    "tracked_count": 1,
                    "owned_count": 1,
                    "watchlist_count": 0,
                    "guide_only_count": 0,
                },
            },
            [position],
            analysis_date=analysis_day,
            agent_provider="core",
            agent_label="Analista nocturno",
        )

        company = {
            "ticker": "ACS",
            "company_name": "ACS",
            "quote_symbol": "ACS.MC",
            "sector": "Construccion",
        }
        workbook_snapshot = {
            "available": False,
            "path": "",
            "companies": [],
        }

        with (
            patch("equities.views.find_ibex_universe_company", return_value=(company, workbook_snapshot)),
            patch("equities.views.build_ibex_universe_card", side_effect=AssertionError("no deberia recalcular en vivo")),
        ):
            response = self.client.get(reverse("equities:ibex_detail", kwargs={"ticker": "ACS"}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ACS")
        self.assertContains(response, "Lectura IA nocturna")
        self.assertContains(response, "La lectura IA ve una oportunidad razonable")

    def test_can_create_equity_position_from_page_form(self):
        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "create_position",
                "position_kind": EquityPosition.PositionKind.OWNED,
                "ownership_category": AssetOwnershipCategory.XIMO,
                "broker": "Interactive Brokers",
                "ticker": "ibe",
                "company_name": "Iberdrola",
                "quote_symbol": "ibe.mc",
                "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
                "benchmark_symbol": "^ibex",
                "benchmark_name": "IBEX 35",
                "shares": "125,5000",
                "average_cost_per_share": "10,2500",
                "current_price_per_share": "",
                "annual_dividend_income": "72,50",
                "annual_maintenance_cost": "18,75",
                "notes": "Posicion principal",
            },
        )

        self.assertRedirects(response, reverse("equities:list"))
        position = EquityPosition.objects.get(ticker="IBE", broker="Interactive Brokers")
        self.assertEqual(position.company_name, "Iberdrola")
        self.assertEqual(position.quote_symbol, "IBE.MC")
        self.assertEqual(position.benchmark_symbol, "^IBEX")
        self.assertEqual(position.ownership_category, AssetOwnershipCategory.XIMO)
        self.assertEqual(position.position_kind, EquityPosition.PositionKind.OWNED)
        self.assertEqual(position.trade_channel, EquityPosition.TradeChannel.APP)
        self.assertEqual(position.shares, Decimal("125.5000"))
        self.assertEqual(position.average_cost_per_share, Decimal("10.2500"))
        self.assertEqual(position.current_price_per_share, Decimal("10.2500"))
        self.assertEqual(position.annual_dividend_income, Decimal("72.50"))
        self.assertEqual(position.annual_maintenance_cost, Decimal("18.75"))

    def test_create_owned_position_captures_purchase_forecast_baseline(self):
        analysis_day = date(2026, 4, 17)
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=analysis_day,
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        cached_position = EquityPosition(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        EquityNightlyAnalysisSnapshot.objects.create(
            run=run,
            analysis_date=analysis_day,
            scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
            analysis_key="ibex:IBE",
            ticker="IBE",
            quote_symbol="IBE.MC",
            company_name="Iberdrola",
            status_key="ibex",
            sector_label="Utilities",
            agent_provider="core",
            analysis_payload=serialize_cached_value(
                {
                    "position": cached_position,
                    "reference_label": "IBEX 35",
                    "projection": {
                        "available": True,
                        "projected_price": Decimal("11.2500"),
                        "base_return_pct": Decimal("12.50"),
                        "safety_score": Decimal("68.00"),
                    },
                    "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
                    "trade_alert": {"label": "Comprar"},
                    "cycle_projection_5y": {
                        "available": True,
                        "path": [
                            {"label": "1A", "projected_price": Decimal("11.0000")},
                            {"label": "2A", "projected_price": Decimal("12.3750")},
                            {"label": "3A", "projected_price": Decimal("13.6125")},
                            {"label": "4A", "projected_price": Decimal("14.9738")},
                            {"label": "5A", "projected_price": Decimal("16.4712")},
                        ],
                    },
                }
            ),
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "create_position",
                "position_kind": EquityPosition.PositionKind.OWNED,
                "ownership_category": AssetOwnershipCategory.XIMO,
                "broker": "Interactive Brokers",
                "ticker": "IBE",
                "company_name": "Iberdrola",
                "quote_symbol": "IBE.MC",
                "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
                "benchmark_symbol": "^IBEX",
                "benchmark_name": "IBEX 35",
                "opened_on": "2026-04-17",
                "shares": "125,5000",
                "average_cost_per_share": "10,2500",
                "current_price_per_share": "",
                "annual_dividend_income": "72,50",
                "annual_maintenance_cost": "18,75",
                "notes": "Posicion principal",
            },
        )

        self.assertRedirects(response, reverse("equities:list"))
        position = EquityPosition.objects.get(ticker="IBE", broker="Interactive Brokers")
        baseline = EquityPurchaseForecastBaseline.objects.get(position=position)
        self.assertEqual(baseline.source_analysis_date, analysis_day)
        self.assertEqual(baseline.projected_return_pct_1y, Decimal("12.50"))
        self.assertEqual(baseline.projected_price_5y, Decimal("16.4712"))

    def test_import_monica_equity_positions_creates_positions_with_today_cost_basis(self):
        analysis_date = date(2026, 4, 19)

        call_command("import_monica_equity_positions", "--as-of", analysis_date.isoformat(), "--broker", "Broker Monica")

        positions = EquityPosition.objects.filter(
            ownership_category=AssetOwnershipCategory.MONICA,
            broker="Broker Monica",
        ).order_by("ticker")

        self.assertEqual(positions.count(), len(MONICA_EQUITY_POSITIONS))
        self.assertTrue(
            positions.filter(
                ticker="SAB",
                company_name="Banco de Sabadell",
                quote_symbol="SAB.MC",
                opened_on=analysis_date,
                shares=Decimal("12552.0000"),
                average_cost_per_share=Decimal("3.3650"),
                current_price_per_share=Decimal("3.3650"),
            ).exists()
        )
        self.assertTrue(
            positions.filter(
                ticker="REP",
                quote_symbol="REP.MC",
                opened_on=analysis_date,
                average_cost_per_share=Decimal("19.7200"),
                current_price_per_share=Decimal("19.7200"),
            ).exists()
        )
        self.assertTrue(
            positions.filter(
                ticker="SAN",
                quote_symbol="SAN.MC",
                opened_on=analysis_date,
                average_cost_per_share=Decimal("11.0420"),
                current_price_per_share=Decimal("11.0420"),
            ).exists()
        )
        self.assertTrue(
            positions.filter(
                ticker="ELE",
                quote_symbol="ELE.MC",
                opened_on=analysis_date,
                average_cost_per_share=Decimal("36.8800"),
                current_price_per_share=Decimal("36.8800"),
            ).exists()
        )
        cuotas_cam = positions.get(ticker="CAM")
        self.assertEqual(cuotas_cam.company_name, "Cuotas CAM")
        self.assertEqual(cuotas_cam.opened_on, analysis_date)
        self.assertEqual(cuotas_cam.average_cost_per_share, Decimal("0.0000"))
        self.assertEqual(cuotas_cam.current_price_per_share, Decimal("0.0000"))
        self.assertIn("ISIN: ES0114400007", cuotas_cam.notes)
        self.assertEqual(
            EquityPriceHistory.objects.filter(
                position__broker="Broker Monica",
                position__ownership_category=AssetOwnershipCategory.MONICA,
            ).count(),
            4,
        )
        self.assertEqual(
            EquityTicketSnapshot.objects.filter(
                position__broker="Broker Monica",
                position__ownership_category=AssetOwnershipCategory.MONICA,
                snapshot_date=analysis_date,
            ).count(),
            len(MONICA_EQUITY_POSITIONS),
        )

    def test_import_monica_equity_positions_updates_existing_rows_without_duplicates(self):
        analysis_date = date(2026, 4, 19)
        existing = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.MONICA,
            broker="Broker Monica",
            ticker="SAN",
            quote_symbol="SAN.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Banco Santander antiguo",
            trade_channel=EquityPosition.TradeChannel.APP,
            opened_on=date(2025, 1, 1),
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("1.0000"),
            current_price_per_share=Decimal("1.0000"),
            notes="Antigua",
        )
        EquityTicketSnapshot.objects.create(
            position=existing,
            snapshot_date=date(2026, 4, 18),
            invested_amount=Decimal("1.00"),
            current_value=Decimal("1.00"),
        )
        EquityPurchaseForecastBaseline.objects.create(
            position=existing,
            source_analysis_date=date(2026, 4, 18),
            baseline_date=date(2026, 4, 18),
            analysis_scope="ibex",
            analysis_key="ibex:SAN",
            baseline_price=Decimal("1.0000"),
            projected_price_1y=Decimal("1.1000"),
            projected_price_5y=Decimal("1.5000"),
            projected_return_pct_1y=Decimal("10.00"),
            projected_return_pct_5y=Decimal("50.00"),
            projected_path_5y=[
                {"label": "1A", "projected_price": "1.1000", "projected_date": "2027-04-18"},
                {"label": "5A", "projected_price": "1.5000", "projected_date": "2031-04-18"},
            ],
        )

        call_command("import_monica_equity_positions", "--as-of", analysis_date.isoformat(), "--broker", "Broker Monica")

        self.assertEqual(
            EquityPosition.objects.filter(
                ownership_category=AssetOwnershipCategory.MONICA,
                broker="Broker Monica",
                ticker="SAN",
            ).count(),
            1,
        )
        existing.refresh_from_db()
        self.assertEqual(existing.company_name, "Banco Santander")
        self.assertEqual(existing.opened_on, analysis_date)
        self.assertEqual(existing.shares, Decimal("2550.0000"))
        self.assertEqual(existing.average_cost_per_share, Decimal("11.0420"))
        self.assertEqual(existing.current_price_per_share, Decimal("11.0420"))
        self.assertEqual(existing.ticket_snapshots.count(), 2)
        self.assertTrue(existing.ticket_snapshots.filter(snapshot_date=date(2026, 4, 18)).exists())
        self.assertTrue(existing.ticket_snapshots.filter(snapshot_date=analysis_date).exists())
        baseline = EquityPurchaseForecastBaseline.objects.get(position=existing)
        self.assertEqual(baseline.source_analysis_date, date(2026, 4, 18))
        self.assertEqual(baseline.projected_price_5y, Decimal("1.5000"))

    def test_can_create_equity_position_by_only_typing_indra(self):
        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "create_position",
                "position_kind": EquityPosition.PositionKind.OWNED,
                "ownership_category": AssetOwnershipCategory.XIMO,
                "broker": "Interactive Brokers",
                "ticker": "",
                "company_name": "Indra",
                "quote_symbol": "",
                "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
                "benchmark_symbol": "",
                "benchmark_name": "",
                "shares": "20",
                "average_cost_per_share": "18,5000",
                "current_price_per_share": "",
                "annual_dividend_income": "0",
                "annual_maintenance_cost": "5,00",
                "notes": "Alta rapida",
            },
        )

        self.assertRedirects(response, reverse("equities:list"))
        position = EquityPosition.objects.get(ticker="IDR", broker="Interactive Brokers")
        self.assertEqual(position.ticker, "IDR")
        self.assertEqual(position.quote_symbol, "IDR.MC")
        self.assertEqual(position.benchmark_symbol, "^IBEX")

    def test_can_delete_watchlist_position(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.XIMO,
            broker="Seguimiento",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("12.0000"),
            current_price_per_share=Decimal("12.0000"),
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "delete_position",
                "position_id": str(position.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('equities:list')}#equity-ibex")
        self.assertFalse(EquityPosition.objects.filter(pk=position.id).exists())

    def test_can_close_owned_position_and_move_it_to_sales_history(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.XIMO,
            broker="Interactive Brokers",
            trade_channel=EquityPosition.TradeChannel.APP,
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2025, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.3000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "close_position",
                "position_id": str(position.id),
                "closed_on": "2026-04-12",
                "sale_price_per_share": "12,8000",
                "notes": "Venta completa",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('equities:list')}#equity-journey")
        self.assertFalse(EquityPosition.objects.filter(pk=position.id).exists())
        closed = EquityClosedPosition.objects.get(ticker="IBE")
        self.assertEqual(closed.sale_price_per_share, Decimal("12.8000"))
        self.assertEqual(closed.closed_on, date(2026, 4, 12))
        self.assertEqual(closed.notes, "Venta completa")

    def test_bank_company_name_switches_default_reference_to_euribor(self):
        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "create_position",
                "position_kind": EquityPosition.PositionKind.OWNED,
                "ownership_category": AssetOwnershipCategory.JOINT,
                "broker": "Banco Sabadell",
                "ticker": "",
                "company_name": "Banco Santander",
                "quote_symbol": "",
                "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
                "benchmark_symbol": "",
                "benchmark_name": "",
                "shares": "15",
                "average_cost_per_share": "4,2500",
                "current_price_per_share": "",
                "annual_dividend_income": "20",
                "annual_maintenance_cost": "0",
                "notes": "",
            },
        )

        self.assertRedirects(response, reverse("equities:list"))
        position = EquityPosition.objects.get(ticker="SAN", broker="Banco Sabadell")
        self.assertEqual(position.reference_profile, EquityPosition.ReferenceProfile.EURIBOR_12M)
        self.assertEqual(position.benchmark_symbol, EURIBOR_REFERENCE_SYMBOL)

    def test_watchlist_can_be_saved_without_shares_or_prices(self):
        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "create_position",
                "position_kind": EquityPosition.PositionKind.WATCHLIST,
                "ownership_category": AssetOwnershipCategory.JOINT,
                "broker": "",
                "ticker": "",
                "company_name": "Iberdrola",
                "quote_symbol": "",
                "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
                "benchmark_symbol": "",
                "benchmark_name": "",
                "shares": "",
                "average_cost_per_share": "",
                "current_price_per_share": "",
                "annual_dividend_income": "",
                "annual_maintenance_cost": "",
                "notes": "",
            },
        )

        self.assertRedirects(response, reverse("equities:list"))
        position = EquityPosition.objects.get(ticker="IBE", broker="Seguimiento")
        self.assertEqual(position.position_kind, EquityPosition.PositionKind.WATCHLIST)
        self.assertEqual(position.shares, Decimal("0"))
        self.assertEqual(position.average_cost_per_share, Decimal("0"))
        self.assertEqual(position.current_price_per_share, Decimal("0"))
        self.assertEqual(position.reference_profile, EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND)

    def test_can_change_reference_from_analysis_card(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Seguimiento",
            ticker="IBE",
            quote_symbol="IBE.MC",
            reference_profile=EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("0"),
            current_price_per_share=Decimal("0"),
        )

        with patch("equities.views.sync_equity_market_data") as mocked_sync:
            response = self.client.post(
                reverse("equities:list"),
                {
                    "action": "change_reference",
                    "position_id": str(position.id),
                    "reference_profile": EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
                    "benchmark_symbol": SPAIN_ELECTRICITY_DEMAND_SYMBOL,
                    "benchmark_name": SPAIN_ELECTRICITY_DEMAND_NAME,
                },
            )

        self.assertRedirects(response, reverse("equities:list"))
        position.refresh_from_db()
        self.assertEqual(position.reference_profile, EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND)
        self.assertEqual(position.benchmark_symbol, SPAIN_ELECTRICITY_DEMAND_SYMBOL)
        mocked_sync.assert_called_once()

    def test_can_store_same_ticker_for_same_broker_with_different_owner(self):
        EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.XIMO,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("9.0000"),
            current_price_per_share=Decimal("10.0000"),
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "create_position",
                "position_kind": EquityPosition.PositionKind.OWNED,
                "ownership_category": AssetOwnershipCategory.MONICA,
                "broker": "Interactive Brokers",
                "ticker": "IBE",
                "company_name": "Iberdrola",
                "quote_symbol": "IBE.MC",
                "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
                "benchmark_symbol": "^IBEX",
                "benchmark_name": "IBEX 35",
                "shares": "25",
                "average_cost_per_share": "10,5000",
                "current_price_per_share": "11,2500",
                "annual_dividend_income": "20,00",
                "annual_maintenance_cost": "4,50",
                "notes": "",
            },
        )

        self.assertRedirects(response, reverse("equities:list"))
        self.assertEqual(EquityPosition.objects.filter(broker="Interactive Brokers", ticker="IBE").count(), 2)
        self.assertTrue(
            EquityPosition.objects.filter(
                broker="Interactive Brokers",
                ticker="IBE",
                ownership_category=AssetOwnershipCategory.MONICA,
                shares=Decimal("25.0000"),
                annual_maintenance_cost=Decimal("4.50"),
            ).exists()
        )

    def test_updating_position_does_not_fail_if_duplicate_rows_already_exist(self):
        first = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.XIMO,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola antigua",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("9.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.XIMO,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola duplicada",
            shares=Decimal("12.0000"),
            average_cost_per_share=Decimal("9.5000"),
            current_price_per_share=Decimal("10.5000"),
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "create_position",
                "position_kind": EquityPosition.PositionKind.OWNED,
                "ownership_category": AssetOwnershipCategory.XIMO,
                "broker": "Interactive Brokers",
                "ticker": "IBE",
                "company_name": "Iberdrola revisada",
                "quote_symbol": "IBE.MC",
                "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
                "benchmark_symbol": "^IBEX",
                "benchmark_name": "IBEX 35",
                "shares": "25",
                "average_cost_per_share": "10,5000",
                "current_price_per_share": "11,2500",
                "annual_dividend_income": "20,00",
                "annual_maintenance_cost": "3,10",
                "notes": "Actualizada",
            },
        )

        self.assertRedirects(response, reverse("equities:list"))
        self.assertEqual(EquityPosition.objects.filter(broker="Interactive Brokers", ticker="IBE").count(), 2)
        first.refresh_from_db()
        self.assertEqual(first.company_name, "Iberdrola antigua")
        latest = EquityPosition.objects.order_by("-updated_at", "-id").first()
        self.assertEqual(latest.company_name, "Iberdrola revisada")
        self.assertEqual(latest.shares, Decimal("25.0000"))
        self.assertEqual(latest.annual_maintenance_cost, Decimal("3.10"))

    def test_can_prefill_equity_form_from_xls_document(self):
        document = SimpleUploadedFile(
            "posicion.xls",
            b"""
<html>
  <body>
    <table>
      <tr><td>Titular: Monica</td></tr>
      <tr>
        <td>Broker</td>
        <td>Ticker</td>
        <td>Empresa</td>
        <td>Acciones</td>
        <td>Coste medio</td>
        <td>Precio actual</td>
        <td>Dividendo anual</td>
      </tr>
      <tr>
        <td>ING</td>
        <td>ENG</td>
        <td>Enagas</td>
        <td>150,0000</td>
        <td>13,4500</td>
        <td>14,2000</td>
        <td>95,00</td>
      </tr>
    </table>
  </body>
</html>
""",
            content_type="application/vnd.ms-excel",
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "prefill_from_document",
                "document": document,
                "default_broker": "",
                "default_ownership_category": AssetOwnershipCategory.JOINT,
            },
        )

        self.assertEqual(response.status_code, 200)
        form = response.context["position_form"]
        self.assertEqual(form.initial["broker"], "ING")
        self.assertEqual(form.initial["ticker"], "ENG")
        self.assertEqual(form.initial["company_name"], "Enagas")
        self.assertEqual(form.initial["ownership_category"], AssetOwnershipCategory.MONICA)
        self.assertEqual(form.initial["position_kind"], EquityPosition.PositionKind.OWNED)
        self.assertEqual(form.initial["shares"], Decimal("150.0000"))
        self.assertEqual(form.initial["average_cost_per_share"], Decimal("13.4500"))
        self.assertEqual(form.initial["current_price_per_share"], Decimal("14.2000"))
        self.assertEqual(form.initial["annual_dividend_income"], Decimal("95.00"))
        self.assertEqual(response.context["prefill_source_filename"], "posicion.xls")

    def test_can_prefill_equity_form_from_pdf_document(self):
        document = SimpleUploadedFile(
            "posicion.pdf",
            b"%PDF-1.4 fake",
            content_type="application/pdf",
        )

        with patch(
            "equities.services.read_pdf_pages",
            return_value=[
                "\n".join(
                    [
                        "Titular:\tXimo",
                        "Broker\tSelf Bank",
                        "Ticker\tIBE",
                        "Empresa\tIberdrola",
                        "Acciones\t125,5000",
                        "Coste medio\t10,2500",
                        "Precio actual\t11,0000",
                        "Dividendos anuales\t72,50",
                    ]
                )
            ],
        ):
            response = self.client.post(
                reverse("equities:list"),
                {
                    "action": "prefill_from_document",
                    "document": document,
                    "default_broker": "",
                    "default_ownership_category": AssetOwnershipCategory.JOINT,
                },
            )

        self.assertEqual(response.status_code, 200)
        form = response.context["position_form"]
        self.assertEqual(form.initial["broker"], "Self Bank")
        self.assertEqual(form.initial["ticker"], "IBE")
        self.assertEqual(form.initial["company_name"], "Iberdrola")
        self.assertEqual(form.initial["ownership_category"], AssetOwnershipCategory.XIMO)
        self.assertEqual(form.initial["position_kind"], EquityPosition.PositionKind.OWNED)
        self.assertEqual(form.initial["shares"], Decimal("125.5000"))
        self.assertEqual(form.initial["average_cost_per_share"], Decimal("10.2500"))
        self.assertEqual(form.initial["current_price_per_share"], Decimal("11.0000"))
        self.assertEqual(form.initial["annual_dividend_income"], Decimal("72.50"))

    def test_equities_page_renders_without_auto_sync_when_disabled(self):
        EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Banco Sabadell",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
            annual_dividend_income=Decimal("40.00"),
            annual_maintenance_cost=Decimal("6.00"),
        )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cockpit de acciones")
        self.assertContains(response, "Canal de compra")
        self.assertContains(response, "Indra Sistemas, S.A.")

    @override_settings(EQUITIES_AUTO_SYNC_ON_VIEW=True)
    def test_optimizer_request_skips_auto_sync_even_when_enabled(self):
        EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )

        with patch("equities.views.sync_all_equities_market_data") as mocked_sync:
            response = self.client.get(
                reverse("equities:list"),
                {
                    "total_investment": "400000",
                    "max_company_pct": "20",
                    "max_sector_positions": "1",
                },
            )

        self.assertEqual(response.status_code, 200)
        mocked_sync.assert_not_called()

    @override_settings(EQUITIES_AUTO_SYNC_ON_VIEW=True)
    def test_auto_sync_on_page_load_only_targets_owned_positions_with_quote_symbol(self):
        owned = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ENG",
            quote_symbol="",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Enagas",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Seguimiento",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("0.0000"),
            average_cost_per_share=Decimal("0.0000"),
            current_price_per_share=Decimal("0.0000"),
        )

        with patch("equities.views.sync_all_equities_market_data", return_value=[(owned, None)]) as mocked_sync:
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        mocked_sync.assert_called_once()
        synced_positions = mocked_sync.call_args.args[0]
        self.assertEqual([position.id for position in synced_positions], [owned.id])

    @override_settings(EQUITIES_IBEX_UNIVERSE_ANALYSIS=False, EQUITIES_IBEX_UNIVERSE_LIMIT=3)
    def test_optimizer_request_forces_full_ibex_analysis_even_if_page_limit_is_disabled(self):
        EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )

        with patch("equities.views.build_equity_analysis_dashboard") as mocked_dashboard:
            mocked_dashboard.return_value = {
                "overview": {
                    "owned_positions_count": 0,
                    "watchlist_positions_count": 0,
                    "invested_amount": Decimal("0"),
                    "current_value": Decimal("0"),
                    "net_annual_income_total": Decimal("0"),
                },
                "history_cards": [],
                "owned_positions": [],
                "watchlist_positions": [],
                "owned_history_cards": [],
                "watchlist_history_cards": [],
                "decision_rows": [],
                "ibex_universe_rows": [],
                "ibex_universe_summary": {
                    "available": False,
                    "analyzed_count": 0,
                    "buy_alert_count": 0,
                    "sell_alert_count": 0,
                    "watch_alert_count": 0,
                    "failed_count": 0,
                    "failures": [],
                    "broker_assumption": "",
                    "trade_channel_label": "",
                    "top_pick": None,
                },
                "tracked_reference_rows": [],
                "reference_guide_rows": [],
                "reference_guide_summary": {"workbook_loaded": False},
                "optimizer_cards": [],
            }

            response = self.client.get(
                reverse("equities:list"),
                {
                    "total_investment": "400000",
                    "max_company_pct": "20",
                    "max_sector_positions": "1",
                },
            )

        self.assertEqual(response.status_code, 200)
        mocked_dashboard.assert_called_once()
        self.assertTrue(mocked_dashboard.call_args.kwargs["include_ibex_universe"])
        self.assertIsNone(mocked_dashboard.call_args.kwargs["ibex_company_limit"])

    def test_equities_page_uses_ibex_radar_as_master_list_for_saved_watchlist(self):
        with patch("equities.views.build_equity_analysis_dashboard") as mocked_dashboard:
            mocked_dashboard.return_value = {
                "overview": {
                    "owned_positions_count": 1,
                    "watchlist_positions_count": 1,
                    "invested_amount": Decimal("1000"),
                    "current_value": Decimal("1100"),
                    "annual_dividends_total": Decimal("30"),
                    "net_dividends_total": Decimal("24"),
                    "annual_maintenance_total": Decimal("6"),
                    "purchase_cost_total": Decimal("4"),
                    "net_annual_income_total": Decimal("20"),
                    "unrealized_gain_total": Decimal("50"),
                    "unrealized_return_pct": Decimal("5"),
                    "weighted_projected_return_12m": Decimal("8"),
                    "weighted_safety_score": Decimal("62"),
                    "weighted_periods": [],
                    "best_decision": None,
                },
                "history_cards": [],
                "owned_positions": [],
                "watchlist_positions": [],
                "owned_history_cards": [],
                "watchlist_history_cards": [],
                "decision_rows": [],
                "ibex_universe_rows": [
                    {
                        "ticker": "IBE",
                        "company_name": "Iberdrola",
                        "status_label": "En seguimiento",
                        "status_note": "La tienes guardada",
                        "trade_alert_label": "Comprar",
                        "trade_alert_tone": "buy",
                        "trade_alert_trigger": "Pendiente positiva",
                        "reference_label": "IBEX 35",
                        "best_reference_label": "IBEX 35",
                        "correlation": Decimal("0.72"),
                        "years_covered": Decimal("10.00"),
                        "projected_return_pct": Decimal("8.50"),
                        "safety_score": Decimal("70.00"),
                        "reliability_label": "Alta",
                        "benefit_risk_ratio": Decimal("1.80"),
                        "cycle_phase": "Expansion",
                    }
                ],
                "ibex_universe_summary": {
                    "available": True,
                    "analyzed_count": 35,
                    "buy_alert_count": 10,
                    "sell_alert_count": 5,
                    "watch_alert_count": 20,
                    "registered_count": 2,
                    "registered_owned_count": 1,
                    "registered_watchlist_count": 1,
                    "radar_only_count": 33,
                    "failed_count": 0,
                    "failures": [],
                    "broker_assumption": "Interactive Brokers",
                    "trade_channel_label": "App",
                    "top_pick": {"ticker": "IBE", "company_name": "Iberdrola"},
                },
                "tracked_reference_rows": [],
                "reference_guide_rows": [],
                "reference_guide_summary": {"workbook_loaded": False},
                "optimizer_cards": [],
            }

            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Esta es ahora la lista maestra del IBEX")
        self.assertContains(response, "seguimientos guardados")
        self.assertContains(response, "La tienes guardada")
        self.assertNotContains(response, "Acciones en seguimiento")

    def test_can_launch_background_optimizer_run_from_page(self):
        EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        runs = [
            EquityOptimizationRun(
                reference_code="OPT-TEST-LAUNCH-12M",
                label="Cartera defensiva - 12M principal",
                total_investment=Decimal("100000"),
                max_company_pct=Decimal("20"),
                max_total_positions=0,
                max_sector_positions=1,
                status=EquityOptimizationRun.Status.PENDING,
            ),
            EquityOptimizationRun(
                reference_code="OPT-TEST-LAUNCH-5A",
                label="Cartera defensiva - 5A principal",
                total_investment=Decimal("100000"),
                max_company_pct=Decimal("20"),
                max_total_positions=0,
                max_sector_positions=1,
                status=EquityOptimizationRun.Status.PENDING,
            ),
        ]

        with patch("equities.views.launch_equity_optimization_run_pair", return_value=runs) as mocked_launch:
            response = self.client.post(
                reverse("equities:list"),
                {
                    "action": "launch_optimizer_run",
                    "reference_label": "Cartera defensiva",
                    "total_investment": "400000",
                    "max_company_pct": "20",
                    "max_total_positions": "8",
                    "max_sector_positions": "1",
                    "selected_sectors": ["Electrica", "Banca"],
                    "selected_owned_tickers": ["IBE"],
                    "restrictions_note": "Maximo una empresa por sector",
                },
            )

        self.assertRedirects(response, f"{reverse('equities:list')}?optimizer_status=1#equity-optimizer")
        mocked_launch.assert_called_once()
        self.assertEqual(mocked_launch.call_args.kwargs["max_total_positions"], 8)
        self.assertEqual(mocked_launch.call_args.kwargs["selected_sectors"], ["Electrica", "Banca"])
        self.assertEqual(mocked_launch.call_args.kwargs["selected_owned_tickers"], ["IBE"])
        self.assertTrue(mocked_launch.call_args.kwargs["selected_owned_tickers_applied"])

    def test_equities_page_renders_optimizer_sector_selection(self):
        EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sectores donde si comprar")
        self.assertContains(response, "Acciones compradas que si pueden entrar")
        self.assertContains(response, "Banca")
        self.assertContains(response, "Iberdrola (IBE)")

    def test_equities_page_defers_full_ibex_radar_while_optimization_is_active(self):
        EquityOptimizationRun.objects.create(
            reference_code="OPT-RUNNING-001",
            label="Ejecucion activa",
            total_investment=Decimal("250000"),
            max_company_pct=Decimal("20"),
            max_total_positions=6,
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.RUNNING,
        )
        dashboard = {
            "overview": {
                "invested_amount": Decimal("0"),
                "current_value": Decimal("0"),
                "net_annual_income_total": Decimal("0"),
                "owned_positions_count": 0,
                "watchlist_positions_count": 0,
                "unrealized_return_pct": None,
                "unrealized_gain_total": Decimal("0"),
                "weighted_projected_return_12m": None,
                "weighted_safety_score": None,
                "weighted_periods": [],
                "best_decision": None,
            },
            "history_cards": [],
            "owned_positions": [],
            "watchlist_positions": [],
            "owned_history_cards": [],
            "watchlist_history_cards": [],
            "decision_rows": [],
            "ibex_universe_rows": [],
            "ibex_universe_summary": {"available": False, "analyzed_count": 0},
            "tracked_reference_rows": [],
            "reference_guide_rows": [],
            "reference_guide_summary": {"workbook_loaded": False},
            "optimizer_cards": [],
        }

        with patch("equities.views.build_equity_analysis_dashboard", return_value=dashboard) as mocked_dashboard:
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(mocked_dashboard.call_args.kwargs["include_ibex_universe"])
        self.assertContains(response, "se aplaza temporalmente mientras hay optimizaciones activas")

    def test_equities_page_keeps_single_optimization_history_table(self):
        first = EquityOptimizationRun.objects.create(
            reference_code="OPT-COMP-001",
            label="Cartera defensiva",
            total_investment=Decimal("300000"),
            max_company_pct=Decimal("20"),
            max_total_positions=6,
            max_sector_positions=1,
            selected_sectors=["Electrica", "Tecnologia y defensa"],
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html>uno</html>",
            summary_data={
                "available": True,
                "created_at_label": "2026-04-10 08:00",
                "completed_at_label": "2026-04-10 08:12",
                "total_investment": 300000,
                "max_company_pct": 20,
                "max_total_positions": 6,
                "max_sector_positions": 1,
                "selected_sectors": ["Electrica", "Tecnologia y defensa"],
                "projected_gain_total": 42000,
                "weighted_return_pct": 14.0,
                "weighted_low_return_pct": -4.5,
                "weighted_safety_score": 76,
                "weighted_reliability_score": 71,
                "net_dividend_income_total": 4200,
                "annual_cost_total": 350,
                "roundtrip_cost_total": 980,
                "cash_reserve_amount": 12000,
                "allocations_count": 4,
                "top_pick_name": "Iberdrola",
            },
            allocations_data=[
                {"company_name": "Iberdrola", "ticker": "IBE", "sector_label": "Electrica"},
                {"company_name": "Indra", "ticker": "IDR", "sector_label": "Tecnologia y defensa"},
            ],
        )
        second = EquityOptimizationRun.objects.create(
            reference_code="OPT-COMP-002",
            label="Cartera agresiva",
            total_investment=Decimal("300000"),
            max_company_pct=Decimal("25"),
            max_total_positions=8,
            max_sector_positions=2,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html>dos</html>",
            summary_data={
                "available": True,
                "created_at_label": "2026-04-11 09:00",
                "completed_at_label": "2026-04-11 09:18",
                "total_investment": 300000,
                "max_company_pct": 25,
                "max_total_positions": 8,
                "max_sector_positions": 2,
                "selected_sectors": [],
                "projected_gain_total": 57000,
                "weighted_return_pct": 19.0,
                "weighted_low_return_pct": -9.0,
                "weighted_safety_score": 62,
                "weighted_reliability_score": 66,
                "net_dividend_income_total": 2600,
                "annual_cost_total": 410,
                "roundtrip_cost_total": 1320,
                "cash_reserve_amount": 0,
                "allocations_count": 5,
                "top_pick_name": "Indra",
            },
            allocations_data=[
                {"company_name": "Indra", "ticker": "IDR", "sector_label": "Tecnologia y defensa"},
                {"company_name": "Repsol", "ticker": "REP", "sector_label": "Energia"},
            ],
        )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Optimizaciones realizadas")
        self.assertContains(response, "equity-optimization-history")
        self.assertContains(response, first.display_label)
        self.assertContains(response, second.display_label)
        self.assertContains(response, "Iberdrola")
        self.assertContains(response, "Repsol")
        self.assertNotContains(response, "Comparacion entre optimizaciones cerradas")

    def test_optimization_history_table_shows_company_weights(self):
        EquityOptimizationRun.objects.create(
            reference_code="OPT-HIST-001",
            label="Ticket visible",
            total_investment=Decimal("250000"),
            max_company_pct=Decimal("20"),
            max_total_positions=5,
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html>hist</html>",
            summary_data={
                "available": True,
                "projected_gain_total": 12000,
                "weighted_return_pct": 8.4,
                "allocations_count": 3,
                "selected_sectors": ["Electrica", "Energia"],
            },
            allocations_data=[
                {"company_name": "Iberdrola", "ticker": "IBE", "allocated_weight_pct": 35.0},
                {"company_name": "Repsol", "ticker": "REP", "allocated_weight_pct": 25.0},
                {"company_name": "Endesa", "ticker": "ELE", "allocated_weight_pct": 20.0},
            ],
        )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Plan propuesto")
        self.assertContains(response, "Iberdrola")
        self.assertContains(response, "35,0 %")
        self.assertContains(response, "Repsol")
        self.assertContains(response, "25,0 %")

    def test_optimization_history_table_shows_compact_timeline_and_annualized_trade(self):
        history_day = timezone.localdate() - timedelta(days=6)
        EquityOptimizationRun.objects.create(
            reference_code="OPT-HIST-TIMELINE",
            label="Ticket con agenda",
            total_investment=Decimal("250000"),
            max_company_pct=Decimal("20"),
            max_total_positions=5,
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html>hist</html>",
            summary_data={
                "available": True,
                "projected_gain_total": 12000,
                "weighted_return_pct": 8.4,
                "allocations_count": 2,
            },
            allocations_data=[
                {
                    "rank": 1,
                    "company_name": "ACS, Actividades de Construccion y Servicios, S.A.",
                    "ticker": "ACS",
                    "allocated_weight_pct": 35.0,
                    "current_price_per_share": 40.10,
                    "latest_price_date": timezone.localdate().isoformat(),
                    "purchase_timing": {
                        "available": True,
                        "buy_date": (history_day + timedelta(days=12)).isoformat(),
                        "buy_window_label": "mayo 2026",
                        "buy_price": 39.25,
                        "exit_date": (history_day + timedelta(days=220)).isoformat(),
                        "exit_window_label": "noviembre 2026",
                        "exit_price": 43.80,
                        "interval_window_label": "mayo 2026 -> noviembre 2026",
                        "interval_return_pct": 11.6,
                        "holding_months": 7,
                    },
                },
                {
                    "rank": 2,
                    "company_name": "Repsol, S.A.",
                    "ticker": "REP",
                    "allocated_weight_pct": 25.0,
                    "current_price_per_share": 13.40,
                    "latest_price_date": timezone.localdate().isoformat(),
                    "purchase_timing": {
                        "available": True,
                        "buy_date": (history_day + timedelta(days=35)).isoformat(),
                        "buy_window_label": "junio 2026",
                        "buy_price": 13.10,
                        "exit_date": (history_day + timedelta(days=300)).isoformat(),
                        "exit_window_label": "febrero 2027",
                        "exit_price": 14.25,
                        "interval_window_label": "junio 2026 -> febrero 2027",
                        "interval_return_pct": 7.4,
                        "holding_months": 9,
                    },
                },
            ],
        )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "equity-mini-optimization")
        self.assertContains(response, "ACS, Actividades de Construccion y Servicios, S.A.")
        self.assertContains(response, "E mayo 2026")
        self.assertContains(response, "S noviembre 2026")
        self.assertContains(response, "Hoy 40,1000 EUR")
        self.assertContains(response, "%/a")

    def test_equities_page_shows_purchase_recommendation_and_gantt_for_optimization(self):
        scheduled_day = timezone.localdate() - timedelta(days=3)
        EquityOptimizationRun.objects.create(
            reference_code="OPT-SCHED-BUY-12M",
            label="Programada 12M",
            total_investment=Decimal("200000"),
            max_company_pct=Decimal("30"),
            max_total_positions=5,
            max_sector_positions=2,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html>scheduled</html>",
            progress_data={
                "strategy_label": "12M principal",
                "strategy_mode": "12m_primary",
                "schedule_kind": "nightly",
                "scheduled_run_key": f"scheduled-optimization:{scheduled_day.isoformat()}",
                "scheduled_analysis_date": scheduled_day.isoformat(),
                "scheduled_weekdays_label": "martes y jueves",
            },
            summary_data={
                "available": True,
                "strategy_label": "12M principal",
                "scheduled_analysis_date": scheduled_day.isoformat(),
            },
            allocations_data=[
                {
                    "rank": 1,
                    "company_name": "ACS",
                    "ticker": "ACS",
                    "allocated_amount": 22000.0,
                    "allocated_weight_pct": 22.0,
                    "current_price_per_share": 40.10,
                    "latest_price_date": timezone.localdate().isoformat(),
                    "net_projected_return_pct": 8.0,
                    "cycle_return_5y_pct": 40.0,
                    "reliability_label": "Media",
                    "reliability_score": 60.0,
                    "purchase_timing": {
                        "available": True,
                        "buy_date": (scheduled_day + timedelta(days=12)).isoformat(),
                        "buy_window_label": "mayo 2026",
                        "buy_price": 39.25,
                        "exit_date": (scheduled_day + timedelta(days=220)).isoformat(),
                        "exit_window_label": "noviembre 2026",
                        "exit_price": 43.80,
                        "interval_window_label": "mayo 2026 -> noviembre 2026",
                        "interval_return_pct": 11.6,
                        "holding_months": 7,
                        "mode_label": "Comprar ya",
                    },
                    "cycle_yearly_margins": [
                        {"year_number": 1, "label": "ANO 1", "margin_pct": 8.0},
                    ],
                },
            ],
        )
        EquityOptimizationRun.objects.create(
            reference_code="OPT-LATEST-BUY",
            label="Ultima optimizacion",
            total_investment=Decimal("180000"),
            max_company_pct=Decimal("30"),
            max_total_positions=5,
            max_sector_positions=2,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html>latest</html>",
            summary_data={
                "available": True,
                "strategy_label": "12M principal",
                "projected_gain_total": 18000,
                "weighted_return_pct": 10.0,
                "weighted_low_return_pct": -4.0,
                "allocations_count": 1,
                "top_pick_name": "ACS",
                "top_pick_buy_window_label": "mayo 2026",
                "top_pick_buy_price": 39.25,
                "top_pick_exit_window_label": "noviembre 2026",
                "top_pick_exit_price": 43.80,
                "top_pick_interval_return_pct": 11.6,
            },
            allocations_data=[
                {
                    "rank": 1,
                    "company_name": "ACS",
                    "ticker": "ACS",
                    "current_price_per_share": 40.10,
                    "latest_price_date": timezone.localdate().isoformat(),
                    "status_label": "Radar IBEX",
                    "sector_label": "Infraestructuras",
                    "trade_alert_label": "Comprar",
                    "reference_label": "IBEX 35",
                    "external_signal_label": "Prensa neutra",
                    "allocated_amount": 24000.0,
                    "allocated_weight_pct": 24.0,
                    "net_projected_return_pct": 11.0,
                    "low_return_pct": -3.0,
                    "expected_net_dividend_income": 200.0,
                    "annual_cost_used": 12.0,
                    "roundtrip_total_cost": 40.0,
                    "purchase_timing": {
                        "available": True,
                        "buy_date": (timezone.localdate() + timedelta(days=10)).isoformat(),
                        "buy_window_label": "mayo 2026",
                        "buy_price": 39.25,
                        "exit_date": (timezone.localdate() + timedelta(days=220)).isoformat(),
                        "exit_window_label": "noviembre 2026",
                        "exit_price": 43.80,
                        "interval_window_label": "mayo 2026 -> noviembre 2026",
                        "interval_return_pct": 11.6,
                        "holding_months": 7,
                        "mode_label": "Comprar ya",
                        "expected_trade_return_pct": 11.4,
                    },
                },
            ],
        )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Primer tramo tactico")
        self.assertContains(response, "mayo 2026")
        self.assertContains(response, "noviembre 2026")
        self.assertContains(response, "Entradas durante un ano, salidas cuando el tramo da mas")
        self.assertContains(response, "39,2500 EUR")
        self.assertContains(response, "Hoy en mercado")
        self.assertContains(response, "40,1000 EUR")

    def helper_equities_page_shows_scheduled_optimization_persistence_panel_legacy(self):
        today = timezone.localdate()
        scheduled_day = today - timedelta(days=3)
        for strategy_label, strategy_mode in (("12M principal", "12m_primary"), ("5A principal", "5y_primary")):
            run = EquityOptimizationRun.objects.create(
                reference_code=f"OPT-PERSIST-{strategy_mode}",
                label=f"Programada {strategy_label}",
                total_investment=Decimal("200000"),
                max_company_pct=Decimal("30"),
                max_total_positions=5,
                max_sector_positions=2,
                status=EquityOptimizationRun.Status.COMPLETED,
                progress_data={
                    "strategy_label": strategy_label,
                    "strategy_mode": strategy_mode,
                    "schedule_kind": "nightly",
                    "scheduled_run_key": f"scheduled-optimization:{scheduled_day.isoformat()}",
                    "scheduled_analysis_date": scheduled_day.isoformat(),
                    "scheduled_weekdays_label": "martes y jueves",
                },
                summary_data={
                    "available": True,
                    "strategy_label": strategy_label,
                    "scheduled_analysis_date": scheduled_day.isoformat(),
                },
                allocations_data=[
                    {
                        "rank": 1,
                        "company_name": "Iberdrola",
                        "ticker": "IBE",
                        "net_projected_return_pct": 12.5,
                        "cycle_return_5y_pct": 68.0,
                        "reliability_label": "Alta",
                        "reliability_score": 82.0,
                        "cycle_yearly_margins": [
                            {"year_number": 1, "label": "AÑO 1", "margin_pct": 12.5},
                            {"year_number": 2, "label": "AÑO 2", "margin_pct": 8.0},
                            {"year_number": 3, "label": "AÑO 3", "margin_pct": 9.0},
                            {"year_number": 4, "label": "AÑO 4", "margin_pct": 10.0},
                            {"year_number": 5, "label": "AÑO 5", "margin_pct": 11.0},
                        ],
                    },
                    {
                        "rank": 3,
                        "company_name": "Endesa",
                        "ticker": "ELE",
                        "net_projected_return_pct": 9.0,
                        "cycle_return_5y_pct": 44.0,
                        "reliability_label": "Media",
                        "reliability_score": 64.0,
                        "cycle_yearly_margins": [
                            {"year_number": 1, "label": "AÑO 1", "margin_pct": 9.0},
                            {"year_number": 2, "label": "AÑO 2", "margin_pct": 6.0},
                        ],
                    },
                ],
            )
            EquityOptimizationRun.objects.filter(pk=run.pk).update(
                created_at=timezone.make_aware(datetime.combine(scheduled_day, datetime.min.time())),
                completed_at=timezone.make_aware(datetime.combine(scheduled_day, datetime.min.time())),
            )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Valores que mas se repiten en optimizaciones automáticas")
        self.assertContains(response, "Iberdrola")
        self.assertContains(response, "100 %")
        self.assertContains(response, "2/2 ejec.")
        self.assertContains(response, "12M medio")
        self.assertContains(response, "Fiabilidad")
        self.assertContains(response, "A&Ntilde;O 1")
        self.assertContains(response, "5A acumulada")
        self.assertContains(response, "Max 5 empresas")

    def test_equities_page_shows_presence_percentage_in_scheduled_optimization_persistence_panel(self):
        today = timezone.localdate()
        scheduled_day = today - timedelta(days=3)
        for strategy_label, strategy_mode in (("12M principal", "12m_primary"), ("5A principal", "5y_primary")):
            run = EquityOptimizationRun.objects.create(
                reference_code=f"OPT-PERSIST-V2-{strategy_mode}",
                label=f"Programada {strategy_label}",
                total_investment=Decimal("200000"),
                max_company_pct=Decimal("30"),
                max_total_positions=5,
                max_sector_positions=2,
                status=EquityOptimizationRun.Status.COMPLETED,
                progress_data={
                    "strategy_label": strategy_label,
                    "strategy_mode": strategy_mode,
                    "schedule_kind": "nightly",
                    "scheduled_run_key": f"scheduled-optimization:{scheduled_day.isoformat()}",
                    "scheduled_analysis_date": scheduled_day.isoformat(),
                    "scheduled_weekdays_label": "martes y jueves",
                },
                summary_data={
                    "available": True,
                    "strategy_label": strategy_label,
                    "scheduled_analysis_date": scheduled_day.isoformat(),
                },
                allocations_data=[
                    {
                        "rank": 1,
                        "company_name": "Iberdrola",
                        "ticker": "IBE",
                        "net_projected_return_pct": 12.5,
                        "cycle_return_5y_pct": 68.0,
                        "reliability_label": "Alta",
                        "reliability_score": 82.0,
                        "cycle_yearly_margins": [
                            {"year_number": 1, "label": "AÃ‘O 1", "margin_pct": 12.5},
                            {"year_number": 2, "label": "AÃ‘O 2", "margin_pct": 8.0},
                            {"year_number": 3, "label": "AÃ‘O 3", "margin_pct": 9.0},
                            {"year_number": 4, "label": "AÃ‘O 4", "margin_pct": 10.0},
                            {"year_number": 5, "label": "AÃ‘O 5", "margin_pct": 11.0},
                        ],
                    },
                    {
                        "rank": 3,
                        "company_name": "Endesa",
                        "ticker": "ELE",
                        "net_projected_return_pct": 9.0,
                        "cycle_return_5y_pct": 44.0,
                        "reliability_label": "Media",
                        "reliability_score": 64.0,
                        "cycle_yearly_margins": [
                            {"year_number": 1, "label": "AÃ‘O 1", "margin_pct": 9.0},
                            {"year_number": 2, "label": "AÃ‘O 2", "margin_pct": 6.0},
                        ],
                    },
                ],
            )
            EquityOptimizationRun.objects.filter(pk=run.pk).update(
                created_at=timezone.make_aware(datetime.combine(scheduled_day, datetime.min.time())),
                completed_at=timezone.make_aware(datetime.combine(scheduled_day, datetime.min.time())),
            )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Valores mas consistentes en optimizaciones automaticas")
        self.assertContains(response, "Iberdrola")
        self.assertContains(response, "100 %")
        self.assertContains(response, "2/2 ejec.")
        self.assertContains(response, "12M medio")
        self.assertContains(response, "Fiabilidad")
        self.assertContains(response, "A&Ntilde;O 1")
        self.assertContains(response, "5A acumulada")
        self.assertContains(response, "Max 5 empresas")

    def test_completed_optimization_run_can_be_deleted_from_history(self):
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-DELETE-001",
            label="Borrar historico",
            total_investment=Decimal("75000"),
            max_company_pct=Decimal("20"),
            max_total_positions=5,
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html>borrar</html>",
            summary_data={"available": True, "projected_gain_total": 8200, "allocations_count": 2},
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "delete_optimization_run",
                "run_id": str(run.id),
            },
        )

        self.assertRedirects(response, f"{reverse('equities:list')}#equity-optimizer")
        self.assertFalse(EquityOptimizationRun.objects.filter(pk=run.pk).exists())

    def test_non_admin_cannot_delete_optimization_history_or_launch_financial_actions(self):
        admin = self.user
        viewer = get_user_model().objects.create_user(
            username="equity-viewer",
            password="StrongPass123!",
        )
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-DELETE-LOCKED",
            label="Borrar historico",
            requested_by=admin,
            total_investment=Decimal("75000"),
            max_company_pct=Decimal("20"),
            max_total_positions=5,
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html>borrar</html>",
        )

        self.client.force_login(viewer)
        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "delete_optimization_run",
                "run_id": str(run.id),
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(EquityOptimizationRun.objects.filter(pk=run.pk).exists())
        self.assertContains(response, "Solo un administrador puede modificar posiciones, sincronizar mercado o lanzar optimizaciones.")

    def test_running_optimization_run_cannot_be_deleted_from_history(self):
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-DELETE-002",
            label="No borrar",
            total_investment=Decimal("75000"),
            max_company_pct=Decimal("20"),
            max_total_positions=5,
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.RUNNING,
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "delete_optimization_run",
                "run_id": str(run.id),
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(EquityOptimizationRun.objects.filter(pk=run.pk).exists())
        self.assertContains(response, "Solo puedes borrar optimizaciones ya cerradas.")

    def test_completed_optimization_run_can_render_and_download_report(self):
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-TEST-REPORT",
            label="Informe guardado",
            total_investment=Decimal("50000"),
            max_company_pct=Decimal("20"),
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html><body>Informe optimizado</body></html>",
            summary_data={"available": True, "projected_gain_total": 5000},
        )

        report_response = self.client.get(reverse("equities:optimization_report", args=[run.id]))
        download_response = self.client.get(reverse("equities:optimization_download", args=[run.id]))
        download_html_response = self.client.get(reverse("equities:optimization_download_html", args=[run.id]))

        self.assertEqual(report_response.status_code, 200)
        self.assertContains(report_response, "Informe optimizado")
        self.assertEqual(download_response.status_code, 200)
        self.assertEqual(download_response["Content-Type"], "application/pdf")
        self.assertIn("attachment;", download_response.headers["Content-Disposition"])
        self.assertIn("opt-test-report.pdf", download_response.headers["Content-Disposition"])
        self.assertTrue(download_response.content.startswith(b"%PDF"))
        self.assertEqual(download_html_response.status_code, 200)
        self.assertIn("attachment;", download_html_response.headers["Content-Disposition"])
        self.assertIn("opt-test-report.html", download_html_response.headers["Content-Disposition"])

    def test_non_admin_cannot_open_or_poll_other_users_optimization_runs(self):
        owner = get_user_model().objects.create_user(
            username="optimization-owner",
            password="StrongPass123!",
        )
        viewer = get_user_model().objects.create_user(
            username="optimization-viewer",
            password="StrongPass123!",
        )
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-PRIVATE-001",
            label="Informe privado",
            requested_by=owner,
            total_investment=Decimal("50000"),
            max_company_pct=Decimal("20"),
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html><body>Informe privado</body></html>",
            summary_data={"available": True, "projected_gain_total": 5000},
        )

        self.client.force_login(viewer)

        report_response = self.client.get(reverse("equities:optimization_report", args=[run.id]))
        progress_response = self.client.get(reverse("equities:optimization_progress", args=[run.id]))

        self.assertEqual(report_response.status_code, 404)
        self.assertEqual(progress_response.status_code, 404)

    def test_non_admin_list_only_shows_own_optimization_runs(self):
        owner = get_user_model().objects.create_user(
            username="run-owner",
            password="StrongPass123!",
        )
        other_user = get_user_model().objects.create_user(
            username="run-other",
            password="StrongPass123!",
        )
        own_run = EquityOptimizationRun.objects.create(
            reference_code="OPT-OWN-001",
            label="Solo mia",
            requested_by=owner,
            total_investment=Decimal("50000"),
            max_company_pct=Decimal("20"),
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            summary_data={"available": True, "projected_gain_total": 1200},
        )
        EquityOptimizationRun.objects.create(
            reference_code="OPT-OTHER-001",
            label="Ajena",
            requested_by=other_user,
            total_investment=Decimal("50000"),
            max_company_pct=Decimal("20"),
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            summary_data={"available": True, "projected_gain_total": 2200},
        )

        self.client.force_login(owner)
        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        visible_run_ids = [run.id for run in response.context["optimization_runs"]]
        self.assertEqual(visible_run_ids, [own_run.id])

    def test_pdf_download_falls_back_to_simplified_template_when_rich_pdf_fails(self):
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-TEST-PDF-FALLBACK",
            label="Fallback PDF",
            total_investment=Decimal("50000"),
            max_company_pct=Decimal("20"),
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.COMPLETED,
            report_html="<html><body>Informe optimizado</body></html>",
            report_pdf_html="<html><body>Informe rico</body></html>",
            summary_data={"available": True, "projected_gain_total": 5000},
        )

        with (
            patch("equities.views.render_report_pdf", side_effect=[ValueError("rich failed"), b"%PDF-fallback"]),
            patch("equities.views.build_fallback_report_pdf_html", return_value="<html><body>Fallback</body></html>") as mocked_fallback,
        ):
            response = self.client.get(reverse("equities:optimization_download", args=[run.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))
        mocked_fallback.assert_called_once_with(run)

    def test_progress_endpoint_returns_live_payload(self):
        run = EquityOptimizationRun.objects.create(
            reference_code="OPT-TEST-PROGRESS",
            label="Seguimiento visible",
            total_investment=Decimal("50000"),
            max_company_pct=Decimal("20"),
            max_sector_positions=1,
            status=EquityOptimizationRun.Status.RUNNING,
            status_note="Analizando el IBEX",
            progress_data={
                "percent": 47,
                "stage_key": "ibex",
                "stage_label": "Analizando IBEX",
                "note": "Analizando el IBEX: 18/35",
                "current_step": 18,
                "total_steps": 35,
                "current_label": "Iberdrola",
                "stages": [
                    {"key": "sync", "label": "Sincronizando cartera", "status": "completed"},
                    {"key": "dashboard", "label": "Construyendo base de mercado", "status": "completed"},
                    {"key": "ibex", "label": "Analizando IBEX", "status": "active"},
                ],
                "preview_candidates": [
                    {
                        "ticker": "IBE",
                        "company_name": "Iberdrola",
                        "sector_label": "Energia",
                        "trade_alert_label": "Comprar",
                        "net_return_pct": 8.5,
                        "safety_score": 74,
                    }
                ],
                "preview_allocations": [],
                "events": [
                    {"label": "Analizando el IBEX: 18/35", "detail": "Iberdrola", "recorded_at": "16:22:10"}
                ],
                "updated_at_label": "2026-04-12 16:22:10",
            },
        )

        response = self.client.get(reverse("equities:optimization_progress", args=[run.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["reference_code"], "OPT-TEST-PROGRESS")
        self.assertEqual(payload["status"], EquityOptimizationRun.Status.RUNNING)
        self.assertFalse(payload["finished"])
        self.assertEqual(payload["progress"]["percent"], 47)
        self.assertEqual(payload["progress"]["current_label"], "Iberdrola")
        self.assertEqual(payload["progress"]["events"][0]["detail"], "Iberdrola")

    def test_equities_page_renders_ticket_tracking_section(self):
        for ticker, company_name, average_cost, current_price in (
            ("IBE", "Iberdrola", Decimal("10.0000"), Decimal("12.0000")),
            ("ENG", "Enagas", Decimal("14.0000"), Decimal("15.5000")),
        ):
            position = EquityPosition.objects.create(
                ownership_category=AssetOwnershipCategory.JOINT,
                broker="Interactive Brokers",
                ticker=ticker,
                quote_symbol=f"{ticker}.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name=company_name,
                shares=Decimal("20.0000"),
                average_cost_per_share=average_cost,
                current_price_per_share=current_price,
            )
            stock_series = build_compound_market_series(
                f"{ticker}.MC",
                company_name,
                growth=Decimal("1.0180"),
                start_price=average_cost,
            )
            reference_series = build_compound_market_series(
                "^IBEX",
                "IBEX 35",
                growth=Decimal("1.0070"),
                start_price=Decimal("100.0000"),
            )
            for stock_point, reference_point in zip(stock_series.points, reference_series.points):
                position.price_history.create(
                    price_date=stock_point["date"],
                    open_price=stock_point["open"],
                    high_price=stock_point["high"],
                    low_price=stock_point["low"],
                    close_price=stock_point["close"],
                    benchmark_close=reference_point["close"],
                )

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            initial_cards = build_equity_history_cards(
                list(EquityPosition.objects.prefetch_related("price_history"))
            )
            capture_equity_ticket_snapshots(initial_cards, snapshot_date=date(2026, 4, 20))

        with (
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
            patch("equities.views.build_equity_investment_journey_context", return_value={"available": False}),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        page = response.content.decode("utf-8")
        self.assertContains(response, "Seguimiento desde")
        self.assertContains(response, "Valor actual cartera")
        self.assertContains(response, "Capital invertido")
        self.assertContains(response, "Proyeccion cartera")
        self.assertContains(response, "Esperanza cartera")
        self.assertContains(response, "3M")
        self.assertContains(response, "6M")
        self.assertContains(response, "9M")
        self.assertContains(response, "1A")
        self.assertContains(response, "2A")
        self.assertContains(response, "3A")
        self.assertContains(response, "4A")
        self.assertContains(response, "5A")
        self.assertContains(response, "Cartera reescalada vs IBEX")
        self.assertContains(response, "Rentabilidad neta 1A")
        self.assertContains(response, "Rentabilidad neta 5A")
        self.assertContains(response, "% rentabilidad 1A")
        self.assertContains(response, "% rentabilidad 5A")
        self.assertContains(response, "equivalente")
        self.assertContains(response, "Acciones compradas")
        self.assertContains(response, "Resumen cartera y tablas")
        self.assertContains(response, "Prevision 12M")
        self.assertContains(response, "Ciclo 5A")
        self.assertContains(response, "Prediccion frente a realidad")
        self.assertContains(response, "Ticket IBE")
        self.assertContains(response, "Ticket ENG")
        self.assertContains(response, "Beneficio neto")
        self.assertContains(response, "Rentabilidad sobre base")
        self.assertContains(response, "Rentabilidad anualizada")
        self.assertContains(response, 'id="tracked-ticket-tab-', html=False)
        self.assertContains(response, 'aria-label="Tickets comprados"', html=False)
        self.assertContains(response, 'id="equity-portfolio-summary"', html=False)
        self.assertContains(response, 'id="equity-decision"', html=False)
        self.assertContains(response, 'id="equity-analysis"', html=False)
        self.assertContains(response, 'href="#tracked-ticket-', html=False)
        self.assertLess(page.index('class="equity-hero"'), page.index('id="equity-ticket-tracking"'))
        self.assertLess(page.index("Esperanza cartera"), page.index('id="equity-ticket-tracking"'))
        self.assertLess(page.index("Rentabilidad neta 1A"), page.index("Acciones compradas"))
        self.assertLess(page.index('id="equity-ticket-tracking"'), page.index('id="equity-portfolio-summary"'))
        self.assertLess(page.index('id="equity-portfolio-summary"'), page.index('id="equity-decision"'))
        self.assertLess(page.index('id="equity-decision"'), page.index('id="equity-ibex"'))
        self.assertEqual(EquityTicketSnapshot.objects.count(), 4)
        self.assertContains(response, "Base por accion")
        self.assertContains(response, "Precio por accion")

    def test_equities_page_renders_portfolio_correlation_section(self):
        close_sets = (
            ("IBE", "Iberdrola", [Decimal("20.0000"), Decimal("19.4000"), Decimal("20.1760"), Decimal("19.7725"), Decimal("20.7611"), Decimal("20.1383"), Decimal("21.3466"), Decimal("20.9197")]),
            ("ENG", "Enagas", [Decimal("18.0000"), Decimal("17.8200"), Decimal("18.3546"), Decimal("18.1709"), Decimal("18.7160"), Decimal("18.3417"), Decimal("19.0754"), Decimal("18.6940")]),
            ("BBVA", "BBVA", [Decimal("12.0000"), Decimal("13.0200"), Decimal("12.4341"), Decimal("13.2413"), Decimal("13.0427"), Decimal("14.0209"), Decimal("13.4601"), Decimal("14.2677")]),
            ("SAN", "Banco Santander", [Decimal("10.0000"), Decimal("11.0000"), Decimal("10.4500"), Decimal("11.2860"), Decimal("11.0603"), Decimal("11.9451"), Decimal("11.3478"), Decimal("12.1421")]),
        )

        for ticker, company_name, closes in close_sets:
            position = EquityPosition.objects.create(
                ownership_category=AssetOwnershipCategory.JOINT,
                broker="Interactive Brokers",
                ticker=ticker,
                quote_symbol=f"{ticker}.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name=company_name,
                shares=Decimal("10.0000"),
                average_cost_per_share=closes[0],
                current_price_per_share=closes[-1],
            )
            populate_position_history_from_closes(position, closes)

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with (
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
            patch("equities.views.build_equity_investment_journey_context", return_value={"available": False}),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        page = response.content.decode("utf-8")
        self.assertContains(response, "Correlacion entre tus acciones y riesgo de concentracion")
        self.assertContains(response, "Curva orientativa correlacion vs probabilidad de perder dinero")
        self.assertContains(response, "Matriz de correlacion entre tus acciones")
        self.assertContains(response, "BBVA / SAN")
        self.assertContains(response, "IBE / SAN")
        self.assertContains(response, 'id="equity-correlation"', html=False)
        self.assertLess(page.index('id="equity-ticket-tracking"'), page.index('id="equity-correlation"'))
        self.assertLess(page.index('id="equity-correlation"'), page.index('id="equity-portfolio-summary"'))

    def test_equities_page_refreshes_today_snapshot_even_with_nightly_status_available(self):
        position = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
            annual_dividend_income=Decimal("18.00"),
            annual_maintenance_cost=Decimal("4.00"),
        )
        populate_position_history(position, months=60)
        EquityNightlyAnalysisRun.objects.create(
            analysis_date=timezone.localdate(),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )

        with (
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
            patch("equities.views.capture_equity_ticket_snapshots", wraps=capture_equity_ticket_snapshots) as mocked_capture,
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        mocked_capture.assert_called_once()
        position.refresh_from_db()
        today_snapshot = EquityTicketSnapshot.objects.get(
            position=position,
            snapshot_date=timezone.localdate(),
        )
        self.assertEqual(today_snapshot.current_value, position.current_value.quantize(Decimal("0.01")))

    def test_equities_page_shows_per_in_main_decision_table(self):
        position = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="SAN",
            quote_symbol="SAN.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Banco Santander",
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("4.5000"),
            current_price_per_share=Decimal("4.9000"),
        )
        populate_position_history(position, growth=Decimal("1.0120"), benchmark_growth=Decimal("1.0060"), months=60)
        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        workbook_snapshot = {
            "available": True,
            "companies": [
                {
                    "ticker": "SAN",
                    "company_name": "Banco Santander",
                    "sector": "Banca",
                    "per_2025": Decimal("8.00"),
                }
            ],
            "companies_by_key": {
                "SAN": {
                    "ticker": "SAN",
                    "company_name": "Banco Santander",
                    "sector": "Banca",
                    "per_2025": Decimal("8.00"),
                },
                "BANCO SANTANDER": {
                    "ticker": "SAN",
                    "company_name": "Banco Santander",
                    "sector": "Banca",
                    "per_2025": Decimal("8.00"),
                },
                "SAN MC": {
                    "ticker": "SAN",
                    "company_name": "Banco Santander",
                    "sector": "Banca",
                    "per_2025": Decimal("8.00"),
                },
            },
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=workbook_snapshot),
            patch("equities.services.should_fetch_equity_fundamentals", return_value=False),
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
            patch("equities.views.build_equity_investment_journey_context", return_value={"available": False}),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-tooltip="Relacion entre precio y beneficio. Cuanto mas alto es, mas crecimiento exige la accion para justificar su precio."', html=False)
        self.assertContains(response, "Banco Santander")
        self.assertContains(response, "Ajustada")
        self.assertContains(response, "8.0")
        self.assertContains(response, 'class="good">8,0</strong>', html=False)

    def test_equities_page_backfills_missing_ticket_snapshots_for_new_owned_positions(self):
        iberdrola = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 10),
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        santander = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.MONICA,
            broker="Cartera Monica",
            ticker="SAN",
            quote_symbol="SAN.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Banco Santander",
            opened_on=date(2026, 4, 19),
            shares=Decimal("2550.0000"),
            average_cost_per_share=Decimal("11.0420"),
            current_price_per_share=Decimal("11.0420"),
        )
        populate_position_history(iberdrola)
        populate_position_history(santander)

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            iberdrola_card = build_equity_history_cards([iberdrola])[0]
        capture_equity_ticket_snapshots([iberdrola_card], snapshot_date=date(2026, 4, 18))

        nightly_status = {
            "available": True,
            "status_key": "completed",
            "status_badge": "OK",
            "completed_at_label": "2026-04-19 08:00",
            "agent_label": "Analista nocturno",
            "cache_available": False,
            "status_note": "",
            "error_message": "",
            "llm": {
                "enabled": False,
                "reason": "",
                "completed_count": 0,
                "total_count": 0,
                "estimated_cost_usd": "0",
                "refresh_performed": False,
                "refresh_failed_count": 0,
                "pending_count": 0,
                "skipped_budget_count": 0,
            },
        }
        with (
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
            patch("equities.views.build_equity_investment_journey_context", return_value={"available": False}),
            patch("equities.views.build_nightly_analysis_status", return_value=nightly_status),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Banco Santander")
        self.assertEqual(
            EquityTicketSnapshot.objects.filter(position=santander, snapshot_date=timezone.localdate()).count(),
            1,
        )

    def test_equities_page_tracking_shows_recommended_sale_and_reentry(self):
        position = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2023, 1, 31),
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        populate_position_history(position, growth=Decimal("1.0180"), benchmark_growth=Decimal("1.0070"), months=48)
        alternative = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Seguimiento",
            ticker="ELE",
            quote_symbol="ELE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Endesa",
            shares=Decimal("0.0000"),
            average_cost_per_share=Decimal("18.0000"),
            current_price_per_share=Decimal("18.0000"),
        )
        populate_position_history(alternative, growth=Decimal("1.0220"), benchmark_growth=Decimal("1.0070"), months=48)
        EquityPurchaseForecastBaseline.objects.create(
            position=position,
            source_analysis_date=date(2026, 4, 16),
            baseline_date=date(2026, 4, 16),
            reference_label="IBEX 35",
            trade_alert_label="Comprar",
            reliability_label="Alta",
            baseline_price=Decimal("10.0000"),
            projected_price_1y=Decimal("11.2000"),
            projected_price_2y=Decimal("9.9000"),
            projected_price_3y=Decimal("11.7000"),
            projected_price_4y=Decimal("12.4000"),
            projected_price_5y=Decimal("13.2000"),
            projected_return_pct_1y=Decimal("12.00"),
            projected_return_pct_2y=Decimal("-1.00"),
            projected_return_pct_3y=Decimal("17.00"),
            projected_return_pct_4y=Decimal("24.00"),
            projected_return_pct_5y=Decimal("32.00"),
        )

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        trade_plan_payload = {
            "available": True,
            "mode": "sale_reentry",
            "analysis_basis_label": "Pendiente 5M desestacionalizada sobre la senda 5A vigente",
            "sale_month_number": 12,
            "sale_year_number": 1,
            "sale_window_label": "abril 2027 (mes 12)",
            "sale_date": date(2027, 4, 17),
            "sale_date_label": "2027-04-17",
            "reentry_month_number": 26,
            "reentry_year_number": 3,
            "reentry_window_label": "junio 2028 (mes 26)",
            "reentry_date": date(2028, 6, 17),
            "reentry_date_label": "2028-06-17",
            "summary": "La pendiente desestacionalizada de 5 meses gira a negativo y luego vuelve a positivo.",
            "signal_label": "Pendiente 5M negativa",
            "signal_value_pct": Decimal("-0.13"),
            "monthly_rows": [],
            "yearly_rows": [],
            "drawdown_month_number": 12,
            "drawdown_year_number": 1,
            "drawdown_margin_pct": Decimal("-0.13"),
            "pre_sale_return_pct": Decimal("8.00"),
        }
        with (
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
            patch("equities.views.build_equity_investment_journey_context", return_value={"available": False}),
            patch("equities.services.build_owned_cycle_trade_timing_plan", return_value=trade_plan_payload),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Salida tactica")
        self.assertContains(response, "abril 2027 (mes 12)")
        self.assertContains(response, "Reentrada tactica")
        self.assertContains(response, "junio 2028 (mes 26)")
        self.assertContains(response, "Agenda de ventas 24M")
        self.assertContains(response, "Gantt tactico de salidas y neto estimado")
        self.assertContains(response, "Neto estimado 24M")
        self.assertContains(response, "Pendiente 5M")
        self.assertContains(response, "Rotacion radar")
        self.assertContains(response, "Rotar a ELE")

    def test_equities_page_backfills_purchase_baseline_for_existing_owned_position(self):
        position = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2025, 12, 15),
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.3000"),
        )
        populate_position_history(position, growth=Decimal("1.0180"), benchmark_growth=Decimal("1.0070"), months=48)
        alternative = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Seguimiento",
            ticker="ELE",
            quote_symbol="ELE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Endesa",
            shares=Decimal("0.0000"),
            average_cost_per_share=Decimal("18.0000"),
            current_price_per_share=Decimal("18.0000"),
        )
        populate_position_history(alternative, growth=Decimal("1.0220"), benchmark_growth=Decimal("1.0070"), months=48)
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=date(2026, 4, 17),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        cached_position = EquityPosition(
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        card = {
            "position": cached_position,
            "reference_label": "IBEX 35",
            "projection": {
                "available": True,
                "projected_price": Decimal("11.2000"),
                "base_return_pct": Decimal("12.00"),
                "safety_score": Decimal("68.00"),
            },
            "projection_reliability": {"label": "Alta", "score": Decimal("79.00")},
            "trade_alert": {"label": "Comprar"},
            "cycle_projection_5y": {
                "available": True,
                "path": [
                    {"label": "1A", "projected_price": Decimal("11.2000")},
                    {"label": "2A", "projected_price": Decimal("9.9000")},
                    {"label": "3A", "projected_price": Decimal("11.7000")},
                    {"label": "4A", "projected_price": Decimal("12.4000")},
                    {"label": "5A", "projected_price": Decimal("13.2000")},
                ],
            },
        }
        EquityNightlyAnalysisSnapshot.objects.create(
            run=run,
            analysis_date=run.analysis_date,
            scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
            analysis_key="ibex:IBE",
            ticker="IBE",
            quote_symbol="IBE.MC",
            company_name="Iberdrola",
            status_key="ibex",
            sector_label="Utilities",
            agent_provider="core",
            analysis_payload=serialize_cached_value(card),
        )
        cards = build_equity_history_cards(list(EquityPosition.objects.prefetch_related("price_history")))
        capture_equity_ticket_snapshots(
            [item for item in cards if item["position"].ticker == "IBE"],
            snapshot_date=date(2026, 4, 17),
        )

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        trade_plan_payload = {
            "available": True,
            "mode": "sale_reentry",
            "analysis_basis_label": "Pendiente 5M desestacionalizada sobre la senda 5A vigente",
            "sale_month_number": 12,
            "sale_year_number": 1,
            "sale_window_label": "abril 2027 (mes 12)",
            "sale_date": date(2027, 4, 17),
            "sale_date_label": "2027-04-17",
            "reentry_month_number": 26,
            "reentry_year_number": 3,
            "reentry_window_label": "junio 2028 (mes 26)",
            "reentry_date": date(2028, 6, 17),
            "reentry_date_label": "2028-06-17",
            "summary": "La pendiente desestacionalizada de 5 meses gira a negativo y luego vuelve a positivo.",
            "signal_label": "Pendiente 5M negativa",
            "signal_value_pct": Decimal("-0.13"),
            "monthly_rows": [],
            "yearly_rows": [],
            "drawdown_month_number": 12,
            "drawdown_year_number": 1,
            "drawdown_margin_pct": Decimal("-0.13"),
            "pre_sale_return_pct": Decimal("8.00"),
        }
        with (
            patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series),
            patch("equities.views.build_equity_investment_journey_context", return_value={"available": False}),
            patch("equities.services.build_owned_cycle_trade_timing_plan", return_value=trade_plan_payload),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        baseline = EquityPurchaseForecastBaseline.objects.get(position=position)
        self.assertEqual(baseline.source_analysis_date, date(2026, 4, 17))
        self.assertEqual(baseline.baseline_date, date(2026, 4, 17))
        self.assertContains(response, "Salida tactica")
        self.assertContains(response, "abril 2027 (mes 12)")

    def test_equities_page_can_render_round_investment_plan(self):
        owned = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("1000.0000"),
            average_cost_per_share=Decimal("20.0000"),
            current_price_per_share=Decimal("22.0000"),
        )
        watchlist = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Seguimiento",
            ticker="IDR",
            quote_symbol="IDR.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Indra",
            shares=Decimal("0.0000"),
            average_cost_per_share=Decimal("18.0000"),
            current_price_per_share=Decimal("18.0000"),
        )
        populate_position_history(owned, growth=Decimal("1.0180"), benchmark_growth=Decimal("1.0070"), months=48)
        populate_position_history(watchlist, growth=Decimal("1.0240"), benchmark_growth=Decimal("1.0070"), months=48)

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            response = self.client.get(
                f"{reverse('equities:list')}?round_target_total_capital=70000&round_max_round_amount=10000&round_max_company_pct=30"
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Plan por rondas")
        self.assertContains(response, "Despliegue escalonado")
        self.assertContains(response, "Paquete objetivo")
        self.assertContains(response, timezone.localdate().isoformat())
        self.assertContains(response, "martes y jueves")

    def test_equities_page_keeps_watchlist_analysis_separate_from_owned_tabs(self):
        owned = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        watchlist = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("0.0000"),
            average_cost_per_share=Decimal("38.0000"),
            current_price_per_share=Decimal("40.0000"),
        )

        for position, company_name, start_price, growth in (
            (owned, "Iberdrola", Decimal("10.0000"), Decimal("1.0180")),
            (watchlist, "ACS", Decimal("38.0000"), Decimal("1.0150")),
        ):
            stock_series = build_compound_market_series(
                position.quote_symbol,
                company_name,
                growth=growth,
                start_price=start_price,
            )
            reference_series = build_compound_market_series(
                "^IBEX",
                "IBEX 35",
                growth=Decimal("1.0070"),
                start_price=Decimal("100.0000"),
            )
            for stock_point, reference_point in zip(stock_series.points, reference_series.points):
                position.price_history.create(
                    price_date=stock_point["date"],
                    open_price=stock_point["open"],
                    high_price=stock_point["high"],
                    low_price=stock_point["low"],
                    close_price=stock_point["close"],
                    benchmark_close=reference_point["close"],
                )

        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )
        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        page = response.content.decode("utf-8")
        self.assertContains(response, "Seguimientos guardados")
        self.assertContains(response, 'id="equity-analysis"', html=False)
        self.assertContains(response, 'aria-label="Seguimientos guardados"', html=False)
        self.assertContains(response, f'id="stock-tab-{watchlist.id}"', html=False)
        self.assertNotContains(response, f'id="stock-tab-{owned.id}"', html=False)
        self.assertIn(f'href="#tracked-ticket-{owned.id}"', page)

    def test_equities_page_highlights_visible_12m_thesis_on_analysis_cards(self):
        position = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        populate_position_history(position, growth=Decimal("1.0180"), benchmark_growth=Decimal("1.0070"), months=48)
        benchmark_series = build_compound_market_series(
            "^IBEX",
            "IBEX 35",
            growth=Decimal("1.0060"),
            start_price=Decimal("100.0000"),
        )

        with patch("equities.services.fetch_reference_series_for_choice", return_value=benchmark_series):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tesis 12M visible")
        self.assertContains(response, "Cierre 12M visible")
        self.assertContains(response, "Rango 12M del modelo")
        self.assertContains(response, "Alerta cuantitativa")

    def test_equities_page_groups_major_sections_into_page_tabs(self):
        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-equity-page-tabs', html=False)
        self.assertContains(response, 'id="equity-page-tab-cartera"', html=False)
        self.assertContains(response, 'id="equity-page-tab-mercado"', html=False)
        self.assertContains(response, 'id="equity-page-tab-bondad"', html=False)
        self.assertContains(response, 'id="equity-page-tab-optimizacion"', html=False)
        self.assertContains(response, 'id="equity-page-tab-seguimiento"', html=False)
        self.assertContains(response, 'id="equity-page-tab-operativa"', html=False)
        self.assertContains(response, 'id="equity-view-cartera"', html=False)
        self.assertContains(response, 'id="equity-view-mercado"', html=False)
        self.assertContains(response, 'id="equity-view-bondad"', html=False)
        self.assertContains(response, 'id="equity-view-optimizacion"', html=False)
        self.assertContains(response, 'id="equity-view-seguimiento"', html=False)
        self.assertContains(response, 'id="equity-view-operativa"', html=False)
        self.assertContains(response, 'id="equity-decision"', html=False)
        self.assertContains(response, 'id="equity-bondad"', html=False)
        self.assertContains(response, 'id="equity-optimizer"', html=False)
        self.assertContains(response, 'id="equity-ops"', html=False)
        self.assertContains(response, "Cartera")
        self.assertContains(response, "Mercado")
        self.assertContains(response, "Bondad")
        self.assertContains(response, "Optimizacion")
        self.assertContains(response, "Seguimiento")
        self.assertContains(response, "Operativa")

    def test_equities_page_renders_expectation_review_tabs(self):
        expectation_dashboard = {
            "available": True,
            "companies_count": 1,
            "reviews_count": 4,
            "equation_ready_count": 1,
            "last_review_date_label": "2026-04-24",
            "scope_note": "Comparativa programada.",
            "equation_note": "Ecuacion correctora activa.",
            "companies": [
                {
                    "available": True,
                    "ticker": "ANA",
                    "company_name": "Acciona",
                    "tab_key": "bondad-ana",
                    "reviews_count": 4,
                    "matured_reviews_count": 1,
                    "current_price": Decimal("118.0000"),
                    "current_price_date_label": "2026-04-26",
                    "average_gap_pct": Decimal("-2.50"),
                    "average_expected_return_pct": Decimal("6.00"),
                    "average_actual_return_pct": Decimal("3.50"),
                    "latest_row": {
                        "analysis_date_label": "2026-04-24",
                        "expected_return_pct": Decimal("1.20"),
                        "actual_return_pct": Decimal("-0.80"),
                        "gap_pct": Decimal("-2.00"),
                        "window_label": "2 dias",
                    },
                    "chart": {
                        "available": True,
                        "actual_line": "18,180 622,120",
                        "expected_line": "18,170 622,132",
                        "benchmark_line": "18,176 622,126",
                        "has_actual_series": True,
                        "has_expected_series": True,
                        "has_benchmark_series": True,
                        "actual_display_points": [],
                        "expected_display_points": [],
                        "benchmark_display_points": [],
                        "x_markers": [],
                        "grid_markers": [],
                        "min_label": "-4.0 %",
                        "max_label": "8.0 %",
                        "start_label": "2026-01-15",
                        "latest_label": "2026-04-24",
                        "zero_y": "146.0",
                    },
                    "equation": {
                        "available": True,
                        "formula_label": "Real observado ~= -1.0 + 0.80 x Esperanza observada",
                        "short_formula_label": "-1.0 + 0.80x",
                        "interpretation": "La correccion enfria el sesgo.",
                        "sample_count": 4,
                        "r_squared": Decimal("0.72"),
                    },
                    "preview_mode": False,
                    "preview_note": "",
                    "rows": [
                        {
                            "analysis_date_label": "2026-04-24",
                            "window_end_date_label": "2026-04-26",
                            "window_label": "2 dias",
                            "matured": False,
                            "expected_return_pct": Decimal("1.20"),
                            "corrected_return_pct": Decimal("0.00"),
                            "actual_return_pct": Decimal("-0.80"),
                            "gap_pct": Decimal("-2.00"),
                        }
                    ],
                    "market_error": "",
                }
            ],
        }

        with patch("equities.views.build_expectation_review_dashboard", return_value=expectation_dashboard):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Como se esta comportando la esperanza frente a la realidad")
        self.assertContains(response, 'id="bondad-tab-bondad-ana"', html=False)
        self.assertContains(response, "Real observado")
        self.assertContains(response, "Esperanza corregida")
        self.assertContains(response, "Real observado ~= -1.0 + 0.80 x Esperanza observada")

    def test_equities_page_renders_expectation_preview_chart_when_real_history_is_pending(self):
        expectation_dashboard = {
            "available": True,
            "companies_count": 1,
            "reviews_count": 1,
            "equation_ready_count": 0,
            "last_review_date_label": "2026-04-26",
            "scope_note": "Comparativa programada.",
            "equation_note": "Ecuacion correctora pendiente.",
            "companies": [
                {
                    "available": True,
                    "ticker": "CABK",
                    "company_name": "CaixaBank",
                    "tab_key": "bondad-cabk",
                    "reviews_count": 1,
                    "matured_reviews_count": 0,
                    "current_price": Decimal("10.3700"),
                    "current_price_date_label": "2026-04-26",
                    "average_gap_pct": None,
                    "average_expected_return_pct": Decimal("30.00"),
                    "average_actual_return_pct": None,
                    "latest_row": None,
                    "chart": {
                        "available": True,
                        "actual_line": "",
                        "expected_line": "18,180 180,150 320,110 470,72 622,40",
                        "benchmark_line": "",
                        "has_actual_series": False,
                        "has_expected_series": True,
                        "has_benchmark_series": False,
                        "actual_display_points": [],
                        "expected_display_points": [],
                        "benchmark_display_points": [],
                        "x_markers": [],
                        "grid_markers": [],
                        "min_label": "-1.0 %",
                        "max_label": "58.0 %",
                        "start_label": "2026-04-26",
                        "latest_label": "2026-04-26",
                        "projection_end_label": "2031-04-26",
                        "zero_y": "198.0",
                    },
                    "equation": {
                        "available": False,
                        "formula_label": "Sin muestras suficientes",
                        "short_formula_label": "Sin ecuacion",
                        "interpretation": "Pendiente de mas muestras.",
                        "sample_count": 0,
                        "r_squared": None,
                    },
                    "preview_mode": True,
                    "preview_note": "Todavia no hay realidad suficiente para comparar la bondad. Se muestra la ultima esperanza guardada con sus hitos 1A..5A.",
                    "rows": [],
                    "market_error": "",
                }
            ],
        }

        with patch("equities.views.build_expectation_review_dashboard", return_value=expectation_dashboard):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vista previa de esperanza")
        self.assertContains(response, "Real pendiente")
        self.assertContains(response, "Esperanza prevista")

    def test_equities_page_renders_investment_journey_section(self):
        active = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2024, 1, 15),
            shares=Decimal("20.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
        )
        sold = EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Banco Sabadell",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            opened_on=date(2023, 2, 1),
            shares=Decimal("15.0000"),
            average_cost_per_share=Decimal("20.0000"),
            current_price_per_share=Decimal("24.0000"),
        )
        for position, start_price in ((active, Decimal("10.00")), (sold, Decimal("20.00"))):
            stock_series = build_compound_market_series(
                position.quote_symbol,
                position.company_name,
                growth=Decimal("1.0100"),
                start_price=start_price,
            )
            reference_series = build_compound_market_series(
                "^IBEX",
                "IBEX 35",
                growth=Decimal("1.0060"),
                start_price=Decimal("100.0000"),
            )
            for stock_point, reference_point in zip(stock_series.points, reference_series.points):
                position.price_history.create(
                    price_date=stock_point["date"],
                    open_price=stock_point["open"],
                    high_price=stock_point["high"],
                    low_price=stock_point["low"],
                    close_price=stock_point["close"],
                    benchmark_close=reference_point["close"],
                )
        archive_equity_position_sale(
            sold,
            closed_on=date(2025, 9, 30),
            sale_price_per_share=Decimal("25.0000"),
            notes="Cierre completo",
        )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cuadro de Gestion")
        self.assertContains(response, "Rentabilidad anual por ejercicio")
        self.assertContains(response, "Cuenta de resultados")
        self.assertContains(response, "Resultado neto comparable de las posiciones cerradas")
        self.assertContains(response, "Rentabilidad comparable por ticket abierto")
        self.assertContains(response, "Salida abierta estimada")
        self.assertContains(response, "Como leer los costes:")

    def test_equities_page_surfaces_sale_simulator_and_unfollow_action(self):
        EquityPosition.objects.create(
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Banco Santander",
            trade_channel=EquityPosition.TradeChannel.APP,
            ticker="IBE",
            quote_symbol="IBE.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Iberdrola",
            opened_on=date(2025, 1, 10),
            shares=Decimal("10.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("12.0000"),
            annual_dividend_income=Decimal("12.00"),
            annual_maintenance_cost=Decimal("5.00"),
        )
        EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Seguimiento",
            ticker="ACS",
            quote_symbol="ACS.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="ACS",
            shares=Decimal("0"),
            average_cost_per_share=Decimal("0"),
            current_price_per_share=Decimal("0"),
        )

        response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Si vendes esta posicion, aqui ves el neto real")
        self.assertContains(response, "Dejar de seguir y volver al radar")
        self.assertContains(response, "Si vendo hoy")
        self.assertContains(response, "Mes eq.")
        self.assertContains(response, "Ano eq.")
        self.assertContains(response, "Calcular / vender")

    @override_settings(EQUITIES_REFERENCE_WORKBOOK="")
    def test_equities_page_renders_workbook_reference_guide(self):
        workbook_path = build_test_reference_workbook()
        self.addCleanup(lambda: os.path.exists(workbook_path) and os.remove(workbook_path))
        load_ibex_reference_workbook_snapshot.cache_clear()

        with override_settings(EQUITIES_REFERENCE_WORKBOOK=workbook_path):
            EquityPosition.objects.create(
                ownership_category=AssetOwnershipCategory.JOINT,
                broker="Banco Sabadell",
                ticker="SAN",
                quote_symbol="SAN.MC",
                benchmark_symbol="^IBEX",
                benchmark_name="IBEX 35",
                company_name="Banco Santander",
                shares=Decimal("10.0000"),
                average_cost_per_share=Decimal("4.0000"),
                current_price_per_share=Decimal("5.0000"),
            )

            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Guia completa IBEX del Excel")
        self.assertContains(response, "Euribor 12m (%)")

    @override_settings(EQUITIES_IBEX_UNIVERSE_ANALYSIS=True, EQUITIES_IBEX_UNIVERSE_LIMIT=1)
    def test_ibex_rows_open_detail_in_new_tab(self):
        acs = find_equity_company_profile("ACS")
        company = {
            "ticker": acs["ticker"],
            "company_name": acs["company_name"],
            "quote_symbol": acs["quote_symbol"],
            "sector": acs["sector_label"],
            "per_2025": Decimal("11.50"),
            "dividend_yield": Decimal("3.10"),
            "catalog_profile": acs,
        }
        workbook_snapshot = {
            "available": True,
            "path": "",
            "companies": [company],
            "companies_by_key": {
                acs["ticker"]: company,
                acs["company_name"].upper(): company,
                acs["quote_symbol"].upper().replace(".", " "): company,
            },
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        def fake_market_series(symbol, range_key="10y", interval="1d"):
            return build_compound_market_series(symbol, symbol, growth=Decimal("1.0200"), start_price=Decimal("12.0000"))

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            return build_compound_market_series(
                benchmark_symbol or "^IBEX",
                benchmark_name or "Referencia",
                growth=Decimal("1.0060"),
                start_price=Decimal("100.0000"),
            )

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=workbook_snapshot),
            patch("equities.services.build_ibex_universe_companies", return_value=[company]),
            patch("equities.services.fetch_market_series", side_effect=fake_market_series),
            patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("equities:ibex_detail", kwargs={"ticker": acs["ticker"]}))
        self.assertContains(response, 'target="_blank"', html=False)
        self.assertContains(response, "Abrir analisis completo")

    @override_settings(EQUITIES_IBEX_UNIVERSE_ANALYSIS=True, EQUITIES_IBEX_UNIVERSE_LIMIT=1)
    def test_ibex_table_shows_five_year_projection_columns(self):
        santander = find_equity_company_profile("Banco Santander")
        company = {
            "ticker": santander["ticker"],
            "company_name": santander["company_name"],
            "quote_symbol": santander["quote_symbol"],
            "sector": "Banca",
            "per_2025": Decimal("11.50"),
            "dividend_yield": Decimal("3.10"),
            "catalog_profile": santander,
        }
        workbook_snapshot = {
            "available": True,
            "path": "",
            "companies": [company],
            "companies_by_key": {
                santander["ticker"]: company,
                santander["company_name"].upper(): company,
                santander["quote_symbol"].upper().replace(".", " "): company,
            },
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        def fake_market_series(symbol, range_key="10y", interval="1d"):
            return build_compound_market_series(symbol, symbol, growth=Decimal("1.0200"), months=120, start_price=Decimal("12.0000"))

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            return build_compound_market_series(
                benchmark_symbol or "^IBEX",
                benchmark_name or "Referencia",
                growth=Decimal("1.0060"),
                months=120,
                start_price=Decimal("100.0000"),
            )

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=workbook_snapshot),
            patch("equities.services.build_ibex_universe_companies", return_value=[company]),
            patch("equities.services.should_fetch_equity_fundamentals", return_value=False),
            patch("equities.services.fetch_market_series", side_effect=fake_market_series),
            patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Relacion entre precio y beneficio")
        self.assertContains(response, "Valoracion")
        self.assertContains(response, "Esperanza 1A")
        self.assertContains(response, "Esperanza 2A")
        self.assertContains(response, "Esperanza 3A")
        self.assertContains(response, "Esperanza 4A")
        self.assertContains(response, "Pred. 5 AÑOS")
        self.assertContains(response, "Esperanza 5A")
        self.assertContains(response, "Márgenes por AÑO")
        self.assertContains(response, "AÑO 1")
        self.assertContains(response, "AÑO 5")
        self.assertContains(response, "Razonable")
        self.assertContains(response, 'class="neutral">11,5</strong>', html=False)

    @override_settings(EQUITIES_IBEX_UNIVERSE_ANALYSIS=True, EQUITIES_IBEX_UNIVERSE_LIMIT=1)
    def test_ibex_table_shows_buy_and_sell_recommendation_dates(self):
        acs = find_equity_company_profile("ACS")
        company = {
            "ticker": acs["ticker"],
            "company_name": acs["company_name"],
            "quote_symbol": acs["quote_symbol"],
            "sector": acs["sector_label"],
            "dividend_yield": Decimal("3.10"),
            "catalog_profile": acs,
        }
        empty_workbook = {
            "available": False,
            "path": "",
            "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        for analysis_day, label in (
            (date(2026, 4, 14), "Vigilar"),
            (date(2026, 4, 15), "Comprar"),
            (date(2026, 4, 16), "Comprar"),
            (date(2026, 4, 17), "Vender"),
        ):
            run = EquityNightlyAnalysisRun.objects.create(
                analysis_date=analysis_day,
                status=EquityNightlyAnalysisRun.Status.COMPLETED,
                started_at=timezone.now(),
                completed_at=timezone.now(),
            )
            EquityNightlyAnalysisSnapshot.objects.create(
                run=run,
                analysis_date=analysis_day,
                scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
                analysis_key=f"ibex:ACS:{analysis_day.isoformat()}",
                ticker="ACS",
                quote_symbol="ACS.MC",
                company_name="ACS",
                status_key="ibex",
                sector_label="Construccion",
                agent_provider="core",
                analysis_payload=serialize_cached_value({"trade_alert": {"label": label}}),
            )

        def fake_market_series(symbol, range_key="10y", interval="1d"):
            return build_compound_market_series(symbol, symbol, growth=Decimal("1.0200"), months=120, start_price=Decimal("12.0000"))

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            return build_compound_market_series(
                benchmark_symbol or "^IBEX",
                benchmark_name or "Referencia",
                growth=Decimal("1.0060"),
                months=120,
                start_price=Decimal("100.0000"),
            )

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=empty_workbook),
            patch("equities.services.build_ibex_universe_companies", return_value=[company]),
            patch("equities.services.fetch_market_series", side_effect=fake_market_series),
            patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series),
        ):
            response = self.client.get(reverse("equities:list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Compra recom.")
        self.assertContains(response, "Venta recom.")
        self.assertContains(response, "2026-04-15")
        self.assertContains(response, "2026-04-17")

    def test_can_open_ibex_detail_page_with_company_title_and_charts(self):
        acs = find_equity_company_profile("ACS")
        company = {
            "ticker": acs["ticker"],
            "company_name": acs["company_name"],
            "quote_symbol": acs["quote_symbol"],
            "sector": acs["sector_label"],
            "dividend_yield": Decimal("3.10"),
            "catalog_profile": acs,
        }
        empty_workbook = {
            "available": False,
            "path": "",
            "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        def fake_market_series(symbol, range_key="10y", interval="1d"):
            return build_compound_market_series(symbol, symbol, growth=Decimal("1.0200"), start_price=Decimal("12.0000"))

        def fake_reference_series(reference_profile, benchmark_symbol="", benchmark_name="", range_key="10y"):
            return build_compound_market_series(
                benchmark_symbol or "^IBEX",
                benchmark_name or "Referencia",
                growth=Decimal("1.0060"),
                start_price=Decimal("100.0000"),
            )

        with (
            patch("equities.services.load_ibex_reference_workbook_snapshot", return_value=empty_workbook),
            patch("equities.services.build_ibex_universe_companies", return_value=[company]),
            patch("equities.services.fetch_market_series", side_effect=fake_market_series),
            patch("equities.services.fetch_reference_series_for_choice", side_effect=fake_reference_series),
        ):
            response = self.client.get(reverse("equities:ibex_detail", kwargs={"ticker": acs["ticker"]}))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "equities/ibex_equity_detail.html")
        self.assertContains(response, f"<title>{acs['company_name']}</title>", html=False)
        self.assertContains(response, "Historico")
        self.assertContains(response, "Mejor correlacion")
        self.assertContains(response, "Prevision 12M")
        self.assertContains(response, "Ciclo 5A")
        self.assertContains(response, "Esperanza 12M")
        self.assertContains(response, "Esperanza 5A")

    def test_ibex_detail_page_uses_visible_12m_hero_values_when_chart_closes_lower(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ANA",
            quote_symbol="ANA.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Acciona",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("10.0000"),
            current_price_per_share=Decimal("10.0000"),
        )
        populate_position_history(position, growth=Decimal("1.0100"), benchmark_growth=Decimal("1.0040"), months=36)
        card = build_equity_history_cards([position])[0]
        latest_date = card["end_date"]
        card["projection"].update(
            {
                "projected_price": Decimal("12.8000"),
                "price_return_pct": Decimal("28.00"),
                "base_return_pct": Decimal("29.20"),
                "net_income_yield_pct": Decimal("1.50"),
                "transaction_drag_pct": Decimal("0.30"),
                "monthly_path": [
                    {
                        "label": "3M",
                        "projected_date": add_calendar_months(latest_date, 3),
                        "projected_price": Decimal("9.7000"),
                    },
                    {
                        "label": "6M",
                        "projected_date": add_calendar_months(latest_date, 6),
                        "projected_price": Decimal("9.3000"),
                    },
                    {
                        "label": "9M",
                        "projected_date": add_calendar_months(latest_date, 9),
                        "projected_price": Decimal("9.0000"),
                    },
                    {
                        "label": "12M",
                        "projected_date": add_calendar_months(latest_date, 12),
                        "projected_price": Decimal("8.8000"),
                    },
                ],
                "quarterly_path": [],
            }
        )
        card["trade_alert"] = {
            **card["trade_alert"],
            "label": "Comprar",
            "tone": "buy",
            "score": Decimal("4.25"),
            "trigger_label": "3 meses con alpha positiva",
            "note": "La pendiente relativa todavia apoya compras.",
        }
        refresh_card_projection_visuals(card)
        company = {
            "ticker": "ANA",
            "company_name": "Acciona",
            "quote_symbol": "ANA.MC",
            "sector": "Infraestructuras",
        }

        with (
            patch("equities.views.find_ibex_universe_company", return_value=(company, {"available": False})),
            patch("equities.views.load_cached_ibex_card", return_value=card),
        ):
            response = self.client.get(reverse("equities:ibex_detail", kwargs={"ticker": "ANA"}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vigilar")
        self.assertContains(response, "La senda visible 12M sigue bajista")
        self.assertNotContains(response, "29.2 %")
        self.assertNotContains(response, "Precio objetivo 12.8000")
        self.assertLess(response.context["card"]["presentation_projection"]["visible_total_return_pct"], ZERO)
        self.assertEqual(response.context["card"]["trade_alert"]["label"], "Vigilar")
        self.assertRegex(response.content.decode(), r"Cierre visible\s+8.8000")

    def test_load_cached_ibex_card_refreshes_stale_snapshot_with_live_position_visuals(self):
        position = EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.WATCHLIST,
            ownership_category=AssetOwnershipCategory.JOINT,
            broker="Interactive Brokers",
            ticker="ANA",
            quote_symbol="ANA.MC",
            benchmark_symbol="^IBEX",
            benchmark_name="IBEX 35",
            company_name="Acciona",
            shares=Decimal("1.0000"),
            average_cost_per_share=Decimal("235.0000"),
            current_price_per_share=Decimal("235.0000"),
        )
        populate_position_history(position, growth=Decimal("1.0100"), benchmark_growth=Decimal("1.0040"), months=36)
        stale_card = build_equity_history_cards([position])[0]
        latest_date = stale_card["end_date"]
        stale_card["projection"].update(
            {
                "projected_price": Decimal("329.2721"),
                "price_return_pct": Decimal("40.30"),
                "base_return_pct": Decimal("40.30"),
                "net_income_yield_pct": Decimal("0.00"),
                "transaction_drag_pct": Decimal("0.00"),
                "monthly_path": [
                    {
                        "label": "3M",
                        "projected_date": add_calendar_months(latest_date, 3),
                        "projected_price": Decimal("224.0000"),
                    },
                    {
                        "label": "6M",
                        "projected_date": add_calendar_months(latest_date, 6),
                        "projected_price": Decimal("212.0000"),
                    },
                    {
                        "label": "9M",
                        "projected_date": add_calendar_months(latest_date, 9),
                        "projected_price": Decimal("201.0000"),
                    },
                    {
                        "label": "12M",
                        "projected_date": add_calendar_months(latest_date, 12),
                        "projected_price": Decimal("188.0000"),
                    },
                ],
                "quarterly_path": [],
            }
        )
        stale_card["presentation_projection"] = {"available": False}
        stale_card["trade_alert"] = {
            **stale_card["trade_alert"],
            "label": "Comprar",
            "tone": "buy",
            "score": Decimal("4.25"),
            "trigger_label": "6 trimestres con alpha positiva",
            "note": "Snapshot antiguo sin refrescar.",
        }
        run = EquityNightlyAnalysisRun.objects.create(
            analysis_date=timezone.localdate(),
            status=EquityNightlyAnalysisRun.Status.COMPLETED,
            started_at=timezone.now(),
            completed_at=timezone.now(),
            summary_data=serialize_cached_value(
                {"tracked_signature": build_positions_analysis_signature([position])}
            ),
        )
        EquityNightlyAnalysisSnapshot.objects.create(
            run=run,
            analysis_date=timezone.localdate(),
            scope=EquityNightlyAnalysisSnapshot.Scope.IBEX,
            analysis_key="ibex:ANA:test",
            position=position,
            ticker="ANA",
            quote_symbol="ANA.MC",
            company_name="Acciona",
            status_key="ibex",
            sector_label="Infraestructuras",
            agent_provider="core",
            analysis_payload=serialize_cached_value(stale_card),
        )

        refreshed_card = load_cached_ibex_card("ANA", [position])

        self.assertIsNotNone(refreshed_card)
        self.assertTrue(refreshed_card["presentation_projection"]["available"])
        self.assertLess(refreshed_card["presentation_projection"]["visible_total_return_pct"], ZERO)
        self.assertEqual(refreshed_card["trade_alert"]["label"], "Vigilar")
        self.assertEqual(refreshed_card["trade_alert"]["trigger_label"], "La senda visible 12M sigue bajista")
        self.assertEqual(
            refreshed_card["presentation_projection"]["visible_projected_price"],
            Decimal("188.0000"),
        )

    def test_can_store_same_ticker_as_owned_and_watchlist_without_collision(self):
        EquityPosition.objects.create(
            position_kind=EquityPosition.PositionKind.OWNED,
            ownership_category=AssetOwnershipCategory.XIMO,
            broker="Interactive Brokers",
            ticker="SAN",
            quote_symbol="SAN.MC",
            reference_profile=EquityPosition.ReferenceProfile.EURIBOR_12M,
            benchmark_symbol="ECB:M.S0.N.C_EUR1Y.E",
            benchmark_name="Euribor 12M",
            company_name="Banco Santander",
            shares=Decimal("30.0000"),
            average_cost_per_share=Decimal("4.2000"),
            current_price_per_share=Decimal("4.7000"),
        )

        response = self.client.post(
            reverse("equities:list"),
            {
                "action": "create_position",
                "position_kind": EquityPosition.PositionKind.WATCHLIST,
                "ownership_category": AssetOwnershipCategory.XIMO,
                "broker": "Interactive Brokers",
                "ticker": "SAN",
                "company_name": "Banco Santander",
                "quote_symbol": "SAN.MC",
                "reference_profile": EquityPosition.ReferenceProfile.EURIBOR_12M,
                "benchmark_symbol": "",
                "benchmark_name": "",
                "shares": "0",
                "average_cost_per_share": "4,5000",
                "current_price_per_share": "4,7000",
                "annual_dividend_income": "0",
                "annual_maintenance_cost": "0",
                "notes": "Seguimiento",
            },
        )

        self.assertRedirects(response, reverse("equities:list"))
        self.assertEqual(
            EquityPosition.objects.filter(broker="Interactive Brokers", ticker="SAN", ownership_category=AssetOwnershipCategory.XIMO).count(),
            2,
        )
