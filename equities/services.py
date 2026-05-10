from __future__ import annotations

import logging
from calendar import monthrange
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from itertools import combinations
import json
import math
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, localcontext
from functools import lru_cache
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.utils import OperationalError, ProgrammingError
from django.utils import timezone as django_timezone

from banking.services import load_rows_from_workbook
from portfolio.ownership import AssetOwnershipCategory

from .broker_costs import estimate_broker_costs, resolve_recurring_cost_used
from .models import (
    EquityClosedPosition,
    EquityExpectationReview,
    EquityPosition,
    EquityPriceHistory,
    EquityPurchaseForecastBaseline,
    EquityTicketSnapshot,
)


ZERO = Decimal("0.00")
ONE_HUNDRED = Decimal("100")
DEFAULT_MARKET_RANGE_KEY = "10y"
MAX_MARKET_RANGE_KEY = "max"
LONG_ANALYSIS_YEARS = 10
LONG_ANALYSIS_DAYS = 365 * LONG_ANALYSIS_YEARS
LONG_MONTHLY_OBSERVATIONS = 132
MONTHLY_CORRELATION_WINDOW = 120
QUARTERLY_CORRELATION_WINDOW = 40
MONTHLY_RECENT_WINDOW = 12
QUARTERLY_RECENT_WINDOW = 4
TRACKING_HORIZON_DAYS = 365
TRACKING_FORECAST_MARKERS = (91, 182, 273, TRACKING_HORIZON_DAYS)
TRACKING_FIVE_YEAR_MARKERS = (1, 2, 3, 4, 5)
TRACKING_MARKER_WEEKDAYS = {0, 3}
TRACKING_WEEKLY_ALPHA_DAYS = 7
TRACKING_EXPECTED_FEEDBACK_MIN_DAYS = 5
TRACKING_EXPECTED_FEEDBACK_FULL_STRENGTH_DAYS = 21
TRACKING_EXPECTED_FEEDBACK_MIN_GAP_PCT = Decimal("1.25")
TRACKING_EXPECTED_FEEDBACK_MAX_WEIGHT = Decimal("0.65")
TRACKING_EXPECTED_FEEDBACK_TARGET_CARRY = Decimal("0.45")
TRACKING_TARGET_MIN_EXCESS_PCT = Decimal("0.75")
TRACKING_TARGET_MAX_ANNUAL_CAP_12M = Decimal("14.00")
TRACKING_TARGET_MAX_ANNUAL_CAP_5Y = Decimal("10.50")
TRACKING_TARGET_MIN_EXCESS_KEEP_RATIO = Decimal("0.18")
TRACKING_TARGET_MAX_EXCESS_KEEP_RATIO = Decimal("0.44")
PORTFOLIO_PROJECTION_HORIZONS = (
    (3, "3M"),
    (6, "6M"),
    (9, "9M"),
    (12, "12M"),
)
PORTFOLIO_EXPECTATION_HORIZONS = (
    (12, "1A", "expected_return_1y_pct"),
    (24, "2A", "expected_return_2y_pct"),
    (36, "3A", "expected_return_3y_pct"),
    (48, "4A", "expected_return_4y_pct"),
    (60, "5A", "expected_return_5y_pct"),
)
PORTFOLIO_CORRELATION_MIN_COMMON_PERIODS = 6
DALIO_CORRELATION_RISK_POINTS = (
    (Decimal("0"), Decimal("11")),
    (Decimal("10"), Decimal("24")),
    (Decimal("20"), Decimal("31")),
    (Decimal("40"), Decimal("34")),
    (Decimal("60"), Decimal("38")),
)
COMPARABLE_MONTH_DAYS = Decimal("30.4375")
OPTIMIZER_MAX_ENTRY_DRAG_PCT = Decimal("1.00")
OPTIMIZER_MAX_ROUNDTRIP_DRAG_PCT = Decimal("2.00")
OPTIMIZER_MIN_GAIN_TO_ROUNDTRIP_MULTIPLE = Decimal("1.80")
OPTIMIZER_TARGET_HOLDING_ANNUALIZED_RETURN_PCT = Decimal("20.00")
OPTIMIZER_TARGET_SHORTFALL_PENALTY_MULTIPLE = Decimal("0.50")
OPTIMIZER_TARGET_EXCESS_BONUS_MULTIPLE = Decimal("0.12")
OPTIMIZER_RISK_PROFILE_LABEL = "Adverso al riesgo"
OPTIMIZER_CONSERVATIVE_MIN_SAFETY_SCORE = Decimal("68.00")
OPTIMIZER_CONSERVATIVE_MIN_RELIABILITY_SCORE = Decimal("64.00")
OPTIMIZER_CONSERVATIVE_MIN_WORST_RETURN_PCT = Decimal("-12.00")
OPTIMIZER_CONSERVATIVE_MIN_STRESS_RETURN_PCT = Decimal("-10.00")
OPTIMIZER_CONSERVATIVE_MAX_VOLATILITY_PCT = Decimal("22.00")
OPTIMIZER_CONSERVATIVE_MAX_UNCERTAINTY_PENALTY_PCT = Decimal("3.25")
OPTIMIZER_EXPECTATION_REVIEW_RECENT_POINTS = 6
OPTIMIZER_EXPECTATION_REVIEW_MEMORY_POINTS = 156
OPTIMIZER_EXPECTATION_REVIEW_MEMORY_HALF_LIFE_DAYS = Decimal("120")
OPTIMIZER_EXPECTATION_REVIEW_MIN_MEMORY_WEIGHT = Decimal("0.08")
OPTIMIZER_EXPECTATION_REVIEW_STABLE_SPREAD_PCT = Decimal("10.00")
OPTIMIZER_EXPECTATION_REVIEW_MAX_BONUS_PCT = Decimal("3.00")
OPTIMIZER_EXPECTATION_REVIEW_MAX_PENALTY_PCT = Decimal("6.00")
EXPECTATION_STABILITY_1Y_MIN_MOVE_PCT = Decimal("1.25")
EXPECTATION_STABILITY_1Y_DAILY_MOVE_PCT = Decimal("0.90")
EXPECTATION_STABILITY_1Y_MAX_MOVE_PCT = Decimal("3.00")
EXPECTATION_STABILITY_5Y_MIN_MOVE_PCT = Decimal("1.00")
EXPECTATION_STABILITY_5Y_DAILY_MOVE_PCT = Decimal("0.55")
EXPECTATION_STABILITY_5Y_MAX_MOVE_PCT = Decimal("2.50")
EXPECTATION_REVIEW_CORRECTION_MIN_POINTS = 3
EXPECTATION_REVIEW_CORRECTION_MIN_ELAPSED_DAYS = 7
PURCHASE_DISCIPLINE_TARGET_SCORE = Decimal("70.00")
REFERENCE_CYCLE_TEMPLATE_RECENT_MONTHS = 18
REFERENCE_CYCLE_TEMPLATE_MAX_MATCHES = 12

logger = logging.getLogger(__name__)
SPANISH_MONTH_LABELS = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}
DEFAULT_EQUITY_ANALYSIS_NOTIONAL = Decimal("10000.00")
DEFAULT_BENCHMARK_SYMBOL = "^IBEX"
DEFAULT_BENCHMARK_NAME = "IBEX 35"
DEFAULT_MARKET_REQUEST_TIMEOUT_SECONDS = 12
DEFAULT_MARKET_DATA_CACHE_MINUTES = 60
DEFAULT_IBEX_UNIVERSE_MAX_WORKERS = 8
SCHEDULED_REVIEW_WEEKDAY_LABELS = {
    1: "lunes",
    2: "martes",
    3: "miercoles",
    4: "jueves",
    5: "viernes",
    6: "sabado",
    7: "domingo",
}
EURIBOR_REFERENCE_SYMBOL = "ECB:M.S0.N.C_EUR1Y.E"
EURIBOR_REFERENCE_NAME = "Euribor 12M"
EURIBOR_FORWARD_REFERENCE_SYMBOL = "ECB:YC.B.U2.EUR.4F.G_N_C.SV_C_YM.IF_5Y"
EURIBOR_FORWARD_REFERENCE_NAME = "Curva BCE forward 5A"
SPAIN_HOUSE_PRICE_SYMBOL = "EUROSTAT:prc_hpi_q:ES:TOTAL:I15_Q"
SPAIN_HOUSE_PRICE_NAME = "Precio vivienda Espana"
SPAIN_ELECTRICITY_DEMAND_SYMBOL = "REE:demand:es:peninsular"
SPAIN_ELECTRICITY_DEMAND_NAME = "Demanda electrica Espana"
SPAIN_GAS_CONSUMPTION_SYMBOL = "EUROSTAT:nrg_cb_gasm:ES:IC_OBS:TJ_GCV"
SPAIN_GAS_CONSUMPTION_NAME = "Consumo de gas Espana"
BRENT_REFERENCE_SYMBOL = "BZ=F"
BRENT_REFERENCE_NAME = "Petroleo Brent"
COPPER_REFERENCE_SYMBOL = "HG=F"
COPPER_REFERENCE_NAME = "Cobre"
EURUSD_REFERENCE_SYMBOL = "EURUSD=X"
EURUSD_REFERENCE_NAME = "Euro / Dolar"
DEFAULT_EQUITY_COLUMN_MAP = {
    "broker": 0,
    "ticker": 1,
    "company_name": 2,
    "shares": 3,
    "average_cost_per_share": 4,
    "current_price_per_share": 5,
}
REFERENCE_PRESETS = {
    "ibex_35": {
        "key": "ibex_35",
        "label": DEFAULT_BENCHMARK_NAME,
        "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
        "benchmark_symbol": DEFAULT_BENCHMARK_SYMBOL,
        "benchmark_name": DEFAULT_BENCHMARK_NAME,
        "description": "Pulso general de la bolsa espanola.",
    },
    "euribor_12m": {
        "key": "euribor_12m",
        "label": EURIBOR_REFERENCE_NAME,
        "reference_profile": EquityPosition.ReferenceProfile.EURIBOR_12M,
        "benchmark_symbol": EURIBOR_REFERENCE_SYMBOL,
        "benchmark_name": EURIBOR_REFERENCE_NAME,
        "description": "Muy util para bancos, deuda y negocios sensibles a tipos.",
    },
    "spain_house_price": {
        "key": "spain_house_price",
        "label": SPAIN_HOUSE_PRICE_NAME,
        "reference_profile": EquityPosition.ReferenceProfile.SPAIN_HOUSE_PRICE,
        "benchmark_symbol": SPAIN_HOUSE_PRICE_SYMBOL,
        "benchmark_name": SPAIN_HOUSE_PRICE_NAME,
        "description": "Sirve para constructoras, promotoras e inmobiliarias.",
    },
    "spain_electricity_demand": {
        "key": "spain_electricity_demand",
        "label": SPAIN_ELECTRICITY_DEMAND_NAME,
        "reference_profile": EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND,
        "benchmark_symbol": SPAIN_ELECTRICITY_DEMAND_SYMBOL,
        "benchmark_name": SPAIN_ELECTRICITY_DEMAND_NAME,
        "description": "Muy util para electricas, redes y negocios sensibles al pulso de demanda.",
    },
    "spain_gas_consumption": {
        "key": "spain_gas_consumption",
        "label": SPAIN_GAS_CONSUMPTION_NAME,
        "reference_profile": EquityPosition.ReferenceProfile.SPAIN_GAS_CONSUMPTION,
        "benchmark_symbol": SPAIN_GAS_CONSUMPTION_SYMBOL,
        "benchmark_name": SPAIN_GAS_CONSUMPTION_NAME,
        "description": "Ayuda a leer gasistas, utilities y negocios expuestos al consumo de gas.",
    },
    "brent": {
        "key": "brent",
        "label": BRENT_REFERENCE_NAME,
        "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
        "benchmark_symbol": BRENT_REFERENCE_SYMBOL,
        "benchmark_name": BRENT_REFERENCE_NAME,
        "description": "Referencia util para energia, transporte y costes industriales.",
    },
    "copper": {
        "key": "copper",
        "label": COPPER_REFERENCE_NAME,
        "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
        "benchmark_symbol": COPPER_REFERENCE_SYMBOL,
        "benchmark_name": COPPER_REFERENCE_NAME,
        "description": "Buen termometro para industria, materiales y ciclo manufacturero.",
    },
    "eurusd": {
        "key": "eurusd",
        "label": EURUSD_REFERENCE_NAME,
        "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
        "benchmark_symbol": EURUSD_REFERENCE_SYMBOL,
        "benchmark_name": EURUSD_REFERENCE_NAME,
        "description": "Ayuda a leer negocios globales con ventas o compras fuera del euro.",
    },
}
IBEX_COMPANY_PROFILES = [
    {
        "company_name": "Acerinox, S.A.",
        "ticker": "ACX",
        "quote_symbol": "ACX.MC",
        "sector_label": "Materiales e industria",
        "aliases": ["ACERINOX"],
        "default_reference_key": "copper",
        "reference_keys": ["copper", "eurusd", "ibex_35"],
    },
    {
        "company_name": "ACS, Actividades de Construccion y Servicios, S.A.",
        "ticker": "ACS",
        "quote_symbol": "ACS.MC",
        "sector_label": "Construccion e infraestructuras",
        "aliases": ["ACS", "ACTIVIDADES DE CONSTRUCCION Y SERVICIOS"],
        "default_reference_key": "spain_house_price",
        "reference_keys": ["spain_house_price", "euribor_12m", "ibex_35"],
    },
    {
        "company_name": "Acciona, S.A.",
        "ticker": "ANA",
        "quote_symbol": "ANA.MC",
        "sector_label": "Infraestructuras y energia",
        "aliases": ["ACCIONA"],
        "default_reference_key": "spain_house_price",
        "reference_keys": ["spain_house_price", "ibex_35", "brent"],
    },
    {
        "company_name": "Acciona Energia, S.A.",
        "ticker": "ANE",
        "quote_symbol": "ANE.MC",
        "sector_label": "Energia renovable",
        "aliases": ["ACCIONA ENERGIA"],
        "default_reference_key": "spain_electricity_demand",
        "reference_keys": ["spain_electricity_demand", "ibex_35", "eurusd", "brent"],
    },
    {
        "company_name": "Aena SME, S.A.",
        "ticker": "AENA",
        "quote_symbol": "AENA.MC",
        "sector_label": "Aeropuertos y movilidad",
        "aliases": ["AENA"],
        "default_reference_key": "brent",
        "reference_keys": ["brent", "eurusd", "ibex_35"],
    },
    {
        "company_name": "Amadeus IT Group, S.A.",
        "ticker": "AMS",
        "quote_symbol": "AMS.MC",
        "sector_label": "Tecnologia de viajes",
        "aliases": ["AMADEUS", "AMADEUS IT GROUP"],
        "default_reference_key": "eurusd",
        "reference_keys": ["eurusd", "brent", "ibex_35"],
    },
    {
        "company_name": "ArcelorMittal, S.A.",
        "ticker": "MTS",
        "quote_symbol": "MTS.MC",
        "sector_label": "Acero e industria",
        "aliases": ["ARCELORMITTAL", "ARCELOR MITTAL"],
        "default_reference_key": "copper",
        "reference_keys": ["copper", "eurusd", "ibex_35"],
    },
    {
        "company_name": "Banco Bilbao Vizcaya Argentaria, S.A.",
        "ticker": "BBVA",
        "quote_symbol": "BBVA.MC",
        "sector_label": "Banca",
        "aliases": ["BBVA", "BANCO BILBAO VIZCAYA ARGENTARIA"],
        "default_reference_key": "euribor_12m",
        "reference_keys": ["euribor_12m", "ibex_35", "eurusd"],
    },
    {
        "company_name": "Banco de Sabadell, S.A.",
        "ticker": "SAB",
        "quote_symbol": "SAB.MC",
        "sector_label": "Banca",
        "aliases": ["SABADELL", "BANCO SABADELL", "BANCO DE SABADELL"],
        "default_reference_key": "euribor_12m",
        "reference_keys": ["euribor_12m", "ibex_35"],
    },
    {
        "company_name": "Banco Santander, S.A.",
        "ticker": "SAN",
        "quote_symbol": "SAN.MC",
        "sector_label": "Banca",
        "aliases": ["SANTANDER", "BANCO SANTANDER"],
        "default_reference_key": "euribor_12m",
        "reference_keys": ["euribor_12m", "ibex_35", "eurusd"],
    },
    {
        "company_name": "Bankinter, S.A.",
        "ticker": "BKT",
        "quote_symbol": "BKT.MC",
        "sector_label": "Banca",
        "aliases": ["BANKINTER"],
        "default_reference_key": "euribor_12m",
        "reference_keys": ["euribor_12m", "ibex_35"],
    },
    {
        "company_name": "CaixaBank, S.A.",
        "ticker": "CABK",
        "quote_symbol": "CABK.MC",
        "sector_label": "Banca",
        "aliases": ["CAIXABANK", "CAIXA BANK"],
        "default_reference_key": "euribor_12m",
        "reference_keys": ["euribor_12m", "ibex_35"],
    },
    {
        "company_name": "Cellnex Telecom, S.A.",
        "ticker": "CLNX",
        "quote_symbol": "CLNX.MC",
        "sector_label": "Telecomunicaciones",
        "aliases": ["CELLNEX", "CELLNEX TELECOM"],
        "default_reference_key": "ibex_35",
        "reference_keys": ["ibex_35", "euribor_12m", "eurusd"],
    },
    {
        "company_name": "Enagas, S.A.",
        "ticker": "ENG",
        "quote_symbol": "ENG.MC",
        "sector_label": "Infraestructura energetica",
        "aliases": ["ENAGAS"],
        "default_reference_key": "spain_gas_consumption",
        "reference_keys": ["spain_gas_consumption", "brent", "ibex_35", "euribor_12m"],
    },
    {
        "company_name": "Endesa, S.A.",
        "ticker": "ELE",
        "quote_symbol": "ELE.MC",
        "sector_label": "Electrica",
        "aliases": ["ENDESA"],
        "default_reference_key": "spain_electricity_demand",
        "reference_keys": ["spain_electricity_demand", "ibex_35", "brent", "euribor_12m"],
    },
    {
        "company_name": "Ferrovial SE",
        "ticker": "FER",
        "quote_symbol": "FER.MC",
        "sector_label": "Infraestructuras y concesiones",
        "aliases": ["FERROVIAL"],
        "default_reference_key": "spain_house_price",
        "reference_keys": ["spain_house_price", "euribor_12m", "ibex_35"],
    },
    {
        "company_name": "Fluidra, S.A.",
        "ticker": "FDR",
        "quote_symbol": "FDR.MC",
        "sector_label": "Equipamiento y consumo ligado al hogar",
        "aliases": ["FLUIDRA"],
        "default_reference_key": "spain_house_price",
        "reference_keys": ["spain_house_price", "ibex_35", "eurusd"],
    },
    {
        "company_name": "Grifols, S.A.",
        "ticker": "GRF",
        "quote_symbol": "GRF.MC",
        "sector_label": "Salud",
        "aliases": ["GRIFOLS"],
        "default_reference_key": "eurusd",
        "reference_keys": ["eurusd", "ibex_35"],
    },
    {
        "company_name": "Iberdrola, S.A.",
        "ticker": "IBE",
        "quote_symbol": "IBE.MC",
        "sector_label": "Electrica",
        "aliases": ["IBERDROLA"],
        "default_reference_key": "spain_electricity_demand",
        "reference_keys": ["spain_electricity_demand", "ibex_35", "brent", "euribor_12m"],
    },
    {
        "company_name": "Inditex, S.A.",
        "ticker": "ITX",
        "quote_symbol": "ITX.MC",
        "sector_label": "Consumo global",
        "aliases": ["INDITEX", "INDUSTRIA DE DISENO TEXTIL", "ZARA"],
        "default_reference_key": "eurusd",
        "reference_keys": ["eurusd", "ibex_35"],
    },
    {
        "company_name": "Indra Sistemas, S.A.",
        "ticker": "IDR",
        "quote_symbol": "IDR.MC",
        "sector_label": "Tecnologia y defensa",
        "aliases": ["INDRA", "INDRA SISTEMAS"],
        "default_reference_key": "ibex_35",
        "reference_keys": ["ibex_35", "eurusd"],
    },
    {
        "company_name": "International Consolidated Airlines Group, S.A.",
        "ticker": "IAG",
        "quote_symbol": "IAG.MC",
        "sector_label": "Aerolineas",
        "aliases": ["IAG", "INTERNATIONAL AIRLINES GROUP"],
        "default_reference_key": "brent",
        "reference_keys": ["brent", "eurusd", "ibex_35"],
    },
    {
        "company_name": "Logista Integral, S.A.",
        "ticker": "LOG",
        "quote_symbol": "LOG.MC",
        "sector_label": "Logistica y distribucion",
        "aliases": ["LOGISTA", "LOGISTA INTEGRAL"],
        "default_reference_key": "brent",
        "reference_keys": ["brent", "ibex_35", "eurusd"],
    },
    {
        "company_name": "Mapfre, S.A.",
        "ticker": "MAP",
        "quote_symbol": "MAP.MC",
        "sector_label": "Seguros",
        "aliases": ["MAPFRE"],
        "default_reference_key": "euribor_12m",
        "reference_keys": ["euribor_12m", "ibex_35"],
    },
    {
        "company_name": "Merlin Properties SOCIMI, S.A.",
        "ticker": "MRL",
        "quote_symbol": "MRL.MC",
        "sector_label": "Inmobiliario",
        "aliases": ["MERLIN", "MERLIN PROPERTIES"],
        "default_reference_key": "spain_house_price",
        "reference_keys": ["spain_house_price", "euribor_12m", "ibex_35"],
    },
    {
        "company_name": "Naturgy Energy Group, S.A.",
        "ticker": "NTGY",
        "quote_symbol": "NTGY.MC",
        "sector_label": "Gas y energia",
        "aliases": ["NATURGY"],
        "default_reference_key": "spain_gas_consumption",
        "reference_keys": ["spain_gas_consumption", "spain_electricity_demand", "brent", "ibex_35", "euribor_12m"],
    },
    {
        "company_name": "Redeia Corporacion, S.A.",
        "ticker": "RED",
        "quote_symbol": "RED.MC",
        "sector_label": "Redes electricas",
        "aliases": ["REDEIA", "RED ELECTRICA", "RED ELECTRICA CORPORACION"],
        "default_reference_key": "spain_electricity_demand",
        "reference_keys": ["spain_electricity_demand", "ibex_35", "euribor_12m"],
    },
    {
        "company_name": "Repsol, S.A.",
        "ticker": "REP",
        "quote_symbol": "REP.MC",
        "sector_label": "Energia",
        "aliases": ["REPSOL"],
        "default_reference_key": "brent",
        "reference_keys": ["brent", "spain_gas_consumption", "eurusd", "ibex_35"],
    },
    {
        "company_name": "Sacyr, S.A.",
        "ticker": "SCYR",
        "quote_symbol": "SCYR.MC",
        "sector_label": "Construccion e infraestructuras",
        "aliases": ["SACYR"],
        "default_reference_key": "spain_house_price",
        "reference_keys": ["spain_house_price", "euribor_12m", "ibex_35"],
    },
    {
        "company_name": "Solaria Energia y Medio Ambiente, S.A.",
        "ticker": "SLR",
        "quote_symbol": "SLR.MC",
        "sector_label": "Solar",
        "aliases": ["SOLARIA", "SOLARIA ENERGIA"],
        "default_reference_key": "spain_electricity_demand",
        "reference_keys": ["spain_electricity_demand", "ibex_35", "eurusd", "euribor_12m"],
    },
    {
        "company_name": "Telefonica, S.A.",
        "ticker": "TEF",
        "quote_symbol": "TEF.MC",
        "sector_label": "Telecomunicaciones",
        "aliases": ["TELEFONICA", "TELEFONICA SA"],
        "default_reference_key": "eurusd",
        "reference_keys": ["eurusd", "ibex_35", "euribor_12m"],
    },
    {
        "company_name": "Unicaja Banco, S.A.",
        "ticker": "UNI",
        "quote_symbol": "UNI.MC",
        "sector_label": "Banca",
        "aliases": ["UNICAJA", "UNICAJA BANCO"],
        "default_reference_key": "euribor_12m",
        "reference_keys": ["euribor_12m", "ibex_35"],
    },
]

IBEX_REFERENCE_WORKBOOK_FILENAME = "IBEX35_Indicadores_Referencia.xlsx"
WORKBOOK_CORRELATION_INDICATOR_ALIASES = {
    "Euribor 12m": "Euribor 12m (%)",
    "Precio electricidad": "Precio electricidad OMIE (EUR/MWh)",
    "Gasto defensa": "Gasto defensa Espana (% PIB)",
    "IPV vivienda": "Indice Precio Vivienda INE (2015=100)",
    "Trafico aereo": "Trafico aereo AENA (M pax)",
    "Brent": "Precio Brent (USD/barril prom.)",
    "Licitacion obra": "Licitacion obra publica (MrdEUR)",
    "Consumo privado": "Consumo privado Espana (var. %)",
}
WORKBOOK_LIVE_REFERENCE_ALIASES = {
    "Euribor 12m (%)": "euribor_12m",
    "IBEX 35 (puntos)": "ibex_35",
    "Consumo electrico Espana (TWh)": "spain_electricity_demand",
    "Indice Precio Vivienda INE (2015=100)": "spain_house_price",
    "Precio Brent (USD/barril prom.)": "brent",
}
WORKBOOK_BASELINE_INDICATORS = [
    "IBEX 35 (puntos)",
]


class MarketDataError(Exception):
    pass


class EquityDocumentImportError(Exception):
    pass


@dataclass
class MarketSeries:
    symbol: str
    name: str
    latest_price: Decimal
    latest_date: date
    points: list[dict]


def clone_market_series(series: MarketSeries | None) -> MarketSeries | None:
    if series is None:
        return None
    return MarketSeries(
        symbol=series.symbol,
        name=series.name,
        latest_price=series.latest_price,
        latest_date=series.latest_date,
        points=[dict(point) for point in series.points],
    )


def build_market_data_cache_bucket(now: datetime | None = None) -> int:
    cache_minutes = max(
        getattr(settings, "EQUITIES_MARKET_DATA_CACHE_MINUTES", DEFAULT_MARKET_DATA_CACHE_MINUTES) or DEFAULT_MARKET_DATA_CACHE_MINUTES,
        1,
    )
    current = now or django_timezone.now()
    return int(current.timestamp() // (cache_minutes * 60))


def market_request_timeout_seconds() -> int:
    return max(
        getattr(settings, "EQUITIES_MARKET_REQUEST_TIMEOUT_SECONDS", DEFAULT_MARKET_REQUEST_TIMEOUT_SECONDS)
        or DEFAULT_MARKET_REQUEST_TIMEOUT_SECONDS,
        3,
    )


def ibex_universe_max_workers() -> int:
    return max(
        getattr(settings, "EQUITIES_IBEX_UNIVERSE_MAX_WORKERS", DEFAULT_IBEX_UNIVERSE_MAX_WORKERS)
        or DEFAULT_IBEX_UNIVERSE_MAX_WORKERS,
        1,
    )


def clear_market_data_caches() -> None:
    _fetch_market_series_cached.cache_clear()
    _fetch_reference_series_for_choice_cached.cache_clear()
    _fetch_equity_fundamentals_cached.cache_clear()
    _fetch_ecb_yield_curve_series_cached.cache_clear()


@dataclass
class EquityDocumentPrefill:
    data: dict
    detected_fields: list[str]
    candidate_count: int
    source_kind: str


@dataclass(frozen=True)
class MarketHistoryPoint:
    price_date: date
    close_price: Decimal


def normalize_company_lookup(value: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^A-Z0-9]+", " ", ascii_text.upper())
    return re.sub(r"\s+", " ", cleaned).strip()


def build_security_lookup_keys(ticker: str = "", company_name: str = "", quote_symbol: str = "") -> set[str]:
    return {
        key
        for key in {
            normalize_company_lookup(ticker),
            normalize_company_lookup(company_name),
            normalize_company_lookup(quote_symbol),
        }
        if key
    }


def resolve_equity_sector_label(company_name: str = "", ticker: str = "", quote_symbol: str = "") -> str:
    profile = (
        find_equity_company_profile(company_name)
        or find_equity_company_profile(ticker)
        or find_equity_company_profile(quote_symbol)
    )
    return profile.get("sector_label", "") if profile else ""


def get_reference_preset(reference_key: str) -> dict:
    preset = REFERENCE_PRESETS.get(reference_key)
    return dict(preset) if preset else {}


def get_equity_company_catalog() -> list[dict]:
    catalog = []
    for profile in IBEX_COMPANY_PROFILES:
        default_reference = get_reference_preset(profile["default_reference_key"])
        reference_suggestions = [get_reference_preset(reference_key) for reference_key in profile["reference_keys"]]
        lookup_keys = {
            normalize_company_lookup(profile["company_name"]),
            normalize_company_lookup(profile["ticker"]),
            normalize_company_lookup(profile["quote_symbol"]),
        }
        lookup_keys.update(normalize_company_lookup(alias) for alias in profile.get("aliases", []))
        catalog.append(
            {
                "company_name": profile["company_name"],
                "ticker": profile["ticker"],
                "quote_symbol": profile["quote_symbol"],
                "sector_label": profile["sector_label"],
                "default_reference": default_reference,
                "reference_suggestions": reference_suggestions,
                "lookup_keys": sorted(key for key in lookup_keys if key),
            }
        )
    return catalog


def get_equity_optimizer_sector_choices() -> list[tuple[str, str]]:
    sectors = sorted(
        {
            str(entry.get("sector_label") or "").strip()
            for entry in get_equity_company_catalog()
            if str(entry.get("sector_label") or "").strip()
        }
    )
    return [(sector, sector) for sector in sectors]


def find_equity_company_profile(query: str) -> dict | None:
    normalized_query = normalize_company_lookup(query)
    if not normalized_query:
        return None

    catalog = get_equity_company_catalog()
    for entry in catalog:
        if normalized_query in entry["lookup_keys"]:
            return entry

    partial_matches = []
    for entry in catalog:
        best_key = None
        for lookup_key in entry["lookup_keys"]:
            if normalized_query and (lookup_key.startswith(normalized_query) or normalized_query in lookup_key):
                best_key = lookup_key
                break
        if best_key:
            partial_matches.append((len(best_key), entry["company_name"], entry))
    if partial_matches:
        partial_matches.sort(key=lambda item: (item[0], item[1]))
        return partial_matches[0][2]
    return None


def build_reference_suggestions_for_equity(company_name: str = "", ticker: str = "") -> list[dict]:
    profile = find_equity_company_profile(company_name) or find_equity_company_profile(ticker)
    if profile:
        return [dict(item) for item in profile["reference_suggestions"]]
    return [get_reference_preset("ibex_35")]


def apply_equity_company_defaults(data: dict, override_generic_reference: bool = True) -> dict:
    result = dict(data)
    profile = (
        find_equity_company_profile(result.get("company_name"))
        or find_equity_company_profile(result.get("ticker"))
        or find_equity_company_profile(result.get("quote_symbol"))
    )
    if not profile:
        return result

    result["company_name"] = result.get("company_name") or profile["company_name"]
    result["ticker"] = result.get("ticker") or profile["ticker"]
    result["quote_symbol"] = result.get("quote_symbol") or profile["quote_symbol"]

    generic_reference = (
        result.get("reference_profile") in {None, "", EquityPosition.ReferenceProfile.MARKET_INDEX}
        and (result.get("benchmark_symbol") in {None, "", DEFAULT_BENCHMARK_SYMBOL})
        and (result.get("benchmark_name") in {None, "", DEFAULT_BENCHMARK_NAME})
    )
    if generic_reference and override_generic_reference:
        default_reference = profile["default_reference"]
        result["reference_profile"] = default_reference["reference_profile"]
        result["benchmark_symbol"] = default_reference["benchmark_symbol"]
        result["benchmark_name"] = default_reference["benchmark_name"]

    return result


def workbook_text(value) -> str:
    return str(value or "").strip()


def workbook_decimal(value) -> Decimal | None:
    if value in {None, ""}:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = workbook_text(value)
    if not text or text in {"-", "—"}:
        return None
    text = (
        text.replace("\xa0", "")
        .replace("%", "")
        .replace("EUR", "")
        .replace("Mrd", "")
        .replace("€", "")
        .replace(",", ".")
    )
    try:
        return Decimal(text)
    except Exception:
        return None


def describe_reference_score(score: Decimal | None) -> str:
    if score is None:
        return "Sin referencia"
    if score >= Decimal("0.70"):
        return "Alta"
    if score >= Decimal("0.50"):
        return "Buena"
    if score >= Decimal("0.35"):
        return "Media"
    return "Baja"


def resolve_equities_reference_workbook_path() -> Path | None:
    configured_path = getattr(settings, "EQUITIES_REFERENCE_WORKBOOK", "")
    candidates = []
    if configured_path:
        candidates.append(Path(configured_path).expanduser())
    candidates.append(Path(settings.BASE_DIR) / IBEX_REFERENCE_WORKBOOK_FILENAME)
    candidates.append(Path(settings.BASE_DIR).parent / IBEX_REFERENCE_WORKBOOK_FILENAME)
    candidates.append(Path.cwd() / IBEX_REFERENCE_WORKBOOK_FILENAME)
    candidates.append(Path.home() / "Downloads" / IBEX_REFERENCE_WORKBOOK_FILENAME)

    for candidate in candidates:
        try:
            resolved = candidate.expanduser()
        except Exception:
            continue
        if resolved.exists():
            return resolved
    return None


def find_workbook_sheet(workbook, keyword: str):
    normalized_keyword = normalize_company_lookup(keyword)
    for worksheet in workbook.worksheets:
        if normalized_keyword in normalize_company_lookup(worksheet.title):
            return worksheet
    return None


def find_workbook_header_index(rows: list[list], first_header: str) -> int | None:
    expected = normalize_company_lookup(first_header)
    for index, row in enumerate(rows):
        if not row:
            continue
        current = normalize_company_lookup(row[0])
        if current == expected or current.startswith(expected):
            return index
    return None


def extract_year_values(row: list, start_index: int = 3) -> dict[int, Decimal]:
    values = {}
    for offset, year in enumerate(range(2019, 2026), start=start_index):
        if len(row) <= offset:
            continue
        numeric_value = workbook_decimal(row[offset])
        if numeric_value is not None:
            values[year] = numeric_value
    return values


def resolve_workbook_live_reference(indicator_name: str) -> dict | None:
    normalized_name = normalize_company_lookup(indicator_name)
    for alias, reference_key in WORKBOOK_LIVE_REFERENCE_ALIASES.items():
        if normalize_company_lookup(alias) == normalized_name:
            preset = get_reference_preset(reference_key)
            return preset or None
    return None


def match_workbook_indicator_name(raw_label: str, workbook_snapshot: dict) -> str | None:
    normalized_label = normalize_company_lookup(raw_label)
    if not normalized_label:
        return None

    candidates = []
    for short_name, full_name in workbook_snapshot.get("indicator_name_by_short", {}).items():
        for candidate_name in (short_name, full_name):
            normalized_candidate = normalize_company_lookup(candidate_name)
            if not normalized_candidate:
                continue
            if normalized_candidate in normalized_label or normalized_label in normalized_candidate:
                candidates.append((len(normalized_candidate), full_name))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def calculate_series_change_correlation(primary_series: dict[int, Decimal], secondary_series: dict[int, Decimal]) -> tuple[Decimal | None, int]:
    shared_years = sorted(set(primary_series.keys()) & set(secondary_series.keys()))
    if len(shared_years) < 4:
        return None, 0

    primary_changes = []
    secondary_changes = []
    for previous_year, current_year in zip(shared_years, shared_years[1:]):
        primary_previous = primary_series.get(previous_year)
        primary_current = primary_series.get(current_year)
        secondary_previous = secondary_series.get(previous_year)
        secondary_current = secondary_series.get(current_year)
        if None in {primary_previous, primary_current, secondary_previous, secondary_current}:
            continue
        primary_changes.append(primary_current - primary_previous)
        secondary_changes.append(secondary_current - secondary_previous)

    if len(primary_changes) < 3:
        return None, len(primary_changes)
    return pearson_correlation(primary_changes, secondary_changes), len(primary_changes)


@lru_cache(maxsize=1)
def load_ibex_reference_workbook_snapshot() -> dict:
    workbook_path = resolve_equities_reference_workbook_path()
    if workbook_path is None:
        return {
            "available": False,
            "path": "",
            "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "available": False,
            "path": str(workbook_path),
            "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        summary_sheet = find_workbook_sheet(workbook, "Resumen")
        quotes_sheet = find_workbook_sheet(workbook, "Cotizaciones")
        indicators_sheet = find_workbook_sheet(workbook, "Indicadores")
        correlations_sheet = find_workbook_sheet(workbook, "Correlaciones")

        if not summary_sheet or not quotes_sheet or not indicators_sheet or not correlations_sheet:
            return {
                "available": False,
                "path": str(workbook_path),
                "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        summary_rows = [list(row) for row in summary_sheet.iter_rows(values_only=True)]
        quotes_rows = [list(row) for row in quotes_sheet.iter_rows(values_only=True)]
        indicators_rows = [list(row) for row in indicators_sheet.iter_rows(values_only=True)]
        correlations_rows = [list(row) for row in correlations_sheet.iter_rows(values_only=True)]

        summary_header_index = find_workbook_header_index(summary_rows, "Ticker")
        quotes_header_index = find_workbook_header_index(quotes_rows, "Ticker")
        indicators_header_index = find_workbook_header_index(indicators_rows, "Indicador")
        correlations_header_index = find_workbook_header_index(correlations_rows, "Sector")
        if None in {summary_header_index, quotes_header_index, indicators_header_index, correlations_header_index}:
            return {
                "available": False,
                "path": str(workbook_path),
                "companies": [],
            "companies_by_key": {},
            "indicators_by_name": {},
            "indicators_by_key": {},
            "indicator_name_by_short": {},
            "sector_map": {},
        }

        companies = []
        companies_by_key = {}
        for row in summary_rows[summary_header_index + 1 :]:
            ticker = workbook_text(row[0] if len(row) > 0 else "")
            company_name = workbook_text(row[1] if len(row) > 1 else "")
            if not ticker or not company_name:
                continue
            company = {
                "ticker": ticker.upper(),
                "company_name": company_name,
                "sector": workbook_text(row[2] if len(row) > 2 else ""),
                "return_2025": (
                    workbook_decimal(row[3] if len(row) > 3 else None) * Decimal("100")
                    if workbook_decimal(row[3] if len(row) > 3 else None) is not None
                    else None
                ),
                "primary_reference": workbook_text(row[4] if len(row) > 4 else ""),
                "reference_source": workbook_text(row[5] if len(row) > 5 else ""),
                "primary_correlation": workbook_decimal(row[6] if len(row) > 6 else None),
                "per_2025": workbook_decimal(row[7] if len(row) > 7 else None),
                "dividend_yield": (
                    workbook_decimal(row[8] if len(row) > 8 else None) * Decimal("100")
                    if workbook_decimal(row[8] if len(row) > 8 else None) is not None
                    else None
                ),
                "market_cap_billion": workbook_decimal(row[9] if len(row) > 9 else None),
                "notes": workbook_text(row[10] if len(row) > 10 else ""),
                "quote_symbol": workbook_text(row[11] if len(row) > 11 else ""),
                "price_history": {},
            }
            companies.append(company)
            for lookup_key in {
                normalize_company_lookup(company["ticker"]),
                normalize_company_lookup(company["company_name"]),
                normalize_company_lookup(company["quote_symbol"]),
            }:
                if lookup_key:
                    companies_by_key[lookup_key] = company

        quote_prices_by_key = {}
        for row in quotes_rows[quotes_header_index + 1 :]:
            ticker = workbook_text(row[0] if len(row) > 0 else "")
            company_name = workbook_text(row[1] if len(row) > 1 else "")
            quote_symbol = workbook_text(row[15] if len(row) > 15 else "")
            price_history = extract_year_values(row)
            for lookup_key in {
                normalize_company_lookup(ticker),
                normalize_company_lookup(company_name),
                normalize_company_lookup(quote_symbol),
            }:
                if lookup_key:
                    quote_prices_by_key[lookup_key] = price_history

        for company in companies:
            for lookup_key in (
                normalize_company_lookup(company["ticker"]),
                normalize_company_lookup(company["company_name"]),
                normalize_company_lookup(company["quote_symbol"]),
            ):
                if lookup_key in quote_prices_by_key:
                    company["price_history"] = quote_prices_by_key[lookup_key]
                    break

        indicators_by_name = {}
        indicators_by_key = {}
        for row in indicators_rows[indicators_header_index + 1 :]:
            indicator_name = workbook_text(row[0] if len(row) > 0 else "")
            if not indicator_name:
                continue
            indicator_entry = {
                "name": indicator_name,
                "source": workbook_text(row[1] if len(row) > 1 else ""),
                "sector_label": workbook_text(row[2] if len(row) > 2 else ""),
                "values": extract_year_values(row),
            }
            indicators_by_name[indicator_name] = indicator_entry
            indicators_by_key[normalize_company_lookup(indicator_name)] = indicator_entry

        correlation_headers = [
            workbook_text(value)
            for value in correlations_rows[correlations_header_index][1:]
            if workbook_text(value)
        ]
        indicator_name_by_short = {}
        for short_name in correlation_headers:
            full_name = WORKBOOK_CORRELATION_INDICATOR_ALIASES.get(short_name, short_name)
            indicator_name_by_short[short_name] = full_name

        sector_map = {}
        for row in correlations_rows[correlations_header_index + 1 :]:
            sector_label = workbook_text(row[0] if len(row) > 0 else "")
            if not sector_label or sector_label.upper().startswith("NOTAS"):
                continue
            scores = {}
            for column_index, short_name in enumerate(correlation_headers, start=1):
                score = workbook_decimal(row[column_index] if len(row) > column_index else None)
                if score is not None:
                    scores[short_name] = score
            if scores:
                sector_map[sector_label] = scores

        return {
            "available": True,
            "path": str(workbook_path),
            "companies": companies,
            "companies_by_key": companies_by_key,
            "indicators_by_name": indicators_by_name,
            "indicators_by_key": indicators_by_key,
            "indicator_name_by_short": indicator_name_by_short,
            "sector_map": sector_map,
        }
    finally:
        workbook.close()


def build_workbook_reference_candidates(company: dict, workbook_snapshot: dict, position: EquityPosition | None = None) -> list[dict]:
    sector_scores = workbook_snapshot.get("sector_map", {}).get(company.get("sector", ""), {})
    selected_reference_key = None
    if position is not None:
        selected_reference_key = (
            position.reference_profile,
            position.benchmark_symbol,
            position.benchmark_name,
        )

    primary_indicator_name = match_workbook_indicator_name(company.get("primary_reference", ""), workbook_snapshot)
    candidate_names = []
    for short_name in sector_scores.keys():
        full_name = workbook_snapshot.get("indicator_name_by_short", {}).get(short_name, short_name)
        candidate_names.append(full_name)
    if primary_indicator_name:
        candidate_names.append(primary_indicator_name)
    candidate_names.extend(WORKBOOK_BASELINE_INDICATORS)

    seen = set()
    candidates = []
    for indicator_name in candidate_names:
        normalized_indicator = normalize_company_lookup(indicator_name)
        if not normalized_indicator or normalized_indicator in seen:
            continue
        seen.add(normalized_indicator)

        short_name = next(
            (
                item_short_name
                for item_short_name, item_full_name in workbook_snapshot.get("indicator_name_by_short", {}).items()
                if normalize_company_lookup(item_full_name) == normalized_indicator
            ),
            indicator_name,
        )
        sector_score = sector_scores.get(short_name)
        if primary_indicator_name and normalize_company_lookup(primary_indicator_name) == normalized_indicator:
            if company.get("primary_correlation") is not None:
                sector_score = max(sector_score or ZERO, company["primary_correlation"])

        indicator_entry = workbook_snapshot.get("indicators_by_key", {}).get(normalized_indicator)
        historical_coefficient = None
        observations_count = 0
        if company.get("price_history") and indicator_entry:
            historical_coefficient, observations_count = calculate_series_change_correlation(
                company["price_history"],
                indicator_entry.get("values", {}),
            )

        live_reference = resolve_workbook_live_reference(indicator_name)
        is_active_reference = False
        if selected_reference_key and live_reference:
            candidate_key = (
                live_reference["reference_profile"],
                live_reference["benchmark_symbol"],
                live_reference["benchmark_name"],
            )
            is_active_reference = candidate_key == selected_reference_key

        sector_component = sector_score or ZERO
        history_component = abs(historical_coefficient) if historical_coefficient is not None else ZERO
        composite_score = (sector_component * Decimal("0.65")) + (history_component * Decimal("0.35"))
        if normalize_company_lookup(primary_indicator_name or "") == normalized_indicator:
            composite_score += Decimal("0.05")

        candidates.append(
            {
                "name": indicator_name,
                "short_name": short_name,
                "sector_score": sector_score,
                "sector_score_label": describe_reference_score(sector_score),
                "historical_coefficient": historical_coefficient,
                "historical_label": describe_correlation(historical_coefficient),
                "observations_count": observations_count,
                "live_reference": live_reference,
                "supports_chart": bool(live_reference),
                "is_active_reference": is_active_reference,
                "is_primary_reference": normalize_company_lookup(primary_indicator_name or "") == normalized_indicator,
                "composite_score": composite_score,
            }
        )

    candidates.sort(
        key=lambda item: (
            item["composite_score"],
            item["sector_score"] or ZERO,
            abs(item["historical_coefficient"]) if item["historical_coefficient"] is not None else ZERO,
            item["name"],
        ),
        reverse=True,
    )

    best_marked = False
    for candidate in candidates:
        candidate["is_best"] = not best_marked
        if not best_marked:
            best_marked = True
    return candidates


def build_reference_playbook_from_card(card: dict) -> dict:
    candidates = []
    for suggestion in card.get("suggested_references", []):
        live_reference = {
            "reference_profile": suggestion["reference_profile"],
            "benchmark_symbol": suggestion["benchmark_symbol"],
            "benchmark_name": suggestion["benchmark_name"],
        }
        historical_coefficient = suggestion["correlation"].get("coefficient")
        candidates.append(
            {
                "name": suggestion["benchmark_name"],
                "short_name": suggestion["benchmark_name"],
                "sector_score": None,
                "sector_score_label": "Sin guia sectorial",
                "historical_coefficient": historical_coefficient,
                "historical_label": suggestion["correlation"].get("label") or describe_correlation(historical_coefficient),
                "observations_count": suggestion["correlation"].get("observations_count", 0),
                "live_reference": live_reference,
                "supports_chart": True,
                "is_active_reference": suggestion.get("is_selected", False),
                "is_primary_reference": suggestion.get("is_best", False),
                "is_best": suggestion.get("is_best", False),
                "composite_score": abs(historical_coefficient) if historical_coefficient is not None else ZERO,
            }
        )

    return {
        "available": bool(candidates),
        "source_label": "Historico de la app",
        "company_sector": "",
        "current_reference_label": card.get("reference_label"),
        "current_candidate": next((candidate for candidate in candidates if candidate["is_active_reference"]), None),
        "best_candidate": next((candidate for candidate in candidates if candidate["is_best"]), None),
        "candidates": candidates[:4],
        "return_2025": None,
        "per_2025": None,
        "dividend_yield": None,
        "notes": card["position"].notes,
    }


def build_workbook_reference_playbook(company: dict, workbook_snapshot: dict, card: dict | None = None) -> dict:
    position = card["position"] if card else None
    ranked_candidates = build_workbook_reference_candidates(company, workbook_snapshot, position=position)
    current_candidate = next((candidate for candidate in ranked_candidates if candidate["is_active_reference"]), None)
    candidates = ranked_candidates[:4]
    if current_candidate and current_candidate not in candidates:
        candidates = [current_candidate, *candidates[:3]]
    current_candidate = next((candidate for candidate in candidates if candidate["is_active_reference"]), None)
    if current_candidate is None and card is None:
        current_candidate = next((candidate for candidate in candidates if candidate["is_primary_reference"]), None)

    return {
        "available": bool(candidates),
        "source_label": "Excel IBEX 2019-2025",
        "company_sector": company.get("sector", ""),
        "current_reference_label": card.get("reference_label") if card else company.get("primary_reference"),
        "current_candidate": current_candidate,
        "best_candidate": next((candidate for candidate in candidates if candidate["is_best"]), None),
        "candidates": candidates,
        "return_2025": company.get("return_2025"),
        "per_2025": company.get("per_2025"),
        "dividend_yield": company.get("dividend_yield"),
        "notes": company.get("notes", ""),
    }


def find_history_card_for_workbook_company(company: dict, history_cards: list[dict]) -> dict | None:
    lookup_keys = build_security_lookup_keys(
        ticker=company.get("ticker", ""),
        company_name=company.get("company_name", ""),
        quote_symbol=company.get("quote_symbol", ""),
    )
    for card in history_cards:
        position = card["position"]
        position_keys = build_security_lookup_keys(
            ticker=position.ticker,
            company_name=position.company_name,
            quote_symbol=position.quote_symbol,
        )
        if lookup_keys & position_keys:
            return card
    return None


def build_reference_guide_row(playbook: dict, card: dict | None = None, company: dict | None = None) -> dict:
    current_candidate = playbook.get("current_candidate")
    position = card["position"] if card else None
    status_key = "guide"
    status_label = "Solo guia"
    detail_anchor = ""
    current_live_correlation = None
    one_year_return = None

    if position is not None:
        status_key = "owned" if position.is_owned else "watchlist"
        status_label = position.get_position_kind_display()
        detail_anchor = f"stock-{position.id}"
        current_live_correlation = card.get("correlation", {}).get("coefficient")
        one_year_snapshot = next((snapshot for snapshot in card.get("period_snapshots", []) if snapshot.get("label") == "1Y"), None)
        if one_year_snapshot and one_year_snapshot.get("available"):
            one_year_return = one_year_snapshot.get("stock_return_pct")

    ticker = company.get("ticker") if company else position.ticker
    company_name = company.get("company_name") if company else position.company_name
    sector = playbook.get("company_sector") or (company.get("sector") if company else "")

    return {
        "ticker": ticker,
        "company_name": company_name,
        "sector": sector,
        "status_key": status_key,
        "status_label": status_label,
        "is_tracked": bool(position),
        "detail_anchor": detail_anchor,
        "current_reference_label": playbook.get("current_reference_label"),
        "current_candidate": current_candidate,
        "current_live_correlation": current_live_correlation,
        "best_candidate": playbook.get("best_candidate"),
        "top_candidates": playbook.get("candidates", [])[:3],
        "return_2025": playbook.get("return_2025"),
        "per_2025": playbook.get("per_2025"),
        "dividend_yield": playbook.get("dividend_yield"),
        "notes": playbook.get("notes"),
        "one_year_return": one_year_return,
        "source_label": playbook.get("source_label"),
    }


def find_workbook_company_for_position(position: EquityPosition) -> dict | None:
    workbook_snapshot = load_ibex_reference_workbook_snapshot()
    if not workbook_snapshot.get("available"):
        return None
    companies_by_key = workbook_snapshot.get("companies_by_key") or {}
    lookup_keys = build_security_lookup_keys(
        ticker=position.ticker,
        company_name=position.company_name,
        quote_symbol=position.quote_symbol,
    )
    for lookup_key in lookup_keys:
        company = companies_by_key.get(lookup_key)
        if company:
            return company
    return None


def resolve_per_valuation_thresholds(sector_label: str) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    normalized_sector = normalize_company_lookup(sector_label)
    if any(keyword in normalized_sector for keyword in ("BANCA", "SEGUROS")):
        return Decimal("7.00"), Decimal("10.00"), Decimal("13.00"), Decimal("16.00")
    if any(keyword in normalized_sector for keyword in ("ELECTRICA", "REDES", "INFRAESTRUCTURAS", "CONCESIONES", "TELECOMUNICACIONES", "GAS")):
        return Decimal("10.00"), Decimal("15.00"), Decimal("20.00"), Decimal("26.00")
    if any(keyword in normalized_sector for keyword in ("SOLAR", "RENOVABLE", "TECNOLOGIA", "VIAJES", "DEFENSA", "SALUD")):
        return Decimal("12.00"), Decimal("18.00"), Decimal("26.00"), Decimal("38.00")
    if any(keyword in normalized_sector for keyword in ("CONSUMO", "INMOBILIARIO")):
        return Decimal("10.00"), Decimal("16.00"), Decimal("23.00"), Decimal("32.00")
    return Decimal("8.00"), Decimal("13.00"), Decimal("20.00"), Decimal("28.00")


def build_equity_per_valuation(
    position: EquityPosition,
    fundamentals: dict | None = None,
    *,
    sector_label: str = "",
) -> dict:
    fundamentals = fundamentals or {}
    workbook_company = find_workbook_company_for_position(position)
    workbook_per = workbook_company.get("per_2025") if workbook_company else None
    derived_per = fundamentals.get("derived_per")
    latest_net_income_value = fundamentals.get("latest_net_income_value")

    per_value = workbook_per if workbook_per is not None else derived_per
    source_label = ""
    source_short_label = ""
    source_kind = ""
    if workbook_per is not None:
        source_label = "PER 2025e de la guia IBEX"
        source_short_label = "2025e"
        source_kind = "workbook"
    elif derived_per is not None:
        source_label = "PER derivado de capitalizacion y beneficio neto"
        source_short_label = "actual"
        source_kind = "fundamentals"

    lower_value, fair_value, demanding_value, stretched_value = resolve_per_valuation_thresholds(sector_label)
    if per_value is None:
        if latest_net_income_value is not None and latest_net_income_value <= ZERO:
            return {
                "available": True,
                "per_value": None,
                "source_label": "Sin PER util por beneficio neto negativo",
                "source_short_label": "sin PER",
                "source_kind": "negative_earnings",
                "label": "Sin PER util",
                "tone": "warn",
                "score": Decimal("-4.00"),
                "note": "El beneficio neto es negativo o nulo, asi que el PER no sirve para valorar bien esta accion ahora mismo.",
            }
        return {
            "available": False,
            "per_value": None,
            "source_label": "",
            "source_short_label": "",
            "source_kind": "",
            "label": "Sin PER",
            "tone": "neutral",
            "score": ZERO,
            "note": "No hay un PER fiable disponible para usarlo como filtro de valoracion.",
        }

    if per_value <= ZERO:
        label = "Sin PER util"
        tone = "warn"
        score = Decimal("-4.00")
        note = "El PER no es util porque el beneficio es demasiado debil o no comparable."
    elif per_value <= lower_value:
        label = "Barata"
        tone = "good"
        score = Decimal("6.00")
        note = "El PER se mueve en una zona baja para su sector y da apoyo a la valoracion."
    elif per_value <= fair_value:
        label = "Ajustada"
        tone = "good"
        score = Decimal("3.00")
        note = "El PER sigue razonable para su sector y no frena la tesis."
    elif per_value <= demanding_value:
        label = "Razonable"
        tone = "neutral"
        score = ZERO
        note = "El PER queda en una zona intermedia y no cambia mucho la lectura."
    elif per_value <= stretched_value:
        label = "Exigente"
        tone = "warn"
        score = Decimal("-4.00")
        note = "El PER ya exige bastante crecimiento y enfria parte de la valoracion."
    else:
        label = "Muy exigente"
        tone = "warn"
        score = Decimal("-7.00")
        note = "El PER esta muy estirado para su sector y resta margen de seguridad."

    if source_kind == "fundamentals":
        score = quantize_decimal(score * Decimal("0.85")) or ZERO
        note += " Esta lectura sale del ultimo beneficio neto disponible, no de una estimacion 2025e."

    return {
        "available": True,
        "per_value": quantize_decimal(per_value, "0.01"),
        "source_label": source_label,
        "source_short_label": source_short_label,
        "source_kind": source_kind,
        "label": label,
        "tone": tone,
        "score": quantize_decimal(score),
        "note": note,
        "sector_thresholds": {
            "cheap_limit": lower_value,
            "fair_limit": fair_value,
            "demanding_limit": demanding_value,
            "stretched_limit": stretched_value,
        },
    }


def apply_per_valuation_overlay(projection: dict, valuation: dict | None = None) -> dict:
    valuation = valuation or {}
    if not projection.get("available"):
        return projection

    valuation_score = valuation.get("score") or ZERO
    decision_score = projection.get("decision_score")
    if decision_score is not None:
        decision_score = quantize_decimal(decision_score + (valuation_score * Decimal("1.15")))
    explanation = str(projection.get("explanation") or "").strip()
    if valuation.get("available"):
        per_value = valuation.get("per_value")
        per_label = f"{per_value:.1f}" if per_value is not None else "sin PER"
        valuation_note = valuation.get("note") or ""
        if explanation:
            explanation += " "
        explanation += f"En valoracion, el PER {per_label} deja la accion {str(valuation.get('label') or '').lower()}. {valuation_note}".strip()
    return {
        **projection,
        "decision_score": decision_score,
        "valuation_score": quantize_decimal(valuation_score),
        "valuation_label": valuation.get("label") or "Sin PER",
        "valuation_tone": valuation.get("tone") or "neutral",
        "per_value": valuation.get("per_value"),
        "per_source_label": valuation.get("source_label") or "",
        "explanation": explanation,
    }


def build_equity_reference_guide(history_cards: list[dict]) -> dict:
    workbook_snapshot = load_ibex_reference_workbook_snapshot()
    rows = []
    matched_position_ids = set()

    if workbook_snapshot.get("available"):
        for company in workbook_snapshot.get("companies", []):
            card = find_history_card_for_workbook_company(company, history_cards)
            if card is not None:
                matched_position_ids.add(card["position"].id)
            playbook = build_workbook_reference_playbook(company, workbook_snapshot, card=card)
            if card is not None:
                card["reference_playbook"] = playbook
            rows.append(build_reference_guide_row(playbook, card=card, company=company))

    for card in history_cards:
        if card["position"].id in matched_position_ids or card.get("reference_playbook"):
            continue
        playbook = build_reference_playbook_from_card(card)
        card["reference_playbook"] = playbook
        rows.append(build_reference_guide_row(playbook, card=card))

    rows.sort(
        key=lambda item: (
            0 if item["status_key"] == "owned" else 1 if item["status_key"] == "watchlist" else 2,
            item["sector"],
            item["company_name"],
        )
    )
    tracked_rows = [row for row in rows if row["is_tracked"]]
    if not tracked_rows:
        tracked_rows = rows[:6]

    return {
        "rows": rows,
        "tracked_rows": tracked_rows,
        "summary": {
            "available": bool(rows),
            "workbook_loaded": workbook_snapshot.get("available", False),
            "source_label": Path(workbook_snapshot["path"]).name if workbook_snapshot.get("path") else "",
            "tracked_count": len([row for row in rows if row["is_tracked"]]),
            "owned_count": len([row for row in rows if row["status_key"] == "owned"]),
            "watchlist_count": len([row for row in rows if row["status_key"] == "watchlist"]),
            "guide_only_count": len([row for row in rows if row["status_key"] == "guide"]),
        },
    }


@lru_cache(maxsize=512)
def _fetch_market_series_cached(
    symbol: str,
    range_key: str = DEFAULT_MARKET_RANGE_KEY,
    interval: str = "1d",
    cache_bucket: int = 0,
) -> MarketSeries:
    params = urlencode({"range": range_key, "interval": interval, "includePrePost": "false"})
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?{params}"
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=market_request_timeout_seconds()) as response:
        payload = json.load(response)

    error = payload.get("chart", {}).get("error")
    if error:
        raise MarketDataError(error.get("description", f"No se han podido cargar los datos de mercado de {symbol}."))

    result = payload["chart"]["result"][0]
    meta = result["meta"]
    timestamps = result.get("timestamp", [])
    quote_data = result["indicators"]["quote"][0]
    opens = quote_data.get("open", [])
    highs = quote_data.get("high", [])
    lows = quote_data.get("low", [])
    closes = quote_data.get("close", [])
    points = []

    for index, timestamp in enumerate(timestamps):
        close = closes[index] if index < len(closes) else None
        if close is None:
            continue
        open_value = opens[index] if index < len(opens) else None
        high_value = highs[index] if index < len(highs) else None
        low_value = lows[index] if index < len(lows) else None
        points.append(
            {
                "date": datetime.fromtimestamp(timestamp, tz=timezone.utc).date(),
                "open": Decimal(str(round(open_value, 4))) if open_value is not None else None,
                "high": Decimal(str(round(high_value, 4))) if high_value is not None else None,
                "low": Decimal(str(round(low_value, 4))) if low_value is not None else None,
                "close": Decimal(str(round(close, 4))),
            }
        )

    if not points:
        raise MarketDataError(f"No se han recibido precios historicos para {symbol}.")

    latest_raw = meta.get("regularMarketPrice")
    latest_timestamp = meta.get("regularMarketTime") or timestamps[-1]
    latest_price = Decimal(str(round(latest_raw if latest_raw is not None else float(points[-1]["close"]), 4)))
    latest_date = datetime.fromtimestamp(latest_timestamp, tz=timezone.utc).date()

    return MarketSeries(
        symbol=symbol,
        name=meta.get("longName") or meta.get("shortName") or symbol,
        latest_price=latest_price,
        latest_date=latest_date,
        points=points,
    )


def fetch_market_series(symbol: str, range_key: str = DEFAULT_MARKET_RANGE_KEY, interval: str = "1d") -> MarketSeries:
    normalized_symbol = clean_symbol(symbol)
    return clone_market_series(
        _fetch_market_series_cached(
            normalized_symbol,
            range_key,
            interval,
            build_market_data_cache_bucket(),
        )
    )


def should_fetch_equity_fundamentals() -> bool:
    return bool(getattr(settings, "EQUITIES_FETCH_FUNDAMENTALS", True))


def build_fundamentals_period_bounds(now: datetime | None = None) -> tuple[int, int]:
    current = now or django_timezone.now()
    period2 = int(current.timestamp())
    period1 = int((current - timedelta(days=365 * 5)).timestamp())
    return period1, period2


def parse_yahoo_timeseries_metric_rows(rows: list[dict] | None) -> list[dict]:
    parsed_rows = []
    for row in rows or []:
        as_of_date_raw = row.get("asOfDate")
        reported_value = row.get("reportedValue", {})
        raw_value = reported_value.get("raw")
        if not as_of_date_raw or raw_value in {None, ""}:
            continue
        try:
            as_of_date = date.fromisoformat(as_of_date_raw)
            value = Decimal(str(raw_value))
        except Exception:
            continue
        parsed_rows.append(
            {
                "as_of_date": as_of_date,
                "period_type": row.get("periodType", ""),
                "currency_code": row.get("currencyCode", ""),
                "value": value,
            }
        )
    parsed_rows.sort(key=lambda item: item["as_of_date"])
    return parsed_rows


def clone_equity_fundamentals_snapshot(snapshot: dict) -> dict:
    return {
        **snapshot,
        "net_income_rows": [dict(row) for row in snapshot.get("net_income_rows", [])],
        "market_cap_rows": [dict(row) for row in snapshot.get("market_cap_rows", [])],
    }


@lru_cache(maxsize=256)
def _fetch_equity_fundamentals_cached(symbol: str, cache_bucket: int = 0) -> dict:
    period1, period2 = build_fundamentals_period_bounds()
    requested_types = ",".join(
        (
            "annualNetIncome",
            "trailingMarketCap",
            "annualMarketCap",
            "quarterlyMarketCap",
        )
    )
    url = (
        "https://query1.finance.yahoo.com/ws/fundamentals-timeseries/v1/finance/timeseries/"
        f"{symbol}?type={requested_types}&period1={period1}&period2={period2}"
    )
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=market_request_timeout_seconds()) as response:
        payload = json.load(response)

    timeseries = payload.get("timeseries", {})
    if timeseries.get("error"):
        description = timeseries["error"].get("description") or f"No se han podido cargar los fundamentales de {symbol}."
        raise MarketDataError(description)

    parsed_metrics = {}
    for item in timeseries.get("result", []):
        metric_keys = [key for key in item.keys() if key not in {"meta", "timestamp"}]
        for metric_key in metric_keys:
            parsed_metrics[metric_key] = parse_yahoo_timeseries_metric_rows(item.get(metric_key))

    net_income_rows = parsed_metrics.get("annualNetIncome", [])
    market_cap_rows = (
        parsed_metrics.get("trailingMarketCap")
        or parsed_metrics.get("quarterlyMarketCap")
        or parsed_metrics.get("annualMarketCap")
        or []
    )
    if not net_income_rows and not market_cap_rows:
        raise MarketDataError(f"No se han recibido fundamentales suficientes para {symbol}.")

    latest_market_cap = market_cap_rows[-1] if market_cap_rows else {}
    currency_code = (
        latest_market_cap.get("currency_code")
        or (net_income_rows[-1].get("currency_code") if net_income_rows else "")
        or ""
    )
    return {
        "symbol": symbol,
        "available": True,
        "currency_code": currency_code,
        "net_income_rows": net_income_rows[-3:],
        "market_cap_rows": market_cap_rows[-3:],
        "market_cap": latest_market_cap.get("value"),
        "market_cap_as_of_date": latest_market_cap.get("as_of_date"),
    }


def fetch_equity_fundamentals(symbol: str) -> dict:
    normalized_symbol = clean_symbol(symbol)
    return clone_equity_fundamentals_snapshot(
        _fetch_equity_fundamentals_cached(
            normalized_symbol,
            build_market_data_cache_bucket(),
        )
    )


def fetch_ecb_reference_series(
    series_key: str,
    series_name: str,
    last_n_observations: int = LONG_MONTHLY_OBSERVATIONS,
) -> MarketSeries:
    url = (
        "https://data-api.ecb.europa.eu/service/data/RTD/"
        f"{series_key}?format=jsondata&lastNObservations={last_n_observations}"
    )
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=20) as response:
        payload = json.load(response)

    time_values = payload["structure"]["dimensions"]["observation"][0]["values"]
    series_map = payload["dataSets"][0]["series"]
    if not series_map:
        raise MarketDataError(f"No se han recibido observaciones del BCE para {series_name}.")

    observations = next(iter(series_map.values()))["observations"]
    points = []
    for raw_index, values in observations.items():
        period_label = time_values[int(raw_index)]["id"]
        year, month = map(int, period_label.split("-"))
        points.append(
            {
                "date": date(year, month, 1),
                "open": None,
                "high": None,
                "low": None,
                "close": Decimal(str(values[0])),
            }
        )

    if not points:
        raise MarketDataError(f"No se han recibido observaciones del BCE para {series_name}.")

    latest = points[-1]
    return MarketSeries(
        symbol=f"ECB:{series_key}",
        name=series_name,
        latest_price=latest["close"],
        latest_date=latest["date"],
        points=points,
    )


def parse_sdmx_period_label_to_date(period_label: str) -> date:
    text = str(period_label or "").strip()
    if re.fullmatch(r"\d{4}-Q[1-4]", text):
        return parse_eurostat_quarter_label(text)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return date.fromisoformat(text)
    if re.fullmatch(r"\d{4}-\d{2}", text):
        year, month = map(int, text.split("-"))
        return date(year, month, 1)
    if re.fullmatch(r"\d{4}", text):
        return date(int(text), 12, 31)
    raise MarketDataError(f"No se ha podido interpretar el periodo SDMX '{text}'.")


def build_market_series_from_sdmx_payload(payload: dict, *, symbol: str, name: str) -> MarketSeries:
    time_values = payload["structure"]["dimensions"]["observation"][0]["values"]
    series_map = payload["dataSets"][0]["series"]
    if not series_map:
        raise MarketDataError(f"No se han recibido observaciones SDMX para {name}.")

    observations = next(iter(series_map.values()))["observations"]
    points = []
    for raw_index, values in observations.items():
        point_date = parse_sdmx_period_label_to_date(time_values[int(raw_index)]["id"])
        points.append(
            {
                "date": point_date,
                "open": None,
                "high": None,
                "low": None,
                "close": Decimal(str(values[0])),
            }
        )

    if not points:
        raise MarketDataError(f"No se han recibido observaciones SDMX para {name}.")

    latest = points[-1]
    return MarketSeries(
        symbol=symbol,
        name=name,
        latest_price=latest["close"],
        latest_date=latest["date"],
        points=points,
    )


@lru_cache(maxsize=32)
def _fetch_ecb_yield_curve_series_cached(
    series_key: str,
    series_name: str,
    last_n_observations: int = LONG_MONTHLY_OBSERVATIONS,
    cache_bucket: int = 0,
) -> MarketSeries:
    url = (
        "https://data-api.ecb.europa.eu/service/data/YC/"
        f"{series_key}?format=jsondata&lastNObservations={last_n_observations}"
    )
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=20) as response:
        payload = json.load(response)

    return build_market_series_from_sdmx_payload(
        payload,
        symbol=f"ECB:YC.{series_key}",
        name=series_name,
    )


def fetch_ecb_yield_curve_series(
    series_key: str,
    series_name: str,
    last_n_observations: int = LONG_MONTHLY_OBSERVATIONS,
) -> MarketSeries:
    return clone_market_series(
        _fetch_ecb_yield_curve_series_cached(
            series_key,
            series_name,
            last_n_observations,
            build_market_data_cache_bucket(),
        )
    )


def parse_eurostat_quarter_label(value: str) -> date:
    year_text, quarter_text = value.split("-Q")
    year = int(year_text)
    quarter = int(quarter_text)
    month = quarter * 3
    month_day = {3: 31, 6: 30, 9: 30, 12: 31}[month]
    return date(year, month, month_day)


def fetch_eurostat_house_price_series() -> MarketSeries:
    url = (
        "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"
        "prc_hpi_q?geo=ES&purchase=TOTAL&unit=I15_Q"
    )
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=20) as response:
        payload = json.load(response)

    time_index = payload["dimension"]["time"]["category"]["index"]
    value_map = payload.get("value", {})
    points = []
    for label, raw_index in sorted(time_index.items(), key=lambda item: item[1]):
        if str(raw_index) not in value_map:
            continue
        points.append(
            {
                "date": parse_eurostat_quarter_label(label),
                "open": None,
                "high": None,
                "low": None,
                "close": Decimal(str(value_map[str(raw_index)])),
            }
        )

    if not points:
        raise MarketDataError("No se han recibido observaciones del indice de vivienda de Eurostat.")

    latest = points[-1]
    return MarketSeries(
        symbol=SPAIN_HOUSE_PRICE_SYMBOL,
        name=SPAIN_HOUSE_PRICE_NAME,
        latest_price=latest["close"],
        latest_date=latest["date"],
        points=points,
    )


def parse_datetime_to_date(value: str) -> date:
    normalized = value.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized).date()
    except ValueError:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()


def fetch_ree_electricity_demand_series() -> MarketSeries:
    end_date = datetime.now(timezone.utc).date()
    start_date = end_date - timedelta(days=LONG_ANALYSIS_DAYS)
    params = urlencode(
        {
            "start_date": f"{start_date.isoformat()}T00:00",
            "end_date": f"{end_date.isoformat()}T23:59",
            "time_trunc": "month",
            "geo_trunc": "electric_system",
            "geo_limit": "peninsular",
            "geo_ids": "8741",
        }
    )
    url = f"https://apidatos.ree.es/es/datos/demanda/evolucion?{params}"
    request = Request(url, headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"})
    with urlopen(request, timeout=20) as response:
        payload = json.load(response)

    series = (payload.get("included") or [{}])[0]
    values = series.get("attributes", {}).get("values", [])
    points = [
        {
            "date": parse_datetime_to_date(item["datetime"]),
            "open": None,
            "high": None,
            "low": None,
            "close": Decimal(str(item["value"])),
        }
        for item in values
        if item.get("value") is not None and item.get("datetime")
    ]
    if not points:
        raise MarketDataError("No se han recibido observaciones de demanda electrica de REE.")

    latest = points[-1]
    return MarketSeries(
        symbol=SPAIN_ELECTRICITY_DEMAND_SYMBOL,
        name=SPAIN_ELECTRICITY_DEMAND_NAME,
        latest_price=latest["close"],
        latest_date=latest["date"],
        points=points,
    )


def fetch_eurostat_gas_consumption_series() -> MarketSeries:
    url = (
        "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"
        "nrg_cb_gasm?freq=M&nrg_bal=IC_OBS&siec=G3000&unit=TJ_GCV&geo=ES"
    )
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=20) as response:
        payload = json.load(response)

    time_index = payload["dimension"]["time"]["category"]["index"]
    value_map = payload.get("value", {})
    points = []
    for label, raw_index in sorted(time_index.items(), key=lambda item: item[1]):
        raw_value = value_map.get(str(raw_index))
        if raw_value is None:
            continue
        year, month = map(int, label.split("-"))
        points.append(
            {
                "date": date(year, month, 1),
                "open": None,
                "high": None,
                "low": None,
                "close": Decimal(str(raw_value)),
            }
        )

    if not points:
        raise MarketDataError("No se han recibido observaciones del consumo de gas de Eurostat.")

    latest = points[-1]
    return MarketSeries(
        symbol=SPAIN_GAS_CONSUMPTION_SYMBOL,
        name=SPAIN_GAS_CONSUMPTION_NAME,
        latest_price=latest["close"],
        latest_date=latest["date"],
        points=points,
    )


@lru_cache(maxsize=256)
def _fetch_reference_series_for_choice_cached(
    reference_profile: str,
    benchmark_symbol: str = "",
    benchmark_name: str = "",
    range_key: str = DEFAULT_MARKET_RANGE_KEY,
    cache_bucket: int = 0,
) -> MarketSeries | None:
    if reference_profile == EquityPosition.ReferenceProfile.EURIBOR_12M:
        return fetch_ecb_reference_series("M.S0.N.C_EUR1Y.E", EURIBOR_REFERENCE_NAME)
    if reference_profile == EquityPosition.ReferenceProfile.SPAIN_HOUSE_PRICE:
        return fetch_eurostat_house_price_series()
    if reference_profile == EquityPosition.ReferenceProfile.SPAIN_ELECTRICITY_DEMAND:
        return fetch_ree_electricity_demand_series()
    if reference_profile == EquityPosition.ReferenceProfile.SPAIN_GAS_CONSUMPTION:
        return fetch_eurostat_gas_consumption_series()
    if benchmark_symbol:
        return _fetch_market_series_cached(
            clean_symbol(benchmark_symbol),
            range_key,
            "1d",
            cache_bucket,
        )
    return None


def fetch_reference_series_for_choice(
    reference_profile: str,
    benchmark_symbol: str = "",
    benchmark_name: str = "",
    range_key: str = DEFAULT_MARKET_RANGE_KEY,
) -> MarketSeries | None:
    return clone_market_series(
        _fetch_reference_series_for_choice_cached(
            reference_profile,
            benchmark_symbol,
            benchmark_name,
            range_key,
            build_market_data_cache_bucket(),
        )
    )


def fetch_reference_series(position: EquityPosition, range_key: str = DEFAULT_MARKET_RANGE_KEY) -> MarketSeries | None:
    return fetch_reference_series_for_choice(
        position.reference_profile,
        benchmark_symbol=position.benchmark_symbol,
        benchmark_name=position.benchmark_name,
        range_key=range_key,
    )


def normalize_document_text(value: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_text).strip().upper()


def parse_document_decimal(value) -> Decimal | None:
    text = str(value or "").strip().replace("\xa0", "").replace(" ", "")
    text = text.replace("EUR", "").replace("€", "").replace("%", "")
    if not text:
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]
    if text.startswith("-"):
        negative = True
        text = text[1:]
    if text.endswith("-"):
        negative = True
        text = text[:-1]

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            normalized = text.replace(".", "").replace(",", ".")
        else:
            normalized = text.replace(",", "")
    elif "," in text:
        normalized = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        normalized = text.replace(".", "")
    else:
        normalized = text

    try:
        number = Decimal(normalized)
    except Exception:
        return None
    return -number if negative else number


def clean_company_name(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def clean_symbol(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().upper())


def clean_ticker(value: str) -> str:
    symbol = clean_symbol(value)
    if "." in symbol:
        symbol = symbol.split(".", 1)[0]
    return symbol


def infer_ownership_category_from_text(text: str, default_category: str) -> str:
    normalized = normalize_document_text(text)
    has_ximo = "XIMO" in normalized
    has_monica = "MONICA" in normalized
    if has_ximo and has_monica:
        return AssetOwnershipCategory.JOINT
    if has_ximo:
        return AssetOwnershipCategory.XIMO
    if has_monica:
        return AssetOwnershipCategory.MONICA
    return default_category


def label_matches(label: str, *tokens: str) -> bool:
    return any(token in label for token in tokens)


def build_equity_column_map(row: list[str]) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for index, cell in enumerate(row):
        label = normalize_document_text(cell)
        if not label:
            continue
        if "broker" not in column_map and label_matches(label, "BROKER", "ENTIDAD", "CUSTODIO", "INTERMEDIARIO"):
            column_map["broker"] = index
        elif "ticker" not in column_map and label_matches(label, "TICKER", "SIMBOLO", "SYMBOL", "CODIGO VALOR"):
            column_map["ticker"] = index
        elif "quote_symbol" not in column_map and label_matches(label, "SIMBOLO DE MERCADO", "MARKET SYMBOL", "COTIZACION"):
            column_map["quote_symbol"] = index
        elif "company_name" not in column_map and label_matches(
            label,
            "EMPRESA",
            "COMPANIA",
            "COMPANY",
            "DESCRIPCION",
            "INSTRUMENTO",
            "VALOR",
            "NOMBRE",
        ):
            column_map["company_name"] = index
        elif "shares" not in column_map and label_matches(label, "ACCIONES", "TITULOS", "SHARES", "UNIDADES", "CANTIDAD"):
            column_map["shares"] = index
        elif "average_cost_per_share" not in column_map and label_matches(
            label,
            "COSTE MEDIO",
            "PRECIO MEDIO",
            "AVERAGE COST",
            "COST BASIS",
        ):
            column_map["average_cost_per_share"] = index
        elif "current_price_per_share" not in column_map and label_matches(
            label,
            "PRECIO ACTUAL",
            "ULTIMO PRECIO",
            "LAST PRICE",
            "MARKET PRICE",
            "COTIZACION",
        ):
            column_map["current_price_per_share"] = index
        elif "current_value" not in column_map and label_matches(
            label,
            "VALOR ACTUAL",
            "VALOR MERCADO",
            "MARKET VALUE",
            "VALORACION",
        ):
            column_map["current_value"] = index
        elif "invested_amount" not in column_map and label_matches(
            label,
            "IMPORTE INVERTIDO",
            "COSTE TOTAL",
            "TOTAL COSTE",
            "INVESTED AMOUNT",
        ):
            column_map["invested_amount"] = index
        elif "annual_dividend_income" not in column_map and label_matches(
            label,
            "DIVIDENDO ANUAL",
            "ANNUAL DIVIDEND",
            "DIVIDENDOS ANUALES",
        ):
            column_map["annual_dividend_income"] = index
    return column_map


def has_useful_equity_columns(column_map: dict[str, int]) -> bool:
    return (
        ("ticker" in column_map or "company_name" in column_map)
        and "shares" in column_map
        and (
            "average_cost_per_share" in column_map
            or "current_price_per_share" in column_map
            or "invested_amount" in column_map
            or "current_value" in column_map
        )
    )


def get_row_cell(row: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return str(row[index]).strip()


def build_equity_data_from_table_row(
    row: list[str],
    column_map: dict[str, int],
    default_broker: str,
    default_ownership_category: str,
) -> dict:
    shares = parse_document_decimal(get_row_cell(row, column_map.get("shares")))
    if shares in {None, ZERO}:
        return {}

    invested_amount = parse_document_decimal(get_row_cell(row, column_map.get("invested_amount")))
    current_value = parse_document_decimal(get_row_cell(row, column_map.get("current_value")))
    average_cost = parse_document_decimal(get_row_cell(row, column_map.get("average_cost_per_share")))
    current_price = parse_document_decimal(get_row_cell(row, column_map.get("current_price_per_share")))

    if average_cost is None and invested_amount is not None and shares:
        average_cost = invested_amount / shares
    if current_price is None and current_value is not None and shares:
        current_price = current_value / shares
    if current_price is None and average_cost is not None:
        current_price = average_cost

    ticker = clean_ticker(get_row_cell(row, column_map.get("ticker")))
    quote_symbol = clean_symbol(get_row_cell(row, column_map.get("quote_symbol")))
    if not ticker and quote_symbol:
        ticker = clean_ticker(quote_symbol)

    company_name = clean_company_name(get_row_cell(row, column_map.get("company_name")))
    broker = clean_company_name(get_row_cell(row, column_map.get("broker"))) or default_broker
    annual_dividend_income = parse_document_decimal(get_row_cell(row, column_map.get("annual_dividend_income"))) or ZERO
    ownership_category = infer_ownership_category_from_text(
        " ".join(str(cell) for cell in row),
        default_ownership_category,
    )

    if not ticker and not company_name:
        return {}

    return apply_equity_company_defaults(
        {
            "position_kind": EquityPosition.PositionKind.OWNED,
            "ownership_category": ownership_category,
            "broker": broker,
            "ticker": ticker,
            "company_name": company_name,
        "quote_symbol": quote_symbol,
        "reference_profile": EquityPosition.ReferenceProfile.MARKET_INDEX,
        "benchmark_symbol": DEFAULT_BENCHMARK_SYMBOL,
        "benchmark_name": DEFAULT_BENCHMARK_NAME,
        "shares": shares,
        "average_cost_per_share": average_cost,
            "current_price_per_share": current_price,
            "annual_dividend_income": annual_dividend_income,
            "notes": "",
        }
    )


def assign_prefill_value(target: dict, label: str, value: str):
    if not value:
        return
    if label_matches(label, "BROKER", "ENTIDAD", "CUSTODIO", "INTERMEDIARIO"):
        target["broker"] = clean_company_name(value)
    elif label_matches(label, "TICKER", "SIMBOLO", "SYMBOL", "CODIGO VALOR"):
        target["ticker"] = clean_ticker(value)
    elif label_matches(label, "SIMBOLO DE MERCADO", "MARKET SYMBOL", "COTIZACION"):
        target["quote_symbol"] = clean_symbol(value)
    elif label_matches(label, "EMPRESA", "COMPANIA", "COMPANY", "DESCRIPCION", "INSTRUMENTO", "NOMBRE"):
        target["company_name"] = clean_company_name(value)
    elif label_matches(label, "ACCIONES", "TITULOS", "SHARES", "UNIDADES", "CANTIDAD"):
        number = parse_document_decimal(value)
        if number is not None:
            target["shares"] = number
    elif label_matches(label, "COSTE MEDIO", "PRECIO MEDIO", "AVERAGE COST", "COST BASIS"):
        number = parse_document_decimal(value)
        if number is not None:
            target["average_cost_per_share"] = number
    elif label_matches(label, "PRECIO ACTUAL", "ULTIMO PRECIO", "LAST PRICE", "MARKET PRICE", "COTIZACION"):
        number = parse_document_decimal(value)
        if number is not None:
            target["current_price_per_share"] = number
    elif label_matches(label, "VALOR ACTUAL", "VALOR MERCADO", "MARKET VALUE", "VALORACION"):
        number = parse_document_decimal(value)
        if number is not None:
            target["current_value"] = number
    elif label_matches(label, "IMPORTE INVERTIDO", "COSTE TOTAL", "TOTAL COSTE", "INVESTED AMOUNT"):
        number = parse_document_decimal(value)
        if number is not None:
            target["invested_amount"] = number
    elif label_matches(label, "DIVIDENDO ANUAL", "ANNUAL DIVIDEND", "DIVIDENDOS ANUALES"):
        number = parse_document_decimal(value)
        if number is not None:
            target["annual_dividend_income"] = number


def build_equity_data_from_key_value_rows(
    rows: list[list[str]],
    default_broker: str,
    default_ownership_category: str,
) -> dict:
    parsed: dict = {}
    for row in rows[:80]:
        non_empty = [str(cell).strip() for cell in row if str(cell).strip()]
        if len(non_empty) < 2:
            continue
        pairs = [(non_empty[0], non_empty[1])]
        if len(non_empty) >= 4:
            pairs.append((non_empty[2], non_empty[3]))
        for raw_label, raw_value in pairs:
            assign_prefill_value(parsed, normalize_document_text(raw_label), raw_value)

    if not parsed:
        return {}

    if "quote_symbol" in parsed and "ticker" not in parsed:
        parsed["ticker"] = clean_ticker(parsed["quote_symbol"])

    shares = parsed.get("shares")
    invested_amount = parsed.get("invested_amount")
    current_value = parsed.get("current_value")
    average_cost = parsed.get("average_cost_per_share")
    current_price = parsed.get("current_price_per_share")

    if average_cost is None and shares and invested_amount is not None:
        parsed["average_cost_per_share"] = invested_amount / shares
    if current_price is None and shares and current_value is not None:
        parsed["current_price_per_share"] = current_value / shares
    if parsed.get("current_price_per_share") is None and parsed.get("average_cost_per_share") is not None:
        parsed["current_price_per_share"] = parsed["average_cost_per_share"]

    parsed.setdefault("broker", default_broker)
    parsed.setdefault("position_kind", EquityPosition.PositionKind.OWNED)
    parsed.setdefault("ownership_category", default_ownership_category)
    parsed.setdefault("reference_profile", EquityPosition.ReferenceProfile.MARKET_INDEX)
    parsed.setdefault("benchmark_symbol", DEFAULT_BENCHMARK_SYMBOL)
    parsed.setdefault("benchmark_name", DEFAULT_BENCHMARK_NAME)
    parsed.setdefault("annual_dividend_income", ZERO)
    parsed.setdefault("notes", "")

    if not parsed.get("ticker") and not parsed.get("company_name"):
        return {}
    if parsed.get("shares") in {None, ZERO}:
        return {}
    if parsed.get("average_cost_per_share") is None:
        return {}

    return apply_equity_company_defaults(parsed)


def score_equity_data(data: dict) -> int:
    return sum(
        1
        for field_name in (
            "broker",
            "ticker",
            "company_name",
            "quote_symbol",
            "shares",
            "average_cost_per_share",
            "current_price_per_share",
            "annual_dividend_income",
        )
        if data.get(field_name) not in {None, "", ZERO}
    )


def finalize_equity_prefill(data: dict, default_ownership_category: str, document_text: str) -> dict:
    result = apply_equity_company_defaults(data)
    result["ownership_category"] = infer_ownership_category_from_text(
        document_text,
        result.get("ownership_category") or default_ownership_category,
    )
    result.setdefault("position_kind", EquityPosition.PositionKind.OWNED)
    result.setdefault("reference_profile", EquityPosition.ReferenceProfile.MARKET_INDEX)
    result.setdefault("benchmark_symbol", DEFAULT_BENCHMARK_SYMBOL)
    result.setdefault("benchmark_name", DEFAULT_BENCHMARK_NAME)
    result.setdefault("annual_dividend_income", ZERO)
    result.setdefault("notes", "")
    result.setdefault("broker", "")
    result.setdefault("quote_symbol", "")
    if result.get("quote_symbol") and not result.get("ticker"):
        result["ticker"] = clean_ticker(result["quote_symbol"])
    if result.get("current_price_per_share") is None and result.get("average_cost_per_share") is not None:
        result["current_price_per_share"] = result["average_cost_per_share"]
    return result


def extract_equity_prefill_from_rows(
    rows: list[list[str]],
    default_broker: str,
    default_ownership_category: str,
) -> tuple[dict, int]:
    document_text = " ".join(" ".join(str(cell) for cell in row) for row in rows)
    candidates = []

    key_value_data = build_equity_data_from_key_value_rows(rows, default_broker, default_ownership_category)
    if key_value_data:
        candidates.append(key_value_data)

    for index, row in enumerate(rows):
        column_map = build_equity_column_map(row)
        if not has_useful_equity_columns(column_map):
            continue
        table_candidates = []
        for data_row in rows[index + 1 :]:
            candidate = build_equity_data_from_table_row(
                data_row,
                column_map,
                default_broker,
                default_ownership_category,
            )
            if candidate:
                table_candidates.append(candidate)
        if table_candidates:
            candidates.extend(table_candidates)
            break

    if not candidates:
        for index, row in enumerate(rows):
            candidate = build_equity_data_from_table_row(
                row,
                DEFAULT_EQUITY_COLUMN_MAP,
                default_broker,
                default_ownership_category,
            )
            if candidate:
                candidates.append(candidate)
                for data_row in rows[index + 1 :]:
                    next_candidate = build_equity_data_from_table_row(
                        data_row,
                        DEFAULT_EQUITY_COLUMN_MAP,
                        default_broker,
                        default_ownership_category,
                    )
                    if next_candidate:
                        candidates.append(next_candidate)
                break

    if not candidates:
        return {}, 0

    best = max(candidates, key=score_equity_data)
    return finalize_equity_prefill(best, default_ownership_category, document_text), len(candidates)


def read_pdf_pages(file_source) -> list[str]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise EquityDocumentImportError("Se necesita pypdf para leer documentos PDF de acciones.") from exc

    if hasattr(file_source, "open"):
        file_source.open("rb")
    try:
        if hasattr(file_source, "seek"):
            file_source.seek(0)
        reader = PdfReader(file_source)
        return [(page.extract_text() or "") for page in reader.pages]
    finally:
        if hasattr(file_source, "seek"):
            file_source.seek(0)
        if hasattr(file_source, "close"):
            file_source.close()


def extract_equity_position_prefill(
    uploaded_file,
    default_broker: str = "",
    default_ownership_category: str = AssetOwnershipCategory.JOINT,
) -> EquityDocumentPrefill:
    source_name = str(getattr(uploaded_file, "name", "documento")).lower()
    suffix = Path(source_name).suffix.lower()

    if suffix in {".xls", ".xlsx"}:
        try:
            rows = load_rows_from_workbook(uploaded_file)
        except ValidationError as exc:
            raise EquityDocumentImportError(str(exc)) from exc
        data, candidate_count = extract_equity_prefill_from_rows(rows, default_broker, default_ownership_category)
        source_kind = "XLS"
    elif suffix == ".pdf":
        pages = read_pdf_pages(uploaded_file)
        rows = []
        for page in pages:
            for line in page.splitlines():
                cells = [segment.strip() for segment in re.split(r"\s{2,}|\t+", line) if segment.strip()]
                if cells:
                    rows.append(cells)
        data, candidate_count = extract_equity_prefill_from_rows(rows, default_broker, default_ownership_category)
        source_kind = "PDF"
    else:
        raise EquityDocumentImportError("Tipo de fichero no compatible. Sube un XLS, XLSX o PDF.")

    if not data:
        raise EquityDocumentImportError(
            "No se han reconocido suficientes datos de la posicion en el documento. Revisa el fichero o completa el formulario manualmente."
        )

    detected_fields = [
        field_name
        for field_name in (
            "ownership_category",
            "broker",
            "ticker",
            "company_name",
            "quote_symbol",
            "shares",
            "average_cost_per_share",
            "current_price_per_share",
            "annual_dividend_income",
        )
        if data.get(field_name) not in {None, ""}
    ]

    return EquityDocumentPrefill(
        data=data,
        detected_fields=detected_fields,
        candidate_count=candidate_count,
        source_kind=source_kind,
    )


def align_reference_points(position_points: list[dict], reference_points: list[dict]) -> dict[date, Decimal | None]:
    ordered_reference = sorted(reference_points, key=lambda item: item["date"])
    aligned: dict[date, Decimal | None] = {}
    current_reference = None
    reference_index = 0

    for point in sorted(position_points, key=lambda item: item["date"]):
        while reference_index < len(ordered_reference) and ordered_reference[reference_index]["date"] <= point["date"]:
            current_reference = ordered_reference[reference_index]["close"]
            reference_index += 1
        aligned[point["date"]] = current_reference
    return aligned


def sync_equity_market_data(position: EquityPosition) -> EquityPosition:
    if not position.quote_symbol:
        raise MarketDataError(f"{position.ticker} no tiene configurado un simbolo de cotizacion.")

    position_series = fetch_market_series(position.quote_symbol)
    benchmark_series = fetch_reference_series(position)
    benchmark_map = align_reference_points(position_series.points, benchmark_series.points) if benchmark_series else {}
    point_dates = {point["date"] for point in position_series.points}

    with transaction.atomic():
        EquityPriceHistory.objects.filter(position=position).exclude(price_date__in=point_dates).delete()
        for point in position_series.points:
            EquityPriceHistory.objects.update_or_create(
                position=position,
                price_date=point["date"],
                defaults={
                    "open_price": point.get("open"),
                    "high_price": point.get("high"),
                    "low_price": point.get("low"),
                    "close_price": point["close"],
                    "benchmark_close": benchmark_map.get(point["date"]),
                },
            )

        position.current_price_per_share = position_series.latest_price
        position.latest_price_date = position_series.latest_date
        position.last_synced_at = django_timezone.now()
        if benchmark_series:
            position.benchmark_name = benchmark_series.name
            position.benchmark_symbol = benchmark_series.symbol
        position.save(
            update_fields=[
                "current_price_per_share",
                "latest_price_date",
                "last_synced_at",
                "benchmark_symbol",
                "benchmark_name",
            ]
        )

    return position


def sync_all_equities_market_data(positions) -> list[tuple[EquityPosition, str | None]]:
    results = []
    for position in positions:
        try:
            sync_equity_market_data(position)
            results.append((position, None))
        except Exception as exc:
            results.append((position, str(exc)))
    return results


def build_svg_polyline(values, width: int = 640, height: int = 220, padding: int = 18) -> str:
    filtered = [value for value in values if value is not None]
    if len(filtered) < 2:
        return ""

    min_value = min(filtered)
    max_value = max(filtered)
    if max_value == min_value:
        max_value += 1

    span_x = width - 2 * padding
    span_y = height - 2 * padding
    points = []
    total_points = len(values) - 1 or 1
    for index, value in enumerate(values):
        if value is None:
            continue
        x = padding + (span_x * index / total_points)
        normalized = (value - min_value) / (max_value - min_value)
        y = height - padding - (normalized * span_y)
        points.append(f"{x:.1f},{y:.1f}")
    return " ".join(points)


def format_axis_value(value: Decimal | None) -> str:
    if value is None:
        return "-"
    absolute = abs(value)
    if absolute >= Decimal("1000000"):
        return f"{(value / Decimal('1000000')):.1f}M"
    if absolute >= Decimal("1000"):
        return f"{(value / Decimal('1000')):.1f}k"
    if absolute >= Decimal("100"):
        return f"{value:.0f}"
    if absolute >= Decimal("10"):
        return f"{value:.1f}"
    return f"{value:.2f}"


def format_percentage_axis_value(value: Decimal | None) -> str:
    if value is None:
        return "-"
    absolute = abs(value)
    if absolute >= Decimal("1000"):
        return f"{value:.0f}"
    if absolute >= Decimal("100"):
        return f"{value:.0f}"
    if absolute >= Decimal("10"):
        return f"{value:.1f}"
    return f"{value:.2f}"


def build_time_axis_markers(
    point_dates: list[date],
    width: int = 640,
    height: int = 220,
    padding: int = 18,
) -> list[dict]:
    filtered_dates = sorted({point_date for point_date in point_dates if point_date is not None})
    if len(filtered_dates) < 2:
        return []

    min_date = filtered_dates[0]
    max_date = filtered_dates[-1]
    total_days = max((max_date - min_date).days, 1)
    span_x = width - (padding * 2)

    def to_x(point_date: date) -> float:
        return padding + (span_x * ((point_date - min_date).days / total_days))

    markers = []
    if min_date.year != max_date.year:
        for year in range(min_date.year, max_date.year + 1):
            marker_date = date(year, 1, 1)
            if marker_date < min_date or marker_date > max_date:
                continue
            markers.append(
                {
                    "x": f"{to_x(marker_date):.1f}",
                    "label": str(year),
                    "date": marker_date.isoformat(),
                    "y1": str(height - padding),
                    "y2": str(height - padding + 6),
                    "text_y": str(height - 2),
                }
            )

    if len(markers) < 2:
        fallback_dates = [min_date]
        if min_date.year != max_date.year:
            fallback_dates.append(max_date)
        else:
            midpoint = min_date + timedelta(days=total_days // 2)
            fallback_dates.extend([midpoint, max_date])

        seen_labels = set()
        markers = []
        for marker_date in fallback_dates:
            label = str(marker_date.year) if min_date.year != max_date.year else marker_date.strftime("%Y-%m")
            if label in seen_labels:
                continue
            seen_labels.add(label)
            markers.append(
                {
                    "x": f"{to_x(marker_date):.1f}",
                    "label": label,
                    "date": marker_date.isoformat(),
                    "y1": str(height - padding),
                    "y2": str(height - padding + 6),
                    "text_y": str(height - 2),
                }
            )

    return markers


def build_month_axis_markers(
    point_dates: list[date],
    width: int = 640,
    height: int = 220,
    padding: int = 18,
    max_labels: int = 8,
) -> list[dict]:
    filtered_dates = sorted({point_date for point_date in point_dates if point_date is not None})
    if len(filtered_dates) < 2:
        return []

    min_date = filtered_dates[0]
    max_date = filtered_dates[-1]
    total_days = max((max_date - min_date).days, 1)
    span_x = width - (padding * 2)

    def to_x(point_date: date) -> float:
        return padding + (span_x * ((point_date - min_date).days / total_days))

    markers: list[dict] = [
        {
            "marker_date": min_date,
            "draw_grid": False,
        }
    ]
    cursor_year = min_date.year
    cursor_month = min_date.month + 1
    if cursor_month == 13:
        cursor_month = 1
        cursor_year += 1

    while True:
        marker_date = date(cursor_year, cursor_month, 1)
        if marker_date > max_date:
            break
        markers.append(
            {
                "marker_date": marker_date,
                "draw_grid": True,
            }
        )
        cursor_month += 1
        if cursor_month == 13:
            cursor_month = 1
            cursor_year += 1

    if len(markers) < 2:
        return build_time_axis_markers(
            point_dates,
            width=width,
            height=height,
            padding=padding,
        )

    label_step = max(1, int(math.ceil(len(markers) / max(max_labels, 1))))
    serialized_markers = []
    for index, marker in enumerate(markers):
        marker_date = marker["marker_date"]
        short_month_label = SPANISH_MONTH_LABELS.get(marker_date.month, str(marker_date.month))[:3].capitalize()
        serialized_markers.append(
            {
                "x": f"{to_x(marker_date):.1f}",
                "label": f"{short_month_label} {str(marker_date.year)[2:]}",
                "date": marker_date.isoformat(),
                "y1": str(height - padding),
                "y2": str(height - padding + 6),
                "text_y": str(height - 2),
                "grid_y1": str(padding),
                "grid_y2": str(height - padding),
                "show_label": index % label_step == 0 or index == len(markers) - 1,
                "draw_grid": marker["draw_grid"],
                "is_major": marker_date.month == 1,
                "anchor": "start" if index == 0 else ("end" if index == len(markers) - 1 else "middle"),
            }
        )
    return serialized_markers


def build_dual_axis_chart(
    stock_points,
    reference_points,
    projection_points=None,
    width: int = 640,
    height: int = 220,
    padding: int = 18,
) -> dict:
    projection_points = projection_points or []

    def normalize_points(points):
        normalized = []
        for point in points:
            raw_date = point.get("date") if isinstance(point, dict) else None
            raw_value = point.get("value") if isinstance(point, dict) else None
            if raw_date is None or raw_value is None:
                continue
            normalized.append((raw_date, Decimal(str(raw_value))))
        return normalized

    stock_filtered = normalize_points(stock_points)
    reference_filtered = normalize_points(reference_points)
    projection_filtered = normalize_points(projection_points)
    if len(stock_filtered) < 2:
        return {
            "stock_line": "",
            "reference_line": "",
            "projection_line": "",
            "stock_min_label": "-",
            "stock_max_label": "-",
            "reference_min_label": "-",
            "reference_max_label": "-",
            "x_markers": [],
        }

    all_stock_values = [value for _, value in stock_filtered]
    if projection_filtered:
        all_stock_values.extend(value for _, value in projection_filtered)

    stock_min = min(all_stock_values)
    stock_max = max(all_stock_values)
    if stock_min == stock_max:
        stock_max += Decimal("1")

    if reference_filtered:
        reference_values_only = [value for _, value in reference_filtered]
        reference_min = min(reference_values_only)
        reference_max = max(reference_values_only)
        if reference_min == reference_max:
            reference_max += Decimal("1")
    else:
        reference_min = reference_max = None

    all_dates = [point_date for point_date, _ in stock_filtered]
    all_dates.extend(point_date for point_date, _ in reference_filtered)
    all_dates.extend(point_date for point_date, _ in projection_filtered)
    min_date = min(all_dates)
    max_date = max(all_dates)
    total_days = max((max_date - min_date).days, 1)

    def scale_series(points, series_min: Decimal | None, series_max: Decimal | None) -> str:
        if series_min is None or series_max is None:
            return ""
        span_x = width - 2 * padding
        span_y = height - 2 * padding
        line_points = []
        for point_date, value in points:
            x = padding + (span_x * ((point_date - min_date).days / total_days))
            normalized = (value - series_min) / (series_max - series_min)
            y = height - padding - (normalized * span_y)
            line_points.append(f"{x:.1f},{y:.1f}")
        return " ".join(line_points)

    return {
        "stock_line": scale_series(stock_filtered, stock_min, stock_max),
        "reference_line": scale_series(reference_filtered, reference_min, reference_max),
        "projection_line": scale_series(projection_filtered, stock_min, stock_max),
        "stock_min_label": format_axis_value(stock_min),
        "stock_max_label": format_axis_value(stock_max),
        "reference_min_label": format_axis_value(reference_min),
        "reference_max_label": format_axis_value(reference_max),
        "x_markers": build_time_axis_markers(all_dates, width=width, height=height, padding=padding),
    }


def build_stock_history_chart(history) -> dict:
    if len(history) < 2:
        return {"available": False}

    stock_series = [{"date": point.price_date, "value": point.close_price} for point in history]
    chart = build_dual_axis_chart(stock_series, [])
    return {
        "available": bool(chart.get("stock_line")),
        "stock_line": chart.get("stock_line", ""),
        "stock_min_label": chart.get("stock_min_label", "-"),
        "stock_max_label": chart.get("stock_max_label", "-"),
        "x_markers": chart.get("x_markers", []),
        "start_label": history[0].price_date.isoformat(),
        "end_label": history[-1].price_date.isoformat(),
        "points_count": len(stock_series),
    }


def build_same_axis_multi_line_chart(
    series_map: dict[str, list[dict]],
    width: int = 640,
    height: int = 220,
    padding: int = 18,
) -> dict:
    normalized_map: dict[str, list[tuple[date, Decimal]]] = {}
    all_values: list[Decimal] = []
    all_dates: list[date] = []
    for key, points in (series_map or {}).items():
        normalized_points = []
        for point in points or []:
            raw_date = point.get("date") if isinstance(point, dict) else None
            raw_value = point.get("value") if isinstance(point, dict) else None
            if raw_date is None or raw_value is None:
                continue
            normalized_points.append((raw_date, Decimal(str(raw_value))))
        if normalized_points:
            normalized_map[key] = normalized_points
            all_values.extend(value for _, value in normalized_points)
            all_dates.extend(point_date for point_date, _ in normalized_points)

    if len(all_values) < 2 or len(all_dates) < 2:
        return {
            "available": False,
            "min_label": "-",
            "max_label": "-",
            "x_markers": [],
            "start_label": "",
            "end_label": "",
            "points_count": 0,
            "lines": {},
        }

    series_min = min(all_values)
    series_max = max(all_values)
    if series_min == series_max:
        series_max += Decimal("1")

    min_date = min(all_dates)
    max_date = max(all_dates)
    total_days = max((max_date - min_date).days, 1)
    span_x = width - (padding * 2)
    span_y = height - (padding * 2)

    def scale_series(points: list[tuple[date, Decimal]]) -> str:
        line_points = []
        for point_date, value in points:
            x = padding + (span_x * ((point_date - min_date).days / total_days))
            normalized = (value - series_min) / (series_max - series_min)
            y = height - padding - (normalized * span_y)
            line_points.append(f"{x:.1f},{y:.1f}")
        return " ".join(line_points)

    return {
        "available": True,
        "min_label": format_axis_value(series_min),
        "max_label": format_axis_value(series_max),
        "x_markers": build_time_axis_markers(all_dates, width=width, height=height, padding=padding),
        "start_label": min_date.isoformat(),
        "end_label": max_date.isoformat(),
        "points_count": max(len(points) for points in normalized_map.values()),
        "lines": {key: scale_series(points) for key, points in normalized_map.items()},
    }


def build_reference_projection_detail_chart(
    historical_points: list[dict],
    raw_projection_points: list[dict],
    adjusted_projection_points: list[dict],
) -> dict:
    chart = build_same_axis_multi_line_chart(
        {
            "history": historical_points,
            "raw": raw_projection_points,
            "adjusted": adjusted_projection_points,
        }
    )
    return {
        "available": chart.get("available", False),
        "history_line": chart.get("lines", {}).get("history", ""),
        "raw_line": chart.get("lines", {}).get("raw", ""),
        "adjusted_line": chart.get("lines", {}).get("adjusted", ""),
        "min_label": chart.get("min_label", "-"),
        "max_label": chart.get("max_label", "-"),
        "x_markers": chart.get("x_markers", []),
        "start_label": chart.get("start_label", ""),
        "end_label": chart.get("end_label", ""),
        "points_count": chart.get("points_count", 0),
    }


def build_reference_coefficient_trend_chart(rolling_rows: list[dict], baseline: Decimal | None) -> dict:
    if len(rolling_rows or []) < 2:
        return {"available": False}

    coefficient_points = [
        {"date": row["end_date"], "value": row["coefficient"]}
        for row in rolling_rows
        if row.get("end_date") and row.get("coefficient") is not None
    ]
    baseline_points = []
    if baseline is not None and coefficient_points:
        baseline_points = [
            {"date": point["date"], "value": baseline}
            for point in coefficient_points
        ]
    chart = build_same_axis_multi_line_chart(
        {
            "coefficient": coefficient_points,
            "baseline": baseline_points,
        }
    )
    return {
        "available": chart.get("available", False),
        "coefficient_line": chart.get("lines", {}).get("coefficient", ""),
        "baseline_line": chart.get("lines", {}).get("baseline", ""),
        "min_label": chart.get("min_label", "-"),
        "max_label": chart.get("max_label", "-"),
        "x_markers": chart.get("x_markers", []),
        "start_label": chart.get("start_label", ""),
        "end_label": chart.get("end_label", ""),
        "points_count": chart.get("points_count", 0),
    }


def build_cycle_projection_comparison_chart(
    history,
    base_path: list[dict],
    factor_path: list[dict],
    final_path: list[dict],
) -> dict:
    if len(history) < 2:
        return {"available": False}

    latest_date = history[-1].price_date
    recent_history = filter_history_window(
        history,
        start_date=latest_date - timedelta(days=365 * 5),
        end_date=latest_date,
    )
    if len(recent_history) < 2:
        recent_history = history
    latest_history_price = recent_history[-1].close_price
    stock_series = [{"date": point.price_date, "value": point.close_price} for point in recent_history]
    base_series = [{"date": latest_date, "value": latest_history_price}]
    base_series.extend(
        {"date": step["projected_date"], "value": step["projected_price"]}
        for step in base_path or []
        if step.get("projected_date") and step.get("projected_price") is not None
    )
    factor_series = [{"date": latest_date, "value": latest_history_price}]
    factor_series.extend(
        {"date": step["projected_date"], "value": step["projected_price"]}
        for step in factor_path or []
        if step.get("projected_date") and step.get("projected_price") is not None
    )
    final_series = [{"date": latest_date, "value": latest_history_price}]
    final_series.extend(
        {"date": step["projected_date"], "value": step["projected_price"]}
        for step in final_path or []
        if step.get("projected_date") and step.get("projected_price") is not None
    )
    chart = build_same_axis_multi_line_chart(
        {
            "stock": stock_series,
            "base": base_series,
            "factor": factor_series,
            "final": final_series,
        }
    )
    return {
        "available": chart.get("available", False),
        "stock_line": chart.get("lines", {}).get("stock", ""),
        "base_line": chart.get("lines", {}).get("base", ""),
        "factor_line": chart.get("lines", {}).get("factor", ""),
        "final_line": chart.get("lines", {}).get("final", ""),
        "min_label": chart.get("min_label", "-"),
        "max_label": chart.get("max_label", "-"),
        "x_markers": chart.get("x_markers", []),
        "start_label": chart.get("start_label", ""),
        "end_label": chart.get("end_label", ""),
        "points_count": chart.get("points_count", 0),
    }


def build_best_correlation_chart(history, suggestions: list[dict], reference_cache: dict) -> dict:
    best_suggestion = next(
        (
            suggestion
            for suggestion in suggestions
            if suggestion.get("correlation", {}).get("coefficient") is not None
        ),
        None,
    )
    if best_suggestion is None or len(history) < 2:
        return {"available": False}

    cache_key = (
        best_suggestion["reference_profile"],
        best_suggestion["benchmark_symbol"],
        best_suggestion["benchmark_name"],
    )
    cached_value = reference_cache.get(cache_key)
    if cached_value is None:
        try:
            cached_value = fetch_reference_series_for_choice(
                best_suggestion["reference_profile"],
                benchmark_symbol=best_suggestion["benchmark_symbol"],
                benchmark_name=best_suggestion["benchmark_name"],
            )
            reference_cache[cache_key] = cached_value
        except Exception as exc:
            reference_cache[cache_key] = exc
            cached_value = exc
    if isinstance(cached_value, Exception):
        return {"available": False}

    stock_series = [{"date": point.price_date, "value": point.close_price} for point in history]
    aligned_reference = align_reference_points(
        [{"date": point.price_date} for point in history],
        cached_value.points,
    )
    reference_series = [
        {"date": point.price_date, "value": aligned_reference.get(point.price_date)}
        for point in history
        if aligned_reference.get(point.price_date) is not None
    ]
    chart = build_dual_axis_chart(stock_series, reference_series)
    coefficient = best_suggestion["correlation"].get("coefficient")
    return {
        "available": bool(chart.get("stock_line") and chart.get("reference_line")),
        "stock_line": chart.get("stock_line", ""),
        "reference_line": chart.get("reference_line", ""),
        "stock_min_label": chart.get("stock_min_label", "-"),
        "stock_max_label": chart.get("stock_max_label", "-"),
        "reference_min_label": chart.get("reference_min_label", "-"),
        "reference_max_label": chart.get("reference_max_label", "-"),
        "reference_label": best_suggestion["benchmark_name"],
        "coefficient": coefficient,
        "x_markers": chart.get("x_markers", []),
        "start_label": history[0].price_date.isoformat(),
        "end_label": history[-1].price_date.isoformat(),
        "points_count": len(stock_series),
        "is_selected_reference": best_suggestion.get("is_selected", False),
    }


def resolve_projection_tracking_path(projection: dict | None) -> list[dict]:
    projection = projection or {}
    return projection.get("monthly_path") or projection.get("quarterly_path") or []


def build_projection_12m_chart(history, projection: dict) -> dict:
    if len(history) < 2 or not projection.get("available"):
        return {"available": False}

    latest_date = history[-1].price_date
    recent_history = filter_history_window(
        history,
        start_date=latest_date - timedelta(days=365),
        end_date=latest_date,
    )
    if len(recent_history) < 2:
        recent_history = history
    projection_path = resolve_projection_tracking_path(projection)
    stock_series = [{"date": point.price_date, "value": point.close_price} for point in recent_history]
    projection_series = [{"date": latest_date, "value": recent_history[-1].close_price}]
    projection_series.extend(
        {
            "date": step["projected_date"],
            "value": step["projected_price"],
        }
        for step in projection_path
        if step.get("projected_date") and step.get("projected_price") is not None
    )
    chart = build_dual_axis_chart(stock_series, [], projection_points=projection_series)
    end_projection_step = projection_path[-1] if projection_path else None
    return {
        "available": bool(chart.get("stock_line") and chart.get("projection_line")),
        "stock_line": chart.get("stock_line", ""),
        "projection_line": chart.get("projection_line", ""),
        "stock_min_label": chart.get("stock_min_label", "-"),
        "stock_max_label": chart.get("stock_max_label", "-"),
        "x_markers": chart.get("x_markers", []),
        "start_label": recent_history[0].price_date.isoformat(),
        "end_label": latest_date.isoformat(),
        "projection_end_label": end_projection_step.get("projected_date").isoformat() if end_projection_step and end_projection_step.get("projected_date") else "",
        "points_count": len(stock_series),
        "history_window_label": projection.get("history_window_label", "Ultimo ano visible"),
        "projection_window_label": projection.get("path_source_label", "Escenario central 12M"),
        "model_window_label": projection.get("path_model_window_label", ""),
    }


def build_projection_presentation_summary(card: dict) -> dict:
    position = card.get("position")
    projection = card.get("projection") or {}
    technical_signal = card.get("technical_signal") or {}
    if position is None or not projection.get("available"):
        return {"available": False}

    monthly_rows = build_projection_monthly_trend_rows(card)
    current_price = quantize_decimal(
        projection.get("latest_price") or getattr(position, "current_price_per_share", None),
        "0.0001",
    )
    anchor_date = projection.get("latest_date") or getattr(position, "latest_price_date", None)
    fallback_end_price = quantize_decimal(projection.get("projected_price"), "0.0001")
    if current_price in {None, ZERO}:
        return {"available": False}

    if not monthly_rows and fallback_end_price not in {None, ZERO}:
        monthly_rows = [
            {
                "month_number": 0,
                "projected_date": anchor_date,
                "projected_price": current_price,
            },
            {
                "month_number": 12,
                "projected_date": add_calendar_months(anchor_date, 12) if anchor_date else None,
                "projected_price": fallback_end_price,
            },
        ]
    if len(monthly_rows) < 2:
        return {"available": False}

    current_row = monthly_rows[0]
    future_rows = [row for row in monthly_rows[1:] if row.get("projected_price") not in {None, ZERO}]
    if not future_rows:
        return {"available": False}

    end_row = future_rows[-1]
    min_row = min(future_rows, key=lambda row: (row.get("projected_price") or ZERO, row.get("month_number") or 999))
    max_row = max(future_rows, key=lambda row: (row.get("projected_price") or ZERO, -(row.get("month_number") or 999)))
    current_price = quantize_decimal(current_row.get("projected_price"), "0.0001") or current_price
    end_price = quantize_decimal(end_row.get("projected_price"), "0.0001")
    min_price = quantize_decimal(min_row.get("projected_price"), "0.0001")
    max_price = quantize_decimal(max_row.get("projected_price"), "0.0001")
    if end_price in {None, ZERO}:
        return {"available": False}

    visible_price_return_pct = quantize_decimal(percentage_change(end_price, current_price), "0.01")
    net_income_yield_pct = quantize_decimal(projection.get("net_income_yield_pct"), "0.01") or ZERO
    transaction_drag_pct = quantize_decimal(projection.get("transaction_drag_pct"), "0.01") or ZERO
    visible_total_return_pct = quantize_decimal(
        (visible_price_return_pct or ZERO) + net_income_yield_pct - transaction_drag_pct,
        "0.01",
    )
    raw_base_return_pct = quantize_decimal(projection.get("base_return_pct"), "0.01")
    base_shift_pct = (
        quantize_decimal(visible_total_return_pct - raw_base_return_pct, "0.01")
        if raw_base_return_pct is not None and visible_total_return_pct is not None
        else ZERO
    )
    scenario_low_return_pct = quantize_decimal(projection.get("low_return_pct"), "0.01")
    if scenario_low_return_pct is not None and raw_base_return_pct is not None and visible_total_return_pct is not None:
        scenario_low_return_pct = quantize_decimal(scenario_low_return_pct + base_shift_pct, "0.01")
    if scenario_low_return_pct is None:
        scenario_low_return_pct = visible_total_return_pct
    scenario_high_return_pct = quantize_decimal(projection.get("high_return_pct"), "0.01")
    if scenario_high_return_pct is not None and raw_base_return_pct is not None and visible_total_return_pct is not None:
        scenario_high_return_pct = quantize_decimal(scenario_high_return_pct + base_shift_pct, "0.01")
    if scenario_high_return_pct is None:
        scenario_high_return_pct = visible_total_return_pct
    scenario_low_price = derive_projected_price_from_total_return(
        current_price,
        scenario_low_return_pct,
        net_income_yield_pct=net_income_yield_pct,
        transaction_drag_pct=transaction_drag_pct,
    ) or quantize_decimal(projection.get("low_price"), "0.0001")
    scenario_high_price = derive_projected_price_from_total_return(
        current_price,
        scenario_high_return_pct,
        net_income_yield_pct=net_income_yield_pct,
        transaction_drag_pct=transaction_drag_pct,
    ) or quantize_decimal(projection.get("high_price"), "0.0001")

    future_prices = [quantize_decimal(row.get("projected_price"), "0.0001") or ZERO for row in future_rows]
    is_non_decreasing = all(current >= previous for previous, current in zip(future_prices, future_prices[1:]))
    is_non_increasing = all(current <= previous for previous, current in zip(future_prices, future_prices[1:]))
    future_min_below_current = min_price not in {None, ZERO} and min_price < current_price
    future_max_above_current = max_price not in {None, ZERO} and max_price > current_price

    if future_min_below_current and min_row.get("month_number", 0) < end_row.get("month_number", 0) and end_price > min_price:
        shape_label = "Corrige y recupera"
        shape_note = (
            f"La senda visible cae primero hasta {min_price:.4f} en "
            f"{format_projection_month_window(min_row.get('projected_date'), min_row.get('month_number'))} "
            f"y luego rebota para cerrar 12M en {end_price:.4f}."
        )
        key_level_label = "Suelo probable"
        key_level_row = min_row
    elif future_max_above_current and max_row.get("month_number", 0) < end_row.get("month_number", 0) and end_price < max_price:
        shape_label = "Sube y corrige"
        shape_note = (
            f"La senda visible marca antes un techo en {max_price:.4f} en "
            f"{format_projection_month_window(max_row.get('projected_date'), max_row.get('month_number'))} "
            f"y despues se enfria hasta cerrar 12M en {end_price:.4f}."
        )
        key_level_label = "Techo intermedio"
        key_level_row = max_row
    elif is_non_decreasing and (visible_price_return_pct or ZERO) > ZERO:
        shape_label = "Subida progresiva"
        shape_note = (
            f"La senda visible mantiene una pendiente alcista bastante continua y cierra 12M "
            f"en {end_price:.4f}."
        )
        key_level_label = "Mejor nivel visible"
        key_level_row = max_row
    elif is_non_increasing and (visible_price_return_pct or ZERO) < ZERO:
        shape_label = "Caida progresiva"
        shape_note = (
            f"La senda visible pierde nivel de forma bastante continua y termina 12M "
            f"en {end_price:.4f}."
        )
        key_level_label = "Peor nivel visible"
        key_level_row = min_row
    elif abs(visible_price_return_pct or ZERO) <= Decimal("2.50"):
        shape_label = "Lateral"
        shape_note = (
            f"La senda visible oscila sin una direccion limpia y cierra 12M muy cerca de hoy, "
            f"en {end_price:.4f}."
        )
        key_level_label = "Tramo clave"
        key_level_row = min_row if future_min_below_current else max_row
    else:
        bias_label = "alcista" if (visible_price_return_pct or ZERO) > ZERO else "bajista"
        shape_label = f"Volatil con sesgo {bias_label}"
        shape_note = (
            f"La senda visible se mueve con idas y vueltas, pero el cierre 12M queda sesgado "
            f"a {end_price:.4f}."
        )
        key_level_label = "Tramo clave"
        key_level_row = min_row if future_min_below_current else max_row

    key_level_price = quantize_decimal(key_level_row.get("projected_price"), "0.0001") if key_level_row else None
    key_level_return_pct = (
        quantize_decimal(percentage_change(key_level_price, current_price), "0.01")
        if key_level_price not in {None, ZERO}
        else None
    )

    support_level = quantize_decimal(technical_signal.get("support_level"), "0.0001")
    resistance_level = quantize_decimal(technical_signal.get("resistance_level"), "0.0001")
    support_gap_pct = (
        quantize_decimal(percentage_change(support_level, current_price), "0.01")
        if support_level not in {None, ZERO}
        else None
    )
    resistance_gap_pct = (
        quantize_decimal(percentage_change(resistance_level, current_price), "0.01")
        if resistance_level not in {None, ZERO}
        else None
    )
    if support_level not in {None, ZERO} and resistance_level not in {None, ZERO}:
        decision_level_note = (
            f"Perder {support_level:.4f} enfriaria la tesis y romper {resistance_level:.4f} la mejoraria."
        )
    elif support_level not in {None, ZERO}:
        decision_level_note = f"Perder {support_level:.4f} seria la primera senal clara de deterioro."
    elif resistance_level not in {None, ZERO}:
        decision_level_note = f"Romper {resistance_level:.4f} ayudaria a confirmar una mejora real."
    else:
        decision_level_note = technical_signal.get("note") or "Sin niveles tecnicos claros todavia."

    average_cost = quantize_decimal(getattr(position, "average_cost_per_share", None), "0.0001")
    current_vs_cost_pct = (
        quantize_decimal(percentage_change(current_price, average_cost), "0.01")
        if position.is_owned and average_cost not in {None, ZERO}
        else None
    )
    visible_vs_cost_pct = (
        quantize_decimal(percentage_change(end_price, average_cost), "0.01")
        if position.is_owned and average_cost not in {None, ZERO}
        else None
    )
    if position.is_owned and average_cost not in {None, ZERO}:
        owned_position_note = (
            f"Hoy va {current_vs_cost_pct:.1f} % frente a tu compra y el cierre 12M visible quedaria "
            f"en {visible_vs_cost_pct:.1f} % sobre tu precio medio."
        )
    else:
        owned_position_note = (
            "La lectura 12M visible usa el precio actual como base y no una recta separada del grafico."
        )

    raw_projected_price = quantize_decimal(projection.get("projected_price"), "0.0001")
    raw_gap_pct = (
        quantize_decimal(percentage_change(end_price, raw_projected_price), "0.01")
        if raw_projected_price not in {None, ZERO}
        else None
    )
    consistency_note = "La rentabilidad y el cierre 12M salen de la misma senda naranja que ves en el grafico."
    if projection.get("uses_cycle_zoom_shape"):
        consistency_note += " El 12M se ancla al primer ano del patron 5A para que ambos horizontes cuenten la misma historia."

    return {
        "available": True,
        "shape_label": shape_label,
        "shape_note": shape_note,
        "consistency_note": consistency_note,
        "current_price": current_price,
        "current_date": current_row.get("projected_date") or anchor_date,
        "visible_projected_price": end_price,
        "visible_projected_date": end_row.get("projected_date"),
        "visible_projected_window_label": format_projection_month_window(
            end_row.get("projected_date"),
            end_row.get("month_number"),
        ),
        "visible_price_return_pct": visible_price_return_pct,
        "visible_total_return_pct": visible_total_return_pct,
        "scenario_low_return_pct": scenario_low_return_pct,
        "scenario_high_return_pct": scenario_high_return_pct,
        "scenario_low_price": scenario_low_price,
        "scenario_high_price": scenario_high_price,
        "key_level_label": key_level_label,
        "key_level_price": key_level_price,
        "key_level_date": key_level_row.get("projected_date") if key_level_row else None,
        "key_level_window_label": format_projection_month_window(
            key_level_row.get("projected_date") if key_level_row else None,
            key_level_row.get("month_number") if key_level_row else None,
        ),
        "key_level_return_pct": key_level_return_pct,
        "support_level": support_level,
        "resistance_level": resistance_level,
        "support_gap_pct": support_gap_pct,
        "resistance_gap_pct": resistance_gap_pct,
        "decision_level_note": decision_level_note,
        "owned_position_note": owned_position_note,
        "current_vs_cost_pct": current_vs_cost_pct,
        "visible_vs_cost_pct": visible_vs_cost_pct,
        "raw_projected_price": raw_projected_price,
        "raw_gap_pct": raw_gap_pct,
    }


def get_card_presentation_projection(card: dict) -> dict:
    presentation_projection = card.get("presentation_projection") or {}
    if presentation_projection.get("available"):
        return presentation_projection
    position = card.get("position")
    projection = card.get("projection") or {}
    if position is None or not projection.get("available"):
        return {"available": False}
    presentation_projection = build_projection_presentation_summary(card)
    card["presentation_projection"] = presentation_projection
    return presentation_projection


def derive_projected_price_from_total_return(
    current_price: Decimal | None,
    total_return_pct: Decimal | None,
    *,
    net_income_yield_pct: Decimal | None = None,
    transaction_drag_pct: Decimal | None = None,
) -> Decimal | None:
    current_price = quantize_decimal(current_price, "0.0001")
    total_return_pct = quantize_decimal(total_return_pct, "0.01")
    net_income_yield_pct = quantize_decimal(net_income_yield_pct, "0.01") or ZERO
    transaction_drag_pct = quantize_decimal(transaction_drag_pct, "0.01") or ZERO
    if current_price in {None, ZERO} or total_return_pct is None:
        return None
    price_return_pct = quantize_decimal(
        total_return_pct - net_income_yield_pct + transaction_drag_pct,
        "0.01",
    )
    return quantize_decimal(
        current_price * (Decimal("1") + ((price_return_pct or ZERO) / ONE_HUNDRED)),
        "0.0001",
    )


def resolve_effective_projection_metrics(card: dict) -> dict:
    position = card.get("position")
    projection = card.get("projection") or {}
    if position is None or not projection.get("available"):
        return {"available": False}

    net_income_yield_pct = quantize_decimal(projection.get("net_income_yield_pct"), "0.01") or ZERO
    transaction_drag_pct = quantize_decimal(projection.get("transaction_drag_pct"), "0.01") or ZERO
    current_price = quantize_decimal(
        projection.get("latest_price") or getattr(position, "current_price_per_share", None),
        "0.0001",
    )
    raw_base_return_pct = quantize_decimal(projection.get("base_return_pct"), "0.01")
    raw_price_return_pct = quantize_decimal(projection.get("price_return_pct"), "0.01")
    if raw_price_return_pct is None and raw_base_return_pct is not None:
        raw_price_return_pct = quantize_decimal(
            raw_base_return_pct - net_income_yield_pct + transaction_drag_pct,
            "0.01",
        )

    presentation_projection = get_card_presentation_projection(card)
    visible_total_return_pct = (
        quantize_decimal(presentation_projection.get("visible_total_return_pct"), "0.01")
        if presentation_projection.get("available")
        else None
    )
    visible_price_return_pct = (
        quantize_decimal(presentation_projection.get("visible_price_return_pct"), "0.01")
        if presentation_projection.get("available")
        else None
    )
    effective_base_return_pct = visible_total_return_pct if visible_total_return_pct is not None else raw_base_return_pct
    effective_price_return_pct = visible_price_return_pct if visible_price_return_pct is not None else raw_price_return_pct
    base_shift_pct = (
        quantize_decimal(effective_base_return_pct - raw_base_return_pct, "0.01")
        if effective_base_return_pct is not None and raw_base_return_pct is not None
        else ZERO
    )

    raw_low_return_pct = quantize_decimal(projection.get("low_return_pct"), "0.01")
    raw_high_return_pct = quantize_decimal(projection.get("high_return_pct"), "0.01")
    raw_price_low_return_pct = quantize_decimal(projection.get("price_low_return_pct"), "0.01")
    raw_price_high_return_pct = quantize_decimal(projection.get("price_high_return_pct"), "0.01")
    if raw_low_return_pct is None and raw_price_low_return_pct is not None:
        raw_low_return_pct = quantize_decimal(
            raw_price_low_return_pct + net_income_yield_pct - transaction_drag_pct,
            "0.01",
        )
    if raw_high_return_pct is None and raw_price_high_return_pct is not None:
        raw_high_return_pct = quantize_decimal(
            raw_price_high_return_pct + net_income_yield_pct - transaction_drag_pct,
            "0.01",
        )
    effective_low_return_pct = (
        quantize_decimal(raw_low_return_pct + base_shift_pct, "0.01")
        if raw_low_return_pct is not None and raw_base_return_pct is not None and effective_base_return_pct is not None
        else raw_low_return_pct
    )
    effective_high_return_pct = (
        quantize_decimal(raw_high_return_pct + base_shift_pct, "0.01")
        if raw_high_return_pct is not None and raw_base_return_pct is not None and effective_base_return_pct is not None
        else raw_high_return_pct
    )
    if effective_low_return_pct is None:
        effective_low_return_pct = effective_base_return_pct
    if effective_high_return_pct is None:
        effective_high_return_pct = effective_base_return_pct

    effective_projected_price = (
        quantize_decimal(presentation_projection.get("visible_projected_price"), "0.0001")
        if presentation_projection.get("available")
        else None
    )
    if effective_projected_price is None:
        effective_projected_price = quantize_decimal(projection.get("projected_price"), "0.0001")
    if effective_projected_price is None and current_price not in {None, ZERO} and effective_price_return_pct is not None:
        effective_projected_price = quantize_decimal(
            current_price * (Decimal("1") + (effective_price_return_pct / ONE_HUNDRED)),
            "0.0001",
        )

    price_low_return_pct = (
        quantize_decimal(effective_low_return_pct - net_income_yield_pct + transaction_drag_pct, "0.01")
        if effective_low_return_pct is not None
        else quantize_decimal(projection.get("price_low_return_pct"), "0.01")
    )
    price_high_return_pct = (
        quantize_decimal(effective_high_return_pct - net_income_yield_pct + transaction_drag_pct, "0.01")
        if effective_high_return_pct is not None
        else quantize_decimal(projection.get("price_high_return_pct"), "0.01")
    )

    return {
        "available": True,
        "current_price": current_price,
        "net_income_yield_pct": net_income_yield_pct,
        "transaction_drag_pct": transaction_drag_pct,
        "raw_base_return_pct": raw_base_return_pct,
        "base_shift_pct": base_shift_pct,
        "base_return_pct": effective_base_return_pct,
        "price_return_pct": effective_price_return_pct,
        "projected_price": effective_projected_price,
        "low_return_pct": effective_low_return_pct,
        "high_return_pct": effective_high_return_pct,
        "price_low_return_pct": price_low_return_pct,
        "price_high_return_pct": price_high_return_pct,
        "low_price": derive_projected_price_from_total_return(
            current_price,
            effective_low_return_pct,
            net_income_yield_pct=net_income_yield_pct,
            transaction_drag_pct=transaction_drag_pct,
        ) or quantize_decimal(projection.get("low_price"), "0.0001"),
        "high_price": derive_projected_price_from_total_return(
            current_price,
            effective_high_return_pct,
            net_income_yield_pct=net_income_yield_pct,
            transaction_drag_pct=transaction_drag_pct,
        ) or quantize_decimal(projection.get("high_price"), "0.0001"),
    }


def build_effective_projection_scenarios(
    card: dict,
    *,
    effective_metrics: dict | None = None,
) -> list[dict]:
    projection = card.get("projection") or {}
    effective_metrics = effective_metrics or resolve_effective_projection_metrics(card)
    scenarios = []
    base_shift_pct = quantize_decimal(effective_metrics.get("base_shift_pct"), "0.01") or ZERO
    current_price = quantize_decimal(effective_metrics.get("current_price"), "0.0001")
    net_income_yield_pct = quantize_decimal(effective_metrics.get("net_income_yield_pct"), "0.01") or ZERO
    transaction_drag_pct = quantize_decimal(effective_metrics.get("transaction_drag_pct"), "0.01") or ZERO

    for row in list(projection.get("scenarios") or []):
        adjusted_row = dict(row)
        total_return_pct = row.get("total_return_pct")
        adjusted_total_return_pct = (
            quantize_decimal(Decimal(str(total_return_pct)) + base_shift_pct, "0.01")
            if total_return_pct is not None
            else None
        )
        if adjusted_total_return_pct is not None:
            adjusted_row["total_return_pct"] = adjusted_total_return_pct
            adjusted_projected_price = derive_projected_price_from_total_return(
                current_price,
                adjusted_total_return_pct,
                net_income_yield_pct=net_income_yield_pct,
                transaction_drag_pct=transaction_drag_pct,
            )
            if adjusted_projected_price is not None:
                adjusted_row["projected_price"] = adjusted_projected_price
        scenarios.append(adjusted_row)
    return scenarios


INFORMATION_THEME_LABELS = {
    "geopolitica": "Geopolitica",
    "tipos": "Tipos e inflacion",
    "energia": "Energia",
    "vivienda": "Vivienda",
    "regulacion": "Regulacion",
    "resultados": "Resultados",
    "contratos": "Contratos",
    "empresa": "Empresa",
    "mercado": "Mercado",
    "wall_street": "Wall Street",
    "bridgewater": "Bridgewater",
    "experto": "Consenso experto",
}


INFORMATION_THEME_PRIORITY = {
    "geopolitica": 9,
    "tipos": 8,
    "energia": 7,
    "regulacion": 7,
    "vivienda": 6,
    "wall_street": 6,
    "bridgewater": 6,
    "resultados": 5,
    "contratos": 5,
    "empresa": 4,
    "experto": 4,
    "mercado": 3,
}


INFORMATION_THEME_NOTES = {
    "geopolitica": "Puede mover timing y rango del escenario.",
    "tipos": "Afecta descuento, financiacion y valoracion.",
    "energia": "Afecta costes y sensibilidad macro.",
    "vivienda": "Afecta demanda y ciclo domestico.",
    "regulacion": "Puede cambiar visibilidad o margenes.",
    "resultados": "Afecta ejecucion y confianza en la tesis.",
    "contratos": "Afecta cartera futura e ingresos esperados.",
    "empresa": "Impacta la tesis propia de la compania.",
    "mercado": "Mueve apetito por riesgo y beta del valor.",
    "wall_street": "Funciona como termometro global de riesgo.",
    "bridgewater": "Anade lectura macro de fondo.",
    "experto": "Refuerza o cuestiona la tesis con opinion externa.",
}


def resolve_information_theme_key(tags: list[str] | None, *, fallback_scope: str = "") -> str:
    normalized_tags = [str(tag or "").strip().lower() for tag in (tags or []) if str(tag or "").strip()]
    for candidate in ("geopolitica", "tipos", "energia", "vivienda", "regulacion", "resultados", "contratos"):
        if candidate in normalized_tags:
            return candidate
    normalized_scope = str(fallback_scope or "").strip().lower()
    if normalized_scope in {"company", "empresa"}:
        return "empresa"
    if normalized_scope in {"market", "mercado"}:
        return "mercado"
    if normalized_scope in {"wall_street", "wall street"}:
        return "wall_street"
    if normalized_scope == "bridgewater":
        return "bridgewater"
    if normalized_scope in {"expert", "experto", "expert_consensus"}:
        return "experto"
    return "mercado"


def build_information_source_label(item: dict) -> str:
    expert_source = str(item.get("expert_source") or "").strip()
    source = str(item.get("source") or "").strip()
    if expert_source and source and source.lower() != expert_source.lower():
        return f"{expert_source} via {source}"
    return expert_source or source or "Fuente agregada"


def resolve_information_tone_label(tone: str | None, score: Decimal | None = None) -> str:
    normalized_tone = str(tone or "").strip().lower()
    numeric_score = Decimal(str(score or ZERO))
    if normalized_tone == "positive" or numeric_score > ZERO:
        return "Apoya"
    if normalized_tone == "negative" or numeric_score < ZERO:
        return "Presiona"
    return "Mixto"


def build_information_evidence_row(
    *,
    title: str,
    source_label: str,
    published_label: str = "",
    tags: list[str] | None = None,
    tone: str | None = None,
    score: Decimal | None = None,
    scope_label: str = "",
    fallback_scope: str = "",
) -> dict | None:
    cleaned_title = " ".join(str(title or "").strip().split())
    if not cleaned_title:
        return None
    theme_key = resolve_information_theme_key(tags, fallback_scope=fallback_scope)
    theme_label = INFORMATION_THEME_LABELS.get(theme_key, "Mercado")
    tone_label = resolve_information_tone_label(tone, score)
    base_note = INFORMATION_THEME_NOTES.get(theme_key, "Aporta contexto adicional a la tesis.")
    numeric_score = Decimal(str(score or ZERO))
    impact_note = (
        f"{tone_label} la prevision. {base_note}"
        if tone_label != "Mixto"
        else f"Introduce una lectura menos limpia. {base_note}"
    )
    return {
        "theme_key": theme_key,
        "theme_label": theme_label,
        "tone_label": tone_label,
        "title": cleaned_title,
        "source_label": source_label or "Fuente agregada",
        "published_label": str(published_label or "").strip(),
        "scope_label": scope_label or "",
        "impact_note": impact_note,
        "tags": list(tags or [])[:4],
        "priority": INFORMATION_THEME_PRIORITY.get(theme_key, 1),
        "score": numeric_score,
        "score_abs": abs(numeric_score),
    }


def build_information_basis_summary(card: dict) -> dict:
    news_context = card.get("news_context") or {}
    expert_consensus = card.get("expert_consensus") or {}
    rows = []
    seen_keys = set()

    def append_row(row: dict | None):
        if not row:
            return
        row_key = (
            str(row.get("title") or "").strip().lower(),
            str(row.get("source_label") or "").strip().lower(),
            str(row.get("theme_key") or "").strip().lower(),
        )
        if row_key in seen_keys:
            return
        seen_keys.add(row_key)
        rows.append(row)

    for item in list(news_context.get("top_items") or [])[:4]:
        append_row(
            build_information_evidence_row(
                title=item.get("title") or news_context.get("material_note") or news_context.get("note"),
                source_label=build_information_source_label(item),
                published_label=item.get("published_label") or "",
                tags=item.get("tags") or [],
                tone=item.get("tone"),
                score=Decimal(str(item.get("score") or ZERO)),
                scope_label="Prensa reciente",
                fallback_scope="market",
            )
        )

    for item in list(expert_consensus.get("top_items") or [])[:4]:
        append_row(
            build_information_evidence_row(
                title=item.get("title") or expert_consensus.get("note"),
                source_label=build_information_source_label(item),
                published_label=item.get("published_label") or "",
                tags=item.get("tags") or [],
                tone=item.get("tone"),
                score=Decimal(str(item.get("score") or ZERO)),
                scope_label="Prevision externa",
                fallback_scope="expert",
            )
        )

    for signal_key, signal_label in (
        ("wall_street_signal", "Wall Street"),
        ("bridgewater_signal", "Bridgewater"),
    ):
        signal = expert_consensus.get(signal_key) or {}
        if not signal.get("available"):
            continue
        first_source_row = (signal.get("source_rows") or [{}])[0]
        append_row(
            build_information_evidence_row(
                title=signal.get("note") or signal.get("label"),
                source_label=str(first_source_row.get("source") or signal_label),
                published_label=(signal.get("items") or [{}])[0].get("published_label") or expert_consensus.get("captured_at_label") or "",
                tags=signal.get("top_tags") or [],
                tone="positive" if Decimal(str(signal.get("score") or ZERO)) > ZERO else "negative" if Decimal(str(signal.get("score") or ZERO)) < ZERO else "neutral",
                score=Decimal(str(signal.get("score") or ZERO)),
                scope_label=signal_label,
                fallback_scope=signal_key,
            )
        )

    rows = sorted(
        rows,
        key=lambda row: (
            -int(row.get("theme_key") == "geopolitica"),
            -int(row.get("theme_key") in {"tipos", "energia", "vivienda", "regulacion", "mercado", "wall_street", "bridgewater"}),
            -int(row.get("priority") or 0),
            -Decimal(str(row.get("score_abs") or ZERO)),
            str(row.get("source_label") or "").strip().lower(),
        ),
    )[:6]
    if not rows:
        return {"available": False, "rows": []}

    theme_labels = []
    for row in rows:
        theme_label = row.get("theme_label")
        if theme_label and theme_label not in theme_labels:
            theme_labels.append(theme_label)
    source_labels = []
    for row in rows:
        source_label = row.get("source_label")
        if source_label and source_label not in source_labels:
            source_labels.append(source_label)

    pressure_rows = [row for row in rows if row.get("tone_label") == "Presiona"]
    support_rows = [row for row in rows if row.get("tone_label") == "Apoya"]
    geopolitical_flag = any(row.get("theme_key") == "geopolitica" for row in rows)
    macro_flag = any(
        row.get("theme_key") in {"tipos", "energia", "vivienda", "regulacion", "mercado", "wall_street", "bridgewater"}
        for row in rows
    )

    def compact_row_label(row: dict) -> str:
        theme_label = str(row.get("theme_label") or "Mercado").strip()
        source_label = str(row.get("source_label") or "Fuente agregada").strip()
        return f"{theme_label} ({source_label})"

    def build_bullet_point(row: dict) -> str:
        theme_label = str(row.get("theme_label") or "Mercado").strip()
        tone_label = str(row.get("tone_label") or "Mixto").strip().lower()
        source_label = str(row.get("source_label") or "Fuente agregada").strip()
        title = str(row.get("title") or "").strip()
        published_label = str(row.get("published_label") or "").strip()
        point = f"{theme_label}: {tone_label} segun {source_label}"
        if published_label:
            point += f" ({published_label})"
        if title:
            point += f". {title}"
        return point

    geopolitical_sources = []
    macro_sources = []
    for row in rows:
        source_label = row.get("source_label")
        if row.get("theme_key") == "geopolitica" and source_label and source_label not in geopolitical_sources:
            geopolitical_sources.append(source_label)
        if row.get("theme_key") in {"tipos", "energia", "vivienda", "regulacion", "mercado", "wall_street", "bridgewater"}:
            if source_label and source_label not in macro_sources:
                macro_sources.append(source_label)

    summary_bits = [f"La prevision usa {len(rows)} evidencias informativas recientes."]
    if pressure_rows:
        summary_bits.append(f"Hoy presionan {', '.join(compact_row_label(row) for row in pressure_rows[:2])}.")
    if support_rows:
        summary_bits.append(f"Apoyan {', '.join(compact_row_label(row) for row in support_rows[:2])}.")
    elif theme_labels:
        summary_bits.append(f"Hoy pesan sobre todo {', '.join(theme_labels[:3]).lower()}.")
    if geopolitical_flag:
        summary_bits.append("Hay senal geopolitica que puede cambiar el timing.")
    if macro_flag:
        summary_bits.append("Tambien pesa el bloque macro y de mercado.")
    if source_labels:
        summary_bits.append(f"Fuentes clave: {', '.join(source_labels[:3])}.")

    return {
        "available": True,
        "summary": " ".join(summary_bits),
        "rows": rows,
        "bullet_points": [build_bullet_point(row) for row in rows[:3]],
        "theme_labels": theme_labels,
        "source_labels": source_labels,
        "geopolitical_flag": geopolitical_flag,
        "macro_flag": macro_flag,
        "geopolitical_sources": geopolitical_sources,
        "macro_sources": macro_sources,
    }


def build_scenario_expectation_table(
    scenarios: list[dict] | None,
    *,
    return_key: str,
    fallback_value: Decimal | None = None,
    projected_price_key: str = "projected_price",
    secondary_return_key: str | None = None,
) -> dict:
    normalized_rows = []
    for row in list(scenarios or []):
        return_pct = row.get(return_key)
        if return_pct is None:
            continue
        normalized_rows.append(
            {
                "key": str(row.get("key") or "").strip().lower(),
                "label": str(row.get("label") or "").strip() or "Escenario",
                "return_pct": Decimal(str(return_pct)),
                "raw_probability_pct": Decimal(str(row.get("probability_pct") or "0")),
                "projected_price": quantize_decimal(row.get(projected_price_key), "0.0001"),
                "secondary_return_pct": (
                    quantize_decimal(row.get(secondary_return_key), "0.01")
                    if secondary_return_key
                    else None
                ),
            }
        )

    if not normalized_rows:
        return {
            "available": False,
            "rows": [],
            "expected_return_pct": fallback_value,
            "expected_projected_price": None,
            "expected_secondary_return_pct": None,
        }

    total_probability_pct = sum((item["raw_probability_pct"] for item in normalized_rows), ZERO)
    if total_probability_pct <= ZERO:
        total_probability_pct = ONE_HUNDRED

    summary = build_optimizer_scenario_summary(
        scenarios,
        return_key=return_key,
        fallback_value=fallback_value,
    )
    rows = []
    expected_projected_price = ZERO
    has_projected_price = False
    expected_secondary_return_pct = ZERO
    has_secondary_return = False
    for item in normalized_rows:
        effective_probability_pct = quantize_decimal(
            (item["raw_probability_pct"] * ONE_HUNDRED) / total_probability_pct,
            "0.1",
        ) or ZERO
        contribution_return_pct = quantize_decimal(
            (item["return_pct"] * item["raw_probability_pct"]) / total_probability_pct,
            "0.01",
        )
        projected_price = item["projected_price"]
        if projected_price is not None:
            expected_projected_price += (projected_price * item["raw_probability_pct"]) / total_probability_pct
            has_projected_price = True
        secondary_return_pct = item["secondary_return_pct"]
        if secondary_return_pct is not None:
            expected_secondary_return_pct += (secondary_return_pct * item["raw_probability_pct"]) / total_probability_pct
            has_secondary_return = True
        rows.append(
            {
                "key": item["key"],
                "label": item["label"],
                "probability_pct": effective_probability_pct,
                "return_pct": quantize_decimal(item["return_pct"], "0.01"),
                "contribution_return_pct": contribution_return_pct,
                "projected_price": projected_price,
                "secondary_return_pct": secondary_return_pct,
            }
        )

    expected_return_pct = quantize_decimal(summary.get("expected_return_pct"), "0.01")
    if rows and expected_return_pct is not None:
        accumulated_contribution = sum(
            (row["contribution_return_pct"] for row in rows[:-1] if row.get("contribution_return_pct") is not None),
            ZERO,
        )
        rows[-1]["contribution_return_pct"] = quantize_decimal(
            expected_return_pct - accumulated_contribution,
            "0.01",
        )

    return {
        **summary,
        "available": True,
        "rows": rows,
        "expected_projected_price": quantize_decimal(expected_projected_price, "0.0001") if has_projected_price else None,
        "expected_secondary_return_pct": quantize_decimal(expected_secondary_return_pct, "0.01") if has_secondary_return else None,
    }


def resolve_cycle_scenario_return_pct_for_months(
    scenario: dict | None,
    *,
    current_price: Decimal | None,
    months: int,
) -> Decimal | None:
    scenario = scenario or {}
    key_map = {
        12: "year_1_return_pct",
        24: "year_2_return_pct",
        36: "year_3_return_pct",
        48: "year_4_return_pct",
        60: "year_5_return_pct",
    }
    direct_key = key_map.get(months)
    direct_return_pct = quantize_decimal(scenario.get(direct_key), "0.01") if direct_key else None
    if direct_return_pct is None and months == 60:
        direct_return_pct = quantize_decimal(scenario.get("five_year_return_pct"), "0.01")
    if direct_return_pct is not None:
        return direct_return_pct

    current_price = quantize_decimal(current_price, "0.0001")
    for point in list(scenario.get("path") or []):
        if parse_projection_label_months(point.get("label")) != months:
            continue
        projected_price = quantize_decimal(point.get("projected_price"), "0.0001")
        if projected_price is not None and current_price not in {None, ZERO}:
            return quantize_decimal(percentage_change(projected_price, current_price), "0.01")

    annual_return_pct = quantize_decimal(scenario.get("annual_return_pct"), "0.01")
    if annual_return_pct is None:
        return None
    base = 1 + (float(annual_return_pct) / 100)
    if base <= 0:
        return Decimal("-100.00")
    years = months / 12
    cumulative_return_pct = (base**years - 1) * 100
    return Decimal(str(round(cumulative_return_pct, 4))).quantize(Decimal("0.01"))


def build_cycle_horizon_expectation_map(
    scenarios: list[dict] | None,
    *,
    current_price: Decimal | None,
) -> dict[int, Decimal]:
    normalized_rows = list(scenarios or [])
    if not normalized_rows:
        return {}

    expectations = {}
    for months in (24, 36, 48, 60):
        weighted_sum = ZERO
        available_probability_pct = ZERO
        for row in normalized_rows:
            probability_pct = Decimal(str(row.get("probability_pct") or "0"))
            return_pct = resolve_cycle_scenario_return_pct_for_months(
                row,
                current_price=current_price,
                months=months,
            )
            if return_pct is None or probability_pct <= ZERO:
                continue
            weighted_sum += return_pct * probability_pct
            available_probability_pct += probability_pct
        if available_probability_pct > ZERO:
            expectations[months] = quantize_decimal(weighted_sum / available_probability_pct, "0.01")
    return expectations


def build_card_scenario_tables(card: dict) -> dict:
    projection = card.get("projection") or {}
    cycle_projection = card.get("cycle_projection_5y") or {}
    effective_projection = resolve_effective_projection_metrics(card)
    return {
        "projection_12m": build_scenario_expectation_table(
            build_effective_projection_scenarios(card, effective_metrics=effective_projection),
            return_key="total_return_pct",
            fallback_value=quantize_decimal(effective_projection.get("base_return_pct"), "0.01"),
        ),
        "cycle_5y": build_scenario_expectation_table(
            cycle_projection.get("scenarios"),
            return_key="five_year_return_pct",
            fallback_value=quantize_decimal(cycle_projection.get("five_year_return_pct"), "0.01"),
            secondary_return_key="annual_return_pct",
        ),
    }


def reconcile_trade_alert_with_visible_projection(trade_alert: dict, presentation_projection: dict) -> dict:
    if not trade_alert or not presentation_projection.get("available"):
        return trade_alert

    tone = trade_alert.get("tone")
    visible_price_return_pct = quantize_decimal(presentation_projection.get("visible_price_return_pct"), "0.01")
    visible_total_return_pct = quantize_decimal(presentation_projection.get("visible_total_return_pct"), "0.01")
    visible_projected_price = quantize_decimal(presentation_projection.get("visible_projected_price"), "0.0001")
    shape_label = presentation_projection.get("shape_label") or ""
    display_price_return_pct = visible_price_return_pct if visible_price_return_pct is not None else ZERO
    display_total_return_pct = visible_total_return_pct if visible_total_return_pct is not None else ZERO
    display_projected_price = visible_projected_price if visible_projected_price is not None else ZERO

    closes_lower_visibly = (
        (visible_price_return_pct is not None and visible_price_return_pct < ZERO)
        or (visible_total_return_pct is not None and visible_total_return_pct < ZERO)
        or shape_label == "Caida progresiva"
    )
    closes_higher_visibly = (
        (visible_price_return_pct is not None and visible_price_return_pct > ZERO)
        and (visible_total_return_pct is not None and visible_total_return_pct > ZERO)
    )

    if tone == "buy" and closes_lower_visibly:
        score = quantize_decimal(trade_alert.get("score"), "0.01")
        return {
            **trade_alert,
            "label": "Vigilar",
            "tone": "watch",
            "score": min(score, Decimal("3.10")) if score is not None else Decimal("3.10"),
            "trigger_label": "La senda visible 12M sigue bajista",
            "note": (
                "La senal relativa venia apoyando compras, pero la senda visible 12M cierra por debajo de hoy "
                f"({display_price_return_pct:.1f} % en precio y {display_total_return_pct:.1f} % neto, "
                f"hasta {display_projected_price:.4f}). Conviene vigilar y esperar un giro claro antes de comprar."
            ),
            "coherence_adjusted": True,
        }

    if tone == "sell" and closes_higher_visibly:
        score = quantize_decimal(trade_alert.get("score"), "0.01")
        return {
            **trade_alert,
            "label": "Vigilar",
            "tone": "watch",
            "score": max(score, Decimal("-3.10")) if score is not None else Decimal("-3.10"),
            "trigger_label": "La senda visible 12M aun cierra al alza",
            "note": (
                "La lectura relativa salia debil, pero la senda visible 12M sigue cerrando por encima de hoy "
                f"({display_price_return_pct:.1f} % en precio y {display_total_return_pct:.1f} % neto, "
                f"hasta {display_projected_price:.4f}). Conviene vigilar antes de activar venta."
            ),
            "coherence_adjusted": True,
        }

    return trade_alert


def reconcile_trade_alert_with_expected_return(
    trade_alert: dict,
    *,
    expected_return_pct: Decimal | None,
    horizon_label: str = "1A",
) -> dict:
    if not trade_alert or expected_return_pct is None:
        return trade_alert

    tone = trade_alert.get("tone")
    expected_return_pct = quantize_decimal(expected_return_pct, "0.01")
    display_expected_return_pct = expected_return_pct if expected_return_pct is not None else ZERO

    if tone == "buy" and expected_return_pct < ZERO:
        score = quantize_decimal(trade_alert.get("score"), "0.01")
        return {
            **trade_alert,
            "label": "Vigilar",
            "tone": "watch",
            "score": min(score, Decimal("3.10")) if score is not None else Decimal("3.10"),
            "trigger_label": f"La esperanza {horizon_label} sigue negativa",
            "note": (
                "La tendencia relativa venia apoyando compras, pero la media ponderada de escenarios "
                f"a {horizon_label} sigue en {display_expected_return_pct:.1f} %. "
                f"Conviene vigilar y esperar a que la esperanza {horizon_label} vuelva a positivo "
                "antes de activar compra."
            ),
            "coherence_adjusted": True,
        }

    if tone == "sell" and expected_return_pct > ZERO:
        score = quantize_decimal(trade_alert.get("score"), "0.01")
        return {
            **trade_alert,
            "label": "Vigilar",
            "tone": "watch",
            "score": max(score, Decimal("-3.10")) if score is not None else Decimal("-3.10"),
            "trigger_label": f"La esperanza {horizon_label} aun sigue positiva",
            "note": (
                "La senal relativa venia pidiendo reducir, pero la media ponderada de escenarios "
                f"a {horizon_label} sigue en {display_expected_return_pct:.1f} %. "
                f"Conviene vigilar y esperar a que la esperanza {horizon_label} se deteriore "
                "de forma clara antes de activar venta."
            ),
            "coherence_adjusted": True,
        }

    return trade_alert


def refresh_card_projection_visuals(card: dict, history=None, *, include_visuals: bool = True) -> dict:
    position = card.get("position")
    if position is None:
        card["presentation_projection"] = {"available": False}
        card["information_basis"] = {"available": False, "rows": []}
        card["scenario_tables"] = {
            "projection_12m": {"available": False, "rows": []},
            "cycle_5y": {"available": False, "rows": []},
        }
        return card

    if history is None:
        if getattr(position, "pk", None):
            history = list(position.price_history.order_by("price_date"))
        else:
            history = []

    if include_visuals and history and len(history) >= 2:
        card["projection_12m_chart"] = build_projection_12m_chart(history, card.get("projection") or {})
        card["cycle_projection_5y_chart"] = build_cycle_projection_5y_chart(history, card.get("cycle_projection_5y") or {})
    else:
        card.setdefault("projection_12m_chart", {"available": False})
        card.setdefault("cycle_projection_5y_chart", {"available": False})
    card["presentation_projection"] = build_projection_presentation_summary(card)
    card["scenario_tables"] = build_card_scenario_tables(card)
    trade_alert = reconcile_trade_alert_with_visible_projection(
        card.get("trade_alert") or {},
        card["presentation_projection"],
    )
    trade_alert = reconcile_trade_alert_with_expected_return(
        trade_alert,
        expected_return_pct=(card.get("scenario_tables") or {}).get("projection_12m", {}).get("expected_return_pct"),
        horizon_label="1A",
    )
    card["trade_alert"] = trade_alert
    card["information_basis"] = build_information_basis_summary(card)
    return card


def collapse_market_points_to_frequency(points: list[dict], frequency: str) -> list[dict]:
    buckets = {}
    for point in points:
        point_date = point.get("date")
        if point_date is None:
            continue
        buckets[bucket_label_for_date(point_date, frequency)] = point
    return [buckets[label] for label in sorted(buckets.keys())]


def calculate_native_annual_change(
    series_points: list[dict],
    *,
    periods_back: int,
    change_mode: str,
) -> Decimal | None:
    if len(series_points) < 2:
        return None
    end_point = series_points[-1]
    start_index = max(len(series_points) - 1 - periods_back, 0)
    start_point = series_points[start_index]
    years_covered = calculate_years_between(start_point.get("date"), end_point.get("date"))
    if years_covered in {None, ZERO}:
        return None
    start_value = start_point.get("close")
    end_value = end_point.get("close")
    if change_mode == "absolute":
        if start_value is None or end_value is None:
            return None
        return Decimal(str(round(float((end_value - start_value) / years_covered), 4)))
    return calculate_series_cagr_pct(start_value, end_value, years_covered)


def calculate_reference_projection_change(
    projected_value: Decimal | None,
    current_value: Decimal | None,
    *,
    change_mode: str,
) -> Decimal | None:
    return calculate_series_change(projected_value, current_value, change_mode=change_mode)


def resolve_reference_projection_bounds(
    reference_profile: str,
    latest_value: Decimal,
) -> tuple[Decimal | None, Decimal | None]:
    if reference_profile == EquityPosition.ReferenceProfile.EURIBOR_12M:
        return Decimal("-1.50"), Decimal("7.50")
    if latest_value <= ZERO:
        return None, None
    return None, None


def clamp_reference_projected_value(
    value: Decimal,
    *,
    lower_bound: Decimal | None,
    upper_bound: Decimal | None,
) -> Decimal:
    if lower_bound is not None:
        value = max(value, lower_bound)
    if upper_bound is not None:
        value = min(value, upper_bound)
    return value


def resolve_reference_forward_signal(reference_profile: str) -> dict:
    if reference_profile != EquityPosition.ReferenceProfile.EURIBOR_12M:
        return {
            "available": False,
            "source_label": "",
            "symbol": "",
            "target_value": None,
            "latest_date": None,
            "note": "",
        }

    try:
        forward_series = fetch_ecb_yield_curve_series(
            "B.U2.EUR.4F.G_N_C.SV_C_YM.IF_5Y",
            EURIBOR_FORWARD_REFERENCE_NAME,
            last_n_observations=260,
        )
    except Exception as exc:
        return {
            "available": False,
            "source_label": "Curva BCE forward 5A",
            "symbol": EURIBOR_FORWARD_REFERENCE_SYMBOL,
            "target_value": None,
            "latest_date": None,
            "note": f"No se ha podido cargar la senal adelantada oficial del BCE: {exc}",
        }

    return {
        "available": True,
        "source_label": EURIBOR_FORWARD_REFERENCE_NAME,
        "symbol": EURIBOR_FORWARD_REFERENCE_SYMBOL,
        "target_value": quantize_decimal(forward_series.latest_price),
        "latest_date": forward_series.latest_date,
        "note": "La proyeccion incorpora la curva forward 5A del BCE como ancla de tipos futuros.",
    }


def build_reference_projection_5y(
    history,
    reference_choice: dict,
    reference_cache: dict | None = None,
    include_visuals: bool = True,
) -> dict:
    reference_cache = reference_cache if reference_cache is not None else {}
    cache_key = (
        reference_choice["reference_profile"],
        reference_choice["benchmark_symbol"],
        reference_choice["benchmark_name"],
    )
    try:
        cached_value = reference_cache.get(cache_key)
        if cached_value is None:
            cached_value = fetch_reference_series_for_choice(
                reference_choice["reference_profile"],
                benchmark_symbol=reference_choice["benchmark_symbol"],
                benchmark_name=reference_choice["benchmark_name"],
            )
            reference_cache[cache_key] = cached_value
        if isinstance(cached_value, Exception):
            raise cached_value
        reference_series = cached_value
    except Exception as exc:
        reference_cache[cache_key] = exc
        return {
            "available": False,
            "reference_label": reference_choice["benchmark_name"],
            "note": f"No se ha podido leer la serie de referencia: {exc}",
        }

    if reference_series is None:
        return {
            "available": False,
            "reference_label": reference_choice["benchmark_name"],
            "note": "No hay serie disponible para esta referencia.",
        }

    correlation = build_reference_correlation_for_series(
        history,
        reference_series,
        reference_choice["reference_profile"],
    )
    frequency = infer_reference_frequency_from_profile(reference_choice["reference_profile"])
    change_mode = infer_reference_change_mode(reference_choice["reference_profile"])
    collapsed_points = collapse_market_points_to_frequency(reference_series.points, frequency)
    minimum_points = 8 if frequency == "monthly" else 6
    if len(collapsed_points) < minimum_points:
        return {
            "available": False,
            "reference_label": reference_choice["benchmark_name"],
            "note": "No hay suficiente historico de la referencia para proyectarla a 5 anos.",
        }

    latest_point = collapsed_points[-1]
    latest_value = latest_point.get("close")
    latest_date = latest_point.get("date")
    if latest_value is None or latest_date is None:
        return {
            "available": False,
            "reference_label": reference_choice["benchmark_name"],
            "note": "La referencia no tiene un ultimo valor valido.",
        }

    long_periods = 60 if frequency == "monthly" else 20
    medium_periods = 36 if frequency == "monthly" else 12
    recent_periods = 12 if frequency == "monthly" else 4
    long_annual_change = calculate_native_annual_change(
        collapsed_points,
        periods_back=long_periods,
        change_mode=change_mode,
    )
    medium_annual_change = calculate_native_annual_change(
        collapsed_points,
        periods_back=medium_periods,
        change_mode=change_mode,
    )
    recent_annual_change = calculate_native_annual_change(
        collapsed_points,
        periods_back=recent_periods,
        change_mode=change_mode,
    )

    change_components = []
    recent_weight_used = Decimal("0.00")
    recent_sign_conflict = False
    if long_annual_change is not None:
        change_components.append(long_annual_change * Decimal("0.55"))
    if medium_annual_change is not None:
        change_components.append(medium_annual_change * Decimal("0.30"))
    if recent_annual_change is not None:
        recent_weight = Decimal("0.15")
        if long_annual_change is not None and recent_annual_change * long_annual_change < ZERO:
            recent_weight = Decimal("0.08")
            recent_sign_conflict = True
        recent_weight_used = recent_weight
        change_components.append(recent_annual_change * recent_weight)
    if not change_components:
        return {
            "available": False,
            "reference_label": reference_choice["benchmark_name"],
            "note": "No se ha podido inferir una tendencia anual util para esta referencia.",
        }

    base_annual_change = sum(change_components, ZERO)
    raw_annual_change = base_annual_change
    if change_mode == "absolute":
        base_annual_change = clamp_decimal(base_annual_change, Decimal("-1.50"), Decimal("1.50"))
    else:
        base_annual_change = clamp_decimal(base_annual_change, Decimal("-14.00"), Decimal("14.00"))
    annual_change_clamped = base_annual_change != raw_annual_change

    forward_signal = resolve_reference_forward_signal(reference_choice["reference_profile"])
    lower_bound, upper_bound = resolve_reference_projection_bounds(
        reference_choice["reference_profile"],
        latest_value,
    )

    beta = correlation.get("beta")
    current_projected_value = latest_value
    current_raw_projected_value = latest_value
    path = []
    raw_path = []
    correction_rows = []
    for year in range(1, 6):
        decay = Decimal(str(round(0.90 ** (year - 1), 4)))
        if change_mode == "absolute":
            yearly_trend = base_annual_change * decay
            trend_value = current_projected_value + yearly_trend
            raw_yearly_trend = raw_annual_change * decay
            raw_trend_value = current_raw_projected_value + raw_yearly_trend
        else:
            yearly_trend = base_annual_change * decay
            trend_value = project_price_from_return(current_projected_value, yearly_trend) or current_projected_value
            raw_yearly_trend = raw_annual_change * decay
            raw_trend_value = project_price_from_return(current_raw_projected_value, raw_yearly_trend) or current_raw_projected_value

        projected_value = trend_value
        anchor_weight = ZERO
        anchor_path_value = None
        if forward_signal.get("available") and forward_signal.get("target_value") is not None:
            anchor_weight = Decimal("0.35") + (Decimal(str(year - 1)) * Decimal("0.02"))
            anchor_path_value = latest_value + (
                (Decimal(str(forward_signal["target_value"])) - latest_value) * Decimal(str(year)) / Decimal("5")
            )
            projected_value = (trend_value * (Decimal("1.00") - anchor_weight)) + (anchor_path_value * anchor_weight)

        unclamped_projected_value = projected_value
        projected_value = clamp_reference_projected_value(
            projected_value,
            lower_bound=lower_bound,
            upper_bound=upper_bound,
        )
        bounds_adjustment_applied = projected_value != unclamped_projected_value
        cumulative_change = calculate_reference_projection_change(
            projected_value,
            latest_value,
            change_mode=change_mode,
        )
        raw_cumulative_change = calculate_reference_projection_change(
            raw_trend_value,
            latest_value,
            change_mode=change_mode,
        )
        implied_stock_return_pct = None
        raw_implied_stock_return_pct = None
        if beta is not None and cumulative_change is not None:
            implied_stock_return_pct = clamp_decimal(beta * cumulative_change, Decimal("-80.00"), Decimal("140.00"))
        if beta is not None and raw_cumulative_change is not None:
            raw_implied_stock_return_pct = clamp_decimal(beta * raw_cumulative_change, Decimal("-80.00"), Decimal("140.00"))
        point_date = latest_date + timedelta(days=365 * year)
        raw_path.append(
            {
                "label": f"{year}A",
                "projected_date": point_date,
                "projected_value": quantize_decimal(raw_trend_value, "0.0001"),
                "cumulative_change": quantize_decimal(raw_cumulative_change),
                "implied_stock_return_pct": quantize_decimal(raw_implied_stock_return_pct),
            }
        )
        path.append(
            {
                "label": f"{year}A",
                "projected_date": point_date,
                "projected_value": quantize_decimal(projected_value, "0.0001"),
                "cumulative_change": quantize_decimal(cumulative_change),
                "implied_stock_return_pct": quantize_decimal(implied_stock_return_pct),
            }
        )
        correction_rows.append(
            {
                "label": f"{year}A",
                "projected_date": point_date,
                "raw_projected_value": quantize_decimal(raw_trend_value, "0.0001"),
                "trend_projected_value": quantize_decimal(trend_value, "0.0001"),
                "adjusted_projected_value": quantize_decimal(projected_value, "0.0001"),
                "anchor_path_value": quantize_decimal(anchor_path_value, "0.0001") if anchor_path_value is not None else None,
                "anchor_weight_pct": quantize_decimal(anchor_weight * ONE_HUNDRED),
                "bounds_adjustment_applied": bounds_adjustment_applied,
                "raw_implied_stock_return_pct": quantize_decimal(raw_implied_stock_return_pct),
                "adjusted_implied_stock_return_pct": quantize_decimal(implied_stock_return_pct),
            }
        )
        current_projected_value = projected_value
        current_raw_projected_value = raw_trend_value

    projected_value_5y = path[-1]["projected_value"]
    implied_stock_return_5y_pct = path[-1]["implied_stock_return_pct"]
    implied_stock_annual_return_pct = annualize_return_pct(implied_stock_return_5y_pct, 60)
    rolling_rows = build_reference_rolling_correlation_rows_for_series(
        history,
        reference_series,
        reference_choice["reference_profile"],
    )
    history_periods = 60 if frequency == "monthly" else 20
    historical_points = [
        {"date": point["date"], "value": point["close"]}
        for point in collapsed_points[-history_periods:]
        if point.get("date") and point.get("close") is not None
    ]
    projection_chart = (
        build_reference_projection_detail_chart(
            historical_points,
            [{"date": row["projected_date"], "value": row["projected_value"]} for row in raw_path],
            [{"date": row["projected_date"], "value": row["projected_value"]} for row in path],
        )
        if include_visuals
        else {"available": False}
    )
    coefficient_chart = (
        build_reference_coefficient_trend_chart(rolling_rows, correlation.get("coefficient"))
        if include_visuals
        else {"available": False}
    )
    return {
        "available": True,
        "reference_label": reference_choice["benchmark_name"],
        "reference_profile": reference_choice["reference_profile"],
        "benchmark_symbol": reference_choice["benchmark_symbol"],
        "change_mode": change_mode,
        "change_unit_label": change_unit_label(change_mode),
        "current_value": quantize_decimal(latest_value, "0.0001"),
        "projected_value_5y": projected_value_5y,
        "raw_annual_change": quantize_decimal(raw_annual_change),
        "annual_change": quantize_decimal(base_annual_change),
        "projected_change_5y": path[-1]["cumulative_change"],
        "coefficient": correlation.get("coefficient"),
        "recent_coefficient": correlation.get("recent_coefficient"),
        "beta": beta,
        "recent_beta": correlation.get("recent_beta"),
        "observations_count": correlation.get("observations_count", 0),
        "stability_label": correlation.get("stability_label"),
        "path": path,
        "raw_path": raw_path,
        "rolling_rows": rolling_rows,
        "projection_chart": projection_chart,
        "coefficient_chart": coefficient_chart,
        "long_annual_change": quantize_decimal(long_annual_change),
        "medium_annual_change": quantize_decimal(medium_annual_change),
        "recent_annual_change": quantize_decimal(recent_annual_change),
        "recent_weight_pct": quantize_decimal(recent_weight_used * ONE_HUNDRED),
        "recent_sign_conflict": recent_sign_conflict,
        "annual_change_clamped": annual_change_clamped,
        "lower_bound": quantize_decimal(lower_bound) if lower_bound is not None else None,
        "upper_bound": quantize_decimal(upper_bound) if upper_bound is not None else None,
        "correction_rows": correction_rows,
        "forward_signal": forward_signal,
        "implied_stock_return_5y_pct": implied_stock_return_5y_pct,
        "implied_stock_annual_return_pct": quantize_decimal(implied_stock_annual_return_pct),
        "note": (
            f"Se proyecta {reference_choice['benchmark_name']} a 5A con serie web oficial, "
            f"mezclando tendencia larga/media/corta y usando senal adelantada cuando existe."
        ),
    }


def build_multifactor_reference_projection_bundle(
    history,
    position: EquityPosition,
    reference_cache: dict | None = None,
    include_visuals: bool = True,
) -> dict:
    reference_cache = reference_cache if reference_cache is not None else {}
    factors = []
    for reference_choice in build_reference_suggestions_for_equity(position.company_name, position.ticker):
        factor = build_reference_projection_5y(
            history,
            reference_choice,
            reference_cache=reference_cache,
            include_visuals=include_visuals,
        )
        if not factor.get("available"):
            continue
        coefficient = factor.get("coefficient")
        beta = factor.get("beta")
        if coefficient is None or beta is None or factor.get("implied_stock_return_5y_pct") is None:
            continue
        factors.append(factor)

    if not factors:
        return {
            "available": False,
            "factors": [],
            "note": "No hay suficientes coeficientes consistentes para levantar un modelo multifactor 5A.",
            "forward_signal_count": 0,
        }

    factors.sort(
        key=lambda item: (
            -(abs(item.get("coefficient") or ZERO)),
            -(item.get("observations_count") or 0),
            0 if item.get("forward_signal", {}).get("available") else 1,
            item.get("reference_label") or "",
        )
    )
    selected_factors = factors[:3]
    raw_weights = []
    for factor in selected_factors:
        coefficient = abs(factor.get("coefficient") or ZERO)
        observations_count = Decimal(str(min(int(factor.get("observations_count") or 0), 60)))
        coefficient_component = coefficient
        observations_component = observations_count / Decimal("100")
        raw_weight = coefficient_component + observations_component
        stability_multiplier = Decimal("1.00")
        if factor.get("stability_label") == "Estable":
            stability_multiplier = Decimal("1.08")
        elif factor.get("stability_label") == "Cambiante":
            stability_multiplier = Decimal("0.82")
        raw_weight *= stability_multiplier
        forward_bonus = ZERO
        if factor.get("forward_signal", {}).get("available"):
            forward_bonus = Decimal("0.10")
            raw_weight += forward_bonus
        factor["weight_inputs"] = {
            "coefficient_component": quantize_decimal(coefficient_component),
            "observations_component": quantize_decimal(observations_component),
            "stability_multiplier": quantize_decimal(stability_multiplier),
            "forward_bonus": quantize_decimal(forward_bonus),
        }
        factor["raw_weight"] = quantize_decimal(raw_weight)
        raw_weights.append(max(raw_weight, Decimal("0.10")))

    total_weight = sum(raw_weights, ZERO)
    if total_weight <= ZERO:
        total_weight = Decimal(str(len(selected_factors) or 1))
        raw_weights = [Decimal("1.00")] * len(selected_factors)

    weighted_five_year_return_pct = ZERO
    weighted_abs_correlation = ZERO
    forward_signal_count = 0
    for factor, raw_weight in zip(selected_factors, raw_weights):
        weight_ratio = raw_weight / total_weight
        factor["weight_pct"] = quantize_decimal(weight_ratio * ONE_HUNDRED)
        weighted_five_year_return_pct += (factor.get("implied_stock_return_5y_pct") or ZERO) * weight_ratio
        weighted_abs_correlation += abs(factor.get("coefficient") or ZERO) * weight_ratio
        if factor.get("forward_signal", {}).get("available"):
            forward_signal_count += 1

    weighted_five_year_return_pct = clamp_decimal(weighted_five_year_return_pct, Decimal("-65.00"), Decimal("120.00"))
    weighted_annual_return_pct = annualize_return_pct(weighted_five_year_return_pct, 60)
    blend_ratio_pct = clamp_decimal(
        Decimal("18.00")
        + (weighted_abs_correlation * Decimal("28.00"))
        + (Decimal("4.00") * Decimal(str(max(len(selected_factors) - 1, 0))))
        + (Decimal("6.00") if forward_signal_count else ZERO),
        Decimal("18.00"),
        Decimal("42.00"),
    )
    factor_labels = ", ".join(factor.get("reference_label") or "" for factor in selected_factors)
    note = (
        f"El ajuste multifactor incorpora {len(selected_factors)} coeficiente(s): {factor_labels}. "
        f"Se ponderan por correlacion, muestra historica y estabilidad, y explican un {blend_ratio_pct:.0f} % de la senda 5A final."
    )
    if forward_signal_count:
        note += " Cuando existe una senal adelantada oficial en la web, se usa como ancla adicional del escenario."
    return {
        "available": True,
        "factors": selected_factors,
        "five_year_return_pct": quantize_decimal(weighted_five_year_return_pct),
        "annual_return_pct": quantize_decimal(weighted_annual_return_pct),
        "blend_ratio_pct": quantize_decimal(blend_ratio_pct),
        "weighted_abs_correlation": quantize_decimal(weighted_abs_correlation),
        "forward_signal_count": forward_signal_count,
        "note": note,
    }


def build_base_five_year_cycle_projection(
    history,
    position: EquityPosition,
    correlation: dict,
    cycle_metrics: dict | None = None,
) -> dict:
    if len(history) < 2:
        return {"available": False}

    raw_latest_date = history[-1].price_date
    analysis_history = trim_current_partial_month_for_long_horizon(
        [point for point in history if point.price_date <= raw_latest_date],
        min_points=24,
    )
    latest_date = analysis_history[-1].price_date
    monthly_history = collapse_history_to_frequency(analysis_history, "monthly")
    if len(monthly_history) < 24:
        return {"available": False}

    cycle_metrics = build_cycle_metrics(analysis_history, max_days=None)
    latest_price = monthly_history[-1].close_price
    three_year_snapshot = build_period_snapshot(
        analysis_history,
        "3Y",
        start_date=latest_date - timedelta(days=365 * 3),
        end_date=latest_date,
        reference_profile=position.reference_profile,
    )
    five_year_snapshot = build_period_snapshot(
        analysis_history,
        "5Y",
        start_date=latest_date - timedelta(days=365 * 5),
        end_date=latest_date,
        reference_profile=position.reference_profile,
    )
    coefficient = correlation.get("coefficient")
    beta = correlation.get("beta")
    annual_return_components = []

    cagr_pct = cycle_metrics.get("cagr_pct")
    if cagr_pct is not None:
        annual_return_components.append(cagr_pct * Decimal("0.50"))
    if five_year_snapshot.get("available") and five_year_snapshot.get("stock_return_pct") is not None:
        five_year_signal = annualize_return_pct(five_year_snapshot["stock_return_pct"], 60)
        if five_year_signal is not None:
            annual_return_components.append(five_year_signal * Decimal("0.30"))
    if three_year_snapshot.get("available") and three_year_snapshot.get("stock_return_pct") is not None:
        three_year_signal = annualize_return_pct(three_year_snapshot["stock_return_pct"], 36)
        if three_year_signal is not None:
            annual_return_components.append(three_year_signal * Decimal("0.20"))
    reference_five_year_change = five_year_snapshot.get("benchmark_change") if five_year_snapshot.get("available") else None
    if beta is not None and reference_five_year_change is not None:
        reference_five_year_signal = annualize_return_pct(beta * reference_five_year_change, 60)
        if reference_five_year_signal is not None:
            annual_return_components.append(reference_five_year_signal * Decimal("0.08"))
    elif coefficient is not None and five_year_snapshot.get("benchmark_return_pct") is not None:
        reference_five_year_signal = annualize_return_pct(five_year_snapshot["benchmark_return_pct"], 60)
        if reference_five_year_signal is not None:
            annual_return_components.append(reference_five_year_signal * coefficient * Decimal("0.08"))

    current_drawdown_pct = cycle_metrics.get("current_drawdown_pct")
    if current_drawdown_pct is not None:
        if current_drawdown_pct <= Decimal("-20.00"):
            annual_return_components.append(Decimal("3.00"))
        elif current_drawdown_pct <= Decimal("-8.00"):
            annual_return_components.append(Decimal("1.50"))
        elif current_drawdown_pct >= Decimal("-3.00"):
            annual_return_components.append(Decimal("-1.20"))

    annual_return_components.append(
        {
            "Correccion": Decimal("2.25"),
            "Recuperacion": Decimal("1.25"),
            "Expansion": Decimal("0.25"),
            "Transicion": Decimal("-0.50"),
        }.get(cycle_metrics.get("cycle_phase") or "Transicion", ZERO)
    )

    positive_year_ratio_pct = cycle_metrics.get("positive_year_ratio_pct")
    if positive_year_ratio_pct is not None:
        annual_return_components.append(
            clamp_decimal((positive_year_ratio_pct - Decimal("50.00")) * Decimal("0.04"), Decimal("-0.80"), Decimal("0.80"))
        )

    if not annual_return_components:
        return {"available": False}

    annual_return_pct = sum(annual_return_components, ZERO)
    years_covered = cycle_metrics.get("years_covered", ZERO)
    if years_covered < Decimal("5.00"):
        annual_return_pct *= Decimal("0.85")
    elif years_covered >= Decimal("8.00"):
        annual_return_pct *= Decimal("0.96")
    annual_return_pct = clamp_decimal(annual_return_pct, Decimal("-10.00"), Decimal("16.00"))

    half_year_returns = []
    for index in range(6, len(monthly_history), 6):
        half_year_return = percentage_change(monthly_history[index].close_price, monthly_history[index - 6].close_price)
        if half_year_return is not None:
            half_year_returns.append(half_year_return)
    positive_half_year_returns = [value for value in half_year_returns if value > ZERO]
    negative_half_year_returns = [value for value in half_year_returns if value < ZERO]
    base_half_year_return = Decimal(
        str(round(((max(0.01, 1 + (float(annual_return_pct) / 100)) ** 0.5) - 1) * 100, 4))
    )
    upside_return = median_decimal(positive_half_year_returns)
    if upside_return is None:
        upside_return = clamp_decimal(base_half_year_return * Decimal("1.40"), Decimal("2.50"), Decimal("10.00"))
    downside_return = median_decimal(negative_half_year_returns)
    downside_floor = -clamp_decimal(
        max(
            abs(cycle_metrics.get("max_drawdown_pct") or ZERO) * Decimal("0.22"),
            (cycle_metrics.get("annualized_volatility_pct") or Decimal("16.00")) * Decimal("0.32"),
            Decimal("2.50"),
        ),
        Decimal("2.50"),
        Decimal("10.00"),
    )
    if downside_return is None:
        downside_return = downside_floor
    else:
        downside_return = min(downside_return, downside_floor)
    mild_up = average_decimal([base_half_year_return, upside_return]) or base_half_year_return
    mild_down = average_decimal([base_half_year_return, downside_return]) or downside_return
    flat_return = base_half_year_return * Decimal("0.35")
    recovery_return = average_decimal([upside_return, mild_up]) or mild_up
    correction_return = average_decimal([downside_return, mild_down]) or mild_down
    phase_sequence = {
        "Expansion": [
            recovery_return,
            correction_return,
            mild_up,
            downside_return,
            recovery_return,
            mild_down,
            upside_return,
            correction_return,
            mild_up,
            flat_return,
        ],
        "Recuperacion": [
            upside_return,
            mild_up,
            correction_return,
            recovery_return,
            mild_down,
            upside_return,
            correction_return,
            mild_up,
            flat_return,
            mild_down,
        ],
        "Correccion": [
            downside_return,
            mild_up,
            correction_return,
            recovery_return,
            mild_down,
            upside_return,
            downside_return,
            mild_up,
            correction_return,
            recovery_return,
        ],
        "Transicion": [
            correction_return,
            mild_up,
            flat_return,
            downside_return,
            recovery_return,
            mild_down,
            upside_return,
            correction_return,
            mild_up,
            flat_return,
        ],
    }
    step_return_pcts = list(phase_sequence.get(cycle_metrics.get("cycle_phase") or "Transicion", phase_sequence["Transicion"]))
    schedule_average = average_decimal(step_return_pcts) or ZERO
    schedule_shift = base_half_year_return - schedule_average
    step_return_pcts = [
        clamp_decimal(value + schedule_shift, Decimal("-12.00"), Decimal("12.00"))
        for value in step_return_pcts
    ]
    if all(value >= ZERO for value in step_return_pcts):
        step_return_pcts[3] = min(downside_return, Decimal("-1.50"))
    if all(value <= ZERO for value in step_return_pcts):
        step_return_pcts[1] = max(upside_return, Decimal("1.50"))

    reference_cycle_template = build_reference_cycle_template(
        position,
        latest_date=latest_date,
        years=5,
        step_months=6,
    )
    if reference_cycle_template.get("available"):
        step_return_pcts = [
            Decimal(str(value))
            for value in (reference_cycle_template.get("step_return_pcts") or step_return_pcts)
        ]

    path = build_cycle_projection_path(
        latest_price,
        annual_return_pct,
        annualized_volatility_pct=cycle_metrics.get("annualized_volatility_pct"),
        current_drawdown_pct=current_drawdown_pct,
        cycle_phase=cycle_metrics.get("cycle_phase") or "Transicion",
        anchor_date=latest_date,
        years=5,
        step_months=6,
        step_return_pcts=step_return_pcts,
    )
    if not path:
        return {"available": False}

    projected_price = path[-1]["projected_price"]
    five_year_return_pct = percentage_change(projected_price, latest_price)
    analysis_years_used = years_covered or ZERO
    confidence = build_projection_confidence(
        correlation.get("coefficient"),
        correlation.get("observations_count", 0),
        cycle_metrics.get("monthly_observations_count", 0),
        years_covered=years_covered,
        positive_year_ratio_pct=cycle_metrics.get("positive_year_ratio_pct"),
        stability_gap=correlation.get("stability_gap"),
    )
    safety = build_safety_score(cycle_metrics, correlation)
    scenario_spread_annual_pct = resolve_cycle_projection_scenario_spread(
        cycle_metrics.get("annualized_volatility_pct"),
        confidence["label"],
        safety["label"],
    )
    scenarios = build_five_year_projection_scenarios(
        latest_price,
        latest_date=latest_date,
        annual_return_pct=annual_return_pct,
        scenario_spread_annual_pct=scenario_spread_annual_pct,
        step_return_pcts=step_return_pcts,
        annualized_volatility_pct=cycle_metrics.get("annualized_volatility_pct"),
        current_drawdown_pct=current_drawdown_pct,
        cycle_phase=cycle_metrics.get("cycle_phase") or "Transicion",
        confidence_label=confidence["label"],
    )
    explanation = (
        f"Esta vista 5A usa {analysis_years_used:.1f} anos del historico disponible para leer el ciclo de {position.company_name}. "
        f"Combina CAGR del ciclo, ritmo a 3-5 anos, drawdown actual y fase {cycle_metrics.get('cycle_phase', 'sin ciclo').lower()} "
        "para dibujar una senda larga de orientacion."
    )
    if raw_latest_date != latest_date:
        explanation += (
            f" Para evitar saltos de pocos dias, el 5A queda anclado al ultimo mes cerrado ({latest_date:%Y-%m-%d}) "
            f"y no al movimiento parcial de {raw_latest_date:%Y-%m-%d}."
        )
    if reference_cycle_template.get("available"):
        explanation += (
            f" La forma de la senda sale de {reference_cycle_template.get('years_covered') or ZERO:.1f} anos de "
            f"{reference_cycle_template.get('reference_label') or position.analysis_reference_label}, buscando ventanas historicas parecidas al punto actual para incluir ciclos economicos completos."
        )
    else:
        explanation += " Como no hay suficiente referencia larga comparable, el modelo recurre a una plantilla sintetica de ciclo."
    return {
        "available": True,
        "annual_return_pct": quantize_decimal(annual_return_pct),
        "projected_price": projected_price,
        "five_year_return_pct": quantize_decimal(five_year_return_pct),
        "path": path,
        "cycle_phase": cycle_metrics.get("cycle_phase"),
        "analysis_years_used": analysis_years_used,
        "history_window_label": "Ultimos 5 anos visibles",
        "model_window_label": f"{analysis_years_used:.1f} anos de historico",
        "step_return_pcts": [quantize_decimal(value) or ZERO for value in step_return_pcts],
        "latest_price": quantize_decimal(latest_price, "0.0001"),
        "latest_date": latest_date,
        "raw_latest_date": raw_latest_date,
        "uses_completed_month_anchor": raw_latest_date != latest_date,
        "annualized_volatility_pct": cycle_metrics.get("annualized_volatility_pct"),
        "current_drawdown_pct": current_drawdown_pct,
        "confidence_label": confidence["label"],
        "confidence_score_pct": confidence["score_pct"],
        "safety_label": safety["label"],
        "safety_score": safety["score"],
        "scenario_spread_annual_pct": quantize_decimal(scenario_spread_annual_pct),
        "scenarios": scenarios,
        "reference_cycle_template": reference_cycle_template,
        "base_explanation": explanation,
        "explanation": explanation,
    }


def build_five_year_cycle_projection(
    history,
    position: EquityPosition,
    correlation: dict,
    cycle_metrics: dict | None = None,
    reference_cache: dict | None = None,
    include_visuals: bool = True,
) -> dict:
    base_projection = build_base_five_year_cycle_projection(
        history,
        position,
        correlation,
        cycle_metrics=cycle_metrics,
    )
    if not base_projection.get("available"):
        return base_projection

    factor_bundle = build_multifactor_reference_projection_bundle(
        history,
        position,
        reference_cache=reference_cache,
        include_visuals=include_visuals,
    )
    if not factor_bundle.get("available"):
        return {
            **base_projection,
            "factor_model_available": False,
            "factor_blend_ratio_pct": ZERO,
            "factor_five_year_return_pct": None,
            "factor_annual_return_pct": None,
            "factors": [],
            "factor_model_label": "Solo ciclo historico",
        }

    latest_price = base_projection.get("latest_price")
    latest_date = base_projection.get("latest_date")
    base_annual_return_pct = base_projection.get("annual_return_pct") or ZERO
    factor_annual_return_pct = factor_bundle.get("annual_return_pct") or ZERO
    blend_ratio_pct = factor_bundle.get("blend_ratio_pct") or ZERO
    blend_ratio = blend_ratio_pct / ONE_HUNDRED
    final_annual_return_pct = clamp_decimal(
        (base_annual_return_pct * (Decimal("1.00") - blend_ratio)) + (factor_annual_return_pct * blend_ratio),
        Decimal("-10.00"),
        Decimal("18.00"),
    )

    original_step_returns = [Decimal(str(value)) for value in (base_projection.get("step_return_pcts") or [])]
    current_average = average_decimal(original_step_returns) or ZERO
    target_average = Decimal(
        str(round(((max(0.01, 1 + (float(final_annual_return_pct) / 100)) ** 0.5) - 1) * 100, 4))
    )
    step_shift = target_average - current_average
    factor_target_average = Decimal(
        str(round(((max(0.01, 1 + (float(factor_annual_return_pct) / 100)) ** 0.5) - 1) * 100, 4))
    )
    factor_step_shift = factor_target_average - current_average
    factor_step_returns = [
        clamp_decimal(step_return + factor_step_shift, Decimal("-12.00"), Decimal("12.00"))
        for step_return in original_step_returns
    ]
    adjusted_step_returns = [
        clamp_decimal(step_return + step_shift, Decimal("-12.00"), Decimal("12.00"))
        for step_return in original_step_returns
    ]

    factor_path = build_cycle_projection_path(
        latest_price,
        factor_annual_return_pct,
        annualized_volatility_pct=base_projection.get("annualized_volatility_pct"),
        current_drawdown_pct=base_projection.get("current_drawdown_pct"),
        cycle_phase=base_projection.get("cycle_phase") or "Transicion",
        anchor_date=latest_date,
        years=5,
        step_months=6,
        step_return_pcts=factor_step_returns,
    )

    final_path, adjusted_step_returns, step_shift = build_cycle_projection_path_for_target(
        latest_price,
        annual_return_pct=final_annual_return_pct,
        step_return_pcts=original_step_returns,
        annualized_volatility_pct=base_projection.get("annualized_volatility_pct"),
        current_drawdown_pct=base_projection.get("current_drawdown_pct"),
        cycle_phase=base_projection.get("cycle_phase") or "Transicion",
        anchor_date=latest_date,
        years=5,
        step_months=6,
    )
    if not final_path:
        return base_projection

    projected_price = final_path[-1]["projected_price"]
    five_year_return_pct = percentage_change(projected_price, latest_price)
    scenario_spread_annual_pct = resolve_cycle_projection_scenario_spread(
        base_projection.get("annualized_volatility_pct"),
        base_projection.get("confidence_label") or "Media",
        base_projection.get("safety_label") or "Media",
    )
    scenarios = build_five_year_projection_scenarios(
        latest_price,
        latest_date=latest_date,
        annual_return_pct=final_annual_return_pct,
        scenario_spread_annual_pct=scenario_spread_annual_pct,
        step_return_pcts=adjusted_step_returns,
        annualized_volatility_pct=base_projection.get("annualized_volatility_pct"),
        current_drawdown_pct=base_projection.get("current_drawdown_pct"),
        cycle_phase=base_projection.get("cycle_phase") or "Transicion",
        confidence_label=base_projection.get("confidence_label") or "Media",
    )
    explanation = (
        f"{base_projection.get('base_explanation', '').strip()} "
        f"{factor_bundle.get('note', '').strip()} "
        f"La senda final mezcla un {max(Decimal('0.00'), Decimal('100.00') - blend_ratio_pct):.0f} % de ciclo propio "
        f"y un {blend_ratio_pct:.0f} % de lectura multifactor."
    ).strip()
    comparison_chart = (
        build_cycle_projection_comparison_chart(
            history,
            base_projection.get("path") or [],
            factor_path,
            final_path,
        )
        if include_visuals
        else {"available": False}
    )
    return {
        **base_projection,
        "annual_return_pct": quantize_decimal(final_annual_return_pct),
        "projected_price": projected_price,
        "five_year_return_pct": quantize_decimal(five_year_return_pct),
        "path": final_path,
        "step_return_pcts": [quantize_decimal(value) or ZERO for value in adjusted_step_returns],
        "base_annual_return_pct": quantize_decimal(base_annual_return_pct),
        "base_projected_price": base_projection.get("projected_price"),
        "base_path": base_projection.get("path") or [],
        "factor_path": factor_path,
        "factor_model_available": True,
        "factor_model_label": f"{len(factor_bundle.get('factors', []))} coeficientes",
        "factor_blend_ratio_pct": quantize_decimal(blend_ratio_pct),
        "factor_five_year_return_pct": factor_bundle.get("five_year_return_pct"),
        "factor_annual_return_pct": factor_bundle.get("annual_return_pct"),
        "final_step_shift": quantize_decimal(step_shift),
        "factor_step_shift": quantize_decimal(factor_step_shift),
        "factors": factor_bundle.get("factors", []),
        "uses_forward_web_signals": bool(factor_bundle.get("forward_signal_count")),
        "forward_signal_count": factor_bundle.get("forward_signal_count", 0),
        "weighted_abs_correlation": factor_bundle.get("weighted_abs_correlation"),
        "scenario_spread_annual_pct": quantize_decimal(scenario_spread_annual_pct),
        "scenarios": scenarios,
        "comparison_chart": comparison_chart,
        "explanation": explanation,
    }


def build_cycle_projection_5y_chart(history, cycle_projection: dict) -> dict:
    if len(history) < 2 or not cycle_projection.get("available"):
        return {"available": False}

    latest_date = history[-1].price_date
    recent_history = filter_history_window(
        history,
        start_date=latest_date - timedelta(days=365 * 5),
        end_date=latest_date,
    )
    if len(recent_history) < 2:
        recent_history = history
    stock_series = [{"date": point.price_date, "value": point.close_price} for point in recent_history]
    projection_series = [{"date": latest_date, "value": recent_history[-1].close_price}]
    projection_series.extend(
        {
            "date": step["projected_date"],
            "value": step["projected_price"],
        }
        for step in cycle_projection.get("path", [])
        if step.get("projected_date") and step.get("projected_price") is not None
    )
    chart = build_dual_axis_chart(stock_series, [], projection_points=projection_series)
    end_projection_step = cycle_projection.get("path", [])[-1] if cycle_projection.get("path") else None
    return {
        "available": bool(chart.get("stock_line") and chart.get("projection_line")),
        "stock_line": chart.get("stock_line", ""),
        "projection_line": chart.get("projection_line", ""),
        "stock_min_label": chart.get("stock_min_label", "-"),
        "stock_max_label": chart.get("stock_max_label", "-"),
        "x_markers": chart.get("x_markers", []),
        "start_label": recent_history[0].price_date.isoformat(),
        "end_label": latest_date.isoformat(),
        "projection_end_label": end_projection_step.get("projected_date").isoformat() if end_projection_step and end_projection_step.get("projected_date") else "",
        "points_count": len(stock_series),
        "history_window_label": cycle_projection.get("history_window_label", "Ultimos 5 anos"),
        "model_window_label": cycle_projection.get("model_window_label", ""),
    }


def build_ticket_snapshot_projection_values(card: dict) -> dict:
    position = card["position"]
    effective_projection = resolve_effective_projection_metrics(card)
    projected_price = effective_projection.get("projected_price")
    projected_market_value_12m = None
    projected_total_value_12m = None
    if projected_price is not None:
        projected_market_value_12m = quantize_decimal(position.shares * projected_price, "0.01")
    if effective_projection.get("base_return_pct") is not None:
        projected_total_value_12m = quantize_decimal(
            position.current_value * (Decimal("1") + (effective_projection["base_return_pct"] / ONE_HUNDRED)),
            "0.01",
        )
    if projected_total_value_12m is None:
        projected_total_value_12m = projected_market_value_12m
    return {
        "projected_price_12m": projected_price,
        "projected_market_value_12m": projected_market_value_12m,
        "projected_total_value_12m": projected_total_value_12m,
    }


def capture_equity_ticket_snapshots(history_cards: list[dict], snapshot_date: date | None = None) -> list[EquityTicketSnapshot]:
    snapshot_date = snapshot_date or django_timezone.localdate()
    snapshots = []
    owned_cards = [card for card in history_cards if card["position"].is_owned]
    with transaction.atomic():
        for card in owned_cards:
            position = card["position"]
            projection_values = build_ticket_snapshot_projection_values(card)
            snapshot, _ = EquityTicketSnapshot.objects.update_or_create(
                position=position,
                snapshot_date=snapshot_date,
                defaults={
                    "invested_amount": quantize_decimal(position.invested_amount, "0.01") or ZERO,
                    "current_value": quantize_decimal(position.current_value, "0.01") or ZERO,
                    "projected_market_value_12m": projection_values["projected_market_value_12m"],
                    "projected_total_value_12m": projection_values["projected_total_value_12m"],
                    "projected_price_12m": projection_values["projected_price_12m"],
                },
            )
            snapshots.append(snapshot)
    return snapshots


def capture_missing_equity_ticket_snapshots(
    history_cards: list[dict],
    snapshot_date: date | None = None,
) -> list[EquityTicketSnapshot]:
    owned_cards = [
        card
        for card in history_cards
        if card["position"].is_owned and card["position"].id
    ]
    if not owned_cards:
        return []

    owned_position_ids = [card["position"].id for card in owned_cards]
    existing_position_ids = set(
        EquityTicketSnapshot.objects.filter(position_id__in=owned_position_ids)
        .values_list("position_id", flat=True)
        .distinct()
    )
    has_missing_cards = any(
        card["position"].id not in existing_position_ids
        for card in owned_cards
    )
    if not has_missing_cards:
        return []
    return capture_equity_ticket_snapshots(owned_cards, snapshot_date=snapshot_date)


def project_expected_value_on_date(
    start_value: Decimal | None,
    target_value: Decimal | None,
    start_date: date | None,
    point_date: date | None,
    horizon_days: int = TRACKING_HORIZON_DAYS,
) -> Decimal | None:
    if start_value is None or target_value is None or start_date is None or point_date is None:
        return None
    if point_date <= start_date:
        return quantize_decimal(start_value, "0.01")

    elapsed_days = min(max((point_date - start_date).days, 0), horizon_days)
    ratio = Decimal(str(elapsed_days / horizon_days))
    if start_value > ZERO and target_value > ZERO:
        multiplier = float(target_value / start_value)
        projected_multiplier = Decimal(str(multiplier ** float(ratio)))
        return quantize_decimal(start_value * projected_multiplier, "0.01")
    return quantize_decimal(start_value + ((target_value - start_value) * ratio), "0.01")


def normalize_tracking_series(points: list[dict] | None) -> list[dict]:
    grouped = {}
    for point in points or []:
        point_date = point.get("date")
        value = point.get("value")
        if point_date is None or value is None:
            continue
        grouped[point_date] = {
            **point,
            "date": point_date,
            "value": Decimal(str(value)),
        }
    return [grouped[point_date] for point_date in sorted(grouped)]


def build_tracking_series_difference(
    primary_series: list[dict],
    comparison_series: list[dict],
) -> list[dict]:
    primary_points = normalize_tracking_series(primary_series)
    comparison_points = normalize_tracking_series(comparison_series)
    if not primary_points or not comparison_points:
        return []

    primary_by_date = {point["date"]: point["value"] for point in primary_points}
    comparison_by_date = {point["date"]: point["value"] for point in comparison_points}
    combined_dates = sorted(set(primary_by_date) | set(comparison_by_date))
    current_primary = None
    current_comparison = None
    difference_series = []

    for point_date in combined_dates:
        if point_date in primary_by_date:
            current_primary = primary_by_date[point_date]
        if point_date in comparison_by_date:
            current_comparison = comparison_by_date[point_date]
        if current_primary is None or current_comparison is None:
            continue
        difference_value = quantize_decimal(current_primary - current_comparison, "0.01")
        if difference_value is None:
            continue
        difference_series.append({"date": point_date, "value": difference_value})

    return difference_series


def build_tracking_trailing_delta_series(
    series: list[dict],
    trailing_days: int = TRACKING_WEEKLY_ALPHA_DAYS,
) -> list[dict]:
    normalized_points = normalize_tracking_series(series)
    if len(normalized_points) < 2 or trailing_days <= 0:
        return []

    anchor_cursor = 0
    anchor_value = None
    trailing_series = []

    for point in normalized_points:
        anchor_date = point["date"] - timedelta(days=trailing_days)
        while anchor_cursor < len(normalized_points) and normalized_points[anchor_cursor]["date"] <= anchor_date:
            anchor_value = normalized_points[anchor_cursor]["value"]
            anchor_cursor += 1
        if anchor_value is None:
            continue
        delta_value = quantize_decimal(point["value"] - anchor_value, "0.01")
        if delta_value is None:
            continue
        trailing_series.append({"date": point["date"], "value": delta_value})

    return trailing_series


def build_value_tracking_chart(
    actual_series: list[dict],
    expected_series: list[dict],
    benchmark_series: list[dict] | None = None,
    value_suffix: str = "EUR",
    axis_formatter=None,
    time_marker_mode: str = "auto",
    grid_marker_mode: str = "none",
    allow_expected_only: bool = False,
    width: int = 640,
    height: int = 220,
    padding: int = 18,
) -> dict:
    axis_formatter = axis_formatter or format_axis_value
    min_marker_distance = 10.0
    expected_marker_distance = 22.0

    def format_chart_value(value: Decimal | None) -> str:
        formatted = axis_formatter(value)
        if formatted == "-":
            return formatted
        return f"{formatted} {value_suffix}".strip()

    actual_points = normalize_tracking_series(actual_series)
    expected_points = normalize_tracking_series(expected_series)
    benchmark_points = normalize_tracking_series(benchmark_series or [])
    primary_points = actual_points or expected_points or benchmark_points
    if (
        not primary_points
        or max(len(actual_points), len(expected_points), len(benchmark_points)) < 2
        or (not actual_points and not allow_expected_only)
    ):
        return {
            "available": False,
            "actual_line": "",
            "expected_line": "",
            "benchmark_line": "",
            "actual_points": [],
            "actual_display_points": [],
            "expected_points": [],
            "expected_display_points": [],
            "benchmark_points": [],
            "benchmark_display_points": [],
            "min_label": "-",
            "max_label": "-",
            "start_label": "",
            "latest_label": "",
            "projection_end_label": "",
            "points_count": 0,
            "zero_y": None,
            "x_markers": [],
            "grid_markers": [],
            "segmented_time_axis": False,
            "projection_zone_start_x": None,
            "projection_zone_width": None,
            "scale_note": "",
            "latest_gap_line": None,
            "latest_expected_point": None,
        }

    all_values = (
        [point["value"] for point in actual_points]
        + [point["value"] for point in expected_points]
        + [point["value"] for point in benchmark_points]
    )
    series_min = min(all_values)
    series_max = max(all_values)
    if series_min == series_max:
        series_max += Decimal("1")

    all_dates = [point["date"] for point in primary_points]
    if expected_points and expected_points is not primary_points:
        all_dates.extend(point["date"] for point in expected_points)
    if benchmark_points and benchmark_points is not primary_points:
        all_dates.extend(point["date"] for point in benchmark_points)
    if actual_points and actual_points is not primary_points:
        all_dates.extend(point["date"] for point in actual_points)
    min_date = min(all_dates)
    max_date = max(all_dates)
    total_days = max((max_date - min_date).days, 1)
    span_x = width - (padding * 2)
    span_y = height - (padding * 2)
    latest_actual_date = actual_points[-1]["date"] if actual_points else primary_points[0]["date"]
    projection_end_date = expected_points[-1]["date"] if expected_points else latest_actual_date
    observed_days = max((latest_actual_date - min_date).days, 0)
    future_days = max((projection_end_date - latest_actual_date).days, 0)
    segmented_time_axis = bool(
        actual_points
        and
        expected_points
        and future_days >= 45
        and future_days >= max(observed_days * 2, 30)
    )
    observed_ratio = None
    if segmented_time_axis:
        observed_ratio = max(
            0.30,
            min(0.46, 0.30 + ((observed_days / max(total_days, 1)) * 1.8)),
        )
    projection_zone_start_x = None

    def scale_x(point_date: date) -> float:
        nonlocal projection_zone_start_x
        if not segmented_time_axis:
            return padding + (span_x * ((point_date - min_date).days / total_days))

        observed_span_x = span_x * float(observed_ratio or 0.34)
        future_span_x = span_x - observed_span_x
        if point_date <= latest_actual_date:
            observed_denominator = max(observed_days, 1)
            x_value = padding + (observed_span_x * ((point_date - min_date).days / observed_denominator))
        else:
            future_denominator = max(future_days, 1)
            x_value = padding + observed_span_x + (future_span_x * ((point_date - latest_actual_date).days / future_denominator))
        if projection_zone_start_x is None:
            projection_zone_start_x = f"{padding + observed_span_x:.1f}"
        return x_value

    def scale_point(point_date: date, value: Decimal) -> tuple[float, float]:
        x = scale_x(point_date)
        normalized_value = (value - series_min) / (series_max - series_min)
        y = height - padding - (normalized_value * span_y)
        return x, y

    def build_series(points: list[dict], prefix: str) -> tuple[str, list[dict]]:
        line_points = []
        point_rows = []
        for point in points:
            x, y = scale_point(point["date"], point["value"])
            line_points.append(f"{x:.1f},{y:.1f}")
            point_rows.append(
                {
                    "x": f"{x:.1f}",
                    "y": f"{y:.1f}",
                    "date": point["date"],
                    "value_label": format_chart_value(point["value"]),
                    "date_label": point["date"].isoformat(),
                    "tooltip": f"{point['date'].isoformat()} | {format_chart_value(point['value'])}",
                    "key": f"{prefix}-{point['date'].isoformat()}",
                    "label": point.get("label"),
                    "is_anchor": bool(point.get("is_anchor")),
                }
            )
        return " ".join(line_points) if len(line_points) >= 2 else "", point_rows

    def build_display_points(point_rows: list[dict]) -> list[dict]:
        if len(point_rows) <= 2:
            return [
                {
                    **point,
                    "cluster_size": 1,
                    "is_latest": point["key"] == point_rows[-1]["key"],
                    "radius": "4.5" if point["key"] == point_rows[-1]["key"] else "4.0",
                }
                for point in point_rows
            ]

        selected_keys = {
            point_rows[0]["key"],
            point_rows[-1]["key"],
        }
        for point in point_rows:
            if point["date"].weekday() in TRACKING_MARKER_WEEKDAYS:
                selected_keys.add(point["key"])
        selected_points = [point for point in point_rows if point["key"] in selected_keys]

        def distance(left: dict, right: dict) -> float:
            delta_x = float(left["x"]) - float(right["x"])
            delta_y = float(left["y"]) - float(right["y"])
            return (delta_x**2 + delta_y**2) ** 0.5

        clusters: list[list[dict]] = []
        current_cluster = [selected_points[0]]
        for point in selected_points[1:]:
            if distance(point, current_cluster[-1]) < min_marker_distance:
                current_cluster.append(point)
            else:
                clusters.append(current_cluster)
                current_cluster = [point]
        clusters.append(current_cluster)

        latest_key = selected_points[-1]["key"]
        display_points = []
        for cluster in clusters:
            representative = dict(cluster[-1])
            representative["cluster_size"] = len(cluster)
            representative["is_latest"] = representative["key"] == latest_key
            representative["radius"] = "5.5" if representative["is_latest"] else "4.2"
            if len(cluster) > 1:
                start_label = cluster[0]["date_label"]
                end_label = cluster[-1]["date_label"]
                if start_label == end_label:
                    representative["tooltip"] = (
                        f"{end_label} | {len(cluster)} lecturas | ultimo {cluster[-1]['value_label']}"
                    )
                else:
                    representative["tooltip"] = (
                        f"{start_label} a {end_label} | {len(cluster)} lecturas | ultimo {cluster[-1]['value_label']}"
                    )
            display_points.append(representative)
        return display_points

    def reduce_rows(rows: list[dict], max_points: int) -> list[dict]:
        if len(rows) <= max_points:
            return rows
        if max_points <= 2:
            return [rows[0], rows[-1]]
        step = max(1, int(math.ceil((len(rows) - 2) / max(max_points - 2, 1))))
        reduced = [rows[0]]
        reduced.extend(row for index, row in enumerate(rows[1:-1], start=1) if index % step == 0)
        if rows[-1]["key"] != reduced[-1]["key"]:
            reduced.append(rows[-1])
        return reduced[: max_points - 1] + [rows[-1]] if len(reduced) > max_points else reduced

    def build_expected_display_points(point_rows: list[dict]) -> list[dict]:
        if not point_rows:
            return []

        selected_rows = [point_rows[0], point_rows[-1]]
        current_expected_row = next(
            (point for point in reversed(point_rows) if point["date"] <= latest_actual_date),
            None,
        )
        if current_expected_row is not None:
            selected_rows.append(current_expected_row)

        anchor_rows = [point for point in point_rows if point.get("is_anchor") and point["date"] > latest_actual_date]
        if anchor_rows:
            max_anchor_points = 6 if future_days <= 450 else 5
            if future_days > 900:
                yearly_anchor_rows = [point for point in anchor_rows if str(point.get("label") or "").endswith("A")]
                anchor_rows = yearly_anchor_rows or anchor_rows
            selected_rows.extend(reduce_rows(anchor_rows, max_anchor_points))

        deduped_rows = []
        seen_keys = set()
        for row in selected_rows:
            if row["key"] in seen_keys:
                continue
            seen_keys.add(row["key"])
            deduped_rows.append(row)
        deduped_rows.sort(key=lambda row: row["date"])

        if len(deduped_rows) == 1:
            row = dict(deduped_rows[0])
            row["radius"] = "4.2"
            row["is_latest_expected"] = True
            return [row]

        clusters: list[list[dict]] = []
        current_cluster = [deduped_rows[0]]

        def distance(left: dict, right: dict) -> float:
            delta_x = float(left["x"]) - float(right["x"])
            delta_y = float(left["y"]) - float(right["y"])
            return (delta_x**2 + delta_y**2) ** 0.5

        for row in deduped_rows[1:]:
            if distance(row, current_cluster[-1]) < expected_marker_distance:
                current_cluster.append(row)
            else:
                clusters.append(current_cluster)
                current_cluster = [row]
        clusters.append(current_cluster)

        latest_key = deduped_rows[-1]["key"]
        display_points = []
        for cluster in clusters:
            representative = dict(cluster[-1])
            representative["cluster_size"] = len(cluster)
            representative["is_latest_expected"] = representative["key"] == latest_key
            representative["radius"] = "4.8" if representative["is_latest_expected"] else "3.6"
            if len(cluster) > 1:
                representative["tooltip"] = (
                    f"{cluster[0]['date_label']} a {cluster[-1]['date_label']} | "
                    f"{len(cluster)} hitos esperados | ultimo {cluster[-1]['value_label']}"
                )
            display_points.append(representative)
        return display_points

    def build_segmented_axis_markers(
        actual_point_rows: list[dict],
        expected_point_rows: list[dict],
    ) -> list[dict]:
        if not segmented_time_axis:
            return []

        def short_date_label(point_date: date) -> str:
            month_label = SPANISH_MONTH_LABELS.get(point_date.month, str(point_date.month))[:3].capitalize()
            return f"{month_label} {str(point_date.year)[2:]}"

        marker_rows = []
        if actual_point_rows:
            marker_rows.append(
                {
                    **actual_point_rows[0],
                    "marker_label": short_date_label(actual_point_rows[0]["date"]),
                    "is_major": True,
                }
            )
            marker_rows.append(
                {
                    **actual_point_rows[-1],
                    "marker_label": "Hoy",
                    "is_major": True,
                }
            )

        future_anchor_rows = [point for point in expected_point_rows if point.get("is_anchor") and point["date"] > latest_actual_date]
        if future_anchor_rows:
            if future_days > 900:
                yearly_rows = [point for point in future_anchor_rows if str(point.get("label") or "").endswith("A")]
                future_anchor_rows = yearly_rows or future_anchor_rows
            future_anchor_rows = reduce_rows(future_anchor_rows, 5 if future_days <= 450 else 6)
            for row in future_anchor_rows:
                marker_rows.append(
                    {
                        **row,
                        "marker_label": str(row.get("label") or short_date_label(row["date"])).strip(),
                        "is_major": str(row.get("label") or "").endswith("A"),
                    }
                )

        serialized_markers = []
        seen_dates = set()
        for row in sorted(marker_rows, key=lambda item: item["date"]):
            marker_date = row["date"]
            if marker_date in seen_dates:
                continue
            seen_dates.add(marker_date)
            serialized_markers.append(
                {
                    "x": row["x"],
                    "label": row["marker_label"],
                    "date": marker_date.isoformat(),
                    "y1": str(height - padding),
                    "y2": str(height - padding + 7),
                    "text_y": str(height - 2),
                    "draw_grid": bool(row.get("is_anchor")) and marker_date > latest_actual_date,
                    "grid_y1": str(padding),
                    "grid_y2": str(height - padding),
                    "is_major": bool(row.get("is_major")),
                    "show_label": True,
                    "anchor": "start" if marker_date == min_date else ("end" if marker_date == projection_end_date else "middle"),
                }
            )
        return serialized_markers

    actual_line, actual_point_rows = build_series(actual_points, "actual")
    expected_line, expected_point_rows = build_series(expected_points, "expected")
    benchmark_line, benchmark_point_rows = build_series(benchmark_points, "benchmark")
    actual_display_points = build_display_points(actual_point_rows)
    expected_display_points = build_expected_display_points(expected_point_rows)
    benchmark_display_points = build_display_points(benchmark_point_rows) if benchmark_point_rows else []
    zero_y = None
    if series_min <= ZERO <= series_max:
        zero_y = f"{scale_point(min_date, ZERO)[1]:.1f}"

    if segmented_time_axis:
        x_markers = build_segmented_axis_markers(actual_point_rows, expected_point_rows)
    elif time_marker_mode == "month":
        x_markers = build_month_axis_markers(
            all_dates,
            width=width,
            height=height,
            padding=padding,
        )
    else:
        x_markers = build_time_axis_markers(
            all_dates,
            width=width,
            height=height,
            padding=padding,
        )

    if segmented_time_axis:
        grid_markers = [marker for marker in x_markers if marker.get("draw_grid")]
    elif grid_marker_mode == "month":
        grid_markers = [marker for marker in x_markers if marker.get("draw_grid")]
    else:
        grid_markers = []

    latest_expected_point = next(
        (point for point in reversed(expected_point_rows) if point["date"] <= latest_actual_date),
        None,
    )
    latest_gap_line = None
    if latest_expected_point is not None and actual_point_rows:
        latest_actual_point = actual_point_rows[-1]
        latest_gap_line = {
            "x": latest_actual_point["x"],
            "y1": latest_actual_point["y"],
            "y2": latest_expected_point["y"],
            "actual_value_label": latest_actual_point["value_label"],
            "expected_value_label": latest_expected_point["value_label"],
        }

    return {
        "available": True,
        "actual_line": actual_line,
        "expected_line": expected_line,
        "benchmark_line": benchmark_line,
        "actual_points": actual_point_rows,
        "actual_display_points": actual_display_points,
        "expected_points": expected_point_rows,
        "expected_display_points": expected_display_points,
        "benchmark_points": benchmark_point_rows,
        "benchmark_display_points": benchmark_display_points,
        "min_label": format_chart_value(series_min),
        "max_label": format_chart_value(series_max),
        "start_label": min_date.isoformat(),
        "latest_label": latest_actual_date.isoformat(),
        "projection_end_label": (
            expected_points[-1]["date"].isoformat()
            if expected_points
            else latest_actual_date.isoformat()
        ),
        "points_count": len(primary_points),
        "zero_y": zero_y,
        "x_markers": x_markers,
        "grid_markers": grid_markers,
        "segmented_time_axis": segmented_time_axis,
        "projection_zone_start_x": projection_zone_start_x,
        "projection_zone_width": (
            f"{max((width - padding) - float(projection_zone_start_x), 0):.1f}"
            if projection_zone_start_x is not None
            else None
        ),
        "scale_note": (
            "Tramo real ampliado para separar mejor los primeros dias frente al horizonte proyectado."
            if segmented_time_axis
            else ""
        ),
        "latest_gap_line": latest_gap_line,
        "latest_expected_point": latest_expected_point,
        "has_actual_series": bool(actual_points),
        "has_expected_series": bool(expected_points),
        "has_benchmark_series": bool(benchmark_points),
    }


def build_ticket_expected_series(
    snapshots: list[EquityTicketSnapshot],
    target_value: Decimal | None,
    card: dict | None = None,
) -> tuple[list[dict], Decimal | None, date | None]:
    if not snapshots:
        return [], None, None
    baseline = snapshots[0]
    projected_end_date = baseline.snapshot_date + timedelta(days=TRACKING_HORIZON_DAYS)
    projection = (card or {}).get("projection") or {}
    projection_path = resolve_projection_tracking_path(projection)
    baseline_value = quantize_decimal(baseline.current_value, "0.01") or ZERO
    normalized_target_value = quantize_decimal(target_value, "0.01") or baseline_value

    anchors = []
    initial_unit_price = quantize_decimal(projection.get("latest_price"), "0.0001")
    if initial_unit_price is None:
        initial_unit_price = quantize_decimal(
            getattr(card.get("position") if card else None, "current_price_per_share", None),
            "0.0001",
        )
    final_projection_price = quantize_decimal(projection.get("projected_price"), "0.0001")

    for step_index, step in enumerate(projection_path, start=1):
        step_price = quantize_decimal(step.get("projected_price"), "0.0001")
        step_date = step.get("projected_date")
        if step_date is None:
            fallback_days = int(round((TRACKING_HORIZON_DAYS * step_index) / max(len(projection_path), 1)))
            step_date = baseline.snapshot_date + timedelta(days=fallback_days)
        progress_ratio = None
        if (
            initial_unit_price not in {None, ZERO}
            and final_projection_price not in {None, ZERO}
            and step_price not in {None, ZERO}
            and final_projection_price != initial_unit_price
        ):
            try:
                denominator = math.log(float(final_projection_price / initial_unit_price))
                numerator = math.log(float(step_price / initial_unit_price))
                if denominator != 0:
                    progress_ratio = Decimal(str(numerator / denominator))
            except (ValueError, ZeroDivisionError):
                progress_ratio = None
        if progress_ratio is None:
            progress_ratio = Decimal(str(fallback_days / TRACKING_HORIZON_DAYS))
        progress_ratio = clamp_decimal(progress_ratio, ZERO, Decimal("1.00"))
        if baseline_value > ZERO and normalized_target_value > ZERO:
            anchor_multiplier = float(normalized_target_value / baseline_value) ** float(progress_ratio)
            anchor_value = quantize_decimal(baseline_value * Decimal(str(anchor_multiplier)), "0.01")
        else:
            anchor_value = quantize_decimal(
                baseline_value + ((normalized_target_value - baseline_value) * progress_ratio),
                "0.01",
            )
        if anchor_value is not None:
            anchors.append(
                {
                    "date": step_date,
                    "value": anchor_value,
                    "label": step.get("label"),
                    "is_anchor": True,
                }
            )

    if anchors:
        series, latest_expected_value, projected_end_date = build_segmented_expected_series(
            snapshots,
            anchors,
        )
    else:
        series_dates = {baseline.snapshot_date, projected_end_date}
        for days in TRACKING_FORECAST_MARKERS:
            series_dates.add(baseline.snapshot_date + timedelta(days=days))
        for snapshot in snapshots:
            series_dates.add(snapshot.snapshot_date)

        series = []
        for point_date in sorted(series_dates):
            expected_value = project_expected_value_on_date(
                baseline.current_value,
                target_value or baseline.current_value,
                baseline.snapshot_date,
                point_date,
            )
            if expected_value is not None:
                series.append({"date": point_date, "value": expected_value})

        latest_snapshot_date = snapshots[-1].snapshot_date
        latest_expected_value = next(
            (point["value"] for point in reversed(series) if point["date"] <= latest_snapshot_date),
            None,
        )
    return series, latest_expected_value, projected_end_date


def resolve_tracking_projection_reliability_score(card: dict | None) -> Decimal:
    reliability = (card or {}).get("projection_reliability") or {}
    raw_score = reliability.get("score")
    if raw_score not in {None, ""}:
        try:
            return clamp_decimal(Decimal(str(raw_score)), ZERO, ONE_HUNDRED)
        except Exception:
            pass

    label = str(reliability.get("label") or "").strip()
    return {
        "Alta": Decimal("82.00"),
        "Media": Decimal("64.00"),
        "Baja": Decimal("42.00"),
    }.get(label, Decimal("50.00"))


def resolve_tracking_projection_backtest_score(card: dict | None) -> Decimal | None:
    card = card or {}
    if "projection_backtest" not in card:
        return None
    backtest = card.get("projection_backtest") or {}
    if backtest.get("available"):
        return {
            "Alta": Decimal("82.00"),
            "Media": Decimal("64.00"),
            "Baja": Decimal("42.00"),
        }.get(str(backtest.get("precision_label") or "").strip(), Decimal("42.00"))
    return Decimal("32.00")


def project_return_pct_from_annualized_rate(annualized_return_pct: Decimal | None, months: int) -> Decimal | None:
    if annualized_return_pct is None or months <= 0:
        return None
    years = Decimal(str(months)) / Decimal("12")
    growth_base = max(0.01, 1 + (float(annualized_return_pct) / 100))
    projected_return = ((growth_base ** float(years)) - 1) * 100
    return Decimal(str(round(projected_return, 4)))


def moderate_tracking_target_value(
    baseline_value: Decimal | None,
    target_value: Decimal | None,
    *,
    months: int,
    card: dict | None = None,
) -> tuple[Decimal | None, dict]:
    backtest_score = resolve_tracking_projection_backtest_score(card)
    base_payload = {
        "available": False,
        "quality_score": None,
        "annualized_target_pct": None,
        "annualized_cap_pct": None,
        "moderated_annualized_pct": None,
        "moderated_return_pct": None,
    }
    if baseline_value in {None, ZERO} or target_value is None or months <= 0 or backtest_score is None:
        return target_value, base_payload

    raw_target_return_pct = percentage_change(target_value, baseline_value)
    if raw_target_return_pct is None or raw_target_return_pct <= ZERO:
        return target_value, {
            **base_payload,
            "annualized_target_pct": quantize_decimal(raw_target_return_pct),
        }

    annualized_target_pct = (
        raw_target_return_pct
        if months <= 12
        else annualize_return_pct(raw_target_return_pct, months)
    )
    if annualized_target_pct is None or annualized_target_pct <= ZERO:
        return target_value, {
            **base_payload,
            "annualized_target_pct": quantize_decimal(annualized_target_pct),
        }

    reliability_score = resolve_tracking_projection_reliability_score(card)
    quality_score = average_decimal([reliability_score, backtest_score]) or reliability_score
    projection = (card or {}).get("projection") or {}
    raw_safety_score = projection.get("safety_score")
    try:
        safety_score = clamp_decimal(
            Decimal(str(raw_safety_score if raw_safety_score not in {None, ""} else Decimal("55.00"))),
            Decimal("15.00"),
            Decimal("92.00"),
        )
    except Exception:
        safety_score = Decimal("55.00")
    cagr_pct = projection.get("cagr_pct")
    cagr_component = ZERO
    if cagr_pct is not None:
        try:
            cagr_component = clamp_decimal(Decimal(str(cagr_pct)) * Decimal("0.15"), Decimal("-1.00"), Decimal("2.00"))
        except Exception:
            cagr_component = ZERO

    annualized_cap_pct = clamp_decimal(
        Decimal("3.80")
        + ((quality_score / ONE_HUNDRED) * Decimal("7.00"))
        + ((safety_score / ONE_HUNDRED) * Decimal("2.50"))
        + cagr_component,
        Decimal("5.00"),
        TRACKING_TARGET_MAX_ANNUAL_CAP_12M if months <= 12 else TRACKING_TARGET_MAX_ANNUAL_CAP_5Y,
    )
    if annualized_target_pct <= annualized_cap_pct + TRACKING_TARGET_MIN_EXCESS_PCT:
        return target_value, {
            **base_payload,
            "quality_score": quantize_decimal(quality_score),
            "annualized_target_pct": quantize_decimal(annualized_target_pct),
            "annualized_cap_pct": quantize_decimal(annualized_cap_pct),
        }

    excess_keep_ratio = clamp_decimal(
        TRACKING_TARGET_MIN_EXCESS_KEEP_RATIO + ((quality_score / ONE_HUNDRED) * Decimal("0.24")),
        TRACKING_TARGET_MIN_EXCESS_KEEP_RATIO,
        TRACKING_TARGET_MAX_EXCESS_KEEP_RATIO,
    )
    moderated_annualized_pct = annualized_cap_pct + ((annualized_target_pct - annualized_cap_pct) * excess_keep_ratio)
    moderated_return_pct = (
        moderated_annualized_pct
        if months <= 12
        else project_return_pct_from_annualized_rate(moderated_annualized_pct, months)
    )
    if moderated_return_pct is None:
        return target_value, {
            **base_payload,
            "quality_score": quantize_decimal(quality_score),
            "annualized_target_pct": quantize_decimal(annualized_target_pct),
            "annualized_cap_pct": quantize_decimal(annualized_cap_pct),
        }
    moderated_target_value = quantize_decimal(
        baseline_value * (Decimal("1") + (moderated_return_pct / ONE_HUNDRED)),
        "0.01",
    )
    if moderated_target_value is None:
        moderated_target_value = target_value

    return moderated_target_value, {
        "available": moderated_target_value != target_value,
        "quality_score": quantize_decimal(quality_score),
        "annualized_target_pct": quantize_decimal(annualized_target_pct),
        "annualized_cap_pct": quantize_decimal(annualized_cap_pct),
        "moderated_annualized_pct": quantize_decimal(moderated_annualized_pct),
        "moderated_return_pct": quantize_decimal(moderated_return_pct),
    }


def build_tracking_expected_series_calibration(
    snapshots: list[EquityTicketSnapshot],
    expected_series: list[dict],
    target_value: Decimal | None,
    card: dict | None = None,
) -> dict:
    baseline = snapshots[0] if snapshots else None
    latest = snapshots[-1] if snapshots else None
    raw_latest_expected_value = (
        next(
            (
                point["value"]
                for point in reversed(expected_series)
                if latest is not None and point["date"] <= latest.snapshot_date
            ),
            None,
        )
        if expected_series
        else None
    )
    base_payload = {
        "available": False,
        "tracked_days": (
            max((latest.snapshot_date - baseline.snapshot_date).days, 0)
            if baseline is not None and latest is not None
            else 0
        ),
        "latest_gap_value": None,
        "latest_gap_pct": None,
        "correction_weight": Decimal("0.00"),
        "target_carry_weight": Decimal("0.00"),
        "reliability_score": resolve_tracking_projection_reliability_score(card),
        "raw_latest_expected_value": raw_latest_expected_value,
        "adjusted_latest_expected_value": raw_latest_expected_value,
        "adjusted_target_value": quantize_decimal(target_value, "0.01"),
        "label": "Sin recalibracion",
    }
    if (
        baseline is None
        or latest is None
        or len(snapshots) < 2
        or raw_latest_expected_value is None
        or latest.current_value is None
    ):
        return base_payload

    tracked_days = base_payload["tracked_days"]
    if tracked_days < TRACKING_EXPECTED_FEEDBACK_MIN_DAYS:
        return base_payload

    latest_gap_value = quantize_decimal(latest.current_value - raw_latest_expected_value, "0.01")
    latest_gap_pct = percentage_change(latest.current_value, raw_latest_expected_value)
    latest_gap_pct_abs = abs(latest_gap_pct) if latest_gap_pct is not None else None
    if latest_gap_pct_abs is not None and latest_gap_pct_abs < TRACKING_EXPECTED_FEEDBACK_MIN_GAP_PCT:
        return {
            **base_payload,
            "latest_gap_value": latest_gap_value,
            "latest_gap_pct": quantize_decimal(latest_gap_pct),
            "label": "Gap asumible",
        }

    tracked_days_ratio = clamp_decimal(
        Decimal(str(tracked_days)) / Decimal(str(TRACKING_EXPECTED_FEEDBACK_FULL_STRENGTH_DAYS)),
        ZERO,
        Decimal("1.00"),
    )
    reliability_score = base_payload["reliability_score"]
    reliability_flex = clamp_decimal(
        Decimal("1.00") - (reliability_score / ONE_HUNDRED),
        Decimal("0.18"),
        Decimal("0.70"),
    )
    correction_weight = clamp_decimal(
        tracked_days_ratio * (Decimal("0.35") + reliability_flex),
        ZERO,
        TRACKING_EXPECTED_FEEDBACK_MAX_WEIGHT,
    )
    target_carry_weight = clamp_decimal(
        correction_weight * TRACKING_EXPECTED_FEEDBACK_TARGET_CARRY,
        ZERO,
        Decimal("0.30"),
    )
    adjusted_latest_expected_value = quantize_decimal(
        raw_latest_expected_value + ((latest.current_value - raw_latest_expected_value) * correction_weight),
        "0.01",
    )
    adjusted_target_value = quantize_decimal(target_value, "0.01")
    if adjusted_target_value is not None and latest_gap_value is not None:
        adjusted_target_value = quantize_decimal(
            max(ZERO, adjusted_target_value + ((latest.current_value - raw_latest_expected_value) * target_carry_weight)),
            "0.01",
        )

    return {
        "available": True,
        "tracked_days": tracked_days,
        "latest_gap_value": latest_gap_value,
        "latest_gap_pct": quantize_decimal(latest_gap_pct),
        "correction_weight": correction_weight,
        "target_carry_weight": target_carry_weight,
        "reliability_score": reliability_score,
        "raw_latest_expected_value": raw_latest_expected_value,
        "adjusted_latest_expected_value": adjusted_latest_expected_value,
        "adjusted_target_value": adjusted_target_value,
        "label": "Acelera" if latest_gap_value is not None and latest_gap_value > ZERO else "Enfria",
    }


def apply_tracking_expected_series_calibration(
    snapshots: list[EquityTicketSnapshot],
    expected_series: list[dict],
    target_value: Decimal | None,
    projected_end_date: date | None,
    card: dict | None = None,
) -> tuple[list[dict], Decimal | None, dict]:
    calibration = build_tracking_expected_series_calibration(
        snapshots,
        expected_series,
        target_value,
        card=card,
    )
    if not expected_series:
        return expected_series, None, calibration
    if not calibration.get("available"):
        return expected_series, calibration.get("raw_latest_expected_value"), calibration

    baseline = snapshots[0]
    latest = snapshots[-1]
    adjusted_latest_expected_value = calibration.get("adjusted_latest_expected_value")
    if adjusted_latest_expected_value is None:
        return expected_series, calibration.get("raw_latest_expected_value"), calibration

    tracked_days = max((latest.snapshot_date - baseline.snapshot_date).days, 1)
    effective_projected_end_date = projected_end_date or expected_series[-1]["date"]
    future_horizon_days = max((effective_projected_end_date - latest.snapshot_date).days, 1)
    adjusted_target_value = calibration.get("adjusted_target_value")

    calibrated_series = []
    for point in expected_series:
        point_date = point["date"]
        point_value = point.get("value")
        if point_date <= latest.snapshot_date:
            calibrated_value = project_expected_value_on_date(
                quantize_decimal(baseline.current_value, "0.01") or ZERO,
                adjusted_latest_expected_value,
                baseline.snapshot_date,
                point_date,
                horizon_days=tracked_days,
            )
        elif adjusted_target_value is not None:
            calibrated_value = project_expected_value_on_date(
                adjusted_latest_expected_value,
                adjusted_target_value,
                latest.snapshot_date,
                point_date,
                horizon_days=future_horizon_days,
            )
        else:
            calibrated_value = point_value
        if calibrated_value is not None:
            calibrated_series.append({"date": point_date, "value": calibrated_value})

    latest_expected_value = next(
        (point["value"] for point in reversed(calibrated_series) if point["date"] <= latest.snapshot_date),
        None,
    )
    return calibrated_series, latest_expected_value, calibration


def build_segmented_expected_series(
    snapshots: list[EquityTicketSnapshot],
    anchors: list[dict],
) -> tuple[list[dict], Decimal | None, date | None]:
    if not snapshots:
        return [], None, None

    baseline = snapshots[0]
    normalized_anchors = [
        {
            "date": baseline.snapshot_date,
            "value": quantize_decimal(baseline.current_value, "0.01") or ZERO,
            "label": "Hoy",
            "is_anchor": True,
        }
    ]
    for anchor in sorted(anchors, key=lambda item: item.get("date") or baseline.snapshot_date):
        anchor_date = anchor.get("date")
        anchor_value = anchor.get("value")
        if anchor_date is None or anchor_value is None or anchor_date <= baseline.snapshot_date:
            continue
        normalized_anchors.append(
            {
                "date": anchor_date,
                "value": quantize_decimal(anchor_value, "0.01") or ZERO,
                "label": anchor.get("label"),
                "is_anchor": bool(anchor.get("is_anchor", True)),
            }
        )

    if len(normalized_anchors) < 2:
        return [], None, None

    series_dates = {baseline.snapshot_date}
    for snapshot in snapshots:
        series_dates.add(snapshot.snapshot_date)
    for anchor in normalized_anchors[1:]:
        series_dates.add(anchor["date"])
    anchor_meta_by_date = {
        anchor["date"]: {
            "label": anchor.get("label"),
            "is_anchor": bool(anchor.get("is_anchor")),
        }
        for anchor in normalized_anchors
    }

    def project_on_anchor_path(point_date: date) -> Decimal | None:
        if point_date <= normalized_anchors[0]["date"]:
            return normalized_anchors[0]["value"]
        for index in range(1, len(normalized_anchors)):
            previous_anchor = normalized_anchors[index - 1]
            next_anchor = normalized_anchors[index]
            if point_date <= next_anchor["date"]:
                segment_days = max((next_anchor["date"] - previous_anchor["date"]).days, 1)
                return project_expected_value_on_date(
                    previous_anchor["value"],
                    next_anchor["value"],
                    previous_anchor["date"],
                    point_date,
                    horizon_days=segment_days,
                )
        return normalized_anchors[-1]["value"]

    series = []
    for point_date in sorted(series_dates):
        expected_value = project_on_anchor_path(point_date)
        if expected_value is not None:
            anchor_meta = anchor_meta_by_date.get(point_date, {})
            series.append(
                {
                    "date": point_date,
                    "value": expected_value,
                    "label": anchor_meta.get("label"),
                    "is_anchor": bool(anchor_meta.get("is_anchor")),
                }
            )

    latest_snapshot_date = snapshots[-1].snapshot_date
    latest_expected_value = next(
        (point["value"] for point in reversed(series) if point["date"] <= latest_snapshot_date),
        None,
    )
    return series, latest_expected_value, normalized_anchors[-1]["date"]


def build_ticket_expected_series_5y(
    card: dict,
    snapshots: list[EquityTicketSnapshot],
    purchase_baseline: EquityPurchaseForecastBaseline | None = None,
) -> tuple[list[dict], Decimal | None, date | None, Decimal | None]:
    if not snapshots:
        return [], None, None, None

    baseline_snapshot = snapshots[0]
    position = card["position"]
    baseline_current_value = quantize_decimal(baseline_snapshot.current_value, "0.01") or ZERO
    anchors = []

    def add_anchor(months: int | None, value: Decimal | None, label: str | None = None):
        if months is None or months <= 0 or value is None:
            return
        anchor_date = add_calendar_months(baseline_snapshot.snapshot_date, months)
        if anchor_date is None:
            return
        moderated_value, _ = moderate_tracking_target_value(
            baseline_current_value,
            quantize_decimal(value, "0.01") or ZERO,
            months=months,
            card=card,
        )
        anchors.append(
            {
                "date": anchor_date,
                "value": moderated_value if moderated_value is not None else (quantize_decimal(value, "0.01") or ZERO),
                "label": label or (f"{months // 12}A" if months % 12 == 0 else f"{months}M"),
                "is_anchor": True,
            }
        )

    if purchase_baseline:
        for step in purchase_baseline.projected_path_5y or []:
            label = str(step.get("label") or "").strip()
            raw_projected_price = step.get("projected_price")
            try:
                projected_price = Decimal(str(raw_projected_price)) if raw_projected_price not in {None, ""} else None
            except Exception:
                projected_price = None
            raw_projected_date = str(step.get("projected_date") or "").strip()
            projected_date = None
            if raw_projected_date:
                try:
                    projected_date = date.fromisoformat(raw_projected_date)
                except ValueError:
                    projected_date = None
            months = parse_projection_label_months(label)
            if projected_price is None:
                continue
            projected_value = quantize_decimal(position.shares * projected_price, "0.01")
            if projected_date and projected_date > baseline_snapshot.snapshot_date:
                months = max((projected_date.year - baseline_snapshot.snapshot_date.year) * 12 + (projected_date.month - baseline_snapshot.snapshot_date.month), 1)
                moderated_value, _ = moderate_tracking_target_value(
                    baseline_current_value,
                    projected_value,
                    months=months,
                    card=card,
                )
                anchors.append(
                    {
                        "date": projected_date,
                        "value": moderated_value if moderated_value is not None else projected_value,
                        "label": label or (f"{months // 12}A" if months % 12 == 0 else f"{months}M"),
                        "is_anchor": True,
                    }
                )
            else:
                add_anchor(months, projected_value, label=label)

    if purchase_baseline and not anchors:
        for year in TRACKING_FIVE_YEAR_MARKERS:
            projected_price = getattr(purchase_baseline, f"projected_price_{year}y", None)
            projected_return_pct = getattr(purchase_baseline, f"projected_return_pct_{year}y", None)
            projected_value = None
            if projected_price is not None:
                projected_value = quantize_decimal(position.shares * projected_price, "0.01")
            elif projected_return_pct is not None:
                projected_value = quantize_decimal(
                    baseline_current_value * (Decimal("1") + (projected_return_pct / ONE_HUNDRED)),
                    "0.01",
                )
            add_anchor(year * 12, projected_value, label=f"{year}A")

    if not anchors:
        cycle_projection = card.get("cycle_projection_5y") or {}
        path_rows = cycle_projection.get("path") or []
        for row in path_rows:
            months = parse_projection_label_months(row.get("label"))
            projected_price = row.get("projected_price")
            if projected_price is None:
                continue
            add_anchor(
                months,
                quantize_decimal(position.shares * projected_price, "0.01"),
                label=row.get("label"),
            )

        if not anchors:
            projected_price = cycle_projection.get("projected_price")
            if projected_price is not None:
                add_anchor(60, quantize_decimal(position.shares * projected_price, "0.01"), label="5A")
            else:
                five_year_return_pct = cycle_projection.get("five_year_return_pct")
                if five_year_return_pct is not None:
                    add_anchor(
                        60,
                        quantize_decimal(
                            baseline_current_value * (Decimal("1") + (five_year_return_pct / ONE_HUNDRED)),
                            "0.01",
                        ),
                        label="5A",
                    )

    series, latest_expected_value, projected_end_date = build_segmented_expected_series(snapshots, anchors)
    expected_total_value_5y = series[-1]["value"] if series else None
    return series, latest_expected_value, projected_end_date, expected_total_value_5y


def densify_projected_tracking_series(series: list[dict]) -> list[dict]:
    normalized_rows = {}
    for point in series or []:
        point_date = point.get("date")
        point_value = point.get("value")
        if point_date is None or point_value is None:
            continue
        normalized_rows[point_date] = {
            **point,
            "date": point_date,
            "value": quantize_decimal(Decimal(str(point_value)), "0.01") or ZERO,
        }

    normalized_series = [
        normalized_rows[point_date]
        for point_date in sorted(normalized_rows)
    ]
    if len(normalized_series) < 2:
        return normalized_series

    dense_series = [dict(normalized_series[0])]
    for previous_point, next_point in zip(normalized_series, normalized_series[1:]):
        previous_date = previous_point["date"]
        next_date = next_point["date"]
        previous_value = previous_point["value"]
        next_value = next_point["value"]
        if next_date <= previous_date:
            continue
        segment_days = max((next_date - previous_date).days, 1)
        for day_offset in range(1, segment_days):
            point_date = previous_date + timedelta(days=day_offset)
            interpolated_value = project_expected_value_on_date(
                previous_value,
                next_value,
                previous_date,
                point_date,
                horizon_days=segment_days,
            )
            if interpolated_value is None:
                continue
            dense_series.append(
                {
                    "date": point_date,
                    "value": quantize_decimal(interpolated_value, "0.01") or ZERO,
                }
            )
        dense_series.append(dict(next_point))
    return dense_series


def filter_ticket_tracking_snapshots(
    snapshots: list[EquityTicketSnapshot],
    tracking_anchor_date: date | None = None,
) -> list[EquityTicketSnapshot]:
    if not snapshots or tracking_anchor_date is None:
        return snapshots
    filtered = [snapshot for snapshot in snapshots if snapshot.snapshot_date >= tracking_anchor_date]
    return filtered or snapshots


def resolve_ticket_tracking_anchor_date(grouped_snapshots: dict[int, list[EquityTicketSnapshot]]) -> date | None:
    earliest_dates = [snapshots[0].snapshot_date for snapshots in grouped_snapshots.values() if snapshots]
    return max(earliest_dates) if earliest_dates else None


def build_normalized_reference_tracking_series(
    reference_points: list[dict],
    tracking_dates: list[date],
    baseline_value: Decimal,
) -> list[dict]:
    normalized_points = [
        {
            "date": point.get("date"),
            "close": Decimal(str(point.get("close"))),
        }
        for point in reference_points
        if point.get("date") is not None and point.get("close") is not None
    ]
    requested_dates = sorted({point_date for point_date in tracking_dates if point_date is not None})
    if not normalized_points or not requested_dates:
        return []

    normalized_points.sort(key=lambda item: item["date"])
    point_index = 0
    last_close = None
    baseline_reference_close = None
    series = []

    for point_date in requested_dates:
        while point_index < len(normalized_points) and normalized_points[point_index]["date"] <= point_date:
            last_close = normalized_points[point_index]["close"]
            point_index += 1

        if baseline_reference_close is None:
            baseline_reference_close = last_close
            if baseline_reference_close is None:
                future_point = next(
                    (point for point in normalized_points if point["date"] >= point_date),
                    None,
                )
                if future_point is None:
                    continue
                baseline_reference_close = future_point["close"]
                last_close = future_point["close"]

        current_close = last_close or baseline_reference_close
        if current_close is None or baseline_reference_close in {None, ZERO}:
            continue

        normalized_value = quantize_decimal(
            baseline_value * (current_close / baseline_reference_close),
            "0.01",
        ) or ZERO
        series.append({"date": point_date, "value": normalized_value})

    return series


def build_reference_close_tracking_series(
    reference_points: list[dict],
    tracking_dates: list[date],
) -> list[dict]:
    normalized_points = [
        {
            "date": point.get("date"),
            "close": Decimal(str(point.get("close"))),
        }
        for point in reference_points
        if point.get("date") is not None and point.get("close") is not None
    ]
    requested_dates = sorted({point_date for point_date in tracking_dates if point_date is not None})
    if not normalized_points or not requested_dates:
        return []

    normalized_points.sort(key=lambda item: item["date"])
    point_index = 0
    last_close = None
    series = []

    for point_date in requested_dates:
        while point_index < len(normalized_points) and normalized_points[point_index]["date"] <= point_date:
            last_close = normalized_points[point_index]["close"]
            point_index += 1

        current_close = last_close
        if current_close is None:
            future_point = next(
                (point for point in normalized_points if point["date"] >= point_date),
                None,
            )
            if future_point is None:
                continue
            current_close = future_point["close"]
            last_close = future_point["close"]

        series.append(
            {
                "date": point_date,
                "value": quantize_decimal(current_close, "0.01") or ZERO,
            }
        )

    return series


def build_tracking_benchmark_context(
    tracking_dates: list[date],
    baseline_value: Decimal,
) -> dict:
    if not tracking_dates:
        return {
            "available": False,
            "label": DEFAULT_BENCHMARK_NAME,
            "series": [],
            "latest_value": None,
            "actual_change_pct": None,
        }

    try:
        benchmark_series = fetch_reference_series_for_choice(
            EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol=DEFAULT_BENCHMARK_SYMBOL,
            benchmark_name=DEFAULT_BENCHMARK_NAME,
        )
    except Exception:
        benchmark_series = None

    if benchmark_series is None:
        return {
            "available": False,
            "label": DEFAULT_BENCHMARK_NAME,
            "series": [],
            "latest_value": None,
            "actual_change_pct": None,
        }

    series = build_normalized_reference_tracking_series(
        benchmark_series.points,
        tracking_dates,
        baseline_value,
    )
    close_series = build_reference_close_tracking_series(
        benchmark_series.points,
        tracking_dates,
    )
    latest_value = series[-1]["value"] if series else None
    return {
        "available": bool(series),
        "label": benchmark_series.name or DEFAULT_BENCHMARK_NAME,
        "series": series,
        "close_series": close_series,
        "latest_value": latest_value,
        "actual_change_pct": percentage_change(latest_value, series[0]["value"]) if series else None,
    }


def build_aggregated_tracking_benchmark_context(ticket_items: list[dict]) -> dict:
    if not ticket_items:
        return {
            "available": False,
            "label": DEFAULT_BENCHMARK_NAME,
            "series": [],
            "close_series": [],
            "latest_value": None,
            "actual_change_pct": None,
        }

    try:
        benchmark_series = fetch_reference_series_for_choice(
            EquityPosition.ReferenceProfile.MARKET_INDEX,
            benchmark_symbol=DEFAULT_BENCHMARK_SYMBOL,
            benchmark_name=DEFAULT_BENCHMARK_NAME,
        )
    except Exception:
        benchmark_series = None

    if benchmark_series is None:
        return {
            "available": False,
            "label": DEFAULT_BENCHMARK_NAME,
            "series": [],
            "close_series": [],
            "latest_value": None,
            "actual_change_pct": None,
        }

    tracking_dates = sorted(
        {
            point["date"]
            for item in ticket_items
            for point in item.get("actual_series", [])
            if point.get("date") is not None
        }
    )
    baseline_total = ZERO
    benchmark_items = []
    for item in ticket_items:
        baseline = item.get("baseline_snapshot")
        baseline_value = getattr(baseline, "current_value", None)
        if baseline_value in {None, ZERO}:
            continue
        baseline_total += baseline_value
        tracking_dates = [point["date"] for point in item.get("actual_series", []) if point.get("date") is not None]
        normalized_series = build_normalized_reference_tracking_series(
            benchmark_series.points,
            tracking_dates,
            baseline_value,
        )
        benchmark_items.append(
            {
                "baseline_snapshot": baseline,
                "benchmark_series": normalized_series,
            }
        )

    series = build_aggregated_ticket_series(benchmark_items, "benchmark_series")
    close_series = build_reference_close_tracking_series(
        benchmark_series.points,
        tracking_dates,
    )
    if not series:
        return {
            "available": False,
            "label": benchmark_series.name or DEFAULT_BENCHMARK_NAME,
            "series": [],
            "close_series": close_series,
            "latest_value": None,
            "actual_change_pct": None,
        }
    latest_value = series[-1]["value"] if series else None
    return {
        "available": bool(series),
        "label": benchmark_series.name or DEFAULT_BENCHMARK_NAME,
        "series": series,
        "close_series": close_series,
        "latest_value": latest_value,
        "actual_change_pct": percentage_change(latest_value, baseline_total) if baseline_total else None,
    }


def build_aggregated_series_entries(series_items: list[dict]) -> list[dict]:
    prepared_items = []
    all_dates = set()
    for item in series_items:
        baseline_date = item.get("baseline_date")
        normalized_series = {}
        for point in item.get("series", []):
            point_date = point.get("date")
            point_value = point.get("value")
            if point_date is None or point_value is None:
                continue
            normalized_series[point_date] = {
                **point,
                "date": point_date,
                "value": Decimal(str(point_value)),
            }
        if not normalized_series:
            continue
        sorted_series = sorted(
            (point_date, point["value"])
            for point_date, point in normalized_series.items()
        )
        all_dates.update(point_date for point_date, _ in sorted_series)
        prepared_items.append(
            {
                "baseline_date": baseline_date,
                "series": sorted_series,
                "series_meta": normalized_series,
                "cursor": 0,
                "current_value": None,
            }
        )

    if not all_dates or not prepared_items:
        return []

    aggregated = []
    for point_date in sorted(all_dates):
        total_value = ZERO
        has_value = False
        has_anchor = False
        labels = []
        for prepared in prepared_items:
            baseline_date = prepared["baseline_date"]
            if baseline_date is not None and point_date < baseline_date:
                continue
            series = prepared["series"]
            cursor = prepared["cursor"]
            while cursor < len(series) and series[cursor][0] <= point_date:
                prepared["current_value"] = series[cursor][1]
                cursor += 1
            prepared["cursor"] = cursor
            current_value = prepared["current_value"]
            if current_value is None:
                continue
            total_value += current_value
            has_value = True
            point_meta = prepared["series_meta"].get(point_date) or {}
            if point_meta.get("is_anchor"):
                has_anchor = True
                label = str(point_meta.get("label") or "").strip()
                if label and label not in labels:
                    labels.append(label)
        if has_value:
            aggregated.append(
                {
                    "date": point_date,
                    "value": quantize_decimal(total_value, "0.01") or ZERO,
                    "is_anchor": has_anchor,
                    "label": " / ".join(labels[:2]) if labels else None,
                }
            )
    return aggregated


def build_aggregated_ticket_series(ticket_items: list[dict], series_key: str) -> list[dict]:
    return build_aggregated_series_entries(
        [
            {
                "baseline_date": getattr(item.get("baseline_snapshot"), "snapshot_date", None),
                "series": item.get(series_key, []),
            }
            for item in ticket_items
        ]
    )


def build_aggregated_ticket_actual_series(ticket_items: list[dict]) -> list[dict]:
    return build_aggregated_ticket_series(ticket_items, "actual_series")


def build_global_ticket_daily_change_pct(ticket_items: list[dict]) -> Decimal | None:
    comparable_base = ZERO
    daily_change_value = ZERO
    has_comparable_history = False
    for item in ticket_items:
        actual_series = item.get("actual_series", [])
        latest_snapshot = item.get("latest_snapshot")
        if latest_snapshot is None:
            continue
        latest_value = getattr(latest_snapshot, "current_value", None) or ZERO
        previous_value = None
        if len(actual_series) >= 2:
            previous_value = actual_series[-2].get("value")
        if previous_value is not None:
            previous_value = Decimal(str(previous_value))
            comparable_base += previous_value
            daily_change_value += latest_value - previous_value
            has_comparable_history = True
        else:
            comparable_base += latest_value

    if not has_comparable_history or comparable_base <= ZERO:
        return None
    return quantize_decimal(
        percentage_change(comparable_base + daily_change_value, comparable_base),
        "0.01",
    )


def resolve_tracking_target_annual_return_pct() -> Decimal:
    raw_target = getattr(settings, "EQUITIES_TRACKING_TARGET_ANNUAL_RETURN_PCT", Decimal("10.00"))
    try:
        target = Decimal(str(raw_target))
    except Exception:
        target = Decimal("10.00")
    normalized_target = quantize_decimal(max(target, ZERO), "0.01")
    return normalized_target if normalized_target is not None else Decimal("10.00")


def project_tracking_target_value(
    baseline_value: Decimal | None,
    baseline_date: date | None,
    target_date: date | None,
    annual_return_pct: Decimal | None = None,
) -> Decimal | None:
    normalized_baseline = quantize_decimal(baseline_value, "0.01")
    if normalized_baseline is None:
        return None
    if normalized_baseline <= ZERO or baseline_date is None or target_date is None:
        return normalized_baseline

    elapsed_days = max((target_date - baseline_date).days, 0)
    target_pct = resolve_tracking_target_annual_return_pct() if annual_return_pct is None else annual_return_pct
    if elapsed_days <= 0 or target_pct <= ZERO:
        return normalized_baseline

    growth_base = max(0.01, 1 + (float(target_pct) / 100))
    target_multiplier = growth_base ** (elapsed_days / 365)
    return quantize_decimal(normalized_baseline * Decimal(str(target_multiplier)), "0.01") or normalized_baseline


def build_tracking_target_series(
    ticket_items: list[dict],
    annual_return_pct: Decimal | None = None,
) -> list[dict]:
    target_pct = resolve_tracking_target_annual_return_pct() if annual_return_pct is None else annual_return_pct
    tracking_dates = sorted(
        {
            point["date"]
            for item in ticket_items
            for point in item.get("actual_series", [])
            if point.get("date") is not None
        }
    )
    if not tracking_dates:
        return []

    target_items = []
    for item in ticket_items:
        baseline = item.get("baseline_snapshot")
        baseline_date = getattr(baseline, "snapshot_date", None)
        baseline_value = getattr(baseline, "current_value", None)
        if baseline_date is None or baseline_value in {None, ZERO}:
            continue
        series = []
        for point_date in tracking_dates:
            if point_date < baseline_date:
                continue
            target_value = project_tracking_target_value(
                baseline_value,
                baseline_date,
                point_date,
                target_pct,
            )
            if target_value is not None:
                series.append({"date": point_date, "value": target_value})
        if series:
            target_items.append({"baseline_date": baseline_date, "series": series})

    return build_aggregated_series_entries(target_items)


def build_tracking_objective_status(
    net_gain_value: Decimal | None,
    annualized_return_pct: Decimal | None,
    target_annual_return_pct: Decimal,
) -> dict:
    target_label = f"{target_annual_return_pct:.0f}"
    if net_gain_value is not None and net_gain_value < ZERO:
        return {
            "label": "Perdiendo dinero",
            "tone": "warn",
            "detail": "Primero volver a positivo; IBEX y escenarios solo son referencias secundarias.",
        }
    if annualized_return_pct is not None and annualized_return_pct >= target_annual_return_pct:
        return {
            "label": f"Cumple {target_label} % anual",
            "tone": "good",
            "detail": "La cartera va al ritmo objetivo o por encima.",
        }
    if net_gain_value is not None and net_gain_value >= ZERO:
        return {
            "label": "Gana, pero bajo objetivo",
            "tone": "warn",
            "detail": f"Hay beneficio, aunque todavia no llega al {target_label} % anual.",
        }
    return {
        "label": "Sin lectura completa",
        "tone": "warn",
        "detail": "Faltan datos suficientes para medir el objetivo anual.",
    }


def resolve_global_tracking_baseline_value_on_date(ticket_items: list[dict], point_date: date | None) -> Decimal:
    if point_date is None:
        return ZERO
    return quantize_decimal(
        sum(
            (
                getattr(item.get("baseline_snapshot"), "current_value", None) or ZERO
            )
            for item in ticket_items
            if getattr(item.get("baseline_snapshot"), "snapshot_date", None) is not None
            and item["baseline_snapshot"].snapshot_date <= point_date
        ),
        "0.01",
    ) or ZERO


def build_tracking_series_vs_dynamic_baseline(
    series: list[dict],
    ticket_items: list[dict],
    *,
    as_percentage: bool = False,
) -> list[dict]:
    transformed = []
    for point in series:
        point_date = point.get("date")
        point_value = point.get("value")
        if point_date is None or point_value is None:
            continue
        baseline_value = resolve_global_tracking_baseline_value_on_date(ticket_items, point_date)
        if baseline_value <= ZERO:
            continue
        point_value_decimal = Decimal(str(point_value))
        transformed_value = (
            percentage_change(point_value_decimal, baseline_value)
            if as_percentage
            else quantize_decimal(point_value_decimal - baseline_value, "0.01")
        )
        if transformed_value is None:
            continue
        transformed.append(
            {
                "date": point_date,
                "value": quantize_decimal(transformed_value, "0.01") or ZERO,
            }
        )
    return transformed


def build_tracking_series_vs_static_baseline(
    series: list[dict],
    baseline_value: Decimal | None = None,
    *,
    as_percentage: bool = False,
) -> list[dict]:
    normalized_points = normalize_tracking_series(series)
    if not normalized_points:
        return []

    resolved_baseline = baseline_value
    if resolved_baseline is None:
        resolved_baseline = normalized_points[0]["value"]
    if resolved_baseline is None:
        return []
    resolved_baseline = quantize_decimal(Decimal(str(resolved_baseline)), "0.01")
    if resolved_baseline is None or resolved_baseline <= ZERO:
        return []

    transformed = []
    for point in normalized_points:
        point_value = Decimal(str(point["value"]))
        transformed_value = (
            percentage_change(point_value, resolved_baseline)
            if as_percentage
            else quantize_decimal(point_value - resolved_baseline, "0.01")
        )
        if transformed_value is None:
            continue
        transformed.append(
            {
                **point,
                "value": quantize_decimal(transformed_value, "0.01") or ZERO,
            }
        )
    return transformed


def resolve_position_price_history_points(position: EquityPosition) -> list[EquityPriceHistory]:
    prefetched = getattr(position, "_prefetched_objects_cache", {}).get("price_history")
    if prefetched is not None:
        return sorted(prefetched, key=lambda point: point.price_date)
    return list(position.price_history.order_by("price_date"))


def build_position_market_value_history_series(
    position: EquityPosition,
    history_points: list[EquityPriceHistory],
    *,
    start_date: date,
    end_date: date,
) -> list[dict]:
    rows = []
    last_before_start = None
    for point in history_points:
        if point.price_date is None or point.close_price is None:
            continue
        if point.price_date < start_date:
            last_before_start = point
            continue
        if point.price_date > end_date:
            break
        rows.append(
            {
                "date": point.price_date,
                "value": quantize_decimal(position.shares * point.close_price, "0.01") or ZERO,
            }
        )

    if last_before_start is not None and (not rows or rows[0]["date"] > start_date):
        rows.insert(
            0,
            {
                "date": start_date,
                "value": quantize_decimal(position.shares * last_before_start.close_price, "0.01") or ZERO,
            },
        )

    return normalize_tracking_series(rows)


def build_position_benchmark_value_series(
    position: EquityPosition,
    history_points: list[EquityPriceHistory],
    *,
    start_date: date,
    end_date: date,
) -> list[dict]:
    rows = []
    last_before_start = None
    for point in history_points:
        if point.price_date is None or point.close_price is None or point.benchmark_close is None:
            continue
        if point.price_date < start_date:
            last_before_start = point
            continue
        if point.price_date > end_date:
            break
        rows.append(
            {
                "date": point.price_date,
                "close_price": Decimal(str(point.close_price)),
                "benchmark_close": Decimal(str(point.benchmark_close)),
            }
        )

    if last_before_start is not None and (not rows or rows[0]["date"] > start_date):
        rows.insert(
            0,
            {
                "date": start_date,
                "close_price": Decimal(str(last_before_start.close_price)),
                "benchmark_close": Decimal(str(last_before_start.benchmark_close)),
            },
        )

    if not rows:
        return []

    base_market_value = quantize_decimal(position.shares * rows[0]["close_price"], "0.01")
    base_benchmark_close = quantize_decimal(rows[0]["benchmark_close"], "0.0001")
    if base_market_value is None or base_market_value <= ZERO or base_benchmark_close is None or base_benchmark_close <= ZERO:
        return []

    benchmark_series = []
    for row in rows:
        benchmark_close = quantize_decimal(row["benchmark_close"], "0.0001")
        if benchmark_close is None or benchmark_close <= ZERO:
            continue
        benchmark_value = quantize_decimal(
            base_market_value * (benchmark_close / base_benchmark_close),
            "0.01",
        )
        if benchmark_value is None:
            continue
        benchmark_series.append({"date": row["date"], "value": benchmark_value})

    return normalize_tracking_series(benchmark_series)


def build_position_projection_market_value_series(
    card: dict,
    actual_series: list[dict],
    *,
    horizon: str,
) -> list[dict]:
    actual_points = normalize_tracking_series(actual_series)
    if not actual_points:
        return []

    latest_actual = actual_points[-1]
    position = card["position"]
    projection_series = [
        {
            "date": latest_actual["date"],
            "value": latest_actual["value"],
            "is_anchor": True,
            "label": "Hoy",
        }
    ]
    if horizon == "12m":
        projection_path = resolve_projection_tracking_path(card.get("projection"))
    else:
        projection_path = (card.get("cycle_projection_5y") or {}).get("path") or []

    for index, step in enumerate(projection_path, start=1):
        projected_date = step.get("projected_date")
        projected_price = step.get("projected_price")
        if projected_date is None or projected_price is None or projected_date <= latest_actual["date"]:
            continue
        projection_series.append(
            {
                "date": projected_date,
                "value": quantize_decimal(position.shares * Decimal(str(projected_price)), "0.01") or ZERO,
                "is_anchor": True,
                "label": str(step.get("label") or (f"{index}M" if horizon == "12m" else f"{index}A")).strip(),
            }
        )

    if len(projection_series) == 1:
        if horizon == "12m":
            fallback_offsets = (("1A", 365),)
        else:
            fallback_offsets = tuple((f"{year}A", 365 * year) for year in range(1, 6))
        for label, offset_days in fallback_offsets:
            projection_series.append(
                {
                    "date": latest_actual["date"] + timedelta(days=offset_days),
                    "value": latest_actual["value"],
                    "is_anchor": True,
                    "label": label,
                }
            )

    return densify_projected_tracking_series(projection_series)


def build_portfolio_weight_mix_label(history_cards: list[dict], max_items: int = 3) -> str:
    weighted_rows = []
    total_value = ZERO
    for card in history_cards:
        position = card["position"]
        current_value = quantize_decimal(position.current_value, "0.01") or ZERO
        if current_value <= ZERO:
            continue
        total_value += current_value
        weighted_rows.append(
            {
                "ticker": position.ticker,
                "company_name": position.company_name,
                "value": current_value,
            }
        )

    if total_value <= ZERO or not weighted_rows:
        return ""

    labels = []
    for row in sorted(weighted_rows, key=lambda item: (-item["value"], item["company_name"]))[:max_items]:
        weight_pct = quantize_decimal((row["value"] / total_value) * ONE_HUNDRED, "0.1") or ZERO
        labels.append(f"{row['ticker']} {weight_pct}%")
    return f"Peso actual: {', '.join(labels)}"


def build_portfolio_summary_horizon_context(
    history_cards: list[dict],
    *,
    horizon: str,
) -> dict:
    horizon_days = TRACKING_HORIZON_DAYS if horizon == "12m" else TRACKING_HORIZON_DAYS * 5
    cards_with_history = []
    latest_dates = []
    history_by_position = {}

    for card in history_cards:
        position = card["position"]
        history_points = resolve_position_price_history_points(position)
        if len(history_points) < 2:
            continue
        history_by_position[position.id or id(position)] = history_points
        latest_dates.append(history_points[-1].price_date)
        cards_with_history.append(card)

    if not cards_with_history or not latest_dates:
        return {"available": False}

    history_end = max(latest_dates)
    requested_start = history_end - timedelta(days=horizon_days)
    history_start = max(
        next(
            (point.price_date for point in history_by_position[card["position"].id or id(card["position"])] if point.price_date >= requested_start),
            history_by_position[card["position"].id or id(card["position"])][0].price_date,
        )
        for card in cards_with_history
    )

    actual_entries = []
    expected_entries = []
    benchmark_entries = []
    for card in cards_with_history:
        position = card["position"]
        history_points = history_by_position[position.id or id(position)]
        actual_series = build_position_market_value_history_series(
            position,
            history_points,
            start_date=history_start,
            end_date=history_end,
        )
        if not actual_series:
            continue
        actual_entries.append(
            {
                "baseline_date": actual_series[0]["date"],
                "series": actual_series,
            }
        )
        expected_series = build_position_projection_market_value_series(
            card,
            actual_series,
            horizon=horizon,
        )
        if expected_series:
            expected_entries.append(
                {
                    "baseline_date": expected_series[0]["date"],
                    "series": expected_series,
                }
            )
        benchmark_series = build_position_benchmark_value_series(
            position,
            history_points,
            start_date=history_start,
            end_date=history_end,
        )
        if benchmark_series:
            benchmark_entries.append(
                {
                    "baseline_date": benchmark_series[0]["date"],
                    "series": benchmark_series,
                }
            )

    actual_series = build_aggregated_series_entries(actual_entries)
    expected_series = build_aggregated_series_entries(expected_entries)
    benchmark_series = build_aggregated_series_entries(benchmark_entries)
    if len(actual_series) < 2:
        return {"available": False}

    baseline_value = actual_series[0]["value"]
    net_series = build_tracking_series_vs_static_baseline(actual_series, baseline_value)
    net_expected_series = build_tracking_series_vs_static_baseline(expected_series, baseline_value)
    return_series = build_tracking_series_vs_static_baseline(
        actual_series,
        baseline_value,
        as_percentage=True,
    )
    return_expected_series = build_tracking_series_vs_static_baseline(
        expected_series,
        baseline_value,
        as_percentage=True,
    )
    return_benchmark_series = build_tracking_series_vs_static_baseline(
        benchmark_series,
        baseline_value,
        as_percentage=True,
    )
    net_chart = build_value_tracking_chart(
        net_series,
        net_expected_series,
        time_marker_mode="month",
        grid_marker_mode="month",
    )
    return_chart = build_value_tracking_chart(
        return_series,
        return_expected_series,
        benchmark_series=return_benchmark_series,
        value_suffix="%",
        axis_formatter=format_percentage_axis_value,
        time_marker_mode="month",
        grid_marker_mode="month",
    )
    return {
        "available": bool(net_chart.get("available") or return_chart.get("available")),
        "history_start": actual_series[0]["date"],
        "history_end": actual_series[-1]["date"],
        "projection_end": expected_series[-1]["date"] if expected_series else actual_series[-1]["date"],
        "range_label": f"Historico comun desde {actual_series[0]['date'].isoformat()}",
        "baseline_value": baseline_value,
        "actual_series": actual_series,
        "expected_series": expected_series,
        "benchmark_series": benchmark_series,
        "net_series": net_series,
        "net_expected_series": net_expected_series,
        "return_series": return_series,
        "return_expected_series": return_expected_series,
        "return_benchmark_series": return_benchmark_series,
        "net_chart": net_chart,
        "return_chart": return_chart,
    }


def build_portfolio_summary_context(
    history_cards: list[dict],
    *,
    benchmark_label: str | None = None,
) -> dict:
    owned_cards = [card for card in history_cards if card["position"].is_owned]
    if not owned_cards:
        return {"available": False}

    horizon_12m = build_portfolio_summary_horizon_context(owned_cards, horizon="12m")
    horizon_5y = build_portfolio_summary_horizon_context(owned_cards, horizon="5y")
    return {
        "available": bool(horizon_12m.get("available") or horizon_5y.get("available")),
        "benchmark_label": benchmark_label or DEFAULT_BENCHMARK_NAME,
        "weight_mix_label": build_portfolio_weight_mix_label(owned_cards),
        "actual_series_12m": horizon_12m.get("actual_series", []),
        "expected_series_12m": horizon_12m.get("expected_series", []),
        "actual_series_5y": horizon_5y.get("actual_series", []),
        "expected_series_5y": horizon_5y.get("expected_series", []),
        "net_chart_12m": horizon_12m.get("net_chart", {"available": False}),
        "net_chart_5y": horizon_5y.get("net_chart", {"available": False}),
        "return_chart_12m": horizon_12m.get("return_chart", {"available": False}),
        "return_chart_5y": horizon_5y.get("return_chart", {"available": False}),
        "range_label_12m": horizon_12m.get("range_label", ""),
        "range_label_5y": horizon_5y.get("range_label", ""),
    }


def build_tracking_rebased_comparison_series(
    actual_series: list[dict],
    invested_series: list[dict],
    benchmark_series: list[dict],
) -> dict:
    actual_points = normalize_tracking_series(actual_series)
    invested_points = normalize_tracking_series(invested_series)
    benchmark_points = normalize_tracking_series(benchmark_series)
    if len(actual_points) < 2 or len(benchmark_points) < 2:
        return {
            "available": False,
            "portfolio_series": [],
            "benchmark_series": [],
            "rebase_dates": [],
            "capital_change_dates": [],
            "latest_gap_value": None,
            "latest_gap_pct": None,
        }

    invested_by_date = {point["date"]: point["value"] for point in invested_points}
    benchmark_by_date = {point["date"]: point["value"] for point in benchmark_points}
    current_invested = None
    previous_invested = None
    current_benchmark = None
    scale_factor = None
    rebased_portfolio_series = []
    aligned_benchmark_series = []
    rebase_dates = []

    for point in actual_points:
        point_date = point["date"]
        point_value = point["value"]
        if point_date in benchmark_by_date:
            current_benchmark = benchmark_by_date[point_date]
        if point_date in invested_by_date:
            current_invested = invested_by_date[point_date]
        if current_benchmark in {None, ZERO} or point_value in {None, ZERO}:
            continue

        should_rebase = scale_factor is None
        if previous_invested is not None and current_invested is not None and current_invested != previous_invested:
            should_rebase = True

        if should_rebase:
            scale_factor = current_benchmark / point_value
            rebase_dates.append(point_date)

        scaled_value = quantize_decimal(point_value * scale_factor, "0.01")
        if scaled_value is None:
            continue

        rebased_portfolio_series.append({"date": point_date, "value": scaled_value})
        aligned_benchmark_series.append(
            {"date": point_date, "value": quantize_decimal(current_benchmark, "0.01") or ZERO}
        )
        previous_invested = current_invested

    if len(rebased_portfolio_series) < 2 or len(aligned_benchmark_series) < 2:
        return {
            "available": False,
            "portfolio_series": rebased_portfolio_series,
            "benchmark_series": aligned_benchmark_series,
            "rebase_dates": rebase_dates,
            "capital_change_dates": rebase_dates[1:],
            "latest_gap_value": None,
            "latest_gap_pct": None,
        }

    latest_gap_value = quantize_decimal(
        rebased_portfolio_series[-1]["value"] - aligned_benchmark_series[-1]["value"],
        "0.01",
    )
    return {
        "available": True,
        "portfolio_series": rebased_portfolio_series,
        "benchmark_series": aligned_benchmark_series,
        "rebase_dates": rebase_dates,
        "capital_change_dates": rebase_dates[1:],
        "latest_gap_value": latest_gap_value,
        "latest_gap_pct": quantize_decimal(
            percentage_change(rebased_portfolio_series[-1]["value"], aligned_benchmark_series[-1]["value"]),
            "0.01",
        ),
    }


def resolve_tracking_trade_plan_projected_sale_price(
    position: EquityPosition,
    trade_plan: dict,
) -> Decimal | None:
    sale_month_number = trade_plan.get("sale_month_number")
    for row in trade_plan.get("monthly_rows") or []:
        if row.get("month_number") != sale_month_number:
            continue
        projected_price = row.get("projected_price")
        if projected_price not in {None, ZERO}:
            return quantize_decimal(Decimal(str(projected_price)), "0.0001")

    current_price = quantize_decimal(getattr(position, "current_price_per_share", None), "0.0001")
    if current_price in {None, ZERO}:
        return None

    pre_sale_return_pct = trade_plan.get("pre_sale_return_pct")
    if pre_sale_return_pct is None:
        return current_price
    return quantize_decimal(
        current_price * (Decimal("1") + (Decimal(str(pre_sale_return_pct)) / ONE_HUNDRED)),
        "0.0001",
    )


def build_tracking_sale_timeline_context(
    ticket_items: list[dict],
    *,
    horizon_months: int = 24,
) -> dict:
    if not ticket_items:
        return {"available": False}

    latest_snapshot_dates = [
        getattr(item.get("latest_snapshot"), "snapshot_date", None)
        for item in ticket_items
        if item.get("latest_snapshot") is not None
    ]
    reference_date = max((point_date for point_date in latest_snapshot_dates if point_date is not None), default=None)
    if reference_date is None:
        reference_date = django_timezone.localdate()
    horizon_end = add_calendar_months(reference_date, horizon_months) or reference_date
    total_days = max((horizon_end - reference_date).days, 1)

    def clamp_pct(value: float) -> str:
        return f"{max(0.0, min(value, 100.0)):.2f}"

    markers = []
    for month_offset in range(0, horizon_months + 1, 3):
        marker_date = add_calendar_months(reference_date, month_offset)
        if marker_date is None:
            continue
        marker_days = max((marker_date - reference_date).days, 0)
        marker_left_pct = (marker_days / total_days) * 100
        month_label = SPANISH_MONTH_LABELS.get(marker_date.month, str(marker_date.month))
        markers.append(
            {
                "left_pct": clamp_pct(marker_left_pct),
                "label": f"{month_label[:3].capitalize()} {str(marker_date.year)[2:]}",
            }
        )

    rows = []
    unscheduled_rows = []
    alert_rows = []

    for item in ticket_items:
        position = item.get("position")
        trade_plan = item.get("trade_plan") or {}
        if position is None:
            continue

        sale_date = trade_plan.get("sale_date")
        if not trade_plan.get("available") or trade_plan.get("mode") not in {"sale_reentry", "sale_review"} or sale_date is None:
            unscheduled_rows.append(
                {
                    "ticker": position.ticker,
                    "company_name": position.company_name,
                    "reason": "Sin venta clara en 24M" if trade_plan.get("mode") == "hold" else "Sin ventana tactica cerrada",
                }
            )
            continue

        projected_sale_price = resolve_tracking_trade_plan_projected_sale_price(position, trade_plan)
        sale_preview = build_equity_sale_preview(
            position,
            sale_price_per_share=projected_sale_price,
            closed_on=sale_date,
        )
        estimated_net_result = sale_preview.get("net_result") if sale_preview.get("available") else None
        days_until_sale = (sale_date - reference_date).days
        if days_until_sale <= 0:
            status_key = "urgent"
            status_label = "Vender ya"
        elif days_until_sale <= 30:
            status_key = "soon"
            status_label = "Venta <= 30 dias"
        elif days_until_sale <= 90:
            status_key = "watch"
            status_label = "Preparar venta"
        else:
            status_key = "scheduled"
            status_label = "Venta programada"

        if sale_date.year == reference_date.year and sale_date.month == reference_date.month and sale_date < reference_date:
            window_start = reference_date
            window_end = min(horizon_end, reference_date + timedelta(days=10))
        else:
            window_start = sale_date.replace(day=1)
            window_end = sale_date.replace(day=monthrange(sale_date.year, sale_date.month)[1])

        if window_start > horizon_end:
            unscheduled_rows.append(
                {
                    "ticker": position.ticker,
                    "company_name": position.company_name,
                    "reason": "Ventana fuera del horizonte 24M",
                }
            )
            continue

        clamped_start = max(window_start, reference_date)
        clamped_end = min(window_end, horizon_end)
        if clamped_end < reference_date:
            clamped_start = reference_date
            clamped_end = min(horizon_end, reference_date + timedelta(days=10))

        start_days = max((clamped_start - reference_date).days, 0)
        end_days = max((clamped_end - reference_date).days, start_days + 1)
        sale_days = max((sale_date - reference_date).days, 0)

        left_pct = (start_days / total_days) * 100
        width_pct = max(((end_days - start_days + 1) / total_days) * 100, 2.8)
        pin_left_pct = (sale_days / total_days) * 100

        row = {
            "ticker": position.ticker,
            "company_name": position.company_name,
            "sale_date": sale_date,
            "sale_window_label": trade_plan.get("sale_window_label") or sale_date.isoformat(),
            "status_key": status_key,
            "status_label": status_label,
            "days_until_sale": days_until_sale,
            "bar_left_pct": clamp_pct(left_pct),
            "bar_width_pct": clamp_pct(min(left_pct + width_pct, 100.0) - left_pct),
            "pin_left_pct": clamp_pct(pin_left_pct),
            "estimated_net_result": quantize_decimal(estimated_net_result, "0.01") if estimated_net_result is not None else None,
            "projected_sale_price": projected_sale_price,
            "sale_mode_label": "Venta y reentrada" if trade_plan.get("mode") == "sale_reentry" else "Venta y revision",
        }
        rows.append(row)
        if days_until_sale <= 90:
            alert_rows.append(row)

    rows.sort(key=lambda row: (row.get("sale_date") or date.max, row["company_name"]))
    alert_rows.sort(key=lambda row: (row.get("sale_date") or date.max, row["company_name"]))
    unscheduled_rows.sort(key=lambda row: row["company_name"])

    if not rows and not unscheduled_rows:
        return {"available": False}

    projected_net_result_total = quantize_decimal(
        sum(((row.get("estimated_net_result") or ZERO) for row in rows), ZERO),
        "0.01",
    ) or ZERO
    next_row = rows[0] if rows else None
    return {
        "available": True,
        "horizon_months": horizon_months,
        "reference_date": reference_date,
        "horizon_end": horizon_end,
        "markers": markers,
        "rows": rows,
        "alert_rows": alert_rows[:4],
        "unscheduled_rows": unscheduled_rows[:6],
        "next_row": next_row,
        "scheduled_count": len(rows),
        "alert_count": len(alert_rows),
        "unscheduled_count": len(unscheduled_rows),
        "projected_net_result_total": projected_net_result_total,
    }


def add_calendar_years(value: date | None, years: int) -> date | None:
    if value is None:
        return None
    try:
        return value.replace(year=value.year + years)
    except ValueError:
        return value.replace(month=2, day=28, year=value.year + years)


def add_calendar_months(value: date | None, months: int) -> date | None:
    if value is None:
        return None
    month_index = (value.month - 1) + months
    year = value.year + (month_index // 12)
    month = (month_index % 12) + 1
    day = min(value.day, monthrange(year, month)[1])
    return value.replace(year=year, month=month, day=day)


def parse_projection_label_months(label: str | None) -> int | None:
    normalized = str(label or "").strip().upper()
    if not normalized:
        return None
    year_match = re.fullmatch(r"(\d+)A", normalized)
    if year_match:
        return int(year_match.group(1)) * 12
    month_match = re.fullmatch(r"(\d+)M", normalized)
    if month_match:
        return int(month_match.group(1))
    return None


def format_projection_month_window(projected_date: date | None, month_number: int | None) -> str:
    if projected_date:
        month_label = SPANISH_MONTH_LABELS.get(projected_date.month, str(projected_date.month))
        if month_number:
            return f"{month_label} {projected_date.year} (mes {month_number})"
        return f"{month_label} {projected_date.year}"
    if month_number:
        return f"Mes {month_number}"
    return ""


def build_purchase_baseline_yearly_rows(
    purchase_baseline: EquityPurchaseForecastBaseline | None,
) -> list[dict]:
    if purchase_baseline is None or purchase_baseline.baseline_date is None:
        return []

    rows = []
    previous_price = purchase_baseline.baseline_price
    previous_cumulative_return = ZERO

    for year_number in range(1, 6):
        projected_price = getattr(purchase_baseline, f"projected_price_{year_number}y", None)
        cumulative_return_pct = getattr(
            purchase_baseline,
            f"projected_return_pct_{year_number}y",
            None,
        )
        if projected_price is None and cumulative_return_pct is None:
            continue

        margin_pct = None
        if projected_price not in {None, ZERO} and previous_price not in {None, ZERO}:
            if year_number == 1 and cumulative_return_pct is not None:
                margin_pct = cumulative_return_pct
            else:
                margin_pct = percentage_change(projected_price, previous_price)
        elif cumulative_return_pct is not None:
            if year_number == 1:
                margin_pct = cumulative_return_pct
            else:
                margin_pct = quantize_decimal(cumulative_return_pct - previous_cumulative_return)

        rows.append(
            {
                "year_number": year_number,
                "projected_date": add_calendar_years(purchase_baseline.baseline_date, year_number),
                "reentry_date": add_calendar_years(purchase_baseline.baseline_date, year_number - 1),
                "projected_price": projected_price,
                "cumulative_return_pct": cumulative_return_pct,
                "margin_pct": quantize_decimal(margin_pct),
            }
        )
        if projected_price is not None:
            previous_price = projected_price
        if cumulative_return_pct is not None:
            previous_cumulative_return = cumulative_return_pct

    return rows


def build_purchase_forecast_trade_plan(
    purchase_baseline: EquityPurchaseForecastBaseline | None,
) -> dict:
    yearly_rows = build_purchase_baseline_yearly_rows(purchase_baseline)
    if not yearly_rows:
        return {"available": False}

    best_row = max(
        yearly_rows,
        key=lambda item: (
            item["cumulative_return_pct"]
            if item["cumulative_return_pct"] is not None
            else Decimal("-9999"),
            -item["year_number"],
        ),
    )
    sale_row = None
    drawdown_row = None
    for row in yearly_rows:
        if row["year_number"] <= 1:
            continue
        margin_pct = row.get("margin_pct")
        previous_row = next(
            (item for item in yearly_rows if item["year_number"] == row["year_number"] - 1),
            None,
        )
        previous_return_pct = previous_row.get("cumulative_return_pct") if previous_row else None
        if (
            margin_pct is not None
            and margin_pct <= -OPTIMIZER_MAX_ROUNDTRIP_DRAG_PCT
            and previous_row is not None
            and previous_return_pct is not None
            and previous_return_pct >= OPTIMIZER_MAX_ENTRY_DRAG_PCT
        ):
            sale_row = previous_row
            drawdown_row = row
            break

    if sale_row is None:
        sale_window_label = (
            f"Cierre A\u00d1O {best_row['year_number']}"
            if best_row.get("year_number")
            else "Mantener"
        )
        return {
            "available": True,
            "mode": "hold",
            "sale_year_number": best_row["year_number"],
            "sale_date": best_row["projected_date"],
            "sale_date_label": "",
            "sale_window_label": sale_window_label,
            "reentry_year_number": None,
            "reentry_date": None,
            "reentry_date_label": "",
            "reentry_window_label": "",
            "summary": (
                "La foto de compra no muestra un retroceso anual lo bastante fuerte como para justificar "
                f"una salida tactica. Mantener hasta {sale_window_label.lower()} sigue siendo la lectura base. "
                "La recomendacion se expresa en ventanas anuales, no en dias exactos."
            )
            if best_row.get("year_number")
            else (
                "La foto de compra no muestra un retroceso anual fuerte y la lectura base sigue siendo mantener. "
                "La recomendacion se expresa en ventanas anuales, no en dias exactos."
            ),
            "yearly_rows": yearly_rows,
            "drawdown_year_number": None,
            "drawdown_margin_pct": None,
        }

    reentry_row = next(
        (
            item
            for item in yearly_rows
            if item["year_number"] > drawdown_row["year_number"]
            and item.get("margin_pct") is not None
            and item["margin_pct"] >= OPTIMIZER_MAX_ENTRY_DRAG_PCT
        ),
        None,
    )
    sale_window_label = f"Cierre A\u00d1O {sale_row['year_number']}"
    reentry_date = reentry_row.get("reentry_date") if reentry_row else None
    reentry_window_label = f"Inicio A\u00d1O {reentry_row['year_number']}" if reentry_row else ""
    if reentry_row and reentry_window_label:
        summary = (
            f"Se propone salir en {sale_window_label.lower()} antes de una caida prevista "
            f"del {drawdown_row['margin_pct']:.1f} %. Si la tesis sigue viva, la reentrada sugerida "
            f"se revisa en {reentry_window_label.lower()}. La recomendacion se expresa en ventanas anuales, "
            "no en dias exactos."
        )
    else:
        summary = (
            f"Se propone salir en {sale_window_label.lower()} antes de una caida prevista "
            f"del {drawdown_row['margin_pct']:.1f} %. Despues conviene revisar de nuevo el radar antes de volver. "
            "La recomendacion se expresa en ventanas anuales, no en dias exactos."
        )

    return {
        "available": True,
        "mode": "sale_reentry" if reentry_row else "sale_review",
        "sale_year_number": sale_row["year_number"],
        "sale_date": sale_row["projected_date"],
        "sale_date_label": "",
        "sale_window_label": sale_window_label,
        "reentry_year_number": reentry_row["year_number"] if reentry_row else None,
        "reentry_date": reentry_date,
        "reentry_date_label": "",
        "reentry_window_label": reentry_window_label,
        "summary": summary,
        "yearly_rows": yearly_rows,
        "drawdown_year_number": drawdown_row["year_number"] if drawdown_row else None,
        "drawdown_margin_pct": drawdown_row["margin_pct"] if drawdown_row else None,
    }


def build_interpolated_projection_monthly_rows(
    anchor_points: list[tuple[int, Decimal, date]],
) -> list[dict]:
    if len(anchor_points) < 2:
        return []

    monthly_rows = []
    for (start_month, start_price, start_date), (end_month, end_price, _) in zip(anchor_points, anchor_points[1:]):
        if start_price in {None, ZERO} or end_price in {None, ZERO} or end_month <= start_month:
            continue
        for month_number in range(start_month, end_month):
            if month_number < start_month:
                continue
            fraction = Decimal(str(month_number - start_month)) / Decimal(str(end_month - start_month))
            price_ratio = float(Decimal(str(end_price)) / Decimal(str(start_price)))
            interpolated_price = Decimal(str(round(float(start_price) * (price_ratio ** float(fraction)), 4)))
            monthly_rows.append(
                {
                    "month_number": month_number,
                    "projected_date": add_calendar_months(start_date, month_number - start_month),
                    "projected_price": quantize_decimal(interpolated_price, "0.0001"),
                }
            )

    final_month, final_price, final_date = anchor_points[-1]
    monthly_rows.append(
        {
            "month_number": final_month,
            "projected_date": final_date,
            "projected_price": quantize_decimal(final_price, "0.0001"),
        }
    )

    monthly_rows.sort(key=lambda item: item["month_number"])
    deduped_rows = []
    seen_months = set()
    for row in monthly_rows:
        if row["month_number"] in seen_months:
            continue
        seen_months.add(row["month_number"])
        deduped_rows.append(row)

    for index, row in enumerate(deduped_rows):
        if index < 2 or index > len(deduped_rows) - 3:
            row["smoothed_price"] = None
            row["smoothed_slope_value"] = None
            row["smoothed_slope_pct"] = None
            continue
        smoothed_price = average_decimal(
            [item["projected_price"] for item in deduped_rows[index - 2:index + 3] if item.get("projected_price") is not None]
        )
        row["smoothed_price"] = quantize_decimal(smoothed_price, "0.0001")
        previous_smoothed = deduped_rows[index - 1].get("smoothed_price")
        if previous_smoothed in {None, ZERO} or row["smoothed_price"] is None:
            row["smoothed_slope_value"] = None
            row["smoothed_slope_pct"] = None
        else:
            row["smoothed_slope_value"] = quantize_decimal(row["smoothed_price"] - previous_smoothed, "0.0001")
            row["smoothed_slope_pct"] = quantize_decimal(
                percentage_change(row["smoothed_price"], previous_smoothed),
                "0.01",
            )

    return deduped_rows


def build_projection_monthly_trend_rows(card: dict) -> list[dict]:
    position = card.get("position")
    projection = card.get("projection") or {}
    if position is None or not projection.get("available"):
        return []

    current_price = quantize_decimal(
        projection.get("latest_price") or getattr(position, "current_price_per_share", None),
        "0.0001",
    )
    path = resolve_projection_tracking_path(projection)
    if current_price in {None, ZERO} or not path:
        return []

    anchor_date = (
        projection.get("latest_date")
        or getattr(position, "latest_price_date", None)
        or django_timezone.localdate()
    )
    anchor_points = [(0, current_price, anchor_date)]
    for index, step in enumerate(path, start=1):
        projected_price = quantize_decimal(step.get("projected_price"), "0.0001")
        months_offset = parse_projection_label_months(step.get("label"))
        projected_date = step.get("projected_date")
        if months_offset is None and projected_date is not None:
            months_offset = max(
                ((projected_date.year - anchor_date.year) * 12)
                + (projected_date.month - anchor_date.month),
                index,
            )
        if months_offset is None:
            months_offset = index
        if projected_price in {None, ZERO} or months_offset <= 0:
            continue
        anchor_points.append(
            (
                months_offset,
                projected_price,
                projected_date or add_calendar_months(anchor_date, months_offset),
            )
        )
    return build_interpolated_projection_monthly_rows(anchor_points)


def build_candidate_purchase_timing_plan(
    card: dict,
    strategy_mode: str = "12m_primary",
) -> dict:
    position = card.get("position")
    projection = card.get("projection") or {}
    cycle_projection = card.get("cycle_projection_5y") or {}
    if position is None or not projection.get("available"):
        return {"available": False}

    strategy = get_optimizer_strategy_config(strategy_mode)
    twelve_month_rows = build_projection_monthly_trend_rows(card)
    cycle_rows = build_cycle_projection_monthly_trend_rows(card)

    path_candidates = []
    if twelve_month_rows:
        twelve_month_max_month = max((int(item.get("month_number") or 0) for item in twelve_month_rows), default=0)
        path_candidates.append(
            {
                "basis_key": "projection_12m",
                "basis_label": "Senda 12M neta del modelo",
                "rows": twelve_month_rows,
                "entry_horizon_months": min(12, max(twelve_month_max_month, 1)),
                "exit_horizon_months": max(twelve_month_max_month, 1),
            }
        )
    if cycle_rows:
        cycle_max_month = max((int(item.get("month_number") or 0) for item in cycle_rows), default=0)
        cycle_basis_label = "Patron de ciclo 5A desestacionalizado"
        if build_projection_monthly_trend_rows(card):
            cycle_basis_label = "Patron de ciclo 5A con entrada afinada por la senda 12M"
        path_candidates.append(
            {
                "basis_key": "cycle_5y",
                "basis_label": cycle_basis_label,
                "rows": cycle_rows,
                "entry_horizon_months": min(12, max(cycle_max_month, 1)),
                "exit_horizon_months": max(cycle_max_month, 1),
            }
        )
    if not path_candidates:
        return {"available": False}

    net_income_yield_pct = quantize_decimal(projection.get("net_income_yield_pct"), "0.01") or ZERO
    transaction_drag_pct = quantize_decimal(projection.get("transaction_drag_pct"), "0.01") or ZERO
    target_annualized_return_pct = OPTIMIZER_TARGET_HOLDING_ANNUALIZED_RETURN_PCT

    def interval_meets_target(interval: dict | None) -> bool:
        if not interval:
            return False
        annualized_return_pct = quantize_decimal(interval.get("holding_annualized_return_pct"), "0.01")
        return annualized_return_pct is not None and annualized_return_pct >= target_annualized_return_pct

    def exit_preference_key(interval: dict) -> tuple:
        trade_return_pct = quantize_decimal(interval.get("trade_return_pct"), "0.01")
        holding_annualized_return_pct = quantize_decimal(interval.get("holding_annualized_return_pct"), "0.01")
        exit_row = interval.get("exit_row") or {}
        exit_month_number = int(interval.get("exit_month_number") or exit_row.get("month_number") or 0)
        holding_months = int(interval.get("holding_months") or 0)
        if interval_meets_target(interval):
            primary_metric = trade_return_pct if trade_return_pct is not None else Decimal("-9999")
            secondary_metric = (
                holding_annualized_return_pct
                if holding_annualized_return_pct is not None
                else Decimal("-9999")
            )
        else:
            primary_metric = (
                holding_annualized_return_pct
                if holding_annualized_return_pct is not None
                else Decimal("-9999")
            )
            secondary_metric = trade_return_pct if trade_return_pct is not None else Decimal("-9999")
        return (
            Decimal("1") if interval_meets_target(interval) else ZERO,
            primary_metric,
            secondary_metric,
            exit_month_number,
            holding_months,
        )

    def entry_preference_key(interval: dict) -> tuple:
        trade_return_pct = quantize_decimal(interval.get("trade_return_pct"), "0.01")
        holding_annualized_return_pct = quantize_decimal(interval.get("holding_annualized_return_pct"), "0.01")
        calendar_adjusted_return_pct = quantize_decimal(interval.get("calendar_adjusted_return_pct"), "0.01")
        discount_vs_now_pct = quantize_decimal(interval.get("discount_vs_now_pct"), "0.01")
        exit_row = interval.get("exit_row") or {}
        exit_month_number = int(interval.get("exit_month_number") or exit_row.get("month_number") or 0)
        entry_month_number = int(interval.get("entry_month_number") or 0)
        holding_months = int(interval.get("holding_months") or 0)
        return (
            Decimal("1") if interval_meets_target(interval) else ZERO,
            calendar_adjusted_return_pct if calendar_adjusted_return_pct is not None else Decimal("-9999"),
            holding_annualized_return_pct if holding_annualized_return_pct is not None else Decimal("-9999"),
            trade_return_pct if trade_return_pct is not None else Decimal("-9999"),
            -(entry_month_number or 0),
            discount_vs_now_pct if discount_vs_now_pct is not None else Decimal("-9999"),
            exit_month_number,
            holding_months,
        )

    def evaluate_path(candidate_path: dict) -> dict | None:
        rows = candidate_path["rows"]
        if len(rows) < 2:
            return None
        current_row = rows[0]
        current_price = current_row.get("projected_price")
        if current_price in {None, ZERO}:
            return None
        entry_horizon_months = int(candidate_path.get("entry_horizon_months") or 12)
        exit_horizon_months = int(candidate_path.get("exit_horizon_months") or entry_horizon_months)
        entry_candidates = []
        for entry_row in rows:
            entry_month_number = int(entry_row.get("month_number") or 0)
            entry_price = entry_row.get("projected_price")
            if entry_price in {None, ZERO}:
                continue
            if entry_month_number > entry_horizon_months or entry_month_number >= exit_horizon_months:
                continue
            best_exit = None
            for exit_row in rows:
                exit_month_number = int(exit_row.get("month_number") or 0)
                exit_price = exit_row.get("projected_price")
                if (
                    exit_month_number <= entry_month_number
                    or exit_price in {None, ZERO}
                    or exit_month_number > exit_horizon_months
                ):
                    continue
                holding_months = exit_month_number - entry_month_number
                if holding_months <= 0:
                    continue
                price_return_pct = quantize_decimal(
                    percentage_change(exit_price, entry_price),
                    "0.01",
                )
                if price_return_pct is None:
                    continue
                income_support_pct = quantize_decimal(
                    net_income_yield_pct * Decimal(str(holding_months / 12)),
                    "0.01",
                ) or ZERO
                trade_return_pct = quantize_decimal(
                    price_return_pct + income_support_pct - transaction_drag_pct,
                    "0.01",
                )
                if trade_return_pct is None:
                    continue
                holding_annualized_return_pct = annualize_return_pct(
                    trade_return_pct,
                    holding_months,
                )
                calendar_adjusted_return_pct = annualize_return_pct(
                    trade_return_pct,
                    max(exit_month_number, 1),
                )
                exit_candidate = {
                    "exit_row": exit_row,
                    "exit_month_number": exit_month_number,
                    "price_return_pct": price_return_pct,
                    "income_support_pct": income_support_pct,
                    "trade_return_pct": trade_return_pct,
                    "holding_annualized_return_pct": quantize_decimal(holding_annualized_return_pct, "0.01"),
                    "calendar_adjusted_return_pct": quantize_decimal(calendar_adjusted_return_pct, "0.01"),
                    "holding_months": holding_months,
                }
                if best_exit is None or exit_preference_key(exit_candidate) > exit_preference_key(best_exit):
                    best_exit = exit_candidate
            if best_exit is None:
                continue
            discount_vs_now_pct = None
            if current_price not in {None, ZERO}:
                discount_vs_now_pct = quantize_decimal(
                    ((current_price - entry_price) / current_price) * ONE_HUNDRED,
                    "0.01",
                )
            entry_candidates.append(
                {
                    "entry_row": entry_row,
                    "entry_month_number": entry_month_number,
                    "entry_price": entry_price,
                    "entry_date": entry_row.get("projected_date"),
                    "buy_window_label": format_projection_month_window(
                        entry_row.get("projected_date"),
                        entry_month_number,
                    ),
                    "discount_vs_now_pct": discount_vs_now_pct,
                    **best_exit,
                }
            )
        if not entry_candidates:
            return None
        return max(entry_candidates, key=entry_preference_key)

    preferred_basis_key = "cycle_5y" if strategy["mode"] == OPTIMIZER_STRATEGY_5Y_PRIMARY else "projection_12m"
    primary_path = next((item for item in path_candidates if item["basis_key"] == preferred_basis_key), None) or path_candidates[0]
    primary_plan = evaluate_path(primary_path)
    secondary_path = next((item for item in path_candidates if item["basis_key"] != primary_path["basis_key"]), None)
    secondary_plan = evaluate_path(secondary_path) if secondary_path else None
    chosen_plan = primary_plan or secondary_plan
    chosen_path = primary_path if primary_plan else secondary_path
    if primary_plan is not None and primary_path is not None and secondary_plan is not None and secondary_path is not None:
        primary_trade_return = primary_plan.get("trade_return_pct")
        secondary_trade_return = secondary_plan.get("trade_return_pct")
        primary_holding_annualized = primary_plan.get("holding_annualized_return_pct")
        secondary_holding_annualized = secondary_plan.get("holding_annualized_return_pct")
        primary_meets_target = interval_meets_target(primary_plan)
        secondary_meets_target = interval_meets_target(secondary_plan)
        improvement_threshold = Decimal("1.50") if strategy["mode"] == OPTIMIZER_STRATEGY_12M_PRIMARY else Decimal("1.00")
        annualized_threshold = Decimal("2.00")
        secondary_is_materially_better = False
        if secondary_meets_target and not primary_meets_target:
            secondary_is_materially_better = True
        elif secondary_meets_target and primary_meets_target:
            secondary_is_materially_better = (
                secondary_trade_return is not None
                and (
                    primary_trade_return is None
                    or secondary_trade_return >= (primary_trade_return + improvement_threshold)
                    or (
                        secondary_trade_return == primary_trade_return
                        and secondary_holding_annualized is not None
                        and (
                            primary_holding_annualized is None
                            or secondary_holding_annualized >= (primary_holding_annualized + annualized_threshold)
                        )
                    )
                )
            )
        elif (
            secondary_holding_annualized is not None
            and (
                primary_holding_annualized is None
                or secondary_holding_annualized >= (primary_holding_annualized + annualized_threshold)
                or (
                    secondary_holding_annualized == primary_holding_annualized
                    and secondary_trade_return is not None
                    and (
                        primary_trade_return is None
                        or secondary_trade_return >= (primary_trade_return + improvement_threshold)
                    )
                )
            )
        ):
            secondary_is_materially_better = True
        if secondary_is_materially_better:
            chosen_plan = secondary_plan
            chosen_path = secondary_path
    if chosen_plan is None or chosen_path is None:
        return {"available": False}

    buy_month_number = chosen_plan["entry_month_number"]
    buy_date = chosen_plan.get("entry_date")
    buy_price = chosen_plan.get("entry_price")
    exit_row = chosen_plan.get("exit_row") or {}
    expected_exit_date = exit_row.get("projected_date")
    exit_month_number = int(exit_row.get("month_number") or 0)
    expected_exit_price = exit_row.get("projected_price")
    if buy_month_number <= 1:
        mode = "buy_now"
        mode_label = "Comprar ya"
    elif (chosen_plan.get("discount_vs_now_pct") or ZERO) >= Decimal("1.00"):
        mode = "wait_pullback"
        mode_label = "Esperar correccion"
    else:
        mode = "entrada_gradual"
        mode_label = "Entrada gradual"

    entry_horizon_months = int(chosen_path.get("entry_horizon_months") or 12)
    exit_horizon_months = int(chosen_path.get("exit_horizon_months") or entry_horizon_months)
    summary = (
        f"Con entrada acotada a los proximos {entry_horizon_months} meses, el mejor tramo sale entrando en {chosen_plan['buy_window_label'].lower()} alrededor de "
        f"{buy_price:.4f} EUR y saliendo en {format_projection_month_window(expected_exit_date, exit_month_number).lower()} "
        f"cerca de {expected_exit_price:.4f} EUR. Ese intervalo deja {chosen_plan['trade_return_pct']:.2f} % neto"
        f" y {chosen_plan['holding_annualized_return_pct']:.2f} % anualizado."
    )
    if chosen_plan.get("holding_annualized_return_pct") is not None:
        if interval_meets_target(chosen_plan):
            summary += (
                f" Cumple el objetivo minimo del {target_annualized_return_pct:.0f} %/a exigido por el optimizador."
            )
        else:
            summary += (
                f" No alcanza el objetivo minimo del {target_annualized_return_pct:.0f} %/a,"
                " asi que solo sirve como referencia tactica y no entraria en la cartera propuesta."
            )
    if exit_month_number > entry_horizon_months:
        summary += (
            f" La salida se deja correr mas alla del ano de entrada porque el tramo mejora hasta el mes {exit_month_number}."
        )
    if secondary_plan is not None and secondary_path is not None and secondary_plan.get("buy_window_label"):
        secondary_delta = abs(
            int(secondary_plan.get("entry_month_number") or 0) - int(chosen_plan.get("entry_month_number") or 0)
        )
        if chosen_path["basis_key"] == secondary_path["basis_key"]:
            summary += (
                f" El contraste {secondary_path['basis_label'].lower()} mejora materialmente el recorrido y por eso se usa como base del tramo."
            )
        elif secondary_delta <= 2:
            summary += (
                f" El contraste {secondary_path['basis_label'].lower()} confirma una ventana similar en "
                f"{secondary_plan['buy_window_label'].lower()}."
            )
        else:
            summary += (
                f" El contraste {secondary_path['basis_label'].lower()} empuja a revisar tambien "
                f"{secondary_plan['buy_window_label'].lower()}."
            )

    return {
        "available": True,
        "mode": mode,
        "mode_label": mode_label,
        "plan_horizon_months": entry_horizon_months,
        "entry_horizon_months": entry_horizon_months,
        "exit_horizon_months": exit_horizon_months,
        "analysis_basis_key": chosen_path["basis_key"],
        "analysis_basis_label": chosen_path["basis_label"],
        "entry_month_number": buy_month_number,
        "entry_date": buy_date,
        "entry_date_label": buy_date.isoformat() if buy_date else "",
        "entry_window_label": chosen_plan.get("buy_window_label") or "",
        "entry_price": buy_price,
        "buy_month_number": buy_month_number,
        "buy_date": buy_date,
        "buy_date_label": buy_date.isoformat() if buy_date else "",
        "buy_window_label": chosen_plan.get("buy_window_label") or "",
        "buy_price": buy_price,
        "discount_vs_now_pct": chosen_plan.get("discount_vs_now_pct"),
        "exit_month_number": exit_month_number,
        "exit_date": expected_exit_date,
        "exit_date_label": expected_exit_date.isoformat() if expected_exit_date else "",
        "exit_window_label": format_projection_month_window(expected_exit_date, exit_month_number),
        "exit_price": expected_exit_price,
        "expected_exit_month_number": exit_month_number,
        "expected_exit_date": expected_exit_date,
        "expected_exit_date_label": expected_exit_date.isoformat() if expected_exit_date else "",
        "expected_exit_window_label": format_projection_month_window(expected_exit_date, exit_month_number),
        "expected_exit_price": expected_exit_price,
        "expected_holding_months": chosen_plan.get("holding_months"),
        "holding_months": chosen_plan.get("holding_months"),
        "price_return_pct": chosen_plan.get("price_return_pct"),
        "income_support_pct": chosen_plan.get("income_support_pct"),
        "interval_return_pct": chosen_plan.get("trade_return_pct"),
        "expected_trade_return_pct": chosen_plan.get("trade_return_pct"),
        "holding_annualized_return_pct": chosen_plan.get("holding_annualized_return_pct"),
        "calendar_adjusted_return_pct": chosen_plan.get("calendar_adjusted_return_pct"),
        "interval_window_label": (
            f"{chosen_plan.get('buy_window_label') or ''} -> {format_projection_month_window(expected_exit_date, exit_month_number)}"
        ).strip(" ->"),
        "summary": summary,
        "primary_basis_label": primary_path.get("basis_label") if primary_path else "",
        "secondary_basis_label": secondary_path.get("basis_label") if secondary_path else "",
    }


def build_cycle_projection_monthly_trend_rows(card: dict) -> list[dict]:
    position = card.get("position")
    projection = card.get("projection") or {}
    cycle_projection = card.get("cycle_projection_5y") or {}
    if position is None or not cycle_projection.get("available"):
        return []

    current_price = getattr(position, "current_price_per_share", None)
    path = cycle_projection.get("path") or []
    if current_price in {None, ZERO} or not path:
        return []

    anchor_date = getattr(position, "latest_price_date", None) or django_timezone.localdate()
    anchor_points_by_month = {
        0: (
            quantize_decimal(current_price, "0.0001") or Decimal(str(current_price)),
            anchor_date,
        )
    }
    detailed_first_year_rows = build_projection_monthly_trend_rows(card)
    for step in detailed_first_year_rows:
        month_number = int(step.get("month_number") or 0)
        projected_price = step.get("projected_price")
        projected_date = step.get("projected_date")
        if month_number <= 0 or month_number > 12 or projected_price in {None, ZERO}:
            continue
        anchor_points_by_month[month_number] = (
            quantize_decimal(projected_price, "0.0001") or Decimal(str(projected_price)),
            projected_date or add_calendar_months(anchor_date, month_number),
        )
    for index, step in enumerate(path, start=1):
        projected_price = step.get("projected_price")
        months_offset = parse_projection_label_months(step.get("label")) or (index * 6)
        if projected_price in {None, ZERO} or months_offset <= 0:
            continue
        if months_offset <= 12 and months_offset in anchor_points_by_month:
            continue
        anchor_points_by_month[months_offset] = (
            quantize_decimal(projected_price, "0.0001") or Decimal(str(projected_price)),
            step.get("projected_date") or add_calendar_months(anchor_date, months_offset),
        )

    anchor_points = [
        (month_number, price, projected_date)
        for month_number, (price, projected_date) in sorted(anchor_points_by_month.items(), key=lambda item: item[0])
    ]
    return build_interpolated_projection_monthly_rows(anchor_points)


def build_owned_cycle_trade_timing_plan(card: dict) -> dict:
    position = card.get("position")
    if position is None or not getattr(position, "is_owned", False):
        return {"available": False}

    monthly_rows = build_cycle_projection_monthly_trend_rows(card)
    if len(monthly_rows) < 7:
        return {"available": False}

    current_price = getattr(position, "current_price_per_share", None)
    valid_rows = [row for row in monthly_rows if row.get("smoothed_slope_pct") is not None]
    if not valid_rows:
        return {"available": False}

    sell_row = None
    previous_row = None
    for row in valid_rows:
        if previous_row is not None:
            previous_slope = previous_row.get("smoothed_slope_pct")
            current_slope = row.get("smoothed_slope_pct")
            if previous_slope is not None and current_slope is not None and previous_slope >= ZERO and current_slope < ZERO:
                sell_row = row
                break
        previous_row = row

    if sell_row is None:
        return {
            "available": True,
            "mode": "hold",
            "analysis_basis_label": "Pendiente 5M desestacionalizada sobre la senda 5A vigente",
            "sale_month_number": None,
            "sale_window_label": "Mantener",
            "sale_date": None,
            "sale_date_label": "",
            "reentry_month_number": None,
            "reentry_window_label": "",
            "reentry_date": None,
            "reentry_date_label": "",
            "summary": (
                "La pendiente desestacionalizada de 5 meses sigue positiva en toda la curva 5A vigente. "
                "No aparece un giro bajista claro que justifique salir por ahora."
            ),
            "signal_label": "Pendiente 5M positiva",
            "signal_value_pct": valid_rows[-1].get("smoothed_slope_pct"),
            "monthly_rows": monthly_rows,
            "drawdown_month_number": None,
            "drawdown_margin_pct": None,
            "pre_sale_return_pct": None,
        }

    reentry_row = None
    previous_row = sell_row
    for row in valid_rows:
        if row["month_number"] <= sell_row["month_number"]:
            previous_row = row
            continue
        previous_slope = previous_row.get("smoothed_slope_pct")
        current_slope = row.get("smoothed_slope_pct")
        if previous_slope is not None and current_slope is not None and previous_slope <= ZERO and current_slope > ZERO:
            reentry_row = row
            break
        previous_row = row

    pre_sale_return_pct = (
        quantize_decimal(percentage_change(sell_row.get("projected_price"), current_price), "0.01")
        if current_price not in {None, ZERO}
        else None
    )
    sale_window_label = format_projection_month_window(
        sell_row.get("projected_date"),
        sell_row.get("month_number"),
    )
    reentry_window_label = format_projection_month_window(
        reentry_row.get("projected_date") if reentry_row else None,
        reentry_row.get("month_number") if reentry_row else None,
    )
    summary = (
        f"La pendiente desestacionalizada de 5 meses gira a negativo en {sale_window_label.lower()}. "
        f"Antes de ese giro la senda aun deja un {pre_sale_return_pct:.2f} % de recorrido estimado."
        if pre_sale_return_pct is not None
        else f"La pendiente desestacionalizada de 5 meses gira a negativo en {sale_window_label.lower()}."
    )
    if reentry_row:
        summary += (
            f" Si despues vuelve a cruzar a positivo, la primera ventana clara aparece en "
            f"{reentry_window_label.lower()}."
        )
    else:
        summary += " Despues no aparece una reentrada clara en la curva 5A actual."

    return {
        "available": True,
        "mode": "sale_reentry" if reentry_row else "sale_review",
        "analysis_basis_label": "Pendiente 5M desestacionalizada sobre la senda 5A vigente",
        "sale_month_number": sell_row["month_number"],
        "sale_year_number": max(int(math.ceil(sell_row["month_number"] / 12)), 1),
        "sale_window_label": sale_window_label,
        "sale_date": sell_row.get("projected_date"),
        "sale_date_label": sell_row.get("projected_date").isoformat() if sell_row.get("projected_date") else "",
        "reentry_month_number": reentry_row["month_number"] if reentry_row else None,
        "reentry_year_number": max(int(math.ceil(reentry_row["month_number"] / 12)), 1) if reentry_row else None,
        "reentry_window_label": reentry_window_label,
        "reentry_date": reentry_row.get("projected_date") if reentry_row else None,
        "reentry_date_label": reentry_row.get("projected_date").isoformat() if reentry_row and reentry_row.get("projected_date") else "",
        "summary": summary,
        "signal_label": "Pendiente 5M negativa" if sell_row.get("smoothed_slope_pct") is not None else "Pendiente 5M",
        "signal_value_pct": sell_row.get("smoothed_slope_pct"),
        "monthly_rows": monthly_rows,
        "yearly_rows": [],
        "drawdown_month_number": sell_row["month_number"],
        "drawdown_year_number": max(int(math.ceil(sell_row["month_number"] / 12)), 1),
        "drawdown_margin_pct": sell_row.get("smoothed_slope_pct"),
        "pre_sale_return_pct": pre_sale_return_pct,
    }


def build_purchase_trade_rotation_guidance(
    trade_plan: dict,
    position: EquityPosition,
    optimizer_cards: list[dict] | None,
    *,
    comparison_amount: Decimal = Decimal("10000"),
) -> dict:
    if not trade_plan.get("available") or not optimizer_cards:
        return {"available": False}

    comparison_amount = comparison_amount if comparison_amount > ZERO else Decimal("10000")
    candidates = [
        candidate
        for candidate in (
            build_equity_optimizer_candidate(card, OPTIMIZER_STRATEGY_12M_PRIMARY)
            for card in optimizer_cards or []
        )
        if candidate and candidate["position"].ticker != position.ticker
    ]
    ranked_candidates = rank_optimizer_candidates(
        filter_positive_optimizer_candidates(candidates, OPTIMIZER_STRATEGY_12M_PRIMARY)
    )

    selected_candidate = None
    selected_scenario = None
    for candidate in ranked_candidates:
        scenario = build_optimizer_allocation_scenario(candidate, comparison_amount)
        review = review_optimizer_ticket_efficiency(candidate, comparison_amount, scenario)
        if review.get("keep"):
            selected_candidate = candidate
            selected_scenario = scenario
            break

    if selected_candidate is None or selected_scenario is None:
        return {"available": False}

    alternative_return_pct = selected_scenario.get("net_projected_return_pct")
    action = "mantener"
    summary = (
        f"La mejor alternativa actual del radar es {selected_candidate['position'].company_name} "
        f"({selected_candidate['position'].ticker}) con retorno neto 12M del {alternative_return_pct:.2f} %."
    )

    tactical_weak_period_pct = trade_plan.get("pre_sale_return_pct")
    if tactical_weak_period_pct is None:
        tactical_weak_period_pct = ZERO
    if (
        trade_plan.get("mode") in {"sale_reentry", "sale_review"}
        and alternative_return_pct is not None
        and alternative_return_pct >= tactical_weak_period_pct + OPTIMIZER_MAX_ROUNDTRIP_DRAG_PCT
    ):
        action = "rotar"
        summary = (
            f"En la foto actual compensa mas rotar hacia {selected_candidate['position'].company_name} "
            f"({selected_candidate['position'].ticker}), con retorno neto 12M del {alternative_return_pct:.2f} %, "
            "que esperar el giro bajista previsto en la tendencia 5A de esta posicion."
        )
    elif trade_plan.get("mode") == "sale_reentry":
        action = "esperar_reentrada"
        summary += " Hoy sigue pesando mas esperar la ventana de salida y la posible reentrada que rotar ahora."
    elif trade_plan.get("mode") == "sale_review":
        action = "revisar_en_venta"
        summary += " Hoy sigue pesando mas esperar la ventana de salida y revisar entonces el radar."

    return {
        "available": True,
        "action": action,
        "summary": summary,
        "alternative_ticker": selected_candidate["position"].ticker,
        "alternative_company_name": selected_candidate["position"].company_name,
        "alternative_return_pct": alternative_return_pct,
        "alternative_reference_label": selected_candidate.get("reference_label") or "",
        "alternative_trade_alert_label": selected_candidate.get("trade_alert_label") or "",
        "alternative_reliability_label": selected_candidate.get("reliability_label") or "",
        "alternative_strategy_label": selected_candidate.get("strategy_label") or "",
        "hold_remaining_return_pct": quantize_decimal(tactical_weak_period_pct),
        "reentry_remaining_return_pct": None,
    }


def resolve_tracking_snapshot_unit_price(
    snapshot: EquityTicketSnapshot | None,
    shares: Decimal | None,
) -> Decimal | None:
    if snapshot is None or shares in {None, ZERO}:
        return None
    try:
        current_value = Decimal(str(snapshot.current_value))
        shares_value = Decimal(str(shares))
    except Exception:
        return None
    if shares_value == ZERO:
        return None
    return quantize_decimal(current_value / shares_value, "0.0001")


def build_tracking_ticket_unit_price_context(
    position: EquityPosition,
    baseline_snapshot: EquityTicketSnapshot,
    latest_snapshot: EquityTicketSnapshot,
    purchase_baseline: EquityPurchaseForecastBaseline | None = None,
) -> dict:
    initial_unit_price = quantize_decimal(getattr(purchase_baseline, "baseline_price", None), "0.0001")
    initial_unit_price_date = getattr(purchase_baseline, "baseline_date", None)
    initial_unit_price_note = "Compra o alta web"
    if initial_unit_price is None:
        initial_unit_price = resolve_tracking_snapshot_unit_price(baseline_snapshot, position.shares)
        initial_unit_price_date = getattr(baseline_snapshot, "snapshot_date", None)
        initial_unit_price_note = "Primera foto guardada"
    if initial_unit_price is None:
        initial_unit_price = quantize_decimal(getattr(position, "average_cost_per_share", None), "0.0001")
        initial_unit_price_date = getattr(position, "opened_on", None) or initial_unit_price_date
        initial_unit_price_note = "Precio medio de compra"

    current_unit_price = quantize_decimal(getattr(position, "current_price_per_share", None), "0.0001")
    current_unit_price_date = getattr(position, "latest_price_date", None)
    current_unit_price_note = "Cotizacion mas reciente"
    if current_unit_price is None:
        current_unit_price = resolve_tracking_snapshot_unit_price(latest_snapshot, position.shares)
        current_unit_price_date = getattr(latest_snapshot, "snapshot_date", None)
        current_unit_price_note = "Ultima foto guardada"

    return {
        "initial_unit_price": initial_unit_price,
        "initial_unit_price_date": initial_unit_price_date,
        "initial_unit_price_note": initial_unit_price_note,
        "current_unit_price": current_unit_price,
        "current_unit_price_date": current_unit_price_date,
        "current_unit_price_note": current_unit_price_note,
    }


def build_equity_ticket_tracking_item(
    card: dict,
    snapshots: list[EquityTicketSnapshot],
    tracking_anchor_date: date | None = None,
    purchase_baseline: EquityPurchaseForecastBaseline | None = None,
    optimizer_cards: list[dict] | None = None,
) -> dict | None:
    if not snapshots:
        return None

    relevant_snapshots = filter_ticket_tracking_snapshots(snapshots, tracking_anchor_date)
    if not relevant_snapshots:
        relevant_snapshots = snapshots

    position = card["position"]
    baseline = snapshots[0]
    latest = snapshots[-1]
    unit_price_context = build_tracking_ticket_unit_price_context(
        position,
        baseline,
        latest,
        purchase_baseline=purchase_baseline,
    )
    previous = snapshots[-2] if len(snapshots) >= 2 else None
    raw_expected_market_value_12m = baseline.projected_market_value_12m or baseline.current_value
    raw_expected_total_value_12m = baseline.projected_total_value_12m or raw_expected_market_value_12m
    expected_market_value_12m, expected_target_moderation_12m = moderate_tracking_target_value(
        quantize_decimal(baseline.current_value, "0.01") or ZERO,
        quantize_decimal(raw_expected_market_value_12m, "0.01") or ZERO,
        months=12,
        card=card,
    )
    expected_total_value_12m, expected_total_target_moderation_12m = moderate_tracking_target_value(
        quantize_decimal(baseline.current_value, "0.01") or ZERO,
        quantize_decimal(raw_expected_total_value_12m, "0.01") or ZERO,
        months=12,
        card=card,
    )
    actual_series = [{"date": snapshot.snapshot_date, "value": snapshot.current_value} for snapshot in snapshots]
    invested_series = [{"date": snapshot.snapshot_date, "value": snapshot.invested_amount} for snapshot in snapshots]
    expected_series, current_expected_value, projected_end_date = build_ticket_expected_series(
        snapshots,
        expected_market_value_12m,
        card=card,
    )
    expected_series, current_expected_value, expected_series_calibration = apply_tracking_expected_series_calibration(
        snapshots,
        expected_series,
        expected_market_value_12m,
        projected_end_date,
        card=card,
    )
    adjusted_expected_market_value_12m = (
        expected_series_calibration.get("adjusted_target_value")
        if expected_series_calibration.get("available")
        else quantize_decimal(expected_market_value_12m, "0.01")
    ) or quantize_decimal(expected_market_value_12m, "0.01") or ZERO
    adjusted_expected_total_value_12m = quantize_decimal(expected_total_value_12m, "0.01")
    if (
        adjusted_expected_total_value_12m is not None
        and expected_market_value_12m is not None
        and adjusted_expected_market_value_12m is not None
    ):
        adjusted_expected_total_value_12m = quantize_decimal(
            adjusted_expected_total_value_12m + (adjusted_expected_market_value_12m - expected_market_value_12m),
            "0.01",
    )
    expected_series_5y, current_expected_value_5y, projected_end_date_5y, expected_total_value_5y = build_ticket_expected_series_5y(
        card,
        snapshots,
        purchase_baseline=purchase_baseline,
    )
    dense_expected_series = densify_projected_tracking_series(expected_series)
    dense_expected_series_5y = densify_projected_tracking_series(expected_series_5y)
    chart = build_value_tracking_chart(actual_series, dense_expected_series)
    shared_baseline = relevant_snapshots[0]
    raw_shared_expected_market_value_12m = shared_baseline.projected_market_value_12m or shared_baseline.current_value
    raw_shared_expected_total_value_12m = shared_baseline.projected_total_value_12m or raw_shared_expected_market_value_12m
    shared_expected_market_value_12m, shared_expected_target_moderation_12m = moderate_tracking_target_value(
        quantize_decimal(shared_baseline.current_value, "0.01") or ZERO,
        quantize_decimal(raw_shared_expected_market_value_12m, "0.01") or ZERO,
        months=12,
        card=card,
    )
    shared_expected_total_value_12m, shared_expected_total_target_moderation_12m = moderate_tracking_target_value(
        quantize_decimal(shared_baseline.current_value, "0.01") or ZERO,
        quantize_decimal(raw_shared_expected_total_value_12m, "0.01") or ZERO,
        months=12,
        card=card,
    )
    shared_actual_series = [{"date": snapshot.snapshot_date, "value": snapshot.current_value} for snapshot in relevant_snapshots]
    shared_expected_series, shared_current_expected_value, shared_projection_end_date = build_ticket_expected_series(
        relevant_snapshots,
        shared_expected_market_value_12m,
        card=card,
    )
    shared_expected_series, shared_current_expected_value, shared_expected_series_calibration = apply_tracking_expected_series_calibration(
        relevant_snapshots,
        shared_expected_series,
        shared_expected_market_value_12m,
        shared_projection_end_date,
        card=card,
    )
    adjusted_shared_expected_market_value_12m = (
        shared_expected_series_calibration.get("adjusted_target_value")
        if shared_expected_series_calibration.get("available")
        else quantize_decimal(shared_expected_market_value_12m, "0.01")
    ) or quantize_decimal(shared_expected_market_value_12m, "0.01") or ZERO
    adjusted_shared_expected_total_value_12m = quantize_decimal(shared_expected_total_value_12m, "0.01")
    if (
        adjusted_shared_expected_total_value_12m is not None
        and shared_expected_market_value_12m is not None
        and adjusted_shared_expected_market_value_12m is not None
    ):
        adjusted_shared_expected_total_value_12m = quantize_decimal(
            adjusted_shared_expected_total_value_12m
            + (adjusted_shared_expected_market_value_12m - shared_expected_market_value_12m),
            "0.01",
        )
    shared_expected_series_5y, shared_current_expected_value_5y, shared_projection_end_date_5y, shared_expected_total_value_5y = build_ticket_expected_series_5y(
        card,
        relevant_snapshots,
        purchase_baseline=purchase_baseline,
    )
    shared_dense_expected_series = densify_projected_tracking_series(shared_expected_series)
    shared_dense_expected_series_5y = densify_projected_tracking_series(shared_expected_series_5y)
    gap_value = (
        quantize_decimal(latest.current_value - current_expected_value, "0.01")
        if current_expected_value is not None
        else None
    )
    gap_pct = percentage_change(latest.current_value, current_expected_value) if current_expected_value is not None else None
    daily_change_pct = (
        percentage_change(latest.current_value, previous.current_value)
        if previous and previous.current_value
        else None
    )
    try:
        trade_plan = build_owned_cycle_trade_timing_plan(card)
    except Exception:
        logger.exception(
            "No se pudo construir el plan tactico de compra para %s",
            position.ticker,
        )
        trade_plan = {"available": False}
    try:
        rotation_plan = build_purchase_trade_rotation_guidance(
            trade_plan,
            position,
            optimizer_cards,
        )
    except Exception:
        logger.exception(
            "No se pudo construir la sugerencia de rotacion para %s",
            position.ticker,
        )
        rotation_plan = {"available": False}
    current_projection = card.get("presentation_projection") or {}
    current_projection_return_pct = None
    if current_projection.get("available"):
        current_projection_return_pct = quantize_decimal(
            current_projection.get("visible_total_return_pct"),
            "0.01",
        )
    if current_projection_return_pct is None:
        current_projection_return_pct = quantize_decimal(
            (card.get("projection") or {}).get("base_return_pct"),
            "0.01",
        )
    stored_projection_return_pct = percentage_change(adjusted_expected_total_value_12m, baseline.current_value)
    projection_drift_pct = (
        quantize_decimal(current_projection_return_pct - stored_projection_return_pct, "0.01")
        if current_projection_return_pct is not None and stored_projection_return_pct is not None
        else None
    )
    projection_direction_changed = (
        current_projection_return_pct is not None
        and stored_projection_return_pct is not None
        and (
            (current_projection_return_pct < ZERO and stored_projection_return_pct >= ZERO)
            or (current_projection_return_pct >= ZERO and stored_projection_return_pct < ZERO)
        )
    )
    projection_source = {
        "available": stored_projection_return_pct is not None,
        "baseline_date": baseline.snapshot_date,
        "stored_return_12m_pct": quantize_decimal(stored_projection_return_pct, "0.01"),
        "current_return_12m_pct": current_projection_return_pct,
        "drift_pct": projection_drift_pct,
        "direction_changed": projection_direction_changed,
        "conflict": bool(projection_direction_changed or (projection_drift_pct is not None and abs(projection_drift_pct) >= Decimal("8.00"))),
    }
    if projection_source["conflict"]:
        projection_source["label"] = "La tesis actual ha cambiado"
        projection_source["note"] = (
            "La linea naranja es la prevision guardada al iniciar el ticket; "
            "los escenarios de abajo son la lectura actual recalculada."
        )
    else:
        projection_source["label"] = "Misma tesis de seguimiento"
        projection_source["note"] = "La senda guardada y la lectura actual no muestran una ruptura grande."
    return {
        "position": position,
        "card": card,
        "baseline_snapshot": baseline,
        "shared_baseline_snapshot": shared_baseline,
        "latest_snapshot": latest,
        "snapshot_count": len(snapshots),
        "shared_snapshot_count": len(relevant_snapshots),
        "days_tracked": max((latest.snapshot_date - baseline.snapshot_date).days, 0),
        "actual_series": actual_series,
        "invested_series": invested_series,
        "expected_series": expected_series,
        "expected_series_dense": dense_expected_series,
        "expected_target_moderation_12m": expected_target_moderation_12m,
        "expected_total_target_moderation_12m": expected_total_target_moderation_12m,
        "expected_series_calibration": expected_series_calibration,
        "chart": chart,
        "current_expected_value": current_expected_value,
        "expected_market_value_12m": adjusted_expected_market_value_12m,
        "expected_total_value_12m": adjusted_expected_total_value_12m,
        "expected_series_5y": expected_series_5y,
        "expected_series_5y_dense": dense_expected_series_5y,
        "current_expected_value_5y": current_expected_value_5y,
        "expected_total_value_5y": expected_total_value_5y,
        "projection_end_date_5y": projected_end_date_5y,
        "shared_actual_series": shared_actual_series,
        "shared_expected_series": shared_expected_series,
        "shared_expected_series_dense": shared_dense_expected_series,
        "shared_expected_target_moderation_12m": shared_expected_target_moderation_12m,
        "shared_expected_total_target_moderation_12m": shared_expected_total_target_moderation_12m,
        "shared_expected_series_calibration": shared_expected_series_calibration,
        "shared_current_expected_value": shared_current_expected_value,
        "shared_expected_market_value_12m": adjusted_shared_expected_market_value_12m,
        "shared_expected_total_value_12m": adjusted_shared_expected_total_value_12m,
        "shared_projection_end_date": shared_projection_end_date,
        "shared_expected_series_5y": shared_expected_series_5y,
        "shared_expected_series_5y_dense": shared_dense_expected_series_5y,
        "shared_current_expected_value_5y": shared_current_expected_value_5y,
        "shared_expected_total_value_5y": shared_expected_total_value_5y,
        "shared_projection_end_date_5y": shared_projection_end_date_5y,
        "actual_change_pct": percentage_change(latest.current_value, baseline.current_value),
        "expected_change_pct": percentage_change(current_expected_value, baseline.current_value)
        if current_expected_value is not None
        else None,
        "gap_value": gap_value,
        "gap_pct": gap_pct,
        "gap_tone": "good" if gap_value is not None and gap_value >= ZERO else "warn",
        "daily_change_pct": daily_change_pct,
        "projection_end_date": projected_end_date,
        "trade_plan": trade_plan,
        "rotation_plan": rotation_plan,
        "projection_source": projection_source,
        **unit_price_context,
    }


def build_global_equity_ticket_tracking_item(
    ticket_items: list[dict],
    history_cards: list[dict] | None = None,
) -> dict:
    actual_series = build_aggregated_ticket_actual_series(ticket_items)
    if not actual_series:
        return {"available": False}

    invested_series = build_aggregated_ticket_series(ticket_items, "invested_series")
    expected_series = build_aggregated_ticket_series(ticket_items, "expected_series_dense")
    expected_series_5y = build_aggregated_ticket_series(ticket_items, "expected_series_5y_dense")
    baseline_value = quantize_decimal(
        sum(
            (
                getattr(item.get("baseline_snapshot"), "current_value", None) or ZERO
            )
            for item in ticket_items
        ),
        "0.01",
    ) or ZERO
    benchmark = build_aggregated_tracking_benchmark_context(ticket_items)
    portfolio_summary = build_portfolio_summary_context(
        history_cards or [],
        benchmark_label=benchmark.get("label"),
    )
    rebased_comparison = build_tracking_rebased_comparison_series(
        actual_series,
        invested_series,
        benchmark.get("close_series", []),
    )
    chart = (
        build_value_tracking_chart(
            rebased_comparison["portfolio_series"],
            [],
            benchmark_series=rebased_comparison["benchmark_series"],
            value_suffix="",
            axis_formatter=format_axis_value,
            time_marker_mode="month",
            grid_marker_mode="month",
        )
        if rebased_comparison.get("available")
        else {"available": False}
    )
    chart_5y = build_value_tracking_chart(actual_series, expected_series_5y) if expected_series_5y else {"available": False}
    latest_actual = actual_series[-1]["value"]
    latest_date = actual_series[-1]["date"]
    target_annual_return_pct = resolve_tracking_target_annual_return_pct()
    target_series = build_tracking_target_series(ticket_items, target_annual_return_pct)
    target_today_value = next(
        (point["value"] for point in reversed(target_series) if point["date"] <= latest_date),
        None,
    )
    current_expected_value = next(
        (point["value"] for point in reversed(expected_series) if point["date"] <= latest_date),
        None,
    )
    current_expected_value_5y = next(
        (point["value"] for point in reversed(expected_series_5y) if point["date"] <= latest_date),
        None,
    )
    expected_total_value_12m = sum(
        (item.get("expected_total_value_12m") or item.get("expected_market_value_12m") or ZERO)
        for item in ticket_items
    )
    expected_total_value_5y = sum((item.get("expected_total_value_5y") or ZERO) for item in ticket_items)
    gap_value = (
        quantize_decimal(latest_actual - current_expected_value, "0.01")
        if current_expected_value is not None
        else None
    )
    tracked_days = max((latest_date - actual_series[0]["date"]).days, 0)
    net_gain_value = quantize_decimal(
        sum(
            (
                getattr(item.get("latest_snapshot"), "current_value", None) or ZERO
            )
            - (
                getattr(item.get("baseline_snapshot"), "current_value", None) or ZERO
            )
            for item in ticket_items
        ),
        "0.01",
    )
    actual_change_pct = percentage_change(
        baseline_value + (net_gain_value or ZERO),
        baseline_value,
    )
    comparable_returns = build_comparable_return_metrics(
        actual_change_pct,
        tracked_days,
    )
    daily_change_pct = build_global_ticket_daily_change_pct(ticket_items)
    net_series_12m = build_tracking_series_vs_dynamic_baseline(actual_series, ticket_items)
    net_expected_series_12m = build_tracking_series_vs_dynamic_baseline(expected_series, ticket_items)
    net_target_series = build_tracking_series_vs_dynamic_baseline(target_series, ticket_items)
    net_series_5y = build_tracking_series_vs_dynamic_baseline(actual_series, ticket_items)
    net_expected_series_5y = build_tracking_series_vs_dynamic_baseline(expected_series_5y, ticket_items)
    return_series_12m = build_tracking_series_vs_dynamic_baseline(actual_series, ticket_items, as_percentage=True)
    return_expected_series_12m = build_tracking_series_vs_dynamic_baseline(expected_series, ticket_items, as_percentage=True)
    return_target_series = build_tracking_series_vs_dynamic_baseline(target_series, ticket_items, as_percentage=True)
    return_benchmark_series = build_tracking_series_vs_dynamic_baseline(
        benchmark.get("series", []),
        ticket_items,
        as_percentage=True,
    )
    return_series_5y = build_tracking_series_vs_dynamic_baseline(actual_series, ticket_items, as_percentage=True)
    return_expected_series_5y = build_tracking_series_vs_dynamic_baseline(
        expected_series_5y,
        ticket_items,
        as_percentage=True,
    )
    cumulative_alpha_series = build_tracking_series_difference(
        return_series_12m,
        return_benchmark_series,
    )
    weekly_alpha_series = build_tracking_trailing_delta_series(cumulative_alpha_series)
    net_chart_12m = build_value_tracking_chart(
        net_series_12m,
        net_expected_series_12m,
    )
    target_chart = build_value_tracking_chart(
        actual_series,
        target_series,
    )
    target_net_chart = build_value_tracking_chart(
        net_series_12m,
        net_target_series,
    )
    target_return_chart = build_value_tracking_chart(
        return_series_12m,
        return_target_series,
        benchmark_series=return_benchmark_series,
        value_suffix="%",
        axis_formatter=format_percentage_axis_value,
    )
    net_chart_5y = build_value_tracking_chart(
        net_series_5y,
        net_expected_series_5y,
    )
    return_chart_12m = build_value_tracking_chart(
        return_series_12m,
        return_expected_series_12m,
        benchmark_series=return_benchmark_series,
        value_suffix="%",
        axis_formatter=format_percentage_axis_value,
    )
    return_chart_5y = build_value_tracking_chart(
        return_series_5y,
        return_expected_series_5y,
        benchmark_series=return_benchmark_series,
        value_suffix="%",
        axis_formatter=format_percentage_axis_value,
    )
    cumulative_alpha_chart = build_value_tracking_chart(
        cumulative_alpha_series,
        [],
        value_suffix="%",
        axis_formatter=format_percentage_axis_value,
    )
    weekly_alpha_chart = build_value_tracking_chart(
        weekly_alpha_series,
        [],
        value_suffix="%",
        axis_formatter=format_percentage_axis_value,
    )
    expected_net_value_12m = (
        quantize_decimal((expected_total_value_12m or ZERO) - baseline_value, "0.01")
        if expected_total_value_12m is not None
        else None
    )
    expected_net_value_5y = (
        quantize_decimal((expected_total_value_5y or ZERO) - baseline_value, "0.01")
        if expected_total_value_5y is not None
        else None
    )
    expected_return_pct_12m = percentage_change(expected_total_value_12m, baseline_value)
    expected_return_pct_5y = percentage_change(expected_total_value_5y, baseline_value)
    target_gap_value = (
        quantize_decimal(latest_actual - target_today_value, "0.01")
        if target_today_value is not None
        else None
    )
    annualized_return_pct = quantize_decimal(comparable_returns.get("annual_equivalent_return_pct"), "0.01")
    objective_status = build_tracking_objective_status(
        net_gain_value,
        annualized_return_pct,
        target_annual_return_pct,
    )
    return {
        "available": True,
        "chart": chart,
        "target_chart": target_chart,
        "target_net_chart": target_net_chart,
        "target_return_chart": target_return_chart,
        "chart_5y": chart_5y,
        "net_chart_12m": net_chart_12m,
        "net_chart_5y": net_chart_5y,
        "return_chart_12m": return_chart_12m,
        "return_chart_5y": return_chart_5y,
        "cumulative_alpha_chart": cumulative_alpha_chart,
        "weekly_alpha_chart": weekly_alpha_chart,
        "baseline_value": baseline_value,
        "latest_value": latest_actual,
        "target_annual_return_pct": target_annual_return_pct,
        "target_today_value": target_today_value,
        "target_gap_value": target_gap_value,
        "target_gap_pct": percentage_change(latest_actual, target_today_value) if target_today_value is not None else None,
        "target_gap_tone": "good" if target_gap_value is not None and target_gap_value >= ZERO else "warn",
        "objective_status": objective_status,
        "expected_today_value": current_expected_value,
        "expected_today_value_5y": current_expected_value_5y,
        "expected_market_value_12m": expected_series[-1]["value"] if expected_series else baseline_value,
        "expected_total_value_12m": quantize_decimal(expected_total_value_12m, "0.01") or ZERO,
        "expected_total_value_5y": quantize_decimal(expected_total_value_5y, "0.01") or ZERO,
        "expected_net_value_12m": expected_net_value_12m,
        "expected_net_value_5y": expected_net_value_5y,
        "expected_return_pct_12m": quantize_decimal(expected_return_pct_12m, "0.01"),
        "expected_return_pct_5y": quantize_decimal(expected_return_pct_5y, "0.01"),
        "actual_change_pct": actual_change_pct,
        "expected_change_pct": percentage_change(current_expected_value, baseline_value)
        if current_expected_value is not None
        else None,
        "rebased_comparison": rebased_comparison,
        "net_gain_value": net_gain_value,
        "net_gain_tone": "good" if net_gain_value is not None and net_gain_value >= ZERO else "warn",
        "invested_return_pct": quantize_decimal(actual_change_pct, "0.01"),
        "annualized_return_pct": annualized_return_pct,
        "annualized_return_tone": "good"
        if annualized_return_pct is not None and annualized_return_pct >= target_annual_return_pct
        else "warn",
        "benchmark": benchmark,
        "portfolio_summary": portfolio_summary,
        "gap_value": gap_value,
        "gap_pct": percentage_change(latest_actual, current_expected_value) if current_expected_value is not None else None,
        "gap_tone": "good" if gap_value is not None and gap_value >= ZERO else "warn",
        "daily_change_pct": daily_change_pct,
        "tracked_days": tracked_days,
    }


def build_equity_ticket_tracking_context(
    history_cards: list[dict],
    optimizer_cards: list[dict] | None = None,
) -> dict:
    owned_cards = [card for card in history_cards if card["position"].is_owned]
    if not owned_cards:
        return {
            "available": False,
            "tickets": [],
            "tracked_ticket_count": 0,
            "snapshot_days_count": 0,
            "sale_timeline": {"available": False},
            "global": {"available": False},
        }

    cards_by_id = {card["position"].id: card for card in owned_cards if card["position"].id}
    snapshots = list(
        EquityTicketSnapshot.objects.filter(position_id__in=cards_by_id.keys())
        .select_related("position")
        .order_by("snapshot_date", "position__company_name", "position__ticker")
    )
    purchase_baselines = {
        baseline.position_id: baseline
        for baseline in EquityPurchaseForecastBaseline.objects.filter(position_id__in=cards_by_id.keys())
    }
    grouped_snapshots: dict[int, list[EquityTicketSnapshot]] = defaultdict(list)
    for snapshot in snapshots:
        grouped_snapshots[snapshot.position_id].append(snapshot)

    tracking_anchor_date = resolve_ticket_tracking_anchor_date(grouped_snapshots)
    ticket_items = []
    for card in owned_cards:
        position_id = card["position"].id
        if not position_id:
            continue
        item = build_equity_ticket_tracking_item(
            card,
            grouped_snapshots.get(position_id, []),
            tracking_anchor_date=tracking_anchor_date,
            purchase_baseline=purchase_baselines.get(position_id),
            optimizer_cards=optimizer_cards,
        )
        if item:
            ticket_items.append(item)

    ticket_items.sort(
        key=lambda item: (
            -(item["latest_snapshot"].current_value if item.get("latest_snapshot") else ZERO),
            item["position"].company_name,
        )
    )
    earliest_baseline_date = min(
        (
            item["baseline_snapshot"].snapshot_date
            for item in ticket_items
            if item.get("baseline_snapshot") is not None
        ),
        default=None,
    )
    snapshot_days = sorted(
        {
            point["date"]
            for item in ticket_items
            for point in item.get("actual_series", [])
        }
    )
    return {
        "available": bool(ticket_items),
        "tickets": ticket_items,
        "tracked_ticket_count": len(ticket_items),
        "snapshot_days_count": len(snapshot_days),
        "anchor_date": earliest_baseline_date,
        "shared_anchor_date": tracking_anchor_date,
        "sale_timeline": build_tracking_sale_timeline_context(ticket_items),
        "global": build_global_equity_ticket_tracking_item(ticket_items, history_cards=owned_cards) if ticket_items else {"available": False},
    }


def serialize_equity_market_value_history(
    position: EquityPosition,
    closed_on: date | None = None,
    sale_price_per_share: Decimal | None = None,
) -> list[dict]:
    history = list(position.price_history.all().order_by("price_date"))
    serialized = []
    for point in history:
        if closed_on and point.price_date > closed_on:
            break
        serialized.append(
            {
                "date": point.price_date.isoformat(),
                "market_value": str(quantize_decimal(position.shares * point.close_price, "0.01") or ZERO),
            }
        )
    if closed_on and sale_price_per_share is not None:
        sale_market_value = quantize_decimal(position.shares * sale_price_per_share, "0.01") or ZERO
        if serialized and serialized[-1]["date"] == closed_on.isoformat():
            serialized[-1]["market_value"] = str(sale_market_value)
        else:
            serialized.append({"date": closed_on.isoformat(), "market_value": str(sale_market_value)})
    return serialized


def build_single_value_history_chart(series: list[dict]) -> dict:
    filtered = [
        {"date": point.get("date"), "value": Decimal(str(point.get("value")))}
        for point in series
        if point.get("date") is not None and point.get("value") is not None
    ]
    if len(filtered) < 2:
        return {"available": False}
    chart = build_dual_axis_chart(filtered, [])
    return {
        "available": bool(chart.get("stock_line")),
        "line": chart.get("stock_line", ""),
        "min_label": chart.get("stock_min_label", "-"),
        "max_label": chart.get("stock_max_label", "-"),
        "x_markers": chart.get("x_markers", []),
        "start_label": filtered[0]["date"].isoformat(),
        "end_label": filtered[-1]["date"].isoformat(),
        "points_count": len(filtered),
    }


def build_annual_result_chart(rows: list[dict], width: int = 640, height: int = 240, padding: int = 26) -> dict:
    if not rows:
        return {"available": False, "bars": []}

    values = [Decimal(str(row.get("net_result") or ZERO)) for row in rows]
    min_value = min([*values, ZERO])
    max_value = max([*values, ZERO])
    if min_value == max_value:
        max_value += Decimal("1")
        min_value -= Decimal("1")

    span_y = height - (padding * 2)
    span_x = width - (padding * 2)
    slot_width = span_x / max(len(rows), 1)
    bar_width = max(slot_width * 0.48, 18)

    def to_y(value: Decimal) -> float:
        normalized = (value - min_value) / (max_value - min_value)
        return float(height - padding - (normalized * span_y))

    zero_y = to_y(ZERO)
    bars = []
    for index, row in enumerate(rows):
        value = Decimal(str(row.get("net_result") or ZERO))
        center_x = padding + (slot_width * index) + (slot_width / 2)
        value_y = to_y(value)
        top_y = min(value_y, zero_y)
        bar_height = max(abs(zero_y - value_y), 1.5)
        bars.append(
            {
                "x": f"{center_x - (bar_width / 2):.1f}",
                "y": f"{top_y:.1f}",
                "width": f"{bar_width:.1f}",
                "height": f"{bar_height:.1f}",
                "year": row["year"],
                "value_label": f'{format_axis_value(value)} EUR',
                "tone": "good" if value >= ZERO else "warn",
                "text_x": f"{center_x:.1f}",
                "text_y": f"{height - 8:.1f}",
            }
        )

    return {
        "available": True,
        "bars": bars,
        "zero_y": f"{zero_y:.1f}",
        "min_label": f"{format_axis_value(min_value)} EUR",
        "max_label": f"{format_axis_value(max_value)} EUR",
        "years_count": len(rows),
    }


def collapse_date_value_series(series: list[dict], frequency: str = "monthly") -> list[dict]:
    buckets = {}
    for point in series:
        point_date = point.get("date")
        value = point.get("value")
        if point_date is None or value is None:
            continue
        label = bucket_label_for_date(point_date, frequency)
        buckets[label] = {"date": point_date, "value": Decimal(str(value))}
    return [buckets[label] for label in sorted(buckets.keys())]


def resolve_position_start_date(
    opened_on: date | None,
    fallback_date: date | None,
    default_date: date | None = None,
) -> tuple[date, str, bool]:
    if opened_on:
        return opened_on, "Fecha de compra indicada", False
    if fallback_date:
        return fallback_date, "Primer dato historico disponible", True
    if default_date:
        return default_date, "Fecha aproximada sin historico", True
    today = django_timezone.localdate()
    return today, "Fecha aproximada sin historico", True


def estimate_period_totals(
    annual_recurring_cost: Decimal,
    annual_net_dividend_income: Decimal,
    start_date: date,
    end_date: date,
) -> tuple[Decimal, Decimal, int]:
    elapsed_days = max((end_date - start_date).days, 0)
    year_fraction = Decimal(str(elapsed_days)) / Decimal("365")
    maintenance_total = quantize_decimal(annual_recurring_cost * year_fraction, "0.01") or ZERO
    dividend_total = quantize_decimal(annual_net_dividend_income * year_fraction, "0.01") or ZERO
    return maintenance_total, dividend_total, elapsed_days


def build_equity_sale_preview(
    position: EquityPosition,
    sale_price_per_share: Decimal | None = None,
    closed_on: date | None = None,
) -> dict:
    if not position.is_owned:
        return {"available": False}

    history = sorted(position.price_history.all(), key=lambda point: point.price_date)
    fallback_start_date = history[0].price_date if history else position.latest_price_date
    requested_closed_on = closed_on or django_timezone.localdate()
    start_date, start_label, is_estimated = resolve_position_start_date(
        position.opened_on,
        fallback_start_date,
        requested_closed_on,
    )
    effective_closed_on = max(requested_closed_on, start_date)
    sale_price = quantize_decimal(
        Decimal(str(sale_price_per_share if sale_price_per_share is not None else position.current_price_per_share or ZERO)),
        "0.0001",
    ) or ZERO
    gross_sale_value = quantize_decimal(position.shares * sale_price, "0.01") or ZERO
    sale_costs = estimate_broker_costs(
        broker_name=position.broker,
        trade_channel=position.trade_channel,
        trade_amount=gross_sale_value,
        valuation_amount=gross_sale_value,
        annual_dividend_income=position.annual_dividend_income,
        quote_symbol=position.quote_symbol,
    )
    sale_total_cost = quantize_decimal(sale_costs.get("sale_total_cost", ZERO), "0.01") or ZERO
    purchase_cost = quantize_decimal(position.purchase_total_cost, "0.01") or ZERO
    maintenance_total, dividend_total, holding_days = estimate_period_totals(
        position.recurring_cost_used,
        position.net_dividend_income,
        start_date,
        effective_closed_on,
    )
    invested_amount = quantize_decimal(position.invested_amount, "0.01") or ZERO
    committed_capital = quantize_decimal(invested_amount + purchase_cost, "0.01") or ZERO
    net_exit_value = quantize_decimal(gross_sale_value - sale_total_cost, "0.01") or ZERO
    total_costs = quantize_decimal(purchase_cost + sale_total_cost + maintenance_total, "0.01") or ZERO
    net_result = quantize_decimal(
        net_exit_value - invested_amount - purchase_cost - maintenance_total + dividend_total,
        "0.01",
    ) or ZERO
    cumulative_margin_pct = (
        quantize_decimal((net_result / committed_capital) * ONE_HUNDRED, "0.01")
        if committed_capital
        else ZERO
    ) or ZERO
    comparable_returns = build_comparable_return_metrics(cumulative_margin_pct, holding_days)
    annualized_margin_pct = quantize_decimal(comparable_returns["annual_equivalent_return_pct"], "0.01") or ZERO
    monthly_equivalent_return_pct = quantize_decimal(comparable_returns["monthly_equivalent_return_pct"], "0.01") or ZERO

    return {
        "available": True,
        "start_date": start_date,
        "start_date_label": start_date.isoformat(),
        "start_date_source_label": start_label,
        "start_date_is_estimated": is_estimated,
        "closed_on": effective_closed_on,
        "closed_on_label": effective_closed_on.isoformat(),
        "holding_days": holding_days,
        "shares": position.shares,
        "sale_price_per_share": sale_price,
        "invested_amount": invested_amount,
        "purchase_cost": purchase_cost,
        "sale_total_cost": sale_total_cost,
        "annual_recurring_cost": quantize_decimal(position.recurring_cost_used, "0.01") or ZERO,
        "annual_net_dividend_income": quantize_decimal(position.net_dividend_income, "0.01") or ZERO,
        "maintenance_total": maintenance_total,
        "dividend_total": dividend_total,
        "gross_sale_value": gross_sale_value,
        "net_exit_value": net_exit_value,
        "total_costs": total_costs,
        "net_result": net_result,
        "committed_capital": committed_capital,
        "cumulative_margin_pct": cumulative_margin_pct,
        "monthly_equivalent_return_pct": monthly_equivalent_return_pct,
        "annualized_margin_pct": annualized_margin_pct,
        "cost_profile_label": sale_costs.get("profile_label") or position.broker,
        "market_scope_label": sale_costs.get("market_scope_label") or "",
        "trade_channel_label": sale_costs.get("trade_channel_label") or position.get_trade_channel_display(),
    }


def build_active_equity_investment_ticket(position: EquityPosition) -> dict | None:
    history = sorted(position.price_history.all(), key=lambda point: point.price_date)
    fallback_start_date = history[0].price_date if history else position.latest_price_date
    default_end_date = position.latest_price_date or (history[-1].price_date if history else django_timezone.localdate())
    start_date, start_label, is_estimated = resolve_position_start_date(position.opened_on, fallback_start_date, default_end_date)
    end_date = max(default_end_date, start_date)
    purchase_cost = quantize_decimal(position.purchase_total_cost, "0.01") or ZERO
    sale_cost_estimate = quantize_decimal(position.sale_total_cost_estimate, "0.01") or ZERO
    maintenance_total, dividend_total, holding_days = estimate_period_totals(
        position.recurring_cost_used,
        position.net_dividend_income,
        start_date,
        end_date,
    )

    live_series = [{"date": start_date, "value": quantize_decimal(position.invested_amount - purchase_cost, "0.01") or ZERO}]
    monthly_history = collapse_date_value_series(
        [
            {
                "date": point.price_date,
                "value": quantize_decimal(
                    (position.shares * point.close_price)
                    - purchase_cost
                    - (quantize_decimal(position.recurring_cost_used * (Decimal(str(max((point.price_date - start_date).days, 0))) / Decimal("365")), "0.01") or ZERO)
                    + (quantize_decimal(position.net_dividend_income * (Decimal(str(max((point.price_date - start_date).days, 0))) / Decimal("365")), "0.01") or ZERO),
                    "0.01",
                )
                or ZERO,
            }
            for point in history
            if point.price_date >= start_date
        ]
    )
    live_series.extend(monthly_history)
    latest_live_value = quantize_decimal(position.current_value - purchase_cost - maintenance_total + dividend_total, "0.01") or ZERO
    if not live_series or live_series[-1]["date"] != end_date:
        live_series.append({"date": end_date, "value": latest_live_value})
    else:
        live_series[-1]["value"] = latest_live_value

    committed_capital = quantize_decimal(position.invested_amount + purchase_cost, "0.01") or ZERO
    net_exit_value = quantize_decimal(latest_live_value - sale_cost_estimate, "0.01") or ZERO
    net_result = quantize_decimal(net_exit_value - position.invested_amount, "0.01") or ZERO
    cumulative_margin_pct = ((net_result / committed_capital) * ONE_HUNDRED) if committed_capital else ZERO
    comparable_returns = build_comparable_return_metrics(cumulative_margin_pct, holding_days)
    annualized_margin_pct = comparable_returns["annual_equivalent_return_pct"] or ZERO
    monthly_equivalent_return_pct = comparable_returns["monthly_equivalent_return_pct"] or ZERO
    return {
        "status": "active",
        "status_label": "En cartera",
        "ticker": position.ticker,
        "company_name": position.company_name,
        "broker": position.broker,
        "position": position,
        "start_date": start_date,
        "end_date": end_date,
        "start_date_label": start_date.isoformat(),
        "end_date_label": end_date.isoformat(),
        "start_date_source_label": start_label,
        "start_date_is_estimated": is_estimated,
        "committed_capital": committed_capital,
        "purchase_cost": purchase_cost,
        "sale_cost": sale_cost_estimate,
        "maintenance_total": maintenance_total,
        "dividend_total": dividend_total,
        "costs_total": quantize_decimal(purchase_cost + sale_cost_estimate + maintenance_total, "0.01") or ZERO,
        "current_or_sale_value": quantize_decimal(position.current_value, "0.01") or ZERO,
        "net_live_value": latest_live_value,
        "net_exit_value": net_exit_value,
        "net_result": net_result,
        "cumulative_margin_pct": quantize_decimal(cumulative_margin_pct, "0.01") or ZERO,
        "monthly_equivalent_return_pct": quantize_decimal(monthly_equivalent_return_pct, "0.01") or ZERO,
        "annualized_margin_pct": quantize_decimal(annualized_margin_pct, "0.01") or ZERO,
        "holding_days": holding_days,
        "value_series": live_series,
        "value_chart": build_single_value_history_chart(live_series),
        "notes": "La rentabilidad actual descuenta compra y mantenimiento. La salida neta incluye tambien el coste estimado de venta si cerraras hoy.",
    }


def build_closed_equity_investment_ticket(position: EquityClosedPosition) -> dict:
    archived_series = collapse_date_value_series(
        [
            {
                "date": date.fromisoformat(str(point.get("date"))),
                "value": Decimal(str(point.get("market_value") or "0")),
            }
            for point in (position.archived_price_history or [])
            if point.get("date")
        ]
    )
    fallback_start_date = archived_series[0]["date"] if archived_series else position.closed_on
    start_date, start_label, is_estimated = resolve_position_start_date(position.opened_on, fallback_start_date, position.closed_on)
    end_date = position.closed_on
    total_days = max((end_date - start_date).days, 0)
    live_series = [{"date": start_date, "value": quantize_decimal(position.invested_amount - position.purchase_total_cost, "0.01") or ZERO}]
    for point in archived_series:
        if point["date"] < start_date:
            continue
        elapsed_ratio = Decimal(str(max((point["date"] - start_date).days, 0))) / Decimal(str(max(total_days, 1)))
        accrued_maintenance = quantize_decimal(position.maintenance_cost_total * elapsed_ratio, "0.01") or ZERO
        accrued_dividends = quantize_decimal(position.net_dividend_income_total * elapsed_ratio, "0.01") or ZERO
        live_series.append(
            {
                "date": point["date"],
                "value": quantize_decimal(point["value"] - position.purchase_total_cost - accrued_maintenance + accrued_dividends, "0.01") or ZERO,
            }
        )
    net_sale_value = quantize_decimal(position.net_sale_value - position.purchase_total_cost - position.maintenance_cost_total + position.net_dividend_income_total, "0.01") or ZERO
    if not live_series or live_series[-1]["date"] != end_date:
        live_series.append({"date": end_date, "value": net_sale_value})
    else:
        live_series[-1]["value"] = net_sale_value

    committed_capital = quantize_decimal(position.committed_capital, "0.01") or ZERO
    net_result = quantize_decimal(position.net_result, "0.01") or ZERO
    cumulative_margin_pct = quantize_decimal(position.cumulative_margin_pct, "0.01") or ZERO
    comparable_returns = build_comparable_return_metrics(cumulative_margin_pct, max((end_date - start_date).days, 0))
    annualized_margin_pct = quantize_decimal(comparable_returns["annual_equivalent_return_pct"], "0.01") or ZERO
    monthly_equivalent_return_pct = quantize_decimal(comparable_returns["monthly_equivalent_return_pct"], "0.01") or ZERO
    return {
        "status": "closed",
        "status_label": "Vendida",
        "ticker": position.ticker,
        "company_name": position.company_name,
        "broker": position.broker,
        "position": position,
        "start_date": start_date,
        "end_date": end_date,
        "start_date_label": start_date.isoformat(),
        "end_date_label": end_date.isoformat(),
        "start_date_source_label": start_label,
        "start_date_is_estimated": is_estimated,
        "committed_capital": committed_capital,
        "purchase_cost": quantize_decimal(position.purchase_total_cost, "0.01") or ZERO,
        "sale_cost": quantize_decimal(position.sale_total_cost, "0.01") or ZERO,
        "maintenance_total": quantize_decimal(position.maintenance_cost_total, "0.01") or ZERO,
        "dividend_total": quantize_decimal(position.net_dividend_income_total, "0.01") or ZERO,
        "costs_total": quantize_decimal(position.purchase_total_cost + position.sale_total_cost + position.maintenance_cost_total, "0.01") or ZERO,
        "current_or_sale_value": quantize_decimal(position.gross_sale_value, "0.01") or ZERO,
        "net_live_value": net_sale_value,
        "net_exit_value": net_sale_value,
        "net_result": net_result,
        "cumulative_margin_pct": cumulative_margin_pct,
        "monthly_equivalent_return_pct": monthly_equivalent_return_pct,
        "annualized_margin_pct": annualized_margin_pct,
        "holding_days": max((end_date - start_date).days, 0),
        "value_series": live_series,
        "value_chart": build_single_value_history_chart(live_series),
        "notes": "Resultado cerrado. La venta ya descuenta compra, mantenimiento y el coste real estimado de salida.",
    }


def build_equity_investment_journey_context(
    open_positions: list[EquityPosition],
    closed_positions: list[EquityClosedPosition],
) -> dict:
    active_tickets = [ticket for ticket in (build_active_equity_investment_ticket(position) for position in open_positions if position.is_owned) if ticket]
    closed_tickets = [build_closed_equity_investment_ticket(position) for position in closed_positions]
    all_tickets = [*active_tickets, *closed_tickets]
    if not all_tickets:
        return {
            "available": False,
            "active_tickets": [],
            "closed_tickets": [],
            "value_chart": {"available": False},
            "profit_chart": {"available": False},
        }

    date_points = sorted({point["date"] for ticket in all_tickets for point in ticket["value_series"]})
    aggregated_value_series = []
    aggregated_profit_series = []
    yearly_snapshot_buckets: dict[int, dict] = {}
    yearly_committed_points: dict[int, list[Decimal]] = defaultdict(list)

    for point_date in date_points:
        total_value = ZERO
        total_committed = ZERO
        for ticket in all_tickets:
            applicable_points = [point for point in ticket["value_series"] if point["date"] <= point_date]
            if not applicable_points or point_date < ticket["start_date"]:
                continue
            latest_point = applicable_points[-1]
            total_value += latest_point["value"]
            total_committed += ticket["committed_capital"]
        total_profit = total_value - total_committed
        aggregated_value_series.append({"date": point_date, "value": quantize_decimal(total_value, "0.01") or ZERO})
        aggregated_profit_series.append({"date": point_date, "value": quantize_decimal(total_profit, "0.01") or ZERO})
        yearly_snapshot_buckets[point_date.year] = {
            "date": point_date,
            "value": quantize_decimal(total_value, "0.01") or ZERO,
            "profit": quantize_decimal(total_profit, "0.01") or ZERO,
            "committed": quantize_decimal(total_committed, "0.01") or ZERO,
        }
        yearly_committed_points[point_date.year].append(quantize_decimal(total_committed, "0.01") or ZERO)

    total_committed_capital = sum((ticket["committed_capital"] for ticket in all_tickets), ZERO)
    total_net_result = sum((ticket["net_result"] for ticket in all_tickets), ZERO)
    total_realized_result = sum((ticket["net_result"] for ticket in closed_tickets), ZERO)
    total_live_result = sum((ticket["net_result"] for ticket in active_tickets), ZERO)
    total_purchase_costs = sum((ticket["purchase_cost"] for ticket in all_tickets), ZERO)
    total_sale_costs = sum((ticket["sale_cost"] for ticket in all_tickets), ZERO)
    total_maintenance_costs = sum((ticket["maintenance_total"] for ticket in all_tickets), ZERO)
    closed_sale_cost_total = sum((ticket["sale_cost"] for ticket in closed_tickets), ZERO)
    open_sale_cost_reserve_total = sum((ticket["sale_cost"] for ticket in active_tickets), ZERO)
    costs_paid_total = total_purchase_costs + total_maintenance_costs + closed_sale_cost_total
    total_costs = sum((ticket["costs_total"] for ticket in all_tickets), ZERO)
    total_dividends = sum((ticket["dividend_total"] for ticket in all_tickets), ZERO)
    earliest_start = min((ticket["start_date"] for ticket in all_tickets), default=django_timezone.localdate())
    latest_end = max((ticket["end_date"] for ticket in all_tickets), default=django_timezone.localdate())
    total_days = max((latest_end - earliest_start).days, 1)
    years_operating = Decimal(str(total_days)) / Decimal("365")
    cumulative_margin_pct = ((total_net_result / total_committed_capital) * ONE_HUNDRED) if total_committed_capital else ZERO
    comparable_returns = build_comparable_return_metrics(cumulative_margin_pct, total_days)
    monthly_equivalent_return_pct = comparable_returns["monthly_equivalent_return_pct"] or ZERO
    annualized_margin_pct = comparable_returns["annual_equivalent_return_pct"] or ZERO
    annual_rows = []
    previous_profit = ZERO
    for year in sorted(yearly_snapshot_buckets):
        snapshot = yearly_snapshot_buckets[year]
        year_profit = snapshot["profit"]
        annual_result = quantize_decimal(year_profit - previous_profit, "0.01") or ZERO
        average_committed_capital = average_decimal(yearly_committed_points.get(year, [])) or snapshot["committed"] or ZERO
        annual_margin_pct = (
            quantize_decimal((annual_result / average_committed_capital) * ONE_HUNDRED, "0.01")
            if average_committed_capital
            else ZERO
        )
        annual_rows.append(
            {
                "year": year,
                "year_end_value": snapshot["value"],
                "year_end_profit": year_profit,
                "average_committed_capital": quantize_decimal(average_committed_capital, "0.01") or ZERO,
                "net_result": annual_result,
                "margin_pct": annual_margin_pct,
            }
        )
        previous_profit = year_profit

    average_annual_result = (
        quantize_decimal(total_net_result / years_operating, "0.01")
        if years_operating
        else ZERO
    ) or ZERO
    cost_ratio_pct = (
        quantize_decimal((total_costs / total_committed_capital) * ONE_HUNDRED, "0.01")
        if total_committed_capital
        else ZERO
    ) or ZERO
    dividend_yield_pct = (
        quantize_decimal((total_dividends / total_committed_capital) * ONE_HUNDRED, "0.01")
        if total_committed_capital
        else ZERO
    ) or ZERO
    current_year_row = annual_rows[-1] if annual_rows else None
    previous_year_row = annual_rows[-2] if len(annual_rows) >= 2 else None
    best_year = max(annual_rows, key=lambda row: row["net_result"], default=None)
    worst_year = min(annual_rows, key=lambda row: row["net_result"], default=None)

    active_tickets.sort(key=lambda ticket: (-ticket["current_or_sale_value"], ticket["company_name"]))
    closed_tickets.sort(key=lambda ticket: (ticket["end_date"], ticket["company_name"]), reverse=True)
    return {
        "available": True,
        "active_tickets": active_tickets,
        "closed_tickets": closed_tickets,
        "tickets_count": len(all_tickets),
        "active_count": len(active_tickets),
        "closed_count": len(closed_tickets),
        "value_chart": build_single_value_history_chart(aggregated_value_series),
        "profit_chart": build_single_value_history_chart(aggregated_profit_series),
        "current_net_value": aggregated_value_series[-1]["value"] if aggregated_value_series else ZERO,
        "cumulative_net_result": quantize_decimal(total_net_result, "0.01") or ZERO,
        "cumulative_margin_pct": quantize_decimal(cumulative_margin_pct, "0.01") or ZERO,
        "monthly_equivalent_return_pct": quantize_decimal(monthly_equivalent_return_pct, "0.01") or ZERO,
        "annualized_margin_pct": quantize_decimal(annualized_margin_pct, "0.01") or ZERO,
        "years_operating": quantize_decimal(years_operating, "0.01") or ZERO,
        "average_annual_result": average_annual_result,
        "cost_ratio_pct": cost_ratio_pct,
        "dividend_yield_pct": dividend_yield_pct,
        "realized_net_result": quantize_decimal(total_realized_result, "0.01") or ZERO,
        "live_net_result": quantize_decimal(total_live_result, "0.01") or ZERO,
        "committed_capital_total": quantize_decimal(total_committed_capital, "0.01") or ZERO,
        "costs_total": quantize_decimal(total_costs, "0.01") or ZERO,
        "costs_paid_total": quantize_decimal(costs_paid_total, "0.01") or ZERO,
        "purchase_cost_total": quantize_decimal(total_purchase_costs, "0.01") or ZERO,
        "sale_cost_total": quantize_decimal(total_sale_costs, "0.01") or ZERO,
        "closed_sale_cost_total": quantize_decimal(closed_sale_cost_total, "0.01") or ZERO,
        "open_sale_cost_reserve_total": quantize_decimal(open_sale_cost_reserve_total, "0.01") or ZERO,
        "maintenance_cost_total": quantize_decimal(total_maintenance_costs, "0.01") or ZERO,
        "dividends_total": quantize_decimal(total_dividends, "0.01") or ZERO,
        "holding_start_label": earliest_start.isoformat() if earliest_start else "",
        "holding_end_label": latest_end.isoformat() if latest_end else "",
        "annual_rows": annual_rows,
        "annual_result_chart": build_annual_result_chart(annual_rows),
        "current_year_row": current_year_row,
        "previous_year_row": previous_year_row,
        "best_year": best_year,
        "worst_year": worst_year,
        "estimated_start_count": sum(1 for ticket in all_tickets if ticket["start_date_is_estimated"]),
    }


@transaction.atomic
def archive_equity_position_sale(
    position: EquityPosition,
    closed_on: date,
    sale_price_per_share: Decimal,
    notes: str = "",
) -> EquityClosedPosition:
    fallback_start_date = position.price_history.order_by("price_date").values_list("price_date", flat=True).first()
    start_date = position.opened_on or fallback_start_date or closed_on
    maintenance_total, dividend_total, _ = estimate_period_totals(
        position.recurring_cost_used,
        position.net_dividend_income,
        start_date,
        closed_on,
    )
    sale_trade_amount = quantize_decimal(position.shares * sale_price_per_share, "0.01") or ZERO
    sale_costs = estimate_broker_costs(
        broker_name=position.broker,
        trade_channel=position.trade_channel,
        trade_amount=sale_trade_amount,
        valuation_amount=sale_trade_amount,
        annual_dividend_income=position.annual_dividend_income,
        quote_symbol=position.quote_symbol,
    )
    archived = EquityClosedPosition.objects.create(
        ownership_category=position.ownership_category,
        broker=position.broker,
        ticker=position.ticker,
        quote_symbol=position.quote_symbol,
        company_name=position.company_name,
        trade_channel=position.trade_channel,
        benchmark_symbol=position.benchmark_symbol,
        benchmark_name=position.benchmark_name,
        opened_on=position.opened_on,
        closed_on=closed_on,
        shares=position.shares,
        average_cost_per_share=position.average_cost_per_share,
        sale_price_per_share=sale_price_per_share,
        purchase_total_cost=quantize_decimal(position.purchase_total_cost, "0.01") or ZERO,
        sale_total_cost=quantize_decimal(sale_costs.get("sale_total_cost", ZERO), "0.01") or ZERO,
        maintenance_cost_total=maintenance_total,
        net_dividend_income_total=dividend_total,
        archived_price_history=serialize_equity_market_value_history(position, closed_on, sale_price_per_share),
        notes="\n\n".join(filter(None, [position.notes, notes])),
    )
    position.delete()
    return archived


def average_decimal(values: list[Decimal]) -> Decimal | None:
    filtered = [value for value in values if value is not None]
    if not filtered:
        return None
    return sum(filtered, ZERO) / Decimal(len(filtered))


def weighted_average_decimal(values: list[tuple[Decimal | None, Decimal | None]]) -> Decimal | None:
    filtered = [
        (value, weight)
        for value, weight in values
        if value is not None and weight is not None and weight > ZERO
    ]
    if not filtered:
        return None
    total_weight = sum((weight for _, weight in filtered), ZERO)
    if total_weight <= ZERO:
        return None
    return sum((value * weight for value, weight in filtered), ZERO) / total_weight


def median_decimal(values: list[Decimal]) -> Decimal | None:
    filtered = sorted(value for value in values if value is not None)
    if not filtered:
        return None
    middle = len(filtered) // 2
    if len(filtered) % 2:
        return filtered[middle]
    return (filtered[middle - 1] + filtered[middle]) / Decimal("2")


def percentage_change(current: Decimal | None, reference: Decimal | None) -> Decimal | None:
    if current is None or reference in {None, ZERO}:
        return None
    return ((current / reference) - Decimal("1")) * Decimal("100")


def infer_reference_change_mode(reference_profile: str) -> str:
    if reference_profile == EquityPosition.ReferenceProfile.EURIBOR_12M:
        return "absolute"
    return "pct"


def calculate_series_change(
    current: Decimal | None,
    reference: Decimal | None,
    *,
    change_mode: str = "pct",
) -> Decimal | None:
    if current is None or reference is None:
        return None
    if change_mode == "absolute":
        return current - reference
    return percentage_change(current, reference)


def change_unit_label(change_mode: str) -> str:
    return "pp" if change_mode == "absolute" else "%"


def format_compact_currency_value(value: Decimal | None, currency_code: str = "") -> str:
    if value is None:
        return "-"

    absolute = abs(value)
    if absolute >= Decimal("1000000000"):
        scaled_value = value / Decimal("1000000000")
        label = f"{scaled_value:.1f} mM"
    elif absolute >= Decimal("1000000"):
        scaled_value = value / Decimal("1000000")
        label = f"{scaled_value:.1f} M"
    elif absolute >= Decimal("1000"):
        scaled_value = value / Decimal("1000")
        label = f"{scaled_value:.1f} k"
    else:
        label = f"{value:.0f}"
    return f"{label} {currency_code}".strip()


def describe_net_income_trend(net_income_rows: list[dict]) -> tuple[str, str]:
    if len(net_income_rows) < 2:
        return "Sin serie suficiente", "Todavia no hay tres cierres completos para leer la trayectoria del beneficio."

    values = [row["value"] for row in net_income_rows]
    latest_value = values[-1]
    previous_value = values[-2]
    oldest_value = values[0]

    if latest_value > ZERO and all(current > previous for previous, current in zip(values, values[1:])):
        return "Mejora", "Los beneficios netos encadenan varios ejercicios al alza."
    if latest_value <= ZERO:
        return "Presion", "El ultimo cierre entra en beneficio negativo y debilita la lectura del ciclo."
    if all(current < previous for previous, current in zip(values, values[1:])):
        return "Deterioro", "El beneficio neto viene enfriandose de forma continuada."
    if latest_value > previous_value and previous_value <= oldest_value:
        return "Recuperacion", "El beneficio habia flojeado, pero el ultimo ejercicio vuelve a mejorar."
    if latest_value < previous_value and previous_value >= oldest_value:
        return "Enfriamiento", "El ultimo ejercicio enfria una etapa de mejora previa."
    return "Mixto", "La serie de beneficios no marca una direccion limpia todavia."


def build_equity_fundamentals_summary(position: EquityPosition, force_live_fetch: bool = False) -> dict:
    if not force_live_fetch and not should_fetch_equity_fundamentals():
        return {
            "available": False,
            "note": "La carga de fundamentales en vivo esta desactivada.",
            "net_income_rows": [],
            "derived_per": None,
        }

    if not position.quote_symbol:
        return {
            "available": False,
            "note": "No hay simbolo de cotizacion para buscar los fundamentales.",
            "net_income_rows": [],
            "derived_per": None,
        }

    try:
        snapshot = fetch_equity_fundamentals(position.quote_symbol)
    except Exception as exc:
        return {
            "available": False,
            "note": f"No se han podido cargar los fundamentales recientes: {exc}",
            "net_income_rows": [],
            "derived_per": None,
        }

    net_income_rows = list(snapshot.get("net_income_rows", []))
    if not net_income_rows and snapshot.get("market_cap") is None:
        return {
            "available": False,
            "note": "La fuente externa no ha devuelto beneficio neto ni capitalizacion reciente.",
            "net_income_rows": [],
            "derived_per": None,
        }

    currency_code = snapshot.get("currency_code") or ""
    trend_label, trend_note = describe_net_income_trend(net_income_rows)
    net_income_delta_pct = None
    if len(net_income_rows) >= 2:
        net_income_delta_pct = percentage_change(net_income_rows[-1]["value"], net_income_rows[0]["value"])

    prepared_rows = []
    for row in reversed(net_income_rows):
        prepared_rows.append(
            {
                **row,
                "value_label": format_compact_currency_value(row.get("value"), row.get("currency_code") or currency_code),
                "year_label": str(row["as_of_date"].year),
            }
        )

    market_cap = snapshot.get("market_cap")
    market_cap_as_of_date = snapshot.get("market_cap_as_of_date")
    latest_net_income_row = max(net_income_rows, key=lambda item: item["as_of_date"]) if net_income_rows else None
    latest_net_income_value = latest_net_income_row.get("value") if latest_net_income_row else None
    derived_per = None
    if market_cap is not None and latest_net_income_value not in {None, ZERO} and latest_net_income_value > ZERO:
        derived_per = quantize_decimal(market_cap / latest_net_income_value, "0.01")
    return {
        "available": True,
        "currency_code": currency_code,
        "net_income_rows": prepared_rows,
        "trend_label": trend_label,
        "trend_note": trend_note,
        "net_income_delta_pct": net_income_delta_pct,
        "market_cap": market_cap,
        "market_cap_label": format_compact_currency_value(market_cap, currency_code),
        "market_cap_as_of_date": market_cap_as_of_date,
        "market_cap_date_label": market_cap_as_of_date.isoformat() if market_cap_as_of_date else "",
        "latest_net_income_value": latest_net_income_value,
        "latest_net_income_as_of_date": latest_net_income_row.get("as_of_date") if latest_net_income_row else None,
        "derived_per": derived_per,
        "derived_per_label": f"{derived_per:.1f}" if derived_per is not None else "",
        "note": "Beneficio neto anual y capitalizacion segun la serie fundamental mas reciente disponible.",
    }


def infer_reference_frequency_from_profile(reference_profile: str) -> str:
    if reference_profile == EquityPosition.ReferenceProfile.SPAIN_HOUSE_PRICE:
        return "quarterly"
    return "monthly"


def infer_reference_frequency(position: EquityPosition) -> str:
    return infer_reference_frequency_from_profile(position.reference_profile)


def bucket_label_for_date(value: date, frequency: str) -> str:
    if frequency == "yearly":
        return f"{value.year}"
    if frequency == "quarterly":
        quarter = ((value.month - 1) // 3) + 1
        return f"{value.year}-Q{quarter}"
    return value.strftime("%Y-%m")


def collapse_history_to_frequency(history, frequency: str) -> list:
    buckets = {}
    for point in history:
        buckets[bucket_label_for_date(point.price_date, frequency)] = point
    return [buckets[label] for label in sorted(buckets.keys())]


def month_end_date(value: date) -> date:
    return date(value.year, value.month, monthrange(value.year, value.month)[1])


def month_start_date(value: date) -> date:
    return date(value.year, value.month, 1)


def trim_current_partial_month_for_long_horizon(history, *, as_of: date | None = None, min_points: int = 24) -> list:
    rows = list(history or [])
    if len(rows) <= min_points:
        return rows
    as_of = as_of or django_timezone.localdate()
    latest_date = rows[-1].price_date
    if latest_date is None:
        return rows
    current_month_start = month_start_date(as_of)
    latest_month_start = month_start_date(latest_date)
    current_month_is_open = as_of < month_end_date(as_of)
    if latest_date > as_of or latest_month_start != current_month_start or not current_month_is_open:
        return rows
    trimmed = [point for point in rows if point.price_date < current_month_start]
    return trimmed if len(trimmed) >= min_points else rows


def build_return_series_from_collapsed_rows(
    collapsed_rows: list[dict],
    primary_key: str,
    secondary_key: str,
    *,
    secondary_change_mode: str = "pct",
) -> tuple[list[Decimal], list[Decimal]]:
    primary_returns = []
    secondary_returns = []
    for previous, current in zip(collapsed_rows, collapsed_rows[1:]):
        primary_return = percentage_change(current.get(primary_key), previous.get(primary_key))
        secondary_return = calculate_series_change(
            current.get(secondary_key),
            previous.get(secondary_key),
            change_mode=secondary_change_mode,
        )
        if primary_return is None or secondary_return is None:
            continue
        primary_returns.append(primary_return)
        secondary_returns.append(secondary_return)
    return primary_returns, secondary_returns


def correlation_window_size(frequency: str) -> int:
    if frequency == "quarterly":
        return QUARTERLY_CORRELATION_WINDOW
    return MONTHLY_CORRELATION_WINDOW


def recent_correlation_window_size(frequency: str) -> int:
    if frequency == "quarterly":
        return QUARTERLY_RECENT_WINDOW
    return MONTHLY_RECENT_WINDOW


def calculate_years_between(start_date: date | None, end_date: date | None) -> Decimal:
    if not start_date or not end_date or end_date <= start_date:
        return ZERO
    return Decimal(str(round((end_date - start_date).days / 365.25, 2)))


def calculate_series_cagr_pct(
    start_value: Decimal | None,
    end_value: Decimal | None,
    years_covered: Decimal | None,
) -> Decimal | None:
    if start_value in {None, ZERO} or end_value is None or years_covered in {None, ZERO}:
        return None
    base = float(end_value / start_value)
    years_float = float(years_covered)
    if base <= 0 or years_float <= 0:
        return None
    annualized = (base ** (1 / years_float) - 1) * 100
    return Decimal(str(round(annualized, 4)))


def calculate_max_drawdown_pct(values: list[Decimal]) -> tuple[Decimal | None, Decimal | None]:
    filtered = [value for value in values if value is not None]
    if len(filtered) < 2:
        return None, None

    peak = filtered[0]
    max_drawdown_pct = ZERO
    for value in filtered:
        if value > peak:
            peak = value
        drawdown_pct = percentage_change(value, peak)
        if drawdown_pct is not None and drawdown_pct < max_drawdown_pct:
            max_drawdown_pct = drawdown_pct

    current_peak = max(filtered)
    current_drawdown_pct = percentage_change(filtered[-1], current_peak)
    return max_drawdown_pct, current_drawdown_pct


def describe_cycle_phase(
    current_drawdown_pct: Decimal | None,
    one_year_return_pct: Decimal | None,
    cagr_pct: Decimal | None,
) -> str:
    if current_drawdown_pct is None:
        return "Sin ciclo"
    one_year_return_pct = one_year_return_pct or ZERO
    cagr_pct = cagr_pct or ZERO
    if current_drawdown_pct <= Decimal("-20.00") and one_year_return_pct < 0:
        return "Correccion"
    if current_drawdown_pct <= Decimal("-8.00") and one_year_return_pct >= 0:
        return "Recuperacion"
    if current_drawdown_pct >= Decimal("-6.00") and one_year_return_pct >= 0 and cagr_pct >= 0:
        return "Expansion"
    return "Transicion"


def build_cycle_metrics(history, max_days: int | None = LONG_ANALYSIS_DAYS) -> dict:
    if not history:
        return {
            "available": False,
            "years_covered": ZERO,
            "cycle_phase": "Sin ciclo",
        }

    latest_date = history[-1].price_date
    if max_days is None:
        window = [point for point in history if point.price_date <= latest_date]
    else:
        window = filter_history_window(
            history,
            start_date=latest_date - timedelta(days=max_days),
            end_date=latest_date,
        )
    monthly_history = collapse_history_to_frequency(window, "monthly")
    if len(monthly_history) < 6:
        return {
            "available": False,
            "years_covered": calculate_years_between(window[0].price_date, window[-1].price_date) if len(window) >= 2 else ZERO,
            "cycle_phase": "Sin ciclo",
        }

    monthly_returns = []
    for previous, current in zip(monthly_history, monthly_history[1:]):
        period_return = percentage_change(current.close_price, previous.close_price)
        if period_return is not None:
            monthly_returns.append(period_return)

    monthly_volatility_pct = standard_deviation_decimal(monthly_returns)
    annualized_volatility_pct = (
        monthly_volatility_pct * Decimal(str(round(math.sqrt(12), 4)))
        if monthly_volatility_pct is not None
        else None
    )
    years_covered = calculate_years_between(monthly_history[0].price_date, monthly_history[-1].price_date)
    cagr_pct = calculate_series_cagr_pct(
        monthly_history[0].close_price,
        monthly_history[-1].close_price,
        years_covered,
    )
    max_drawdown_pct, current_drawdown_pct = calculate_max_drawdown_pct(
        [point.close_price for point in monthly_history]
    )
    one_year_snapshot = build_period_snapshot(
        monthly_history,
        "1Y",
        start_date=latest_date - timedelta(days=365),
        end_date=latest_date,
    )
    three_year_snapshot = build_period_snapshot(
        monthly_history,
        "3Y",
        start_date=latest_date - timedelta(days=365 * 3),
        end_date=latest_date,
    )
    yearly_history = collapse_history_to_frequency(window, "yearly")
    yearly_returns = []
    for previous, current in zip(yearly_history, yearly_history[1:]):
        yearly_return = percentage_change(current.close_price, previous.close_price)
        if yearly_return is not None:
            yearly_returns.append(yearly_return)
    positive_year_ratio_pct = None
    if yearly_returns:
        positive_year_ratio_pct = Decimal(sum(1 for value in yearly_returns if value > 0)) * ONE_HUNDRED / Decimal(
            len(yearly_returns)
        )

    return {
        "available": True,
        "start_date": monthly_history[0].price_date,
        "end_date": monthly_history[-1].price_date,
        "years_covered": years_covered,
        "monthly_observations_count": len(monthly_history),
        "yearly_observations_count": len(yearly_returns),
        "monthly_volatility_pct": monthly_volatility_pct,
        "annualized_volatility_pct": annualized_volatility_pct,
        "cagr_pct": cagr_pct,
        "max_drawdown_pct": max_drawdown_pct,
        "current_drawdown_pct": current_drawdown_pct,
        "cycle_total_return_pct": percentage_change(monthly_history[-1].close_price, monthly_history[0].close_price),
        "one_year_return_pct": one_year_snapshot.get("stock_return_pct") if one_year_snapshot.get("available") else None,
        "three_year_return_pct": three_year_snapshot.get("stock_return_pct") if three_year_snapshot.get("available") else None,
        "positive_year_ratio_pct": positive_year_ratio_pct,
        "cycle_phase": describe_cycle_phase(
            current_drawdown_pct,
            one_year_snapshot.get("stock_return_pct") if one_year_snapshot.get("available") else None,
            cagr_pct,
        ),
    }


def build_market_history_points(points: list[dict], end_date: date | None = None) -> list[MarketHistoryPoint]:
    history = []
    for point in points or []:
        point_date = point.get("date")
        close_price = point.get("close")
        if point_date is None or close_price in {None, ""}:
            continue
        if end_date is not None and point_date > end_date:
            continue
        try:
            history.append(
                MarketHistoryPoint(
                    price_date=point_date,
                    close_price=Decimal(str(close_price)),
                )
            )
        except Exception:
            continue
    return history


def build_normalized_history_path(window: list[MarketHistoryPoint]) -> list[Decimal]:
    if len(window) < 2 or window[0].close_price in {None, ZERO}:
        return []

    baseline_price = window[0].close_price
    path = [ZERO]
    for point in window[1:]:
        change_pct = percentage_change(point.close_price, baseline_price)
        path.append(change_pct if change_pct is not None else ZERO)
    return path


def build_history_step_returns(window: list[MarketHistoryPoint]) -> list[Decimal]:
    step_returns = []
    for previous, current in zip(window, window[1:]):
        period_return = percentage_change(current.close_price, previous.close_price)
        step_returns.append(period_return if period_return is not None else ZERO)
    return step_returns


def average_absolute_difference(left: list[Decimal], right: list[Decimal]) -> Decimal | None:
    if len(left) != len(right) or not left:
        return None
    differences = [abs(left_value - right_value) for left_value, right_value in zip(left, right)]
    return average_decimal(differences)


def build_reference_cycle_template_from_series(
    reference_series: MarketSeries | None,
    *,
    latest_date: date | None,
    years: int = 5,
    step_months: int = 6,
) -> dict:
    if reference_series is None or years <= 0 or step_months <= 0:
        return {"available": False}

    monthly_history = collapse_history_to_frequency(
        build_market_history_points(reference_series.points, end_date=latest_date),
        "monthly",
    )
    step_count = int((years * 12) / step_months)
    recent_months = max(REFERENCE_CYCLE_TEMPLATE_RECENT_MONTHS, step_months * 2)
    minimum_points = recent_months + (step_count * step_months) + 1
    if len(monthly_history) < minimum_points:
        return {
            "available": False,
            "reference_label": reference_series.name or reference_series.symbol,
            "years_covered": calculate_years_between(
                monthly_history[0].price_date if monthly_history else None,
                monthly_history[-1].price_date if monthly_history else None,
            ),
            "match_windows_count": 0,
            "selected_windows_count": 0,
        }

    current_metrics = build_cycle_metrics(monthly_history, max_days=None)
    current_window = monthly_history[-(recent_months + 1) :]
    current_shape_path = build_normalized_history_path(current_window)
    current_return_shape = build_history_step_returns(current_window)
    current_recent_return_pct = percentage_change(
        current_window[-1].close_price,
        current_window[0].close_price,
    ) or ZERO
    current_one_year_return_pct = current_metrics.get("one_year_return_pct") or ZERO
    current_drawdown_pct = current_metrics.get("current_drawdown_pct") or ZERO
    current_volatility_pct = current_metrics.get("annualized_volatility_pct") or ZERO
    current_phase = current_metrics.get("cycle_phase") or "Sin ciclo"
    matches = []

    last_anchor_index = len(monthly_history) - (step_count * step_months) - 1
    for anchor_index in range(recent_months, last_anchor_index + 1):
        anchor_history = monthly_history[: anchor_index + 1]
        anchor_metrics = build_cycle_metrics(anchor_history, max_days=None)
        if not anchor_metrics.get("available"):
            continue
        anchor_window = monthly_history[(anchor_index - recent_months) : (anchor_index + 1)]
        if len(anchor_window) != len(current_window):
            continue

        anchor_shape_path = build_normalized_history_path(anchor_window)
        anchor_return_shape = build_history_step_returns(anchor_window)
        shape_gap = average_absolute_difference(anchor_shape_path, current_shape_path)
        return_shape_gap = average_absolute_difference(anchor_return_shape, current_return_shape)
        recent_return_gap = abs(
            (percentage_change(anchor_window[-1].close_price, anchor_window[0].close_price) or ZERO)
            - current_recent_return_pct
        )
        if shape_gap is None or return_shape_gap is None:
            continue

        phase_penalty = (
            Decimal("0.00")
            if (anchor_metrics.get("cycle_phase") or "Sin ciclo") == current_phase
            else Decimal("5.00")
        )
        score = (
            (shape_gap * Decimal("1.60"))
            + (return_shape_gap * Decimal("0.95"))
            + (recent_return_gap * Decimal("0.55"))
            + abs((anchor_metrics.get("one_year_return_pct") or ZERO) - current_one_year_return_pct)
            + (abs((anchor_metrics.get("current_drawdown_pct") or ZERO) - current_drawdown_pct) * Decimal("1.35"))
            + (abs((anchor_metrics.get("annualized_volatility_pct") or ZERO) - current_volatility_pct) * Decimal("0.35"))
            + phase_penalty
        )

        step_return_pcts = []
        for step_number in range(1, step_count + 1):
            previous_index = anchor_index + ((step_number - 1) * step_months)
            next_index = anchor_index + (step_number * step_months)
            if next_index >= len(monthly_history):
                step_return_pcts = []
                break
            step_return_pct = percentage_change(
                monthly_history[next_index].close_price,
                monthly_history[previous_index].close_price,
            )
            if step_return_pct is None:
                step_return_pcts = []
                break
            step_return_pcts.append(step_return_pct)

        if len(step_return_pcts) != step_count:
            continue

        matches.append(
            {
                "score": score,
                "step_return_pcts": step_return_pcts,
                "anchor_date": monthly_history[anchor_index].price_date,
                "cycle_phase": anchor_metrics.get("cycle_phase") or "Sin ciclo",
                "shape_gap": shape_gap,
                "return_shape_gap": return_shape_gap,
            }
        )

    if not matches:
        return {
            "available": False,
            "reference_label": reference_series.name or reference_series.symbol,
            "years_covered": calculate_years_between(monthly_history[0].price_date, monthly_history[-1].price_date),
            "match_windows_count": 0,
            "selected_windows_count": 0,
        }

    selected_matches = sorted(matches, key=lambda item: (item["score"], item["anchor_date"]))[
        : min(REFERENCE_CYCLE_TEMPLATE_MAX_MATCHES, len(matches))
    ]
    averaged_steps = []
    for step_index in range(step_count):
        weighted_total = ZERO
        total_weight = ZERO
        for match in selected_matches:
            weight = Decimal("1.00") / (Decimal("1.00") + (match["score"] / Decimal("10.00")))
            weighted_total += match["step_return_pcts"][step_index] * weight
            total_weight += weight
        averaged_steps.append((weighted_total / total_weight) if total_weight else ZERO)

    return {
        "available": True,
        "reference_label": reference_series.name or reference_series.symbol,
        "years_covered": calculate_years_between(monthly_history[0].price_date, monthly_history[-1].price_date),
        "match_windows_count": len(matches),
        "selected_windows_count": len(selected_matches),
        "selected_anchor_dates": [match["anchor_date"] for match in selected_matches],
        "recent_months": recent_months,
        "shape_window_start_date": current_window[0].price_date,
        "shape_window_end_date": current_window[-1].price_date,
        "selected_shape_gaps": [quantize_decimal(match["shape_gap"]) or ZERO for match in selected_matches],
        "step_return_pcts": [quantize_decimal(value) or ZERO for value in averaged_steps],
        "current_cycle_phase": current_phase,
    }


def build_reference_cycle_template(
    position: EquityPosition,
    *,
    latest_date: date | None,
    years: int = 5,
    step_months: int = 6,
) -> dict:
    try:
        reference_series = fetch_reference_series(
            position,
            range_key=MAX_MARKET_RANGE_KEY,
        )
    except Exception:
        logger.exception("No se pudo cargar la referencia larga para %s", position.ticker)
        return {"available": False}
    return build_reference_cycle_template_from_series(
        reference_series,
        latest_date=latest_date,
        years=years,
        step_months=step_months,
    )


def pearson_correlation(xs: list[Decimal], ys: list[Decimal]) -> Decimal | None:
    if len(xs) != len(ys) or len(xs) < 3:
        return None

    x_values = [float(value) for value in xs]
    y_values = [float(value) for value in ys]
    mean_x = sum(x_values) / len(x_values)
    mean_y = sum(y_values) / len(y_values)
    covariance = sum((x - mean_x) * (y - mean_y) for x, y in zip(x_values, y_values))
    variance_x = sum((x - mean_x) ** 2 for x in x_values)
    variance_y = sum((y - mean_y) ** 2 for y in y_values)
    if variance_x <= 0 or variance_y <= 0:
        return None
    correlation = covariance / math.sqrt(variance_x * variance_y)
    return Decimal(str(round(correlation, 4)))


def calculate_regression_beta(xs: list[Decimal], ys: list[Decimal]) -> Decimal | None:
    if len(xs) != len(ys) or len(xs) < 3:
        return None

    x_values = [float(value) for value in xs]
    y_values = [float(value) for value in ys]
    mean_x = sum(x_values) / len(x_values)
    mean_y = sum(y_values) / len(y_values)
    variance_x = sum((x - mean_x) ** 2 for x in x_values)
    if variance_x <= 0:
        return None
    covariance = sum((x - mean_x) * (y - mean_y) for x, y in zip(x_values, y_values))
    beta = covariance / variance_x
    return Decimal(str(round(beta, 4)))


def describe_correlation(coefficient: Decimal | None) -> str:
    if coefficient is None:
        return "Sin correlacion suficiente"
    if coefficient >= Decimal("0.70"):
        return "Relacion positiva alta"
    if coefficient >= Decimal("0.35"):
        return "Relacion positiva moderada"
    if coefficient <= Decimal("-0.35"):
        return "Relacion inversa"
    return "Relacion debil"


def build_correlation_from_rows(rows: list[dict], frequency: str, *, secondary_change_mode: str = "pct") -> dict:
    buckets = {}
    for row in rows:
        buckets[bucket_label_for_date(row["date"], frequency)] = row
    collapsed = [buckets[label] for label in sorted(buckets.keys())]

    stock_returns, reference_returns = build_return_series_from_collapsed_rows(
        collapsed,
        "stock_close",
        "reference_close",
        secondary_change_mode=secondary_change_mode,
    )
    window_size = correlation_window_size(frequency)
    recent_window = recent_correlation_window_size(frequency)
    coefficient = pearson_correlation(stock_returns[-window_size:], reference_returns[-window_size:])
    recent_coefficient = pearson_correlation(stock_returns[-recent_window:], reference_returns[-recent_window:])
    beta = calculate_regression_beta(reference_returns[-window_size:], stock_returns[-window_size:])
    recent_beta = calculate_regression_beta(reference_returns[-recent_window:], stock_returns[-recent_window:])
    stability_gap = abs(coefficient - recent_coefficient) if coefficient is not None and recent_coefficient is not None else None
    if stability_gap is None or stability_gap <= Decimal("0.15"):
        stability_label = "Estable"
    elif stability_gap <= Decimal("0.35"):
        stability_label = "Aceptable"
    else:
        stability_label = "Cambiante"
    return {
        "frequency": frequency,
        "change_mode": secondary_change_mode,
        "coefficient": coefficient,
        "recent_coefficient": recent_coefficient,
        "beta": beta,
        "recent_beta": recent_beta,
        "label": describe_correlation(coefficient),
        "observations_count": len(stock_returns),
        "years_covered": calculate_years_between(
            collapsed[0]["date"] if collapsed else None,
            collapsed[-1]["date"] if collapsed else None,
        ),
        "window_observations": min(len(stock_returns), window_size),
        "stability_gap": stability_gap,
        "stability_label": stability_label,
    }


def build_reference_correlation(history, position: EquityPosition) -> dict:
    frequency = infer_reference_frequency(position)
    rows = [
        {
            "date": point.price_date,
            "stock_close": point.close_price,
            "reference_close": point.benchmark_close,
        }
        for point in history
    ]
    return build_correlation_from_rows(
        rows,
        frequency,
        secondary_change_mode=infer_reference_change_mode(position.reference_profile),
    )


def build_reference_correlation_for_series(history, reference_series: MarketSeries, reference_profile: str) -> dict:
    aligned = align_reference_points(
        [{"date": point.price_date} for point in history],
        reference_series.points,
    )
    rows = [
        {
            "date": point.price_date,
            "stock_close": point.close_price,
            "reference_close": aligned.get(point.price_date),
        }
        for point in history
    ]
    return build_correlation_from_rows(
        rows,
        infer_reference_frequency_from_profile(reference_profile),
        secondary_change_mode=infer_reference_change_mode(reference_profile),
    )


def build_reference_rolling_correlation_rows_for_series(
    history,
    reference_series: MarketSeries,
    reference_profile: str,
) -> list[dict]:
    frequency = infer_reference_frequency_from_profile(reference_profile)
    aligned = align_reference_points(
        [{"date": point.price_date} for point in history],
        reference_series.points,
    )
    rows = [
        {
            "date": point.price_date,
            "stock_close": point.close_price,
            "reference_close": aligned.get(point.price_date),
        }
        for point in history
        if aligned.get(point.price_date) is not None
    ]
    buckets = {}
    for row in rows:
        buckets[bucket_label_for_date(row["date"], frequency)] = row
    collapsed = [buckets[label] for label in sorted(buckets.keys())]
    stock_returns, reference_returns = build_return_series_from_collapsed_rows(
        collapsed,
        "stock_close",
        "reference_close",
        secondary_change_mode=infer_reference_change_mode(reference_profile),
    )
    recent_window = recent_correlation_window_size(frequency)
    if len(stock_returns) < recent_window + 2:
        return []

    rolling_rows = []
    for end_index in range(recent_window, len(stock_returns) + 1):
        rolling_coefficient = pearson_correlation(
            stock_returns[end_index - recent_window : end_index],
            reference_returns[end_index - recent_window : end_index],
        )
        if rolling_coefficient is None:
            continue
        rolling_rows.append(
            {
                "end_date": collapsed[end_index]["date"],
                "coefficient": rolling_coefficient,
            }
        )
    return rolling_rows


def collapse_equity_history_to_close_map(history, frequency: str = "monthly") -> dict[str, dict]:
    collapsed_rows = collapse_history_to_frequency(history, frequency)
    return {
        bucket_label_for_date(point.price_date, frequency): {
            "date": point.price_date,
            "close": point.close_price,
        }
        for point in collapsed_rows
        if point.price_date is not None and point.close_price is not None
    }


def resolve_equity_sector_relationship(left_sector_label: str = "", right_sector_label: str = "") -> dict:
    normalized_left = normalize_company_lookup(left_sector_label)
    normalized_right = normalize_company_lookup(right_sector_label)
    if not normalized_left or not normalized_right:
        return {
            "key": "unknown",
            "label": "Sin sector suficiente",
            "detail": "Falta al menos un sector",
        }
    if normalized_left == normalized_right:
        return {
            "key": "same",
            "label": "Mismo sector",
            "detail": left_sector_label or right_sector_label,
        }

    stop_words = {"Y", "E", "DE", "DEL", "LA", "EL", "AL", "LOS", "LAS", "CON"}
    family_map = {
        "energia": {
            "ENERGIA", "ENERGETICA", "ENERGETICO", "ELECTRICA", "ELECTRICAS",
            "GAS", "SOLAR", "RENOVABLE", "RENOVABLES", "REDES", "UTILITY", "UTILITIES",
        },
        "infraestructuras": {
            "INFRAESTRUCTURA", "INFRAESTRUCTURAS", "CONSTRUCCION", "CONSTRUCCIONES",
            "CONCESIONES", "INMOBILIARIO", "VIVIENDA", "HOGAR",
        },
        "finanzas": {
            "BANCA", "BANCARIO", "BANCARIOS", "SEGUROS", "SEGURO", "FINANCIERO", "FINANCIEROS",
        },
        "tecnologia": {
            "TECNOLOGIA", "TECNOLOGICO", "TECNOLOGICOS", "DEFENSA", "TELECOMUNICACIONES",
        },
        "consumo": {
            "CONSUMO", "DISTRIBUCION", "LOGISTICA", "VIAJES", "MOVILIDAD", "AEROPUERTOS", "AEROLINEAS",
        },
        "industria": {
            "INDUSTRIA", "INDUSTRIAL", "INDUSTRIALES", "MATERIALES", "ACERO", "COBRE", "EQUIPAMIENTO",
        },
        "salud": {"SALUD"},
    }
    family_labels = {
        "energia": "energia y utilities",
        "infraestructuras": "infraestructuras y ciclo inmobiliario",
        "finanzas": "finanzas",
        "tecnologia": "tecnologia y telecom",
        "consumo": "consumo y movilidad",
        "industria": "industria y materiales",
        "salud": "salud",
    }

    def build_tags(sector_label: str) -> set[str]:
        tokens = {
            token
            for token in normalize_company_lookup(sector_label).split()
            if token and token not in stop_words
        }
        tags = set()
        for family, keywords in family_map.items():
            if tokens & keywords:
                tags.add(family)
        if not tags:
            tags = {token.lower() for token in tokens}
        return tags

    shared_tags = sorted(build_tags(left_sector_label) & build_tags(right_sector_label))
    if shared_tags:
        shared_label = ", ".join(family_labels.get(tag, tag.replace("_", " ")) for tag in shared_tags)
        return {
            "key": "related",
            "label": "Sectores afines",
            "detail": shared_label,
        }
    return {
        "key": "distinct",
        "label": "Sectores distintos",
        "detail": f"{left_sector_label} vs {right_sector_label}",
    }


def build_pairwise_equity_correlation(
    left_card: dict,
    right_card: dict,
    frequency: str = "monthly",
) -> dict:
    left_position = left_card["position"]
    right_position = right_card["position"]
    left_history = list(left_position.price_history.all())
    right_history = list(right_position.price_history.all())
    sector_relationship = resolve_equity_sector_relationship(
        left_card.get("sector_label", ""),
        right_card.get("sector_label", ""),
    )
    left_map = collapse_equity_history_to_close_map(left_history, frequency)
    right_map = collapse_equity_history_to_close_map(right_history, frequency)
    shared_labels = sorted(set(left_map.keys()) & set(right_map.keys()))
    if len(shared_labels) < PORTFOLIO_CORRELATION_MIN_COMMON_PERIODS:
        return {
            "available": False,
            "left_ticker": left_position.ticker,
            "right_ticker": right_position.ticker,
            "left_company_name": left_position.company_name,
            "right_company_name": right_position.company_name,
            "left_sector_label": left_card.get("sector_label") or "Sin sector",
            "right_sector_label": right_card.get("sector_label") or "Sin sector",
            "pair_label": f"{left_position.ticker} / {right_position.ticker}",
            "coefficient": None,
            "observations_count": 0,
            "shared_periods_count": len(shared_labels),
            "years_covered": ZERO,
            "start_label": "",
            "end_label": "",
            "sector_relationship_key": sector_relationship["key"],
            "sector_relationship_label": sector_relationship["label"],
            "sector_relationship_detail": sector_relationship["detail"],
        }

    rows = []
    for label in shared_labels:
        left_point = left_map[label]
        right_point = right_map[label]
        rows.append(
            {
                "date": max(left_point["date"], right_point["date"]),
                "stock_close": left_point["close"],
                "reference_close": right_point["close"],
            }
        )
    correlation = build_correlation_from_rows(rows, frequency, secondary_change_mode="pct")
    coefficient = correlation.get("coefficient")
    return {
        "available": coefficient is not None,
        "left_ticker": left_position.ticker,
        "right_ticker": right_position.ticker,
        "left_company_name": left_position.company_name,
        "right_company_name": right_position.company_name,
        "left_sector_label": left_card.get("sector_label") or "Sin sector",
        "right_sector_label": right_card.get("sector_label") or "Sin sector",
        "pair_label": f"{left_position.ticker} / {right_position.ticker}",
        "coefficient": coefficient,
        "observations_count": correlation.get("observations_count", 0),
        "shared_periods_count": len(rows),
        "years_covered": correlation.get("years_covered") or ZERO,
        "start_label": rows[0]["date"].isoformat() if rows else "",
        "end_label": rows[-1]["date"].isoformat() if rows else "",
        "sector_relationship_key": sector_relationship["key"],
        "sector_relationship_label": sector_relationship["label"],
        "sector_relationship_detail": sector_relationship["detail"],
    }


def estimate_dalio_loss_probability(correlation_pct: Decimal | None) -> Decimal | None:
    if correlation_pct is None:
        return None
    points = list(DALIO_CORRELATION_RISK_POINTS)
    if not points:
        return None
    if correlation_pct <= points[0][0]:
        return points[0][1]
    for left_point, right_point in zip(points, points[1:]):
        left_x, left_y = left_point
        right_x, right_y = right_point
        if left_x <= correlation_pct <= right_x:
            span = right_x - left_x
            if span == ZERO:
                return left_y
            ratio = (correlation_pct - left_x) / span
            return quantize_decimal(left_y + ((right_y - left_y) * ratio), "0.01")
    last_left_x, last_left_y = points[-2]
    last_right_x, last_right_y = points[-1]
    tail_span = last_right_x - last_left_x
    if tail_span == ZERO:
        return last_right_y
    tail_ratio = (correlation_pct - last_left_x) / tail_span
    return quantize_decimal(last_left_y + ((last_right_y - last_left_y) * tail_ratio), "0.01")


def build_correlation_risk_curve_chart(
    current_correlation_pct: Decimal | None,
    current_loss_probability_pct: Decimal | None,
    width: int = 640,
    height: int = 220,
    padding: int = 26,
) -> dict:
    points = [
        {
            "correlation_pct": correlation_pct,
            "loss_probability_pct": loss_probability_pct,
        }
        for correlation_pct, loss_probability_pct in DALIO_CORRELATION_RISK_POINTS
    ]
    if len(points) < 2:
        return {"available": False}

    max_curve_x = max(point["correlation_pct"] for point in points)
    max_curve_y = max(point["loss_probability_pct"] for point in points)
    current_x = current_correlation_pct if current_correlation_pct is not None else ZERO
    current_y = current_loss_probability_pct if current_loss_probability_pct is not None else ZERO
    x_max = max(max_curve_x, Decimal(str(int(math.ceil(float(current_x) / 10.0)) * 10))) if current_x else max_curve_x
    y_max_base = max(max_curve_y, current_y)
    y_max = Decimal(str(int(math.ceil(float((y_max_base or ZERO) + Decimal("4")) / 10.0)) * 10 or 10))
    span_x = width - (padding * 2)
    span_y = height - (padding * 2)

    def to_x(value: Decimal) -> float:
        base = float(x_max) if x_max not in {None, ZERO} else 1.0
        return padding + (span_x * (float(value) / base))

    def to_y(value: Decimal) -> float:
        base = float(y_max) if y_max not in {None, ZERO} else 1.0
        return height - padding - (span_y * (float(value) / base))

    curve_line = " ".join(
        f"{to_x(point['correlation_pct']):.1f},{to_y(point['loss_probability_pct']):.1f}"
        for point in points
    )
    x_markers = []
    x_tick = 0
    while x_tick <= int(x_max):
        x_markers.append(
            {
                "x": f"{to_x(Decimal(str(x_tick))):.1f}",
                "label": f"{x_tick} %",
                "y1": str(height - padding),
                "y2": str(height - padding + 6),
                "text_y": str(height - 2),
            }
        )
        x_tick += 10
    y_markers = []
    y_tick = 0
    while y_tick <= int(y_max):
        y_markers.append(
            {
                "y": f"{to_y(Decimal(str(y_tick))):.1f}",
                "label": f"{y_tick} %",
            }
        )
        y_tick += 10
    current_marker = None
    if current_correlation_pct is not None and current_loss_probability_pct is not None:
        current_marker = {
            "x": f"{to_x(current_correlation_pct):.1f}",
            "y": f"{to_y(current_loss_probability_pct):.1f}",
            "correlation_label": f"{current_correlation_pct:.1f} %",
            "loss_probability_label": f"{current_loss_probability_pct:.1f} %",
        }
    return {
        "available": True,
        "curve_line": curve_line,
        "curve_points": [
            {
                "x": f"{to_x(point['correlation_pct']):.1f}",
                "y": f"{to_y(point['loss_probability_pct']):.1f}",
                "label": (
                    f"Corr {point['correlation_pct']:.0f} % | "
                    f"prob. perder dinero {point['loss_probability_pct']:.0f} %"
                ),
            }
            for point in points
        ],
        "current_marker": current_marker,
        "x_markers": x_markers,
        "y_markers": y_markers,
        "max_label": f"{y_max:.0f} %",
        "min_label": "0 %",
        "current_correlation_label": f"{current_correlation_pct:.1f} %" if current_correlation_pct is not None else "-",
        "current_loss_probability_label": (
            f"{current_loss_probability_pct:.1f} %"
            if current_loss_probability_pct is not None
            else "-"
        ),
    }


def build_correlation_heatmap_palette(coefficient: Decimal | None, *, is_self: bool = False) -> dict:
    if is_self:
        return {"background": "#0d5d78", "text": "#ffffff"}
    if coefficient is None:
        return {"background": "#f3f7fa", "text": "#8aa0b1"}
    if coefficient >= Decimal("0.75"):
        return {"background": "#c9653d", "text": "#ffffff"}
    if coefficient >= Decimal("0.50"):
        return {"background": "#e69a64", "text": "#17384e"}
    if coefficient >= Decimal("0.25"):
        return {"background": "#f6d8bf", "text": "#17384e"}
    if coefficient > Decimal("-0.25"):
        return {"background": "#e8f0f7", "text": "#17384e"}
    if coefficient > Decimal("-0.50"):
        return {"background": "#cae6db", "text": "#17384e"}
    return {"background": "#2f7d5a", "text": "#ffffff"}


def build_portfolio_correlation_context(history_cards: list[dict]) -> dict:
    owned_cards = [
        card
        for card in history_cards
        if card.get("position") is not None
        and card["position"].is_owned
        and card.get("has_history")
    ]
    if len(owned_cards) < 2:
        return {"available": False}

    owned_cards = sorted(
        owned_cards,
        key=lambda card: (
            -(card["position"].current_value or ZERO),
            card["position"].company_name,
        ),
    )
    total_current_value = sum((card["position"].current_value or ZERO) for card in owned_cards)
    pair_rows = []
    pairs_by_key = {}
    pair_weight_total = ZERO
    weighted_correlation_total = ZERO
    weighted_positive_correlation_total = ZERO
    sector_relation_counts = defaultdict(int)
    unavailable_pairs_count = 0

    for left_card, right_card in combinations(owned_cards, 2):
        pair = build_pairwise_equity_correlation(left_card, right_card)
        left_value = left_card["position"].current_value or ZERO
        right_value = right_card["position"].current_value or ZERO
        left_weight = (left_value / total_current_value) if total_current_value > ZERO else ZERO
        right_weight = (right_value / total_current_value) if total_current_value > ZERO else ZERO
        pair_weight = Decimal("2") * left_weight * right_weight
        pair["pair_weight"] = pair_weight
        pair["pair_weight_pct"] = quantize_decimal(pair_weight * ONE_HUNDRED, "0.01") or ZERO
        if not pair.get("available"):
            unavailable_pairs_count += 1
            continue
        coefficient = pair.get("coefficient")
        if coefficient is None:
            unavailable_pairs_count += 1
            continue
        pair_rows.append(pair)
        pair_key = tuple(sorted((left_card["position"].id or 0, right_card["position"].id or 0)))
        pairs_by_key[pair_key] = pair
        pair_weight_total += pair_weight
        weighted_correlation_total += coefficient * pair_weight
        weighted_positive_correlation_total += max(coefficient, ZERO) * pair_weight
        sector_relation_counts[pair["sector_relationship_key"]] += 1

    if not pair_rows:
        return {"available": False}

    average_correlation = (
        quantize_decimal(weighted_correlation_total / pair_weight_total, "0.01")
        if pair_weight_total > ZERO
        else average_decimal([pair["coefficient"] for pair in pair_rows])
    )
    dalio_correlation_pct = (
        quantize_decimal((weighted_positive_correlation_total / pair_weight_total) * ONE_HUNDRED, "0.01")
        if pair_weight_total > ZERO
        else quantize_decimal(
            average_decimal([max(pair["coefficient"], ZERO) * ONE_HUNDRED for pair in pair_rows]),
            "0.01",
        )
    )
    estimated_loss_probability_pct = estimate_dalio_loss_probability(dalio_correlation_pct)
    risk_curve_chart = build_correlation_risk_curve_chart(
        dalio_correlation_pct,
        estimated_loss_probability_pct,
    )
    pair_rows.sort(
        key=lambda pair: (
            -(pair.get("coefficient") or Decimal("-9")),
            -(pair.get("pair_weight_pct") or ZERO),
            pair["pair_label"],
        )
    )
    highest_pair = max(pair_rows, key=lambda pair: pair.get("coefficient") or Decimal("-9"))
    most_diversifying_pair = min(pair_rows, key=lambda pair: pair.get("coefficient") or Decimal("9"))
    total_pairs = len(pair_rows)

    heatmap_rows = []
    for left_card in owned_cards:
        left_position = left_card["position"]
        cells = []
        for right_card in owned_cards:
            right_position = right_card["position"]
            if left_position.id == right_position.id:
                palette = build_correlation_heatmap_palette(Decimal("1.00"), is_self=True)
                cells.append(
                    {
                        "label": "1.00",
                        "tooltip": f"{left_position.company_name} | misma accion",
                        "background": palette["background"],
                        "text_color": palette["text"],
                        "is_self": True,
                    }
                )
                continue
            pair = pairs_by_key.get(tuple(sorted((left_position.id or 0, right_position.id or 0))))
            coefficient = pair.get("coefficient") if pair else None
            palette = build_correlation_heatmap_palette(coefficient)
            if pair:
                tooltip = (
                    f"{pair['left_company_name']} vs {pair['right_company_name']} | "
                    f"corr {pair['coefficient']:.2f} | {pair['sector_relationship_label']} | "
                    f"{pair['years_covered']:.1f}A"
                )
            else:
                tooltip = f"{left_position.company_name} vs {right_position.company_name} | sin historico comun suficiente"
            cells.append(
                {
                    "label": f"{coefficient:.2f}" if coefficient is not None else "-",
                    "tooltip": tooltip,
                    "background": palette["background"],
                    "text_color": palette["text"],
                    "is_self": False,
                }
            )
        heatmap_rows.append(
            {
                "ticker": left_position.ticker,
                "company_name": left_position.company_name,
                "sector_label": left_card.get("sector_label") or "Sin sector",
                "cells": cells,
            }
        )

    diversification_label = "Alta"
    if dalio_correlation_pct is not None:
        if dalio_correlation_pct >= Decimal("40"):
            diversification_label = "Ajustada"
        elif dalio_correlation_pct >= Decimal("20"):
            diversification_label = "Media"
        elif dalio_correlation_pct >= Decimal("10"):
            diversification_label = "Buena"

    return {
        "available": True,
        "positions_count": len(owned_cards),
        "pair_count": total_pairs,
        "total_possible_pairs": (len(owned_cards) * (len(owned_cards) - 1)) // 2,
        "unavailable_pairs_count": unavailable_pairs_count,
        "frequency_label": "mensual sobre cierres compartidos",
        "average_correlation": average_correlation,
        "average_positive_correlation_pct": dalio_correlation_pct,
        "estimated_loss_probability_pct": estimated_loss_probability_pct,
        "diversification_label": diversification_label,
        "same_sector_pairs_count": sector_relation_counts.get("same", 0),
        "related_sector_pairs_count": sector_relation_counts.get("related", 0),
        "distinct_sector_pairs_count": sector_relation_counts.get("distinct", 0),
        "highest_pair": highest_pair,
        "most_diversifying_pair": most_diversifying_pair,
        "pairs": pair_rows,
        "heatmap_headers": [
            {
                "ticker": card["position"].ticker,
                "company_name": card["position"].company_name,
            }
            for card in owned_cards
        ],
        "heatmap_rows": heatmap_rows,
        "risk_curve_chart": risk_curve_chart,
        "dalio_source_note": (
            "Curva orientativa basada en los puntos que has indicado: 0, 10, 20, 40 y 60 % de correlacion."
        ),
    }


def filter_history_window(history, start_date: date | None = None, end_date: date | None = None):
    return [
        point
        for point in history
        if (start_date is None or point.price_date >= start_date) and (end_date is None or point.price_date <= end_date)
    ]


def build_period_snapshot(
    history,
    label: str,
    start_date: date | None = None,
    end_date: date | None = None,
    *,
    reference_profile: str | None = None,
) -> dict:
    window = filter_history_window(history, start_date=start_date, end_date=end_date)
    benchmark_change_mode = infer_reference_change_mode(reference_profile or EquityPosition.ReferenceProfile.MARKET_INDEX)
    if len(window) < 2:
        return {
            "label": label,
            "available": False,
            "stock_return_pct": None,
            "benchmark_return_pct": None,
            "benchmark_change": None,
            "benchmark_change_mode": benchmark_change_mode,
            "alpha_pct": None,
            "start_date": window[0].price_date if window else start_date,
            "end_date": window[-1].price_date if window else end_date,
        }

    first_point = window[0]
    last_point = window[-1]
    stock_return_pct = percentage_change(last_point.close_price, first_point.close_price)
    first_benchmark_close = getattr(first_point, "benchmark_close", None)
    last_benchmark_close = getattr(last_point, "benchmark_close", None)
    benchmark_return_pct = percentage_change(last_benchmark_close, first_benchmark_close)
    benchmark_change = calculate_series_change(
        last_benchmark_close,
        first_benchmark_close,
        change_mode=benchmark_change_mode,
    )
    alpha_pct = (
        stock_return_pct - benchmark_return_pct
        if benchmark_change_mode == "pct" and stock_return_pct is not None and benchmark_return_pct is not None
        else None
    )

    return {
        "label": label,
        "available": True,
        "stock_return_pct": stock_return_pct,
        "benchmark_return_pct": benchmark_return_pct,
        "benchmark_change": benchmark_change,
        "benchmark_change_mode": benchmark_change_mode,
        "alpha_pct": alpha_pct,
        "start_date": first_point.price_date,
        "end_date": last_point.price_date,
    }


def quantize_decimal(value: Decimal | None, places: str = "0.01") -> Decimal | None:
    if value is None:
        return None
    if not value.is_finite():
        return None
    quantizer = Decimal(places)
    decimal_places = max(-quantizer.as_tuple().exponent, 0)
    integer_digits = value.adjusted() + 1 if value != 0 else 1
    required_precision = max(28, integer_digits + decimal_places + 2)
    with localcontext() as context:
        context.prec = required_precision
        try:
            return value.quantize(quantizer)
        except InvalidOperation:
            context.prec = max(context.prec, len(value.as_tuple().digits) + decimal_places + 2)
            return value.quantize(quantizer)


def clamp_decimal(value: Decimal, lower: Decimal, upper: Decimal) -> Decimal:
    return max(lower, min(upper, value))


def equity_analysis_notional() -> Decimal:
    configured = getattr(settings, "EQUITIES_ANALYSIS_NOTIONAL", DEFAULT_EQUITY_ANALYSIS_NOTIONAL)
    try:
        amount = Decimal(str(configured))
    except Exception:
        amount = DEFAULT_EQUITY_ANALYSIS_NOTIONAL
    return max(amount, Decimal("1000.00"))


def resolve_projection_analysis_value(position: EquityPosition, latest_price: Decimal | None = None) -> tuple[Decimal, str]:
    current_value = Decimal(position.current_value or 0)
    invested_amount = Decimal(position.invested_amount or 0)
    latest_price = Decimal(latest_price or 0)
    shares = Decimal(position.shares or 0)
    analysis_value = current_value or invested_amount
    if analysis_value <= ZERO and latest_price > ZERO and shares > ZERO:
        analysis_value = latest_price * shares
    if position.is_owned:
        return analysis_value, "actual"
    normalized_floor = equity_analysis_notional()
    if analysis_value <= ZERO or analysis_value < normalized_floor:
        return normalized_floor, "normalized_watchlist"
    return analysis_value, "watchlist_actual"


def build_projection_dividend_income(position: EquityPosition, analysis_value: Decimal) -> Decimal:
    analysis_value = Decimal(analysis_value or 0)
    if analysis_value <= ZERO:
        return ZERO
    current_value = Decimal(position.current_value or 0)
    invested_amount = Decimal(position.invested_amount or 0)
    reference_value = current_value or invested_amount
    annual_dividend_income = Decimal(position.annual_dividend_income or 0)
    if reference_value > ZERO and annual_dividend_income > ZERO:
        return quantize_decimal((annual_dividend_income / reference_value) * analysis_value, "0.01") or ZERO
    return annual_dividend_income


def build_analysis_broker_costs(position: EquityPosition, analysis_value: Decimal, annual_dividend_income: Decimal) -> dict:
    analysis_value = Decimal(analysis_value or 0)
    annual_dividend_income = Decimal(annual_dividend_income or 0)
    broker_costs = estimate_broker_costs(
        broker_name=position.broker,
        trade_channel=position.trade_channel,
        trade_amount=analysis_value,
        valuation_amount=analysis_value,
        annual_dividend_income=annual_dividend_income,
        quote_symbol=position.quote_symbol,
    )
    annual_cost_used, annual_cost_source = resolve_recurring_cost_used(
        position.annual_maintenance_cost,
        broker_costs.get("annual_recurring_cost", ZERO),
    )
    net_dividend_income = annual_dividend_income - broker_costs.get("annual_dividend_fee", ZERO)
    return {
        **broker_costs,
        "annual_cost_used": annual_cost_used,
        "annual_cost_source": annual_cost_source,
        "net_dividend_income": quantize_decimal(net_dividend_income, "0.01") or ZERO,
        "gross_dividend_income": quantize_decimal(annual_dividend_income, "0.01") or ZERO,
    }


def annualize_return_pct(return_pct: Decimal | None, months: int) -> Decimal | None:
    if return_pct is None or months <= 0:
        return None
    base = 1 + (float(return_pct) / 100)
    if base <= 0:
        return Decimal("-100.00")
    annualized = (base ** (12 / months) - 1) * 100
    return Decimal(str(round(annualized, 4)))


def calculate_equivalent_return_pct(
    cumulative_return_pct: Decimal | None,
    elapsed_days: int,
    target_days: Decimal,
) -> Decimal | None:
    if cumulative_return_pct is None or elapsed_days <= 0 or target_days <= 0:
        return None
    base = 1 + (float(cumulative_return_pct) / 100)
    if base <= 0:
        return Decimal("-100.00")
    exponent = float(target_days / Decimal(str(elapsed_days)))
    equivalent = (base ** exponent - 1) * 100
    return Decimal(str(round(equivalent, 4)))


def build_comparable_return_metrics(
    cumulative_return_pct: Decimal | None,
    holding_days: int,
) -> dict:
    return {
        "holding_days": holding_days,
        "monthly_equivalent_return_pct": calculate_equivalent_return_pct(
            cumulative_return_pct,
            holding_days,
            COMPARABLE_MONTH_DAYS,
        ),
        "annual_equivalent_return_pct": calculate_equivalent_return_pct(
            cumulative_return_pct,
            holding_days,
            Decimal("365"),
        ),
    }


def build_projection_signal(return_pct: Decimal | None, months: int) -> Decimal | None:
    if return_pct is None:
        return None
    annualized = annualize_return_pct(return_pct, months)
    if annualized is None:
        return None
    linearized = return_pct * (Decimal("12") / Decimal(months))
    return (annualized * Decimal("0.55")) + (linearized * Decimal("0.45"))


def standard_deviation_decimal(values: list[Decimal]) -> Decimal | None:
    if len(values) < 2:
        return None
    numeric_values = [float(value) for value in values]
    mean_value = sum(numeric_values) / len(numeric_values)
    variance = sum((value - mean_value) ** 2 for value in numeric_values) / (len(numeric_values) - 1)
    return Decimal(str(round(math.sqrt(variance), 4)))


def build_recent_monthly_stock_returns(history, months_back: int = 6) -> list[Decimal]:
    if not history:
        return []
    latest_date = history[-1].price_date
    recent_history = filter_history_window(
        history,
        start_date=latest_date - timedelta(days=35 * (months_back + 1)),
        end_date=latest_date,
    )
    monthly_history = collapse_history_to_frequency(recent_history, "monthly")[-(months_back + 1):]
    monthly_returns = []
    for previous, current in zip(monthly_history, monthly_history[1:]):
        period_return = percentage_change(current.close_price, previous.close_price)
        if period_return is not None:
            monthly_returns.append(period_return)
    return monthly_returns


def decimal_linear_slope(values: list[Decimal]) -> Decimal | None:
    filtered = [value for value in values if value is not None]
    if len(filtered) < 2:
        return None

    xs = list(range(len(filtered)))
    mean_x = sum(xs) / len(xs)
    mean_y = sum(float(value) for value in filtered) / len(filtered)
    numerator = sum((x - mean_x) * (float(value) - mean_y) for x, value in zip(xs, filtered))
    denominator = sum((x - mean_x) ** 2 for x in xs)
    if denominator <= 0:
        return None
    return Decimal(str(round(numerator / denominator, 4)))


def consecutive_tail_count(values: list[Decimal], positive: bool) -> int:
    streak = 0
    for value in reversed(values):
        if value is None:
            continue
        if positive and value > ZERO:
            streak += 1
            continue
        if not positive and value < ZERO:
            streak += 1
            continue
        break
    return streak


def consecutive_tail_matches(values: list, predicate) -> int:
    streak = 0
    for value in reversed(values):
        if predicate(value):
            streak += 1
            continue
        break
    return streak


def build_relative_strength_trend(
    history,
    position: EquityPosition,
    reference_sensitivity: Decimal | None,
    periods_back: int = 8,
) -> dict:
    if not history:
        return {
            "available": False,
            "label": "Sin tendencia relativa",
            "periods_label": "periodos",
        }

    frequency = infer_reference_frequency(position)
    periods_label = "meses" if frequency == "monthly" else "trimestres"
    buffer_days = 35 if frequency == "monthly" else 95
    latest_date = history[-1].price_date
    recent_history = filter_history_window(
        history,
        start_date=latest_date - timedelta(days=buffer_days * (periods_back + 2)),
        end_date=latest_date,
    )
    collapsed_history = collapse_history_to_frequency(recent_history, frequency)[-(periods_back + 1) :]
    change_mode = infer_reference_change_mode(position.reference_profile)

    rows = []
    for previous, current in zip(collapsed_history, collapsed_history[1:]):
        stock_return_pct = percentage_change(current.close_price, previous.close_price)
        reference_change = calculate_series_change(
            current.benchmark_close,
            previous.benchmark_close,
            change_mode=change_mode,
        )
        if stock_return_pct is None or reference_change is None:
            continue
        expected_return_pct = (
            reference_change * reference_sensitivity
            if reference_sensitivity is not None
            else reference_change
        )
        relative_gap_pct = stock_return_pct - expected_return_pct
        rows.append(
            {
                "start_date": previous.price_date,
                "end_date": current.price_date,
                "stock_return_pct": stock_return_pct,
                "reference_return_pct": reference_change,
                "expected_return_pct": expected_return_pct,
                "relative_gap_pct": relative_gap_pct,
            }
        )

    if len(rows) < 3:
        return {
            "available": False,
            "label": "Sin tendencia relativa",
            "periods_label": periods_label,
            "rows": rows,
        }

    gap_values = [row["relative_gap_pct"] for row in rows]
    positive_streak = consecutive_tail_count(gap_values, positive=True)
    negative_streak = consecutive_tail_count(gap_values, positive=False)
    recent_gap_avg_pct = average_decimal(gap_values[-3:])
    overall_gap_avg_pct = average_decimal(gap_values)
    gap_slope_pct = decimal_linear_slope(gap_values)
    stock_slope_pct = decimal_linear_slope([row["stock_return_pct"] for row in rows])
    threshold = 4 if len(gap_values) >= 4 else len(gap_values)
    prolonged_positive = (
        positive_streak >= threshold
        and (recent_gap_avg_pct or ZERO) > Decimal("0.35")
        and (gap_slope_pct or ZERO) >= ZERO
    )
    prolonged_negative = (
        negative_streak >= threshold
        and (recent_gap_avg_pct or ZERO) < Decimal("-0.35")
        and (gap_slope_pct or ZERO) <= ZERO
    )

    if prolonged_positive:
        label = "Pendiente positiva prolongada"
    elif prolonged_negative:
        label = "Pendiente negativa prolongada"
    elif (recent_gap_avg_pct or ZERO) > ZERO:
        label = "Mejora moderada"
    elif (recent_gap_avg_pct or ZERO) < ZERO:
        label = "Debilitamiento moderado"
    else:
        label = "Sin sesgo claro"

    return {
        "available": True,
        "label": label,
        "periods_label": periods_label,
        "rows": rows,
        "periods_count": len(rows),
        "positive_streak": positive_streak,
        "negative_streak": negative_streak,
        "recent_gap_avg_pct": recent_gap_avg_pct,
        "overall_gap_avg_pct": overall_gap_avg_pct,
        "gap_slope_pct": gap_slope_pct,
        "stock_slope_pct": stock_slope_pct,
        "prolonged_positive": prolonged_positive,
        "prolonged_negative": prolonged_negative,
    }


def build_reference_coefficient_alert(
    history,
    position: EquityPosition,
    correlation: dict,
) -> dict:
    base_payload = {
        "available": False,
        "label": "Sin lectura",
        "tone": "neutral",
        "note": "Todavia no hay suficiente serie reciente para leer si el coeficiente de referencia se deteriora o se mantiene.",
        "trigger_label": "Sin ventana reciente suficiente",
        "long_term_coefficient": correlation.get("coefficient"),
        "recent_coefficient": correlation.get("recent_coefficient"),
        "coefficient_gap": None,
        "deterioration_streak": 0,
        "weak_streak": 0,
        "negative_streak": 0,
        "periods_label": "periodos",
        "window_size": 0,
        "windows_count": 0,
        "trend_slope": None,
    }
    if not history:
        return base_payload

    frequency = infer_reference_frequency(position)
    periods_label = "meses" if frequency == "monthly" else "trimestres"
    rows = [
        {
            "date": point.price_date,
            "stock_close": point.close_price,
            "reference_close": point.benchmark_close,
        }
        for point in history
        if point.benchmark_close is not None
    ]
    buckets = {}
    for row in rows:
        buckets[bucket_label_for_date(row["date"], frequency)] = row
    collapsed = [buckets[label] for label in sorted(buckets.keys())]
    stock_returns, reference_returns = build_return_series_from_collapsed_rows(
        collapsed,
        "stock_close",
        "reference_close",
        secondary_change_mode=infer_reference_change_mode(position.reference_profile),
    )
    recent_window = recent_correlation_window_size(frequency)
    if len(stock_returns) < recent_window + 2:
        return {
            **base_payload,
            "periods_label": periods_label,
            "window_size": recent_window,
        }

    rolling_rows = []
    for end_index in range(recent_window, len(stock_returns) + 1):
        rolling_coefficient = pearson_correlation(
            stock_returns[end_index - recent_window : end_index],
            reference_returns[end_index - recent_window : end_index],
        )
        if rolling_coefficient is None:
            continue
        rolling_rows.append(
            {
                "end_date": collapsed[end_index]["date"],
                "coefficient": rolling_coefficient,
            }
        )

    if len(rolling_rows) < 2:
        return {
            **base_payload,
            "periods_label": periods_label,
            "window_size": recent_window,
            "windows_count": len(rolling_rows),
        }

    baseline = correlation.get("coefficient")
    recent_coefficient = correlation.get("recent_coefficient") or rolling_rows[-1]["coefficient"]
    coefficient_gap = (
        baseline - recent_coefficient
        if baseline is not None and recent_coefficient is not None
        else None
    )
    coefficients = [row["coefficient"] for row in rolling_rows]
    weak_streak = consecutive_tail_matches(
        coefficients,
        lambda value: value is not None and value <= Decimal("0.20"),
    )
    negative_streak = consecutive_tail_matches(
        coefficients,
        lambda value: value is not None and value < ZERO,
    )
    deterioration_streak = consecutive_tail_matches(
        coefficients,
        lambda value: (
            value is not None
            and (
                value <= Decimal("0.20")
                or (
                    baseline is not None
                    and value <= baseline - Decimal("0.18")
                )
            )
        ),
    )
    trend_slope = decimal_linear_slope(coefficients[-4:])

    sell_signal = (
        (negative_streak >= 2)
        or (
            baseline is not None
            and baseline >= Decimal("0.35")
            and deterioration_streak >= 3
            and recent_coefficient is not None
            and recent_coefficient <= Decimal("0.20")
            and (coefficient_gap or ZERO) >= Decimal("0.20")
            and (trend_slope is None or trend_slope < ZERO)
        )
    )
    watch_signal = (
        not sell_signal
        and (
            weak_streak >= 2
            or deterioration_streak >= 2
            or (
                coefficient_gap is not None
                and coefficient_gap >= Decimal("0.12")
                and recent_coefficient is not None
                and recent_coefficient <= Decimal("0.30")
            )
            or correlation.get("stability_label") == "Cambiante"
        )
    )

    trigger_label = (
        f"Reciente {recent_coefficient:.2f} vs 10A {baseline:.2f}"
        if baseline is not None and recent_coefficient is not None
        else "Coeficiente reciente sin comparar"
    )
    if sell_signal:
        note = (
            f"El coeficiente frente a {position.analysis_reference_label} se debilita de forma continuada: "
            f"{deterioration_streak} {periods_label} seguidos con peor lectura y ultimo tramo en {recent_coefficient:.2f}. "
            f"Es una senal de venta si la posicion sigue en cartera."
        )
        return {
            **base_payload,
            "available": True,
            "label": "Vender",
            "tone": "sell",
            "note": note,
            "trigger_label": trigger_label,
            "long_term_coefficient": baseline,
            "recent_coefficient": recent_coefficient,
            "coefficient_gap": coefficient_gap,
            "deterioration_streak": deterioration_streak,
            "weak_streak": weak_streak,
            "negative_streak": negative_streak,
            "periods_label": periods_label,
            "window_size": recent_window,
            "windows_count": len(rolling_rows),
            "trend_slope": trend_slope,
        }

    if watch_signal:
        note = (
            f"El coeficiente frente a {position.analysis_reference_label} ha perdido consistencia en el tramo reciente. "
            f"Todavia no fuerza venta inmediata, pero conviene vigilar si el deterioro se prolonga."
        )
        return {
            **base_payload,
            "available": True,
            "label": "Vigilar",
            "tone": "watch",
            "note": note,
            "trigger_label": trigger_label,
            "long_term_coefficient": baseline,
            "recent_coefficient": recent_coefficient,
            "coefficient_gap": coefficient_gap,
            "deterioration_streak": deterioration_streak,
            "weak_streak": weak_streak,
            "negative_streak": negative_streak,
            "periods_label": periods_label,
            "window_size": recent_window,
            "windows_count": len(rolling_rows),
            "trend_slope": trend_slope,
        }

    note = (
        f"El coeficiente frente a {position.analysis_reference_label} sigue estable en el tramo reciente "
        f"y no deja una senal continuada de venta."
    )
    return {
        **base_payload,
        "available": True,
        "label": "Estable",
        "tone": "neutral",
        "note": note,
        "trigger_label": trigger_label,
        "long_term_coefficient": baseline,
        "recent_coefficient": recent_coefficient,
        "coefficient_gap": coefficient_gap,
        "deterioration_streak": deterioration_streak,
        "weak_streak": weak_streak,
        "negative_streak": negative_streak,
        "periods_label": periods_label,
        "window_size": recent_window,
        "windows_count": len(rolling_rows),
        "trend_slope": trend_slope,
    }


def build_projection_confidence(
    coefficient: Decimal | None,
    observations_count: int,
    monthly_returns_count: int,
    years_covered: Decimal = ZERO,
    positive_year_ratio_pct: Decimal | None = None,
    stability_gap: Decimal | None = None,
) -> dict:
    score = 0
    absolute_coefficient = abs(coefficient) if coefficient is not None else None
    if absolute_coefficient is not None:
        if absolute_coefficient >= Decimal("0.20"):
            score += 1
        if absolute_coefficient >= Decimal("0.45"):
            score += 1
    if observations_count >= 12:
        score += 1
    if observations_count >= 36:
        score += 1
    if observations_count >= 72:
        score += 1
    if monthly_returns_count >= 24:
        score += 1
    if years_covered >= Decimal("5.00"):
        score += 1
    if years_covered >= Decimal("8.00"):
        score += 1
    if positive_year_ratio_pct is not None and positive_year_ratio_pct >= Decimal("55.00"):
        score += 1
    if positive_year_ratio_pct is not None and positive_year_ratio_pct >= Decimal("70.00"):
        score += 1
    if stability_gap is not None and stability_gap <= Decimal("0.20"):
        score += 1

    if score >= 8:
        return {
            "label": "Alta",
            "note": "Hay historico largo, ciclo coherente y una relacion bastante estable con la referencia elegida.",
            "score_pct": Decimal("85.00"),
        }
    if score >= 5:
        return {
            "label": "Media",
            "note": "La lectura es util, aunque la fiabilidad sigue siendo orientativa y conviene vigilar la volatilidad.",
            "score_pct": Decimal("65.00"),
        }
    return {
        "label": "Baja",
        "note": "La proyeccion se apoya en una base limitada o en una relacion debil con la referencia, asi que hay que leerla con prudencia.",
        "score_pct": Decimal("40.00"),
    }


def build_safety_score(cycle_metrics: dict, correlation: dict) -> dict:
    score = Decimal("62.00")
    annualized_volatility_pct = cycle_metrics.get("annualized_volatility_pct")
    max_drawdown_pct = cycle_metrics.get("max_drawdown_pct")
    current_drawdown_pct = cycle_metrics.get("current_drawdown_pct")
    positive_year_ratio_pct = cycle_metrics.get("positive_year_ratio_pct")
    cagr_pct = cycle_metrics.get("cagr_pct")
    coefficient = correlation.get("coefficient")

    if annualized_volatility_pct is not None:
        score -= clamp_decimal(annualized_volatility_pct - Decimal("16.00"), ZERO, Decimal("26.00"))
    if max_drawdown_pct is not None:
        score -= clamp_decimal(abs(max_drawdown_pct) * Decimal("0.45"), ZERO, Decimal("18.00"))
    if current_drawdown_pct is not None and current_drawdown_pct >= Decimal("-10.00"):
        score += Decimal("6.00")
    elif current_drawdown_pct is not None and current_drawdown_pct <= Decimal("-25.00"):
        score -= Decimal("6.00")
    if positive_year_ratio_pct is not None:
        score += clamp_decimal((positive_year_ratio_pct - Decimal("50.00")) * Decimal("0.18"), Decimal("-8.00"), Decimal("10.00"))
    if cagr_pct is not None and cagr_pct >= Decimal("6.00"):
        score += Decimal("5.00")
    if coefficient is not None and abs(coefficient) >= Decimal("0.35"):
        score += Decimal("4.00")
    if correlation.get("stability_label") == "Estable":
        score += Decimal("4.00")
    score = clamp_decimal(score, Decimal("15.00"), Decimal("92.00"))

    if score >= Decimal("72.00"):
        label = "Alta"
    elif score >= Decimal("56.00"):
        label = "Media"
    else:
        label = "Baja"
    return {
        "score": score,
        "label": label,
    }


def project_price_from_return(current_price: Decimal | None, return_pct: Decimal | None) -> Decimal | None:
    if current_price in {None, ZERO} or return_pct is None:
        return None
    multiplier = Decimal("1") + (return_pct / Decimal("100"))
    if multiplier <= 0:
        multiplier = Decimal("0.01")
    return current_price * multiplier


def resolve_projection_path_quarter_weights(
    annual_return_pct: Decimal | None,
    cycle_phase: str = "Transicion",
) -> list[Decimal]:
    if (annual_return_pct or ZERO) >= ZERO:
        return {
            "Correccion": [Decimal("0.12"), Decimal("0.20"), Decimal("0.28"), Decimal("0.40")],
            "Recuperacion": [Decimal("0.30"), Decimal("0.28"), Decimal("0.23"), Decimal("0.19")],
            "Expansion": [Decimal("0.21"), Decimal("0.24"), Decimal("0.26"), Decimal("0.29")],
            "Transicion": [Decimal("0.16"), Decimal("0.22"), Decimal("0.27"), Decimal("0.35")],
        }.get(cycle_phase or "Transicion", [Decimal("0.16"), Decimal("0.22"), Decimal("0.27"), Decimal("0.35")])

    return {
        "Correccion": [Decimal("0.42"), Decimal("0.28"), Decimal("0.18"), Decimal("0.12")],
        "Recuperacion": [Decimal("0.34"), Decimal("0.27"), Decimal("0.22"), Decimal("0.17")],
        "Expansion": [Decimal("0.28"), Decimal("0.26"), Decimal("0.24"), Decimal("0.22")],
        "Transicion": [Decimal("0.35"), Decimal("0.28"), Decimal("0.22"), Decimal("0.15")],
    }.get(cycle_phase or "Transicion", [Decimal("0.35"), Decimal("0.28"), Decimal("0.22"), Decimal("0.15")])


def resolve_projection_path_month_weights(
    annual_return_pct: Decimal | None,
    cycle_phase: str = "Transicion",
) -> list[Decimal]:
    quarter_weights = resolve_projection_path_quarter_weights(
        annual_return_pct,
        cycle_phase=cycle_phase,
    )
    if (annual_return_pct or ZERO) >= ZERO:
        month_split = {
            "Correccion": [Decimal("0.14"), Decimal("0.29"), Decimal("0.57")],
            "Recuperacion": [Decimal("0.48"), Decimal("0.31"), Decimal("0.21")],
            "Expansion": [Decimal("0.24"), Decimal("0.33"), Decimal("0.43")],
            "Transicion": [Decimal("0.18"), Decimal("0.31"), Decimal("0.51")],
        }.get(cycle_phase or "Transicion", [Decimal("0.18"), Decimal("0.31"), Decimal("0.51")])
    else:
        month_split = {
            "Correccion": [Decimal("0.54"), Decimal("0.29"), Decimal("0.17")],
            "Recuperacion": [Decimal("0.42"), Decimal("0.33"), Decimal("0.25")],
            "Expansion": [Decimal("0.37"), Decimal("0.33"), Decimal("0.30")],
            "Transicion": [Decimal("0.49"), Decimal("0.30"), Decimal("0.21")],
        }.get(cycle_phase or "Transicion", [Decimal("0.49"), Decimal("0.30"), Decimal("0.21")])

    month_weights = []
    for quarter_weight in quarter_weights:
        for month_weight in month_split:
            month_weights.append(quarter_weight * month_weight)
    return month_weights


def build_projection_path_from_weights(
    current_price: Decimal | None,
    annual_return_pct: Decimal | None,
    *,
    weights: list[Decimal],
    anchor_date: date | None = None,
    months_per_step: int = 1,
) -> list[dict]:
    if current_price in {None, ZERO} or annual_return_pct is None or not weights or months_per_step <= 0:
        return []

    annual_multiplier = max(0.01, 1 + (float(annual_return_pct) / 100))
    cumulative_share = Decimal("0.00")
    path = []
    for step_index, share in enumerate(weights, start=1):
        cumulative_share += share
        projected_multiplier = annual_multiplier ** float(cumulative_share)
        projected_price = Decimal(str(round(float(current_price) * projected_multiplier, 4)))
        months = step_index * months_per_step
        if months % 12 == 0:
            label = f"{months // 12}A"
        else:
            label = f"{months}M"
        path.append(
            {
                "label": label,
                "projected_date": anchor_date + timedelta(days=int(round((365 * months) / 12))) if anchor_date else None,
                "projected_price": projected_price,
            }
        )
    return path


def build_projection_path(
    current_price: Decimal | None,
    annual_return_pct: Decimal | None,
    anchor_date: date | None = None,
    cycle_phase: str = "Transicion",
) -> list[dict]:
    quarter_weights = resolve_projection_path_quarter_weights(annual_return_pct, cycle_phase=cycle_phase)
    return build_projection_path_from_weights(
        current_price,
        annual_return_pct,
        weights=quarter_weights,
        anchor_date=anchor_date,
        months_per_step=3,
    )


def build_monthly_projection_path(
    current_price: Decimal | None,
    annual_return_pct: Decimal | None,
    anchor_date: date | None = None,
    cycle_phase: str = "Transicion",
) -> list[dict]:
    month_weights = resolve_projection_path_month_weights(
        annual_return_pct,
        cycle_phase=cycle_phase,
    )
    return build_projection_path_from_weights(
        current_price,
        annual_return_pct,
        weights=month_weights,
        anchor_date=anchor_date,
        months_per_step=1,
    )


def build_quarterly_projection_path_from_monthly_path(monthly_path: list[dict]) -> list[dict]:
    quarterly_path = []
    for step in monthly_path or []:
        months_offset = parse_projection_label_months(step.get("label"))
        if months_offset in {3, 6, 9, 12}:
            quarterly_path.append(step)
    return quarterly_path


def build_cycle_zoomed_monthly_projection_path(
    current_price: Decimal | None,
    target_price: Decimal | None,
    *,
    anchor_date: date | None,
    cycle_projection: dict | None,
    months: int = 12,
) -> list[dict]:
    cycle_projection = cycle_projection or {}
    if (
        current_price in {None, ZERO}
        or target_price in {None, ZERO}
        or anchor_date is None
        or months <= 0
        or not cycle_projection.get("available")
    ):
        return []

    normalized_current_price = quantize_decimal(current_price, "0.0001") or Decimal(str(current_price))
    normalized_target_price = quantize_decimal(target_price, "0.0001") or Decimal(str(target_price))
    anchor_prices_by_month = {0: normalized_current_price}
    for step in cycle_projection.get("path") or []:
        step_price = quantize_decimal(step.get("projected_price"), "0.0001")
        if step_price in {None, ZERO}:
            continue
        months_offset = parse_projection_label_months(step.get("label"))
        if months_offset is None:
            projected_date = step.get("projected_date")
            if projected_date and anchor_date:
                months_offset = max(
                    ((projected_date.year - anchor_date.year) * 12)
                    + (projected_date.month - anchor_date.month),
                    0,
                )
        if months_offset is None or months_offset <= 0:
            continue
        anchor_prices_by_month[months_offset] = step_price

    anchor_months = sorted(anchor_prices_by_month.keys())
    if len(anchor_months) < 2 or anchor_months[-1] < months:
        return []

    def interpolate_source_value(month_number: int) -> Decimal | None:
        if month_number in anchor_prices_by_month:
            return anchor_prices_by_month[month_number]
        previous_months = [value for value in anchor_months if value < month_number]
        next_months = [value for value in anchor_months if value > month_number]
        if not previous_months or not next_months:
            return None
        start_month = previous_months[-1]
        end_month = next_months[0]
        start_price = anchor_prices_by_month.get(start_month)
        end_price = anchor_prices_by_month.get(end_month)
        if start_price in {None, ZERO} or end_price in {None, ZERO}:
            return None
        segment_months = max(end_month - start_month, 1)
        progress = (month_number - start_month) / segment_months
        if start_price > ZERO and end_price > ZERO:
            projected_value = Decimal(
                str(round(float(start_price) * ((float(end_price / start_price)) ** progress), 4))
            )
        else:
            projected_value = quantize_decimal(
                start_price + ((end_price - start_price) * Decimal(str(progress))),
                "0.0001",
            )
        return projected_value

    raw_monthly_values = []
    for month_number in range(1, months + 1):
        source_value = interpolate_source_value(month_number)
        if source_value in {None, ZERO}:
            return []
        raw_monthly_values.append((month_number, source_value))

    raw_end_price = raw_monthly_values[-1][1]
    if raw_end_price in {None, ZERO}:
        return []

    monthly_path = []
    for month_number, raw_source_value in raw_monthly_values:
        progress = Decimal(str(month_number / months))
        raw_baseline_value = normalized_current_price
        target_baseline_value = normalized_current_price
        if normalized_current_price > ZERO and raw_end_price > ZERO:
            raw_baseline_multiplier = Decimal(
                str(round((float(raw_end_price / normalized_current_price)) ** float(progress), 8))
            )
            raw_baseline_value = normalized_current_price * raw_baseline_multiplier
        if normalized_current_price > ZERO and normalized_target_price > ZERO:
            target_baseline_multiplier = Decimal(
                str(round((float(normalized_target_price / normalized_current_price)) ** float(progress), 8))
            )
            target_baseline_value = normalized_current_price * target_baseline_multiplier
        raw_excess_factor = (
            raw_source_value / raw_baseline_value
            if raw_baseline_value not in {None, ZERO}
            else Decimal("1.00")
        )
        raw_excess_factor = clamp_decimal(raw_excess_factor, Decimal("0.70"), Decimal("1.35"))
        projected_price = quantize_decimal(target_baseline_value * raw_excess_factor, "0.0001") or target_baseline_value
        projected_date = add_calendar_months(anchor_date, month_number)
        if projected_date is None:
            projected_date = anchor_date + timedelta(days=int(round((365 * month_number) / 12)))
        monthly_path.append(
            {
                "label": f"{month_number}M" if month_number < 12 else "1A",
                "projected_date": projected_date,
                "projected_price": projected_price,
            }
        )
    return monthly_path


def resolve_cycle_projection_step(
    cycle_projection: dict | None,
    *,
    months: int,
    anchor_date: date | None = None,
) -> dict | None:
    cycle_projection = cycle_projection or {}
    if months <= 0:
        return None

    exact_match = None
    dated_candidates = []
    for step in cycle_projection.get("path") or []:
        step_price = step.get("projected_price")
        if step_price in {None, ZERO}:
            continue
        months_offset = parse_projection_label_months(step.get("label"))
        if months_offset == months:
            exact_match = step
            break
        projected_date = step.get("projected_date")
        if projected_date is not None and anchor_date is not None:
            months_offset = max(
                ((projected_date.year - anchor_date.year) * 12)
                + (projected_date.month - anchor_date.month),
                0,
            )
            dated_candidates.append((abs(months_offset - months), months_offset, step))
    if exact_match is not None:
        return exact_match
    if dated_candidates:
        dated_candidates.sort(key=lambda item: (item[0], item[1]))
        return dated_candidates[0][2]
    return None


def align_projection_with_cycle_year_one(
    projection: dict | None,
    *,
    current_price: Decimal | None,
    cycle_projection: dict | None,
    anchor_date: date | None = None,
) -> dict:
    projection = projection or {}
    cycle_year_one_step = resolve_cycle_projection_step(cycle_projection, months=12, anchor_date=anchor_date)
    cycle_year_one_price = quantize_decimal(
        cycle_year_one_step.get("projected_price"),
        "0.0001",
    ) if cycle_year_one_step else None
    if (
        not projection.get("available")
        or current_price in {None, ZERO}
        or cycle_year_one_price in {None, ZERO}
    ):
        return projection

    projection["cycle_sync"] = {
        "applied": True,
        "source_label": "1A del patron 5A",
        "projected_price": cycle_year_one_price,
        "note": "El cierre de 12 meses replica el punto 1A de la curva 5A para que ambos graficos cuenten la misma historia.",
    }
    return projection


def synchronize_projection_path_with_cycle_zoom(
    projection: dict | None,
    cycle_projection: dict | None,
    *,
    current_price: Decimal | None,
    anchor_date: date | None,
) -> dict:
    projection = projection or {}
    cycle_projection = cycle_projection or {}
    if not projection.get("available"):
        return projection

    projection["history_window_label"] = "Ultimo ano visible"
    projection["path_source_label"] = "Escenario central 12M"
    projection["path_model_window_label"] = ""
    projection["uses_cycle_zoom_shape"] = False

    cycle_year_one_step = resolve_cycle_projection_step(cycle_projection, months=12, anchor_date=anchor_date)
    cycle_year_one_price = quantize_decimal(
        cycle_year_one_step.get("projected_price"),
        "0.0001",
    ) if cycle_year_one_step else None
    zoomed_monthly_path = build_cycle_zoomed_monthly_projection_path(
        current_price,
        cycle_year_one_price or projection.get("projected_price"),
        anchor_date=anchor_date,
        cycle_projection=cycle_projection,
        months=12,
    )
    if not zoomed_monthly_path:
        projection["quarterly_path"] = (
            build_quarterly_projection_path_from_monthly_path(projection.get("monthly_path") or [])
            or projection.get("quarterly_path")
            or []
        )
        return projection

    align_projection_with_cycle_year_one(
        projection,
        current_price=current_price,
        cycle_projection=cycle_projection,
        anchor_date=anchor_date,
    )
    projection["monthly_path"] = zoomed_monthly_path
    projection["quarterly_path"] = build_quarterly_projection_path_from_monthly_path(zoomed_monthly_path)
    projection["path_source_label"] = "Zoom 12M del patron 5A"
    projection["path_model_window_label"] = cycle_projection.get("model_window_label", "")
    projection["uses_cycle_zoom_shape"] = True
    return projection


def build_cycle_projection_path(
    current_price: Decimal | None,
    annual_return_pct: Decimal | None,
    annualized_volatility_pct: Decimal | None = None,
    current_drawdown_pct: Decimal | None = None,
    cycle_phase: str = "Transicion",
    anchor_date: date | None = None,
    years: int = 5,
    step_months: int = 6,
    step_return_pcts: list[Decimal] | None = None,
) -> list[dict]:
    if current_price in {None, ZERO} or annual_return_pct is None or years <= 0 or step_months <= 0:
        return []

    steps = int((years * 12) / step_months)
    annual_multiplier = max(0.01, 1 + (float(annual_return_pct) / 100))
    base_step_multiplier = annual_multiplier ** (step_months / 12)
    projected_price = Decimal(str(current_price))
    path = []

    for step in range(1, steps + 1):
        if step_return_pcts and step <= len(step_return_pcts):
            step_return_pct = step_return_pcts[step - 1]
            step_multiplier = max(0.80, 1 + (float(step_return_pct) / 100))
        else:
            amplitude_pct = clamp_decimal(
                (annualized_volatility_pct or Decimal("16.00")) * Decimal("0.18"),
                Decimal("1.50"),
                Decimal("5.50"),
            )
            phase_offset = {
                "Correccion": -math.pi / 2,
                "Recuperacion": -math.pi / 4,
                "Expansion": math.pi / 6,
                "Transicion": math.pi / 2,
            }.get(cycle_phase or "Transicion", math.pi / 2)
            cycle_length_steps = max(int(round((12 * 3) / step_months)), 4)
            angle_increment = (2 * math.pi) / cycle_length_steps
            cycle_wave = Decimal(str(round(math.sin(phase_offset + ((step - 1) * angle_increment)), 4)))
            cycle_adjustment_pct = cycle_wave * amplitude_pct * Decimal("0.40")
            if current_drawdown_pct is not None and current_drawdown_pct <= Decimal("-8.00") and step <= 2:
                rebound_bonus_pct = clamp_decimal(abs(current_drawdown_pct) * Decimal("0.08"), ZERO, Decimal("1.80"))
                cycle_adjustment_pct += rebound_bonus_pct / Decimal(str(step))
            elif current_drawdown_pct is not None and current_drawdown_pct >= Decimal("-3.00") and step == 1:
                cycle_adjustment_pct -= Decimal("0.80")
            step_multiplier = base_step_multiplier * max(0.85, 1 + (float(cycle_adjustment_pct) / 100))
        projected_price = Decimal(str(round(float(projected_price) * step_multiplier, 4)))
        months = step * step_months
        if months % 12 == 0:
            label = f"{months // 12}A"
        else:
            label = f"{months}M"
        path.append(
            {
                "label": label,
                "projected_date": anchor_date + timedelta(days=int(round((365 * months) / 12))) if anchor_date else None,
                "projected_price": projected_price,
            }
        )

    return path


def confidence_label_from_score(score_pct: Decimal | None) -> str:
    score_pct = score_pct if score_pct is not None else Decimal("40.00")
    if score_pct >= Decimal("75.00"):
        return "Alta"
    if score_pct >= Decimal("58.00"):
        return "Media"
    return "Baja"


def safety_label_from_score(score: Decimal | None) -> str:
    score = score if score is not None else Decimal("55.00")
    if score >= Decimal("72.00"):
        return "Alta"
    if score >= Decimal("56.00"):
        return "Media"
    return "Baja"


def build_scenario_probability_weights(
    confidence_label: str,
    *,
    shock_adjusted: bool = False,
    sentiment_score: Decimal | None = None,
) -> dict[str, Decimal]:
    if confidence_label == "Alta":
        bear_weight = Decimal("22.00")
        base_weight = Decimal("56.00")
        bull_weight = Decimal("22.00")
    elif confidence_label == "Media":
        bear_weight = Decimal("27.00")
        base_weight = Decimal("46.00")
        bull_weight = Decimal("27.00")
    else:
        bear_weight = Decimal("33.00")
        base_weight = Decimal("34.00")
        bull_weight = Decimal("33.00")

    if shock_adjusted:
        bear_weight += Decimal("5.00")
        bull_weight += Decimal("5.00")
        base_weight -= Decimal("10.00")

    if sentiment_score is not None:
        bounded_sentiment = clamp_decimal(Decimal(str(sentiment_score)), Decimal("-6.00"), Decimal("6.00"))
        sentiment_bias = clamp_decimal(abs(bounded_sentiment) * Decimal("1.40"), ZERO, Decimal("8.00"))
        if bounded_sentiment < ZERO:
            bear_weight += sentiment_bias
            bull_weight -= sentiment_bias
        elif bounded_sentiment > ZERO:
            bull_weight += sentiment_bias
            bear_weight -= sentiment_bias

    bear_weight = max(Decimal("12.00"), bear_weight)
    base_weight = max(Decimal("22.00"), base_weight)
    bull_weight = max(Decimal("12.00"), bull_weight)
    total_weight = bear_weight + base_weight + bull_weight
    if total_weight <= ZERO:
        return {
            "bear": Decimal("33.00"),
            "base": Decimal("34.00"),
            "bull": Decimal("33.00"),
        }
    return {
        "bear": quantize_decimal((bear_weight * ONE_HUNDRED) / total_weight, "0.1") or Decimal("33.0"),
        "base": quantize_decimal((base_weight * ONE_HUNDRED) / total_weight, "0.1") or Decimal("34.0"),
        "bull": quantize_decimal((bull_weight * ONE_HUNDRED) / total_weight, "0.1") or Decimal("33.0"),
    }


def build_one_year_projection_scenarios(
    current_price: Decimal | None,
    *,
    price_return_pct: Decimal | None,
    price_low_return_pct: Decimal | None,
    price_high_return_pct: Decimal | None,
    base_return_pct: Decimal | None,
    low_return_pct: Decimal | None,
    high_return_pct: Decimal | None,
    confidence_label: str,
    shock_adjusted: bool = False,
    sentiment_score: Decimal | None = None,
) -> list[dict]:
    probability_weights = build_scenario_probability_weights(
        confidence_label,
        shock_adjusted=shock_adjusted,
        sentiment_score=sentiment_score,
    )
    scenario_rows = [
        {
            "key": "bear",
            "label": "Bajista",
            "probability_pct": probability_weights["bear"],
            "price_return_pct": price_low_return_pct,
            "total_return_pct": low_return_pct,
            "projected_price": project_price_from_return(current_price, price_low_return_pct),
        },
        {
            "key": "base",
            "label": "Base",
            "probability_pct": probability_weights["base"],
            "price_return_pct": price_return_pct,
            "total_return_pct": base_return_pct,
            "projected_price": project_price_from_return(current_price, price_return_pct),
        },
        {
            "key": "bull",
            "label": "Alcista",
            "probability_pct": probability_weights["bull"],
            "price_return_pct": price_high_return_pct,
            "total_return_pct": high_return_pct,
            "projected_price": project_price_from_return(current_price, price_high_return_pct),
        },
    ]
    for row in scenario_rows:
        row["price_return_pct"] = quantize_decimal(row.get("price_return_pct"))
        row["total_return_pct"] = quantize_decimal(row.get("total_return_pct"))
        row["projected_price"] = quantize_decimal(row.get("projected_price"), "0.0001")
    return scenario_rows


def build_half_year_target_average(annual_return_pct: Decimal | None) -> Decimal:
    if annual_return_pct is None:
        return ZERO
    return Decimal(
        str(round(((max(0.01, 1 + (float(annual_return_pct) / 100)) ** 0.5) - 1) * 100, 4))
    )


def build_cycle_projection_path_for_target(
    current_price: Decimal | None,
    *,
    annual_return_pct: Decimal | None,
    step_return_pcts: list[Decimal],
    annualized_volatility_pct: Decimal | None,
    current_drawdown_pct: Decimal | None,
    cycle_phase: str,
    anchor_date: date | None,
    years: int = 5,
    step_months: int = 6,
) -> tuple[list[dict], list[Decimal], Decimal]:
    source_steps = [Decimal(str(value)) for value in (step_return_pcts or [])]
    current_average = average_decimal(source_steps) or ZERO
    target_average = build_half_year_target_average(annual_return_pct)
    step_shift = target_average - current_average
    adjusted_steps = [
        clamp_decimal(step_return + step_shift, Decimal("-14.00"), Decimal("14.00"))
        for step_return in source_steps
    ]
    path = build_cycle_projection_path(
        current_price,
        annual_return_pct,
        annualized_volatility_pct=annualized_volatility_pct,
        current_drawdown_pct=current_drawdown_pct,
        cycle_phase=cycle_phase,
        anchor_date=anchor_date,
        years=years,
        step_months=step_months,
        step_return_pcts=adjusted_steps,
    )
    return path, adjusted_steps, step_shift


def resolve_cycle_projection_scenario_spread(
    annualized_volatility_pct: Decimal | None,
    confidence_label: str,
    safety_label: str = "Media",
) -> Decimal:
    spread_pct = (
        annualized_volatility_pct * Decimal("0.18")
        if annualized_volatility_pct is not None
        else Decimal("3.60")
    )
    if confidence_label == "Alta":
        spread_pct -= Decimal("0.80")
    elif confidence_label == "Baja":
        spread_pct += Decimal("1.10")
    if safety_label == "Alta":
        spread_pct -= Decimal("0.45")
    elif safety_label == "Baja":
        spread_pct += Decimal("0.85")
    return clamp_decimal(spread_pct, Decimal("1.80"), Decimal("7.00"))


def build_five_year_projection_scenarios(
    current_price: Decimal | None,
    *,
    latest_date: date | None,
    annual_return_pct: Decimal | None,
    scenario_spread_annual_pct: Decimal,
    step_return_pcts: list[Decimal],
    annualized_volatility_pct: Decimal | None,
    current_drawdown_pct: Decimal | None,
    cycle_phase: str,
    confidence_label: str,
    shock_adjusted: bool = False,
    sentiment_score: Decimal | None = None,
) -> list[dict]:
    def extract_horizon_return_pct(path: list[dict], months: int) -> Decimal | None:
        for step in path:
            if parse_projection_label_months(step.get("label")) != months:
                continue
            return quantize_decimal(percentage_change(step.get("projected_price"), current_price), "0.01")
        return None

    probability_weights = build_scenario_probability_weights(
        confidence_label,
        shock_adjusted=shock_adjusted,
        sentiment_score=sentiment_score,
    )
    lower_annual_return_pct = clamp_decimal(
        (annual_return_pct or ZERO) - scenario_spread_annual_pct,
        Decimal("-14.00"),
        Decimal("20.00"),
    )
    upper_annual_return_pct = clamp_decimal(
        (annual_return_pct or ZERO) + scenario_spread_annual_pct,
        Decimal("-14.00"),
        Decimal("22.00"),
    )
    scenario_specs = [
        ("bear", "Bajista", probability_weights["bear"], lower_annual_return_pct),
        ("base", "Base", probability_weights["base"], annual_return_pct or ZERO),
        ("bull", "Alcista", probability_weights["bull"], upper_annual_return_pct),
    ]
    scenario_rows = []
    for key, label, probability_pct, scenario_annual_return_pct in scenario_specs:
        scenario_path, _, step_shift = build_cycle_projection_path_for_target(
            current_price,
            annual_return_pct=scenario_annual_return_pct,
            step_return_pcts=step_return_pcts,
            annualized_volatility_pct=annualized_volatility_pct,
            current_drawdown_pct=current_drawdown_pct,
            cycle_phase=cycle_phase,
            anchor_date=latest_date,
            years=5,
            step_months=6,
        )
        projected_price = scenario_path[-1]["projected_price"] if scenario_path else None
        scenario_rows.append(
            {
                "key": key,
                "label": label,
                "probability_pct": probability_pct,
                "annual_return_pct": quantize_decimal(scenario_annual_return_pct),
                "year_1_return_pct": extract_horizon_return_pct(scenario_path, 12),
                "year_2_return_pct": extract_horizon_return_pct(scenario_path, 24),
                "year_3_return_pct": extract_horizon_return_pct(scenario_path, 36),
                "year_4_return_pct": extract_horizon_return_pct(scenario_path, 48),
                "year_5_return_pct": extract_horizon_return_pct(scenario_path, 60),
                "five_year_return_pct": quantize_decimal(percentage_change(projected_price, current_price)),
                "projected_price": quantize_decimal(projected_price, "0.0001"),
                "step_shift": quantize_decimal(step_shift),
            }
        )
    return scenario_rows


def build_backtest_monthly_chart(rows: list[dict], width: int = 640, height: int = 220, padding: int = 18, max_points: int = 60) -> dict:
    normalized_rows = []
    for row in rows[-max_points:]:
        forecast_date = row.get("forecast_date")
        target_date = row.get("target_date")
        forecast_return_pct = row.get("forecast_return_pct")
        actual_return_pct = row.get("actual_return_pct")
        if forecast_date is None or target_date is None or forecast_return_pct is None or actual_return_pct is None:
            continue
        normalized_rows.append(
            {
                "forecast_date": forecast_date,
                "target_date": target_date,
                "forecast_return_pct": Decimal(str(forecast_return_pct)),
                "actual_return_pct": Decimal(str(actual_return_pct)),
            }
        )

    if len(normalized_rows) < 2:
        return {
            "available": False,
            "forecast_line": "",
            "actual_line": "",
            "forecast_points": [],
            "actual_points": [],
            "min_label": "-",
            "max_label": "-",
            "start_label": "",
            "end_label": "",
            "points_count": len(normalized_rows),
        }

    all_values = []
    for row in normalized_rows:
        all_values.append(row["forecast_return_pct"])
        all_values.append(row["actual_return_pct"])
    series_min = min(all_values)
    series_max = max(all_values)
    if series_min == series_max:
        series_max += Decimal("1")

    min_date = normalized_rows[0]["forecast_date"]
    max_date = normalized_rows[-1]["forecast_date"]
    total_days = max((max_date - min_date).days, 1)
    span_x = width - (padding * 2)
    span_y = height - (padding * 2)

    def scale_point(point_date: date, value: Decimal) -> tuple[float, float]:
        x = padding + (span_x * ((point_date - min_date).days / total_days))
        normalized_value = (value - series_min) / (series_max - series_min)
        y = height - padding - (normalized_value * span_y)
        return x, y

    def build_series(series_key: str) -> tuple[str, list[dict]]:
        line_points = []
        point_rows = []
        for row in normalized_rows:
            value = row[series_key]
            x, y = scale_point(row["forecast_date"], value)
            line_points.append(f"{x:.1f},{y:.1f}")
            point_rows.append(
                {
                    "x": f"{x:.1f}",
                    "y": f"{y:.1f}",
                    "value_label": f"{value:.1f} %",
                    "forecast_date_label": row["forecast_date"].isoformat(),
                    "target_date_label": row["target_date"].isoformat(),
                }
            )
        return " ".join(line_points), point_rows

    forecast_line, forecast_points = build_series("forecast_return_pct")
    actual_line, actual_points = build_series("actual_return_pct")
    return {
        "available": True,
        "forecast_line": forecast_line,
        "actual_line": actual_line,
        "forecast_points": forecast_points,
        "actual_points": actual_points,
        "min_label": f"{format_axis_value(series_min)} %",
        "max_label": f"{format_axis_value(series_max)} %",
        "start_label": normalized_rows[0]["forecast_date"].isoformat(),
        "end_label": normalized_rows[-1]["forecast_date"].isoformat(),
        "points_count": len(normalized_rows),
        "x_markers": build_time_axis_markers(
            [row["forecast_date"] for row in normalized_rows],
            width=width,
            height=height,
            padding=padding,
        ),
    }


def build_one_year_projection(
    history,
    position: EquityPosition,
    correlation: dict,
    six_month_snapshot: dict,
    cycle_metrics: dict | None = None,
    technical_signal: dict | None = None,
) -> dict:
    if not history:
        return {"available": False}

    latest_price = history[-1].close_price if history[-1].close_price else position.current_price_per_share
    latest_date = history[-1].price_date
    cycle_metrics = cycle_metrics or build_cycle_metrics(history)
    stock_6m_return_pct = six_month_snapshot.get("stock_return_pct") if six_month_snapshot.get("available") else None
    reference_6m_change = six_month_snapshot.get("benchmark_change") if six_month_snapshot.get("available") else None
    one_year_snapshot = build_period_snapshot(
        history,
        "1Y",
        start_date=latest_date - timedelta(days=365),
        end_date=latest_date,
        reference_profile=position.reference_profile,
    )
    three_year_snapshot = build_period_snapshot(
        history,
        "3Y",
        start_date=latest_date - timedelta(days=365 * 3),
        end_date=latest_date,
        reference_profile=position.reference_profile,
    )
    coefficient = correlation.get("coefficient")
    recent_coefficient = correlation.get("recent_coefficient")
    beta = correlation.get("beta")
    recent_beta = correlation.get("recent_beta")
    observations_count = correlation.get("observations_count", 0)
    price_return_components = []

    stock_6m_signal = build_projection_signal(stock_6m_return_pct, 6)
    if stock_6m_signal is not None:
        price_return_components.append(stock_6m_signal * Decimal("0.20"))
    if one_year_snapshot.get("available") and one_year_snapshot.get("stock_return_pct") is not None:
        price_return_components.append(one_year_snapshot["stock_return_pct"] * Decimal("0.30"))
    if three_year_snapshot.get("available") and three_year_snapshot.get("stock_return_pct") is not None:
        three_year_signal = annualize_return_pct(three_year_snapshot["stock_return_pct"], 36)
        if three_year_signal is not None:
            price_return_components.append(three_year_signal * Decimal("0.20"))
    if cycle_metrics.get("cagr_pct") is not None:
        price_return_components.append(cycle_metrics["cagr_pct"] * Decimal("0.30"))
    if cycle_metrics.get("current_drawdown_pct") is not None:
        current_drawdown_pct = cycle_metrics["current_drawdown_pct"]
        mean_reversion_pct = clamp_decimal(abs(current_drawdown_pct) * Decimal("0.18"), ZERO, Decimal("10.00"))
        if current_drawdown_pct <= Decimal("-8.00"):
            price_return_components.append(mean_reversion_pct)
        elif current_drawdown_pct >= Decimal("-3.00"):
            price_return_components.append(-mean_reversion_pct * Decimal("0.40"))

    reference_one_year_change = one_year_snapshot.get("benchmark_change") if one_year_snapshot.get("available") else None
    if beta is not None and reference_one_year_change is not None:
        price_return_components.append((beta * reference_one_year_change) * Decimal("0.15"))
    if coefficient is not None and recent_coefficient is not None:
        price_return_components.append(
            clamp_decimal((recent_coefficient - coefficient) * Decimal("8.00"), Decimal("-2.50"), Decimal("2.00"))
        )

    if not price_return_components:
        return {"available": False}

    price_return_pct = sum(price_return_components, ZERO)
    annualized_volatility_pct = cycle_metrics.get("annualized_volatility_pct")
    technical_signal = technical_signal or build_candlestick_metrics(history)
    confidence = build_projection_confidence(
        coefficient,
        observations_count,
        cycle_metrics.get("monthly_observations_count", 0),
        years_covered=cycle_metrics.get("years_covered", ZERO),
        positive_year_ratio_pct=cycle_metrics.get("positive_year_ratio_pct"),
        stability_gap=correlation.get("stability_gap"),
    )
    evidence_factor = Decimal("0.76")
    if confidence["label"] == "Alta":
        evidence_factor = Decimal("0.98")
    elif confidence["label"] == "Media":
        evidence_factor = Decimal("0.88")
    safety = build_safety_score(cycle_metrics, correlation)
    safety_multiplier = clamp_decimal(Decimal("0.88") + (safety["score"] / Decimal("600")), Decimal("0.88"), Decimal("1.05"))
    price_return_pct = clamp_decimal(
        price_return_pct * evidence_factor * safety_multiplier,
        Decimal("-35.00"),
        Decimal("40.00"),
    )
    model_price_return_pct = price_return_pct
    technical_return_adjustment_pct = quantize_decimal(technical_signal.get("return_adjustment_pct"), "0.01") or ZERO
    technical_band_multiplier = technical_signal.get("band_multiplier") or Decimal("1.00")
    technical_alignment_label = "Sin lectura"
    if technical_signal.get("available") and technical_signal.get("signal_score") is not None:
        technical_score = Decimal(str(technical_signal.get("signal_score") or ZERO))
        if technical_score * model_price_return_pct > ZERO and abs(technical_score) >= Decimal("1.40"):
            technical_alignment_label = "De apoyo"
        elif technical_score * model_price_return_pct < ZERO and abs(technical_score) >= Decimal("1.40"):
            technical_alignment_label = "En conflicto"
            technical_band_multiplier = clamp_decimal(technical_band_multiplier + Decimal("0.06"), Decimal("0.92"), Decimal("1.16"))
        else:
            technical_alignment_label = "Mixto"
    price_return_pct = clamp_decimal(
        model_price_return_pct + technical_return_adjustment_pct,
        Decimal("-35.00"),
        Decimal("40.00"),
    )

    analysis_value, analysis_value_source = resolve_projection_analysis_value(position, latest_price)
    annual_dividend_income = build_projection_dividend_income(position, analysis_value)
    broker_costs = build_analysis_broker_costs(position, analysis_value, annual_dividend_income)
    net_income_yield_pct = None
    transaction_drag_pct = None
    gross_dividend_yield_pct = None
    annual_cost_used = broker_costs.get("annual_cost_used", ZERO)
    net_dividend_income = broker_costs.get("net_dividend_income", ZERO)
    net_annual_income = net_dividend_income - annual_cost_used
    if analysis_value and analysis_value > 0:
        gross_dividend_yield_pct = (annual_dividend_income / analysis_value) * ONE_HUNDRED
        net_income_yield_pct = (net_annual_income / analysis_value) * ONE_HUNDRED
        transaction_drag_pct = (broker_costs.get("roundtrip_total_cost", ZERO) / analysis_value) * ONE_HUNDRED
    base_return_pct = price_return_pct
    if net_income_yield_pct is not None:
        base_return_pct += net_income_yield_pct
    if transaction_drag_pct is not None:
        base_return_pct -= transaction_drag_pct
    base_return_pct = clamp_decimal(base_return_pct, Decimal("-45.00"), Decimal("45.00"))

    band_pct = annualized_volatility_pct * Decimal("0.65") if annualized_volatility_pct is not None else Decimal("16.00")
    confidence_score_pct = confidence["score_pct"]
    if technical_signal.get("available") and technical_signal.get("signal_score") is not None:
        technical_score = Decimal(str(technical_signal.get("signal_score") or ZERO))
        confidence_delta = ZERO
        if technical_alignment_label == "De apoyo":
            confidence_delta = clamp_decimal(abs(technical_score) * Decimal("1.05"), ZERO, Decimal("5.00"))
        elif technical_alignment_label == "En conflicto":
            confidence_delta = -clamp_decimal(abs(technical_score) * Decimal("1.20"), ZERO, Decimal("6.00"))
        elif technical_alignment_label == "Mixto":
            confidence_delta = clamp_decimal(
                (Decimal(str(technical_signal.get("confidence_score") or Decimal("55.00"))) - Decimal("60.00")) * Decimal("0.12"),
                Decimal("-2.00"),
                Decimal("2.00"),
            )
        confidence_score_pct = clamp_decimal(confidence_score_pct + confidence_delta, Decimal("35.00"), Decimal("90.00"))
    confidence_label = resolve_projection_confidence_label(confidence_score_pct)
    confidence_note = confidence["note"]
    if technical_signal.get("available") and technical_signal.get("signal_label"):
        confidence_note = (
            f"{confidence_note} La lectura tecnica reciente ({technical_signal.get('signal_label', '').lower()}) "
            f"entra como capa adicional y hoy queda {technical_alignment_label.lower()} frente al escenario base."
        )

    if confidence_label == "Alta":
        band_pct -= Decimal("2.50")
    elif confidence_label == "Baja":
        band_pct += Decimal("5.00")
    if safety["label"] == "Alta":
        band_pct -= Decimal("1.50")
    elif safety["label"] == "Baja":
        band_pct += Decimal("3.00")
    band_pct *= technical_band_multiplier
    band_pct = clamp_decimal(band_pct, Decimal("8.00"), Decimal("32.00"))

    low_return_pct = clamp_decimal(base_return_pct - band_pct, Decimal("-80.00"), Decimal("120.00"))
    high_return_pct = clamp_decimal(base_return_pct + band_pct, Decimal("-80.00"), Decimal("140.00"))
    benefit_risk_ratio = None
    if band_pct > 0:
        benefit_risk_ratio = base_return_pct / band_pct
    decision_score = base_return_pct * projection_confidence_multiplier(confidence_label) * (safety["score"] / ONE_HUNDRED)
    years_covered = cycle_metrics.get("years_covered", ZERO)
    if coefficient is None or beta is None or reference_one_year_change is None:
        explanation = (
            f"La proyeccion 12M mezcla el ciclo propio de la accion y su historico de {years_covered:.2f} anos. "
            f"Hoy pesa mas la trayectoria del valor que la referencia porque la relacion con {position.analysis_reference_label} "
            f"todavia no es lo bastante robusta."
        )
    else:
        change_mode = correlation.get("change_mode", "pct")
        reference_change_unit = change_unit_label(change_mode)
        explanation = (
            f"La proyeccion usa hasta {years_covered:.2f} anos de serie, fase {cycle_metrics.get('cycle_phase', 'sin ciclo').lower()} "
            f"y la referencia {position.analysis_reference_label} con correlacion 10A de {coefficient:.2f}. "
            f"El beta historico frente a esa referencia es {beta:.2f} por {reference_change_unit}. "
            f"El retorno total incluye dividendos netos y penaliza comisiones y custodia del broker."
        )
    if coefficient is not None and recent_coefficient is not None:
        if recent_coefficient <= coefficient - Decimal("0.15"):
            explanation += (
                f" El coeficiente reciente cae a {recent_coefficient:.2f}, por debajo del 10A, "
                "asi que el modelo enfria la lectura de ciclo."
            )
        elif recent_coefficient >= coefficient + Decimal("0.15"):
            explanation += (
                f" El coeficiente reciente sube a {recent_coefficient:.2f}, por encima del 10A, "
                "y refuerza la lectura de ciclo actual."
            )
    if analysis_value_source == "normalized_watchlist":
        explanation += (
            f" Como es un valor en seguimiento, los costes y dividendos se normalizan sobre un ticket analitico de "
            f"{analysis_value:.0f} EUR para que una sola accion de muestra no distorsione la lectura."
        )
    if technical_signal.get("available") and technical_signal.get("signal_label"):
        explanation += (
            f" La capa tecnica de velas, tendencia, RSI y soportes/resistencias apunta "
            f"{str(technical_signal.get('signal_label') or '').lower()} "
            f"y queda {technical_alignment_label.lower()} frente al modelo principal."
        )

    price_low_return_pct = clamp_decimal(price_return_pct - band_pct, Decimal("-80.00"), Decimal("120.00"))
    price_high_return_pct = clamp_decimal(price_return_pct + band_pct, Decimal("-80.00"), Decimal("140.00"))
    scenarios = build_one_year_projection_scenarios(
        latest_price,
        price_return_pct=price_return_pct,
        price_low_return_pct=price_low_return_pct,
        price_high_return_pct=price_high_return_pct,
        base_return_pct=base_return_pct,
        low_return_pct=low_return_pct,
        high_return_pct=high_return_pct,
        confidence_label=confidence_label,
    )
    monthly_projection_path = build_monthly_projection_path(
        latest_price,
        price_return_pct,
        anchor_date=history[-1].price_date,
        cycle_phase=cycle_metrics.get("cycle_phase") or "Transicion",
    )
    quarterly_projection_path = build_quarterly_projection_path_from_monthly_path(monthly_projection_path)
    if not quarterly_projection_path:
        quarterly_projection_path = build_projection_path(
            latest_price,
            price_return_pct,
            anchor_date=history[-1].price_date,
            cycle_phase=cycle_metrics.get("cycle_phase") or "Transicion",
        )

    return {
        "available": True,
        "price_return_pct": price_return_pct,
        "price_low_return_pct": price_low_return_pct,
        "price_high_return_pct": price_high_return_pct,
        "base_return_pct": base_return_pct,
        "band_pct": quantize_decimal(band_pct),
        "low_return_pct": low_return_pct,
        "high_return_pct": high_return_pct,
        "projected_price": project_price_from_return(latest_price, price_return_pct),
        "low_price": project_price_from_return(latest_price, price_low_return_pct),
        "high_price": project_price_from_return(latest_price, price_high_return_pct),
        "monthly_path": monthly_projection_path,
        "quarterly_path": quarterly_projection_path,
        "scenarios": scenarios,
        "confidence_label": confidence_label,
        "confidence_note": confidence_note,
        "confidence_score_pct": quantize_decimal(confidence_score_pct),
        "safety_score": safety["score"],
        "safety_label": safety["label"],
        "benefit_risk_ratio": benefit_risk_ratio,
        "decision_score": decision_score,
        "model_price_return_pct": quantize_decimal(model_price_return_pct),
        "stock_6m_return_pct": stock_6m_return_pct,
        "reference_6m_return_pct": reference_6m_change,
        "stock_1y_return_pct": one_year_snapshot.get("stock_return_pct") if one_year_snapshot.get("available") else None,
        "reference_1y_return_pct": reference_one_year_change,
        "coefficient": coefficient,
        "recent_coefficient": recent_coefficient,
        "beta": beta,
        "recent_beta": recent_beta,
        "reference_change_mode": correlation.get("change_mode", "pct"),
        "cycle_phase": cycle_metrics.get("cycle_phase"),
        "years_covered": years_covered,
        "cagr_pct": cycle_metrics.get("cagr_pct"),
        "annualized_volatility_pct": annualized_volatility_pct,
        "max_drawdown_pct": cycle_metrics.get("max_drawdown_pct"),
        "current_drawdown_pct": cycle_metrics.get("current_drawdown_pct"),
        "positive_year_ratio_pct": cycle_metrics.get("positive_year_ratio_pct"),
        "gross_dividend_yield_pct": gross_dividend_yield_pct,
        "net_income_yield_pct": net_income_yield_pct,
        "transaction_drag_pct": transaction_drag_pct,
        "analysis_value_amount": quantize_decimal(analysis_value, "0.01") or ZERO,
        "analysis_value_source": analysis_value_source,
        "annual_dividend_income_used": annual_dividend_income,
        "annual_cost_used": annual_cost_used,
        "annual_cost_source": broker_costs.get("annual_cost_source", "broker"),
        "net_dividend_income": net_dividend_income,
        "broker_costs": broker_costs,
        "reference_label": position.analysis_reference_label,
        "latest_price": quantize_decimal(latest_price, "0.0001"),
        "latest_date": latest_date,
        "history_window_label": "Ultimo ano visible",
        "path_source_label": "Escenario central 12M",
        "path_model_window_label": "",
        "uses_cycle_zoom_shape": False,
        "technical_adjustment": {
            "applied": bool(technical_signal.get("available")),
            "signal_label": technical_signal.get("signal_label"),
            "signal_score": technical_signal.get("signal_score"),
            "confidence_label": technical_signal.get("confidence_label"),
            "alignment_label": technical_alignment_label,
            "return_adjustment_pct": quantize_decimal(technical_return_adjustment_pct),
            "band_multiplier": quantize_decimal(technical_band_multiplier, "0.01"),
            "note": technical_signal.get("note") or "",
        },
        "explanation": explanation,
    }


def find_closest_history_point(history, target_date: date, tolerance_days: int = 20):
    candidates = [point for point in history if abs((point.price_date - target_date).days) <= tolerance_days]
    if not candidates:
        return None
    return min(candidates, key=lambda point: (abs((point.price_date - target_date).days), point.price_date))


def build_projection_backtest(
    history,
    position: EquityPosition,
    max_rows: int = 8,
    include_monthly_chart: bool = True,
) -> dict:
    monthly_points = collapse_history_to_frequency(history, "monthly")
    if len(monthly_points) < 20:
        return {
            "available": False,
            "comparisons_count": 0,
            "rows": [],
            "monthly_chart": {"available": False},
            "mean_absolute_error_pct": None,
            "direction_hit_rate_pct": None,
            "in_range_rate_pct": None,
            "precision_label": "Sin historico suficiente",
            "plain_explanation": "Todavia no hay suficiente historico mensual para comprobar como habria funcionado el modelo en el pasado.",
            "error_explanation": "Error medio: diferencia media entre lo que proyecto el modelo y lo que ocurrio realmente 12 meses despues.",
            "direction_explanation": "Acierto de direccion: porcentaje de veces que el modelo acerto si la accion subia o bajaba.",
            "range_explanation": "Dentro del rango: porcentaje de veces que el resultado real quedo dentro del escenario previsto.",
        }

    rows = []
    for anchor_point in monthly_points:
        anchor_date = anchor_point.price_date
        if anchor_date - monthly_points[0].price_date < timedelta(days=182):
            continue

        future_point = find_closest_history_point(history, anchor_date + timedelta(days=365))
        if future_point is None or future_point.price_date <= anchor_date:
            continue

        visible_history = [point for point in history if point.price_date <= anchor_date]
        six_month_snapshot = build_period_snapshot(
            visible_history,
            "6M",
            start_date=anchor_date - timedelta(days=182),
            end_date=anchor_date,
            reference_profile=position.reference_profile,
        )
        if not six_month_snapshot["available"]:
            continue

        correlation = build_reference_correlation(visible_history, position)
        projection = build_one_year_projection(visible_history, position, correlation, six_month_snapshot)
        if not projection.get("available"):
            continue

        actual_return_pct = percentage_change(future_point.close_price, anchor_point.close_price)
        if actual_return_pct is None:
            continue

        forecast_return_pct = projection.get("price_return_pct")
        if forecast_return_pct is None:
            continue
        absolute_error_pct = abs(forecast_return_pct - actual_return_pct)
        direction_hit = (
            (forecast_return_pct >= 0 and actual_return_pct >= 0)
            or (forecast_return_pct < 0 and actual_return_pct < 0)
        )
        in_range = False
        if projection.get("price_low_return_pct") is not None and projection.get("price_high_return_pct") is not None:
            in_range = projection["price_low_return_pct"] <= actual_return_pct <= projection["price_high_return_pct"]

        rows.append(
            {
                "forecast_date": anchor_date,
                "target_date": future_point.price_date,
                "forecast_return_pct": quantize_decimal(forecast_return_pct),
                "actual_return_pct": quantize_decimal(actual_return_pct),
                "absolute_error_pct": quantize_decimal(absolute_error_pct),
                "direction_hit": direction_hit,
                "in_range": in_range,
                "forecast_price": quantize_decimal(projection.get("projected_price"), "0.0001"),
                "actual_price": quantize_decimal(future_point.close_price, "0.0001"),
                "confidence_label": projection.get("confidence_label"),
            }
        )

    if not rows:
        return {
            "available": False,
            "comparisons_count": 0,
            "rows": [],
            "monthly_chart": {"available": False},
            "mean_absolute_error_pct": None,
            "direction_hit_rate_pct": None,
            "in_range_rate_pct": None,
            "precision_label": "Sin historico suficiente",
            "plain_explanation": "Todavia no hay suficiente historico mensual para comprobar como habria funcionado el modelo en el pasado.",
            "error_explanation": "Error medio: diferencia media entre lo que proyecto el modelo y lo que ocurrio realmente 12 meses despues.",
            "direction_explanation": "Acierto de direccion: porcentaje de veces que el modelo acerto si la accion subia o bajaba.",
            "range_explanation": "Dentro del rango: porcentaje de veces que el resultado real quedo dentro del escenario previsto.",
        }

    recent_rows = list(reversed(rows[-max_rows:]))
    comparisons_count = len(rows)
    monthly_chart = build_backtest_monthly_chart(rows) if include_monthly_chart else {"available": False}
    mean_absolute_error_pct = quantize_decimal(
        sum((row["absolute_error_pct"] for row in rows), ZERO) / Decimal(comparisons_count)
    )
    direction_hit_rate_pct = quantize_decimal(
        Decimal(sum(1 for row in rows if row["direction_hit"])) * Decimal("100") / Decimal(comparisons_count)
    )
    in_range_rate_pct = quantize_decimal(
        Decimal(sum(1 for row in rows if row["in_range"])) * Decimal("100") / Decimal(comparisons_count)
    )

    if mean_absolute_error_pct <= Decimal("8") and direction_hit_rate_pct >= Decimal("70"):
        precision_label = "Alta"
    elif mean_absolute_error_pct <= Decimal("15") and direction_hit_rate_pct >= Decimal("55"):
        precision_label = "Media"
    else:
        precision_label = "Baja"

    return {
        "available": True,
        "comparisons_count": comparisons_count,
        "rows": recent_rows,
        "monthly_chart": monthly_chart,
        "mean_absolute_error_pct": mean_absolute_error_pct,
        "direction_hit_rate_pct": direction_hit_rate_pct,
        "in_range_rate_pct": in_range_rate_pct,
        "precision_label": precision_label,
        "plain_explanation": (
            "Cada punto mensual responde a una pregunta sencilla: si ese mes hubieramos lanzado la proyeccion a 12 meses, "
            "que habria dicho el modelo y que paso realmente un ano despues."
        ),
        "error_explanation": "Error medio: diferencia media entre la prediccion y el resultado real 12 meses despues.",
        "direction_explanation": "Acierto de direccion: veces que el modelo acerto si la accion terminaria subiendo o bajando.",
        "range_explanation": "Dentro del rango: veces que el resultado real quedo dentro del escenario previsto por el modelo.",
    }


def build_projection_reliability(projection: dict, backtest: dict) -> dict:
    projection_score = projection.get("confidence_score_pct") or Decimal("40.00")
    backtest_score_map = {
        "Alta": Decimal("82.00"),
        "Media": Decimal("64.00"),
        "Baja": Decimal("42.00"),
        "Sin historico suficiente": Decimal("45.00"),
    }
    if backtest.get("available"):
        backtest_score = backtest_score_map.get(backtest.get("precision_label"), Decimal("45.00"))
        reliability_score = (projection_score * Decimal("0.45")) + (backtest_score * Decimal("0.55"))
    else:
        reliability_score = projection_score

    if reliability_score >= Decimal("75.00"):
        label = "Alta"
    elif reliability_score >= Decimal("58.00"):
        label = "Media"
    else:
        label = "Baja"
    return {
        "score": quantize_decimal(reliability_score),
        "label": label,
    }


def build_trade_alert(
    position: EquityPosition,
    projection: dict,
    correlation: dict,
    reliability: dict,
    relative_trend: dict,
    six_month_snapshot: dict,
    one_year_snapshot: dict,
    valuation: dict | None = None,
    technical_signal: dict | None = None,
) -> dict:
    technical_signal = technical_signal or {}
    base_payload = {
        "label": "Vigilar",
        "tone": "watch",
        "score": ZERO,
        "note": "No hay una pendiente prolongada lo bastante clara frente a la referencia como para activar compra o venta.",
        "trend_label": relative_trend.get("label", "Sin tendencia relativa"),
        "periods_label": relative_trend.get("periods_label", "periodos"),
        "gap_slope_pct": relative_trend.get("gap_slope_pct"),
        "recent_gap_avg_pct": relative_trend.get("recent_gap_avg_pct"),
        "positive_streak": relative_trend.get("positive_streak", 0),
        "negative_streak": relative_trend.get("negative_streak", 0),
        "trigger_label": "Sin confirmacion prolongada",
        "technical_label": technical_signal.get("signal_label", "Sin lectura tecnica"),
        "technical_score": technical_signal.get("signal_score"),
    }
    if not projection.get("available"):
        base_payload["note"] = "Todavia no hay suficiente historico para activar una alerta operativa."
        return base_payload

    reliability_score = reliability.get("score") or Decimal("40.00")
    safety_score = projection.get("safety_score") or Decimal("55.00")
    projected_return_pct = projection.get("base_return_pct") or ZERO
    valuation_score = (valuation or {}).get("score") or ZERO
    one_year_alpha_pct = one_year_snapshot.get("alpha_pct") if one_year_snapshot.get("available") else None
    six_month_alpha_pct = six_month_snapshot.get("alpha_pct") if six_month_snapshot.get("available") else None
    trade_score = ZERO

    if relative_trend.get("prolonged_positive"):
        trade_score += Decimal("4.00")
    elif relative_trend.get("prolonged_negative"):
        trade_score -= Decimal("4.00")

    recent_gap_avg_pct = relative_trend.get("recent_gap_avg_pct")
    if recent_gap_avg_pct is not None:
        trade_score += clamp_decimal(recent_gap_avg_pct * Decimal("0.40"), Decimal("-2.00"), Decimal("2.00"))

    gap_slope_pct = relative_trend.get("gap_slope_pct")
    if gap_slope_pct is not None:
        trade_score += clamp_decimal(gap_slope_pct * Decimal("2.50"), Decimal("-1.50"), Decimal("1.50"))

    alpha_signal_pct = one_year_alpha_pct if one_year_alpha_pct is not None else six_month_alpha_pct
    if alpha_signal_pct is not None:
        trade_score += clamp_decimal(alpha_signal_pct * Decimal("0.12"), Decimal("-1.50"), Decimal("1.50"))

    trade_score += clamp_decimal((projected_return_pct - Decimal("4.00")) * Decimal("0.08"), Decimal("-1.50"), Decimal("1.50"))
    trade_score += clamp_decimal((safety_score - Decimal("55.00")) * Decimal("0.03"), Decimal("-0.75"), Decimal("0.75"))
    trade_score += clamp_decimal(valuation_score * Decimal("0.22"), Decimal("-1.25"), Decimal("1.25"))
    technical_score = technical_signal.get("signal_score")
    if technical_score is not None:
        technical_score = Decimal(str(technical_score))
        trade_score += clamp_decimal(technical_score * Decimal("0.28"), Decimal("-1.25"), Decimal("1.25"))
        if technical_signal.get("confidence_label") == "Alta" and abs(technical_score) >= Decimal("3.20"):
            trade_score += Decimal("0.35") if technical_score > ZERO else Decimal("-0.35")

    if reliability_score < Decimal("55.00"):
        trade_score *= Decimal("0.78")

    coefficient = correlation.get("coefficient")
    if coefficient is not None and abs(coefficient) >= Decimal("0.35"):
        trade_score += Decimal("0.35") if trade_score > ZERO else Decimal("-0.35") if trade_score < ZERO else ZERO

    positive_streak = relative_trend.get("positive_streak", 0)
    negative_streak = relative_trend.get("negative_streak", 0)
    periods_label = relative_trend.get("periods_label", "periodos")
    trend_label = relative_trend.get("label", "Sin tendencia relativa")
    technical_note = ""
    if technical_signal.get("available") and technical_signal.get("signal_label"):
        technical_note = (
            f" La lectura tecnica de velas y tendencia apunta {str(technical_signal.get('signal_label') or '').lower()}."
        )

    if trade_score >= Decimal("3.20") and projected_return_pct < ZERO:
        note = (
            f"La accion encadena {positive_streak} {periods_label} mejorando frente a su referencia ajustada por coeficiente, "
            f"pero el retorno neto 12M sigue en negativo. Conviene vigilar antes de activar compra."
        )
        note += technical_note
        return {
            **base_payload,
            "score": quantize_decimal(trade_score) or ZERO,
            "note": note,
            "trend_label": trend_label,
            "trigger_label": f"{positive_streak} {periods_label} con mejora relativa, pero neto 12M negativo",
        }

    if trade_score <= Decimal("-3.20") and projected_return_pct > ZERO:
        note = (
            f"La accion encadena {negative_streak} {periods_label} perdiendo fuerza frente a su referencia, "
            f"pero la proyeccion neta 12M todavia es positiva. Conviene vigilar antes de activar venta."
        )
        note += technical_note
        return {
            **base_payload,
            "score": quantize_decimal(trade_score) or ZERO,
            "note": note,
            "trend_label": trend_label,
            "trigger_label": f"{negative_streak} {periods_label} con deterioro relativo, pero neto 12M positivo",
        }

    if trade_score >= Decimal("3.20"):
        note = (
            f"La accion encadena {positive_streak} {periods_label} superando su referencia ajustada por coeficiente. "
            f"La pendiente relativa es positiva y el retorno neto 12M sigue apoyando compras."
        )
        if valuation_score <= Decimal("-4.00"):
            note += " Aun asi, el PER sigue exigente y frena parte del entusiasmo."
        elif valuation_score >= Decimal("3.00"):
            note += " Ademas, el PER acompana y da apoyo extra a la valoracion."
        if position.is_owned:
            note += " Si ya esta en cartera, la lectura es compatible con mantener o ampliar."
        note += technical_note
        return {
            **base_payload,
            "label": "Comprar",
            "tone": "buy",
            "score": quantize_decimal(trade_score) or ZERO,
            "note": note,
            "trend_label": trend_label,
            "trigger_label": f"{positive_streak} {periods_label} con alpha positiva",
        }

    if trade_score <= Decimal("-3.20"):
        note = (
            f"La accion encadena {negative_streak} {periods_label} perdiendo fuerza frente a su referencia ajustada por coeficiente. "
            f"La pendiente relativa es negativa y la proyeccion neta se deteriora."
        )
        if valuation_score >= Decimal("3.00"):
            note += " El PER ayuda un poco, pero no compensa el deterioro operativo."
        if position.is_owned:
            note += " Conviene revisar venta total o parcial."
        note += technical_note
        return {
            **base_payload,
            "label": "Vender",
            "tone": "sell",
            "score": quantize_decimal(trade_score) or ZERO,
            "note": note,
            "trend_label": trend_label,
            "trigger_label": f"{negative_streak} {periods_label} con pendiente negativa",
        }

    note = base_payload["note"]
    if recent_gap_avg_pct is not None:
        direction = "mejorando" if recent_gap_avg_pct > ZERO else "debilitandose" if recent_gap_avg_pct < ZERO else "sin cambio"
        note = (
            f"La lectura relativa esta {direction}, pero todavia no acumula suficiente persistencia para activar compra o venta."
        )
    note += technical_note
    return {
        **base_payload,
        "score": quantize_decimal(trade_score) or ZERO,
        "note": note,
        "trend_label": trend_label,
    }


def apply_news_context_adjustments_to_card(card: dict) -> dict:
    projection = card.get("projection") or {}
    news_context = card.get("news_context") or {}
    if not projection.get("available") or not news_context.get("material_event"):
        return card
    if (projection.get("news_adjustment") or {}).get("applied"):
        return card

    position = card["position"]
    cycle_projection = card.get("cycle_projection_5y") or {}
    aggregate_score = Decimal(str(news_context.get("score") or "0"))
    items_count = int(news_context.get("items_count") or 0)
    top_tags = [str(tag).strip().lower() for tag in (news_context.get("top_tags") or []) if str(tag).strip()]
    severe_tag_bonus = Decimal("0.12") if {"geopolitica", "regulacion", "resultados"} & set(top_tags) else ZERO
    severity = clamp_decimal(
        (abs(aggregate_score) / Decimal("8.00")) + (Decimal(items_count) * Decimal("0.04")) + severe_tag_bonus,
        Decimal("0.25"),
        Decimal("1.00"),
    )
    confidence_penalty = clamp_decimal(
        Decimal("6.00") + (severity * Decimal("8.00")),
        Decimal("6.00"),
        Decimal("14.00"),
    )
    safety_penalty = clamp_decimal(
        Decimal("3.00") + (severity * Decimal("5.00")),
        Decimal("3.00"),
        Decimal("8.00"),
    )
    reliability_penalty = clamp_decimal(
        Decimal("4.00") + (severity * Decimal("6.00")),
        Decimal("4.00"),
        Decimal("10.00"),
    )
    band_multiplier = clamp_decimal(
        Decimal("1.18") + (severity * Decimal("0.22")),
        Decimal("1.18"),
        Decimal("1.40"),
    )
    spread_multiplier = clamp_decimal(
        Decimal("1.16") + (severity * Decimal("0.20")),
        Decimal("1.16"),
        Decimal("1.36"),
    )
    price_return_adjustment_pct = clamp_decimal(
        aggregate_score * Decimal("0.55"),
        Decimal("-4.50"),
        Decimal("2.50"),
    )
    annual_return_adjustment_pct = clamp_decimal(
        price_return_adjustment_pct * Decimal("0.28"),
        Decimal("-1.75"),
        Decimal("1.00"),
    )
    note = (
        str(news_context.get("material_note") or news_context.get("note") or "").strip()
        or "El contexto web reciente detecta un evento material que puede desordenar temporalmente el patron historico."
    )
    latest_price = position.current_price_per_share
    latest_date = card.get("end_date") or position.latest_price_date
    net_income_yield_pct = projection.get("net_income_yield_pct")
    transaction_drag_pct = projection.get("transaction_drag_pct")
    adjusted_price_return_pct = clamp_decimal(
        (projection.get("price_return_pct") or ZERO) + price_return_adjustment_pct,
        Decimal("-45.00"),
        Decimal("45.00"),
    )
    adjusted_base_return_pct = adjusted_price_return_pct
    if net_income_yield_pct is not None:
        adjusted_base_return_pct += net_income_yield_pct
    if transaction_drag_pct is not None:
        adjusted_base_return_pct -= transaction_drag_pct
    adjusted_base_return_pct = clamp_decimal(adjusted_base_return_pct, Decimal("-50.00"), Decimal("50.00"))
    adjusted_band_pct = clamp_decimal(
        (projection.get("band_pct") or Decimal("16.00")) * band_multiplier,
        Decimal("10.00"),
        Decimal("40.00"),
    )
    adjusted_low_return_pct = clamp_decimal(
        adjusted_base_return_pct - adjusted_band_pct,
        Decimal("-80.00"),
        Decimal("120.00"),
    )
    adjusted_high_return_pct = clamp_decimal(
        adjusted_base_return_pct + adjusted_band_pct,
        Decimal("-80.00"),
        Decimal("140.00"),
    )
    adjusted_price_low_return_pct = clamp_decimal(
        adjusted_price_return_pct - adjusted_band_pct,
        Decimal("-80.00"),
        Decimal("120.00"),
    )
    adjusted_price_high_return_pct = clamp_decimal(
        adjusted_price_return_pct + adjusted_band_pct,
        Decimal("-80.00"),
        Decimal("140.00"),
    )
    adjusted_confidence_score_pct = clamp_decimal(
        (projection.get("confidence_score_pct") or projection_reliability_score(projection.get("confidence_label") or "Baja"))
        - confidence_penalty,
        Decimal("18.00"),
        Decimal("92.00"),
    )
    adjusted_confidence_label = confidence_label_from_score(adjusted_confidence_score_pct)
    adjusted_safety_score = clamp_decimal(
        (projection.get("safety_score") or Decimal("55.00")) - safety_penalty,
        Decimal("18.00"),
        Decimal("92.00"),
    )
    adjusted_safety_label = safety_label_from_score(adjusted_safety_score)
    benefit_risk_ratio = None
    if adjusted_band_pct > ZERO:
        benefit_risk_ratio = adjusted_base_return_pct / adjusted_band_pct
    decision_score = (
        adjusted_base_return_pct
        * projection_confidence_multiplier(adjusted_confidence_label)
        * (adjusted_safety_score / ONE_HUNDRED)
    )
    projection["price_return_pct"] = quantize_decimal(adjusted_price_return_pct)
    projection["price_low_return_pct"] = quantize_decimal(adjusted_price_low_return_pct)
    projection["price_high_return_pct"] = quantize_decimal(adjusted_price_high_return_pct)
    projection["base_return_pct"] = quantize_decimal(adjusted_base_return_pct)
    projection["band_pct"] = quantize_decimal(adjusted_band_pct)
    projection["low_return_pct"] = quantize_decimal(adjusted_low_return_pct)
    projection["high_return_pct"] = quantize_decimal(adjusted_high_return_pct)
    projection["projected_price"] = project_price_from_return(latest_price, adjusted_price_return_pct)
    projection["low_price"] = project_price_from_return(latest_price, adjusted_price_low_return_pct)
    projection["high_price"] = project_price_from_return(latest_price, adjusted_price_high_return_pct)
    projection["monthly_path"] = build_monthly_projection_path(
        latest_price,
        adjusted_price_return_pct,
        anchor_date=latest_date,
        cycle_phase=projection.get("cycle_phase") or "Transicion",
    )
    projection["quarterly_path"] = build_quarterly_projection_path_from_monthly_path(projection["monthly_path"])
    if not projection["quarterly_path"]:
        projection["quarterly_path"] = build_projection_path(
            latest_price,
            adjusted_price_return_pct,
            anchor_date=latest_date,
            cycle_phase=projection.get("cycle_phase") or "Transicion",
        )
    projection["confidence_label"] = adjusted_confidence_label
    projection["confidence_score_pct"] = quantize_decimal(adjusted_confidence_score_pct)
    projection["confidence_note"] = (
        f"{str(projection.get('confidence_note') or '').strip()} "
        "El evento material reciente reduce la confianza y obliga a ampliar el rango del escenario."
    ).strip()
    projection["safety_score"] = quantize_decimal(adjusted_safety_score)
    projection["safety_label"] = adjusted_safety_label
    projection["benefit_risk_ratio"] = quantize_decimal(benefit_risk_ratio)
    projection["decision_score"] = quantize_decimal(decision_score)
    projection["scenarios"] = build_one_year_projection_scenarios(
        latest_price,
        price_return_pct=adjusted_price_return_pct,
        price_low_return_pct=adjusted_price_low_return_pct,
        price_high_return_pct=adjusted_price_high_return_pct,
        base_return_pct=adjusted_base_return_pct,
        low_return_pct=adjusted_low_return_pct,
        high_return_pct=adjusted_high_return_pct,
        confidence_label=adjusted_confidence_label,
        shock_adjusted=True,
        sentiment_score=aggregate_score,
    )
    projection["news_adjustment"] = {
        "applied": True,
        "aggregate_score": quantize_decimal(aggregate_score, "0.01"),
        "severity": quantize_decimal(severity, "0.01"),
        "price_return_adjustment_pct": quantize_decimal(price_return_adjustment_pct),
        "annual_return_adjustment_pct": quantize_decimal(annual_return_adjustment_pct),
        "confidence_penalty_pct": quantize_decimal(confidence_penalty),
        "safety_penalty_pct": quantize_decimal(safety_penalty),
        "band_multiplier": quantize_decimal(band_multiplier, "0.01"),
        "note": note,
    }
    projection["explanation"] = f"{str(projection.get('explanation') or '').strip()} {note}".strip()

    reliability = card.get("projection_reliability") or {"label": "Baja", "score": Decimal("40.00")}
    adjusted_reliability_score = clamp_decimal(
        (reliability.get("score") or projection_reliability_score(reliability.get("label") or "Baja")) - reliability_penalty,
        Decimal("18.00"),
        Decimal("92.00"),
    )
    reliability["score"] = quantize_decimal(adjusted_reliability_score)
    reliability["label"] = confidence_label_from_score(adjusted_reliability_score)
    reliability["news_adjustment"] = {
        "applied": True,
        "reliability_penalty_pct": quantize_decimal(reliability_penalty),
        "note": note,
    }
    card["projection_reliability"] = reliability

    if cycle_projection.get("available"):
        adjusted_annual_return_pct = clamp_decimal(
            (cycle_projection.get("annual_return_pct") or ZERO) + annual_return_adjustment_pct,
            Decimal("-12.00"),
            Decimal("18.00"),
        )
        scenario_spread_annual_pct = clamp_decimal(
            (cycle_projection.get("scenario_spread_annual_pct") or Decimal("3.50")) * spread_multiplier,
            Decimal("2.00"),
            Decimal("9.00"),
        )
        cycle_path, adjusted_steps, news_step_shift = build_cycle_projection_path_for_target(
            latest_price,
            annual_return_pct=adjusted_annual_return_pct,
            step_return_pcts=[Decimal(str(value)) for value in (cycle_projection.get("step_return_pcts") or [])],
            annualized_volatility_pct=cycle_projection.get("annualized_volatility_pct"),
            current_drawdown_pct=cycle_projection.get("current_drawdown_pct"),
            cycle_phase=cycle_projection.get("cycle_phase") or "Transicion",
            anchor_date=cycle_projection.get("latest_date") or latest_date,
            years=5,
            step_months=6,
        )
        if cycle_path:
            projected_price = cycle_path[-1]["projected_price"]
            cycle_projection["annual_return_pct"] = quantize_decimal(adjusted_annual_return_pct)
            cycle_projection["projected_price"] = projected_price
            cycle_projection["five_year_return_pct"] = quantize_decimal(percentage_change(projected_price, latest_price))
            cycle_projection["path"] = cycle_path
            cycle_projection["step_return_pcts"] = [quantize_decimal(value) or ZERO for value in adjusted_steps]
            cycle_projection["news_step_shift"] = quantize_decimal(news_step_shift)
            cycle_projection["scenario_spread_annual_pct"] = quantize_decimal(scenario_spread_annual_pct)
            cycle_projection["scenarios"] = build_five_year_projection_scenarios(
                latest_price,
                latest_date=cycle_projection.get("latest_date") or latest_date,
                annual_return_pct=adjusted_annual_return_pct,
                scenario_spread_annual_pct=scenario_spread_annual_pct,
                step_return_pcts=adjusted_steps,
                annualized_volatility_pct=cycle_projection.get("annualized_volatility_pct"),
                current_drawdown_pct=cycle_projection.get("current_drawdown_pct"),
                cycle_phase=cycle_projection.get("cycle_phase") or "Transicion",
                confidence_label=adjusted_confidence_label,
                shock_adjusted=True,
                sentiment_score=aggregate_score,
            )
            cycle_projection["news_adjustment"] = {
                "applied": True,
                "aggregate_score": quantize_decimal(aggregate_score, "0.01"),
                "annual_return_adjustment_pct": quantize_decimal(annual_return_adjustment_pct),
                "spread_multiplier": quantize_decimal(spread_multiplier, "0.01"),
                "note": note,
            }
            cycle_projection["explanation"] = f"{str(cycle_projection.get('explanation') or '').strip()} {note}".strip()

    synchronize_projection_path_with_cycle_zoom(
        projection,
        cycle_projection,
        current_price=latest_price,
        anchor_date=latest_date,
    )

    one_year_snapshot = next(
        (snapshot for snapshot in (card.get("period_snapshots") or []) if snapshot.get("label") == "1Y"),
        {"available": False},
    )
    six_month_snapshot = card.get("six_month_snapshot") or {"available": False}
    trade_alert = build_trade_alert(
        position,
        projection,
        card.get("correlation") or {},
        reliability,
        card.get("relative_trend") or {},
        six_month_snapshot,
        one_year_snapshot,
        technical_signal=card.get("technical_signal") or {},
    )
    trade_alert["note"] = f"{str(trade_alert.get('note') or '').strip()} {note}".strip()
    card["trade_alert"] = trade_alert
    keep_visuals = bool(
        ((card.get("projection_12m_chart") or {}).get("available"))
        or ((card.get("cycle_projection_5y_chart") or {}).get("available"))
    )
    return refresh_card_projection_visuals(card, include_visuals=keep_visuals)


def apply_news_context_adjustments_to_dashboard(dashboard: dict) -> dict:
    history_cards = list(dashboard.get("history_cards") or [])
    ibex_cards = list(dashboard.get("ibex_universe_cards") or [])
    cards = [*history_cards, *ibex_cards]
    adjusted_cards_count = 0
    material_event_count = 0
    for card in cards:
        if (card.get("news_context") or {}).get("material_event"):
            material_event_count += 1
        was_adjusted = bool(((card.get("projection") or {}).get("news_adjustment") or {}).get("applied"))
        apply_news_context_adjustments_to_card(card)
        refresh_card_projection_visuals(card)
        is_adjusted = bool(((card.get("projection") or {}).get("news_adjustment") or {}).get("applied"))
        if is_adjusted and not was_adjusted:
            adjusted_cards_count += 1

    if history_cards:
        dashboard["decision_rows"] = build_equity_decision_rows(history_cards)
        dashboard["optimizer_cards"] = build_optimizer_master_cards(history_cards, ibex_cards)
        positions = [*(dashboard.get("owned_positions") or []), *(dashboard.get("watchlist_positions") or [])]
        if positions and dashboard.get("overview") is not None:
            dashboard["overview"] = build_equity_analysis_overview(
                positions,
                history_cards,
                dashboard["decision_rows"],
                dashboard.get("ibex_universe_summary") or {},
            )
    if ibex_cards:
        dashboard["ibex_universe_rows"] = build_equity_decision_rows(ibex_cards)
        ibex_summary = dashboard.get("ibex_universe_summary") or {}
        ibex_rows = dashboard["ibex_universe_rows"]
        ibex_summary.update(
            {
                "buy_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Comprar"),
                "sell_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Vender"),
                "watch_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Vigilar"),
                "top_pick": ibex_rows[0] if ibex_rows else ibex_summary.get("top_pick"),
            }
        )
        dashboard["ibex_universe_summary"] = ibex_summary
    dashboard["news_adjustment_summary"] = {
        "adjusted_cards_count": adjusted_cards_count,
        "material_event_count": material_event_count,
    }
    return dashboard["news_adjustment_summary"]


def resolve_expert_consensus_alignment(
    projected_return_pct: Decimal | None,
    consensus_score: Decimal,
) -> tuple[str, str]:
    projected_return_pct = projected_return_pct or ZERO
    if abs(consensus_score) < Decimal("0.85"):
        return "mixed", "Mixto"
    if abs(projected_return_pct) < Decimal("0.50"):
        return "supportive", "De apoyo"
    if consensus_score * projected_return_pct >= ZERO:
        return "supportive", "De apoyo"
    return "contradictory", "En conflicto"


def apply_expert_consensus_adjustments_to_card(card: dict) -> dict:
    projection = card.get("projection") or {}
    expert_consensus = card.get("expert_consensus") or {}
    if not projection.get("available") or not expert_consensus.get("available"):
        return card
    if (projection.get("expert_adjustment") or {}).get("applied"):
        return card

    position = card["position"]
    cycle_projection = card.get("cycle_projection_5y") or {}
    consensus_score = Decimal(str(expert_consensus.get("score") or ZERO))
    quality_score = Decimal(str(expert_consensus.get("quality_score") or Decimal("56.00")))
    items_count = int(expert_consensus.get("items_count") or 0)
    if items_count <= 0:
        return card

    projected_return_pct = projection.get("price_return_pct") or ZERO
    alignment_key, alignment_label = resolve_expert_consensus_alignment(projected_return_pct, consensus_score)
    quality_factor = clamp_decimal(quality_score / ONE_HUNDRED, Decimal("0.42"), Decimal("0.92"))
    conviction = clamp_decimal(
        (abs(consensus_score) / Decimal("8.00")) + (Decimal(min(items_count, 6)) * Decimal("0.05")),
        Decimal("0.18"),
        Decimal("1.00"),
    )
    price_return_adjustment_pct = clamp_decimal(
        consensus_score * (Decimal("0.12") + (quality_factor * Decimal("0.22"))),
        Decimal("-3.20"),
        Decimal("3.20"),
    )
    annual_return_adjustment_pct = clamp_decimal(
        price_return_adjustment_pct * Decimal("0.24"),
        Decimal("-1.20"),
        Decimal("1.20"),
    )
    if alignment_key == "supportive":
        reliability_shift = clamp_decimal(conviction * quality_factor * Decimal("4.80"), Decimal("0.80"), Decimal("4.80"))
        confidence_shift = clamp_decimal(conviction * quality_factor * Decimal("3.20"), Decimal("0.50"), Decimal("3.20"))
        safety_shift = clamp_decimal(conviction * quality_factor * Decimal("2.20"), Decimal("0.30"), Decimal("2.20"))
        band_multiplier = clamp_decimal(
            Decimal("1.00") - (conviction * quality_factor * Decimal("0.09")),
            Decimal("0.90"),
            Decimal("0.99"),
        )
        spread_multiplier = clamp_decimal(
            Decimal("1.00") - (conviction * quality_factor * Decimal("0.08")),
            Decimal("0.91"),
            Decimal("0.99"),
        )
    elif alignment_key == "contradictory":
        reliability_shift = -clamp_decimal(conviction * quality_factor * Decimal("5.60"), Decimal("1.00"), Decimal("5.60"))
        confidence_shift = -clamp_decimal(conviction * quality_factor * Decimal("4.00"), Decimal("0.70"), Decimal("4.00"))
        safety_shift = -clamp_decimal(conviction * quality_factor * Decimal("2.60"), Decimal("0.40"), Decimal("2.60"))
        band_multiplier = clamp_decimal(
            Decimal("1.00") + (conviction * quality_factor * Decimal("0.13")),
            Decimal("1.02"),
            Decimal("1.15"),
        )
        spread_multiplier = clamp_decimal(
            Decimal("1.00") + (conviction * quality_factor * Decimal("0.14")),
            Decimal("1.03"),
            Decimal("1.18"),
        )
    else:
        reliability_shift = -clamp_decimal(conviction * quality_factor * Decimal("2.20"), ZERO, Decimal("2.20"))
        confidence_shift = -clamp_decimal(conviction * quality_factor * Decimal("1.40"), ZERO, Decimal("1.40"))
        safety_shift = ZERO
        band_multiplier = clamp_decimal(
            Decimal("1.00") + (conviction * quality_factor * Decimal("0.05")),
            Decimal("0.98"),
            Decimal("1.08"),
        )
        spread_multiplier = clamp_decimal(
            Decimal("1.00") + (conviction * quality_factor * Decimal("0.06")),
            Decimal("0.98"),
            Decimal("1.10"),
        )

    best_sources = [str(source).strip() for source in (expert_consensus.get("best_sources") or []) if str(source).strip()]
    note = (
        str(expert_consensus.get("note") or "").strip()
        or "El consenso de expertos se integra como una senal adicional ponderada por su acierto historico."
    )
    if best_sources:
        note = f"{note} Fuentes mejor rankeadas: {', '.join(best_sources[:3])}."

    latest_price = position.current_price_per_share
    latest_date = card.get("end_date") or position.latest_price_date
    net_income_yield_pct = projection.get("net_income_yield_pct")
    transaction_drag_pct = projection.get("transaction_drag_pct")
    adjusted_price_return_pct = clamp_decimal(
        (projection.get("price_return_pct") or ZERO) + price_return_adjustment_pct,
        Decimal("-45.00"),
        Decimal("45.00"),
    )
    adjusted_base_return_pct = adjusted_price_return_pct
    if net_income_yield_pct is not None:
        adjusted_base_return_pct += net_income_yield_pct
    if transaction_drag_pct is not None:
        adjusted_base_return_pct -= transaction_drag_pct
    adjusted_base_return_pct = clamp_decimal(adjusted_base_return_pct, Decimal("-50.00"), Decimal("50.00"))
    adjusted_band_pct = clamp_decimal(
        (projection.get("band_pct") or Decimal("16.00")) * band_multiplier,
        Decimal("9.00"),
        Decimal("40.00"),
    )
    adjusted_low_return_pct = clamp_decimal(
        adjusted_base_return_pct - adjusted_band_pct,
        Decimal("-80.00"),
        Decimal("120.00"),
    )
    adjusted_high_return_pct = clamp_decimal(
        adjusted_base_return_pct + adjusted_band_pct,
        Decimal("-80.00"),
        Decimal("140.00"),
    )
    adjusted_price_low_return_pct = clamp_decimal(
        adjusted_price_return_pct - adjusted_band_pct,
        Decimal("-80.00"),
        Decimal("120.00"),
    )
    adjusted_price_high_return_pct = clamp_decimal(
        adjusted_price_return_pct + adjusted_band_pct,
        Decimal("-80.00"),
        Decimal("140.00"),
    )
    adjusted_confidence_score_pct = clamp_decimal(
        (projection.get("confidence_score_pct") or projection_reliability_score(projection.get("confidence_label") or "Baja"))
        + confidence_shift,
        Decimal("18.00"),
        Decimal("92.00"),
    )
    adjusted_confidence_label = confidence_label_from_score(adjusted_confidence_score_pct)
    adjusted_safety_score = clamp_decimal(
        (projection.get("safety_score") or Decimal("55.00")) + safety_shift,
        Decimal("18.00"),
        Decimal("92.00"),
    )
    adjusted_safety_label = safety_label_from_score(adjusted_safety_score)
    benefit_risk_ratio = None
    if adjusted_band_pct > ZERO:
        benefit_risk_ratio = adjusted_base_return_pct / adjusted_band_pct
    decision_score = (
        adjusted_base_return_pct
        * projection_confidence_multiplier(adjusted_confidence_label)
        * (adjusted_safety_score / ONE_HUNDRED)
    )
    projection["price_return_pct"] = quantize_decimal(adjusted_price_return_pct)
    projection["price_low_return_pct"] = quantize_decimal(adjusted_price_low_return_pct)
    projection["price_high_return_pct"] = quantize_decimal(adjusted_price_high_return_pct)
    projection["base_return_pct"] = quantize_decimal(adjusted_base_return_pct)
    projection["band_pct"] = quantize_decimal(adjusted_band_pct)
    projection["low_return_pct"] = quantize_decimal(adjusted_low_return_pct)
    projection["high_return_pct"] = quantize_decimal(adjusted_high_return_pct)
    projection["projected_price"] = project_price_from_return(latest_price, adjusted_price_return_pct)
    projection["low_price"] = project_price_from_return(latest_price, adjusted_price_low_return_pct)
    projection["high_price"] = project_price_from_return(latest_price, adjusted_price_high_return_pct)
    projection["monthly_path"] = build_monthly_projection_path(
        latest_price,
        adjusted_price_return_pct,
        anchor_date=latest_date,
        cycle_phase=projection.get("cycle_phase") or "Transicion",
    )
    projection["quarterly_path"] = build_quarterly_projection_path_from_monthly_path(projection["monthly_path"])
    if not projection["quarterly_path"]:
        projection["quarterly_path"] = build_projection_path(
            latest_price,
            adjusted_price_return_pct,
            anchor_date=latest_date,
            cycle_phase=projection.get("cycle_phase") or "Transicion",
        )
    projection["confidence_label"] = adjusted_confidence_label
    projection["confidence_score_pct"] = quantize_decimal(adjusted_confidence_score_pct)
    projection["confidence_note"] = (
        f"{str(projection.get('confidence_note') or '').strip()} "
        "La capa de consenso experto modula la confianza segun su track record reciente."
    ).strip()
    projection["safety_score"] = quantize_decimal(adjusted_safety_score)
    projection["safety_label"] = adjusted_safety_label
    projection["benefit_risk_ratio"] = quantize_decimal(benefit_risk_ratio)
    projection["decision_score"] = quantize_decimal(decision_score)
    projection["scenarios"] = build_one_year_projection_scenarios(
        latest_price,
        price_return_pct=adjusted_price_return_pct,
        price_low_return_pct=adjusted_price_low_return_pct,
        price_high_return_pct=adjusted_price_high_return_pct,
        base_return_pct=adjusted_base_return_pct,
        low_return_pct=adjusted_low_return_pct,
        high_return_pct=adjusted_high_return_pct,
        confidence_label=adjusted_confidence_label,
    )
    projection["expert_adjustment"] = {
        "applied": True,
        "aggregate_score": quantize_decimal(consensus_score, "0.01"),
        "quality_score": quantize_decimal(quality_score),
        "alignment_label": alignment_label,
        "price_return_adjustment_pct": quantize_decimal(price_return_adjustment_pct),
        "annual_return_adjustment_pct": quantize_decimal(annual_return_adjustment_pct),
        "band_multiplier": quantize_decimal(band_multiplier, "0.01"),
        "note": note,
    }
    projection["explanation"] = f"{str(projection.get('explanation') or '').strip()} {note}".strip()

    reliability = card.get("projection_reliability") or {"label": "Baja", "score": Decimal("40.00")}
    adjusted_reliability_score = clamp_decimal(
        (reliability.get("score") or projection_reliability_score(reliability.get("label") or "Baja")) + reliability_shift,
        Decimal("18.00"),
        Decimal("92.00"),
    )
    reliability["score"] = quantize_decimal(adjusted_reliability_score)
    reliability["label"] = confidence_label_from_score(adjusted_reliability_score)
    reliability["expert_adjustment"] = {
        "applied": True,
        "quality_score": quantize_decimal(quality_score),
        "alignment_label": alignment_label,
        "reliability_shift_pct": quantize_decimal(reliability_shift),
        "note": note,
    }
    card["projection_reliability"] = reliability

    if cycle_projection.get("available"):
        adjusted_annual_return_pct = clamp_decimal(
            (cycle_projection.get("annual_return_pct") or ZERO) + annual_return_adjustment_pct,
            Decimal("-12.00"),
            Decimal("18.00"),
        )
        scenario_spread_annual_pct = clamp_decimal(
            (cycle_projection.get("scenario_spread_annual_pct") or Decimal("3.50")) * spread_multiplier,
            Decimal("2.00"),
            Decimal("9.00"),
        )
        cycle_path, adjusted_steps, step_shift = build_cycle_projection_path_for_target(
            latest_price,
            annual_return_pct=adjusted_annual_return_pct,
            step_return_pcts=[Decimal(str(value)) for value in (cycle_projection.get("step_return_pcts") or [])],
            annualized_volatility_pct=cycle_projection.get("annualized_volatility_pct"),
            current_drawdown_pct=cycle_projection.get("current_drawdown_pct"),
            cycle_phase=cycle_projection.get("cycle_phase") or "Transicion",
            anchor_date=cycle_projection.get("latest_date") or latest_date,
            years=5,
            step_months=6,
        )
        if cycle_path:
            projected_price = cycle_path[-1]["projected_price"]
            cycle_projection["annual_return_pct"] = quantize_decimal(adjusted_annual_return_pct)
            cycle_projection["projected_price"] = projected_price
            cycle_projection["five_year_return_pct"] = quantize_decimal(percentage_change(projected_price, latest_price))
            cycle_projection["path"] = cycle_path
            cycle_projection["step_return_pcts"] = [quantize_decimal(value) or ZERO for value in adjusted_steps]
            cycle_projection["expert_step_shift"] = quantize_decimal(step_shift)
            cycle_projection["scenario_spread_annual_pct"] = quantize_decimal(scenario_spread_annual_pct)
            cycle_projection["scenarios"] = build_five_year_projection_scenarios(
                latest_price,
                latest_date=cycle_projection.get("latest_date") or latest_date,
                annual_return_pct=adjusted_annual_return_pct,
                scenario_spread_annual_pct=scenario_spread_annual_pct,
                step_return_pcts=adjusted_steps,
                annualized_volatility_pct=cycle_projection.get("annualized_volatility_pct"),
                current_drawdown_pct=cycle_projection.get("current_drawdown_pct"),
                cycle_phase=cycle_projection.get("cycle_phase") or "Transicion",
                confidence_label=adjusted_confidence_label,
            )
            cycle_projection["expert_adjustment"] = {
                "applied": True,
                "aggregate_score": quantize_decimal(consensus_score, "0.01"),
                "quality_score": quantize_decimal(quality_score),
                "alignment_label": alignment_label,
                "annual_return_adjustment_pct": quantize_decimal(annual_return_adjustment_pct),
                "spread_multiplier": quantize_decimal(spread_multiplier, "0.01"),
                "note": note,
            }
            cycle_projection["explanation"] = f"{str(cycle_projection.get('explanation') or '').strip()} {note}".strip()

    synchronize_projection_path_with_cycle_zoom(
        projection,
        cycle_projection,
        current_price=latest_price,
        anchor_date=latest_date,
    )

    one_year_snapshot = next(
        (snapshot for snapshot in (card.get("period_snapshots") or []) if snapshot.get("label") == "1Y"),
        {"available": False},
    )
    six_month_snapshot = card.get("six_month_snapshot") or {"available": False}
    trade_alert = build_trade_alert(
        position,
        projection,
        card.get("correlation") or {},
        reliability,
        card.get("relative_trend") or {},
        six_month_snapshot,
        one_year_snapshot,
        technical_signal=card.get("technical_signal") or {},
    )
    trade_alert["note"] = f"{str(trade_alert.get('note') or '').strip()} {note}".strip()
    card["trade_alert"] = trade_alert
    keep_visuals = bool(
        ((card.get("projection_12m_chart") or {}).get("available"))
        or ((card.get("cycle_projection_5y_chart") or {}).get("available"))
    )
    return refresh_card_projection_visuals(card, include_visuals=keep_visuals)


def apply_expert_consensus_adjustments_to_dashboard(dashboard: dict) -> dict:
    history_cards = list(dashboard.get("history_cards") or [])
    ibex_cards = list(dashboard.get("ibex_universe_cards") or [])
    cards = [*history_cards, *ibex_cards]
    adjusted_cards_count = 0
    strong_consensus_count = 0
    for card in cards:
        expert_consensus = card.get("expert_consensus") or {}
        if abs(Decimal(str(expert_consensus.get("score") or ZERO))) >= Decimal("2.50"):
            strong_consensus_count += 1
        was_adjusted = bool(((card.get("projection") or {}).get("expert_adjustment") or {}).get("applied"))
        apply_expert_consensus_adjustments_to_card(card)
        refresh_card_projection_visuals(card)
        is_adjusted = bool(((card.get("projection") or {}).get("expert_adjustment") or {}).get("applied"))
        if is_adjusted and not was_adjusted:
            adjusted_cards_count += 1

    if history_cards:
        dashboard["decision_rows"] = build_equity_decision_rows(history_cards)
        dashboard["optimizer_cards"] = build_optimizer_master_cards(history_cards, ibex_cards)
        positions = [*(dashboard.get("owned_positions") or []), *(dashboard.get("watchlist_positions") or [])]
        if positions and dashboard.get("overview") is not None:
            dashboard["overview"] = build_equity_analysis_overview(
                positions,
                history_cards,
                dashboard["decision_rows"],
                dashboard.get("ibex_universe_summary") or {},
            )
    if ibex_cards:
        dashboard["ibex_universe_rows"] = build_equity_decision_rows(ibex_cards)
        ibex_summary = dashboard.get("ibex_universe_summary") or {}
        ibex_rows = dashboard["ibex_universe_rows"]
        ibex_summary.update(
            {
                "buy_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Comprar"),
                "sell_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Vender"),
                "watch_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Vigilar"),
                "top_pick": ibex_rows[0] if ibex_rows else ibex_summary.get("top_pick"),
            }
        )
        dashboard["ibex_universe_summary"] = ibex_summary
    dashboard["expert_adjustment_summary"] = {
        "adjusted_cards_count": adjusted_cards_count,
        "strong_consensus_count": strong_consensus_count,
    }
    return dashboard["expert_adjustment_summary"]


def build_decision_action_label(
    position: EquityPosition,
    projected_return_pct: Decimal | None,
    safety_score: Decimal | None,
    reliability_score: Decimal | None,
    valuation_score: Decimal | None = None,
) -> str:
    projected_return_pct = projected_return_pct or ZERO
    safety_score = safety_score or ZERO
    reliability_score = reliability_score or ZERO
    valuation_score = valuation_score or ZERO
    if projected_return_pct >= Decimal("10.00") and safety_score >= Decimal("65.00") and reliability_score >= Decimal("70.00"):
        if valuation_score <= Decimal("-6.00"):
            return "Mantener" if position.is_owned else "Seguir"
        return "Priorizar"
    if projected_return_pct >= Decimal("4.00") and safety_score >= Decimal("55.00"):
        if valuation_score <= Decimal("-6.00") and projected_return_pct < Decimal("8.00"):
            return "Mantener" if position.is_owned else "Vigilar"
        return "Mantener" if position.is_owned else "Seguir"
    if projected_return_pct >= ZERO:
        if valuation_score >= Decimal("5.00") and safety_score >= Decimal("52.00") and reliability_score >= Decimal("58.00"):
            return "Mantener" if position.is_owned else "Seguir"
        return "Vigilar"
    return "Reducir riesgo" if position.is_owned else "Esperar"


def build_cycle_projection_yearly_margins(
    current_price: Decimal | None,
    cycle_projection: dict | None,
    *,
    first_year_projected_price: Decimal | None = None,
    first_year_return_pct: Decimal | None = None,
    max_years: int = 5,
) -> list[dict]:
    if current_price in {None, ZERO} or not cycle_projection or not cycle_projection.get("available"):
        return []

    path = cycle_projection.get("path") or []
    checkpoints_by_label = {
        str(step.get("label")): step
        for step in path
        if step.get("label") and step.get("projected_price") is not None
    }
    margins = []
    previous_price = current_price

    for year_number in range(1, max_years + 1):
        step = checkpoints_by_label.get(f"{year_number}A")
        if year_number == 1 and first_year_projected_price is not None:
            projected_price = first_year_projected_price
        elif step:
            projected_price = step.get("projected_price")
        else:
            projected_price = None
        if projected_price is None:
            continue
        if year_number == 1 and first_year_return_pct is not None:
            margin_pct = first_year_return_pct
        else:
            margin_pct = percentage_change(projected_price, previous_price)
        cumulative_return_pct = percentage_change(projected_price, current_price)
        margins.append(
            {
                "year_number": year_number,
                "label": f"AÑO {year_number}",
                "projected_price": projected_price,
                "margin_pct": quantize_decimal(margin_pct),
                "cumulative_return_pct": quantize_decimal(cumulative_return_pct),
            }
        )
        previous_price = projected_price

    return margins


def build_cycle_projection_return_profile(
    current_price: Decimal | None,
    cycle_projection: dict | None,
    *,
    first_year_projected_price: Decimal | None = None,
    first_year_return_pct: Decimal | None = None,
) -> dict:
    yearly_margins = build_cycle_projection_yearly_margins(
        current_price,
        cycle_projection,
        first_year_projected_price=first_year_projected_price,
        first_year_return_pct=first_year_return_pct,
    )
    five_year_return_pct = None
    five_year_annualized_return_pct = None
    if cycle_projection and cycle_projection.get("available"):
        five_year_return_pct = cycle_projection.get("five_year_return_pct")
        if five_year_return_pct is None:
            five_year_row = next((item for item in yearly_margins if item["year_number"] == 5), None)
            if five_year_row:
                five_year_return_pct = five_year_row.get("cumulative_return_pct")
        if five_year_return_pct is not None:
            five_year_annualized_return_pct = annualize_return_pct(five_year_return_pct, 60)
        if five_year_annualized_return_pct is None:
            five_year_annualized_return_pct = cycle_projection.get("annual_return_pct")

    return {
        "five_year_return_pct": quantize_decimal(five_year_return_pct),
        "five_year_annualized_return_pct": quantize_decimal(five_year_annualized_return_pct),
        "yearly_margins": yearly_margins,
    }


def build_equity_decision_rows(history_cards: list[dict]) -> list[dict]:
    rows = []
    status_order = {"owned": 0, "watchlist": 1, "ibex": 2, "guide": 3}
    for card in history_cards:
        position = card["position"]
        projection = card.get("projection", {})
        if not card.get("has_history") or not projection.get("available"):
            continue
        reliability = card.get("projection_reliability", {"label": "Baja", "score": Decimal("40.00")})
        best_candidate = card.get("reference_playbook", {}).get("best_candidate")
        effective_projection = resolve_effective_projection_metrics(card)
        projected_return_pct = effective_projection.get("base_return_pct")
        coefficient_alert = card.get("coefficient_alert", {})
        valuation = card.get("valuation") or {}
        cycle_projection = card.get("cycle_projection_5y") or {}
        scenario_tables = card.get("scenario_tables") or build_card_scenario_tables(card)
        trade_alert = reconcile_trade_alert_with_expected_return(
            card.get("trade_alert", {}),
            expected_return_pct=(scenario_tables.get("projection_12m") or {}).get("expected_return_pct"),
            horizon_label="1A",
        )
        cycle_horizon_expectations = build_cycle_horizon_expectation_map(
            cycle_projection.get("scenarios"),
            current_price=position.current_price_per_share,
        )
        cycle_return_profile = build_cycle_projection_return_profile(
            position.current_price_per_share,
            cycle_projection,
            first_year_projected_price=effective_projection.get("projected_price"),
            first_year_return_pct=projected_return_pct,
        )
        rows.append(
            {
                "position": position,
                "status_key": card.get("status_key") or ("owned" if position.is_owned else "watchlist"),
                "status_label": card.get("status_label") or position.get_position_kind_display(),
                "status_note": card.get("status_note", ""),
                "company_name": position.company_name,
                "ticker": position.ticker,
                "sector_label": card.get("sector_label") or "Sin sector",
                "detail_anchor": card.get("detail_anchor") or "",
                "reference_label": card.get("reference_label"),
                "best_reference_label": best_candidate.get("name") if best_candidate else card.get("reference_label"),
                "correlation": card.get("correlation", {}).get("coefficient"),
                "years_covered": projection.get("years_covered"),
                "projected_return_pct": projected_return_pct,
                "projected_price": effective_projection.get("projected_price"),
                "expected_return_1y_pct": (scenario_tables.get("projection_12m") or {}).get("expected_return_pct"),
                "expected_return_2y_pct": cycle_horizon_expectations.get(24),
                "expected_return_3y_pct": cycle_horizon_expectations.get(36),
                "expected_return_4y_pct": cycle_horizon_expectations.get(48),
                "expected_return_5y_pct": cycle_horizon_expectations.get(60)
                or (scenario_tables.get("cycle_5y") or {}).get("expected_return_pct"),
                "cycle_return_5y_pct": cycle_return_profile["five_year_return_pct"],
                "cycle_return_annual_pct": cycle_return_profile["five_year_annualized_return_pct"],
                "cycle_yearly_margins": cycle_return_profile["yearly_margins"],
                "safety_score": projection.get("safety_score"),
                "safety_label": projection.get("safety_label"),
                "reliability_score": reliability.get("score"),
                "reliability_label": reliability.get("label"),
                "benefit_risk_ratio": projection.get("benefit_risk_ratio"),
                "cycle_phase": projection.get("cycle_phase"),
                "decision_score": projection.get("decision_score"),
                "valuation_label": valuation.get("label") or "Sin PER",
                "valuation_tone": valuation.get("tone") or "neutral",
                "valuation_note": valuation.get("note") or "",
                "valuation_score": valuation.get("score"),
                "per_value": valuation.get("per_value"),
                "per_source_label": valuation.get("source_short_label") or valuation.get("source_label") or "",
                "trade_alert_label": trade_alert.get("label", "Vigilar"),
                "trade_alert_tone": trade_alert.get("tone", "watch"),
                "trade_alert_note": trade_alert.get("note", ""),
                "trade_alert_score": trade_alert.get("score", ZERO),
                "trade_alert_trigger": trade_alert.get("trigger_label", ""),
                "coefficient_alert_label": coefficient_alert.get("label", "Sin lectura"),
                "coefficient_alert_tone": coefficient_alert.get("tone", "neutral"),
                "coefficient_alert_trigger": coefficient_alert.get("trigger_label", ""),
                "action_label": build_decision_action_label(
                    position,
                    projected_return_pct,
                    projection.get("safety_score"),
                    reliability.get("score"),
                    valuation.get("score"),
                ),
            }
        )

    rows.sort(
        key=lambda item: (
            status_order.get(item["status_key"], 9),
            -(item["decision_score"] if item["decision_score"] is not None else Decimal("-9999")),
            -(item["trade_alert_score"] if item["trade_alert_score"] is not None else Decimal("-9999")),
            item["company_name"],
        )
    )
    return rows


def build_candlestick_svg(history, width: int = 640, height: int = 220, padding: int = 18) -> str:
    candles = history[-32:]
    if len(candles) < 2:
        return ""

    highs = [point.high_price or point.close_price for point in candles]
    lows = [point.low_price or point.close_price for point in candles]
    min_price = min(lows)
    max_price = max(highs)
    if max_price == min_price:
        max_price += Decimal("1")

    span_x = width - 2 * padding
    span_y = height - 2 * padding
    step = span_x / max(len(candles) - 1, 1)
    body_width = max(step * 0.55, 6)

    def to_y(value: Decimal) -> float:
        normalized = float((value - min_price) / (max_price - min_price))
        return float(height - padding - (normalized * span_y))

    fragments = [
        f'<line x1="{padding}" y1="{height - padding}" x2="{width - padding}" y2="{height - padding}" stroke="#d9e4ec" stroke-width="1" />',
        f'<line x1="{padding}" y1="{padding}" x2="{padding}" y2="{height - padding}" stroke="#d9e4ec" stroke-width="1" />',
    ]

    for index, point in enumerate(candles):
        open_price = point.open_price or point.close_price
        close_price = point.close_price
        high_price = point.high_price or max(open_price, close_price)
        low_price = point.low_price or min(open_price, close_price)
        center_x = padding + (step * index)
        wick_top = to_y(high_price)
        wick_bottom = to_y(low_price)
        body_top_price = max(open_price, close_price)
        body_bottom_price = min(open_price, close_price)
        body_top = to_y(body_top_price)
        body_bottom = to_y(body_bottom_price)
        body_height = max(body_bottom - body_top, 2.0)
        tone = "#177245" if close_price >= open_price else "#a85f00"
        body_y = min(body_top, body_bottom)
        fragments.append(
            f'<line x1="{center_x:.1f}" y1="{wick_top:.1f}" x2="{center_x:.1f}" y2="{wick_bottom:.1f}" stroke="{tone}" stroke-width="1.5" />'
        )
        fragments.append(
            f'<rect x="{center_x - (body_width / 2):.1f}" y="{body_y:.1f}" width="{body_width:.1f}" height="{body_height:.1f}" rx="2" fill="{tone}" opacity="0.88" />'
        )

    return "".join(fragments)


def normalize_candle_prices(point) -> dict:
    close_price = point.close_price or ZERO
    open_price = point.open_price or close_price
    high_price = point.high_price or max(open_price, close_price)
    low_price = point.low_price or min(open_price, close_price)
    return {
        "open_price": open_price,
        "high_price": high_price,
        "low_price": low_price,
        "close_price": close_price,
    }


def calculate_rsi(closes: list[Decimal], period: int = 14) -> Decimal | None:
    filtered = [value for value in closes if value is not None]
    if len(filtered) <= period:
        return None

    gains = []
    losses = []
    window = filtered[-(period + 1):]
    for previous_close, current_close in zip(window, window[1:]):
        delta = current_close - previous_close
        gains.append(max(delta, ZERO))
        losses.append(max(-delta, ZERO))

    average_gain = average_decimal(gains) or ZERO
    average_loss = average_decimal(losses) or ZERO
    if average_loss == ZERO:
        if average_gain == ZERO:
            return Decimal("50.00")
        return Decimal("100.00")
    relative_strength = average_gain / average_loss
    rsi = Decimal("100.00") - (Decimal("100.00") / (Decimal("1.00") + relative_strength))
    return quantize_decimal(rsi)


def resolve_projection_confidence_label(score_pct: Decimal | None) -> str:
    score_pct = score_pct or Decimal("40.00")
    if score_pct >= Decimal("75.00"):
        return "Alta"
    if score_pct >= Decimal("58.00"):
        return "Media"
    return "Baja"


def detect_candlestick_pattern(
    previous_candle: dict | None,
    latest_candle: dict,
    *,
    support_level: Decimal | None = None,
    resistance_level: Decimal | None = None,
) -> dict:
    open_price = latest_candle["open_price"]
    close_price = latest_candle["close_price"]
    high_price = latest_candle["high_price"]
    low_price = latest_candle["low_price"]
    body_size = abs(close_price - open_price)
    candle_range = high_price - low_price
    upper_shadow = high_price - max(open_price, close_price)
    lower_shadow = min(open_price, close_price) - low_price

    near_support = False
    near_resistance = False
    if support_level not in {None, ZERO}:
        near_support = abs(low_price - support_level) / support_level <= Decimal("0.025")
    if resistance_level not in {None, ZERO}:
        near_resistance = abs(high_price - resistance_level) / resistance_level <= Decimal("0.025")

    if previous_candle is not None:
        previous_open = previous_candle["open_price"]
        previous_close = previous_candle["close_price"]
        if (
            previous_close < previous_open
            and close_price > open_price
            and open_price <= previous_close
            and close_price >= previous_open
        ):
            return {
                "label": "Envolvente alcista",
                "score": Decimal("1.70"),
                "note": "La ultima vela absorbe la caida previa y deja una senal de compra tactica.",
            }
        if (
            previous_close > previous_open
            and close_price < open_price
            and open_price >= previous_close
            and close_price <= previous_open
        ):
            return {
                "label": "Envolvente bajista",
                "score": Decimal("-1.70"),
                "note": "La ultima vela absorbe el avance previo y deja una senal de venta tactica.",
            }

    if candle_range > ZERO and body_size <= candle_range * Decimal("0.15"):
        return {
            "label": "Doji",
            "score": Decimal("0.20") if near_support else Decimal("-0.20") if near_resistance else ZERO,
            "note": "La ultima vela cierra muy equilibrada y pide confirmacion adicional.",
        }
    if candle_range > ZERO and lower_shadow >= body_size * Decimal("2.20") and upper_shadow <= max(body_size, candle_range * Decimal("0.18")):
        return {
            "label": "Martillo",
            "score": Decimal("1.20") if close_price >= open_price or near_support else Decimal("0.70"),
            "note": "La sombra inferior larga sugiere rechazo de precios bajos y posible giro comprador.",
        }
    if candle_range > ZERO and upper_shadow >= body_size * Decimal("2.20") and lower_shadow <= max(body_size, candle_range * Decimal("0.18")):
        return {
            "label": "Estrella fugaz",
            "score": Decimal("-1.20") if close_price <= open_price or near_resistance else Decimal("-0.70"),
            "note": "La sombra superior larga sugiere rechazo de precios altos y posible giro vendedor.",
        }
    if close_price >= open_price:
        return {
            "label": "Vela alcista",
            "score": Decimal("0.30"),
            "note": "La ultima vela acompana al alza, aunque sin patron fuerte de giro.",
        }
    return {
        "label": "Vela bajista",
        "score": Decimal("-0.30"),
        "note": "La ultima vela acompana a la baja, aunque sin patron fuerte de giro.",
    }


def build_candlestick_metrics(history) -> dict:
    recent = history[-220:]
    if not recent:
        return {
            "available": False,
            "trend_label": "Sin historico",
            "last_candle_label": "Sin vela",
            "average_range_pct": None,
            "support_level": None,
            "resistance_level": None,
            "momentum_label": "Sin momento",
            "pattern_label": "Sin patron",
            "breakout_label": "Sin ruptura",
            "volatility_label": "Sin lectura",
            "signal_label": "Sin senal",
            "signal_score": None,
            "confidence_score": None,
            "confidence_label": "Baja",
            "return_adjustment_pct": ZERO,
            "band_multiplier": Decimal("1.00"),
            "rsi_14": None,
            "rsi_label": "Sin RSI",
            "distance_to_support_pct": None,
            "distance_to_resistance_pct": None,
            "note": "",
            "candlestick_svg": "",
        }

    closes = [point.close_price for point in recent]
    sma20 = average_decimal(closes[-20:]) or recent[-1].close_price
    sma50 = average_decimal(closes[-50:]) or sma20
    sma200 = average_decimal(closes[-200:]) if len(closes) >= 120 else None
    latest_close = recent[-1].close_price
    trend_score = ZERO
    if latest_close > sma20:
        trend_score += Decimal("0.70")
    elif latest_close < sma20:
        trend_score -= Decimal("0.70")
    if sma20 > sma50:
        trend_score += Decimal("1.25")
    elif sma20 < sma50:
        trend_score -= Decimal("1.25")
    if sma200 is not None:
        if sma50 > sma200:
            trend_score += Decimal("1.55")
        elif sma50 < sma200:
            trend_score -= Decimal("1.55")
        if latest_close > sma200:
            trend_score += Decimal("0.85")
        elif latest_close < sma200:
            trend_score -= Decimal("0.85")

    if trend_score >= Decimal("3.00"):
        trend_label = "Tendencia alcista confirmada"
    elif trend_score >= Decimal("1.20"):
        trend_label = "Sesgo alcista"
    elif trend_score <= Decimal("-3.00"):
        trend_label = "Tendencia bajista confirmada"
    elif trend_score <= Decimal("-1.20"):
        trend_label = "Sesgo bajista"
    else:
        trend_label = "Tendencia lateral"

    last_point = recent[-1]
    latest_candle = normalize_candle_prices(last_point)
    previous_candle = normalize_candle_prices(recent[-2]) if len(recent) >= 2 else None
    last_candle_label = "Vela alcista" if latest_candle["close_price"] >= latest_candle["open_price"] else "Vela bajista"
    if latest_candle["high_price"] > latest_candle["low_price"]:
        body_size = abs(latest_candle["close_price"] - latest_candle["open_price"])
        candle_range = latest_candle["high_price"] - latest_candle["low_price"]
        if body_size <= candle_range * Decimal("0.15"):
            last_candle_label = "Doji"

    range_percentages = []
    for point in recent[-20:]:
        base_open = point.open_price or point.close_price
        high_price = point.high_price or point.close_price
        low_price = point.low_price or point.close_price
        if base_open:
            range_percentages.append(((high_price - low_price) / base_open) * Decimal("100"))

    support_level = min((point.low_price or point.close_price for point in recent[-20:]), default=None)
    resistance_level = max((point.high_price or point.close_price for point in recent[-20:]), default=None)
    previous_window = recent[-21:-1] if len(recent) >= 21 else recent[:-1]
    previous_support = min((point.low_price or point.close_price for point in previous_window), default=support_level)
    previous_resistance = max((point.high_price or point.close_price for point in previous_window), default=resistance_level)
    pattern = detect_candlestick_pattern(
        previous_candle,
        latest_candle,
        support_level=support_level,
        resistance_level=resistance_level,
    )

    short_return_pct = percentage_change(latest_close, closes[-6]) if len(closes) >= 6 else None
    medium_return_pct = percentage_change(latest_close, closes[-21]) if len(closes) >= 21 else None
    rsi_14 = calculate_rsi(closes, period=14)
    momentum_score = ZERO
    if short_return_pct is not None:
        momentum_score += clamp_decimal(short_return_pct * Decimal("0.12"), Decimal("-1.10"), Decimal("1.10"))
    if medium_return_pct is not None:
        momentum_score += clamp_decimal(medium_return_pct * Decimal("0.07"), Decimal("-1.40"), Decimal("1.40"))
    if rsi_14 is not None:
        if rsi_14 >= Decimal("65.00"):
            momentum_score += Decimal("0.80")
            rsi_label = "Impulso alcista"
        elif rsi_14 >= Decimal("55.00"):
            momentum_score += Decimal("0.40")
            rsi_label = "Momentum positivo"
        elif rsi_14 <= Decimal("35.00"):
            momentum_score -= Decimal("0.80")
            rsi_label = "Impulso bajista"
        elif rsi_14 <= Decimal("45.00"):
            momentum_score -= Decimal("0.40")
            rsi_label = "Momentum flojo"
        else:
            rsi_label = "Neutral"
    else:
        rsi_label = "Sin RSI"

    if momentum_score >= Decimal("1.40"):
        momentum_label = "Impulso comprador"
    elif momentum_score >= Decimal("0.35"):
        momentum_label = "Momento favorable"
    elif momentum_score <= Decimal("-1.40"):
        momentum_label = "Impulso vendedor"
    elif momentum_score <= Decimal("-0.35"):
        momentum_label = "Momento flojo"
    else:
        momentum_label = "Momento neutro"

    distance_to_support_pct = percentage_change(latest_close, support_level) if support_level not in {None, ZERO} else None
    distance_to_resistance_pct = percentage_change(resistance_level, latest_close) if resistance_level not in {None, ZERO} else None
    breakout_score = ZERO
    if previous_resistance not in {None, ZERO} and latest_close >= previous_resistance * Decimal("1.003"):
        breakout_label = "Ruptura alcista"
        breakout_score = Decimal("1.80")
    elif previous_support not in {None, ZERO} and latest_close <= previous_support * Decimal("0.997"):
        breakout_label = "Ruptura bajista"
        breakout_score = Decimal("-1.80")
    elif distance_to_support_pct is not None and distance_to_support_pct <= Decimal("3.00") and latest_candle["close_price"] >= latest_candle["open_price"]:
        breakout_label = "Apoyo en soporte"
        breakout_score = Decimal("0.45")
    elif distance_to_resistance_pct is not None and distance_to_resistance_pct <= Decimal("3.00") and latest_candle["close_price"] < latest_candle["open_price"]:
        breakout_label = "Freno en resistencia"
        breakout_score = Decimal("-0.45")
    else:
        breakout_label = "Sin ruptura"

    recent_range_avg = average_decimal(range_percentages[-10:]) or average_decimal(range_percentages) or ZERO
    previous_range_avg = average_decimal(range_percentages[:-10]) or recent_range_avg
    volatility_score = ZERO
    if previous_range_avg not in {None, ZERO} and recent_range_avg <= previous_range_avg * Decimal("0.85"):
        volatility_label = "Compresion"
        volatility_score = Decimal("0.30") if breakout_score > ZERO else Decimal("-0.10") if breakout_score < ZERO else ZERO
        band_multiplier = Decimal("0.96")
    elif previous_range_avg not in {None, ZERO} and recent_range_avg >= previous_range_avg * Decimal("1.15"):
        volatility_label = "Expansion"
        volatility_score = Decimal("0.20") if breakout_score > ZERO else Decimal("-0.20") if breakout_score < ZERO else ZERO
        band_multiplier = Decimal("1.08")
    else:
        volatility_label = "Normal"
        band_multiplier = Decimal("1.00")

    signal_score = clamp_decimal(
        trend_score + momentum_score + pattern["score"] + breakout_score + volatility_score,
        Decimal("-8.00"),
        Decimal("8.00"),
    )
    if signal_score >= Decimal("3.20"):
        signal_label = "Compra tecnica"
    elif signal_score >= Decimal("1.40"):
        signal_label = "Sesgo comprador"
    elif signal_score <= Decimal("-3.20"):
        signal_label = "Venta tecnica"
    elif signal_score <= Decimal("-1.40"):
        signal_label = "Sesgo vendedor"
    else:
        signal_label = "Neutral"

    component_signs = [
        value
        for value in (trend_score, momentum_score, pattern["score"], breakout_score)
        if abs(value) >= Decimal("0.35")
    ]
    aligned_positive = sum(1 for value in component_signs if value > ZERO)
    aligned_negative = sum(1 for value in component_signs if value < ZERO)
    alignment_bonus = Decimal(max(aligned_positive, aligned_negative)) * Decimal("4.00")
    confidence_score = clamp_decimal(
        Decimal("48.00") + (abs(signal_score) * Decimal("4.80")) + alignment_bonus,
        Decimal("42.00"),
        Decimal("86.00"),
    )
    confidence_label = resolve_projection_confidence_label(confidence_score)
    return_adjustment_pct = clamp_decimal(
        signal_score * (confidence_score / ONE_HUNDRED) * Decimal("0.48"),
        Decimal("-2.40"),
        Decimal("2.40"),
    )

    note_parts = [trend_label, momentum_label]
    if breakout_label != "Sin ruptura":
        note_parts.append(breakout_label)
    if pattern["label"] not in {"Vela alcista", "Vela bajista"}:
        note_parts.append(pattern["label"])
    note = (
        "Integra velas, medias, RSI, soportes/resistencias y volatilidad. "
        f"Lectura {signal_label.lower()}: {', '.join(note_parts[:3]).lower()}."
    )

    return {
        "available": True,
        "trend_label": trend_label,
        "last_candle_label": last_candle_label,
        "average_range_pct": average_decimal(range_percentages),
        "support_level": support_level,
        "resistance_level": resistance_level,
        "momentum_label": momentum_label,
        "pattern_label": pattern["label"],
        "pattern_note": pattern["note"],
        "breakout_label": breakout_label,
        "volatility_label": volatility_label,
        "signal_label": signal_label,
        "signal_score": quantize_decimal(signal_score),
        "confidence_score": quantize_decimal(confidence_score),
        "confidence_label": confidence_label,
        "return_adjustment_pct": quantize_decimal(return_adjustment_pct),
        "band_multiplier": quantize_decimal(band_multiplier, "0.01"),
        "rsi_14": rsi_14,
        "rsi_label": rsi_label,
        "distance_to_support_pct": quantize_decimal(distance_to_support_pct),
        "distance_to_resistance_pct": quantize_decimal(distance_to_resistance_pct),
        "short_return_pct": quantize_decimal(short_return_pct),
        "medium_return_pct": quantize_decimal(medium_return_pct),
        "note": note,
        "candlestick_svg": build_candlestick_svg(recent),
    }


def build_suggested_reference_cards(history, position: EquityPosition, reference_cache: dict) -> list[dict]:
    suggestions = []
    selected_key = (
        position.reference_profile,
        position.benchmark_symbol,
        position.benchmark_name,
    )
    for reference in build_reference_suggestions_for_equity(position.company_name, position.ticker):
        cache_key = (
            reference["reference_profile"],
            reference["benchmark_symbol"],
            reference["benchmark_name"],
        )
        correlation = {
            "frequency": infer_reference_frequency_from_profile(reference["reference_profile"]),
            "coefficient": None,
            "label": "Sin datos",
            "observations_count": 0,
        }
        try:
            cached_value = reference_cache.get(cache_key)
            if cached_value is None:
                cached_value = fetch_reference_series_for_choice(
                    reference["reference_profile"],
                    benchmark_symbol=reference["benchmark_symbol"],
                    benchmark_name=reference["benchmark_name"],
                )
                reference_cache[cache_key] = cached_value
            if isinstance(cached_value, Exception):
                raise cached_value
            correlation = build_reference_correlation_for_series(
                history,
                cached_value,
                reference["reference_profile"],
            )
        except Exception as exc:
            reference_cache[cache_key] = exc

        suggestions.append(
            {
                **reference,
                "is_selected": cache_key == selected_key,
                "correlation": correlation,
            }
        )
    suggestions.sort(
        key=lambda item: (
            0 if item["correlation"]["coefficient"] is not None else 1,
            -(abs(item["correlation"]["coefficient"])) if item["correlation"]["coefficient"] is not None else Decimal("0"),
            item["benchmark_name"],
        )
    )
    best_marked = False
    for suggestion in suggestions:
        coefficient = suggestion["correlation"]["coefficient"]
        is_best = coefficient is not None and not best_marked
        suggestion["is_best"] = is_best
        if is_best:
            best_marked = True
    return suggestions


def build_equity_history_card(
    position: EquityPosition,
    history,
    reference_cache: dict,
    selected_start_date: date | None = None,
    selected_end_date: date | None = None,
    status_key: str | None = None,
    status_label: str | None = None,
    detail_anchor: str | None = None,
    sector_label: str | None = None,
    include_visuals: bool = True,
    include_reference_suggestions: bool = True,
    include_fundamentals: bool | None = None,
) -> dict:
    resolved_status_key = status_key or ("owned" if position.is_owned else "watchlist")
    resolved_status_label = status_label or position.get_position_kind_display()
    resolved_detail_anchor = detail_anchor if detail_anchor is not None else (
        (f"tracked-ticket-{position.id}" if position.is_owned else f"stock-{position.id}") if position.id else ""
    )
    resolved_sector_label = sector_label or resolve_equity_sector_label(
        company_name=position.company_name,
        ticker=position.ticker,
        quote_symbol=position.quote_symbol,
    )
    sale_preview = build_equity_sale_preview(position)
    resolved_include_fundamentals = should_fetch_equity_fundamentals() if include_fundamentals is None else include_fundamentals
    fundamentals = (
        build_equity_fundamentals_summary(position, force_live_fetch=bool(include_fundamentals))
        if resolved_include_fundamentals
        else {
            "available": False,
            "note": "La lectura fundamental se ha omitido en esta vista para priorizar velocidad.",
            "net_income_rows": [],
            "derived_per": None,
        }
    )
    valuation = build_equity_per_valuation(
        position,
        fundamentals,
        sector_label=resolved_sector_label,
    )

    if not history:
        return {
            "position": position,
            "status_key": resolved_status_key,
            "status_label": resolved_status_label,
            "detail_anchor": resolved_detail_anchor,
            "sector_label": resolved_sector_label,
            "has_history": False,
            "sale_preview": sale_preview,
            "projection": {"available": False},
            "cycle_projection_5y": {"available": False},
            "projection_backtest": {"available": False, "monthly_chart": {"available": False}},
            "projection_reliability": {"label": "Baja", "score": Decimal("40.00")},
            "trade_alert": {
                "label": "Vigilar",
                "tone": "watch",
                "score": ZERO,
                "note": "Todavia no hay historico suficiente para activar una alerta.",
                "trend_label": "Sin tendencia relativa",
                "trigger_label": "Sin historico suficiente",
            },
            "coefficient_alert": {
                "available": False,
                "label": "Sin lectura",
                "tone": "neutral",
                "note": "Sin historico suficiente para leer el coeficiente de referencia.",
                "trigger_label": "Sin historico suficiente",
            },
            "valuation": valuation,
            "fundamentals": fundamentals,
            "technical_signal": {
                "available": False,
                "signal_label": "Sin senal",
                "signal_score": None,
                "confidence_label": "Baja",
                "note": "",
            },
            "reference_playbook": {"available": False, "candidates": []},
            "suggested_references": [],
            "market_history_points": [],
            "historical_chart": {"available": False},
            "projection_12m_chart": {"available": False},
            "cycle_projection_5y_chart": {"available": False},
        }

    first_price = history[0].close_price
    first_benchmark = next((point.benchmark_close for point in history if point.benchmark_close is not None), None)
    latest_point = history[-1]
    if selected_start_date or selected_end_date:
        selected_period = build_period_snapshot(
            history,
            "Periodo elegido",
            start_date=selected_start_date,
            end_date=selected_end_date,
            reference_profile=position.reference_profile,
        )
    else:
        selected_period = {"available": False}
    if not selected_period["available"] and latest_point.price_date:
        selected_period = build_period_snapshot(
            history,
            "1Y",
            start_date=latest_point.price_date - timedelta(days=365),
            end_date=latest_point.price_date,
            reference_profile=position.reference_profile,
        )

    period_snapshots = [
        build_period_snapshot(history, "1Y", start_date=latest_point.price_date - timedelta(days=365), end_date=latest_point.price_date, reference_profile=position.reference_profile),
        build_period_snapshot(history, "3Y", start_date=latest_point.price_date - timedelta(days=365 * 3), end_date=latest_point.price_date, reference_profile=position.reference_profile),
        build_period_snapshot(history, "5Y", start_date=latest_point.price_date - timedelta(days=365 * 5), end_date=latest_point.price_date, reference_profile=position.reference_profile),
        build_period_snapshot(history, "10Y", start_date=latest_point.price_date - timedelta(days=LONG_ANALYSIS_DAYS), end_date=latest_point.price_date, reference_profile=position.reference_profile),
    ]
    one_year_snapshot = next((snapshot for snapshot in period_snapshots if snapshot["label"] == "1Y"), {"available": False})
    correlation = build_reference_correlation(history, position)
    six_month_snapshot = build_period_snapshot(
        history,
        "6M",
        start_date=latest_point.price_date - timedelta(days=182),
        end_date=latest_point.price_date,
        reference_profile=position.reference_profile,
    )
    cycle_metrics = build_cycle_metrics(history)
    technical_signal = build_candlestick_metrics(history)
    projection = build_one_year_projection(
        history,
        position,
        correlation,
        six_month_snapshot,
        cycle_metrics=cycle_metrics,
        technical_signal=technical_signal,
    )
    valuation = build_equity_per_valuation(
        position,
        fundamentals,
        sector_label=resolved_sector_label,
    )
    projection = apply_per_valuation_overlay(projection, valuation)
    cycle_projection_5y = build_five_year_cycle_projection(
        history,
        position,
        correlation,
        cycle_metrics=cycle_metrics,
        reference_cache=reference_cache,
        include_visuals=include_visuals,
    )
    synchronize_projection_path_with_cycle_zoom(
        projection,
        cycle_projection_5y,
        current_price=latest_point.close_price or position.current_price_per_share,
        anchor_date=latest_point.price_date,
    )
    projection_backtest = build_projection_backtest(
        history,
        position,
        include_monthly_chart=include_visuals,
    )
    projection_reliability = (
        build_projection_reliability(projection, projection_backtest)
        if projection.get("available")
        else {"label": "Baja", "score": Decimal("40.00")}
    )
    relative_trend = build_relative_strength_trend(
        history,
        position,
        correlation.get("beta") if correlation.get("beta") is not None else correlation.get("coefficient"),
    )
    coefficient_alert = build_reference_coefficient_alert(history, position, correlation)
    trade_alert = build_trade_alert(
        position,
        projection,
        correlation,
        projection_reliability,
        relative_trend,
        six_month_snapshot,
        one_year_snapshot,
        valuation=valuation,
        technical_signal=technical_signal,
    )
    analysis_value_amount = projection.get("analysis_value_amount") or ZERO
    annual_cost_used = projection.get("annual_cost_used", position.recurring_cost_used) or ZERO
    maintenance_drag_pct = (
        (annual_cost_used / analysis_value_amount) * Decimal("100")
        if analysis_value_amount
        else (
            (position.recurring_cost_used / position.invested_amount) * Decimal("100")
            if position.invested_amount
            else ZERO
        )
    )
    broker_costs = {
        **(projection.get("broker_costs") or position.estimated_broker_costs),
        "annual_cost_used": annual_cost_used,
        "annual_cost_source": projection.get("annual_cost_source", position.recurring_cost_source),
        "net_dividend_income": projection.get("net_dividend_income", position.net_dividend_income),
        "analysis_value_amount": analysis_value_amount,
        "analysis_value_source": projection.get("analysis_value_source", "actual"),
        "annual_dividend_income_used": projection.get("annual_dividend_income_used", position.annual_dividend_income),
    }
    dual_axis_chart = {
        "stock_line": "",
        "reference_line": "",
        "projection_line": "",
        "stock_min_label": "-",
        "stock_max_label": "-",
        "reference_min_label": "-",
        "reference_max_label": "-",
        "x_markers": [],
    }
    historical_chart = {
        "available": False,
        "stock_line": "",
        "stock_min_label": "-",
        "stock_max_label": "-",
        "x_markers": [],
    }
    best_correlation_chart = {
        "available": False,
        "stock_line": "",
        "reference_line": "",
        "stock_min_label": "-",
        "stock_max_label": "-",
        "reference_min_label": "-",
        "reference_max_label": "-",
        "reference_name": "",
        "x_markers": [],
    }
    projection_12m_chart = {
        "available": False,
        "stock_line": "",
        "projection_line": "",
        "stock_min_label": "-",
        "stock_max_label": "-",
        "x_markers": [],
    }
    cycle_projection_5y_chart = {
        "available": False,
        "stock_line": "",
        "projection_line": "",
        "stock_min_label": "-",
        "stock_max_label": "-",
        "x_markers": [],
    }
    if include_visuals:
        stock_series = [{"date": point.price_date, "value": point.close_price} for point in history]
        benchmark_series = [
            {"date": point.price_date, "value": point.benchmark_close}
            for point in history
            if point.benchmark_close is not None
        ]
        projection_series = []
        if projection.get("available"):
            projection_path = resolve_projection_tracking_path(projection)
            projection_series = [{"date": latest_point.price_date, "value": latest_point.close_price}]
            projection_series.extend(
                {
                    "date": step["projected_date"],
                    "value": step["projected_price"],
                }
                for step in projection_path
                if step.get("projected_date") and step.get("projected_price") is not None
            )
        dual_axis_chart = build_dual_axis_chart(stock_series, benchmark_series, projection_points=projection_series)
        historical_chart = build_stock_history_chart(history)
        projection_12m_chart = build_projection_12m_chart(history, projection)
        cycle_projection_5y_chart = build_cycle_projection_5y_chart(history, cycle_projection_5y)

    suggested_references = []
    if include_reference_suggestions:
        suggested_references = build_suggested_reference_cards(history, position, reference_cache)
        if include_visuals:
            best_correlation_chart = build_best_correlation_chart(history, suggested_references, reference_cache)

    card = {
        "position": position,
        "status_key": resolved_status_key,
        "status_label": resolved_status_label,
        "detail_anchor": resolved_detail_anchor,
        "sector_label": resolved_sector_label or "Sin sector",
        "has_history": True,
        "points_count": len(history),
        "start_date": history[0].price_date,
        "end_date": history[-1].price_date,
        "reference_label": position.analysis_reference_label,
        "reference_profile_label": position.get_reference_profile_display(),
        "stock_return_pct": ((history[-1].close_price / first_price) - 1) * Decimal("100") if first_price else ZERO,
        "benchmark_return_pct": (
            ((history[-1].benchmark_close / first_benchmark) - 1) * Decimal("100")
            if first_benchmark and history[-1].benchmark_close
            else None
        ),
        "stock_line": dual_axis_chart["stock_line"],
        "benchmark_line": dual_axis_chart["reference_line"],
        "projection_line": dual_axis_chart["projection_line"],
        "dual_axis_chart": dual_axis_chart,
        "period_snapshots": period_snapshots,
        "six_month_snapshot": six_month_snapshot,
        "selected_period": selected_period,
        "net_unrealized_gain": position.unrealized_gain_after_costs,
        "net_unrealized_return_pct": position.unrealized_return_pct,
        "net_annual_income": position.net_annual_income,
        "maintenance_drag_pct": maintenance_drag_pct,
        "price_vs_cost_pct": percentage_change(position.current_price_per_share, position.average_cost_per_share),
        "broker_costs": broker_costs,
        "correlation": correlation,
        "cycle_metrics": cycle_metrics,
        "sale_preview": sale_preview,
        "projection": projection,
        "cycle_projection_5y": cycle_projection_5y,
        "projection_backtest": projection_backtest,
        "projection_reliability": projection_reliability,
        "relative_trend": relative_trend,
        "coefficient_alert": coefficient_alert,
        "trade_alert": trade_alert,
        "valuation": valuation,
        "fundamentals": fundamentals,
        "technical_signal": technical_signal,
        "suggested_references": suggested_references,
        "market_history_points": [
            {
                "date": point.price_date,
                "close": point.close_price,
            }
            for point in history
            if point.price_date is not None and point.close_price is not None
        ],
        "historical_chart": historical_chart,
        "best_correlation_chart": best_correlation_chart,
        "projection_12m_chart": projection_12m_chart,
        "cycle_projection_5y_chart": cycle_projection_5y_chart,
        "presentation_projection": {"available": False},
        "information_basis": {"available": False, "rows": []},
    }
    card["reference_playbook"] = (
        build_reference_playbook_from_card(card)
        if include_reference_suggestions
        else {
            "available": False,
            "source_label": "",
            "company_sector": resolved_sector_label or "",
            "current_reference_label": card.get("reference_label"),
            "current_candidate": None,
            "best_candidate": None,
            "candidates": [],
            "return_2025": None,
            "per_2025": None,
            "dividend_yield": None,
            "notes": position.notes,
        }
    )
    return refresh_card_projection_visuals(card, history=history, include_visuals=include_visuals)


def build_equity_history_cards(
    positions,
    selected_start_date: date | None = None,
    selected_end_date: date | None = None,
    reference_cache: dict | None = None,
    include_fundamentals: bool | None = None,
) -> list[dict]:
    cards = []
    reference_cache = reference_cache if reference_cache is not None else {}
    for position in positions:
        history = list(position.price_history.order_by("price_date"))
        cards.append(
            build_equity_history_card(
                position,
                history,
                reference_cache,
                selected_start_date=selected_start_date,
                selected_end_date=selected_end_date,
                include_fundamentals=include_fundamentals,
            )
        )
    return cards


def resolve_analysis_broker_profile(positions) -> dict:
    fallback = {
        "broker": "Interactive Brokers",
        "trade_channel": EquityPosition.TradeChannel.APP,
        "ownership_category": AssetOwnershipCategory.JOINT,
    }
    candidates = [position for position in positions if position.is_owned] or list(positions)
    if not candidates:
        return {
            **fallback,
            "trade_channel_label": dict(EquityPosition.TradeChannel.choices).get(EquityPosition.TradeChannel.APP, "App"),
        }

    grouped: dict[tuple[str, str, str], dict] = {}
    for position in candidates:
        key = (position.broker, position.trade_channel, position.ownership_category)
        summary = grouped.setdefault(key, {"count": 0, "current_value": ZERO})
        summary["count"] += 1
        summary["current_value"] += position.current_value if position.is_owned else ZERO

    selected_key, _ = max(
        grouped.items(),
        key=lambda item: (item[1]["count"], item[1]["current_value"]),
    )
    broker, trade_channel, ownership_category = selected_key
    return {
        "broker": broker,
        "trade_channel": trade_channel,
        "ownership_category": ownership_category,
        "trade_channel_label": dict(EquityPosition.TradeChannel.choices).get(trade_channel, trade_channel),
    }


def build_ibex_universe_companies(workbook_snapshot: dict) -> list[dict]:
    merged: dict[str, dict] = {}
    catalog = get_equity_company_catalog()
    catalog_by_ticker = {clean_ticker(entry["ticker"]): entry for entry in catalog}

    for company in workbook_snapshot.get("companies", []):
        profile = (
            catalog_by_ticker.get(clean_ticker(company.get("ticker", "")))
            or find_equity_company_profile(company.get("ticker", ""))
            or find_equity_company_profile(company.get("company_name", ""))
        )
        ticker = clean_ticker(company.get("ticker", "")) or (profile["ticker"] if profile else "")
        if not ticker:
            continue
        merged[ticker] = {
            **company,
            "ticker": ticker,
            "company_name": company.get("company_name") or (profile["company_name"] if profile else ticker),
            "quote_symbol": company.get("quote_symbol") or (profile["quote_symbol"] if profile else ""),
            "sector": company.get("sector") or (profile["sector_label"] if profile else ""),
            "catalog_profile": profile,
        }

    for profile in catalog:
        ticker = clean_ticker(profile["ticker"])
        if ticker in merged:
            merged[ticker].setdefault("quote_symbol", profile["quote_symbol"])
            merged[ticker].setdefault("sector", profile["sector_label"])
            merged[ticker].setdefault("company_name", profile["company_name"])
            merged[ticker]["catalog_profile"] = merged[ticker].get("catalog_profile") or profile
            continue
        merged[ticker] = {
            "ticker": ticker,
            "company_name": profile["company_name"],
            "quote_symbol": profile["quote_symbol"],
            "sector": profile["sector_label"],
            "dividend_yield": None,
            "catalog_profile": profile,
        }

    return sorted(merged.values(), key=lambda item: item.get("company_name") or item.get("ticker") or "")


def resolve_ibex_reference_choice(company: dict, workbook_snapshot: dict) -> tuple[dict, dict | None]:
    workbook_playbook = None
    if workbook_snapshot.get("available"):
        workbook_playbook = build_workbook_reference_playbook(company, workbook_snapshot)
        for candidate in workbook_playbook.get("candidates", []):
            if candidate.get("supports_chart") and candidate.get("live_reference"):
                return dict(candidate["live_reference"]), workbook_playbook

    profile = company.get("catalog_profile") or find_equity_company_profile(company.get("ticker", ""))
    if profile and profile.get("default_reference"):
        return dict(profile["default_reference"]), workbook_playbook
    return get_reference_preset("ibex_35"), workbook_playbook


def build_virtual_ibex_position(
    company: dict,
    reference_choice: dict,
    broker_profile: dict,
    latest_price: Decimal,
) -> EquityPosition:
    dividend_yield_pct = company.get("dividend_yield") or ZERO
    annual_dividend_income = (latest_price * dividend_yield_pct) / ONE_HUNDRED if dividend_yield_pct else ZERO
    return EquityPosition(
        position_kind=EquityPosition.PositionKind.WATCHLIST,
        ownership_category=broker_profile["ownership_category"],
        broker=broker_profile["broker"],
        ticker=clean_ticker(company.get("ticker", "")),
        quote_symbol=company.get("quote_symbol", ""),
        reference_profile=reference_choice["reference_profile"],
        benchmark_symbol=reference_choice["benchmark_symbol"],
        benchmark_name=reference_choice["benchmark_name"],
        company_name=company.get("company_name") or clean_ticker(company.get("ticker", "")),
        trade_channel=broker_profile["trade_channel"],
        shares=Decimal("1.0000"),
        average_cost_per_share=latest_price,
        current_price_per_share=latest_price,
        annual_dividend_income=quantize_decimal(annual_dividend_income, "0.01") or ZERO,
        annual_maintenance_cost=ZERO,
        latest_price_date=None,
        notes="Radar IBEX automatico",
    )


def find_ibex_universe_company(query: str, workbook_snapshot: dict | None = None) -> tuple[dict | None, dict]:
    workbook_snapshot = workbook_snapshot or load_ibex_reference_workbook_snapshot()
    normalized_query = normalize_company_lookup(query)
    if not normalized_query:
        return None, workbook_snapshot

    for company in build_ibex_universe_companies(workbook_snapshot):
        lookup_keys = build_security_lookup_keys(
            ticker=company.get("ticker", ""),
            company_name=company.get("company_name", ""),
            quote_symbol=company.get("quote_symbol", ""),
        )
        if normalized_query in lookup_keys:
            return company, workbook_snapshot
    return None, workbook_snapshot


def build_ibex_universe_card(
    company: dict,
    positions,
    selected_start_date: date | None = None,
    selected_end_date: date | None = None,
    reference_cache: dict | None = None,
    workbook_snapshot: dict | None = None,
    broker_profile: dict | None = None,
    include_visuals: bool = True,
    include_reference_suggestions: bool = True,
    include_fundamentals: bool | None = None,
) -> dict:
    reference_cache = reference_cache if reference_cache is not None else {}
    workbook_snapshot = workbook_snapshot or load_ibex_reference_workbook_snapshot()
    broker_profile = broker_profile or resolve_analysis_broker_profile(positions)

    quote_symbol = company.get("quote_symbol", "")
    if not quote_symbol:
        raise ValueError(f"{company.get('company_name') or company.get('ticker')}: sin simbolo de cotizacion")

    stock_series = fetch_market_series(quote_symbol)
    reference_choice, workbook_playbook = resolve_ibex_reference_choice(company, workbook_snapshot)
    reference_series = fetch_reference_series_for_choice(
        reference_choice["reference_profile"],
        benchmark_symbol=reference_choice["benchmark_symbol"],
        benchmark_name=reference_choice["benchmark_name"],
    )
    benchmark_map = align_reference_points(stock_series.points, reference_series.points)
    position = build_virtual_ibex_position(company, reference_choice, broker_profile, stock_series.latest_price)
    position.latest_price_date = stock_series.latest_date
    position.last_synced_at = django_timezone.now()
    history = [
        EquityPriceHistory(
            position=position,
            price_date=point["date"],
            open_price=point.get("open"),
            high_price=point.get("high"),
            low_price=point.get("low"),
            close_price=point["close"],
            benchmark_close=benchmark_map.get(point["date"]),
        )
        for point in stock_series.points
    ]
    card = build_equity_history_card(
        position,
        history,
        reference_cache,
        selected_start_date=selected_start_date,
        selected_end_date=selected_end_date,
        status_key="ibex",
        status_label="Solo radar",
        detail_anchor="",
        sector_label=company.get("sector", ""),
        include_visuals=include_visuals,
        include_reference_suggestions=include_reference_suggestions,
        include_fundamentals=include_fundamentals,
    )
    if workbook_playbook and workbook_playbook.get("available"):
        card["reference_playbook"] = build_workbook_reference_playbook(company, workbook_snapshot, card=card)
    card["ibex_company"] = company
    return card


def build_ibex_registered_summary(cards: list[dict]) -> dict:
    owned_count = sum(1 for card in cards if card["position"].is_owned)
    watchlist_count = sum(1 for card in cards if not card["position"].is_owned)
    if owned_count and watchlist_count:
        return {
            "status_key": "owned",
            "status_label": "Comprada + seguimiento",
            "status_note": f"{owned_count} compra(s) y {watchlist_count} seguimiento(s) guardados",
        }
    if owned_count:
        return {
            "status_key": "owned",
            "status_label": "Comprada",
            "status_note": f"{owned_count} posicion(es) guardadas" if owned_count > 1 else "Ya esta en cartera",
        }
    return {
        "status_key": "watchlist",
        "status_label": "En seguimiento",
        "status_note": f"{watchlist_count} seguimiento(s) guardados" if watchlist_count > 1 else "La tienes guardada",
    }


def merge_tracked_ibex_card(company: dict, matching_cards: list[dict]) -> dict:
    def card_rank(card: dict) -> tuple[int, int, int]:
        position = card["position"]
        return (
            0 if position.is_owned else 1,
            0 if position.current_price_per_share else 1,
            -(position.id or 0),
        )

    primary = dict(sorted(matching_cards, key=card_rank)[0])
    summary = build_ibex_registered_summary(matching_cards)
    primary["status_key"] = summary["status_key"]
    primary["status_label"] = summary["status_label"]
    primary["status_note"] = summary["status_note"]
    primary["sector_label"] = company.get("sector", "") or primary.get("sector_label") or "Sin sector"
    primary["ibex_company"] = company
    return primary


def build_optimizer_master_cards(history_cards: list[dict], ibex_cards: list[dict]) -> list[dict]:
    ibex_keys = set()
    for card in ibex_cards:
        position = card["position"]
        ibex_keys.update(
            build_security_lookup_keys(
                ticker=position.ticker,
                company_name=position.company_name,
                quote_symbol=position.quote_symbol,
            )
        )
    cards = list(ibex_cards)
    for card in history_cards:
        position = card["position"]
        position_keys = build_security_lookup_keys(
            ticker=position.ticker,
            company_name=position.company_name,
            quote_symbol=position.quote_symbol,
        )
        if position_keys & ibex_keys:
            continue
        cards.append(card)
    cards_by_ticker = {
        clean_ticker(getattr(card.get("position"), "ticker", "")): card
        for card in cards
        if clean_ticker(getattr(card.get("position"), "ticker", ""))
    }
    expectation_signal_map = build_optimizer_expectation_review_signal_map(
        [getattr(card.get("position"), "ticker", "") for card in cards],
        current_cards_by_ticker=cards_by_ticker,
    )
    for card in cards:
        position = card.get("position")
        ticker = clean_ticker(getattr(position, "ticker", ""))
        card["expectation_review_signal"] = expectation_signal_map.get(ticker, {"available": False})
    return cards


def build_ibex_universe_analysis(
    tracked_history_cards: list[dict],
    positions,
    selected_start_date: date | None = None,
    selected_end_date: date | None = None,
    reference_cache: dict | None = None,
    company_limit: int | None = None,
    progress_callback=None,
    include_visuals: bool = False,
    include_reference_suggestions: bool = False,
    include_fundamentals: bool = False,
) -> dict:
    reference_cache = reference_cache if reference_cache is not None else {}
    workbook_snapshot = load_ibex_reference_workbook_snapshot()
    companies = build_ibex_universe_companies(workbook_snapshot)
    tracked_card_map: dict[str, list[dict]] = defaultdict(list)
    for card in tracked_history_cards:
        position = card["position"]
        for key in build_security_lookup_keys(
            ticker=position.ticker,
            company_name=position.company_name,
            quote_symbol=position.quote_symbol,
        ):
            tracked_card_map[key].append(card)

    broker_profile = resolve_analysis_broker_profile(positions)
    tracked_cards = []
    candidate_companies = []
    failures = []

    for company in companies:
        company_keys = build_security_lookup_keys(
            ticker=company.get("ticker", ""),
            company_name=company.get("company_name", ""),
            quote_symbol=company.get("quote_symbol", ""),
        )
        matching_cards = []
        for key in company_keys:
            matching_cards.extend(tracked_card_map.get(key, []))
        unique_matching_cards = []
        seen_card_ids = set()
        for card in matching_cards:
            card_id = id(card)
            if card_id in seen_card_ids:
                continue
            seen_card_ids.add(card_id)
            unique_matching_cards.append(card)
        if unique_matching_cards:
            tracked_cards.append(merge_tracked_ibex_card(company, unique_matching_cards))
            continue
        if company_limit is not None and len(candidate_companies) >= company_limit:
            break
        quote_symbol = company.get("quote_symbol", "")
        if not quote_symbol:
            failures.append(f"{company.get('company_name') or company.get('ticker')}: sin simbolo de cotizacion")
            continue
        candidate_companies.append(company)

    cards = [None] * len(candidate_companies)
    if candidate_companies:
        max_workers = min(ibex_universe_max_workers(), len(candidate_companies))
        if max_workers <= 1:
            for index, company in enumerate(candidate_companies):
                try:
                    cards[index] = build_ibex_universe_card(
                        company,
                        positions,
                        selected_start_date=selected_start_date,
                        selected_end_date=selected_end_date,
                        reference_cache={},
                        workbook_snapshot=workbook_snapshot,
                        broker_profile=broker_profile,
                        include_visuals=include_visuals,
                        include_reference_suggestions=include_reference_suggestions,
                        include_fundamentals=include_fundamentals,
                    )
                except Exception as exc:
                    failures.append(f"{company.get('company_name') or company.get('ticker')}: {exc}")
                if progress_callback:
                    progress_callback(
                        completed_count=index + 1,
                        total_count=len(candidate_companies),
                        company=company,
                        completed_cards=len([card for card in cards if card is not None]),
                        failures_count=len(failures),
                    )
        else:
            with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="ibex-analysis") as executor:
                future_map = {
                    executor.submit(
                        build_ibex_universe_card,
                        company,
                        positions,
                        selected_start_date=selected_start_date,
                        selected_end_date=selected_end_date,
                        reference_cache={},
                        workbook_snapshot=workbook_snapshot,
                        broker_profile=broker_profile,
                        include_visuals=include_visuals,
                        include_reference_suggestions=include_reference_suggestions,
                        include_fundamentals=include_fundamentals,
                    ): (index, company)
                    for index, company in enumerate(candidate_companies)
                }
                completed_count = 0
                for future in as_completed(future_map):
                    index, company = future_map[future]
                    try:
                        cards[index] = future.result()
                    except Exception as exc:
                        failures.append(f"{company.get('company_name') or company.get('ticker')}: {exc}")
                    completed_count += 1
                    if progress_callback:
                        progress_callback(
                            completed_count=completed_count,
                            total_count=len(candidate_companies),
                            company=company,
                            completed_cards=len([card for card in cards if card is not None]),
                            failures_count=len(failures),
                        )

    generated_cards = [card for card in cards if card is not None]
    cards = [*tracked_cards, *generated_cards]
    target_count = len(tracked_cards) + len(candidate_companies)

    rows = build_equity_decision_rows(cards)
    buy_alert_count = sum(1 for card in cards if card.get("trade_alert", {}).get("label") == "Comprar")
    sell_alert_count = sum(1 for card in cards if card.get("trade_alert", {}).get("label") == "Vender")
    registered_owned_count = sum(1 for card in tracked_cards if card.get("status_key") == "owned")
    registered_watchlist_count = sum(1 for card in tracked_cards if card.get("status_key") == "watchlist")

    return {
        "cards": cards,
        "tracked_cards": tracked_cards,
        "generated_cards": generated_cards,
        "rows": rows,
        "summary": {
            "available": bool(cards),
            "workbook_loaded": workbook_snapshot.get("available", False),
            "source_label": Path(workbook_snapshot["path"]).name if workbook_snapshot.get("path") else "Catalogo IBEX",
            "analyzed_count": len(cards),
            "target_count": target_count,
            "buy_alert_count": buy_alert_count,
            "sell_alert_count": sell_alert_count,
            "watch_alert_count": max(len(cards) - buy_alert_count - sell_alert_count, 0),
            "registered_count": len(tracked_cards),
            "registered_owned_count": registered_owned_count,
            "registered_watchlist_count": registered_watchlist_count,
            "radar_only_count": max(len(cards) - len(tracked_cards), 0),
            "failed_count": len(failures),
            "failures": failures[:8],
            "broker_assumption": broker_profile["broker"],
            "trade_channel_label": broker_profile["trade_channel_label"],
            "top_pick": rows[0] if rows else None,
        },
    }


def build_owned_positions_comparable_summary(history_cards: list[dict]) -> dict:
    comparable_cards = [
        card
        for card in history_cards
        if card["position"].is_owned and card.get("sale_preview", {}).get("available")
    ]
    if not comparable_cards:
        return {
            "available": False,
            "positions_count": 0,
        }

    total_committed_capital = sum(
        (card["sale_preview"].get("committed_capital") or ZERO)
        for card in comparable_cards
    )
    weighted_total_return_pct = (
        sum(
            (
                (card["sale_preview"].get("cumulative_margin_pct") or ZERO)
                * (card["sale_preview"].get("committed_capital") or ZERO)
            )
            for card in comparable_cards
        )
        / total_committed_capital
        if total_committed_capital
        else None
    )

    comparable_weight = sum(
        (
            card["sale_preview"].get("committed_capital") or ZERO
        )
        for card in comparable_cards
        if card["sale_preview"].get("monthly_equivalent_return_pct") is not None
    )
    weighted_monthly_return_pct = (
        sum(
            (
                card["sale_preview"]["monthly_equivalent_return_pct"]
                * (card["sale_preview"].get("committed_capital") or ZERO)
            )
            for card in comparable_cards
            if card["sale_preview"].get("monthly_equivalent_return_pct") is not None
        )
        / comparable_weight
        if comparable_weight
        else None
    )
    weighted_annual_return_pct = (
        sum(
            (
                card["sale_preview"]["annualized_margin_pct"]
                * (card["sale_preview"].get("committed_capital") or ZERO)
            )
            for card in comparable_cards
            if card["sale_preview"].get("annualized_margin_pct") is not None
        )
        / comparable_weight
        if comparable_weight
        else None
    )

    best_card = max(
        comparable_cards,
        key=lambda card: card["sale_preview"].get("annualized_margin_pct") or Decimal("-9999"),
    )
    worst_card = min(
        comparable_cards,
        key=lambda card: card["sale_preview"].get("annualized_margin_pct") or Decimal("9999"),
    )
    return {
        "available": True,
        "positions_count": len(comparable_cards),
        "total_committed_capital": quantize_decimal(total_committed_capital, "0.01") or ZERO,
        "weighted_total_return_pct": quantize_decimal(weighted_total_return_pct, "0.01"),
        "weighted_monthly_return_pct": quantize_decimal(weighted_monthly_return_pct, "0.01"),
        "weighted_annual_return_pct": quantize_decimal(weighted_annual_return_pct, "0.01"),
        "best_ticker": best_card["position"].ticker,
        "best_annual_return_pct": quantize_decimal(best_card["sale_preview"].get("annualized_margin_pct"), "0.01"),
        "worst_ticker": worst_card["position"].ticker,
        "worst_annual_return_pct": quantize_decimal(worst_card["sale_preview"].get("annualized_margin_pct"), "0.01"),
        "method_label": "Rentabilidad equivalente compuesta sobre capital comprometido neto si cerraras hoy.",
    }


def build_selected_period_label(
    selected_start_date: date | None = None,
    selected_end_date: date | None = None,
) -> str:
    if selected_start_date and selected_end_date:
        return f"{selected_start_date:%Y-%m-%d} a {selected_end_date:%Y-%m-%d}"
    if selected_start_date:
        return f"Desde {selected_start_date:%Y-%m-%d}"
    if selected_end_date:
        return f"Hasta {selected_end_date:%Y-%m-%d}"
    return "Ultimos 90 dias"


def resolve_projection_path_step(
    projection: dict | None,
    *,
    months: int,
    anchor_date: date | None = None,
) -> dict | None:
    projection = projection or {}
    if months <= 0:
        return None

    exact_match = None
    dated_candidates = []
    for step in resolve_projection_tracking_path(projection):
        step_price = step.get("projected_price")
        if step_price in {None, ZERO}:
            continue
        months_offset = parse_projection_label_months(step.get("label"))
        if months_offset == months:
            exact_match = step
            break
        projected_date = step.get("projected_date")
        if projected_date is not None and anchor_date is not None:
            months_offset = max(
                ((projected_date.year - anchor_date.year) * 12)
                + (projected_date.month - anchor_date.month),
                0,
            )
            dated_candidates.append((abs(months_offset - months), months_offset, step))
    if exact_match is not None:
        return exact_match
    if dated_candidates:
        dated_candidates.sort(key=lambda item: (item[0], item[1]))
        return dated_candidates[0][2]
    return None


def build_projection_horizon_snapshot(card: dict, months: int) -> dict | None:
    projection = card.get("projection") or {}
    position = card["position"]
    current_value = quantize_decimal(position.current_value, "0.01") or ZERO
    if not projection.get("available") or current_value <= ZERO:
        return None

    current_price = quantize_decimal(
        projection.get("latest_price") or position.current_price_per_share,
        "0.0001",
    )
    anchor_date = projection.get("latest_date")
    projection_step = resolve_projection_path_step(
        projection,
        months=months,
        anchor_date=anchor_date,
    )
    projected_price = quantize_decimal(
        projection_step.get("projected_price") if projection_step else None,
        "0.0001",
    )
    if current_price in {None, ZERO} or projected_price in {None, ZERO}:
        return None

    price_return_pct = quantize_decimal(
        percentage_change(projected_price, current_price),
        "0.01",
    )
    if price_return_pct is None:
        return None

    months_ratio = Decimal(str(months / 12))
    total_return_pct = price_return_pct
    net_income_yield_pct = quantize_decimal(projection.get("net_income_yield_pct"), "0.01")
    transaction_drag_pct = quantize_decimal(projection.get("transaction_drag_pct"), "0.01")
    if net_income_yield_pct is not None:
        total_return_pct += net_income_yield_pct * months_ratio
    if transaction_drag_pct is not None:
        total_return_pct -= transaction_drag_pct
    total_return_pct = quantize_decimal(total_return_pct, "0.01")
    if total_return_pct is None:
        return None

    projected_total_value = quantize_decimal(
        current_value * (Decimal("1") + (total_return_pct / ONE_HUNDRED)),
        "0.01",
    )
    projected_market_value = quantize_decimal(position.shares * projected_price, "0.01")
    return {
        "months": months,
        "label": f"{months}M" if months < 12 else "12M",
        "current_value": current_value,
        "projected_total_value": projected_total_value,
        "projected_market_value": projected_market_value,
        "projected_price": projected_price,
        "return_pct": total_return_pct,
        "price_return_pct": price_return_pct,
        "projected_date": projection_step.get("projected_date") if projection_step else None,
    }


def build_portfolio_projection_horizons(history_cards: list[dict]) -> list[dict]:
    owned_projection_cards = [
        card
        for card in history_cards
        if card["position"].is_owned and card.get("projection", {}).get("available")
    ]
    horizon_rows = []
    for months, label in PORTFOLIO_PROJECTION_HORIZONS:
        covered_positions = 0
        current_value_total = ZERO
        projected_total_value = ZERO
        projected_market_value = ZERO
        latest_projected_date = None
        for card in owned_projection_cards:
            snapshot = build_projection_horizon_snapshot(card, months)
            if snapshot is None or snapshot.get("projected_total_value") is None:
                continue
            covered_positions += 1
            current_value_total += snapshot["current_value"]
            projected_total_value += snapshot["projected_total_value"]
            projected_market_value += snapshot.get("projected_market_value") or ZERO
            projected_date = snapshot.get("projected_date")
            if projected_date is not None and (latest_projected_date is None or projected_date > latest_projected_date):
                latest_projected_date = projected_date
        horizon_return_pct = (
            quantize_decimal(percentage_change(projected_total_value, current_value_total), "0.01")
            if current_value_total > ZERO
            else None
        )
        horizon_rows.append(
            {
                "label": label,
                "months": months,
                "return_pct": horizon_return_pct,
                "projected_total_value": quantize_decimal(projected_total_value, "0.01") if current_value_total > ZERO else None,
                "projected_market_value": quantize_decimal(projected_market_value, "0.01") if current_value_total > ZERO else None,
                "positions_count": covered_positions,
                "projection_end_date": latest_projected_date,
                "tone": (
                    "good"
                    if horizon_return_pct is not None and horizon_return_pct >= ZERO
                    else ("warn" if horizon_return_pct is not None else "")
                ),
            }
        )
    return horizon_rows


def build_portfolio_expectation_horizons(
    history_cards: list[dict],
    decision_rows: list[dict],
) -> list[dict]:
    owned_cards = [
        card
        for card in history_cards
        if card["position"].is_owned and (quantize_decimal(card["position"].current_value, "0.01") or ZERO) > ZERO
    ]
    decision_rows_by_ticker = {
        str(row.get("ticker") or "").strip().upper(): row
        for row in list(decision_rows or [])
        if row.get("ticker")
    }

    horizon_rows = []
    for months, label, row_key in PORTFOLIO_EXPECTATION_HORIZONS:
        covered_positions = 0
        current_value_total = ZERO
        expected_total_value = ZERO
        weighted_expected_return = ZERO
        for card in owned_cards:
            position = card["position"]
            row = decision_rows_by_ticker.get((position.ticker or "").strip().upper())
            if row is None:
                continue
            expected_return_pct = quantize_decimal(row.get(row_key), "0.01")
            current_value = quantize_decimal(position.current_value, "0.01") or ZERO
            if expected_return_pct is None or current_value <= ZERO:
                continue
            covered_positions += 1
            current_value_total += current_value
            weighted_expected_return += expected_return_pct * current_value
            expected_total_value += current_value * (Decimal("1") + (expected_return_pct / ONE_HUNDRED))

        horizon_return_pct = (
            quantize_decimal(weighted_expected_return / current_value_total, "0.01")
            if current_value_total > ZERO
            else None
        )
        horizon_rows.append(
            {
                "label": label,
                "months": months,
                "return_pct": horizon_return_pct,
                "projected_total_value": quantize_decimal(expected_total_value, "0.01") if current_value_total > ZERO else None,
                "positions_count": covered_positions,
                "tone": (
                    "good"
                    if horizon_return_pct is not None and horizon_return_pct >= ZERO
                    else ("warn" if horizon_return_pct is not None else "")
                ),
            }
        )
    return horizon_rows


def build_equity_analysis_overview(
    positions,
    history_cards: list[dict],
    decision_rows: list[dict],
    ibex_universe_summary: dict | None = None,
    *,
    selected_start_date: date | None = None,
    selected_end_date: date | None = None,
) -> dict:
    ibex_universe_summary = ibex_universe_summary or {}
    owned_positions = [position for position in positions if position.is_owned]
    watchlist_positions = [position for position in positions if not position.is_owned]
    current_value_total = sum((position.current_value for position in owned_positions), ZERO)
    invested_amount_total = sum((position.invested_amount for position in owned_positions), ZERO)
    annual_dividends_total = sum((position.annual_dividend_income for position in owned_positions), ZERO)
    net_dividends_total = sum((position.net_dividend_income for position in owned_positions), ZERO)
    annual_maintenance_total = sum((position.recurring_cost_used for position in owned_positions), ZERO)
    purchase_cost_total = sum((position.purchase_total_cost for position in owned_positions), ZERO)
    net_annual_income_total = sum((position.net_annual_income for position in owned_positions), ZERO)
    unrealized_gain_total = sum((position.unrealized_gain_after_costs for position in owned_positions), ZERO)
    comparable_summary = build_owned_positions_comparable_summary(history_cards)

    weighted_periods = []
    for label in ("1Y", "3Y", "5Y", "10Y"):
        weighted_stock = ZERO
        weighted_benchmark = ZERO
        weight_total = ZERO
        for card in history_cards:
            if not card["position"].is_owned:
                continue
            snapshot = next((item for item in card.get("period_snapshots", []) if item["label"] == label), None)
            if not snapshot or not snapshot["available"] or snapshot["stock_return_pct"] is None:
                continue
            weight = card["position"].current_value
            weighted_stock += snapshot["stock_return_pct"] * weight
            if snapshot["benchmark_return_pct"] is not None:
                weighted_benchmark += snapshot["benchmark_return_pct"] * weight
            weight_total += weight
        weighted_periods.append(
            {
                "label": label,
                "stock_return_pct": (weighted_stock / weight_total) if weight_total else None,
                "benchmark_return_pct": (weighted_benchmark / weight_total) if weight_total else None,
            }
        )

    latest_sync_at = max((position.last_synced_at for position in positions if position.last_synced_at), default=None)
    latest_price_date = max((position.latest_price_date for position in positions if position.latest_price_date), default=None)
    selected_cards = [
        card
        for card in history_cards
        if card["position"].is_owned and card.get("selected_period", {}).get("available")
    ]
    weighted_selected_return = None
    if selected_cards:
        numerator = sum(
            (card["selected_period"]["stock_return_pct"] or ZERO) * card["position"].current_value
            for card in selected_cards
        )
        denominator = sum((card["position"].current_value for card in selected_cards), ZERO)
        if denominator:
            weighted_selected_return = numerator / denominator

    projection_horizons = build_portfolio_projection_horizons(history_cards)
    expectation_horizons = build_portfolio_expectation_horizons(history_cards, decision_rows)
    weighted_projected_return_12m = next(
        (
            item["return_pct"]
            for item in projection_horizons
            if item.get("months") == 12 and item.get("return_pct") is not None
        ),
        None,
    )
    weighted_safety_score = None
    next_sale_recommendation = {"available": False}
    owned_projection_cards = [
        card
        for card in history_cards
        if card["position"].is_owned and card.get("projection", {}).get("available")
    ]
    if owned_projection_cards:
        projection_weight_total = sum((card["position"].current_value for card in owned_projection_cards), ZERO)
        if projection_weight_total:
            weighted_safety_score = sum(
                (card["projection"].get("safety_score") or ZERO) * card["position"].current_value
                for card in owned_projection_cards
            ) / projection_weight_total
    owned_trade_cards = [card for card in history_cards if card["position"].is_owned]
    next_sale_candidates = []
    for card in owned_trade_cards:
        try:
            trade_plan = build_owned_cycle_trade_timing_plan(card)
        except Exception:
            logger.exception(
                "No se pudo calcular la primera venta tactica para %s",
                card["position"].ticker,
            )
            continue
        if not trade_plan.get("available") or trade_plan.get("mode") not in {"sale_reentry", "sale_review"}:
            continue
        sale_month_number = trade_plan.get("sale_month_number")
        sale_window_label = trade_plan.get("sale_window_label")
        if sale_month_number is None or not sale_window_label:
            continue
        signal_value_pct = trade_plan.get("signal_value_pct")
        next_sale_candidates.append(
            {
                "available": True,
                "ticker": card["position"].ticker,
                "company_name": card["position"].company_name,
                "sale_month_number": sale_month_number,
                "sale_date": trade_plan.get("sale_date"),
                "sale_window_label": sale_window_label,
                "mode": trade_plan.get("mode"),
                "signal_value_pct": quantize_decimal(signal_value_pct, "0.01") if signal_value_pct is not None else None,
                "summary": trade_plan.get("summary") or "",
            }
        )
    if next_sale_candidates:
        next_sale_recommendation = min(
            next_sale_candidates,
            key=lambda item: (
                item.get("sale_date") or date.max,
                item.get("sale_month_number") or 999,
                item.get("signal_value_pct") if item.get("signal_value_pct") is not None else Decimal("9999"),
                item.get("company_name") or "",
            ),
        )

    return {
        "positions_count": len(positions),
        "owned_positions_count": len(owned_positions),
        "watchlist_positions_count": len(watchlist_positions),
        "invested_amount": invested_amount_total,
        "current_value": current_value_total,
        "annual_dividends_total": annual_dividends_total,
        "net_dividends_total": net_dividends_total,
        "annual_maintenance_total": annual_maintenance_total,
        "purchase_cost_total": purchase_cost_total,
        "net_annual_income_total": net_annual_income_total,
        "unrealized_gain_total": unrealized_gain_total,
        "unrealized_return_pct": (
            (unrealized_gain_total / (invested_amount_total + purchase_cost_total)) * ONE_HUNDRED
            if invested_amount_total + purchase_cost_total
            else None
        ),
        "latest_sync_at": latest_sync_at,
        "latest_price_date": latest_price_date,
        "weighted_selected_return": weighted_selected_return,
        "weighted_periods": weighted_periods,
        "projection_horizons": projection_horizons,
        "expectation_horizons": expectation_horizons,
        "weighted_projected_return_12m": weighted_projected_return_12m,
        "weighted_safety_score": weighted_safety_score,
        "selected_period_label": build_selected_period_label(selected_start_date, selected_end_date),
        "watchlist_latest_price_count": sum(1 for position in watchlist_positions if position.current_price_per_share),
        "best_decision": decision_rows[0] if decision_rows else ibex_universe_summary.get("top_pick"),
        "comparable_summary": comparable_summary,
        "next_sale_recommendation": next_sale_recommendation,
    }


def resolve_market_range_key_for_days(total_days: int) -> str:
    if total_days <= 45:
        return "3mo"
    if total_days <= 120:
        return "6mo"
    if total_days <= 370:
        return "1y"
    if total_days <= 740:
        return "2y"
    if total_days <= 1850:
        return "5y"
    if total_days <= LONG_ANALYSIS_DAYS:
        return DEFAULT_MARKET_RANGE_KEY
    return MAX_MARKET_RANGE_KEY


def project_return_pct_to_elapsed_days(
    return_pct: Decimal | None,
    elapsed_days: int,
    *,
    horizon_days: int = TRACKING_HORIZON_DAYS,
) -> Decimal | None:
    if return_pct is None or elapsed_days <= 0 or horizon_days <= 0:
        return None
    clipped_days = min(max(elapsed_days, 0), horizon_days)
    base = 1 + (float(return_pct) / 100)
    if base <= 0:
        return Decimal("-100.00")
    cumulative_return_pct = (base ** (clipped_days / horizon_days) - 1) * 100
    return Decimal(str(round(cumulative_return_pct, 4)))


def find_closest_market_point(points: list[dict] | None, target_date: date, tolerance_days: int = 10) -> dict | None:
    normalized_points = [
        point
        for point in list(points or [])
        if point.get("date") is not None and point.get("close") is not None
    ]
    if not normalized_points:
        return None

    nearby_points = [
        point
        for point in normalized_points
        if abs((point["date"] - target_date).days) <= tolerance_days
    ]
    if nearby_points:
        return min(
            nearby_points,
            key=lambda point: (
                abs((point["date"] - target_date).days),
                0 if point["date"] <= target_date else 1,
                point["date"],
            ),
        )

    before_points = [point for point in normalized_points if point["date"] <= target_date]
    if before_points:
        return max(before_points, key=lambda point: point["date"])

    after_points = [point for point in normalized_points if point["date"] >= target_date]
    if after_points:
        return min(after_points, key=lambda point: point["date"])
    return None


def calculate_regression_intercept(xs: list[Decimal], ys: list[Decimal], beta: Decimal | None) -> Decimal | None:
    if beta is None or len(xs) != len(ys) or len(xs) < EXPECTATION_REVIEW_CORRECTION_MIN_POINTS:
        return None
    mean_x = average_decimal(xs)
    mean_y = average_decimal(ys)
    if mean_x is None or mean_y is None:
        return None
    return quantize_decimal(mean_y - (beta * mean_x), "0.01")


def calculate_regression_r_squared(xs: list[Decimal], ys: list[Decimal]) -> Decimal | None:
    correlation = pearson_correlation(xs, ys)
    if correlation is None:
        return None
    return quantize_decimal(correlation * correlation, "0.01")


def build_expectation_review_company_context(
    *,
    ticker: str,
    reviews: list[EquityExpectationReview],
    market_series: MarketSeries | None,
    market_error: str,
    as_of: date,
) -> dict:
    horizon_preview_fields = (
        (12, "1A", "expected_return_pct_1y"),
        (24, "2A", "expected_return_pct_2y"),
        (36, "3A", "expected_return_pct_3y"),
        (48, "4A", "expected_return_pct_4y"),
        (60, "5A", "expected_return_pct_5y"),
    )

    def build_preview_horizon_markers(chart: dict) -> list[dict]:
        point_rows = list(chart.get("expected_points") or [])
        if not point_rows:
            return []
        markers = []
        total_points = len(point_rows)
        for index, point in enumerate(point_rows):
            markers.append(
                {
                    "x": point["x"],
                    "label": str(point.get("label") or point.get("date_label") or "").strip(),
                    "date": point.get("date_label") or "",
                    "y1": "202",
                    "y2": "208",
                    "text_y": "218",
                    "grid_y1": "18",
                    "grid_y2": "202",
                    "draw_grid": 0 < index < total_points - 1,
                    "show_label": True,
                    "is_major": index in {0, total_points - 1},
                    "anchor": "start" if index == 0 else ("end" if index == total_points - 1 else "middle"),
                }
            )
        return markers
    sorted_reviews = sorted(
        list(reviews or []),
        key=lambda row: (row.analysis_date, row.created_at, row.id or 0),
    )
    company_name = next((row.company_name for row in reversed(sorted_reviews) if row.company_name), ticker)
    market_points = list((market_series.points if market_series else []) or [])
    latest_price = quantize_decimal((market_series.latest_price if market_series else None), "0.0001")
    latest_price_date = market_series.latest_date if market_series else None
    comparison_rows = []

    for review in sorted_reviews:
        window_end_date = min(as_of, review.analysis_date + timedelta(days=TRACKING_HORIZON_DAYS))
        elapsed_days = max((window_end_date - review.analysis_date).days, 0)
        if elapsed_days <= 0:
            continue
        expected_return_pct = quantize_decimal(
            review.expected_return_pct_1y if review.expected_return_pct_1y is not None else review.projected_return_pct_1y,
            "0.01",
        )
        expected_progress_pct = quantize_decimal(
            project_return_pct_to_elapsed_days(expected_return_pct, elapsed_days),
            "0.01",
        )
        start_price = quantize_decimal(review.current_price, "0.0001")
        if start_price is None:
            start_point = find_closest_market_point(market_points, review.analysis_date)
            start_price = quantize_decimal(start_point.get("close") if start_point else None, "0.0001")
        end_point = find_closest_market_point(market_points, window_end_date)
        actual_price = quantize_decimal(end_point.get("close") if end_point else latest_price, "0.0001")
        actual_return_pct = quantize_decimal(percentage_change(actual_price, start_price), "0.01")
        gap_pct = (
            quantize_decimal(actual_return_pct - expected_progress_pct, "0.01")
            if actual_return_pct is not None and expected_progress_pct is not None
            else None
        )
        comparison_rows.append(
            {
                "analysis_date": review.analysis_date,
                "analysis_date_label": review.analysis_date.isoformat(),
                "window_end_date": window_end_date,
                "window_end_date_label": window_end_date.isoformat(),
                "elapsed_days": elapsed_days,
                "window_label": "1A cerrado" if elapsed_days >= TRACKING_HORIZON_DAYS else f"{elapsed_days} dias",
                "matured": elapsed_days >= TRACKING_HORIZON_DAYS,
                "expected_return_pct": expected_progress_pct,
                "actual_return_pct": actual_return_pct,
                "gap_pct": gap_pct,
                "current_price": actual_price,
                "current_price_date_label": end_point["date"].isoformat() if end_point and end_point.get("date") else "",
                "trade_alert_label": review.trade_alert_label,
                "review_kind": review.review_kind,
            }
        )

    calibration_rows = [
        row
        for row in comparison_rows
        if row.get("expected_return_pct") is not None
        and row.get("actual_return_pct") is not None
        and row.get("elapsed_days", 0) >= EXPECTATION_REVIEW_CORRECTION_MIN_ELAPSED_DAYS
    ]
    xs = [Decimal(str(row["expected_return_pct"])) for row in calibration_rows]
    ys = [Decimal(str(row["actual_return_pct"])) for row in calibration_rows]
    beta = quantize_decimal(calculate_regression_beta(xs, ys), "0.01")
    intercept = calculate_regression_intercept(xs, ys, beta)
    r_squared = calculate_regression_r_squared(xs, ys)

    for row in comparison_rows:
        corrected_return_pct = None
        expected_return_pct = row.get("expected_return_pct")
        if beta is not None and intercept is not None and expected_return_pct is not None:
            corrected_return_pct = quantize_decimal(intercept + (beta * Decimal(str(expected_return_pct))), "0.01")
        row["corrected_return_pct"] = corrected_return_pct
        row["corrected_gap_pct"] = (
            quantize_decimal((row.get("actual_return_pct") or ZERO) - corrected_return_pct, "0.01")
            if corrected_return_pct is not None and row.get("actual_return_pct") is not None
            else None
        )

    actual_series = [
        {"date": row["analysis_date"], "value": row["actual_return_pct"], "label": row["window_label"]}
        for row in comparison_rows
        if row.get("actual_return_pct") is not None
    ]
    expected_series = [
        {"date": row["analysis_date"], "value": row["expected_return_pct"], "label": row["window_label"]}
        for row in comparison_rows
        if row.get("expected_return_pct") is not None
    ]
    corrected_series = [
        {"date": row["analysis_date"], "value": row["corrected_return_pct"], "label": row["window_label"]}
        for row in comparison_rows
        if row.get("corrected_return_pct") is not None
    ]

    latest_review = sorted_reviews[-1] if sorted_reviews else None
    preview_mode = False
    preview_note = ""
    if len(comparison_rows) >= 2:
        chart = build_value_tracking_chart(
            actual_series,
            expected_series,
            corrected_series,
            value_suffix="%",
            axis_formatter=format_percentage_axis_value,
            time_marker_mode="month",
            grid_marker_mode="month",
        )
    else:
        preview_expected_series = []
        preview_corrected_series = []
        if latest_review is not None:
            preview_expected_series.append(
                {
                    "date": latest_review.analysis_date,
                    "value": ZERO,
                    "label": "Hoy",
                    "is_anchor": True,
                }
            )
            if beta is not None and intercept is not None:
                preview_corrected_series.append(
                    {
                        "date": latest_review.analysis_date,
                        "value": ZERO,
                        "label": "Hoy",
                        "is_anchor": True,
                    }
                )
            for months, label, field_name in horizon_preview_fields:
                expected_value = quantize_decimal(getattr(latest_review, field_name, None), "0.01")
                horizon_date = add_calendar_months(latest_review.analysis_date, months)
                if expected_value is None or horizon_date is None:
                    continue
                preview_expected_series.append(
                    {
                        "date": horizon_date,
                        "value": expected_value,
                        "label": label,
                        "is_anchor": True,
                    }
                )
                if beta is not None and intercept is not None:
                    preview_corrected_value = quantize_decimal(intercept + (beta * expected_value), "0.01")
                    preview_corrected_series.append(
                        {
                            "date": horizon_date,
                            "value": preview_corrected_value,
                            "label": label,
                            "is_anchor": True,
                        }
                    )
        chart = build_value_tracking_chart(
            [],
            preview_expected_series,
            preview_corrected_series,
            value_suffix="%",
            axis_formatter=format_percentage_axis_value,
            time_marker_mode="auto",
            grid_marker_mode="none",
            allow_expected_only=True,
        )
        preview_mode = bool(chart.get("available"))
        if preview_mode:
            chart["x_markers"] = build_preview_horizon_markers(chart)
            chart["grid_markers"] = [marker for marker in chart["x_markers"] if marker.get("draw_grid")]
            chart["scale_note"] = "Hitos directos desde la ultima revision guardada."
            preview_note = (
                "Todavia no hay realidad suficiente para comparar la bondad. "
                "Se muestra la ultima esperanza guardada con sus hitos 1A..5A."
            )

    average_expected_return_pct = average_decimal(
        [Decimal(str(row["expected_return_pct"])) for row in comparison_rows if row.get("expected_return_pct") is not None]
    )
    average_actual_return_pct = average_decimal(
        [Decimal(str(row["actual_return_pct"])) for row in comparison_rows if row.get("actual_return_pct") is not None]
    )
    average_gap_pct = average_decimal(
        [Decimal(str(row["gap_pct"])) for row in comparison_rows if row.get("gap_pct") is not None]
    )
    latest_row = comparison_rows[-1] if comparison_rows else None

    equation_available = beta is not None and intercept is not None
    if equation_available:
        if (average_gap_pct or ZERO) <= Decimal("-4.00"):
            interpretation = "El modelo ha venido demasiado optimista y la correccion lo enfria."
        elif (average_gap_pct or ZERO) >= Decimal("4.00"):
            interpretation = "El modelo ha venido demasiado prudente y la correccion deja mas margen."
        else:
            interpretation = "La lectura historica va razonablemente cerca de la realidad observada."
        formula_label = (
            f"Real observado ~= {intercept:+.1f} + {beta:.2f} x Esperanza observada"
        )
        short_formula_label = f"{intercept:+.1f} + {beta:.2f}x"
    else:
        if calibration_rows and (average_gap_pct or ZERO) >= Decimal("4.00"):
            interpretation = (
                "Senal provisional: la realidad reciente esta saliendo mejor que la esperanza. "
                f"La ecuacion formal necesita {EXPECTATION_REVIEW_CORRECTION_MIN_POINTS} muestras con recorrido, "
                "pero este valor no deberia tratarse como venta solo por una esperanza negativa temprana."
            )
            formula_label = "Modelo demasiado prudente"
            short_formula_label = "Prudente"
        elif calibration_rows and (average_gap_pct or ZERO) <= Decimal("-4.00"):
            interpretation = (
                "Senal provisional: la realidad reciente esta saliendo peor que la esperanza. "
                f"La ecuacion formal necesita {EXPECTATION_REVIEW_CORRECTION_MIN_POINTS} muestras con recorrido."
            )
            formula_label = "Modelo demasiado optimista"
            short_formula_label = "Optimista"
        else:
            interpretation = (
                f"Hacen falta al menos {EXPECTATION_REVIEW_CORRECTION_MIN_POINTS} revisiones con algo de recorrido "
                "para calcular la ecuacion correctora."
            )
            formula_label = "Sin muestras suficientes"
            short_formula_label = "Sin ecuacion"

    return {
        "available": bool(comparison_rows) or preview_mode,
        "ticker": ticker,
        "company_name": company_name,
        "tab_key": re.sub(r"[^a-z0-9]+", "-", f"bondad-{clean_ticker(ticker).lower()}").strip("-") or clean_ticker(ticker).lower(),
        "scope_label": "Revisiones programadas",
        "market_error": market_error,
        "current_price": latest_price,
        "current_price_date": latest_price_date,
        "current_price_date_label": latest_price_date.isoformat() if latest_price_date else "",
        "reviews_count": len(comparison_rows),
        "matured_reviews_count": sum(1 for row in comparison_rows if row.get("matured")),
        "average_expected_return_pct": quantize_decimal(average_expected_return_pct, "0.01"),
        "average_actual_return_pct": quantize_decimal(average_actual_return_pct, "0.01"),
        "average_gap_pct": quantize_decimal(average_gap_pct, "0.01"),
        "latest_row": latest_row,
        "rows": list(reversed(comparison_rows[-8:])),
        "chart": chart,
        "preview_mode": preview_mode,
        "preview_note": preview_note,
        "equation": {
            "available": equation_available,
            "beta": beta,
            "intercept": intercept,
            "r_squared": r_squared,
            "formula_label": formula_label,
            "short_formula_label": short_formula_label,
            "interpretation": interpretation,
            "sample_count": len(calibration_rows),
        },
    }


def build_expectation_review_dashboard(as_of: date | None = None) -> dict:
    as_of = as_of or django_timezone.localdate()

    try:
        queryset = (
            EquityExpectationReview.objects.filter(
                scope=EquityExpectationReview.Scope.IBEX,
                review_kind=EquityExpectationReview.ReviewKind.SCHEDULED,
            )
            .select_related("position")
            .order_by("company_name", "ticker", "analysis_date", "id")
        )
        review_rows = list(queryset)
        if not review_rows:
            review_rows = list(
                EquityExpectationReview.objects.filter(scope=EquityExpectationReview.Scope.IBEX)
                .select_related("position")
                .order_by("company_name", "ticker", "analysis_date", "id")
            )
    except (OperationalError, ProgrammingError):
        return {
            "available": False,
            "companies": [],
            "reviews_count": 0,
            "companies_count": 0,
            "equation_ready_count": 0,
            "last_review_date_label": "",
            "scope_note": "La tabla historica de bondad aun no esta disponible. Aplica la migracion para activar esta pestana.",
        }

    if not review_rows:
        return {
            "available": False,
            "companies": [],
            "reviews_count": 0,
            "companies_count": 0,
            "equation_ready_count": 0,
            "last_review_date_label": "",
            "scope_note": "Todavia no hay revisiones historicas guardadas para comparar la bondad del modelo.",
        }

    reviews_by_ticker: dict[str, list[EquityExpectationReview]] = defaultdict(list)
    for review in review_rows:
        reviews_by_ticker[clean_ticker(review.ticker)].append(review)

    grouped_rows = []
    series_map: dict[str, MarketSeries | None] = {}
    error_map: dict[str, str] = {}
    fetch_targets = {}
    for ticker, rows in reviews_by_ticker.items():
        earliest_review_date = min(row.analysis_date for row in rows)
        range_key = resolve_market_range_key_for_days(max((as_of - earliest_review_date).days + 14, 30))
        quote_symbol = next((row.quote_symbol for row in reversed(rows) if row.quote_symbol), ticker)
        fetch_targets[ticker] = (quote_symbol, range_key)

    max_workers = min(max(len(fetch_targets), 1), ibex_universe_max_workers())
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(fetch_market_series, quote_symbol, range_key=range_key): ticker
            for ticker, (quote_symbol, range_key) in fetch_targets.items()
        }
        for future in as_completed(future_map):
            ticker = future_map[future]
            try:
                series_map[ticker] = future.result()
                error_map[ticker] = ""
            except Exception as exc:
                series_map[ticker] = None
                error_map[ticker] = str(exc)

    for ticker, rows in sorted(reviews_by_ticker.items(), key=lambda item: (item[1][-1].company_name or item[0], item[0])):
        grouped_rows.append(
            build_expectation_review_company_context(
                ticker=ticker,
                reviews=rows,
                market_series=series_map.get(ticker),
                market_error=error_map.get(ticker, ""),
                as_of=as_of,
            )
        )

    available_rows = [row for row in grouped_rows if row.get("available")]
    last_review_date = max((review.analysis_date for review in review_rows), default=None)
    return {
        "available": bool(available_rows),
        "companies": available_rows,
        "reviews_count": len(review_rows),
        "companies_count": len(available_rows),
        "equation_ready_count": sum(1 for row in available_rows if (row.get("equation") or {}).get("available")),
        "last_review_date": last_review_date,
        "last_review_date_label": last_review_date.isoformat() if last_review_date else "",
        "scope_note": (
            "Cada punto compara la esperanza 1A guardada el martes o jueves frente a la realidad observada "
            "hasta hoy o hasta cumplir 365 dias, lo que ocurra antes."
        ),
        "equation_note": (
            "La ecuacion correctora sale de ajustar la realidad observada contra la esperanza observada "
            "para no fiarnos solo del modelo bruto."
        ),
    }


def extract_expectation_review_return_pct(review: EquityExpectationReview, *, horizon_years: int) -> Decimal | None:
    if horizon_years <= 1:
        return quantize_decimal(
            review.expected_return_pct_1y if review.expected_return_pct_1y is not None else review.projected_return_pct_1y,
            "0.01",
        )
    if horizon_years >= 5:
        accumulated_return_pct = quantize_decimal(
            review.expected_return_pct_5y if review.expected_return_pct_5y is not None else review.projected_return_pct_5y,
            "0.01",
        )
        if accumulated_return_pct is None:
            return None
        return quantize_decimal(annualize_return_pct(accumulated_return_pct, 60), "0.01")
    return None


def select_expectation_review_memory_rows(reviews: list[EquityExpectationReview]) -> list[EquityExpectationReview]:
    return list(reviews or [])[-OPTIMIZER_EXPECTATION_REVIEW_MEMORY_POINTS:]


def build_expectation_review_memory_weight(analysis_date: date, anchor_date: date | None) -> Decimal:
    if anchor_date is None:
        return Decimal("1.0000")
    age_days = max((anchor_date - analysis_date).days, 0)
    half_life_days = float(OPTIMIZER_EXPECTATION_REVIEW_MEMORY_HALF_LIFE_DAYS)
    raw_weight = Decimal(str(0.5 ** (age_days / half_life_days)))
    return quantize_decimal(
        clamp_decimal(raw_weight, OPTIMIZER_EXPECTATION_REVIEW_MIN_MEMORY_WEIGHT, Decimal("1.00")),
        "0.0001",
    ) or Decimal("1.0000")


def resolve_card_memory_market_points(card: dict | None) -> list[dict]:
    card = card or {}
    points = []
    for point in card.get("market_history_points") or []:
        point_date = point.get("date")
        close_price = quantize_decimal(point.get("close"), "0.0001")
        if point_date is None or close_price is None:
            continue
        points.append({"date": point_date, "close": close_price})
    if points:
        return sorted(points, key=lambda item: item["date"])

    position = card.get("position")
    if getattr(position, "pk", None):
        try:
            return [
                {
                    "date": row.price_date,
                    "close": quantize_decimal(row.close_price, "0.0001"),
                }
                for row in position.price_history.order_by("price_date")
                if row.price_date is not None and row.close_price is not None
            ]
        except (OperationalError, ProgrammingError):
            return []
    return []


def build_expectation_review_reality_feedback(
    reviews: list[EquityExpectationReview],
    *,
    horizon_years: int,
    current_price: Decimal | None = None,
    current_date: date | None = None,
    market_points: list[dict] | None = None,
) -> dict:
    current_price = quantize_decimal(current_price, "0.0001")
    if current_price in {None, ZERO} or current_date is None:
        return {"available": False, "sample_count": 0, "rows": []}

    market_points = [
        {
            "date": point.get("date"),
            "close": quantize_decimal(point.get("close"), "0.0001"),
        }
        for point in list(market_points or [])
        if point.get("date") is not None and point.get("close") is not None
    ]
    rows = []
    for review in select_expectation_review_memory_rows(reviews):
        start_price = quantize_decimal(review.current_price, "0.0001")
        if start_price in {None, ZERO}:
            start_point = find_closest_market_point(market_points, review.analysis_date)
            start_price = quantize_decimal(start_point.get("close") if start_point else None, "0.0001")
        if start_price in {None, ZERO}:
            continue
        elapsed_days = (current_date - review.analysis_date).days
        if elapsed_days < EXPECTATION_REVIEW_CORRECTION_MIN_ELAPSED_DAYS:
            continue
        observed_days = min(elapsed_days, TRACKING_HORIZON_DAYS)
        observed_date = review.analysis_date + timedelta(days=observed_days)
        actual_price = current_price
        actual_price_date = current_date
        if elapsed_days > TRACKING_HORIZON_DAYS:
            end_point = find_closest_market_point(market_points, observed_date)
            if not end_point:
                continue
            actual_price = quantize_decimal(end_point.get("close"), "0.0001")
            actual_price_date = end_point.get("date") or observed_date
        if actual_price in {None, ZERO}:
            continue
        expected_return_pct = extract_expectation_review_return_pct(review, horizon_years=horizon_years)
        if expected_return_pct is None:
            continue
        expected_progress_pct = quantize_decimal(
            project_return_pct_to_elapsed_days(expected_return_pct, observed_days),
            "0.01",
        )
        actual_return_pct = quantize_decimal(percentage_change(actual_price, start_price), "0.01")
        if expected_progress_pct is None or actual_return_pct is None:
            continue
        gap_pct = quantize_decimal(actual_return_pct - expected_progress_pct, "0.01")
        direction_hit = (
            (expected_progress_pct >= ZERO and actual_return_pct >= ZERO)
            or (expected_progress_pct < ZERO and actual_return_pct < ZERO)
        )
        elapsed_weight = clamp_decimal(
            Decimal(observed_days) / Decimal(TRACKING_EXPECTED_FEEDBACK_FULL_STRENGTH_DAYS),
            Decimal("0.25"),
            Decimal("1.00"),
        )
        memory_weight = quantize_decimal(
            build_expectation_review_memory_weight(review.analysis_date, current_date) * elapsed_weight,
            "0.0001",
        ) or Decimal("1.0000")
        rows.append(
            {
                "analysis_date": review.analysis_date,
                "elapsed_days": elapsed_days,
                "observed_days": observed_days,
                "observed_date": observed_date,
                "observed_date_label": observed_date.isoformat(),
                "expected_progress_pct": expected_progress_pct,
                "actual_return_pct": actual_return_pct,
                "actual_price": actual_price,
                "actual_price_date": actual_price_date,
                "actual_price_date_label": actual_price_date.isoformat() if actual_price_date else "",
                "gap_pct": gap_pct,
                "absolute_error_pct": abs(gap_pct),
                "direction_hit": direction_hit,
                "memory_weight": memory_weight,
            }
        )

    if not rows:
        return {"available": False, "sample_count": 0, "rows": []}

    average_gap_pct = quantize_decimal(
        weighted_average_decimal([(row["gap_pct"], row["memory_weight"]) for row in rows]),
        "0.01",
    )
    recent_rows = rows[-OPTIMIZER_EXPECTATION_REVIEW_RECENT_POINTS:]
    recent_average_gap_pct = quantize_decimal(
        weighted_average_decimal([(row["gap_pct"], row["memory_weight"]) for row in recent_rows]),
        "0.01",
    )
    latest_gap_pct = rows[-1]["gap_pct"]
    mean_absolute_error_pct = quantize_decimal(
        weighted_average_decimal([(row["absolute_error_pct"], row["memory_weight"]) for row in rows]),
        "0.01",
    )
    average_actual_return_pct = quantize_decimal(
        weighted_average_decimal([(row["actual_return_pct"], row["memory_weight"]) for row in rows]),
        "0.01",
    )
    positive_actual_ratio_pct = quantize_decimal(
        (Decimal(sum(1 for row in rows if row["actual_return_pct"] > ZERO)) * ONE_HUNDRED) / Decimal(len(rows)),
        "0.01",
    )
    direction_hit_rate_pct = quantize_decimal(
        weighted_average_decimal(
            [
                (ONE_HUNDRED if row["direction_hit"] else ZERO, row["memory_weight"])
                for row in rows
            ]
        ),
        "0.01",
    )
    positive_feedback_cap = (
        Decimal("6.00")
        if len(rows) >= 2 and (recent_average_gap_pct or ZERO) >= Decimal("4.00")
        else Decimal("3.00")
    )
    bias_adjustment_pct = clamp_decimal(
        ((average_gap_pct or ZERO) * Decimal("0.50"))
        + ((recent_average_gap_pct or ZERO) * Decimal("0.30"))
        + ((latest_gap_pct or ZERO) * Decimal("0.20")),
        Decimal("-4.50"),
        positive_feedback_cap,
    )
    if (average_gap_pct or ZERO) >= Decimal("2.00"):
        model_bias_label = "Demasiado prudente"
    elif (average_gap_pct or ZERO) <= Decimal("-2.00"):
        model_bias_label = "Demasiado optimista"
    else:
        model_bias_label = "Ajustado"
    return {
        "available": True,
        "sample_count": len(rows),
        "recent_sample_count": len(recent_rows),
        "average_gap_pct": average_gap_pct,
        "recent_average_gap_pct": recent_average_gap_pct,
        "latest_gap_pct": latest_gap_pct,
        "average_actual_return_pct": average_actual_return_pct,
        "latest_actual_return_pct": rows[-1]["actual_return_pct"],
        "positive_actual_ratio_pct": positive_actual_ratio_pct,
        "mean_absolute_error_pct": mean_absolute_error_pct,
        "direction_hit_rate_pct": direction_hit_rate_pct,
        "bias_adjustment_pct": quantize_decimal(bias_adjustment_pct, "0.01"),
        "model_bias_label": model_bias_label,
        "latest_elapsed_days": rows[-1]["elapsed_days"],
        "oldest_analysis_date": rows[0]["analysis_date"],
        "oldest_analysis_date_label": rows[0]["analysis_date"].isoformat(),
        "memory_span_days": (rows[-1]["analysis_date"] - rows[0]["analysis_date"]).days,
        "rows": [
            {
                **row,
                "analysis_date_label": row["analysis_date"].isoformat(),
            }
            for row in rows[-OPTIMIZER_EXPECTATION_REVIEW_RECENT_POINTS:]
        ],
    }


def build_optimizer_expectation_review_horizon_signal(
    reviews: list[EquityExpectationReview],
    *,
    horizon_years: int,
    current_price: Decimal | None = None,
    current_date: date | None = None,
    market_points: list[dict] | None = None,
) -> dict:
    value_rows = []
    memory_reviews = select_expectation_review_memory_rows(reviews)
    anchor_date = current_date or (memory_reviews[-1].analysis_date if memory_reviews else None)
    for review in memory_reviews:
        expected_return_pct = extract_expectation_review_return_pct(review, horizon_years=horizon_years)
        if expected_return_pct is None:
            continue
        memory_weight = build_expectation_review_memory_weight(review.analysis_date, anchor_date)
        value_rows.append(
            {
                "analysis_date": review.analysis_date,
                "value": expected_return_pct,
                "memory_weight": memory_weight,
            }
        )
    reality_feedback = build_expectation_review_reality_feedback(
        reviews,
        horizon_years=horizon_years,
        current_price=current_price,
        current_date=current_date,
        market_points=market_points,
    )
    if not value_rows:
        return {
            "available": False,
            "sample_count": 0,
            "latest_return_pct": None,
            "average_return_pct": None,
            "trend_return_pct": None,
            "recent_delta_pct": None,
            "spread_pct": None,
            "historical_spread_pct": None,
            "reality_feedback": reality_feedback,
        }

    values = [Decimal(str(row["value"])) for row in value_rows]
    recent_values = values[-OPTIMIZER_EXPECTATION_REVIEW_RECENT_POINTS:]
    latest_return_pct = quantize_decimal(values[-1], "0.01")
    average_return_pct = quantize_decimal(
        weighted_average_decimal([(Decimal(str(row["value"])), row["memory_weight"]) for row in value_rows]),
        "0.01",
    )
    recent_average_return_pct = quantize_decimal(average_decimal(recent_values), "0.01")
    trend_return_pct = quantize_decimal(values[-1] - values[0], "0.01") if len(values) >= 2 else None
    recent_delta_pct = quantize_decimal(values[-1] - values[-2], "0.01") if len(values) >= 2 else None
    spread_pct = quantize_decimal(max(recent_values) - min(recent_values), "0.01") if recent_values else None
    historical_spread_pct = quantize_decimal(max(values) - min(values), "0.01") if values else None
    date_span_days = (
        (value_rows[-1]["analysis_date"] - value_rows[0]["analysis_date"]).days
        if len(value_rows) >= 2
        else 0
    )
    return {
        "available": True,
        "horizon_years": horizon_years,
        "sample_count": len(value_rows),
        "recent_sample_count": min(len(value_rows), OPTIMIZER_EXPECTATION_REVIEW_RECENT_POINTS),
        "latest_return_pct": latest_return_pct,
        "average_return_pct": average_return_pct,
        "recent_average_return_pct": recent_average_return_pct,
        "trend_return_pct": trend_return_pct,
        "recent_delta_pct": recent_delta_pct,
        "spread_pct": spread_pct,
        "historical_spread_pct": historical_spread_pct,
        "first_review_date": value_rows[0]["analysis_date"],
        "latest_review_date": value_rows[-1]["analysis_date"],
        "anchor_date": anchor_date,
        "date_span_days": date_span_days,
        "reality_feedback": reality_feedback,
    }


def build_optimizer_expectation_review_signal(
    reviews: list[EquityExpectationReview],
    *,
    current_price: Decimal | None = None,
    current_date: date | None = None,
    market_points: list[dict] | None = None,
) -> dict:
    sorted_reviews = sorted(
        list(reviews or []),
        key=lambda row: (row.analysis_date, row.created_at, row.id or 0),
    )
    if not sorted_reviews:
        return {"available": False}

    scheduled_reviews = [row for row in sorted_reviews if row.review_kind == EquityExpectationReview.ReviewKind.SCHEDULED]
    source_reviews = scheduled_reviews or sorted_reviews
    source_label = "programadas" if scheduled_reviews else "historicas"
    one_year_signal = build_optimizer_expectation_review_horizon_signal(
        source_reviews,
        horizon_years=1,
        current_price=current_price,
        current_date=current_date,
        market_points=market_points,
    )
    five_year_signal = build_optimizer_expectation_review_horizon_signal(
        source_reviews,
        horizon_years=5,
        current_price=current_price,
        current_date=current_date,
        market_points=market_points,
    )
    latest_review = source_reviews[-1]
    return {
        "available": bool(one_year_signal.get("available") or five_year_signal.get("available")),
        "source_label": source_label,
        "sample_count": max(int(one_year_signal.get("sample_count") or 0), int(five_year_signal.get("sample_count") or 0)),
        "latest_analysis_date": latest_review.analysis_date,
        "latest_analysis_date_label": latest_review.analysis_date.isoformat(),
        "1y": one_year_signal,
        "5y": five_year_signal,
    }


def resolve_card_memory_price_and_date(card: dict | None) -> tuple[Decimal | None, date | None]:
    card = card or {}
    position = card.get("position")
    projection = card.get("projection") or {}
    current_price = quantize_decimal(
        projection.get("latest_price") or getattr(position, "current_price_per_share", None),
        "0.0001",
    )
    current_date = (
        projection.get("latest_date")
        or card.get("end_date")
        or getattr(position, "latest_price_date", None)
        or django_timezone.localdate()
    )
    if isinstance(current_date, str):
        try:
            current_date = date.fromisoformat(current_date)
        except ValueError:
            current_date = django_timezone.localdate()
    return current_price, current_date


def build_optimizer_expectation_review_signal_map(
    tickers: list[str] | tuple[str, ...] | set[str],
    *,
    current_cards_by_ticker: dict[str, dict] | None = None,
) -> dict[str, dict]:
    normalized_tickers = sorted({clean_ticker(ticker) for ticker in tickers or [] if clean_ticker(ticker)})
    if not normalized_tickers:
        return {}

    try:
        review_rows = list(
            EquityExpectationReview.objects.filter(ticker__in=normalized_tickers)
            .order_by("ticker", "analysis_date", "id")
        )
    except (OperationalError, ProgrammingError):
        return {}

    grouped_reviews: dict[str, list[EquityExpectationReview]] = defaultdict(list)
    for review in review_rows:
        grouped_reviews[clean_ticker(review.ticker)].append(review)

    signal_map = {}
    current_cards_by_ticker = current_cards_by_ticker or {}
    for ticker in normalized_tickers:
        current_card = current_cards_by_ticker.get(ticker)
        current_price, current_date = resolve_card_memory_price_and_date(current_card)
        signal_map[ticker] = build_optimizer_expectation_review_signal(
            grouped_reviews.get(ticker, []),
            current_price=current_price,
            current_date=current_date,
            market_points=resolve_card_memory_market_points(current_card),
        )
    return signal_map


def resolve_card_optimizer_expectation_review_signal(card: dict) -> dict:
    signal = card.get("expectation_review_signal")
    if isinstance(signal, dict):
        return signal
    position = card.get("position")
    ticker = clean_ticker(getattr(position, "ticker", "") or card.get("ticker") or "")
    if not ticker:
        return {"available": False}
    return build_optimizer_expectation_review_signal_map(
        [ticker],
        current_cards_by_ticker={ticker: card},
    ).get(ticker, {"available": False})


def expectation_stability_move_cap(horizon_years: int, elapsed_days: int | None) -> Decimal:
    elapsed_days = max(int(elapsed_days or 1), 1)
    if int(horizon_years or 1) >= 5:
        return clamp_decimal(
            EXPECTATION_STABILITY_5Y_DAILY_MOVE_PCT * Decimal(elapsed_days),
            EXPECTATION_STABILITY_5Y_MIN_MOVE_PCT,
            EXPECTATION_STABILITY_5Y_MAX_MOVE_PCT,
        )
    return clamp_decimal(
        EXPECTATION_STABILITY_1Y_DAILY_MOVE_PCT * Decimal(elapsed_days),
        EXPECTATION_STABILITY_1Y_MIN_MOVE_PCT,
        EXPECTATION_STABILITY_1Y_MAX_MOVE_PCT,
    )


def stabilize_return_against_latest_expectation(
    raw_return_pct: Decimal | None,
    horizon_signal: dict | None,
) -> dict:
    raw_return_pct = quantize_decimal(raw_return_pct, "0.01")
    if raw_return_pct is None:
        return {
            "available": False,
            "raw_return_pct": None,
            "stabilized_return_pct": None,
            "adjustment_pct": None,
            "applied": False,
        }

    signal = horizon_signal or {}
    latest_return_pct = quantize_decimal(signal.get("latest_return_pct"), "0.01")
    if not signal.get("available") or latest_return_pct is None:
        return {
            "available": False,
            "raw_return_pct": raw_return_pct,
            "stabilized_return_pct": raw_return_pct,
            "adjustment_pct": ZERO,
            "applied": False,
        }

    anchor_date = signal.get("anchor_date")
    latest_review_date = signal.get("latest_review_date")
    elapsed_days = None
    if isinstance(anchor_date, date) and isinstance(latest_review_date, date):
        elapsed_days = max((anchor_date - latest_review_date).days, 1)
    max_move_pct = expectation_stability_move_cap(int(signal.get("horizon_years") or 1), elapsed_days)
    delta_pct = quantize_decimal(raw_return_pct - latest_return_pct, "0.01") or ZERO
    if abs(delta_pct) <= max_move_pct:
        return {
            "available": True,
            "raw_return_pct": raw_return_pct,
            "latest_return_pct": latest_return_pct,
            "stabilized_return_pct": raw_return_pct,
            "adjustment_pct": ZERO,
            "delta_pct": delta_pct,
            "max_move_pct": quantize_decimal(max_move_pct, "0.01"),
            "elapsed_days": elapsed_days,
            "applied": False,
        }

    bounded_delta_pct = max_move_pct if delta_pct > ZERO else -max_move_pct
    stabilized_return_pct = quantize_decimal(latest_return_pct + bounded_delta_pct, "0.01")
    adjustment_pct = quantize_decimal(stabilized_return_pct - raw_return_pct, "0.01")
    return {
        "available": True,
        "raw_return_pct": raw_return_pct,
        "latest_return_pct": latest_return_pct,
        "stabilized_return_pct": stabilized_return_pct,
        "adjustment_pct": adjustment_pct,
        "delta_pct": delta_pct,
        "max_move_pct": quantize_decimal(max_move_pct, "0.01"),
        "elapsed_days": elapsed_days,
        "applied": True,
        "note": (
            "La expectativa se amortigua frente a la ultima revision guardada para evitar cambios bruscos "
            "sin varias sesiones de confirmacion."
        ),
    }


def apply_optimizer_expectation_review_adjustment(
    raw_return_pct: Decimal | None,
    horizon_signal: dict | None,
) -> dict:
    raw_return_pct = quantize_decimal(raw_return_pct, "0.01")
    if raw_return_pct is None:
        return {
            "available": False,
            "raw_return_pct": None,
            "adjusted_return_pct": None,
            "adjustment_pct": None,
        }

    signal = horizon_signal or {}
    if not signal.get("available"):
        return {
            "available": False,
            "raw_return_pct": raw_return_pct,
            "adjusted_return_pct": raw_return_pct,
            "adjustment_pct": ZERO,
        }

    sample_count = int(signal.get("sample_count") or 0)
    horizon_years = int(signal.get("horizon_years") or 0)
    date_span_days = int(signal.get("date_span_days") or 0)
    if horizon_years >= 5 and (sample_count < 4 or date_span_days < 21):
        return {
            "available": True,
            "deferred_for_stability": True,
            "sample_count": sample_count,
            "date_span_days": date_span_days,
            "raw_return_pct": raw_return_pct,
            "adjusted_return_pct": raw_return_pct,
            "adjustment_pct": ZERO,
            "note": "El ajuste 5A se mantiene sin cambios hasta tener al menos 4 lecturas y 21 dias de separacion.",
        }
    latest_return_pct = quantize_decimal(signal.get("latest_return_pct"), "0.01")
    average_return_pct = quantize_decimal(signal.get("average_return_pct"), "0.01")
    trend_return_pct = quantize_decimal(signal.get("trend_return_pct"), "0.01") or ZERO
    recent_delta_pct = quantize_decimal(signal.get("recent_delta_pct"), "0.01") or ZERO
    spread_pct = quantize_decimal(signal.get("spread_pct"), "0.01") or ZERO
    reality_feedback = signal.get("reality_feedback") or {}
    actual_feedback_adjustment_pct = (
        quantize_decimal(reality_feedback.get("bias_adjustment_pct"), "0.01")
        if reality_feedback.get("available")
        else ZERO
    ) or ZERO
    reality_error_penalty_pct = ZERO
    reality_contradiction_floor_pct = None
    if reality_feedback.get("available"):
        mean_absolute_error_pct = quantize_decimal(reality_feedback.get("mean_absolute_error_pct"), "0.01") or ZERO
        direction_hit_rate_pct = quantize_decimal(reality_feedback.get("direction_hit_rate_pct"), "0.01")
        average_gap_pct = quantize_decimal(reality_feedback.get("average_gap_pct"), "0.01") or ZERO
        recent_average_gap_pct = quantize_decimal(reality_feedback.get("recent_average_gap_pct"), "0.01") or ZERO
        latest_gap_pct = quantize_decimal(reality_feedback.get("latest_gap_pct"), "0.01") or ZERO
        average_actual_return_pct = quantize_decimal(reality_feedback.get("average_actual_return_pct"), "0.01") or ZERO
        positive_actual_ratio_pct = quantize_decimal(reality_feedback.get("positive_actual_ratio_pct"), "0.01") or ZERO
        model_too_optimistic = (
            ((average_gap_pct * Decimal("0.50")) + (recent_average_gap_pct * Decimal("0.30")) + (latest_gap_pct * Decimal("0.20")))
            < ZERO
        )
        if model_too_optimistic:
            reality_error_penalty_pct = clamp_decimal(
                max(mean_absolute_error_pct - Decimal("1.25"), ZERO) * Decimal("0.12"),
                ZERO,
                Decimal("1.40"),
            )
            if direction_hit_rate_pct is not None and direction_hit_rate_pct < Decimal("50.00"):
                reality_error_penalty_pct += Decimal("0.45")
        if (
            int(reality_feedback.get("sample_count") or 0) >= 2
            and raw_return_pct < ZERO
            and average_gap_pct >= Decimal("5.00")
            and recent_average_gap_pct >= Decimal("3.00")
            and average_actual_return_pct > ZERO
            and positive_actual_ratio_pct >= Decimal("60.00")
        ):
            reality_contradiction_floor_pct = quantize_decimal(
                clamp_decimal(
                    average_actual_return_pct * Decimal("0.15"),
                    Decimal("0.50"),
                    Decimal("3.00"),
                ),
                "0.01",
            )
    anchor_return_pct = quantize_decimal(
        average_decimal([item for item in (latest_return_pct, average_return_pct) if item is not None]),
        "0.01",
    ) or latest_return_pct or average_return_pct or raw_return_pct
    blend_weight = Decimal("0.18")
    if sample_count >= 4:
        blend_weight = Decimal("0.38")
    elif sample_count >= 2:
        blend_weight = Decimal("0.28")
    anchor_adjustment_pct = (anchor_return_pct - raw_return_pct) * blend_weight
    trend_adjustment_pct = (
        max(trend_return_pct, ZERO) * Decimal("0.08")
        + min(trend_return_pct, ZERO) * Decimal("0.16")
        + max(recent_delta_pct, ZERO) * Decimal("0.05")
        + min(recent_delta_pct, ZERO) * Decimal("0.10")
    )
    stability_penalty_pct = clamp_decimal(
        max(spread_pct - OPTIMIZER_EXPECTATION_REVIEW_STABLE_SPREAD_PCT, ZERO) * Decimal("0.06"),
        ZERO,
        Decimal("1.80"),
    )
    adjustment_pct = clamp_decimal(
        anchor_adjustment_pct
        + trend_adjustment_pct
        + actual_feedback_adjustment_pct
        - stability_penalty_pct
        - reality_error_penalty_pct,
        -OPTIMIZER_EXPECTATION_REVIEW_MAX_PENALTY_PCT,
        OPTIMIZER_EXPECTATION_REVIEW_MAX_BONUS_PCT,
    )
    adjusted_return_pct = clamp_decimal(
        raw_return_pct + adjustment_pct,
        Decimal("-80.00"),
        Decimal("140.00"),
    )
    stability = stabilize_return_against_latest_expectation(adjusted_return_pct, signal)
    if stability.get("applied") and stability.get("stabilized_return_pct") is not None:
        adjusted_return_pct = stability["stabilized_return_pct"]
        adjustment_pct = quantize_decimal(adjusted_return_pct - raw_return_pct, "0.01") or ZERO
    reality_contradiction_applied = False
    if reality_contradiction_floor_pct is not None and adjusted_return_pct < reality_contradiction_floor_pct:
        adjusted_return_pct = reality_contradiction_floor_pct
        adjustment_pct = quantize_decimal(adjusted_return_pct - raw_return_pct, "0.01") or ZERO
        reality_contradiction_applied = True
    return {
        "available": True,
        "sample_count": sample_count,
        "raw_return_pct": raw_return_pct,
        "anchor_return_pct": quantize_decimal(anchor_return_pct, "0.01"),
        "latest_return_pct": latest_return_pct,
        "average_return_pct": average_return_pct,
        "trend_return_pct": quantize_decimal(trend_return_pct, "0.01"),
        "recent_delta_pct": quantize_decimal(recent_delta_pct, "0.01"),
        "spread_pct": quantize_decimal(spread_pct, "0.01"),
        "blend_weight_pct": quantize_decimal(blend_weight * ONE_HUNDRED, "0.1"),
        "anchor_adjustment_pct": quantize_decimal(anchor_adjustment_pct, "0.01"),
        "trend_adjustment_pct": quantize_decimal(trend_adjustment_pct, "0.01"),
        "actual_feedback_adjustment_pct": quantize_decimal(actual_feedback_adjustment_pct, "0.01"),
        "stability_penalty_pct": quantize_decimal(stability_penalty_pct, "0.01"),
        "reality_error_penalty_pct": quantize_decimal(reality_error_penalty_pct, "0.01"),
        "reality_contradiction_floor_pct": reality_contradiction_floor_pct,
        "reality_contradiction_applied": reality_contradiction_applied,
        "adjustment_pct": quantize_decimal(adjustment_pct, "0.01"),
        "adjusted_return_pct": quantize_decimal(adjusted_return_pct, "0.01"),
        "stability": stability,
        "reality_feedback": reality_feedback,
    }


def apply_expectation_review_memory_to_card(card: dict, signal: dict | None = None) -> dict:
    projection = card.get("projection") or {}
    if not projection.get("available"):
        return card
    if (projection.get("historical_memory_adjustment") or {}).get("applied"):
        return card

    position = card.get("position")
    ticker = clean_ticker(getattr(position, "ticker", "") or "")
    if signal is None:
        signal = build_optimizer_expectation_review_signal_map(
            [ticker],
            current_cards_by_ticker={ticker: card} if ticker else {},
        ).get(ticker, {"available": False})
    card["expectation_review_signal"] = signal or {"available": False}
    if not (signal or {}).get("available"):
        return card

    projection_adjustment = apply_optimizer_expectation_review_adjustment(
        projection.get("base_return_pct"),
        signal.get("1y"),
    )
    adjustment_pct = quantize_decimal(projection_adjustment.get("adjustment_pct"), "0.01") or ZERO
    if abs(adjustment_pct) < Decimal("0.05"):
        return card

    latest_price, latest_date = resolve_card_memory_price_and_date(card)
    if latest_price in {None, ZERO}:
        return card

    net_income_yield_pct = quantize_decimal(projection.get("net_income_yield_pct"), "0.01") or ZERO
    transaction_drag_pct = quantize_decimal(projection.get("transaction_drag_pct"), "0.01") or ZERO
    raw_base_return_pct = quantize_decimal(projection.get("base_return_pct"), "0.01")
    raw_price_return_pct = quantize_decimal(projection.get("price_return_pct"), "0.01")
    if raw_price_return_pct is None and raw_base_return_pct is not None:
        raw_price_return_pct = raw_base_return_pct - net_income_yield_pct + transaction_drag_pct
    adjusted_base_return_pct = projection_adjustment.get("adjusted_return_pct")
    if adjusted_base_return_pct is None:
        return card
    return_shift_pct = adjusted_base_return_pct - (raw_base_return_pct or ZERO)
    adjusted_price_return_pct = (raw_price_return_pct or ZERO) + return_shift_pct

    raw_low_return_pct = quantize_decimal(projection.get("low_return_pct"), "0.01")
    raw_high_return_pct = quantize_decimal(projection.get("high_return_pct"), "0.01")
    raw_price_low_return_pct = quantize_decimal(projection.get("price_low_return_pct"), "0.01")
    raw_price_high_return_pct = quantize_decimal(projection.get("price_high_return_pct"), "0.01")
    reality_feedback = projection_adjustment.get("reality_feedback") or {}
    error_pct = quantize_decimal(reality_feedback.get("mean_absolute_error_pct"), "0.01") or ZERO
    spread_pct = quantize_decimal(projection_adjustment.get("spread_pct"), "0.01") or ZERO
    band_multiplier = clamp_decimal(
        Decimal("1.00")
        + (max(error_pct - Decimal("2.50"), ZERO) * Decimal("0.015"))
        + (max(spread_pct - OPTIMIZER_EXPECTATION_REVIEW_STABLE_SPREAD_PCT, ZERO) * Decimal("0.006")),
        Decimal("1.00"),
        Decimal("1.18"),
    )

    def shift_bound(bound_value: Decimal | None, base_value: Decimal | None) -> Decimal | None:
        if bound_value is None:
            return None
        if base_value is None:
            return bound_value + return_shift_pct
        distance = bound_value - base_value
        return adjusted_base_return_pct + (distance * band_multiplier)

    adjusted_low_return_pct = shift_bound(raw_low_return_pct, raw_base_return_pct)
    adjusted_high_return_pct = shift_bound(raw_high_return_pct, raw_base_return_pct)
    adjusted_price_low_return_pct = (
        adjusted_low_return_pct - net_income_yield_pct + transaction_drag_pct
        if adjusted_low_return_pct is not None
        else (raw_price_low_return_pct + return_shift_pct if raw_price_low_return_pct is not None else None)
    )
    adjusted_price_high_return_pct = (
        adjusted_high_return_pct - net_income_yield_pct + transaction_drag_pct
        if adjusted_high_return_pct is not None
        else (raw_price_high_return_pct + return_shift_pct if raw_price_high_return_pct is not None else None)
    )
    confidence_score = projection.get("confidence_score_pct") or projection_reliability_score(projection.get("confidence_label") or "Baja")
    reliability_penalty = clamp_decimal(
        max(error_pct - Decimal("1.00"), ZERO) * Decimal("0.35")
        + max(spread_pct - Decimal("8.00"), ZERO) * Decimal("0.12"),
        ZERO,
        Decimal("6.00"),
    )
    adjusted_confidence_score = clamp_decimal(confidence_score - (reliability_penalty * Decimal("0.55")), Decimal("18.00"), Decimal("92.00"))
    adjusted_confidence_label = confidence_label_from_score(adjusted_confidence_score)
    adjusted_safety_score = clamp_decimal((projection.get("safety_score") or Decimal("55.00")) - (reliability_penalty * Decimal("0.40")), Decimal("18.00"), Decimal("92.00"))
    adjusted_safety_label = safety_label_from_score(adjusted_safety_score)

    projection["base_return_pct"] = quantize_decimal(adjusted_base_return_pct)
    projection["price_return_pct"] = quantize_decimal(adjusted_price_return_pct)
    projection["low_return_pct"] = quantize_decimal(adjusted_low_return_pct)
    projection["high_return_pct"] = quantize_decimal(adjusted_high_return_pct)
    projection["price_low_return_pct"] = quantize_decimal(adjusted_price_low_return_pct)
    projection["price_high_return_pct"] = quantize_decimal(adjusted_price_high_return_pct)
    projection["projected_price"] = quantize_decimal(project_price_from_return(latest_price, adjusted_price_return_pct), "0.0001")
    projection["low_price"] = quantize_decimal(project_price_from_return(latest_price, adjusted_price_low_return_pct), "0.0001")
    projection["high_price"] = quantize_decimal(project_price_from_return(latest_price, adjusted_price_high_return_pct), "0.0001")
    projection["monthly_path"] = build_monthly_projection_path(
        latest_price,
        adjusted_price_return_pct,
        anchor_date=latest_date,
        cycle_phase=projection.get("cycle_phase") or "Transicion",
    )
    projection["quarterly_path"] = build_quarterly_projection_path_from_monthly_path(projection["monthly_path"])
    if not projection["quarterly_path"]:
        projection["quarterly_path"] = build_projection_path(
            latest_price,
            adjusted_price_return_pct,
            anchor_date=latest_date,
            cycle_phase=projection.get("cycle_phase") or "Transicion",
        )
    projection["confidence_score_pct"] = quantize_decimal(adjusted_confidence_score)
    projection["confidence_label"] = adjusted_confidence_label
    projection["safety_score"] = quantize_decimal(adjusted_safety_score)
    projection["safety_label"] = adjusted_safety_label
    projection["scenarios"] = build_one_year_projection_scenarios(
        latest_price,
        price_return_pct=adjusted_price_return_pct,
        price_low_return_pct=adjusted_price_low_return_pct,
        price_high_return_pct=adjusted_price_high_return_pct,
        base_return_pct=adjusted_base_return_pct,
        low_return_pct=adjusted_low_return_pct,
        high_return_pct=adjusted_high_return_pct,
        confidence_label=adjusted_confidence_label,
    )
    projection["historical_memory_adjustment"] = {
        "applied": True,
        "raw_return_pct": raw_base_return_pct,
        "adjusted_return_pct": quantize_decimal(adjusted_base_return_pct),
        "adjustment_pct": quantize_decimal(adjustment_pct),
        "sample_count": projection_adjustment.get("sample_count"),
        "latest_return_pct": projection_adjustment.get("latest_return_pct"),
        "average_return_pct": projection_adjustment.get("average_return_pct"),
        "recent_delta_pct": projection_adjustment.get("recent_delta_pct"),
        "spread_pct": projection_adjustment.get("spread_pct"),
        "actual_feedback_adjustment_pct": projection_adjustment.get("actual_feedback_adjustment_pct"),
        "reality_error_penalty_pct": projection_adjustment.get("reality_error_penalty_pct"),
        "reality_contradiction_floor_pct": projection_adjustment.get("reality_contradiction_floor_pct"),
        "reality_contradiction_applied": projection_adjustment.get("reality_contradiction_applied"),
        "reality_feedback": reality_feedback,
        "note": "Ajuste aplicado con el historico completo de revisiones guardadas y sus desviaciones reales observadas.",
    }
    projection["explanation"] = (
        f"{str(projection.get('explanation') or '').strip()} "
        "La prevision se templa con el historico de revisiones y con la desviacion real observada."
    ).strip()
    card["projection"] = projection

    reliability = card.get("projection_reliability") or {"label": "Baja", "score": Decimal("40.00")}
    reliability_score = reliability.get("score") or projection_reliability_score(reliability.get("label") or "Baja")
    reliability["score"] = quantize_decimal(clamp_decimal(reliability_score - reliability_penalty, Decimal("18.00"), Decimal("92.00")))
    reliability["label"] = confidence_label_from_score(reliability["score"])
    reliability["historical_memory_adjustment"] = {
        "applied": True,
        "reliability_penalty_pct": quantize_decimal(reliability_penalty),
        "note": projection["historical_memory_adjustment"]["note"],
    }
    card["projection_reliability"] = reliability

    cycle_projection = card.get("cycle_projection_5y") or {}
    cycle_adjustment = apply_optimizer_expectation_review_adjustment(
        cycle_projection.get("annual_return_pct"),
        signal.get("5y") if signal.get("available") else None,
    )
    cycle_adjustment_pct = quantize_decimal(cycle_adjustment.get("adjustment_pct"), "0.01") or ZERO
    if cycle_projection.get("available") and abs(cycle_adjustment_pct) >= Decimal("0.05"):
        adjusted_annual_return_pct = clamp_decimal(
            (cycle_projection.get("annual_return_pct") or ZERO) + cycle_adjustment_pct,
            Decimal("-12.00"),
            Decimal("18.00"),
        )
        scenario_spread_annual_pct = clamp_decimal(
            (cycle_projection.get("scenario_spread_annual_pct") or Decimal("3.50")) * band_multiplier,
            Decimal("2.00"),
            Decimal("9.50"),
        )
        cycle_path, adjusted_steps, step_shift = build_cycle_projection_path_for_target(
            latest_price,
            annual_return_pct=adjusted_annual_return_pct,
            step_return_pcts=[Decimal(str(value)) for value in (cycle_projection.get("step_return_pcts") or [])],
            annualized_volatility_pct=cycle_projection.get("annualized_volatility_pct"),
            current_drawdown_pct=cycle_projection.get("current_drawdown_pct"),
            cycle_phase=cycle_projection.get("cycle_phase") or "Transicion",
            anchor_date=cycle_projection.get("latest_date") or latest_date,
            years=5,
            step_months=6,
        )
        if cycle_path:
            projected_price = cycle_path[-1]["projected_price"]
            cycle_projection["annual_return_pct"] = quantize_decimal(adjusted_annual_return_pct)
            cycle_projection["five_year_return_pct"] = quantize_decimal(percentage_change(projected_price, latest_price))
            cycle_projection["projected_price"] = quantize_decimal(projected_price, "0.0001")
            cycle_projection["path"] = cycle_path
            cycle_projection["step_return_pcts"] = [quantize_decimal(value) or ZERO for value in adjusted_steps]
            cycle_projection["scenario_spread_annual_pct"] = quantize_decimal(scenario_spread_annual_pct)
            cycle_projection["scenarios"] = build_five_year_projection_scenarios(
                latest_price,
                latest_date=cycle_projection.get("latest_date") or latest_date,
                annual_return_pct=adjusted_annual_return_pct,
                scenario_spread_annual_pct=scenario_spread_annual_pct,
                step_return_pcts=adjusted_steps,
                annualized_volatility_pct=cycle_projection.get("annualized_volatility_pct"),
                current_drawdown_pct=cycle_projection.get("current_drawdown_pct"),
                cycle_phase=cycle_projection.get("cycle_phase") or "Transicion",
                confidence_label=adjusted_confidence_label,
            )
            cycle_projection["historical_memory_adjustment"] = {
                "applied": True,
                "raw_return_pct": cycle_adjustment.get("raw_return_pct"),
                "adjusted_return_pct": cycle_adjustment.get("adjusted_return_pct"),
                "adjustment_pct": cycle_adjustment.get("adjustment_pct"),
                "sample_count": cycle_adjustment.get("sample_count"),
                "latest_return_pct": cycle_adjustment.get("latest_return_pct"),
                "average_return_pct": cycle_adjustment.get("average_return_pct"),
                "step_shift": quantize_decimal(step_shift),
                "reality_feedback": cycle_adjustment.get("reality_feedback"),
                "note": projection["historical_memory_adjustment"]["note"],
            }
            card["cycle_projection_5y"] = cycle_projection

    synchronize_projection_path_with_cycle_zoom(
        projection,
        card.get("cycle_projection_5y") or {},
        current_price=latest_price,
        anchor_date=latest_date,
    )
    one_year_snapshot = next(
        (snapshot for snapshot in (card.get("period_snapshots") or []) if snapshot.get("label") == "1Y"),
        {"available": False},
    )
    card["trade_alert"] = build_trade_alert(
        position,
        projection,
        card.get("correlation") or {},
        card.get("projection_reliability") or {},
        card.get("relative_trend") or {},
        card.get("six_month_snapshot") or {"available": False},
        one_year_snapshot,
        valuation=card.get("valuation") or {},
        technical_signal=card.get("technical_signal") or {},
    )
    keep_visuals = bool(
        ((card.get("projection_12m_chart") or {}).get("available"))
        or ((card.get("cycle_projection_5y_chart") or {}).get("available"))
    )
    return refresh_card_projection_visuals(card, include_visuals=keep_visuals)


def rebuild_card_trade_alert_from_projection(card: dict) -> dict:
    position = card.get("position")
    projection = card.get("projection") or {}
    if position is None or not projection.get("available"):
        return card
    one_year_snapshot = next(
        (snapshot for snapshot in (card.get("period_snapshots") or []) if snapshot.get("label") == "1Y"),
        {"available": False},
    )
    card["trade_alert"] = build_trade_alert(
        position,
        projection,
        card.get("correlation") or {},
        card.get("projection_reliability") or {},
        card.get("relative_trend") or {},
        card.get("six_month_snapshot") or {"available": False},
        one_year_snapshot,
        valuation=card.get("valuation") or {},
        technical_signal=card.get("technical_signal") or {},
    )
    return card


def apply_projection_expectation_shift(card: dict, shift_pct: Decimal | None) -> bool:
    shift_pct = quantize_decimal(shift_pct, "0.01")
    projection = card.get("projection") or {}
    if shift_pct is None or abs(shift_pct) < Decimal("0.05") or not projection.get("available"):
        return False

    latest_price, latest_date = resolve_card_memory_price_and_date(card)
    if latest_price in {None, ZERO}:
        return False

    def shifted(field_name: str) -> Decimal | None:
        value = quantize_decimal(projection.get(field_name), "0.01")
        return quantize_decimal(value + shift_pct, "0.01") if value is not None else None

    adjusted_base_return_pct = shifted("base_return_pct")
    adjusted_price_return_pct = shifted("price_return_pct")
    adjusted_low_return_pct = shifted("low_return_pct")
    adjusted_high_return_pct = shifted("high_return_pct")
    adjusted_price_low_return_pct = shifted("price_low_return_pct")
    adjusted_price_high_return_pct = shifted("price_high_return_pct")
    if adjusted_base_return_pct is None or adjusted_price_return_pct is None:
        return False

    projection["base_return_pct"] = clamp_decimal(adjusted_base_return_pct, Decimal("-80.00"), Decimal("140.00"))
    projection["price_return_pct"] = clamp_decimal(adjusted_price_return_pct, Decimal("-80.00"), Decimal("140.00"))
    projection["low_return_pct"] = clamp_decimal(adjusted_low_return_pct, Decimal("-80.00"), Decimal("140.00")) if adjusted_low_return_pct is not None else None
    projection["high_return_pct"] = clamp_decimal(adjusted_high_return_pct, Decimal("-80.00"), Decimal("140.00")) if adjusted_high_return_pct is not None else None
    projection["price_low_return_pct"] = clamp_decimal(adjusted_price_low_return_pct, Decimal("-80.00"), Decimal("140.00")) if adjusted_price_low_return_pct is not None else None
    projection["price_high_return_pct"] = clamp_decimal(adjusted_price_high_return_pct, Decimal("-80.00"), Decimal("140.00")) if adjusted_price_high_return_pct is not None else None
    projection["projected_price"] = quantize_decimal(project_price_from_return(latest_price, projection["price_return_pct"]), "0.0001")
    projection["low_price"] = quantize_decimal(project_price_from_return(latest_price, projection.get("price_low_return_pct")), "0.0001")
    projection["high_price"] = quantize_decimal(project_price_from_return(latest_price, projection.get("price_high_return_pct")), "0.0001")
    projection["monthly_path"] = build_monthly_projection_path(
        latest_price,
        projection["price_return_pct"],
        anchor_date=latest_date,
        cycle_phase=projection.get("cycle_phase") or "Transicion",
    )
    projection["quarterly_path"] = build_quarterly_projection_path_from_monthly_path(projection["monthly_path"])
    if not projection["quarterly_path"]:
        projection["quarterly_path"] = build_projection_path(
            latest_price,
            projection["price_return_pct"],
            anchor_date=latest_date,
            cycle_phase=projection.get("cycle_phase") or "Transicion",
        )
    projection["decision_score"] = quantize_decimal(
        projection["base_return_pct"]
        * projection_confidence_multiplier(projection.get("confidence_label") or "Baja")
        * ((projection.get("safety_score") or Decimal("55.00")) / ONE_HUNDRED),
        "0.01",
    )
    projection["scenarios"] = build_one_year_projection_scenarios(
        latest_price,
        price_return_pct=projection.get("price_return_pct"),
        price_low_return_pct=projection.get("price_low_return_pct"),
        price_high_return_pct=projection.get("price_high_return_pct"),
        base_return_pct=projection.get("base_return_pct"),
        low_return_pct=projection.get("low_return_pct"),
        high_return_pct=projection.get("high_return_pct"),
        confidence_label=projection.get("confidence_label") or "Baja",
    )
    projection["expectation_stability_shift_pct"] = shift_pct
    card["projection"] = projection
    rebuild_card_trade_alert_from_projection(card)
    return True


def apply_cycle_expectation_target(card: dict, target_five_year_return_pct: Decimal | None) -> bool:
    target_five_year_return_pct = quantize_decimal(target_five_year_return_pct, "0.01")
    cycle_projection = card.get("cycle_projection_5y") or {}
    if target_five_year_return_pct is None or not cycle_projection.get("available"):
        return False

    latest_price, latest_date = resolve_card_memory_price_and_date(card)
    if latest_price in {None, ZERO}:
        return False

    current_five_year_return_pct = quantize_decimal(cycle_projection.get("five_year_return_pct"), "0.01")
    if current_five_year_return_pct is None:
        current_five_year_return_pct = quantize_decimal(
            (build_scenario_expectation_table(
                cycle_projection.get("scenarios"),
                return_key="five_year_return_pct",
                fallback_value=None,
            ) or {}).get("expected_return_pct"),
            "0.01",
        )
    if current_five_year_return_pct is None:
        return False

    shift_pct = quantize_decimal(target_five_year_return_pct - current_five_year_return_pct, "0.01")
    if shift_pct is None or abs(shift_pct) < Decimal("0.05"):
        return False

    target_annual_return_pct = annualize_return_pct(target_five_year_return_pct, 60)
    if target_annual_return_pct is None:
        return False
    adjusted_annual_return_pct = clamp_decimal(target_annual_return_pct, Decimal("-12.00"), Decimal("18.00"))
    scenario_spread_annual_pct = cycle_projection.get("scenario_spread_annual_pct") or Decimal("3.50")
    cycle_path, adjusted_steps, step_shift = build_cycle_projection_path_for_target(
        latest_price,
        annual_return_pct=adjusted_annual_return_pct,
        step_return_pcts=[Decimal(str(value)) for value in (cycle_projection.get("step_return_pcts") or [])],
        annualized_volatility_pct=cycle_projection.get("annualized_volatility_pct"),
        current_drawdown_pct=cycle_projection.get("current_drawdown_pct"),
        cycle_phase=cycle_projection.get("cycle_phase") or "Transicion",
        anchor_date=cycle_projection.get("latest_date") or latest_date,
        years=5,
        step_months=6,
    )
    if not cycle_path:
        return False

    projected_price = cycle_path[-1]["projected_price"]
    cycle_projection["annual_return_pct"] = quantize_decimal(adjusted_annual_return_pct)
    cycle_projection["five_year_return_pct"] = quantize_decimal(percentage_change(projected_price, latest_price))
    cycle_projection["projected_price"] = quantize_decimal(projected_price, "0.0001")
    cycle_projection["path"] = cycle_path
    cycle_projection["step_return_pcts"] = [quantize_decimal(value) or ZERO for value in adjusted_steps]
    cycle_projection["expectation_stability_step_shift"] = quantize_decimal(step_shift)
    cycle_projection["expectation_stability_shift_pct"] = shift_pct
    cycle_projection["scenarios"] = build_five_year_projection_scenarios(
        latest_price,
        latest_date=cycle_projection.get("latest_date") or latest_date,
        annual_return_pct=adjusted_annual_return_pct,
        scenario_spread_annual_pct=scenario_spread_annual_pct,
        step_return_pcts=adjusted_steps,
        annualized_volatility_pct=cycle_projection.get("annualized_volatility_pct"),
        current_drawdown_pct=cycle_projection.get("current_drawdown_pct"),
        cycle_phase=cycle_projection.get("cycle_phase") or "Transicion",
        confidence_label=(card.get("projection") or {}).get("confidence_label") or "Baja",
    )
    card["cycle_projection_5y"] = cycle_projection
    synchronize_projection_path_with_cycle_zoom(
        card.get("projection") or {},
        cycle_projection,
        current_price=latest_price,
        anchor_date=latest_date,
    )
    rebuild_card_trade_alert_from_projection(card)
    return True


def apply_expectation_stability_to_card(card: dict, signal: dict | None = None) -> dict:
    position = card.get("position")
    projection = card.get("projection") or {}
    if position is None or not projection.get("available"):
        return card

    ticker = clean_ticker(getattr(position, "ticker", "") or "")
    if signal is None:
        signal = build_optimizer_expectation_review_signal_map(
            [ticker],
            current_cards_by_ticker={ticker: card} if ticker else {},
        ).get(ticker, {"available": False})
    if not (signal or {}).get("available"):
        return card

    scenario_tables = build_card_scenario_tables(card)
    one_year_expected = (scenario_tables.get("projection_12m") or {}).get("expected_return_pct")
    one_year_stability = stabilize_return_against_latest_expectation(one_year_expected, signal.get("1y"))
    five_year_expectations = build_cycle_horizon_expectation_map(
        (card.get("cycle_projection_5y") or {}).get("scenarios"),
        current_price=getattr(position, "current_price_per_share", None),
    )
    five_year_expected = (
        five_year_expectations.get(60)
        or (scenario_tables.get("cycle_5y") or {}).get("expected_return_pct")
    )
    five_year_stability = stabilize_return_against_latest_expectation(five_year_expected, signal.get("5y"))
    one_year_applied = apply_projection_expectation_shift(card, one_year_stability.get("adjustment_pct"))
    five_year_applied = (
        apply_cycle_expectation_target(card, five_year_stability.get("stabilized_return_pct"))
        if five_year_stability.get("applied")
        else False
    )
    if not (one_year_applied or five_year_applied):
        return card

    card["expectation_stability"] = {
        "applied": True,
        "one_year": one_year_stability,
        "five_year": five_year_stability,
        "note": "Se limita el cambio entre revisiones para que la tesis se confirme durante varias sesiones antes de girar.",
    }
    keep_visuals = bool(
        ((card.get("projection_12m_chart") or {}).get("available"))
        or ((card.get("cycle_projection_5y_chart") or {}).get("available"))
    )
    return refresh_card_projection_visuals(card, include_visuals=keep_visuals)


def apply_expectation_stability_to_cards(cards: list[dict]) -> dict:
    cards = [card for card in cards or [] if card.get("position") is not None]
    tickers = [clean_ticker(getattr(card["position"], "ticker", "")) for card in cards]
    cards_by_ticker = {
        clean_ticker(getattr(card["position"], "ticker", "")): card
        for card in cards
        if clean_ticker(getattr(card["position"], "ticker", ""))
    }
    signal_map = build_optimizer_expectation_review_signal_map(
        tickers,
        current_cards_by_ticker=cards_by_ticker,
    )
    adjusted_count = 0
    for card in cards:
        ticker = clean_ticker(getattr(card["position"], "ticker", ""))
        before = dict(card.get("expectation_stability") or {})
        apply_expectation_stability_to_card(card, signal=signal_map.get(ticker, {"available": False}))
        after = card.get("expectation_stability") or {}
        if after.get("applied") and not before.get("applied"):
            adjusted_count += 1
    return {
        "available": bool(signal_map),
        "cards_count": len(cards),
        "adjusted_cards_count": adjusted_count,
    }


def apply_expectation_stability_to_dashboard(dashboard: dict) -> dict:
    history_cards = list(dashboard.get("history_cards") or [])
    ibex_cards = list(dashboard.get("ibex_universe_cards") or [])
    summary = apply_expectation_stability_to_cards([*history_cards, *ibex_cards])

    if history_cards:
        dashboard["decision_rows"] = build_equity_decision_rows(history_cards)
        dashboard["optimizer_cards"] = build_optimizer_master_cards(history_cards, ibex_cards)
        positions = [*(dashboard.get("owned_positions") or []), *(dashboard.get("watchlist_positions") or [])]
        if positions and dashboard.get("overview") is not None:
            dashboard["overview"] = build_equity_analysis_overview(
                positions,
                history_cards,
                dashboard["decision_rows"],
                dashboard.get("ibex_universe_summary") or {},
            )
    if ibex_cards:
        dashboard["ibex_universe_rows"] = build_equity_decision_rows(ibex_cards)
        ibex_summary = dashboard.get("ibex_universe_summary") or {}
        ibex_rows = dashboard["ibex_universe_rows"]
        ibex_summary.update(
            {
                "buy_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Comprar"),
                "sell_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Vender"),
                "watch_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Vigilar"),
                "top_pick": ibex_rows[0] if ibex_rows else ibex_summary.get("top_pick"),
            }
        )
        dashboard["ibex_universe_summary"] = ibex_summary
    dashboard["expectation_stability_summary"] = summary
    return summary


def apply_expectation_review_memory_to_cards(cards: list[dict]) -> dict:
    cards = [card for card in cards or [] if card.get("position") is not None]
    tickers = [clean_ticker(getattr(card["position"], "ticker", "")) for card in cards]
    cards_by_ticker = {
        clean_ticker(getattr(card["position"], "ticker", "")): card
        for card in cards
        if clean_ticker(getattr(card["position"], "ticker", ""))
    }
    signal_map = build_optimizer_expectation_review_signal_map(
        tickers,
        current_cards_by_ticker=cards_by_ticker,
    )
    adjusted_count = 0
    reality_feedback_count = 0
    for card in cards:
        ticker = clean_ticker(getattr(card["position"], "ticker", ""))
        signal = signal_map.get(ticker, {"available": False})
        card["expectation_review_signal"] = signal
        if ((signal.get("1y") or {}).get("reality_feedback") or {}).get("available"):
            reality_feedback_count += 1
        was_adjusted = bool(((card.get("projection") or {}).get("historical_memory_adjustment") or {}).get("applied"))
        apply_expectation_review_memory_to_card(card, signal=signal)
        is_adjusted = bool(((card.get("projection") or {}).get("historical_memory_adjustment") or {}).get("applied"))
        if is_adjusted and not was_adjusted:
            adjusted_count += 1
    return {
        "available": bool(signal_map),
        "cards_count": len(cards),
        "adjusted_cards_count": adjusted_count,
        "reality_feedback_count": reality_feedback_count,
    }


def build_equity_analysis_dashboard(
    positions,
    selected_start_date: date | None = None,
    selected_end_date: date | None = None,
    include_ibex_universe: bool = False,
    ibex_company_limit: int | None = None,
    ibex_progress_callback=None,
    ibex_include_visuals: bool = False,
    ibex_include_reference_suggestions: bool = False,
    ibex_include_fundamentals: bool = False,
) -> dict:
    reference_cache: dict = {}
    history_cards = build_equity_history_cards(
        positions,
        selected_start_date=selected_start_date,
        selected_end_date=selected_end_date,
        reference_cache=reference_cache,
    )
    owned_positions = [position for position in positions if position.is_owned]
    watchlist_positions = [position for position in positions if not position.is_owned]
    ibex_universe = {
        "cards": [],
        "rows": [],
        "summary": {
            "available": False,
            "analyzed_count": 0,
            "buy_alert_count": 0,
            "sell_alert_count": 0,
            "watch_alert_count": 0,
            "failed_count": 0,
            "failures": [],
            "broker_assumption": "",
            "trade_channel_label": "",
            "top_pick": None,
        },
    }
    if include_ibex_universe:
        ibex_universe = build_ibex_universe_analysis(
            history_cards,
            positions,
            selected_start_date=selected_start_date,
            selected_end_date=selected_end_date,
            reference_cache=reference_cache,
            company_limit=ibex_company_limit,
            progress_callback=ibex_progress_callback,
            include_visuals=ibex_include_visuals,
            include_reference_suggestions=ibex_include_reference_suggestions,
            include_fundamentals=ibex_include_fundamentals,
        )
    expectation_memory_summary = apply_expectation_review_memory_to_cards(
        [*history_cards, *(ibex_universe.get("cards") or [])]
    )
    reference_guide = build_equity_reference_guide(history_cards)
    decision_rows = build_equity_decision_rows(history_cards)
    if ibex_universe.get("cards"):
        ibex_rows = build_equity_decision_rows(ibex_universe["cards"])
        ibex_universe["rows"] = ibex_rows
        summary = ibex_universe.get("summary") or {}
        summary.update(
            {
                "buy_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Comprar"),
                "sell_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Vender"),
                "watch_alert_count": sum(1 for row in ibex_rows if row.get("trade_alert_label") == "Vigilar"),
                "top_pick": ibex_rows[0] if ibex_rows else summary.get("top_pick"),
            }
        )
        ibex_universe["summary"] = summary
    overview = build_equity_analysis_overview(
        positions,
        history_cards,
        decision_rows,
        ibex_universe["summary"],
        selected_start_date=selected_start_date,
        selected_end_date=selected_end_date,
    )

    return {
        "overview": overview,
        "history_cards": history_cards,
        "owned_positions": owned_positions,
        "watchlist_positions": watchlist_positions,
        "owned_history_cards": [card for card in history_cards if card["position"].is_owned],
        "watchlist_history_cards": [card for card in history_cards if not card["position"].is_owned],
        "decision_rows": decision_rows,
        "ibex_universe_cards": ibex_universe["cards"],
        "ibex_universe_rows": ibex_universe["rows"],
        "ibex_universe_summary": ibex_universe["summary"],
        "optimizer_cards": build_optimizer_master_cards(history_cards, ibex_universe["cards"]),
        "reference_guide_rows": reference_guide["rows"],
        "tracked_reference_rows": reference_guide["tracked_rows"],
        "reference_guide_summary": reference_guide["summary"],
        "expectation_memory_summary": expectation_memory_summary,
    }


def projection_confidence_multiplier(confidence_label: str) -> Decimal:
    if confidence_label == "Alta":
        return Decimal("1.00")
    if confidence_label == "Media":
        return Decimal("0.85")
    return Decimal("0.70")


def projection_reliability_score(label: str) -> Decimal:
    score_map = {
        "Alta": Decimal("82.00"),
        "Media": Decimal("64.00"),
        "Baja": Decimal("42.00"),
        "Sin historico suficiente": Decimal("45.00"),
    }
    return score_map.get(label or "Baja", Decimal("42.00"))


OPTIMIZER_STRATEGY_12M_PRIMARY = "12m_primary"
OPTIMIZER_STRATEGY_5Y_PRIMARY = "5y_primary"
OPTIMIZER_STRATEGIES = {
    OPTIMIZER_STRATEGY_12M_PRIMARY: {
        "label": "12M principal",
        "description": "Prioriza la prevision a 12 meses y usa el ciclo a 5 anos como contraste.",
        "primary_horizon_label": "12 meses",
        "secondary_horizon_label": "5 anos",
        "primary_weight": Decimal("0.78"),
        "secondary_weight": Decimal("0.22"),
    },
    OPTIMIZER_STRATEGY_5Y_PRIMARY: {
        "label": "5A principal",
        "description": "Prioriza el ciclo a 5 anos y usa la prevision a 12 meses como contraste de timing.",
        "primary_horizon_label": "5 anos",
        "secondary_horizon_label": "12 meses",
        "primary_weight": Decimal("0.65"),
        "secondary_weight": Decimal("0.35"),
    },
}
OPTIMIZER_MIN_SAFETY_SCORE = Decimal("55.00")
OPTIMIZER_MIN_RELIABILITY_SCORE = Decimal("58.00")
OPTIMIZER_DECISION_ACTION_ADJUSTMENTS = {
    "Priorizar": Decimal("1.80"),
    "Seguir": Decimal("0.75"),
    "Mantener": Decimal("0.75"),
    "Vigilar": Decimal("-1.85"),
    "Esperar": Decimal("-3.60"),
    "Reducir riesgo": Decimal("-4.40"),
}


def get_optimizer_strategy_config(strategy_mode: str | None = None) -> dict:
    normalized_mode = str(strategy_mode or OPTIMIZER_STRATEGY_12M_PRIMARY).strip().lower()
    strategy = OPTIMIZER_STRATEGIES.get(normalized_mode)
    if strategy is None:
        normalized_mode = OPTIMIZER_STRATEGY_12M_PRIMARY
        strategy = OPTIMIZER_STRATEGIES[normalized_mode]
    return {"mode": normalized_mode, **strategy}


def build_optimizer_quality_adjustments(
    safety_score: Decimal | None,
    reliability_score: Decimal | None,
    decision_action_label: str | None,
) -> dict:
    safety_score = safety_score or ZERO
    reliability_score = reliability_score or ZERO
    decision_action_label = str(decision_action_label or "").strip()
    safety_gap = max(OPTIMIZER_MIN_SAFETY_SCORE - safety_score, ZERO)
    reliability_gap = max(OPTIMIZER_MIN_RELIABILITY_SCORE - reliability_score, ZERO)
    safety_floor_penalty_pct = clamp_decimal(
        safety_gap * Decimal("0.11"),
        ZERO,
        Decimal("5.20"),
    )
    reliability_floor_penalty_pct = clamp_decimal(
        reliability_gap * Decimal("0.09"),
        ZERO,
        Decimal("4.20"),
    )
    quality_floor_penalty_pct = clamp_decimal(
        safety_floor_penalty_pct + reliability_floor_penalty_pct,
        ZERO,
        Decimal("8.20"),
    )
    decision_action_adjustment = OPTIMIZER_DECISION_ACTION_ADJUSTMENTS.get(
        decision_action_label,
        Decimal("-2.75"),
    )
    return {
        "safety_gap": quantize_decimal(safety_gap),
        "reliability_gap": quantize_decimal(reliability_gap),
        "safety_floor_penalty_pct": quantize_decimal(safety_floor_penalty_pct),
        "reliability_floor_penalty_pct": quantize_decimal(reliability_floor_penalty_pct),
        "quality_floor_penalty_pct": quantize_decimal(quality_floor_penalty_pct),
        "decision_action_adjustment": quantize_decimal(decision_action_adjustment),
    }


def build_optimizer_scenario_summary(
    scenarios: list[dict] | None,
    *,
    return_key: str,
    fallback_value: Decimal | None = None,
) -> dict:
    normalized_rows = []
    for row in list(scenarios or []):
        value = row.get(return_key)
        probability_pct = row.get("probability_pct")
        if value is None:
            continue
        normalized_rows.append(
            {
                "key": str(row.get("key") or "").strip().lower(),
                "label": str(row.get("label") or "").strip(),
                "value": Decimal(str(value)),
                "probability_pct": Decimal(str(probability_pct or "0")),
            }
        )

    if not normalized_rows:
        return {
            "available": False,
            "expected_return_pct": fallback_value,
            "base_return_pct": fallback_value,
            "worst_return_pct": fallback_value,
            "best_return_pct": fallback_value,
            "spread_pct": ZERO,
            "bear_probability_pct": None,
            "base_probability_pct": None,
            "bull_probability_pct": None,
        }

    total_probability_pct = sum((item["probability_pct"] for item in normalized_rows), ZERO)
    if total_probability_pct <= ZERO:
        total_probability_pct = ONE_HUNDRED
    expected_return_pct = sum(
        ((item["value"] * item["probability_pct"]) / total_probability_pct for item in normalized_rows),
        ZERO,
    )
    base_row = next((item for item in normalized_rows if item["key"] == "base"), None)
    bear_row = next((item for item in normalized_rows if item["key"] == "bear"), None)
    bull_row = next((item for item in normalized_rows if item["key"] == "bull"), None)
    base_return_pct = base_row["value"] if base_row is not None else fallback_value if fallback_value is not None else expected_return_pct
    worst_return_pct = min((item["value"] for item in normalized_rows), default=base_return_pct)
    best_return_pct = max((item["value"] for item in normalized_rows), default=base_return_pct)
    spread_pct = best_return_pct - worst_return_pct if best_return_pct is not None and worst_return_pct is not None else ZERO
    return {
        "available": True,
        "expected_return_pct": quantize_decimal(expected_return_pct),
        "base_return_pct": quantize_decimal(base_return_pct),
        "worst_return_pct": quantize_decimal(worst_return_pct),
        "best_return_pct": quantize_decimal(best_return_pct),
        "spread_pct": quantize_decimal(spread_pct),
        "bear_probability_pct": quantize_decimal(bear_row["probability_pct"], "0.1") if bear_row is not None else None,
        "base_probability_pct": quantize_decimal(base_row["probability_pct"], "0.1") if base_row is not None else None,
        "bull_probability_pct": quantize_decimal(bull_row["probability_pct"], "0.1") if bull_row is not None else None,
    }


def build_optimizer_uncertainty_profile(
    card: dict,
    projection: dict,
    cycle_projection: dict,
    external_signal: dict,
) -> dict:
    projection_news_adjustment = projection.get("news_adjustment") or {}
    cycle_news_adjustment = cycle_projection.get("news_adjustment") or {}
    news_context = card.get("news_context") or {}
    external_signal_score = Decimal(str(external_signal.get("score", "0") or "0"))
    external_signal_items_count = int(external_signal.get("items_count") or 0)
    model_confidence_penalty_pct = Decimal(str(projection_news_adjustment.get("confidence_penalty_pct") or "0"))
    model_band_multiplier = Decimal(str(projection_news_adjustment.get("band_multiplier") or "1"))
    model_spread_multiplier = Decimal(str(cycle_news_adjustment.get("spread_multiplier") or "1"))
    material_event = bool(
        news_context.get("material_event")
        or projection_news_adjustment.get("applied")
        or cycle_news_adjustment.get("applied")
    )
    model_uncertainty_penalty_pct = clamp_decimal(
        (model_confidence_penalty_pct * Decimal("0.16"))
        + (max(model_band_multiplier - Decimal("1.00"), ZERO) * Decimal("5.50"))
        + (max(model_spread_multiplier - Decimal("1.00"), ZERO) * Decimal("4.50")),
        ZERO,
        Decimal("4.50"),
    )
    external_signal_penalty_pct = clamp_decimal(
        abs(min(external_signal_score, ZERO)) * Decimal("0.42"),
        ZERO,
        Decimal("3.00"),
    )
    coverage_penalty_pct = clamp_decimal(
        Decimal(str(external_signal_items_count)) * Decimal("0.12"),
        ZERO,
        Decimal("0.75"),
    ) if external_signal_score < ZERO else ZERO
    uncertainty_penalty_pct = clamp_decimal(
        model_uncertainty_penalty_pct
        + external_signal_penalty_pct
        + coverage_penalty_pct
        + (Decimal("0.85") if material_event else ZERO),
        ZERO,
        Decimal("6.50"),
    )
    return {
        "material_event": material_event,
        "model_uncertainty_penalty_pct": quantize_decimal(model_uncertainty_penalty_pct),
        "external_signal_penalty_pct": quantize_decimal(external_signal_penalty_pct),
        "coverage_penalty_pct": quantize_decimal(coverage_penalty_pct),
        "uncertainty_penalty_pct": quantize_decimal(uncertainty_penalty_pct),
    }


def score_optimizer_memory_quality(expectation_review_signal: dict | None) -> dict:
    signal = expectation_review_signal or {}
    feedback = ((signal.get("1y") or {}).get("reality_feedback") or {})
    sample_count = int(feedback.get("sample_count") or 0)
    if not feedback.get("available"):
        return {
            "score": Decimal("58.00"),
            "sample_count": sample_count,
            "label": "Sin memoria suficiente",
            "reason": "Aun hay poca historia propia para premiar o penalizar el modelo.",
        }

    mean_absolute_error_pct = quantize_decimal(feedback.get("mean_absolute_error_pct"), "0.01") or ZERO
    average_gap_pct = quantize_decimal(feedback.get("average_gap_pct"), "0.01") or ZERO
    direction_hit_rate_pct = quantize_decimal(feedback.get("direction_hit_rate_pct"), "0.01") or Decimal("50.00")
    score = clamp_decimal(
        Decimal("68.00")
        + (min(Decimal(sample_count), Decimal("18")) * Decimal("0.75"))
        + ((direction_hit_rate_pct - Decimal("50.00")) * Decimal("0.18"))
        - (mean_absolute_error_pct * Decimal("2.85"))
        - (abs(average_gap_pct) * Decimal("1.25")),
        Decimal("18.00"),
        Decimal("92.00"),
    )
    if score >= Decimal("76.00"):
        label = "Memoria fiable"
        reason = "El historico del modelo acompana esta lectura."
    elif score >= Decimal("58.00"):
        label = "Memoria usable"
        reason = "La memoria historica se usa, pero sin darle plena confianza."
    else:
        label = "Memoria penaliza"
        reason = "Las desviaciones pasadas obligan a enfriar la recomendacion."
    return {
        "score": quantize_decimal(score),
        "sample_count": sample_count,
        "label": label,
        "reason": reason,
        "mean_absolute_error_pct": mean_absolute_error_pct,
        "average_gap_pct": average_gap_pct,
        "direction_hit_rate_pct": direction_hit_rate_pct,
    }


def build_purchase_discipline_portfolio_context(cards: list[dict]) -> dict:
    sector_values: dict[str, Decimal] = defaultdict(lambda: ZERO)
    owned_tickers = set()
    owned_positions_count = 0
    total_value = ZERO
    for card in cards or []:
        position = card.get("position")
        if position is None or not getattr(position, "is_owned", False):
            continue
        owned_positions_count += 1
        ticker = clean_ticker(getattr(position, "ticker", ""))
        if ticker:
            owned_tickers.add(ticker)
        current_value = quantize_decimal(getattr(position, "current_value", None), "0.01") or ZERO
        if current_value <= ZERO:
            continue
        sector_label = card.get("sector_label") or resolve_equity_sector_label(
            company_name=getattr(position, "company_name", ""),
            ticker=getattr(position, "ticker", ""),
            quote_symbol=getattr(position, "quote_symbol", ""),
        ) or "Sin sector"
        sector_values[sector_label] += current_value
        total_value += current_value
    return {
        "total_value": total_value,
        "sector_values": dict(sector_values),
        "owned_tickers": owned_tickers,
        "owned_positions_count": owned_positions_count,
    }


def score_optimizer_portfolio_fit(candidate: dict, portfolio_context: dict | None) -> dict:
    portfolio_context = portfolio_context or {}
    position = candidate.get("position")
    sector_label = candidate.get("sector_label") or "Sin sector"
    total_value = quantize_decimal(portfolio_context.get("total_value"), "0.01") or ZERO
    sector_value = quantize_decimal((portfolio_context.get("sector_values") or {}).get(sector_label), "0.01") or ZERO
    ticker = clean_ticker(getattr(position, "ticker", ""))
    owned_tickers = portfolio_context.get("owned_tickers") or set()
    is_owned = bool(getattr(position, "is_owned", False))
    sector_weight_pct = (sector_value / total_value) * ONE_HUNDRED if total_value > ZERO else ZERO
    score = Decimal("72.00")
    if is_owned or ticker in owned_tickers:
        score += Decimal("5.00")
    elif total_value <= ZERO:
        score = Decimal("68.00")
    elif sector_weight_pct <= Decimal("8.00"):
        score += Decimal("7.00")
    elif sector_weight_pct >= Decimal("30.00"):
        score -= Decimal("18.00")
    elif sector_weight_pct >= Decimal("22.00"):
        score -= Decimal("10.00")
    elif sector_weight_pct >= Decimal("15.00"):
        score -= Decimal("4.00")

    if portfolio_context.get("owned_positions_count", 0) <= 2 and not is_owned:
        score += Decimal("4.00")
    score = clamp_decimal(score, Decimal("25.00"), Decimal("90.00"))
    if score >= Decimal("76.00"):
        label = "Encaja bien"
        reason = "Aporta diversificacion o refuerza una posicion ya controlada."
    elif score >= Decimal("60.00"):
        label = "Encaje neutro"
        reason = "No mejora mucho la diversificacion, pero tampoco concentra en exceso."
    else:
        label = "Concentra cartera"
        reason = "El sector ya pesa bastante y exige mas margen de seguridad."
    return {
        "score": quantize_decimal(score),
        "label": label,
        "reason": reason,
        "sector_weight_pct": quantize_decimal(sector_weight_pct),
    }


def score_optimizer_return_quality(candidate: dict) -> dict:
    primary_signal_pct = quantize_decimal(candidate.get("primary_signal_pct"), "0.01") or ZERO
    scenario_expected_return_pct = quantize_decimal(candidate.get("scenario_expected_return_pct"), "0.01")
    holding_gap_pct = quantize_decimal(candidate.get("annualized_target_gap_pct"), "0.01")
    score = Decimal("50.00") + (primary_signal_pct * Decimal("2.05"))
    if scenario_expected_return_pct is not None:
        score += scenario_expected_return_pct * Decimal("0.45")
    if holding_gap_pct is not None:
        score += holding_gap_pct * Decimal("0.35")
    score = clamp_decimal(score, Decimal("18.00"), Decimal("94.00"))
    if score >= Decimal("76.00"):
        label = "Retorno atractivo"
        reason = "El retorno robusto supera claramente el coste de oportunidad."
    elif score >= Decimal("60.00"):
        label = "Retorno correcto"
        reason = "El retorno acompana, pero no permite relajar otros filtros."
    else:
        label = "Retorno justo"
        reason = "La rentabilidad esperada exige prudencia."
    return {"score": quantize_decimal(score), "label": label, "reason": reason}


def score_optimizer_risk_quality(candidate: dict) -> dict:
    safety_score = quantize_decimal(candidate.get("safety_score"), "0.01") or ZERO
    reliability_score = quantize_decimal(candidate.get("reliability_score"), "0.01") or ZERO
    worst_return_pct = quantize_decimal(candidate.get("low_return_pct"), "0.01") or ZERO
    stress_return_pct = quantize_decimal(candidate.get("downside_stress_return_pct"), "0.01") or worst_return_pct
    volatility_pct = quantize_decimal(candidate.get("annualized_volatility_pct"), "0.01") or Decimal("18.00")
    uncertainty_penalty_pct = quantize_decimal(candidate.get("uncertainty_penalty_pct"), "0.01") or ZERO
    score = (
        (safety_score * Decimal("0.34"))
        + (reliability_score * Decimal("0.32"))
        + (clamp_decimal(worst_return_pct + Decimal("20.00"), ZERO, Decimal("35.00")) * Decimal("0.80"))
        + (clamp_decimal(stress_return_pct + Decimal("18.00"), ZERO, Decimal("32.00")) * Decimal("0.45"))
        - (max(volatility_pct - Decimal("14.00"), ZERO) * Decimal("0.70"))
        - (uncertainty_penalty_pct * Decimal("3.50"))
        + Decimal("10.00")
    )
    score = clamp_decimal(score, Decimal("15.00"), Decimal("94.00"))
    if score >= Decimal("76.00"):
        label = "Riesgo controlado"
        reason = "Seguridad, fiabilidad y peor escenario son compatibles con compra."
    elif score >= Decimal("60.00"):
        label = "Riesgo medio"
        reason = "El riesgo es asumible solo si el retorno y el timing compensan."
    else:
        label = "Riesgo alto"
        reason = "La proteccion es floja para abrir compra con conviccion."
    return {"score": quantize_decimal(score), "label": label, "reason": reason}


def score_optimizer_timing_quality(candidate: dict) -> dict:
    purchase_timing = candidate.get("purchase_timing") or {}
    holding_annualized_return_pct = quantize_decimal(candidate.get("holding_annualized_return_pct"), "0.01")
    target_return_pct = quantize_decimal(candidate.get("annualized_target_return_pct"), "0.01") or OPTIMIZER_TARGET_HOLDING_ANNUALIZED_RETURN_PCT
    score = Decimal("58.00")
    if holding_annualized_return_pct is not None:
        score = Decimal("62.00") + ((holding_annualized_return_pct - target_return_pct) * Decimal("0.95"))
    if purchase_timing.get("available") and purchase_timing.get("mode_label") == "Comprar ya":
        score += Decimal("5.00")
    elif purchase_timing.get("available") and purchase_timing.get("mode_label") == "Entrada gradual":
        score += Decimal("2.00")
    elif purchase_timing.get("available") and purchase_timing.get("mode_label") == "Esperar correccion":
        score -= Decimal("3.00")
    score = clamp_decimal(score, Decimal("20.00"), Decimal("92.00"))
    if score >= Decimal("76.00"):
        label = "Timing favorable"
        reason = "La ventana tactica permite actuar sin esperar mucho."
    elif score >= Decimal("60.00"):
        label = "Timing aceptable"
        reason = "Se puede comprar por tramos o esperar precio."
    else:
        label = "Timing debil"
        reason = "La entrada no esta clara todavia."
    return {"score": quantize_decimal(score), "label": label, "reason": reason}


def build_optimizer_purchase_discipline_review(candidate: dict, portfolio_context: dict | None = None) -> dict:
    return_quality = score_optimizer_return_quality(candidate)
    risk_quality = score_optimizer_risk_quality(candidate)
    memory_quality = score_optimizer_memory_quality(candidate.get("expectation_review_signal") or {})
    portfolio_fit = score_optimizer_portfolio_fit(candidate, portfolio_context)
    timing_quality = score_optimizer_timing_quality(candidate)
    score = (
        (return_quality["score"] * Decimal("0.30"))
        + (risk_quality["score"] * Decimal("0.25"))
        + (memory_quality["score"] * Decimal("0.20"))
        + (portfolio_fit["score"] * Decimal("0.15"))
        + (timing_quality["score"] * Decimal("0.10"))
    )
    score = clamp_decimal(score, Decimal("10.00"), Decimal("95.00"))
    weakest = min(
        [
            ("retorno", return_quality),
            ("riesgo", risk_quality),
            ("memoria", memory_quality),
            ("cartera", portfolio_fit),
            ("timing", timing_quality),
        ],
        key=lambda item: item[1]["score"],
    )
    strongest = max(
        [
            ("retorno", return_quality),
            ("riesgo", risk_quality),
            ("memoria", memory_quality),
            ("cartera", portfolio_fit),
            ("timing", timing_quality),
        ],
        key=lambda item: item[1]["score"],
    )
    if score >= Decimal("80.00"):
        label = "Comprar"
        action_label = "Comprar ahora" if (candidate.get("purchase_timing") or {}).get("mode_label") == "Comprar ya" else "Comprar por tramos"
        tone = "buy"
        reason = strongest[1]["reason"]
    elif score >= PURCHASE_DISCIPLINE_TARGET_SCORE:
        label = "Comprar por tramos"
        action_label = "Comprar por tramos"
        tone = "buy"
        reason = strongest[1]["reason"]
    elif score >= Decimal("58.00"):
        label = "Vigilar precio"
        action_label = "Esperar"
        tone = "watch"
        reason = weakest[1]["reason"]
    else:
        label = "No comprar"
        action_label = "Descartar"
        tone = "warn"
        reason = weakest[1]["reason"]
    adjustment_pct = clamp_decimal((score - PURCHASE_DISCIPLINE_TARGET_SCORE) * Decimal("0.055"), Decimal("-2.75"), Decimal("2.00"))
    return {
        "available": True,
        "score": quantize_decimal(score),
        "label": label,
        "action_label": action_label,
        "tone": tone,
        "reason": reason,
        "weakest_key": weakest[0],
        "strongest_key": strongest[0],
        "adjustment_pct": quantize_decimal(adjustment_pct),
        "return_score": return_quality["score"],
        "risk_score": risk_quality["score"],
        "memory_score": memory_quality["score"],
        "portfolio_fit_score": portfolio_fit["score"],
        "timing_score": timing_quality["score"],
        "memory_label": memory_quality["label"],
        "portfolio_fit_label": portfolio_fit["label"],
        "timing_label": timing_quality["label"],
    }


def apply_purchase_discipline_to_optimizer_candidates(
    candidates: list[dict],
    portfolio_context: dict | None = None,
) -> list[dict]:
    for candidate in candidates:
        if candidate.get("purchase_discipline_applied"):
            continue
        base_score = quantize_decimal(candidate.get("optimization_score"), "0.01") or ZERO
        discipline = build_optimizer_purchase_discipline_review(candidate, portfolio_context)
        candidate["base_optimization_score"] = base_score
        candidate["purchase_discipline"] = discipline
        candidate["purchase_discipline_score"] = discipline.get("score")
        candidate["purchase_discipline_label"] = discipline.get("label", "")
        candidate["purchase_discipline_reason"] = discipline.get("reason", "")
        candidate["purchase_discipline_adjustment_pct"] = discipline.get("adjustment_pct") or ZERO
        candidate["optimization_score"] = quantize_decimal(
            base_score + (discipline.get("adjustment_pct") or ZERO),
            "0.01",
        )
        candidate["purchase_discipline_applied"] = True
    return candidates


def build_equity_optimizer_candidate(card: dict, strategy_mode: str = OPTIMIZER_STRATEGY_12M_PRIMARY) -> dict | None:
    projection = card.get("projection") or {}
    effective_projection = resolve_effective_projection_metrics(card)
    if not projection.get("available") or effective_projection.get("projected_price") is None:
        return None

    base_return_pct = effective_projection.get("base_return_pct")
    if base_return_pct is None:
        return None

    confidence_label = projection.get("confidence_label", "Baja")
    reliability = card.get("projection_reliability") or {}
    reliability_label = reliability.get("label") or confidence_label
    reliability_score = reliability.get("score") or projection_reliability_score(reliability_label)
    safety_score = projection.get("safety_score") or Decimal("60.00")
    years_covered = projection.get("years_covered") or ZERO
    positive_year_ratio_pct = projection.get("positive_year_ratio_pct") or Decimal("50.00")
    annualized_volatility_pct = projection.get("annualized_volatility_pct") or Decimal("18.00")
    low_return_pct = effective_projection.get("low_return_pct")
    current_drawdown_pct = projection.get("current_drawdown_pct") or ZERO
    max_drawdown_pct = projection.get("max_drawdown_pct") or ZERO
    net_income_yield_pct = projection.get("net_income_yield_pct") or ZERO
    gross_dividend_yield_pct = projection.get("gross_dividend_yield_pct") or ZERO
    transaction_drag_pct = projection.get("transaction_drag_pct") or ZERO
    cycle_projection = card.get("cycle_projection_5y") or {}
    cycle_return_annual_pct = cycle_projection.get("annual_return_pct")
    cycle_return_5y_pct = cycle_projection.get("five_year_return_pct")
    trade_alert = card.get("trade_alert") or {}
    trade_alert_label = trade_alert.get("label") or "Vigilar"
    external_signal = card.get("external_signal") or {}
    external_signal_score = Decimal(str(external_signal.get("score", "0") or "0"))
    external_signal_label = external_signal.get("label") or "Sin prensa reciente"
    decision_action_label = build_decision_action_label(
        card["position"],
        base_return_pct,
        safety_score,
        reliability_score,
        (card.get("valuation") or {}).get("score"),
    )
    quality_adjustments = build_optimizer_quality_adjustments(
        safety_score,
        reliability_score,
        decision_action_label,
    )
    trade_signal_adjustment = {
        "Comprar": Decimal("1.75"),
        "Vigilar": ZERO,
        "Vender": Decimal("-4.50"),
    }.get(trade_alert_label, ZERO)
    external_signal_adjustment = clamp_decimal(
        external_signal_score * Decimal("0.35"),
        Decimal("-2.50"),
        Decimal("2.50"),
    )
    twelve_month_scenario = build_optimizer_scenario_summary(
        build_effective_projection_scenarios(card, effective_metrics=effective_projection),
        return_key="total_return_pct",
        fallback_value=base_return_pct,
    )
    five_year_scenario = build_optimizer_scenario_summary(
        cycle_projection.get("scenarios"),
        return_key="annual_return_pct",
        fallback_value=cycle_return_annual_pct,
    )
    expectation_review_signal = resolve_card_optimizer_expectation_review_signal(card)
    uncertainty_profile = build_optimizer_uncertainty_profile(card, projection, cycle_projection, external_signal)
    raw_scenario_expected_return_pct = twelve_month_scenario.get("expected_return_pct") or base_return_pct
    scenario_expectation_review = apply_optimizer_expectation_review_adjustment(
        raw_scenario_expected_return_pct,
        (expectation_review_signal.get("1y") if expectation_review_signal.get("available") else None),
    )
    scenario_expected_return_pct = (
        scenario_expectation_review.get("adjusted_return_pct")
        if scenario_expectation_review.get("adjusted_return_pct") is not None
        else raw_scenario_expected_return_pct
    )
    downside_stress_return_pct = twelve_month_scenario.get("worst_return_pct")
    if downside_stress_return_pct is None:
        downside_stress_return_pct = low_return_pct if low_return_pct is not None else base_return_pct
    scenario_spread_pct = twelve_month_scenario.get("spread_pct") or ZERO
    raw_cycle_expected_annual_return_pct = five_year_scenario.get("expected_return_pct")
    cycle_expectation_review = apply_optimizer_expectation_review_adjustment(
        raw_cycle_expected_annual_return_pct,
        (expectation_review_signal.get("5y") if expectation_review_signal.get("available") else None),
    )
    cycle_expected_annual_return_pct = cycle_expectation_review.get("adjusted_return_pct")
    if cycle_expected_annual_return_pct is None:
        cycle_expected_annual_return_pct = cycle_return_annual_pct
    cycle_downside_stress_annual_pct = five_year_scenario.get("worst_return_pct")
    if cycle_downside_stress_annual_pct is None:
        cycle_downside_stress_annual_pct = cycle_return_annual_pct
    cycle_scenario_spread_pct = five_year_scenario.get("spread_pct") or ZERO
    uncertainty_penalty_pct = uncertainty_profile.get("uncertainty_penalty_pct") or ZERO

    downside_return_pct = ZERO
    if low_return_pct is not None and low_return_pct < ZERO:
        downside_return_pct = abs(low_return_pct)

    history_depth_multiplier = clamp_decimal(
        (years_covered / Decimal(str(LONG_ANALYSIS_YEARS))) if years_covered else Decimal("0.55"),
        Decimal("0.55"),
        Decimal("1.00"),
    )
    consistency_multiplier = clamp_decimal(
        Decimal("0.75") + ((positive_year_ratio_pct - Decimal("50.00")) / Decimal("100")),
        Decimal("0.70"),
        Decimal("1.10"),
    )
    confidence_multiplier = projection_confidence_multiplier(confidence_label)
    quality_multiplier = (
        confidence_multiplier
        * (Decimal("0.55") + (safety_score / Decimal("200")))
        * (Decimal("0.55") + (reliability_score / Decimal("200")))
        * history_depth_multiplier
        * consistency_multiplier
    )

    risk_penalty_pct = (
        (downside_return_pct * Decimal("0.32"))
        + (annualized_volatility_pct * Decimal("0.08"))
        + (abs(current_drawdown_pct) * Decimal("0.05") if current_drawdown_pct < ZERO else ZERO)
        + (abs(max_drawdown_pct) * Decimal("0.02") if max_drawdown_pct < ZERO else ZERO)
    )
    income_support_bonus_pct = clamp_decimal(net_income_yield_pct * Decimal("0.30"), Decimal("-2.00"), Decimal("3.00"))
    cost_efficiency_penalty_pct = clamp_decimal(transaction_drag_pct * Decimal("0.40"), ZERO, Decimal("3.00"))
    risk_adjusted_return_pct = (
        (base_return_pct * quality_multiplier)
        - risk_penalty_pct
        + clamp_decimal(external_signal_score * Decimal("0.12"), Decimal("-1.20"), Decimal("1.20"))
    )
    scenario_expectation_adjustment_pct = clamp_decimal(
        ((scenario_expected_return_pct or ZERO) - (base_return_pct or ZERO)) * Decimal("0.70"),
        Decimal("-3.00"),
        Decimal("3.00"),
    )
    scenario_stress_penalty_pct = clamp_decimal(
        (abs(min(downside_stress_return_pct or ZERO, ZERO)) * Decimal("0.14"))
        + (scenario_spread_pct * Decimal("0.05")),
        ZERO,
        Decimal("8.00"),
    )
    robust_return_signal_pct = clamp_decimal(
        risk_adjusted_return_pct
        + scenario_expectation_adjustment_pct
        - scenario_stress_penalty_pct
        - uncertainty_penalty_pct,
        Decimal("-25.00"),
        Decimal("25.00"),
    )
    cycle_support_score = ZERO
    if cycle_projection.get("available") and cycle_return_annual_pct is not None:
        cycle_quality_multiplier = (
            (Decimal("0.70") + (safety_score / Decimal("250")))
            * (Decimal("0.70") + (reliability_score / Decimal("250")))
            * history_depth_multiplier
        )
        cycle_risk_penalty_pct = (
            (abs(current_drawdown_pct) * Decimal("0.03") if current_drawdown_pct < ZERO else ZERO)
            + (abs(max_drawdown_pct) * Decimal("0.015") if max_drawdown_pct < ZERO else ZERO)
        )
        cycle_support_score = clamp_decimal(
            (cycle_return_annual_pct * cycle_quality_multiplier) - cycle_risk_penalty_pct,
            Decimal("-8.00"),
            Decimal("10.00"),
        )
        if cycle_return_5y_pct is not None and cycle_return_5y_pct < ZERO:
            cycle_support_score += clamp_decimal(cycle_return_5y_pct * Decimal("0.05"), Decimal("-3.00"), ZERO)
    cycle_expectation_adjustment_pct = clamp_decimal(
        ((cycle_expected_annual_return_pct or ZERO) - (cycle_return_annual_pct or ZERO)) * Decimal("0.85"),
        Decimal("-2.50"),
        Decimal("2.50"),
    )
    cycle_stress_penalty_pct = clamp_decimal(
        (abs(min(cycle_downside_stress_annual_pct or ZERO, ZERO)) * Decimal("0.45"))
        + (cycle_scenario_spread_pct * Decimal("0.18")),
        ZERO,
        Decimal("6.00"),
    )
    robust_cycle_support_score = clamp_decimal(
        cycle_support_score
        + cycle_expectation_adjustment_pct
        - cycle_stress_penalty_pct
        - (uncertainty_penalty_pct * Decimal("0.55")),
        Decimal("-12.00"),
        Decimal("12.00"),
    )
    strategy = get_optimizer_strategy_config(strategy_mode)
    purchase_timing = build_candidate_purchase_timing_plan(card, strategy["mode"])
    holding_annualized_return_pct = quantize_decimal(
        purchase_timing.get("holding_annualized_return_pct"),
        "0.01",
    ) if purchase_timing.get("available") else None
    annualized_target_return_pct = OPTIMIZER_TARGET_HOLDING_ANNUALIZED_RETURN_PCT
    annualized_target_gap_pct = (
        quantize_decimal(holding_annualized_return_pct - annualized_target_return_pct, "0.01")
        if holding_annualized_return_pct is not None
        else None
    )
    annualized_target_penalty_pct = clamp_decimal(
        max(annualized_target_return_pct - (holding_annualized_return_pct or ZERO), ZERO)
        * OPTIMIZER_TARGET_SHORTFALL_PENALTY_MULTIPLE,
        ZERO,
        Decimal("10.00"),
    ) if holding_annualized_return_pct is not None else ZERO
    annualized_target_bonus_pct = clamp_decimal(
        max((holding_annualized_return_pct or ZERO) - annualized_target_return_pct, ZERO)
        * OPTIMIZER_TARGET_EXCESS_BONUS_MULTIPLE,
        ZERO,
        Decimal("3.00"),
    ) if holding_annualized_return_pct is not None else ZERO
    conservative_risk_penalty_pct = clamp_decimal(
        (abs(min(low_return_pct or ZERO, ZERO)) * Decimal("0.16"))
        + (abs(min(downside_stress_return_pct or ZERO, ZERO)) * Decimal("0.22"))
        + (annualized_volatility_pct * Decimal("0.05"))
        + (uncertainty_penalty_pct * Decimal("0.45")),
        ZERO,
        Decimal("10.00"),
    )
    conservative_quality_bonus_pct = clamp_decimal(
        (max(safety_score - Decimal("70.00"), ZERO) * Decimal("0.05"))
        + (max(reliability_score - Decimal("70.00"), ZERO) * Decimal("0.04")),
        ZERO,
        Decimal("3.50"),
    )
    meets_target_annualized_return = (
        holding_annualized_return_pct >= annualized_target_return_pct
        if holding_annualized_return_pct is not None
        else False
    )
    if strategy["mode"] == OPTIMIZER_STRATEGY_5Y_PRIMARY:
        primary_signal_pct = robust_cycle_support_score
        secondary_signal_pct = robust_return_signal_pct
    else:
        primary_signal_pct = robust_return_signal_pct
        secondary_signal_pct = robust_cycle_support_score
    blended_return_signal_pct = (
        (primary_signal_pct * strategy["primary_weight"])
        + (secondary_signal_pct * strategy["secondary_weight"])
    )
    optimization_score = (
        blended_return_signal_pct
        + income_support_bonus_pct
        - cost_efficiency_penalty_pct
        + annualized_target_bonus_pct
        - annualized_target_penalty_pct
        + conservative_quality_bonus_pct
        - conservative_risk_penalty_pct
        + trade_signal_adjustment
        + external_signal_adjustment
        + (quality_adjustments.get("decision_action_adjustment") or ZERO)
        - (quality_adjustments.get("quality_floor_penalty_pct") or ZERO)
    )

    candidate = {
        "card": card,
        "position": card["position"],
        "projection": projection,
        "effective_projection_metrics": effective_projection,
        "status_key": card.get("status_key") or ("owned" if card["position"].is_owned else "watchlist"),
        "status_label": card.get("status_label") or card["position"].get_position_kind_display(),
        "sector_label": card.get("sector_label") or resolve_equity_sector_label(
            company_name=card["position"].company_name,
            ticker=card["position"].ticker,
            quote_symbol=card["position"].quote_symbol,
        ) or "Sin sector",
        "strategy_mode": strategy["mode"],
        "strategy_label": strategy["label"],
        "strategy_primary_horizon_label": strategy["primary_horizon_label"],
        "strategy_secondary_horizon_label": strategy["secondary_horizon_label"],
        "reference_label": card["reference_label"],
        "confidence_label": confidence_label,
        "reliability_label": reliability_label,
        "reliability_score": reliability_score,
        "safety_score": safety_score,
        "base_return_pct": base_return_pct,
        "low_return_pct": low_return_pct,
        "expectation_review_signal": expectation_review_signal,
        "raw_scenario_expected_return_pct": raw_scenario_expected_return_pct,
        "scenario_expected_return_pct": scenario_expected_return_pct,
        "scenario_expectation_review": scenario_expectation_review,
        "downside_stress_return_pct": downside_stress_return_pct,
        "scenario_spread_pct": scenario_spread_pct,
        "raw_cycle_expected_annual_return_pct": raw_cycle_expected_annual_return_pct,
        "cycle_expected_annual_return_pct": cycle_expected_annual_return_pct,
        "cycle_expectation_review": cycle_expectation_review,
        "cycle_downside_stress_annual_pct": cycle_downside_stress_annual_pct,
        "cycle_scenario_spread_pct": cycle_scenario_spread_pct,
        "primary_signal_pct": primary_signal_pct,
        "secondary_signal_pct": secondary_signal_pct,
        "risk_adjusted_return_pct": risk_adjusted_return_pct,
        "robust_return_signal_pct": robust_return_signal_pct,
        "blended_return_signal_pct": blended_return_signal_pct,
        "optimization_score": optimization_score,
        "downside_return_pct": downside_return_pct,
        "annualized_volatility_pct": annualized_volatility_pct,
        "gross_dividend_yield_pct": gross_dividend_yield_pct,
        "net_income_yield_pct": net_income_yield_pct,
        "transaction_drag_pct": transaction_drag_pct,
        "cycle_projection_available": bool(cycle_projection.get("available")),
        "cycle_return_annual_pct": cycle_return_annual_pct,
        "cycle_return_5y_pct": cycle_return_5y_pct,
        "cycle_support_score": cycle_support_score,
        "robust_cycle_support_score": robust_cycle_support_score,
        "scenario_expectation_adjustment_pct": scenario_expectation_adjustment_pct,
        "scenario_stress_penalty_pct": scenario_stress_penalty_pct,
        "cycle_expectation_adjustment_pct": cycle_expectation_adjustment_pct,
        "cycle_stress_penalty_pct": cycle_stress_penalty_pct,
        "years_covered": years_covered,
        "cycle_phase": projection.get("cycle_phase") or "Sin ciclo",
        "max_drawdown_pct": max_drawdown_pct,
        "current_drawdown_pct": current_drawdown_pct,
        "trade_alert_label": trade_alert_label,
        "trade_alert_tone": trade_alert.get("tone", "watch"),
        "trade_alert_note": trade_alert.get("note", ""),
        "decision_action_label": decision_action_label,
        "decision_action_adjustment": quality_adjustments.get("decision_action_adjustment") or ZERO,
        "safety_gap": quality_adjustments.get("safety_gap") or ZERO,
        "reliability_gap": quality_adjustments.get("reliability_gap") or ZERO,
        "safety_floor_penalty_pct": quality_adjustments.get("safety_floor_penalty_pct") or ZERO,
        "reliability_floor_penalty_pct": quality_adjustments.get("reliability_floor_penalty_pct") or ZERO,
        "quality_floor_penalty_pct": quality_adjustments.get("quality_floor_penalty_pct") or ZERO,
        "trade_signal_adjustment": trade_signal_adjustment,
        "holding_annualized_return_pct": holding_annualized_return_pct,
        "annualized_target_return_pct": annualized_target_return_pct,
        "annualized_target_gap_pct": annualized_target_gap_pct,
        "annualized_target_penalty_pct": annualized_target_penalty_pct,
        "annualized_target_bonus_pct": annualized_target_bonus_pct,
        "annualized_target_applicable": holding_annualized_return_pct is not None,
        "meets_target_annualized_return": meets_target_annualized_return,
        "conservative_risk_penalty_pct": conservative_risk_penalty_pct,
        "conservative_quality_bonus_pct": conservative_quality_bonus_pct,
        "external_signal_label": external_signal_label,
        "external_signal_score": external_signal_score,
        "external_signal_adjustment": external_signal_adjustment,
        "external_signal_note": external_signal.get("note", ""),
        "external_signal_items_count": external_signal.get("items_count", 0),
        "material_event": uncertainty_profile.get("material_event", False),
        "uncertainty_penalty_pct": uncertainty_penalty_pct,
        "model_uncertainty_penalty_pct": uncertainty_profile.get("model_uncertainty_penalty_pct") or ZERO,
        "external_signal_penalty_pct": uncertainty_profile.get("external_signal_penalty_pct") or ZERO,
        "coverage_penalty_pct": uncertainty_profile.get("coverage_penalty_pct") or ZERO,
        "purchase_timing": purchase_timing,
    }
    conservative_profile_review = build_optimizer_conservative_profile_review(candidate)
    candidate["risk_profile_label"] = OPTIMIZER_RISK_PROFILE_LABEL
    candidate["conservative_profile_review"] = conservative_profile_review
    candidate["passes_conservative_profile"] = conservative_profile_review["passes"]
    candidate["conservative_profile_summary"] = conservative_profile_review["summary"]
    candidate["conservative_profile_failures"] = conservative_profile_review["failures"]
    return candidate


def candidate_meets_target_annualized_return(candidate: dict) -> bool:
    annualized_return_pct = quantize_decimal(candidate.get("holding_annualized_return_pct"), "0.01")
    if annualized_return_pct is None:
        return True
    target_return_pct = quantize_decimal(
        candidate.get("annualized_target_return_pct"),
        "0.01",
    ) or OPTIMIZER_TARGET_HOLDING_ANNUALIZED_RETURN_PCT
    return annualized_return_pct >= target_return_pct


def build_optimizer_conservative_profile_review(candidate: dict) -> dict:
    safety_score = quantize_decimal(candidate.get("safety_score"), "0.01") or ZERO
    reliability_score = quantize_decimal(candidate.get("reliability_score"), "0.01") or ZERO
    worst_return_pct = quantize_decimal(candidate.get("low_return_pct"), "0.01")
    stress_return_pct = quantize_decimal(candidate.get("downside_stress_return_pct"), "0.01")
    annualized_volatility_pct = quantize_decimal(candidate.get("annualized_volatility_pct"), "0.01") or ZERO
    uncertainty_penalty_pct = quantize_decimal(candidate.get("uncertainty_penalty_pct"), "0.01") or ZERO
    trade_alert_label = str(candidate.get("trade_alert_label") or "").strip()

    failures = []
    if trade_alert_label != "Comprar":
        failures.append("alerta de compra")
    if safety_score < OPTIMIZER_CONSERVATIVE_MIN_SAFETY_SCORE:
        failures.append("seguridad")
    if reliability_score < OPTIMIZER_CONSERVATIVE_MIN_RELIABILITY_SCORE:
        failures.append("fiabilidad")
    if worst_return_pct is None or worst_return_pct < OPTIMIZER_CONSERVATIVE_MIN_WORST_RETURN_PCT:
        failures.append("peor escenario")
    if stress_return_pct is None or stress_return_pct < OPTIMIZER_CONSERVATIVE_MIN_STRESS_RETURN_PCT:
        failures.append("estres")
    if annualized_volatility_pct > OPTIMIZER_CONSERVATIVE_MAX_VOLATILITY_PCT:
        failures.append("volatilidad")
    if uncertainty_penalty_pct > OPTIMIZER_CONSERVATIVE_MAX_UNCERTAINTY_PENALTY_PCT:
        failures.append("incertidumbre")

    summary = (
        f"Perfil {OPTIMIZER_RISK_PROFILE_LABEL.lower()}: solo entran ideas con alerta Comprar, "
        f"seguridad >= {OPTIMIZER_CONSERVATIVE_MIN_SAFETY_SCORE:.0f}, "
        f"fiabilidad >= {OPTIMIZER_CONSERVATIVE_MIN_RELIABILITY_SCORE:.0f}, "
        f"peor escenario >= {OPTIMIZER_CONSERVATIVE_MIN_WORST_RETURN_PCT:.0f} %, "
        f"estres >= {OPTIMIZER_CONSERVATIVE_MIN_STRESS_RETURN_PCT:.0f} %, "
        f"volatilidad <= {OPTIMIZER_CONSERVATIVE_MAX_VOLATILITY_PCT:.0f} % "
        f"e incertidumbre <= {OPTIMIZER_CONSERVATIVE_MAX_UNCERTAINTY_PENALTY_PCT:.2f} pts."
    )
    if failures:
        summary += " Esta idea no pasa por: " + ", ".join(failures) + "."

    return {
        "passes": not failures,
        "failures": failures,
        "summary": summary,
    }


def candidate_passes_conservative_profile(candidate: dict) -> bool:
    profile_review = candidate.get("conservative_profile_review") or build_optimizer_conservative_profile_review(candidate)
    return bool(profile_review.get("passes"))


def filter_positive_optimizer_candidates(
    candidates: list[dict],
    strategy_mode: str,
) -> list[dict]:
    strategy = get_optimizer_strategy_config(strategy_mode)
    if strategy["mode"] == OPTIMIZER_STRATEGY_5Y_PRIMARY:
        return [
            item
            for item in candidates
            if item["optimization_score"] > ZERO
            and item.get("robust_cycle_support_score", item["cycle_support_score"]) > ZERO
            and candidate_meets_target_annualized_return(item)
            and candidate_passes_conservative_profile(item)
            and item["trade_alert_label"] != "Vender"
        ]
    return [
        item
        for item in candidates
        if item["optimization_score"] > ZERO
        and item.get("robust_return_signal_pct", item["base_return_pct"]) > ZERO
        and candidate_meets_target_annualized_return(item)
        and candidate_passes_conservative_profile(item)
        and item["trade_alert_label"] != "Vender"
    ]


def rank_optimizer_candidates(candidates: list[dict]) -> list[dict]:
    return sorted(
        candidates,
        key=lambda item: (
            item["optimization_score"],
            item["safety_score"],
            item["reliability_score"],
            item.get("downside_stress_return_pct") if item.get("downside_stress_return_pct") is not None else Decimal("-9999"),
            item.get("low_return_pct") if item.get("low_return_pct") is not None else Decimal("-9999"),
            -(item.get("annualized_volatility_pct") or Decimal("9999")),
            -(item.get("uncertainty_penalty_pct") or Decimal("9999")),
            item.get("purchase_discipline_score") or ZERO,
            item["primary_signal_pct"],
            item["secondary_signal_pct"],
            item.get("scenario_expected_return_pct") if item.get("scenario_expected_return_pct") is not None else Decimal("-9999"),
            item.get("downside_stress_return_pct") if item.get("downside_stress_return_pct") is not None else Decimal("-9999"),
            item["base_return_pct"],
            item["safety_score"],
        ),
        reverse=True,
    )


def filter_optimizer_candidates_by_sector(
    candidates: list[dict],
    max_sector_positions: int,
) -> tuple[list[dict], list[dict]]:
    if max_sector_positions <= 0:
        return candidates, []

    sector_counts: dict[str, int] = {}
    accepted = []
    excluded = []
    for candidate in candidates:
        sector_label = candidate.get("sector_label") or "Sin sector"
        current_count = sector_counts.get(sector_label, 0)
        if current_count >= max_sector_positions:
            excluded.append(candidate)
            continue
        sector_counts[sector_label] = current_count + 1
        accepted.append(candidate)
    return accepted, excluded


def distribute_optimizer_amounts(
    candidates: list[dict],
    total_investment: Decimal,
    company_cap_amount: Decimal,
) -> tuple[list[Decimal], Decimal]:
    if not candidates:
        return [], total_investment

    epsilon = Decimal("0.01")
    allocations = [ZERO for _ in candidates]
    open_indexes = list(range(len(candidates)))
    remaining_amount = total_investment

    while remaining_amount > epsilon and open_indexes:
        score_total = sum((candidates[index]["optimization_score"] for index in open_indexes), ZERO)
        if score_total <= ZERO:
            break

        distributed_amount = ZERO
        next_open_indexes = []
        for index in open_indexes:
            capacity = company_cap_amount - allocations[index]
            if capacity <= epsilon:
                continue
            proportional_amount = remaining_amount * (candidates[index]["optimization_score"] / score_total)
            allocated_amount = min(capacity, proportional_amount)
            if allocated_amount <= ZERO:
                continue
            allocations[index] += allocated_amount
            distributed_amount += allocated_amount
            if company_cap_amount - allocations[index] > epsilon:
                next_open_indexes.append(index)

        if distributed_amount <= epsilon:
            break
        remaining_amount -= distributed_amount
        open_indexes = next_open_indexes

    quantized_allocations = [quantize_decimal(amount, "0.01") or ZERO for amount in allocations]
    allocated_total = sum(quantized_allocations, ZERO)
    remaining_amount = quantize_decimal(total_investment - allocated_total, "0.01") or ZERO

    if remaining_amount > ZERO:
        for index, amount in enumerate(quantized_allocations):
            capacity = company_cap_amount - amount
            top_up = min(capacity, remaining_amount)
            if top_up <= ZERO:
                continue
            quantized_allocations[index] = amount + top_up
            remaining_amount -= top_up
            if remaining_amount <= ZERO:
                break

    return quantized_allocations, remaining_amount


def build_optimizer_allocation_scenario(candidate: dict, allocated_amount: Decimal) -> dict:
    position = candidate["position"]
    projection = candidate["projection"]
    effective_projection = candidate.get("effective_projection_metrics") or resolve_effective_projection_metrics(candidate.get("card") or {})
    if allocated_amount <= ZERO:
        return {
            "gross_dividend_income": ZERO,
            "net_dividend_income": ZERO,
            "annual_cost_used": ZERO,
            "roundtrip_total_cost": ZERO,
            "transaction_drag_pct": ZERO,
            "net_income_yield_pct": ZERO,
            "net_projected_return_pct": ZERO,
            "low_return_pct": ZERO,
            "high_return_pct": ZERO,
            "projected_gain_amount": ZERO,
            "downside_amount": ZERO,
        }

    gross_dividend_yield_pct = projection.get("gross_dividend_yield_pct") or ZERO
    gross_dividend_income = (allocated_amount * gross_dividend_yield_pct) / ONE_HUNDRED
    broker_costs = estimate_broker_costs(
        broker_name=position.broker,
        trade_channel=position.trade_channel,
        trade_amount=allocated_amount,
        valuation_amount=allocated_amount,
        annual_dividend_income=gross_dividend_income,
        quote_symbol=position.quote_symbol,
    )
    annual_cost_used, annual_cost_source = resolve_recurring_cost_used(
        position.annual_maintenance_cost,
        broker_costs.get("annual_recurring_cost", ZERO),
    )
    net_dividend_income = gross_dividend_income - broker_costs.get("annual_dividend_fee", ZERO)
    net_annual_income = net_dividend_income - annual_cost_used
    net_income_yield_pct = (net_annual_income / allocated_amount) * ONE_HUNDRED if allocated_amount else ZERO
    transaction_drag_pct = (
        (broker_costs.get("roundtrip_total_cost", ZERO) / allocated_amount) * ONE_HUNDRED
        if allocated_amount
        else ZERO
    )
    price_return_pct = effective_projection.get("price_return_pct")
    if price_return_pct is None:
        fallback_base_return_pct = projection.get("base_return_pct") or ZERO
        fallback_net_income_yield_pct = projection.get("net_income_yield_pct") or ZERO
        fallback_transaction_drag_pct = projection.get("transaction_drag_pct") or ZERO
        price_return_pct = fallback_base_return_pct - fallback_net_income_yield_pct + fallback_transaction_drag_pct
    price_low_return_pct = effective_projection.get("price_low_return_pct")
    if price_low_return_pct is None:
        price_low_return_pct = effective_projection.get("low_return_pct")
    if price_low_return_pct is None:
        price_low_return_pct = price_return_pct
    price_high_return_pct = effective_projection.get("price_high_return_pct")
    if price_high_return_pct is None:
        price_high_return_pct = effective_projection.get("high_return_pct")
    if price_high_return_pct is None:
        price_high_return_pct = price_return_pct
    net_projected_return_pct = clamp_decimal(
        price_return_pct + net_income_yield_pct - transaction_drag_pct,
        Decimal("-45.00"),
        Decimal("45.00"),
    )
    low_return_pct = clamp_decimal(
        price_low_return_pct + net_income_yield_pct - transaction_drag_pct,
        Decimal("-80.00"),
        Decimal("120.00"),
    )
    high_return_pct = clamp_decimal(
        price_high_return_pct + net_income_yield_pct - transaction_drag_pct,
        Decimal("-80.00"),
        Decimal("140.00"),
    )
    projected_gain_amount = (allocated_amount * net_projected_return_pct) / ONE_HUNDRED
    downside_amount = (allocated_amount * low_return_pct) / ONE_HUNDRED

    return {
        "gross_dividend_income": quantize_decimal(gross_dividend_income, "0.01") or ZERO,
        "net_dividend_income": quantize_decimal(net_dividend_income, "0.01") or ZERO,
        "annual_cost_used": quantize_decimal(annual_cost_used, "0.01") or ZERO,
        "annual_cost_source": annual_cost_source,
        "roundtrip_total_cost": quantize_decimal(broker_costs.get("roundtrip_total_cost", ZERO), "0.01") or ZERO,
        "purchase_total_cost": quantize_decimal(broker_costs.get("purchase_total_cost", ZERO), "0.01") or ZERO,
        "annual_dividend_fee": quantize_decimal(broker_costs.get("annual_dividend_fee", ZERO), "0.01") or ZERO,
        "transaction_drag_pct": quantize_decimal(transaction_drag_pct) or ZERO,
        "net_income_yield_pct": quantize_decimal(net_income_yield_pct) or ZERO,
        "net_projected_return_pct": quantize_decimal(net_projected_return_pct) or ZERO,
        "low_return_pct": quantize_decimal(low_return_pct) or ZERO,
        "high_return_pct": quantize_decimal(high_return_pct) or ZERO,
        "projected_gain_amount": quantize_decimal(projected_gain_amount, "0.01") or ZERO,
        "downside_amount": quantize_decimal(downside_amount, "0.01") or ZERO,
    }


def review_optimizer_ticket_efficiency(candidate: dict, allocated_amount: Decimal, scenario: dict) -> dict:
    if allocated_amount <= ZERO:
        return {
            "keep": False,
            "reason": "sin_asignacion",
            "entry_drag_pct": None,
            "roundtrip_drag_pct": None,
            "gain_to_roundtrip_multiple": None,
        }

    purchase_total_cost = scenario.get("purchase_total_cost", ZERO)
    roundtrip_total_cost = scenario.get("roundtrip_total_cost", ZERO)
    projected_gain_amount = scenario.get("projected_gain_amount", ZERO)
    entry_drag_pct = (
        (purchase_total_cost / allocated_amount) * ONE_HUNDRED
        if allocated_amount and purchase_total_cost > ZERO
        else ZERO
    )
    roundtrip_drag_pct = (
        (roundtrip_total_cost / allocated_amount) * ONE_HUNDRED
        if allocated_amount and roundtrip_total_cost > ZERO
        else ZERO
    )
    gain_to_roundtrip_multiple = (
        (projected_gain_amount / roundtrip_total_cost)
        if roundtrip_total_cost > ZERO and projected_gain_amount is not None
        else None
    )

    if purchase_total_cost <= ZERO and roundtrip_total_cost <= ZERO:
        return {
            "keep": True,
            "reason": "",
            "entry_drag_pct": entry_drag_pct,
            "roundtrip_drag_pct": roundtrip_drag_pct,
            "gain_to_roundtrip_multiple": gain_to_roundtrip_multiple,
        }

    if entry_drag_pct > OPTIMIZER_MAX_ENTRY_DRAG_PCT:
        return {
            "keep": False,
            "reason": "entry_drag",
            "entry_drag_pct": entry_drag_pct,
            "roundtrip_drag_pct": roundtrip_drag_pct,
            "gain_to_roundtrip_multiple": gain_to_roundtrip_multiple,
        }

    if roundtrip_drag_pct > OPTIMIZER_MAX_ROUNDTRIP_DRAG_PCT:
        return {
            "keep": False,
            "reason": "roundtrip_drag",
            "entry_drag_pct": entry_drag_pct,
            "roundtrip_drag_pct": roundtrip_drag_pct,
            "gain_to_roundtrip_multiple": gain_to_roundtrip_multiple,
        }

    if (
        gain_to_roundtrip_multiple is not None
        and projected_gain_amount > ZERO
        and gain_to_roundtrip_multiple < OPTIMIZER_MIN_GAIN_TO_ROUNDTRIP_MULTIPLE
    ):
        return {
            "keep": False,
            "reason": "gain_vs_cost",
            "entry_drag_pct": entry_drag_pct,
            "roundtrip_drag_pct": roundtrip_drag_pct,
            "gain_to_roundtrip_multiple": gain_to_roundtrip_multiple,
        }

    return {
        "keep": True,
        "reason": "",
        "entry_drag_pct": entry_drag_pct,
        "roundtrip_drag_pct": roundtrip_drag_pct,
        "gain_to_roundtrip_multiple": gain_to_roundtrip_multiple,
        }


def scheduled_review_iso_weekdays() -> tuple[int, ...]:
    return tuple(
        weekday
        for weekday in getattr(settings, "EQUITIES_SCHEDULED_OPTIMIZATION_ISO_WEEKDAYS", (2, 4))
        if 1 <= int(weekday) <= 7
    )


def build_scheduled_review_weekdays_label(weekdays: tuple[int, ...] | list[int]) -> str:
    labels = [
        SCHEDULED_REVIEW_WEEKDAY_LABELS.get(int(weekday), str(weekday))
        for weekday in weekdays
        if 1 <= int(weekday) <= 7
    ]
    if not labels:
        return "sin dias configurados"
    if len(labels) == 1:
        return labels[0]
    return ", ".join(labels[:-1]) + " y " + labels[-1]


def resolve_next_review_date(
    start_date: date,
    weekdays: tuple[int, ...] | list[int],
    *,
    include_today: bool = False,
) -> date | None:
    normalized_weekdays = tuple(sorted({int(weekday) for weekday in weekdays if 1 <= int(weekday) <= 7}))
    if not normalized_weekdays:
        return None
    current_date = start_date if include_today else (start_date + timedelta(days=1))
    for _ in range(14):
        if current_date.isoweekday() in normalized_weekdays:
            return current_date
        current_date += timedelta(days=1)
    return None


def build_round_review_dates(as_of: date, rounds_count: int) -> list[date]:
    if rounds_count <= 0:
        return []
    dates = [as_of]
    weekdays = scheduled_review_iso_weekdays()
    cursor = as_of
    while len(dates) < rounds_count:
        next_date = resolve_next_review_date(cursor, weekdays, include_today=False)
        if next_date is None:
            next_date = cursor + timedelta(days=7)
        dates.append(next_date)
        cursor = next_date
    return dates


def build_equity_round_investment_plan(
    owned_history_cards: list[dict],
    optimizer_cards: list[dict],
    target_total_capital: Decimal,
    max_round_amount: Decimal,
    max_company_pct: Decimal,
    *,
    strategy_mode: str = OPTIMIZER_STRATEGY_12M_PRIMARY,
    as_of: date | None = None,
) -> dict:
    strategy = get_optimizer_strategy_config(strategy_mode)
    if target_total_capital <= 0 or max_round_amount <= 0 or max_company_pct <= 0:
        return {
            "available": False,
            "reason": "Parametros insuficientes para calcular el plan por rondas.",
            "strategy_label": strategy["label"],
        }

    as_of = as_of or django_timezone.localdate()
    company_cap_amount = quantize_decimal((target_total_capital * max_company_pct) / Decimal("100"), "0.01") or ZERO
    current_allocations_by_ticker: dict[str, Decimal] = defaultdict(lambda: ZERO)
    current_allocations_count_by_ticker: dict[str, int] = defaultdict(int)
    for card in owned_history_cards:
        position = card["position"]
        ticker = position.ticker
        current_allocations_by_ticker[ticker] += quantize_decimal(position.current_value, "0.01") or ZERO
        current_allocations_count_by_ticker[ticker] += 1
    current_invested_total = sum(current_allocations_by_ticker.values(), ZERO)
    capital_to_deploy = quantize_decimal(target_total_capital - current_invested_total, "0.01") or ZERO
    if capital_to_deploy <= ZERO:
        return {
            "available": False,
            "reason": "La cartera ya alcanza o supera el paquete objetivo indicado. No hace falta abrir nuevas rondas.",
            "strategy_label": strategy["label"],
            "current_invested_total": current_invested_total,
            "target_total_capital": target_total_capital,
        }

    candidate_pool = []
    for card in optimizer_cards:
        candidate = build_equity_optimizer_candidate(card, strategy["mode"])
        if candidate is None:
            continue
        if candidate["position"].is_owned:
            trade_plan = build_owned_cycle_trade_timing_plan(candidate["card"])
            if trade_plan.get("mode") in {"sale_reentry", "sale_review"}:
                continue
        candidate_pool.append(candidate)

    positive_candidates = rank_optimizer_candidates(
        filter_positive_optimizer_candidates(candidate_pool, strategy["mode"])
    )
    annualized_target_excluded_count = sum(
        1
        for candidate in candidate_pool
        if candidate["optimization_score"] > ZERO
        and candidate.get("trade_alert_label") != "Vender"
        and candidate.get("holding_annualized_return_pct") is not None
        and not candidate_meets_target_annualized_return(candidate)
    )
    conservative_profile_excluded_count = sum(
        1
        for candidate in candidate_pool
        if candidate["optimization_score"] > ZERO
        and candidate.get("trade_alert_label") != "Vender"
        and candidate_meets_target_annualized_return(candidate)
        and not candidate_passes_conservative_profile(candidate)
    )
    if not positive_candidates:
        return {
            "available": False,
            "reason": (
                "Ahora mismo no hay candidatas con retorno neto positivo suficiente para abrir rondas nuevas."
                + (
                    f" Las ideas tacticas detectadas tampoco alcanzan el objetivo minimo de {OPTIMIZER_TARGET_HOLDING_ANNUALIZED_RETURN_PCT:.0f} %/a."
                    if annualized_target_excluded_count
                    else ""
                )
                + (
                    f" Ademas, el perfil {OPTIMIZER_RISK_PROFILE_LABEL.lower()} deja fuera {conservative_profile_excluded_count} candidata(s) por riesgo excesivo."
                    if conservative_profile_excluded_count
                    else ""
                )
            ),
            "strategy_label": strategy["label"],
        }

    planned_allocations_by_ticker: dict[str, Decimal] = defaultdict(lambda: ZERO, current_allocations_by_ticker)
    rounds = []
    remaining_capital = capital_to_deploy
    max_rounds = max(int(math.ceil(float(capital_to_deploy / max_round_amount))), 1)
    round_dates = build_round_review_dates(as_of, max_rounds)

    for round_index, round_date in enumerate(round_dates, start=1):
        if remaining_capital <= ZERO:
            break
        best_choice = None
        for candidate in positive_candidates:
            ticker = candidate["position"].ticker
            current_amount = planned_allocations_by_ticker.get(ticker, ZERO)
            available_capacity = quantize_decimal(company_cap_amount - current_amount, "0.01") or ZERO
            proposed_amount = min(max_round_amount, remaining_capital, available_capacity)
            if proposed_amount <= ZERO:
                continue
            scenario = build_optimizer_allocation_scenario(candidate, proposed_amount)
            review = review_optimizer_ticket_efficiency(candidate, proposed_amount, scenario)
            if not review.get("keep"):
                continue
            projected_gain_amount = scenario.get("projected_gain_amount", ZERO)
            choice = {
                "candidate": candidate,
                "amount": proposed_amount,
                "scenario": scenario,
                "projected_gain_amount": projected_gain_amount,
                "current_amount": current_amount,
                "current_weight_pct": quantize_decimal((current_amount / target_total_capital) * Decimal("100"))
                if target_total_capital
                else ZERO,
                "post_weight_pct": quantize_decimal(((current_amount + proposed_amount) / target_total_capital) * Decimal("100"))
                if target_total_capital
                else ZERO,
            }
            choice_key = (
                projected_gain_amount,
                scenario.get("net_projected_return_pct", ZERO),
                candidate["optimization_score"],
                Decimal("1") if current_amount == ZERO else ZERO,
            )
            if best_choice is None or choice_key > best_choice["choice_key"]:
                choice["choice_key"] = choice_key
                best_choice = choice

        if best_choice is None:
            break

        candidate = best_choice["candidate"]
        position = candidate["position"]
        amount = best_choice["amount"]
        planned_allocations_by_ticker[position.ticker] += amount
        remaining_capital = quantize_decimal(remaining_capital - amount, "0.01") or ZERO
        is_existing_holding = best_choice["current_amount"] > ZERO
        rounds.append(
            {
                "round_number": round_index,
                "round_date": round_date,
                "round_date_label": round_date.isoformat(),
                "ticker": position.ticker,
                "company_name": position.company_name,
                "action_label": "Ampliar" if is_existing_holding else "Abrir",
                "status_label": candidate.get("status_label") or "",
                "amount": amount,
                "current_weight_pct": best_choice["current_weight_pct"],
                "post_weight_pct": best_choice["post_weight_pct"],
                "net_projected_return_pct": best_choice["scenario"].get("net_projected_return_pct"),
                "cycle_return_5y_pct": candidate.get("cycle_return_5y_pct"),
                "reliability_label": candidate.get("reliability_label") or "",
                "reliability_score": candidate.get("reliability_score"),
                "trade_alert_label": candidate.get("trade_alert_label") or "",
                "reference_label": candidate.get("reference_label") or "",
                "optimization_score": candidate.get("optimization_score"),
                "note": (
                    f"{'Suma' if is_existing_holding else 'Abre'} posicion en {position.company_name} con "
                    f"retorno neto 12M del {best_choice['scenario'].get('net_projected_return_pct', ZERO):.2f} % "
                    f"y peso final del {best_choice['post_weight_pct'] or ZERO:.2f} % del paquete."
                ),
            }
        )

    current_overweights = [
        {
            "ticker": ticker,
            "amount": amount,
            "weight_pct": quantize_decimal((amount / target_total_capital) * Decimal("100")) if target_total_capital else ZERO,
        }
        for ticker, amount in current_allocations_by_ticker.items()
        if amount > company_cap_amount
    ]
    return {
        "available": bool(rounds),
        "reason": "" if rounds else "Con los limites actuales no sale ninguna ronda que compense en esta foto del radar.",
        "strategy_label": strategy["label"],
        "target_total_capital": target_total_capital,
        "current_invested_total": current_invested_total,
        "capital_to_deploy": capital_to_deploy,
        "remaining_capital": remaining_capital,
        "max_round_amount": max_round_amount,
        "max_company_pct": max_company_pct,
        "company_cap_amount": company_cap_amount,
        "rounds": rounds,
        "rounds_count": len(rounds),
        "new_tickets_count": sum(1 for item in rounds if item["action_label"] == "Abrir"),
        "top_up_rounds_count": sum(1 for item in rounds if item["action_label"] == "Ampliar"),
        "current_overweights": current_overweights,
        "current_overweights_count": len(current_overweights),
        "cadence_label": build_scheduled_review_weekdays_label(scheduled_review_iso_weekdays()),
    }


def build_equity_allocation_plan(
    history_cards: list[dict],
    total_investment: Decimal,
    max_company_pct: Decimal,
    max_total_positions: int = 0,
    max_sector_positions: int = 0,
    selected_sectors: list[str] | None = None,
    selected_owned_tickers: list[str] | None = None,
    selected_owned_tickers_applied: bool = False,
    strategy_mode: str = OPTIMIZER_STRATEGY_12M_PRIMARY,
) -> dict:
    strategy = get_optimizer_strategy_config(strategy_mode)
    if total_investment <= 0 or max_company_pct <= 0:
        return {
            "available": False,
            "reason": "Parametros insuficientes para calcular la propuesta.",
            "strategy_mode": strategy["mode"],
            "strategy_label": strategy["label"],
        }

    company_cap_amount = (total_investment * max_company_pct) / Decimal("100")
    portfolio_context = build_purchase_discipline_portfolio_context(history_cards)
    candidates = apply_purchase_discipline_to_optimizer_candidates(
        [candidate for candidate in (build_equity_optimizer_candidate(card, strategy["mode"]) for card in history_cards) if candidate],
        portfolio_context,
    )

    if not candidates:
        return {
            "available": False,
            "reason": f"Todavia no hay suficientes proyecciones para proponer una distribucion con foco en {strategy['primary_horizon_label']}.",
            "strategy_mode": strategy["mode"],
            "strategy_label": strategy["label"],
        }

    positive_candidates = filter_positive_optimizer_candidates(candidates, strategy["mode"])
    annualized_target_excluded_candidates = [
        item
        for item in candidates
        if item["optimization_score"] > ZERO
        and item.get("trade_alert_label") != "Vender"
        and item.get("holding_annualized_return_pct") is not None
        and not candidate_meets_target_annualized_return(item)
    ]
    conservative_profile_excluded_candidates = [
        item
        for item in candidates
        if item["optimization_score"] > ZERO
        and item.get("trade_alert_label") != "Vender"
        and candidate_meets_target_annualized_return(item)
        and not candidate_passes_conservative_profile(item)
    ]
    if not positive_candidates:
        annualized_target_reason = ""
        if annualized_target_excluded_candidates:
            annualized_target_reason = (
                f" Ademas, {len(annualized_target_excluded_candidates)} candidata(s) con tramo tactico claro no alcanzan "
                f"el objetivo minimo de {OPTIMIZER_TARGET_HOLDING_ANNUALIZED_RETURN_PCT:.0f} % anualizado."
            )
        conservative_profile_reason = ""
        if conservative_profile_excluded_candidates:
            conservative_profile_reason = (
                f" El perfil {OPTIMIZER_RISK_PROFILE_LABEL.lower()} tambien aparta {len(conservative_profile_excluded_candidates)} candidata(s) "
                "porque el peor escenario, la volatilidad, la incertidumbre o la seguridad no son lo bastante defensivos."
            )
        return {
            "available": False,
            "reason": (
                "Ahora mismo ninguna accion ofrece retorno robusto positivo "
                f"con prioridad en {strategy['primary_horizon_label']} despues de descontar riesgo, costes y penalizaciones de incertidumbre."
                + annualized_target_reason
                + conservative_profile_reason
            ),
            "strategy_mode": strategy["mode"],
            "strategy_label": strategy["label"],
        }

    owned_positions_available_count = sum(1 for item in candidates if item["position"].is_owned)
    normalized_selected_sectors = []
    seen_selected_sectors = set()
    for sector in selected_sectors or []:
        normalized_sector = str(sector or "").strip()
        if not normalized_sector or normalized_sector in seen_selected_sectors:
            continue
        seen_selected_sectors.add(normalized_sector)
        normalized_selected_sectors.append(normalized_sector)
    selected_sector_set = set(normalized_selected_sectors)
    normalized_selected_owned_tickers = []
    seen_selected_owned_tickers = set()
    for ticker in selected_owned_tickers or []:
        normalized_ticker = str(ticker or "").strip().upper()
        if not normalized_ticker or normalized_ticker in seen_selected_owned_tickers:
            continue
        seen_selected_owned_tickers.add(normalized_ticker)
        normalized_selected_owned_tickers.append(normalized_ticker)
    selected_owned_ticker_set = set(normalized_selected_owned_tickers)
    selected_owned_tickers_applied = bool(selected_owned_tickers_applied and owned_positions_available_count > 0)
    selected_sector_excluded_candidates = []
    selected_owned_ticker_excluded_candidates = []
    if selected_sector_set:
        selected_sector_filtered_candidates = []
        for item in positive_candidates:
            sector_label = item.get("sector_label") or "Sin sector"
            if sector_label in selected_sector_set:
                selected_sector_filtered_candidates.append(item)
            else:
                selected_sector_excluded_candidates.append(item)
        positive_candidates = selected_sector_filtered_candidates

    if selected_owned_tickers_applied:
        owned_ticker_filtered_candidates = []
        for item in positive_candidates:
            ticker = str(item["position"].ticker or "").strip().upper()
            if item["position"].is_owned and ticker not in selected_owned_ticker_set:
                selected_owned_ticker_excluded_candidates.append(item)
            else:
                owned_ticker_filtered_candidates.append(item)
        positive_candidates = owned_ticker_filtered_candidates

    if not positive_candidates and (selected_sector_set or selected_owned_tickers_applied):
        selected_sector_note = ""
        if selected_sector_set:
            selected_sector_note = (
                "La optimizacion se ha limitado a estos sectores: "
                + ", ".join(normalized_selected_sectors)
                + "."
            )
        selected_owned_ticker_note = ""
        if selected_owned_tickers_applied:
            if normalized_selected_owned_tickers:
                selected_owned_ticker_note = (
                    "De las acciones compradas solo se han dejado activas estas: "
                    + ", ".join(normalized_selected_owned_tickers)
                    + "."
                )
            else:
                selected_owned_ticker_note = "Has desactivado todas las acciones compradas para esta optimizacion."
        return {
            "available": False,
            "reason": (
                "Con los filtros manuales actuales no quedan candidatas validas para construir la propuesta. "
                "Prueba ampliando sectores, reactivando acciones compradas o relajando otras restricciones."
            ),
            "strategy_mode": strategy["mode"],
            "strategy_label": strategy["label"],
            "selected_sectors": normalized_selected_sectors,
            "selected_sector_note": selected_sector_note,
            "selected_owned_tickers": normalized_selected_owned_tickers,
            "selected_owned_tickers_applied": selected_owned_tickers_applied,
            "selected_owned_ticker_note": selected_owned_ticker_note,
        }

    ranked_candidates = rank_optimizer_candidates(positive_candidates)
    sector_excluded_candidates = []
    if max_sector_positions > 0:
        ranked_candidates, sector_excluded_candidates = filter_optimizer_candidates_by_sector(
            ranked_candidates,
            max_sector_positions,
        )

    if not ranked_candidates:
        return {
            "available": False,
            "reason": "Con el limite por sector indicado no quedan suficientes candidatos validos para construir la propuesta.",
            "strategy_mode": strategy["mode"],
            "strategy_label": strategy["label"],
        }

    candidate_pool = ranked_candidates
    position_cap_overflow_count = 0
    if max_total_positions > 0:
        position_cap_overflow_count = max(len(candidate_pool) - max_total_positions, 0)
        iteration_candidates = candidate_pool[:max_total_positions]
        next_candidate_index = len(iteration_candidates)
    else:
        iteration_candidates = candidate_pool
        next_candidate_index = len(candidate_pool)

    ticket_filtered_count = 0
    ticket_filter_reasons = {"entry_drag": 0, "roundtrip_drag": 0, "gain_vs_cost": 0}
    remaining_amount = total_investment
    allocated_amounts = []
    kept_scenarios = []
    max_iterations = len(candidate_pool) + 1

    for _ in range(max_iterations):
        allocated_amounts, remaining_amount = distribute_optimizer_amounts(
            iteration_candidates,
            total_investment,
            company_cap_amount,
        )
        rejected_reviews = []
        kept_scenarios = []
        for index, (candidate, allocated_amount) in enumerate(zip(iteration_candidates, allocated_amounts)):
            if allocated_amount <= ZERO:
                kept_scenarios.append(None)
                continue
            scenario = build_optimizer_allocation_scenario(candidate, allocated_amount)
            review = review_optimizer_ticket_efficiency(candidate, allocated_amount, scenario)
            scenario["entry_drag_pct"] = quantize_decimal(review.get("entry_drag_pct")) or ZERO
            scenario["roundtrip_drag_pct"] = quantize_decimal(review.get("roundtrip_drag_pct")) or ZERO
            scenario["gain_to_roundtrip_multiple"] = quantize_decimal(review.get("gain_to_roundtrip_multiple")) if review.get("gain_to_roundtrip_multiple") is not None else None
            if not review["keep"]:
                rejected_reviews.append(
                    {
                        "index": index,
                        "reason": review["reason"],
                        "entry_drag_pct": review.get("entry_drag_pct") or ZERO,
                        "roundtrip_drag_pct": review.get("roundtrip_drag_pct") or ZERO,
                        "gain_to_roundtrip_multiple": review.get("gain_to_roundtrip_multiple"),
                        "allocated_amount": allocated_amount,
                    }
                )
                kept_scenarios.append(None)
                continue
            kept_scenarios.append(scenario)

        if not rejected_reviews:
            break
        rejected_reviews.sort(
            key=lambda item: (
                item["roundtrip_drag_pct"],
                item["entry_drag_pct"],
                -(item["gain_to_roundtrip_multiple"] if item["gain_to_roundtrip_multiple"] is not None else Decimal("999")),
                -item["allocated_amount"],
            ),
            reverse=True,
        )
        rejected_index = rejected_reviews[0]["index"]
        rejected_reason = rejected_reviews[0]["reason"]
        ticket_filtered_count += 1
        if rejected_reason in ticket_filter_reasons:
            ticket_filter_reasons[rejected_reason] += 1
        next_iteration_candidates = [
            candidate
            for index, candidate in enumerate(iteration_candidates)
            if index != rejected_index
        ]
        if max_total_positions > 0:
            while len(next_iteration_candidates) < max_total_positions and next_candidate_index < len(candidate_pool):
                next_iteration_candidates.append(candidate_pool[next_candidate_index])
                next_candidate_index += 1
        iteration_candidates = next_iteration_candidates
        if not iteration_candidates:
            break

    ranked_candidates = iteration_candidates
    allocations = []
    for rank, (candidate, allocated_amount, scenario) in enumerate(zip(ranked_candidates, allocated_amounts, kept_scenarios), start=1):
        if allocated_amount <= ZERO:
            continue
        if scenario is None:
            scenario = build_optimizer_allocation_scenario(candidate, allocated_amount)
        purchase_timing = dict(candidate.get("purchase_timing") or {"available": False})
        allocations.append(
            {
                "rank": rank,
                "position": candidate["position"],
                "status_key": candidate["status_key"],
                "status_label": candidate["status_label"],
                "sector_label": candidate["sector_label"],
                "reference_label": candidate["reference_label"],
                "strategy_mode": candidate["strategy_mode"],
                "strategy_label": candidate["strategy_label"],
                "allocated_amount": allocated_amount,
                "allocated_weight_pct": (allocated_amount / total_investment) * Decimal("100"),
                "base_return_pct": candidate["base_return_pct"],
                "raw_scenario_expected_return_pct": candidate.get("raw_scenario_expected_return_pct"),
                "scenario_expected_return_pct": candidate.get("scenario_expected_return_pct"),
                "scenario_expectation_review": candidate.get("scenario_expectation_review"),
                "downside_stress_return_pct": candidate.get("downside_stress_return_pct"),
                "scenario_spread_pct": candidate.get("scenario_spread_pct"),
                "primary_signal_pct": candidate["primary_signal_pct"],
                "secondary_signal_pct": candidate["secondary_signal_pct"],
                "adjusted_return_pct": candidate["risk_adjusted_return_pct"],
                "robust_return_signal_pct": candidate.get("robust_return_signal_pct"),
                "blended_return_signal_pct": candidate["blended_return_signal_pct"],
                "net_projected_return_pct": scenario["net_projected_return_pct"],
                "low_return_pct": scenario["low_return_pct"],
                "high_return_pct": scenario["high_return_pct"],
                "projected_gain_amount": scenario["projected_gain_amount"],
                "downside_amount": scenario["downside_amount"],
                "expected_net_dividend_income": scenario["net_dividend_income"],
                "expected_gross_dividend_income": scenario["gross_dividend_income"],
                "annual_cost_used": scenario["annual_cost_used"],
                "roundtrip_total_cost": scenario["roundtrip_total_cost"],
                "purchase_total_cost": scenario["purchase_total_cost"],
                "entry_drag_pct": scenario.get("entry_drag_pct", ZERO),
                "roundtrip_drag_pct": scenario.get("roundtrip_drag_pct", ZERO),
                "gain_to_roundtrip_multiple": scenario.get("gain_to_roundtrip_multiple"),
                "transaction_drag_pct": scenario["transaction_drag_pct"],
                "net_income_yield_pct": scenario["net_income_yield_pct"],
                "confidence_label": candidate["confidence_label"],
                "reliability_label": candidate["reliability_label"],
                "reliability_score": candidate["reliability_score"],
                "safety_score": candidate["safety_score"],
                "base_optimization_score": candidate.get("base_optimization_score", candidate["optimization_score"]),
                "optimization_score": candidate["optimization_score"],
                "purchase_discipline": candidate.get("purchase_discipline") or {},
                "purchase_discipline_score": candidate.get("purchase_discipline_score"),
                "purchase_discipline_label": candidate.get("purchase_discipline_label", ""),
                "purchase_discipline_reason": candidate.get("purchase_discipline_reason", ""),
                "purchase_discipline_adjustment_pct": candidate.get("purchase_discipline_adjustment_pct", ZERO),
                "trade_alert_label": candidate["trade_alert_label"],
                "trade_alert_tone": candidate["trade_alert_tone"],
                "external_signal_label": candidate["external_signal_label"],
                "external_signal_score": candidate["external_signal_score"],
                "external_signal_note": candidate["external_signal_note"],
                "external_signal_items_count": candidate["external_signal_items_count"],
                "annualized_volatility_pct": candidate["annualized_volatility_pct"],
                "years_covered": candidate["years_covered"],
                "cycle_phase": candidate["cycle_phase"],
                "cycle_projection_available": candidate["cycle_projection_available"],
                "cycle_return_annual_pct": candidate["cycle_return_annual_pct"],
                "cycle_return_5y_pct": candidate["cycle_return_5y_pct"],
                "raw_cycle_expected_annual_return_pct": candidate.get("raw_cycle_expected_annual_return_pct"),
                "cycle_expected_annual_return_pct": candidate.get("cycle_expected_annual_return_pct"),
                "cycle_expectation_review": candidate.get("cycle_expectation_review"),
                "cycle_downside_stress_annual_pct": candidate.get("cycle_downside_stress_annual_pct"),
                "cycle_scenario_spread_pct": candidate.get("cycle_scenario_spread_pct"),
                "holding_annualized_return_pct": candidate.get("holding_annualized_return_pct"),
                "annualized_target_return_pct": candidate.get("annualized_target_return_pct"),
                "annualized_target_gap_pct": candidate.get("annualized_target_gap_pct"),
                "meets_target_annualized_return": candidate.get("meets_target_annualized_return"),
                "passes_conservative_profile": candidate.get("passes_conservative_profile"),
                "cycle_yearly_margins": build_cycle_projection_yearly_margins(
                    candidate["position"].current_price_per_share,
                    candidate["card"].get("cycle_projection_5y") or {},
                    first_year_projected_price=(candidate.get("effective_projection_metrics") or {}).get("projected_price"),
                    first_year_return_pct=candidate["base_return_pct"],
                ),
                "cycle_support_score": candidate["cycle_support_score"],
                "robust_cycle_support_score": candidate.get("robust_cycle_support_score"),
                "material_event": candidate.get("material_event", False),
                "uncertainty_penalty_pct": candidate.get("uncertainty_penalty_pct", ZERO),
                "max_drawdown_pct": candidate["max_drawdown_pct"],
                "current_drawdown_pct": candidate["current_drawdown_pct"],
                "projected_price": candidate["projection"].get("projected_price"),
                "purchase_timing": purchase_timing,
            }
        )

    allocated_amount_total = sum((item["allocated_amount"] for item in allocations), ZERO)
    projected_gain_total = sum((item["projected_gain_amount"] for item in allocations), ZERO)
    weighted_return_pct = (
        (projected_gain_total / allocated_amount_total) * Decimal("100")
        if allocated_amount_total
        else None
    )
    weighted_low_return_pct = (
        (sum((item["downside_amount"] for item in allocations), ZERO) / allocated_amount_total) * Decimal("100")
        if allocated_amount_total
        else None
    )
    weighted_expected_return_pct = (
        sum(
            (
                (item["scenario_expected_return_pct"] * item["allocated_amount"])
                for item in allocations
                if item.get("scenario_expected_return_pct") is not None
            ),
            ZERO,
        )
        / allocated_amount_total
        if allocated_amount_total
        else None
    )
    weighted_stress_return_pct = (
        sum(
            (
                (item["downside_stress_return_pct"] * item["allocated_amount"])
                for item in allocations
                if item.get("downside_stress_return_pct") is not None
            ),
            ZERO,
        )
        / allocated_amount_total
        if allocated_amount_total
        else None
    )
    weighted_safety_score = (
        sum((item["safety_score"] * item["allocated_amount"] for item in allocations), ZERO) / allocated_amount_total
        if allocated_amount_total
        else None
    )
    weighted_reliability_score = (
        sum((item["reliability_score"] * item["allocated_amount"] for item in allocations), ZERO) / allocated_amount_total
        if allocated_amount_total
        else None
    )
    weighted_purchase_discipline_score = (
        sum(
            (
                ((item.get("purchase_discipline_score") or ZERO) * item["allocated_amount"])
                for item in allocations
                if item.get("purchase_discipline_score") is not None
            ),
            ZERO,
        )
        / allocated_amount_total
        if allocated_amount_total
        else None
    )
    cycle_allocated_amount_total = sum(
        (item["allocated_amount"] for item in allocations if item.get("cycle_return_annual_pct") is not None),
        ZERO,
    )
    weighted_cycle_return_annual_pct = (
        sum(
            (
                (item["cycle_return_annual_pct"] * item["allocated_amount"])
                for item in allocations
                if item.get("cycle_return_annual_pct") is not None
            ),
            ZERO,
        )
        / cycle_allocated_amount_total
        if cycle_allocated_amount_total
        else None
    )
    weighted_cycle_return_5y_pct = (
        sum(
            (
                (item["cycle_return_5y_pct"] * item["allocated_amount"])
                for item in allocations
                if item.get("cycle_return_5y_pct") is not None
            ),
            ZERO,
        )
        / cycle_allocated_amount_total
        if cycle_allocated_amount_total
        else None
    )
    weighted_volatility_pct = (
        sum((item["annualized_volatility_pct"] * item["allocated_amount"] for item in allocations), ZERO) / allocated_amount_total
        if allocated_amount_total
        else None
    )
    weighted_uncertainty_penalty_pct = (
        sum((item["uncertainty_penalty_pct"] * item["allocated_amount"] for item in allocations), ZERO) / allocated_amount_total
        if allocated_amount_total
        else None
    )
    target_annualized_allocated_amount_total = sum(
        (
            item["allocated_amount"]
            for item in allocations
            if (item.get("purchase_timing") or {}).get("holding_annualized_return_pct") is not None
        ),
        ZERO,
    )
    weighted_holding_annualized_return_pct = (
        sum(
            (
                (item["purchase_timing"]["holding_annualized_return_pct"] * item["allocated_amount"])
                for item in allocations
                if (item.get("purchase_timing") or {}).get("holding_annualized_return_pct") is not None
            ),
            ZERO,
        )
        / target_annualized_allocated_amount_total
        if target_annualized_allocated_amount_total
        else None
    )
    target_holding_annualized_return_pct = OPTIMIZER_TARGET_HOLDING_ANNUALIZED_RETURN_PCT
    weighted_holding_annualized_target_gap_pct = (
        quantize_decimal(
            weighted_holding_annualized_return_pct - target_holding_annualized_return_pct,
            "0.01",
        )
        if weighted_holding_annualized_return_pct is not None
        else None
    )
    net_dividend_income_total = sum((item["expected_net_dividend_income"] for item in allocations), ZERO)
    annual_cost_total = sum((item["annual_cost_used"] for item in allocations), ZERO)
    roundtrip_cost_total = sum((item["roundtrip_total_cost"] for item in allocations), ZERO)
    owned_allocations_count = sum(1 for item in allocations if item["position"].is_owned)
    ibex_allocations_count = sum(1 for item in allocations if item["status_key"] == "ibex")
    watchlist_allocations_count = sum(
        1
        for item in allocations
        if not item["position"].is_owned and item["status_key"] != "ibex"
    )
    sectors_used = sorted({item["sector_label"] for item in allocations if item.get("sector_label")})
    external_signal_used_count = sum(1 for item in allocations if item.get("external_signal_items_count"))
    material_event_allocations_count = sum(1 for item in allocations if item.get("material_event"))
    shock_adjusted_allocations_count = sum(1 for item in allocations if (item.get("uncertainty_penalty_pct") or ZERO) > ZERO)
    allocations_with_timing_count = sum(
        1
        for item in allocations
        if (item.get("purchase_timing") or {}).get("holding_annualized_return_pct") is not None
    )
    allocations_meeting_target_count = sum(
        1
        for item in allocations
        if (
            (item.get("purchase_timing") or {}).get("holding_annualized_return_pct") is not None
            and (item["purchase_timing"]["holding_annualized_return_pct"] >= target_holding_annualized_return_pct)
        )
    )
    weighted_target_compliance_pct = (
        quantize_decimal(
            (
                sum(
                    (
                        item["allocated_amount"]
                        for item in allocations
                        if (
                            (item.get("purchase_timing") or {}).get("holding_annualized_return_pct") is not None
                            and item["purchase_timing"]["holding_annualized_return_pct"] >= target_holding_annualized_return_pct
                        )
                    ),
                    ZERO,
                )
                / allocated_amount_total
            ) * ONE_HUNDRED,
            "0.01",
        )
        if allocated_amount_total
        else None
    )
    weighted_conservative_profile_compliance_pct = (
        quantize_decimal(
            (
                sum(
                    (
                        item["allocated_amount"]
                        for item in allocations
                        if candidate_passes_conservative_profile(item)
                    ),
                    ZERO,
                )
                / allocated_amount_total
            ) * ONE_HUNDRED,
            "0.01",
        )
        if allocated_amount_total
        else None
    )

    if remaining_amount > ZERO:
        reserve_reason = (
            "Queda caja en reserva porque el filtro de riesgo/rentabilidad solo deja pasar las ideas con retorno robusto positivo, "
            "riesgo asumible y costes razonables."
        )
        if annualized_target_excluded_candidates:
            reserve_reason += (
                f" Tambien se reserva capital porque {len(annualized_target_excluded_candidates)} candidata(s) no llegan "
                f"al objetivo de {target_holding_annualized_return_pct:.0f} %/a."
            )
        if conservative_profile_excluded_candidates:
            reserve_reason += (
                f" El perfil {OPTIMIZER_RISK_PROFILE_LABEL.lower()} deja ademas fuera {len(conservative_profile_excluded_candidates)} candidata(s) "
                "por no ofrecer suficiente proteccion en peor escenario, volatilidad, incertidumbre o calidad del modelo."
            )
        if max_total_positions > 0 and (company_cap_amount * Decimal(str(max_total_positions))) < total_investment:
            reserve_reason += (
                " Ademas, el maximo total de empresas combinado con el peso maximo por empresa impide asignar todo el capital sin saltarse tus propios limites."
            )
    else:
        reserve_reason = ""
    if ticket_filtered_count:
        ticket_filter_note = (
            "Se han descartado "
            f"{ticket_filtered_count} propuesta(s) porque el coste fijo+variable por operacion se comia demasiado retorno esperado."
        )
    else:
        ticket_filter_note = ""
    if selected_sector_set:
        selected_sector_note = "La optimizacion se limita a estos sectores: " + ", ".join(normalized_selected_sectors) + "."
        if selected_sector_excluded_candidates:
            selected_sector_note += (
                f" {len(selected_sector_excluded_candidates)} candidata(s) han quedado fuera por no pertenecer a los sectores permitidos."
            )
    else:
        selected_sector_note = ""
    if selected_owned_tickers_applied:
        if normalized_selected_owned_tickers:
            selected_owned_ticker_note = (
                "De las acciones compradas solo pueden entrar estas: "
                + ", ".join(normalized_selected_owned_tickers)
                + "."
            )
        else:
            selected_owned_ticker_note = "Se han desactivado todas las acciones compradas para esta optimizacion."
        if selected_owned_ticker_excluded_candidates:
            selected_owned_ticker_note += (
                f" {len(selected_owned_ticker_excluded_candidates)} candidata(s) compradas han quedado fuera por este filtro manual."
            )
    else:
        selected_owned_ticker_note = ""
    if max_total_positions > 0:
        position_limit_note = f"Se aplica un maximo total de {max_total_positions} empresa(s) en la cartera propuesta."
        if position_cap_overflow_count:
            position_limit_note += (
                f" En el ranking inicial habia {position_cap_overflow_count} candidata(s) adicionales fuera del corte por capacidad."
            )
    else:
        position_limit_note = ""
    if max_sector_positions > 0:
        sector_limit_note = (
            f"Se aplica un maximo de {max_sector_positions} empresa(s) por sector."
        )
        if sector_excluded_candidates:
            sector_limit_note += f" {len(sector_excluded_candidates)} candidata(s) han quedado fuera por diversificacion sectorial."
    else:
        sector_limit_note = ""
    if annualized_target_excluded_candidates:
        annualized_target_note = (
            f"Se han dejado fuera {len(annualized_target_excluded_candidates)} candidata(s) cuyo mejor tramo tactico no alcanza "
            f"el objetivo minimo de {target_holding_annualized_return_pct:.0f} % anualizado."
        )
    else:
        annualized_target_note = (
            f"El optimizador exige, cuando existe plan tactico de entrada y salida, un objetivo minimo de "
            f"{target_holding_annualized_return_pct:.0f} % anualizado por compra."
        )
    conservative_profile_note = (
        f"Perfil {OPTIMIZER_RISK_PROFILE_LABEL.lower()}: solo entran ideas con alerta Comprar, seguridad >= "
        f"{OPTIMIZER_CONSERVATIVE_MIN_SAFETY_SCORE:.0f}, fiabilidad >= {OPTIMIZER_CONSERVATIVE_MIN_RELIABILITY_SCORE:.0f}, "
        f"peor escenario >= {OPTIMIZER_CONSERVATIVE_MIN_WORST_RETURN_PCT:.0f} %, estres >= {OPTIMIZER_CONSERVATIVE_MIN_STRESS_RETURN_PCT:.0f} %, "
        f"volatilidad <= {OPTIMIZER_CONSERVATIVE_MAX_VOLATILITY_PCT:.0f} % e incertidumbre <= {OPTIMIZER_CONSERVATIVE_MAX_UNCERTAINTY_PENALTY_PCT:.2f} pts."
    )
    reason = ""
    if not allocations:
        reason = (
            (
                annualized_target_note
                if annualized_target_excluded_candidates
                else ""
            )
            or ticket_filter_note
            or position_limit_note
            or sector_limit_note
            or conservative_profile_note
            or "Con las tarifas actuales del broker, las compras que caben por peso maximo no compensan el coste fijo+variable por operacion."
        )

    top_pick = allocations[0] if allocations else None
    top_pick_purchase_timing = dict(top_pick.get("purchase_timing") or {}) if top_pick else {}
    purchase_discipline_rows = [
        {
            "rank": item["rank"],
            "ticker": item["position"].ticker,
            "company_name": item["position"].company_name,
            "label": item.get("purchase_discipline_label") or "",
            "action_label": (item.get("purchase_discipline") or {}).get("action_label", ""),
            "tone": (item.get("purchase_discipline") or {}).get("tone", ""),
            "score": item.get("purchase_discipline_score"),
            "reason": item.get("purchase_discipline_reason") or "",
            "buy_window_label": (item.get("purchase_timing") or {}).get("buy_window_label", ""),
            "holding_annualized_return_pct": (item.get("purchase_timing") or {}).get("holding_annualized_return_pct"),
        }
        for item in allocations[:5]
    ]
    return {
        "available": bool(allocations),
        "reason": reason,
        "strategy_mode": strategy["mode"],
        "strategy_label": strategy["label"],
        "strategy_description": strategy["description"],
        "primary_horizon_label": strategy["primary_horizon_label"],
        "secondary_horizon_label": strategy["secondary_horizon_label"],
        "allocations": allocations,
        "total_investment": total_investment,
        "max_company_pct": max_company_pct,
        "max_total_positions": max_total_positions,
        "company_cap_amount": company_cap_amount,
        "selected_sectors": normalized_selected_sectors,
        "selected_sector_note": selected_sector_note,
        "selected_owned_tickers": normalized_selected_owned_tickers,
        "selected_owned_tickers_applied": selected_owned_tickers_applied,
        "selected_owned_ticker_note": selected_owned_ticker_note,
        "owned_positions_available_count": owned_positions_available_count,
        "allocated_amount_total": allocated_amount_total,
        "cash_reserve_amount": remaining_amount,
        "projected_gain_total": projected_gain_total,
        "weighted_return_pct": weighted_return_pct,
        "weighted_low_return_pct": weighted_low_return_pct,
        "weighted_expected_return_pct": weighted_expected_return_pct,
        "weighted_stress_return_pct": weighted_stress_return_pct,
        "weighted_safety_score": weighted_safety_score,
        "weighted_reliability_score": weighted_reliability_score,
        "weighted_purchase_discipline_score": weighted_purchase_discipline_score,
        "purchase_discipline_rows": purchase_discipline_rows,
        "weighted_cycle_return_annual_pct": weighted_cycle_return_annual_pct,
        "weighted_cycle_return_5y_pct": weighted_cycle_return_5y_pct,
        "target_holding_annualized_return_pct": target_holding_annualized_return_pct,
        "weighted_holding_annualized_return_pct": weighted_holding_annualized_return_pct,
        "weighted_holding_annualized_target_gap_pct": weighted_holding_annualized_target_gap_pct,
        "allocations_with_timing_count": allocations_with_timing_count,
        "allocations_meeting_target_count": allocations_meeting_target_count,
        "weighted_target_compliance_pct": weighted_target_compliance_pct,
        "weighted_conservative_profile_compliance_pct": weighted_conservative_profile_compliance_pct,
        "weighted_volatility_pct": weighted_volatility_pct,
        "weighted_uncertainty_penalty_pct": weighted_uncertainty_penalty_pct,
        "net_dividend_income_total": net_dividend_income_total,
        "annual_cost_total": annual_cost_total,
        "roundtrip_cost_total": roundtrip_cost_total,
        "owned_allocations_count": owned_allocations_count,
        "watchlist_allocations_count": watchlist_allocations_count,
        "ibex_allocations_count": ibex_allocations_count,
        "max_sector_positions": max_sector_positions,
        "position_cap_filtered_count": position_cap_overflow_count,
        "position_cap_note": position_limit_note,
        "sectors_used": sectors_used,
        "sectors_used_count": len(sectors_used),
        "external_signal_used_count": external_signal_used_count,
        "material_event_allocations_count": material_event_allocations_count,
        "shock_adjusted_allocations_count": shock_adjusted_allocations_count,
        "sector_filtered_count": len(sector_excluded_candidates),
        "sector_filter_note": sector_limit_note,
        "annualized_target_filtered_count": len(annualized_target_excluded_candidates),
        "annualized_target_note": annualized_target_note,
        "risk_profile_label": OPTIMIZER_RISK_PROFILE_LABEL,
        "conservative_profile_filtered_count": len(conservative_profile_excluded_candidates),
        "conservative_profile_note": conservative_profile_note,
        "selected_owned_ticker_filtered_count": len(selected_owned_ticker_excluded_candidates),
        "ticket_filtered_count": ticket_filtered_count,
        "ticket_filter_reasons": ticket_filter_reasons,
        "ticket_filter_note": ticket_filter_note,
        "reserve_reason": reserve_reason,
        "top_pick": top_pick,
        "top_pick_purchase_timing": top_pick_purchase_timing,
        "methodology_note": (
            f"Esta ejecucion usa la estrategia {strategy['label']}: "
            f"prioriza la lectura de {strategy['primary_horizon_label']} y deja {strategy['secondary_horizon_label']} como contraste secundario para reforzar o penalizar la jerarquia final. "
            "La jerarquia ya no depende solo del caso central: combina retorno esperado por escenarios, peor caso, dispersion entre escenarios, seguridad, fiabilidad del modelo, alertas de tendencia, senal externa reciente de prensa y castigo automatico por shocks informativos. "
            f"Como perfil {OPTIMIZER_RISK_PROFILE_LABEL.lower()}, el ranking prioriza preservar capital: favorece mejor peor escenario, menor volatilidad, menos incertidumbre y mayor seguridad/fiabilidad antes de apurar retorno. "
            f"La seguridad ({int(OPTIMIZER_MIN_SAFETY_SCORE)} como referencia), la fiabilidad ({int(OPTIMIZER_MIN_RELIABILITY_SCORE)} como referencia) y la lectura de la ficha ya no expulsan por un corte fijo: penalizan el score de forma progresiva para que el optimizador siga priorizando rentabilidad, pero no ignore el riesgo. "
            f"Ademas, cuando hay plan tactico de compra, se exige un objetivo minimo de {target_holding_annualized_return_pct:.0f} % anualizado; si una idea no llega, se deja fuera y el capital queda en reserva. "
            f"Sobre ese objetivo, el filtro conservador exige alerta Comprar, seguridad >= {OPTIMIZER_CONSERVATIVE_MIN_SAFETY_SCORE:.0f}, fiabilidad >= {OPTIMIZER_CONSERVATIVE_MIN_RELIABILITY_SCORE:.0f}, peor escenario >= {OPTIMIZER_CONSERVATIVE_MIN_WORST_RETURN_PCT:.0f} %, estres >= {OPTIMIZER_CONSERVATIVE_MIN_STRESS_RETURN_PCT:.0f} %, volatilidad <= {OPTIMIZER_CONSERVATIVE_MAX_VOLATILITY_PCT:.0f} % e incertidumbre <= {OPTIMIZER_CONSERVATIVE_MAX_UNCERTAINTY_PENALTY_PCT:.2f} pts. "
            "Las cifras monetarias del informe siguen expresadas a 12 meses para mantener comparables dividendos, costes, caja y peor escenario, aunque la jerarquia de seleccion pueda priorizar 5 anos. "
            "eficiencia real del ticket de compra y, si lo marcas, un maximo total de empresas y diversificacion maxima por sector. En la version robusta se "
            "analiza siempre todo el IBEX, no solo los valores que ya tienes en seguimiento."
        ),
    }
